# -*- coding: utf-8 -*-
# local modules import
from general_helpers import datestr_to_python
# python general modules import
import openpyxl
import os
import io

from general_helpers import NRtoDay
from general_helpers import iso_to_gregorian
from general_helpers import create_locations_dict
from general_helpers import create_classtypes_dict
from general_helpers import max_string_length
from general_helpers import get_last_day_month
from general_helpers import get_lastweek_year
from general_helpers import get_months_list
from general_helpers import set_form_id_and_get_submit_button

from openstudio.os_staff_schedule import StaffSchedule


def get_shiftname(shID):
    record = db.shifts(shID)
    location = db.school_locations(record.school_locations_id).Name
    shift = db.school_shifts(record.school_shifts_id).Name
    Weekday = NRtoDay(record.Week_day)

    return Weekday + " " + \
           location + " " + \
           record.Starttime.strftime('%H:%M') + " " + \
           shift


def shifts_get_back(var=None):
    """
        General back button for shift edit pages
    """
    return os_gui.get_button('back', URL('schedule'))


@auth.requires_login()
def shift_add():
    """
        This function shows an add page for a shift
    """
    response.title = T("Add a new shift")
    response.subtitle = T("")
    response.view = 'general/tabs_menu.html'

    next_url = '/staff/shift_employees_add?shID=[id]&wiz=True'
    return_url = URL('schedule')

    crud.messages.submit_button = T("Next")
    crud.messages.record_created = T("Added shift")
    crud.settings.create_next = next_url
    form = crud.create(db.shifts)

    result = set_form_id_and_get_submit_button(form, 'MainForm')
    form = result['form']
    submit = result['submit']

    back = os_gui.get_button('back', return_url)
    menu = shift_add_get_menu(request.function)

    return dict(content=form, back=back, menu=menu, save=submit)


def shift_add_get_menu(page):
    """
        Returns submenu for adding a workshop
    """
    pages = [ ['shift_add', T('1. Shift'), "#"],
              ['shift_employees_add', T('2. Assign employees'), "#"] ]

    return os_gui.get_submenu(pages, page, horizontal=True, htype='tabs')


def shifts_edit_get_menu(page, shID, date_formatted):
    """
        This function returns a submenu for the class edit pages
    """
    vars = {'shID':shID, 'date':date_formatted}
    pages = [['shift_edit',
              T('Edit'),
              URL('shift_edit', vars=vars)],
             ['shift_staff',
              T('Staff'),
              URL('shift_staff',  vars=vars)]]

    return os_gui.get_submenu(pages, page, horizontal=True, htype='tabs')


@auth.requires_login()
def shift_edit():
    """
        This function shows an edit page for a shift
    """
    response.title = T("Edit shift")
    shID = request.vars['shID']
    date_formatted = request.vars['date']
    response.subtitle = get_shiftname(shID)
    response.view = 'general/tabs_menu.html'

    shift = db.shifts(shID)

    return_url = URL('schedule')

    crud.messages.submit_button = T("Save")
    crud.messages.record_updated = T("Saved")
    crud.settings.update_next = URL(vars=request.vars)
    crud.settings.update_deletable = False
    form = crud.update(db.shifts, shID)

    result = set_form_id_and_get_submit_button(form, 'MainForm')
    form = result['form']
    submit = result['submit']

    links = [ A(SPAN(_class="glyphicon glyphicon-duplicate"), " ", T("Duplicate"),
                  _href=URL('shift_duplicate', vars={'shID':shID}),
                  _title=T("Duplicate this shift")) ]

    tools = os_gui.get_dropdown_menu(links,
                                     '',
                                     btn_size='',
                                     btn_icon='wrench',
                                     menu_class='pull-right'
                                     )

    menu = shifts_edit_get_menu(request.function, shID, date_formatted)
    back = shifts_get_back()

    content = form

    return dict(content=content,
                menu=menu,
                back=back,
                header_tools=tools,
                save=submit)


@auth.requires(auth.has_membership(group_id='Admins') or
               auth.has_permission('update', 'shifts_otc'))
def shift_edit_on_date():
    """
        Edit class on a selected date without editing other classes in range
    """
    shID = request.vars['shID']
    date_formatted = request.vars['date']
    date = datestr_to_python(DATE_FORMAT, date_formatted)

    shiftname = get_shiftname(shID)

    response.title = shiftname
    response.subtitle = date_formatted
    response.view = 'general/only_content.html'

    result = shift_edit_on_date_get_form(shID, date_formatted)
    form = result['form']

    form_id = "MainForm"
    form_element = form.element('form')
    form['_id'] = form_id

    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = form_id

    submit = form.element('input[type=submit]')

    sotcID  = result['sotcID']
    link_remove = shift_edit_on_date_get_link_remove(sotcID)

    content = form

    content = DIV(DIV(form,
                      _class='col-md-9'),
                  DIV(link_remove,
                      _class='col-md-3'))

    back = SPAN(shifts_get_back(), get_week_chooser(request.function, shID, date))

    return dict(content=content,
                week_chooser=get_week_chooser(request.function, shID, date),
                back=back,
                save=submit)


def shift_edit_on_date_get_form(shID, date_formatted):
    """
        Returns add or edit crud form for one time shift change
    """
    date = datestr_to_python(DATE_FORMAT, date_formatted)

    query = (db.shifts_otc.shifts_id == shID) & \
            (db.shifts_otc.ShiftDate == date)
    rows = db(query).select(db.shifts_otc.id)

    next_url = URL('shift_edit_on_date', vars={'shID':shID,
                                               'date':date_formatted})

    sotcID = None
    if len(rows):
        sotcID = rows.first().id

    if sotcID:
        crud.settings.update_next = next_url
        form = crud.update(db.shifts_otc, sotcID)
    else:
        db.shifts_otc.shifts_id.default = shID
        db.shifts_otc.ShiftDate.default = date

        crud.settings.create_next = next_url
        form = crud.create(db.shifts_otc)

    return dict(form=form, sotcID=sotcID)


def shift_edit_on_date_get_link_remove(sotcID):
    """
        @param cotcID: db.shifts_otc.id
        @return: link with confirmation to remove a all changes made to a shift
    """
    if sotcID is None:
        return ''

    onclick = "return confirm('" + \
        T('Do you really want to remove all changes made to the shift on this date?') + "');"

    link = A(SPAN(SPAN(_class=os_gui.get_icon('remove')), ' ',
                  T('Remove changes')),
             _href=URL('shift_edit_on_date_remove_changes',
                       vars={'sotcID':sotcID}),
             _onclick=onclick,
             _class='red')

    return link


@auth.requires(auth.has_membership(group_id='Admins') or
               auth.has_permission('delete', 'shifts_otc'))
def shift_edit_on_date_remove_changes():
    """
        Remove all changes made to a shift
    """
    sotcID = request.vars['sotcID']

    query = (db.shifts_otc.id == sotcID)
    db(query).delete()

    session.flash = T('Removed changes')

    redirect(URL('schedule'))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'shifts_staff'))
def shift_staff():
    """
        Overview with teachers for a class
    """
    shID = request.vars['shID']
    date_formatted = request.vars['date']

    response.title = T("Edit shift")
    response.subtitle = get_shiftname(shID)
    response.view = 'general/tabs_menu.html'

    links = [lambda row: os_gui.get_button('edit', URL('shift_employees_edit',
                                                vars={'shsID':row.id,
                                                      'shID' :shID,
                                                      'date' :date_formatted}))]

    query = (db.shifts_staff.shifts_id == shID)

    fields = [db.shifts_staff.auth_employee_id,
              db.shifts_staff.auth_employee_id2,
              db.shifts_staff.Startdate,
              db.shifts_staff.Enddate,
              ]

    delete_permission = auth.has_membership(group_id='Admins') or \
                        auth.has_permission('delete', 'shifts_staff')

    grid = SQLFORM.grid(query,
                        fields=fields,
                        links=links,
                        details=False,
                        searchable=False,
                        deletable=delete_permission,
                        csv=False,
                        create=False,
                        editable=False,
                        orderby=~db.shifts_staff.Startdate,
                        field_id=db.shifts_staff.id,
                        ui = grid_ui)
    grid.element('.web2py_counter', replace=None) # remove the counter
    grid.elements('span[title=Delete]', replace=None) # remove text from delete button

    add_permission = (auth.has_membership(group_id='Admins') or
                      auth.has_permission('create', 'shifts_staff'))
    if add_permission:
        add = os_gui.get_button('add', URL('shift_employees_add',
                                           vars={'shID' : shID,
                                                 'date' : date_formatted}))
    else:
        add = ''

    menu = shifts_edit_get_menu(request.function, shID, date_formatted)
    back = shifts_get_back()

    return dict(content=grid,
                menu=menu,
                back=back,
                add=add)


@auth.requires_login()
def shift_employees_add():
    """
        This function shows an add page for shift_employees
    """
    shID = request.vars['shID']
    date_formatted = request.vars['date']
    wizzard = True if 'wiz' in request.vars else False

    response.title = T("Assign employees")
    response.subtitle = get_shiftname(shID)

    return_url = shifts_staff_add_edit_get_return_url(shID, date_formatted)

    if wizzard:
        response.view = 'general/tabs_menu.html'
        menu = shift_add_get_menu(request.function)
        back = os_gui.get_button('back', return_url)

    else:
        response.view = 'general/only_content.html'
        menu = ''
        back = os_gui.get_button('back', return_url)

    # set default values for date fields if this will be the first record created
    query = (db.shifts_staff.shifts_id == shID)
    num_records = db(query).count()
    if not num_records:
        shift = db.shifts(shID)
        db.shifts_staff.Startdate.default = shift.Startdate
        db.shifts_staff.Enddate.default = shift.Enddate

    db.shifts_staff.shifts_id.default = shID

    crud.messages.submit_button = T("Save")
    crud.messages.record_created = T("Saved")
    crud.settings.create_next = return_url
    form = crud.create(db.shifts_staff)


    result = set_form_id_and_get_submit_button(form, 'MainForm')
    form = result['form']
    submit = result['submit']

    submit = form.element('input[type=submit]')

    return dict(content=form, back=back, menu=menu, save=submit)


@auth.requires_login()
def shift_employees_edit():
    """
        This function shows an edit page for the employee assigned to a shift
    """
    response.title = T("Edit assigned employees")
    shsID = request.vars['shsID']
    shID  = request.vars['shID']
    date_formatted = request.vars['date']
    response.subtitle = get_shiftname(shID)
    response.view = 'general/only_content.html'

    return_url = shifts_staff_add_edit_get_return_url(shID, date_formatted)

    crud.messages.submit_button = T("Save")
    crud.messages.record_updated = T("Saved")
    crud.settings.update_next = return_url
    crud.settings.update_deletable = False
    form = crud.update(db.shifts_staff, shsID)

    form_id = "MainForm"
    form_element = form.element('form')
    form['_id'] = form_id

    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = form_id

    submit = form.element('input[type=submit]')

    back = os_gui.get_button('back', return_url)

    return dict(content=form,
                save=submit,
                back=back)


def shifts_staff_add_edit_get_return_url(shID, date_formatted):
    """
        Returns return url for adding or editing an employee
    """
    return URL('shift_staff', vars ={'shID':shID,
                                      'date' :date_formatted})


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('delete', 'shifts'))
def shift_delete():
    """
        Removed the selected class and redirect to the manage page
    """
    shID = request.vars['shID']

    query = (db.shifts.id == shID)
    result = db(query).delete()

    session.flash = T('Failed to delete shift')
    if result == 1: # Success
        session.flash = T('Deleted shift')

    redirect(URL('schedule'))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('create', 'shifts'))
def shift_duplicate():
    shID = request.vars['shID']
    row = db.shifts[shID]

    id = db.shifts.insert(
        school_locations_id = row.school_locations_id,
        school_shifts_id    = row.school_shifts_id,
        Week_day            = row.Week_day,
        Starttime           = row.Starttime,
        Endtime             = row.Endtime,
        Startdate           = row.Startdate,
        Enddate             = row.Enddate )

    # duplicate assigned employees
    query = (db.shifts_staff.shifts_id == shID)
    rows = db(query).select(db.shifts_staff.ALL)
    for row in rows:
        db.shifts_staff.insert(
            shifts_id          = id,
            auth_employee_id   = row.auth_employee_id,
            auth_employee_id2  = row.auth_employee_id2,
            Startdate          = row.Startdate,
            Enddate            = row.Enddate
        )

    session.flash = T("You are now editing the duplicated shift")

    redirect(URL('shift_edit', vars={'shID':id}))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'shifts'))
def schedule():
    """
        Main list of shifts
    """
    session.schedule_staff_holidays_back = 'staff_schedule'

    ## sort form begin ##
    if 'sort' in request.vars:
        session.staff_schedule_sort = request.vars['sort']
    elif session.staff_schedule_sort is None or \
         session.staff_schedule_sort == '':
        # check if we have a default value
        sys_property = 'staff_schedule_default_sort'
        sorting = get_sys_property(sys_property)
        if sorting:
            session.staff_schedule_sort = sorting
        else:
            # default to sorting by location
            session.staff_schedule_sort = 'location'

    form_sort = schedule_get_form_sort()

    ## sort form end ##

    if 'schedule_year' in request.vars:
        year = int(request.vars['schedule_year'])
    elif not session.staff_schedule_year is None:
        year = session.staff_schedule_year
    else:
        today = datetime.date.today()
        year = datetime.date.today().year
    session.staff_schedule_year = year
    if 'schedule_week' in request.vars:
        week = int(request.vars['schedule_week'])
    elif not session.staff_schedule_week is None:
        week = session.staff_schedule_week
    else:
        week = today.isocalendar()[1]
        if week == 0:
            week = 1
    session.staff_schedule_week = week
    if not session.schedule_show_date is None:
        date = datestr_to_python(DATE_FORMAT, session.schedule_show_date)
        year = date.year
        week = date.isocalendar()[1]
        session.schedule_show_date = None

    # if we used the jump date box to select a week
    if 'schedule_jump_date' in request.vars:
        jump_date = datestr_to_python(DATE_FORMAT,
                                      request.vars['schedule_jump_date'])
        jump_date_iso = jump_date.isocalendar()
        year = jump_date_iso[0]
        week = jump_date_iso[1]
        session.staff_schedule_year = year
        session.staff_schedule_week = week
    else:
        jump_date = datetime.date.today()

    ## week selection
    form_week = schedule_get_form_week(year, week)
    form_year = schedule_get_form_year(year)
    # jump to date
    form_jump = schedule_get_form_jump(jump_date)

    current_week = A(T('Current week'),
                     _href=URL('schedule_current_week'),
                     _class='btn btn-default full-width',
                     _id='schedule_current_week')

    modals = DIV()
    # schedule weekly status
    schedule_status = schedule_get_status(modals)

    # show schedule status
    week_chooser = schedule_get_week_chooser()

    ## week selection end ##

    # filter
    if 'shift' in request.vars:
        location = request.vars['location']
        employee = request.vars['employee']
        shift = request.vars['shift']
        status = request.vars['status']
        filter_form = schedule_get_filter_form(
            shift, location, employee, status)
        session.staff_schedule_filter_location = location
        session.staff_schedule_filter_employee = employee
        session.staff_schedule_filter_shift = shift
        session.staff_schedule_filter_status = status
    elif not session.schedule_filter_location is None:
        filter_form = schedule_get_filter_form(
            session.staff_schedule_filter_shift,
            session.staff_schedule_filter_location,
            session.staff_schedule_filter_employee,
            session.staff_schedule_filter_status)
    else:
        filter_form = schedule_get_filter_form()
        session.staff_schedule_filter_location = None
        session.staff_schedule_filter_employee = None
        session.staff_schedule_filter_shift = None
        session.staff_schedule_filter_status = None

    title = T('Schedule')
    response.title = T('Studio staff')
    response.subtitle = schedule_get_subtitle(year, week)

    days = dict()
    for day in range(1, 8):
        date = iso_to_gregorian(int(year), int(week), int(day))
        staff_schedule = StaffSchedule(
            date,
            filter_id_school_shifts = session.staff_schedule_filter_shift,
            filter_id_school_locations = session.staff_schedule_filter_location,
            filter_id_employee = session.staff_schedule_filter_employee,
            filter_id_status = session.staff_schedule_filter_status )
        days[day] = staff_schedule.get_day_table()


    schedule_tools = schedule_get_schedule_tools()
    export = schedule_get_export()

    # add new class
    add = ''
    permission = auth.has_membership(group_id='Admins') or \
                 auth.has_permission('create', 'shifts')

    if permission:
        add = os_gui.get_button('add',
                URL('shift_add'),
                tooltip=T("Add a new shift"),
                _class='pull-right')

    mobile_export_title = H4(T("Export"))
    mobile_classes_title = H4(T("Classes"))

    return dict(form_week=form_week,
                form_year=form_year,
                form_jump=form_jump,
                filter_form=filter_form,
                current_week=current_week,
                week_chooser=week_chooser,
                year=year,
                title=title,
                add=add,
                schedule_tools=schedule_tools,
                export=export,
                days=days,
                schedule_status=schedule_status,
                modals=modals,
                )


def schedule_get_subtitle(year, week):
    """
        Returns subtitle for schedule
    """
    return str(year) + " " + T("week") + " " + str(week)


def schedule_get_form_week(year, week):
    """
        Returns week selection form for schedule
    """
    lastweek = get_lastweek_year(year)

    input_week = INPUT(_name='schedule_week',
                       requires=IS_INT_IN_RANGE(1, lastweek),
                       _type='number',
                       _value=week,
                       _min=1,
                       _max=lastweek,
                       _id='no_table_schedule_week')

    submit = INPUT(_type='submit',
                   _value=T('Week'),
                   _class='full-width',
                   _id='no_table_schedule_submit'),

    form = FORM(DIV(DIV(input_week,
                        DIV(submit,
                            _class='input-group-btn'),
                        _class='input-group')),
                _class='form_inline')

    return form


def schedule_get_form_year(year):
    """
        Returns week selection form for schedule
    """
    lastweek = get_lastweek_year(year)

    input_year = INPUT(_name='schedule_year',
                       requires=IS_INT_IN_RANGE(1900, 3000),
                       _type='number',
                       _value=year,
                       _min=datetime.date.today().year-25,
                       _max=datetime.date.today().year+25,
                       _id='no_table_schedule_year')

    submit = INPUT(_type='submit',
                   _value=T('Year'),
                   _class='full-width',
                   _id='no_table_schedule_submit'),

    form = FORM(DIV(DIV(input_year,
                        DIV(submit,
                            _class='input-group-btn'),
                        _class='input-group')),
                _class='form_inline')

    return form


def schedule_get_form_jump(jump_date):
    """
        Returns a form to jump to a date
    """
    form_jump = SQLFORM.factory(
                Field('schedule_jump_date', 'date',
                      requires=IS_DATE_IN_RANGE(
                                format=DATE_FORMAT,
                                minimum=datetime.date(1900,1,1),
                                maximum=datetime.date(2999,1,1)),
                      default=jump_date,
                      label=T(""),
                      widget=os_datepicker_widget_small),
                submit_button=T('Go'),
                )

    submit_jump = form_jump.element('input[type=submit]')
    submit_jump['_class'] = 'full-width'

    form_jump = DIV(form_jump.custom.begin,
                    DIV(form_jump.custom.widget.schedule_jump_date,
                        DIV(form_jump.custom.submit,
                            _class='input-group-btn'),
                        _class='input-group'),
                    form_jump.custom.end,
                    _class='form_inline',
                    _id='form_jump')

    return form_jump


def schedule_get_week_chooser(var=None):
    """
        Returns a week chooser for the schedule
    """
    year = session.staff_schedule_year
    week = session.staff_schedule_week

    lastweek = get_lastweek_year(year)

    if week == 1:
        prev_week  = lastweek
        prev_year  = year - 1
    else:
        prev_week  = week - 1
        prev_year  = year

    if week == lastweek:
        next_week  = 1
        next_year  = year + 1
    else:
        next_week  = week + 1
        next_year  = year


    previous = A(I(_class='fa fa-angle-left'),
                 _href=URL('schedule_set_week', vars={ 'week' : prev_week,
                                                       'year' : prev_year }),
                 _class='btn btn-default')
    nxt = A(I(_class='fa fa-angle-right'),
            _href=URL('schedule_set_week', vars={ 'week' : next_week,
                                                  'year' : next_year }),
            _class='btn btn-default')

    return DIV(previous, nxt, _class='btn-group', _id='schedule_week_chooser')


def schedule_get_schedule_tools(var=None):
    """
        Returns tools for schedule
    """
    schedule_tools = []

    # teacher holidays
    permission = auth.has_membership(group_id='Admins') or \
                 auth.has_permission('read', 'teacher_holidays')

    if permission:
        teacher_holidays = A((os_gui.get_fa_icon('fa-sun-o'),
                              T("Staff holidays")),
                             _href=URL('schedule', 'staff_holidays'),
                             _title=T('Open classes & shifts in a range of dates'))
        schedule_tools.append(teacher_holidays)

    # List open classes
    permission = auth.has_membership(group_id='Admins') or \
                 auth.has_permission('read', 'classes_open')

    if permission:
        shifts_open = A((os_gui.get_fa_icon('fa-bell-o'),
                         T('All open shifts')),
                        _href=URL('shifts_open'),
                        _title=T('List open shifts'))
        schedule_tools.append(shifts_open)

    # Set default sorting
    permission = auth.has_membership(group_id='Admins') or \
                 auth.has_permission('update', 'schedule_set_default_sort')
    if permission:
        set_default_sorting = A(os_gui.get_fa_icon('fa-sort-amount-asc'),
                                T('Set default sorting'),
                                _href=URL('schedule_set_sort_default'),
                                _title=T('Set default sorting for classes'))
        schedule_tools.append(set_default_sorting)

    tools = os_gui.get_dropdown_menu(schedule_tools,
                                     '',
                                     btn_size='btn-sm',
                                     btn_icon='wrench',
                                     menu_class='pull-right'
                                     )

    return tools


def schedule_get_export(var=None):
    """
        Returns export drop down for schedule
    """
    export_locations = A(os_gui.get_fa_icon('fa-calendar'), T("Schedule"),
                          _href=URL('schedule_export_excel',
                                    vars=dict(year=session.staff_schedule_year,
                                              week=session.staff_schedule_week,
                                              export='locations')))

    links = [ export_locations ]

    export = os_gui.get_dropdown_menu(
        links = links,
        btn_text = '',
        btn_size = 'btn-sm',
        btn_icon = 'download',
        menu_class='pull-right')

    return export


def schedule_get_status(modals):
    """
        Returns status for schedule
    """
    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('read', 'schedule_staff_status'):

        schedule_status = SPAN(T("Status"), ": ",
                               SPAN(T("Open"), _class='bold'),
                               ' ',
                               _class='pull-right')
        row = db.schedule_staff_status(ScheduleYear=session.staff_schedule_year,
                                         ScheduleWeek=session.staff_schedule_week)
        if row:
            if row.Status == 'final':
                status = SPAN(T('Final'), _class='green bold')
            schedule_status = DIV(T('Status') + ': ', status + ' ',
                                  _class='pull-right')

        if auth.has_membership(group_id='Admins') or \
           auth.has_permission('update', 'schedule_staff_status'):
            btn_icon = SPAN(_class='glyphicon glyphicon-pencil')

            modal_content = UL(LI(A(T('Open'),
                      _href=URL('schedule_set_week_status',
                                vars={'status':''})),
                      BR(),
                      T('Teachers are being scheduled')),
                   LI(A(SPAN(T('Final'), _class='green'),
                      _href=URL('schedule_set_week_status',
                                vars={'status':'final'})),
                      BR(),
                      T("Schedule for this week has been set")),
                   _class='ul_liststyle_none')
            result = os_gui.get_modal(button_text=XML(btn_icon),
                                      modal_title=T('Staff schedule week status'),
                                      modal_content=modal_content,
                                      modal_class='classes_schedule_status',
                                      button_class='btn-sm')
            modals.append(result['modal'])
            btn_status = result['button']

            schedule_status.append(btn_status)
    else:
        schedule_status = ""

    return schedule_status


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'shifts'))
def schedule_set_week():
    """
        Set the session variables for schedule week and year
    """
    year  = request.vars['year']
    week  = request.vars['week']

    session.staff_schedule_year = int(year)
    session.staff_schedule_week = int(week)

    redirect(URL('schedule'))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'schedule_staff_status'))
def schedule_set_week_status():
    """
        Function to set the weekly status of the schedule
    """
    status = request.vars['status']
    row = db.schedule_staff_status(ScheduleYear=session.staff_schedule_year,
                                   ScheduleWeek=session.staff_schedule_week)
    if row:
        if not status == '':
            # update status
            row.Status = status
            row.update_record()
        else:
            # remove record
            query = (db.schedule_staff_status.id == row.id)
            db(query).delete()
    elif not status == '':
        # add status
        db.schedule_staff_status.insert(ScheduleYear=session.staff_schedule_year,
                                        ScheduleWeek=session.staff_schedule_week,
                                        Status=status)

    redirect(URL('schedule'))


@auth.requires_login()
def schedule_get_sort_options():
    """
        returns a list of sorting options for classes
    """
    sort_options = [ ['location', T('Location')],
                     ['starttime', T('Start time')] ]

    return sort_options


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'staff_schedule_set_default_sort'))
def schedule_set_sort_default():
    """
        Displays a page to edit the default sorting order for the schedule
    """
    response.title = T('Set default sort order for shifts')
    response.subtitle = ''
    response.view = 'general/only_content.html'

    sort_options = schedule_get_sort_options()
    sort = session.staff_schedule_sort

    form = SQLFORM.factory(
        Field('sort',
            requires=IS_IN_SET(sort_options,
                              zero=None),
            default=sort,
            label=""),
        formstyle='divs',
        submit_button=T('Save'),
        )


    return_url = URL('schedule')

    if form.process().accepted:
        session.flash = T('Updated default sort order')
        sort = request.vars['sort']
        session.staff_schedule_sort = sort

        sys_property = 'staff_schedule_default_sort'
        row = db.sys_properties(Property=sys_property)
        if not row:
            db.sys_properties.insert(Property=sys_property,
                                     PropertyValue=sort)
        else:
            row.PropertyValue=sort
            row.update_record()

        # Clear cache
        cache_clear_sys_properties()


        redirect(return_url)

    back = os_gui.get_button('back', return_url)
    return dict(content=form,
                back=back)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'shifts'))
def schedule_current_week():
    session.staff_schedule_week = None
    session.staff_schedule_year = None

    redirect(URL('schedule'))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'shifts'))
def schedule_clear_filter():
    session.staff_schedule_filter_location = None
    session.staff_schedule_filter_employee = None
    session.staff_schedule_filter_shift = None
    session.staff_schedule_filter_status = None

    redirect(URL('schedule'))


def schedule_get_filter_form(school_shifts_id='',
                             school_locations_id='',
                             employee_id='',
                             status=''):

    au_query = (db.auth_user.employee == True) & \
               (db.auth_user.trashed == False)

    form = SQLFORM.factory(
        Field('location',
            requires=IS_IN_DB(db,'school_locations.id', '%(Name)s',
                              zero=T('All locations')),
            default=school_locations_id,
            label=""),
        Field('shift',
            requires=IS_IN_DB(db,'school_shifts.id', '%(Name)s',
                              zero=T('All shifts')),
            default=school_shifts_id,
            label=""),
        Field('employee',
            requires=IS_IN_DB(db(au_query),'auth_user.id',
                              '%(full_name)s',
                              zero=T('All employees')),
            default=employee_id,
            label=""),
        Field('status',
            requires=IS_IN_SET(SHIFT_STATUSES,
                               zero=T('All statuses')),
            default=status,
            label=''),
        submit_button=T('Go'),
        formstyle='divs',
        )

    # sumbit form on change
    selects = form.elements('select')
    for select in selects:
        select.attributes['_onchange'] = "this.form.submit();"

    clear = A(T("Clear"), _class="btn btn-default",
                _href=URL('schedule_clear_filter'),
                _title=T("Reset filter to default"))

    div = DIV(
        form.custom.begin,
        DIV(form.custom.widget.location,
            _class='col-md-2'),
        DIV(form.custom.widget.shift,
            _class='col-md-2'),
        DIV(form.custom.widget.employee,
            _class='col-md-2'),
        DIV(form.custom.widget.status,
            _class='col-md-2'),
        DIV(' ',
            _class='col-md-2'),
        DIV(DIV(form.custom.submit,
                clear,
                _class="pull-right"),
            _class='col-md-2'),
        form.custom.end,
        _id="schedule_filter_form")

    return div


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('read', 'shifts'))
def schedule_export_excel():
    iso_year = request.vars['year']
    iso_week = request.vars['week']
    export = request.vars['export']


    from openpyxl.utils import get_column_letter

    def get_cell_id(col, row):
        """
            Returns cell id for colums / row
        """
        col_letter = get_column_letter(col)
        return col_letter + str(row)


    def writer_location(locID=None):
        """
        This helper function writes the locations schedule
        """
        for day in range(1,8):
            date = iso_to_gregorian(iso_year, iso_week, day)
            staff_schedule = StaffSchedule(date,
                                           filter_id_school_locations = locID)
            rows = staff_schedule.get_day_rows()

            col = day
            dayname = NRtoDay(day)
            c_id = get_cell_id(col, 2)
            ws[c_id] = dayname + " \n" + str(date)
            ws[c_id].alignment = alignment

            r = 3
            for i, row in enumerate(rows):
                repr_row = list(rows[i:i+1].render())[0]
                shID = row.shifts.id

                location = repr_row.shifts.school_locations_id
                shift = repr_row.shifts.school_shifts_id

                employee2 = ''
                if row.shifts_otc.Status == 'sub': # it's a sub shift
                    employee = repr_row.shifts_sub.auth_employee_id
                    if not row.shifts_sub.auth_employee_id2 is None:
                        employee2 = repr_row.shifts_otc.auth_employee_id2
                else:
                    employee = repr_row.shifts_staff.auth_employee_id
                    if not row.shifts_staff.auth_employee_id2 is None:
                        employee2 = repr_row.shifts_otc.auth_employee_id2

                try:
                    employee = employee.decode('utf-8')
                except AttributeError:
                    employee = ''
                try:
                    employee2 = employee2.decode('utf-8')
                except AttributeError:
                    employee2 = ''

                shift_data = ''
                if row.shifts_otc.Status == 'cancelled':
                    shift_data += T("CANCELLED") + "\n"

                time = row.shifts.Starttime.strftime('%H:%M') + "-" + \
                       row.shifts.Endtime.strftime('%H:%M')
                if locID is None:
                    shift_data += location + "\n" + \
                                  time + " \n" + \
                                  shift + " \n" + \
                                  employee + " \n" + \
                                  employee2 + " \n"
                else:
                    shift_data += time + " \n" + \
                                  shift + " \n" + \
                                  employee + " \n" + \
                                  employee2 + " \n"
                c_id = get_cell_id(col, r)
                ws[c_id] = shift_data
                ws[c_id].font = font
                ws[c_id].alignment = alignment
                r += 1

    # Set some general values
    font = openpyxl.styles.Font(size=10)
    alignment = openpyxl.styles.Alignment(wrap_text=True,
                                          shrink_to_fit=True,
                                          vertical='top')

    if export == 'locations':
        # Full schedule
        wb = openpyxl.workbook.Workbook()
        title = "ALL"
        ws = wb.active
        ws.title = title
        ws['A1'] = T("Schedule week") + " " + str(iso_week)
        writer_location()
        # schedule by location
        rows = db().select(db.school_locations.id,
                           db.school_locations.Name,
                           orderby=db.school_locations.Name)
        for row in rows:
            location = locations_dict.get(row.id, None)
            title = location.decode('utf-8')[0:30]
            ws = wb.create_sheet(title=title)
            ws['A1'] = "Schedule" + " " + location + " " + \
                       "week " + str(iso_week)
            writer_location(row.id)

        # create filestream
        stream = io.BytesIO()

        fname = T("Schedule") + '.xlsx'
        wb.save(stream)
        response.headers['Content-Type']='application/vnd.ms-excel'
        response.headers['Content-disposition']='attachment; filename=' + fname

        return stream.getvalue()


def get_week_chooser(function, id, date):
    """
        Returns a week chooser for the schedule
    """
    delta = datetime.timedelta(days=7)
    date_prev = (date-delta).strftime(DATE_FORMAT)
    date_next = (date+delta).strftime(DATE_FORMAT)

    url_prev = URL(function, vars={'shID' : id,
                                   'date'  : date_prev})
    url_next = URL(function, vars={'shID' : id,
                                   'date'  : date_next})

    previous = A(I(_class='fa fa-angle-left'),
                 _href=url_prev,
                 _class='btn btn-default')
    nxt = A(I(_class='fa fa-angle-right'),
            _href=url_next,
            _class='btn btn-default')

    return DIV(previous, nxt, _class='btn-group pull-right')


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'shifts_open'))
def shifts_open():
    """
        List all open shits.
    """
    response.title = T('Open shifts')
    response.subtitle = ''
    response.view = 'general/only_content.html'

    table = TABLE(TR(TH(''), # status marker
                     TH(T('Date')),
                     TH(T('Location')),
                     TH(T('Shift')),
                     TH(T('Time')),
                     TH(),
                     _class='os-table_header'),
                  _class='os-schedule')

    rows = shifts_open_get_rows(TODAY_LOCAL)

    for i, row in enumerate(rows):
        repr_row = list(rows[i:i + 1].render())[0]

        shID = row.shifts.id
        date = row.shifts_otc.ShiftDate
        date_formatted = repr_row.shifts_otc.ShiftDate

        location = max_string_length(repr_row.shifts.school_locations_id, 15)
        shift = max_string_length(repr_row.shifts.school_shifts_id, 24)
        time = SPAN(repr_row.shifts.Starttime, ' - ', repr_row.shifts.Endtime)

        vars = {'shID':shID,
                'date':date_formatted}


        edit = os_gui.get_button('edit_notext',
                                 URL('shift_edit_on_date', vars=vars),
                                 _class='pull-right')


        buttons = ''

        if auth.has_membership(group_id='Admins') or \
            auth.has_permission('update', 'shifts_status'):
            buttons = edit


        row_class = TR(DIV(_class='status_marker bg_red'), # status marker
                       TD(date),
                       TD(location),
                       TD(shift),
                       TD(time),
                       TD(buttons),
                       _class='os-schedule_class')

        table.append(row_class)

    back = os_gui.get_button('back', URL('schedule'))

    return dict(content=DIV(BR(), table),
                back=back)


def shifts_open_get_rows(from_date):
    """
        Return open shift rows
    """
    fields = [
        db.shifts_otc.id,
        db.shifts_otc.ShiftDate,
        db.shifts_otc.Status,
        db.shifts.id,
        db.shifts.school_locations_id,
        db.shifts.school_shifts_id,
        db.shifts.Starttime,
        db.shifts.Endtime
    ]

    query = """
    SELECT sotc.id,
           sotc.ShiftDate,
           sotc.Status,
           sh.id,
           CASE WHEN sotc.school_locations_id IS NOT NULL
                THEN sotc.school_locations_id
                ELSE sh.school_locations_id
                END AS school_locations_id,
           CASE WHEN sotc.school_shifts_id IS NOT NULL
                THEN sotc.school_shifts_id
                ELSE sh.school_shifts_id
                END AS school_shifts_id,
           CASE WHEN sotc.Starttime IS NOT NULL
                THEN sotc.Starttime
                ELSE sh.Starttime
                END AS Starttime,
           CASE WHEN sotc.Endtime IS NOT NULL
                THEN sotc.Endtime
                ELSE sh.Endtime
                END AS Endtime
    FROM shifts_otc sotc
    LEFT JOIN shifts sh ON sh.id = sotc.shifts_id
    WHERE sotc.ShiftDate >= '{date}' AND sotc.Status = 'open'
    ORDER BY sotc.ShiftDate, Starttime
    """.format(date=from_date)

    rows = db.executesql(query, fields=fields)

    return rows


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'classes_notes'))
def notes():
    """
        Add notes to a class
    """
    clsID = request.vars['clsID']
    note_type = request.vars['note_type']
    date_formatted = request.vars['date']
    date = datestr_to_python(DATE_FORMAT, date_formatted)
    response.title = T("Class")
    response.subtitle = get_classname(clsID) + ": " + date_formatted

    response.view = 'general/content_left_sidebar.html'

    backoffice_class = ''
    teachers_class = ''
    all_class = ''

    active_class = 'web2py-menu-active'

    query = (db.classes_notes.classes_id == clsID) & \
            (db.classes_notes.ClassDate == date)

    if note_type is None:
        if session.classes_notes_filter:
            note_type = session.classes_notes_filter
        else:
            note_type = 'teachers'

    if note_type == 'backoffice':
        db.classes_notes.BackofficeNote.default = True
        query &= (db.classes_notes.BackofficeNote == True)

    if note_type == 'teachers':
        db.classes_notes.TeacherNote.default = True
        query &= (db.classes_notes.TeacherNote == True)

    session.classes_notes_filter = note_type


    notes = UL(_id='os-customers_notes')
    rows = db(query).select(db.classes_notes.ALL,
                            orderby=~db.classes_notes.ClassDate|\
                                    ~db.classes_notes.id)
    for row in rows.render():
        row_note_type = ''
        if row.BackofficeNote:
            row_note_type = T('Back office')
        elif row.TeacherNote:
            row_note_type = T('Teachers')

        vars = {'cnID':row.id,
                'date':date_formatted,
                'clsID':clsID}

        buttons = DIV(_class='btn-group right')
        if auth.has_membership(group_id='Admins') or \
           auth.has_permission('update', 'classes_notes'):
            edit = os_gui.get_button('edit_notext',
                              URL('note_edit', vars=vars),
                              cid=request.cid)
            buttons.append(edit)

        if auth.has_membership(group_id='Admins') or \
           auth.has_permission('delete', 'classes_notes'):

            onclick = "return confirm('" + \
                     T('Do you really want to remove this note?') + "');"

            remove = os_gui.get_button('delete_notext',
                                       URL('note_delete', vars=vars),
                                       onclick=onclick)
            buttons.append(remove)

        notes.append(LI(buttons,
                        SPAN(row.ClassDate,
                             _class='bold'),
                        SPAN(' - ',
                             row.auth_user_id,
                             _class='grey'),
                        BR(),
                        XML(row.Note.replace('\n','<br>')),
                        _id='note_' + str(row.id)))


    notes_filter = notes_get_filter_form(session.classes_notes_filter)

    perm = auth.has_membership(group_id='Admins') or \
           auth.has_permission('create', 'classes_notes')
    if perm:
        add = notes_get_add()
        add_title = H4(T('Add a new note'))
    else:
        add = ''
        add_title = ''

    content = DIV(notes_filter, BR())

    # check permissions
    if note_type == 'teachers':
        perm = auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'classes_notes_teachers')
    elif note_type == 'backoffice':
        perm = auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'classes_notes_backoffice')


    if perm:
        content.append(add)
        content.append(notes)

    back = DIV(shifts_get_back(), BR(), BR(),
               get_week_chooser(request.function, clsID, date))
    menu = classes_get_menu(request.function, clsID, date_formatted)

    return dict(content=content,
                back=back,
                menu=menu,
                left_sidebar_enabled=True)


def notes_get_filter_form(page, _class='right'):
    """
        Backoffice / Teachers view selection form
    """
    pages = [('backoffice', T('Back office')), ('teachers', T('Teachers'))]

    if page == 'backoffice':
        value = True
        bo_class = 'btn-info'
    else:
        value = False
        bo_class = 'btn-default'
    input_backoffice = INPUT(_type='radio',
                             _name='note_type',
                             _value='backoffice',
                             _id='radio_bo',
                             _onchange="this.form.submit();",
                             value=value)

    if page == 'teachers':
        value = True
        te_class = 'btn-info'
    else:
        value = False
        te_class = 'btn-default'
    input_teachers = INPUT(_type='radio',
                           _name='note_type',
                           _value='teachers',
                           _id='radio_te',
                           _onchange="this.form.submit();",
                           value=value)

    backoffice_text = T('Back office')
    teachers_text = T('Teachers')

    backoffice = LABEL(backoffice_text, input_backoffice,
                    _class='btn btn-sm ' + bo_class)
    teachers = LABEL(teachers_text, input_teachers,
                     _class='btn btn-sm ' + te_class)

    return FORM(DIV(teachers,
                    backoffice,
                    _class='btn-group',
                    **{'_data-toggle': 'buttons'}),
                _class=_class)


@auth.requires_login()
def note_edit():
    """
        Provides an edit page for a note.
    """
    clsID = request.vars['clsID']
    date_formatted = request.vars['date']
    date = datestr_to_python(DATE_FORMAT, date_formatted)
    response.title = T("Class")
    response.subtitle = get_classname(clsID) + ": " + date_formatted

    response.view = 'general/content_left_sidebar.html'

    cnID = request.vars['cnID']
    request.vars.pop('cnID', None)

    note = db.classes_notes(cnID)
    classes_id = note.classes_id
    date = note.ClassDate

    if note.BackofficeNote:
        note_type = 'backoffice'
    elif note.TeacherNote:
        note_type = 'teachers'

    return_url = URL('notes', vars=request.vars)

    crud.messages.submit_button = T("Save")
    crud.messages.record_updated = T('Saved')
    crud.settings.update_next = return_url
    form = crud.update(db.classes_notes, cnID)

    cancel = A(T('Cancel'),
               _href=return_url,
               _class='btn btn-default')


    form = DIV(form.custom.begin,
               form.custom.widget.Note,
               form.custom.submit,
               cancel,
               form.custom.end,
               _class='os-customers_notes_edit clear')

    back =  os_gui.get_button('back', return_url, _class='left')

    content = DIV(B(T('You are now editing the note below')), form)

    back = DIV(shifts_get_back(), BR(), BR(),
               get_week_chooser(request.function, clsID, date))
    menu = classes_get_menu('notes', clsID, date_formatted)

    return dict(content=content,
                back=back,
                menu=menu,
                left_sidebar_enabled=True)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('delete', 'classes_notes'))
def note_delete():
    """
        Used to remove a note
    """
    response.view = 'generic.json'
    cnID = request.vars['cnID']
    date_formatted = request.vars['date']
    clsID = request.vars['clsID']

    request.vars.pop('cnID', None)

    query = (db.classes_notes.id == cnID)
    result = db(query).delete()

    redirect(URL('notes', vars=request.vars))

def notes_get_add(var=None):
    """
        Provides a page to add a note
        request.vars['note_type'] can be 2 values
            'backoffice' for a backoffice note
            'teachers' for a teachers note
    """
    note_type = request.vars['note_type']
    clsID = request.vars['clsID']
    date_formatted = request.vars['date']
    date = datestr_to_python(DATE_FORMAT, date_formatted)

    vars = {'clsID':clsID,
            'date':date_formatted}

    return_url = URL('notes', vars=vars)

    db.classes_notes.classes_id.default = clsID
    db.classes_notes.ClassDate.default = date

    crud.messages.submit_button = T("Save")
    crud.messages.record_created = T('')
    crud.settings.create_next = return_url
    form = crud.create(db.classes_notes)

    # form.custom.widget.Note['_class'] += ' form-control'

    form_id = 'add_note'
    form = DIV(form.custom.begin,
               form.custom.widget.Note,
               form.custom.submit,
               form.custom.end,
               _class='os-customers_notes_edit collapse',
               _id=form_id)

    add = A(SPAN(SPAN(_class=os_gui.get_icon('plus')), ' ', T('Add')),
            _href='#' + form_id,
            _role='button',
            _class='btn btn-default btn-sm',
            **{'_data-toggle':'collapse',
               '_aria-expanded':'false',
               '_aria-controls':form_id})

    return DIV(add, form)


""" keep this function at the bottom, atom package symbols-tree-view get weird
    from the XML tag and doesn't show any symbols after that """
@auth.requires_login()
def schedule_get_form_sort(var=None):
    """
        Returns sort form for schedule
    """
    sort_options = schedule_get_sort_options()
    sort = session.staff_schedule_sort

    form = SQLFORM.factory(
        Field('sort',
            requires=IS_IN_SET(sort_options,
                              zero=None),
            default=sort,
            label=""),
        formstyle='divs')

    # sumbit form on change
    selects = form.elements('select')
    for select in selects:
        select.attributes['_onchange'] = "this.form.submit();"

    # add permission check to allow setting of default sorting for schedule
    settings = ''
    permission = auth.has_membership(group_id='Admins') or \
                 auth.has_permission('update',
                                     'staff_schedule_set_default_sort')
    if permission:
        settings = SPAN(A(SPAN(_class='glyphicon glyphicon-cog'),
                          _href=URL('schedule_set_sort_default'),
                          _title=T('Set default sorting for classes'),
                          _class='btn btn-default'),
                        _class='input-group-btn')

    form = DIV(H4(T('Sort shifts by'), _class='os-no_margin_top'),
               # this is form.custom.begin with an extra class
               XML('<form action="#" enctype="multipart/form-data" \
                     method="post" class="os-no_margin_bottom">'),
               DIV(form.custom.widget.sort,
                   settings,
                   _class='input-group'),
               form.custom.end,
               _class='os-no_margin select-narrow')

    return form
