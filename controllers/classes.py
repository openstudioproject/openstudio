# -*- coding: utf-8 -*-
# local modules import
from general_helpers import datestr_to_python

# python general modules import
import openpyxl
import os
import cStringIO

from general_helpers import get_badge
from general_helpers import NRtoDay
from general_helpers import iso_to_gregorian
from general_helpers import get_classname
from general_helpers import get_submenu
from general_helpers import classes_get_status
from general_helpers import create_teachers_dict
from general_helpers import create_locations_dict
from general_helpers import create_classtypes_dict
from general_helpers import max_string_length
from general_helpers import class_get_teachers
from general_helpers import get_last_day_month
from general_helpers import get_lastweek_year
from general_helpers import get_months_list
from general_helpers import set_form_id_and_get_submit_button

from openstudio.openstudio import ClasscardsHelper, \
    AttendanceHelper, \
    ReservationHelper, \
    Class, \
    ClassSchedule, \
    CustomersHelper, \
    Invoice, \
    InvoicesHelper

from openstudio.os_customers_subscriptions import CustomerSubscriptions
from openstudio.os_customer import Customer

# helper functions


def classes_get_menu(page, clsID, date_formatted):
    """
        This function returns a menu for the manage page
        page is expected to be a function name showing a page
        class_id is expected to be the class ID
        class_date is expected to be the class date, formatted using DATE_FORMAT
    """
    vars = {'clsID' : clsID,
            'date'  : date_formatted}

    pages = []
    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('read', 'classes_attendance'):
        pages.append([ 'attendance',
                       T('Attendance'),
                       URL('attendance', vars=vars) ])
    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('create', 'classes_attendance_override'):
        pages.append([ 'attendance_override',
                       T("Attendance count"),
                       URL('attendance_override',vars=vars)
                     ])
    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('read', 'classes_reservation'):
        pages.append([ 'reservations',
                       T('Enrollments'),
                       URL('reservations', vars=vars)])
    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('read', 'classes_waitinglist'):
        pages.append([ 'waitinglist',
                        T('Waitinglist'),
                       URL('waitinglist', vars=vars)])
    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('read', 'classes_notes'):
        pages.append([ 'notes',
                        T('Notes'),
                       URL('notes', vars=vars) ])
    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('create', 'classes_otc'):
        pages.append([ 'class_edit_on_date',
                        T('Edit'),
                       URL('class_edit_on_date', vars=vars) ])

    return get_submenu(pages, page, horizontal=True, htype='tabs')


def _check_attendance_remove(row):
    clsID = session.clsID
    date = session.attendance_classdate
    check = db.classes_attendance(customers_id=row.id,
                                  classes_id=clsID,
                                  ClassDate=date)
    if check == None:
        return A(T('Add'),_href=URL("attendance",vars=dict(customers_id=row.id, clsID=clsID, date=date))) # Add to attendance


def _check_attendance_dropin(row):
    clsID = session.clsID
    date = session.attendance_classdate
    return A(T('Drop In'),_href=URL("attendance",vars=dict(customers_id=row.id, clsID=clsID, date=date, attType=2, remove=False)))


def manage_reservation_check_paused(row, value=None):
    if 'customers' in row:
        customers_id = row.auth_user.id
    else:
        customers_id = row.id
    if type(session.register_classdate) != datetime.date:
        date = datestr_to_python(DATE_FORMAT, session.register_classdate)
    else:
        date = session.register_classdate

    return_value = ''
    query = (db.customers_subscriptions.auth_customer_id == customers_id) & \
            (db.customers_subscriptions.Startdate <= date) & \
            ((db.customers_subscriptions.Enddate >= date) |
             (db.customers_subscriptions.Enddate == None))
    current_membership = \
        db(query).select(db.customers_subscriptions.ALL).first()
    if current_membership:
        query = (db.customers_subscriptions_paused.customers_subscriptions_id == current_membership.id) & \
                (db.customers_subscriptions_paused.Startdate <= date) & \
                ((db.customers_subscriptions_paused.Enddate >= date) |
                 (db.customers_subscriptions_paused.Enddate == None))
        check = db(query).select(db.customers_subscriptions_paused.ALL).first()
        if check:
            return_value = T("Yes")
    return return_value


def _check_reservation(row):
    clsID = session.clsID
    check = db.classes_reservation(customers_id=row.id, classes_id=clsID)
    if check is None:
        return os_gui.get_button('add', URL("reserve_class",vars={'cuID':row.id,
                                                            'clsID':clsID}))
    else:
        return ""


def _check_waitinglist(row):
    clsID = session.clsID
    check = db.classes_waitinglist(auth_customer_id=row.id, classes_id=clsID)
    date = session.waitinglist_classdate
    date_formatted = date.strftime(DATE_FORMAT)
    if check is None:
        return os_gui.get_button('add',
                          URL("waitinglist_process",
                              vars=dict(customers_id=row.id,
                                        clsID=clsID,
                                        date=date_formatted)),
                          _class='pull-right')
    else:
        return ""


def get_customers_searchform(clsID,
                             date,
                             name,
                             function,
                             showall=True,
                             placeholder=T('Search...')):
    date_formatted = date.strftime(DATE_FORMAT)
    form = SQLFORM.factory(
        Field('name', default=name, label=T("")),
        submit_button='Search',
        )
    form.element('input[name=name]')['_placeholder'] = placeholder
    form.element('input[name=name]')['_autocomplete'] = 'off'

    if showall:
        button_all = A(T("Show all"),
                       _href=URL(function,
                                 vars=dict(clsID=clsID, date=date_formatted)),
                       _class='btn btn-default')
    else:
        button_all = ''

    form = DIV(form.custom.begin,
               form.custom.widget.name,
               form.custom.submit,
               button_all,
               form.custom.end,
               _class='left form_inline',
               _id='classes_customers_searchform')

    return form


def attendance_list_classcards_count_classes(row):
    '''
        Returns classes remaining for a class card
    '''
    try:
        ccdID = row.customers_classcards.id
    except AttributeError:
        ccdID = row.id

    ccdh = ClasscardsHelper()

    return ccdh.get_classes_remaining(ccdID)

### helpers end ###

def class_get_back(var=None):
    '''
        General back button for class edit pages
    '''
    if session.classes_attendance_back == 'reports_teacher_classes':
        url = URL('reports', 'teacher_classes')
    elif session.classes_attendance_back == 'reports_attendance_classes':
        url = URL('reports', 'attendance_classes')
    elif session.classes_attendance_back == 'classes_schedule':
        url = URL('schedule')
    else: # catch all
        url = URL('schedule')

    return os_gui.get_button('back', url)


def attendance_count_classes(customers_id, date):
    try:
        date = datestr_to_python(DATE_FORMAT, date)
    except:
        pass
    iso_year = date.isocalendar()[0]
    iso_week = date.isocalendar()[1]
    weekstart = iso_to_gregorian(iso_year, iso_week, 1)
    weekend = iso_to_gregorian(iso_year, iso_week, 7)
    query = (db.classes_attendance.auth_customer_id==customers_id) & \
            (db.classes_attendance.ClassDate>=weekstart) & \
            (db.classes_attendance.ClassDate<=weekend)
    return db(query).count()


@auth.requires(auth.has_membership(group_id='Admins') or
               auth.has_permission('delete', 'classes_otc'))
def class_edit_on_date_remove_changes():
    '''
        Remove all changes made to a class
    '''
    clsID = request.vars['clsID']
    date_formatted = request.vars['date']
    cotcID = request.vars['cotcID']

    query = (db.classes_otc.id == cotcID)
    db(query).delete()

    # Clear cache
    cache_clear_classschedule()

    session.flash = T('Removed changes')

    redirect(URL('class_edit_on_date', vars={'clsID':clsID, 'date':date_formatted}))


def class_edit_on_date_get_link_remove(cotcID, clsID, date_formatted):
    '''
        @param cotcID: db.classes_otc.id
        @return: link with confirmation to remove a all changes made to a class
    '''
    if cotcID is None:
        return ''

    onclick = "return confirm('" + \
        T('Do you really want to remove all changes made to the class on this date?') + "');"

    link = A(SPAN(SPAN(_class=os_gui.get_icon('remove')), ' ',
                  T('Remove all changes to this class'),
                  _class='pull-right'),
             _href=URL('class_edit_on_date_remove_changes',
                       vars={'cotcID':cotcID, 'clsID':clsID, 'date':date_formatted}),
             _onclick=onclick,
             _class='red')

    return link


@auth.requires(auth.has_membership(group_id='Admins') or
               auth.has_permission('update', 'classes'))
def class_edit_on_date():
    '''
        Edit class on a selected date without editing other classes in range
    '''
    clsID = request.vars['clsID']
    date_formatted = request.vars['date']
    date = datestr_to_python(DATE_FORMAT, date_formatted)
    response.title = T("Class")
    response.subtitle = get_classname(clsID) + ": " + date_formatted
    response.view = 'general/tabs_menu.html'

    result = class_edit_on_date_get_form(clsID, date_formatted)
    form = result['form']

    cotcID  = result['cotcID']
    link_remove = class_edit_on_date_get_link_remove(cotcID, clsID, date_formatted)

    form_id = "MainForm"
    form_element = form.element('form')
    form['_id'] = form_id

    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = form_id

    submit = form.element('input[type=submit]')

    content = DIV(DIV(H4(T('Edit class on'), ' ', date_formatted),
                      P(T('The changes below will only be applied to the class on'), ' ',
                        date_formatted, '.', _class='grey'), BR(),
                      form,
                      _class='col-md-9'),
                  DIV(link_remove,
                      _class='col-md-3'),
                  _class='row')


    back = SPAN(class_get_back(), classes_get_week_chooser(request.function, clsID, date))
    menu = classes_get_menu(request.function, clsID, date_formatted)

    return dict(content=content,
                back=back,
                menu=menu,
                save=submit)


def class_edit_on_accept(form):
    '''
        :param form: crud form for db.classes_otc
        :return: None
    '''
    cotcID = form.vars.id

    cotc = db.classes_otc(cotcID)

    if cotc.Status == 'cancelled':
        # Cancel reservations and return credits when status is set to cancelled
        ah = AttendanceHelper()
        ah.attendance_cancel_reservations_for_classes([cotc.classes_id], cotc.ClassDate)


def class_edit_on_date_get_form(clsID, date_formatted):
    '''
        Returns add or edit crud form for one time class change
    '''
    date = datestr_to_python(DATE_FORMAT, date_formatted)

    query = (db.classes_otc.classes_id == clsID) & \
            (db.classes_otc.ClassDate == date)
    rows = db(query).select(db.classes_otc.id)

    next_url = URL('class_edit_on_date', vars={'clsID':clsID,
                                               'date':date_formatted})

    cotcID = None
    if len(rows):
        cotcID = rows.first().id

    if cotcID:
        crud.settings.update_next = next_url
        crud.settings.update_onaccept = [cache_clear_classschedule, class_edit_on_accept]
        form = crud.update(db.classes_otc, cotcID)
    else:
        db.classes_otc.classes_id.default = clsID
        db.classes_otc.ClassDate.default = date

        crud.settings.create_next = next_url
        crud.settings.create_onaccept = [cache_clear_classschedule, class_edit_on_accept]
        form = crud.create(db.classes_otc)

    return dict(form=form, cotcID=cotcID)


@auth.requires_login()
def class_add():
    """
        This function shows an add page for a class
    """
    response.title = T("Add a new class")
    response.subtitle = T("")
    response.view = 'general/tabs_menu.html'

    next_url = '/classes/class_teacher_add/?clsID=[id]&wiz=True'
    return_url = URL('schedule')

    crud.messages.submit_button = T("Next")
    crud.messages.record_created = T("Added class")
    crud.settings.create_next = next_url
    crud.settings.create_onaccept = [cache_clear_classschedule]
    form = crud.create(db.classes)

    result = set_form_id_and_get_submit_button(form, 'MainForm')
    form = result['form']
    submit = result['submit']

    back = os_gui.get_button('back', return_url)
    menu = class_add_get_menu(request.function)

    return dict(content=form, save=submit, back=back, menu=menu)


def class_add_get_menu(page):
    '''
        Returns submenu for adding a workshop
    '''
    pages = [ ['class_add', T('1. Class'), "#"],
              ['class_teacher_add', T('2. Teachers'), "#"] ]

    return get_submenu(pages, page, horizontal=True, htype='tabs')


def class_edit_get_menu(page, clsID):
    """
        This function returns a submenu for the class edit pages
    """
    vars = {'clsID':clsID}
    pages = [['class_edit',
              T('Edit'),
              URL('class_edit', vars=vars)],
             ['class_teachers',
              T('Teachers'),
              URL('class_teachers', vars=vars)],
             ['class_prices',
              T('Prices'),
              URL('class_prices', vars=vars)]]

    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('read', 'classes_school_subscriptions'):
        pages.append([ 'class_subscriptions',
                        T('Subscriptions'),
                       URL('class_subscriptions', vars=vars)])

    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('read', 'classes_school_classcards'):
        pages.append([ 'class_classcards',
                        T('Class cards'),
                       URL('class_classcards', vars=vars)])

    return get_submenu(pages, page, horizontal=True, htype='tabs')


@auth.requires_login()
def class_edit():
    """
        This function shows an edit page for a class
        request.args[0] is expected to be the classes_id
    """
    response.title = T("Edit class")
    clsID = request.vars['clsID']
    date_formatted = request.vars['date']
    classname = get_classname(clsID)
    response.subtitle = classname
    response.view = 'general/tabs_menu.html'

    cls = db.classes(clsID)

    return_url = URL('schedule')

    crud.messages.submit_button = T("Save")
    crud.messages.record_updated = T("Updated class")
    crud.messages.record_deleted = T('Deleted class') + ': ' + classname
    crud.settings.update_next = URL(vars=request.vars)
    crud.settings.update_onaccept = [cache_clear_classschedule]
    crud.settings.update_deletable = False
    form = crud.update(db.classes, clsID)

    form_id = "MainForm"
    form_element = form.element('form')
    form['_id'] = form_id

    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = form_id

    submit = form.element('input[type=submit]')

    links = [ A(SPAN(_class="glyphicon glyphicon-duplicate"), " ", T("Duplicate"),
                _href=URL('duplicate_class', args=[clsID]),
                _title=T("Duplicate this class")) ]

    tools = os_gui.get_dropdown_menu(links,
                                     '',
                                     btn_size='',
                                     btn_icon='wrench',
                                     menu_class='pull-right' )


    # check if number of recurring reservations + dropin/trial reservations <= spaces
    spaces = cls.Maxstudents
    res_recur = cls.MaxReservationsRecurring or 0

    warning = ''
    if res_recur > spaces:
        warning = os_gui.get_alert('warning',
            SPAN(B(T('Warning')), BR(),
                 T('Total of max enrollments exceeds available spaces')))

    notification = class_edit_get_notification_no_access_defined(clsID)

    menu = class_edit_get_menu(request.function, clsID)
    back = class_get_back()

    content = DIV(warning, notification, form)

    return dict(content=content,
                menu=menu,
                back=back,
                header_tools = tools,
                save=submit)


def class_edit_get_notification_no_access_defined(clsID):
    '''
        :param clsID: db.classes.id
        :return: warning to notify user that no classcards or subscriptions are defined for access
    '''
    query = (db.classes_school_subscriptions_groups.classes_id == clsID)
    count_subscriptions = db(query).count()

    query = (db.classes_school_classcards_groups.classes_id == clsID)
    count_classcards = db(query).count()

    count = count_subscriptions + count_classcards
    notification = ''
    if count == 0:
        notification = os_gui.get_alert(
            'info',
            SPAN(B(T('Info')), ' ', BR(),
                   T('No subscription groups or class card groups have been granted access to this class.'))
        )

    return notification


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'classes'))
def class_teachers():
    '''
        Overview with teachers for a class
        request.vars[clsID] is required to be classes_id
    '''
    clsID = request.vars['clsID']
    date_formatted = request.vars['date']

    response.title = T("Edit class")
    classname = get_classname(clsID)
    response.subtitle = classname
    response.view = 'general/tabs_menu.html'

    links = [lambda row: os_gui.get_button('edit', URL('class_teacher_edit',
                                                vars={'cltID':row.id,
                                                      'clsID':clsID,
                                                      'date' :date_formatted}))]

    query = (db.classes_teachers.classes_id == clsID)

    fields = [db.classes_teachers.auth_teacher_id,
              db.classes_teachers.teacher_role,
              db.classes_teachers.auth_teacher_id2,
              db.classes_teachers.teacher_role2,
              db.classes_teachers.Startdate,
              db.classes_teachers.Enddate]

    delete_permission = auth.has_membership(group_id='Admins') or \
                        auth.has_permission('delete', 'classes_teachers')

    grid = SQLFORM.grid(query,
                        fields=fields,
                        links=links,
                        details=False,
                        searchable=False,
                        deletable=delete_permission,
                        csv=False,
                        create=False,
                        editable=False,
                        ondelete=cache_clear_classschedule,
                        orderby=~db.classes_teachers.Startdate,
                        field_id=db.classes_teachers.id,
                        ui = grid_ui)
    grid.element('.web2py_counter', replace=None) # remove the counter
    grid.elements('span[title=Delete]', replace=None) # remove text from delete button

    add_permission = (auth.has_membership(group_id='Admins') or
                      auth.has_permission('create', 'classes_teachers'))
    if add_permission:
        add = os_gui.get_button('add', URL('class_teacher_add',
                                           vars={'clsID':clsID,
                                                 'date' :date_formatted}))
    else:
        add = ''

    msg = session.classes_teachers_msg or ''
    menu = class_edit_get_menu(request.function, clsID)
    back = class_get_back()

    content = DIV(msg, grid)

    # reset message panel
    session.classes_teachers_msg = None

    return dict(content=content,
                menu=menu,
                back=back,
                add=add)


@auth.requires_login()
def class_teacher_add():
    """
        This function shows an add page for classes_teachers
    """
    clsID = request.vars['clsID']
    date_formatted = request.vars['date']
    wizzard = True if 'wiz' in request.vars else False

    response.title = T("Add teacher")
    classname = get_classname(clsID)
    response.subtitle = classname

    return_url = class_teachers_add_edit_get_return_url(clsID, date_formatted)

    if wizzard:
        response.view = 'general/tabs_menu.html'
        menu = class_add_get_menu(request.function)
        back = os_gui.get_button('back', return_url)
    else:
        response.view = 'general/only_content.html'
        menu = ''
        back = os_gui.get_button('back', return_url)

    query = (db.classes_teachers.classes_id == clsID)
    teachers_count = db(query).count()
    if teachers_count == 0:
        startdate = db.classes(clsID).Startdate
        enddate = db.classes(clsID).Enddate
        db.classes_teachers.Startdate.default = startdate
        db.classes_teachers.Enddate.default = enddate

    db.classes_teachers.classes_id.default = clsID

    crud.messages.submit_button = T("Save")
    crud.messages.record_created = T("Added teacher")
    crud.settings.create_next = return_url
    crud.settings.create_onaccept = [ class_teachers_check_classtype, cache_clear_classschedule ]
    form = crud.create(db.classes_teachers)

    form_id = "MainForm"
    form_element = form.element('form')
    form['_id'] = form_id

    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = form_id

    submit = form.element('input[type=submit]')

    return dict(content=form, back=back, menu=menu, save=submit)


@auth.requires_login()
def class_teacher_edit():
    """
        This function shows an edit page for a teacher of a class
        request.vars[clsID] is expected to be the classes_id
    """
    response.title = T("Edit teacher")
    cltID = request.vars['cltID']
    clsID = request.vars['clsID']
    date_formatted = request.vars['date']
    classname = get_classname(clsID)
    response.subtitle = classname
    response.view = 'general/only_content.html'

    return_url = class_teachers_add_edit_get_return_url(clsID, date_formatted)

    crud.messages.submit_button = T("Save")
    crud.messages.record_updated = T("Saved")
    crud.settings.update_next = return_url
    crud.settings.update_onaccept = [ class_teachers_check_classtype, cache_clear_classschedule ]
    crud.settings.update_deletable = False
    form = crud.update(db.classes_teachers, cltID)

    result = set_form_id_and_get_submit_button(form, 'MainForm')
    form = result['form']
    submit = result['submit']

    back = os_gui.get_button('back', return_url)

    return dict(content=form,
                save=submit,
                back=back)


def class_teachers_add_edit_get_return_url(clsID, date_formatted):
    '''
        Returns return url for adding or editing a teacher
    '''
    return URL('class_teachers', vars ={'clsID':clsID,
                                        'date' :date_formatted})


def class_teachers_check_classtype(form):
    '''
        Check if this teacher is assigned to this classtype
    '''
    teID = form.vars['auth_teacher_id']
    clsID = request.vars['clsID']

    ctID = db.classes(clsID).school_classtypes_id

    query = (db.teachers_classtypes.auth_user_id == teID) & \
            (db.teachers_classtypes.school_classtypes_id == ctID)

    count = db(query).count()

    if not count:
        # there's no record assigning this teacher to this classtype
        msg =  SPAN(B(T('Note:')), ' ',
                    T("This teacher isn't assigned for this class type."),
                    BR())
        session.classes_teachers_msg = os_gui.get_alert('warning',
                                                        msg,
                                                        icon='info',
                                                        dismissable=True)
    else:
        session.classes_teachers_msg = None


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('delete', 'classes'))
def class_delete():
    '''
        Removed the selected class and redirect to the manage page
        request.args[0] is expected to be the classes_id (clsID)
    '''
    clsID = request.args[0]

    flash_msg = T('Failed to delete class')

    attendance_query = (db.classes_attendance.classes_id == clsID)
    attendance_count = db(attendance_query).count()
    if attendance_count:
        flash_msg = SPAN(SPAN(T('Unable to delete class'), _class='bold'), BR(),
                        T('Customer attendance found for this class.'), ' ',
                        T('To no longer show this class in the schedule, please set an end date.'))
    else:
        # Log deletion
        row = db.classes(clsID)
        db.sys_accounting.insert(
            auth_user_id=auth.user.id,
            table_name='classes',
            record_id=clsID,
            record_data=unicode(row),
            action_name='delete' )

        # No attendance linked to this class
        query = (db.classes.id == clsID)
        result = db(query).delete()

        if result == 1:  # Success
            flash_msg = T('Deleted class')
            cache_clear_classschedule()

    session.flash = flash_msg


    redirect(URL('schedule'))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('create', 'classes'))
def duplicate_class():
    clsID = request.args[0]
    row = db.classes[clsID]

    id = db.classes.insert(school_locations_id=row.school_locations_id,
                           school_classtypes_id=row.school_classtypes_id,
                           Week_day=row.Week_day,
                           Starttime=row.Starttime,
                           Endtime=row.Endtime,
                           Startdate=row.Startdate,
                           Enddate=row.Enddate,
                           Maxstudents=row.Maxstudents,
                           AllowAPI=row.AllowAPI)

    # duplicate teachers
    query = (db.classes_teachers.classes_id == clsID)
    rows = db(query).select(db.classes_teachers.ALL)
    for row in rows:
        db.classes_teachers.insert(
            classes_id       = id,
            auth_teacher_id  = row.auth_teacher_id,
            teacher_role     = row.teacher_role,
            auth_teacher_id2 = row.auth_teacher_id2,
            teacher_role2    = row.teacher_role2,
            Startdate        = row.Startdate,
            Enddate          = row.Enddate
        )

    # duplicate prices
    query = (db.classes_price.classes_id == clsID)
    rows = db(query).select(db.classes_price.ALL)
    for row in rows:
        db.classes_price.insert(
            classes_id          = id,
            Startdate           = row.Startdate,
            Enddate             = row.Enddate,
            Dropin              = row.Dropin,
            tax_rates_id_dropin = row.tax_rates_id_dropin,
            Trial               = row.Trial,
            tax_rates_id_trial  = row.tax_rates_id_trial
        )


    session.flash = T("You are now editing the duplicated class")

    redirect(URL('class_edit', vars={'clsID':id}))


def check_class(form):
    # check if teacher is marked for a class, if he/she isn't display a message.
    cltID = request.vars['school_classtypes_id']
    teID = request.vars['teachers_id']
    teID2 = request.vars['teachers_id2']

    te1_check = db.teachers_classtypes(school_classtypes_id=cltID,
                                       auth_teacher_id=teID)
    if te1_check == None:
        session.flash = T("This teacher is not marked as a teacher for this type of class")
    te2_check = db.teachers_classtypes(school_classtypes_id=cltID,
                                       auth_teacher_id=teID)
    if te2_check == None:
        session.flash = T("This teacher is not marked as a teacher for this type of class")


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'classes'))
def schedule():
    '''
        Main list of classes
    '''
    def get_day(weekday, trend_medium, trend_high):
        '''
            Helper function that returns a dict containing a title for the weekday,
            a date for the class and
            a SQLFORM.grid for a selected day which is within 1 - 7 (ISO standard).
        '''
        title = NRtoDay(weekday)
        date = iso_to_gregorian(int(year), int(week), int(weekday))
        session.schedule_get_day_date = date # used for schedule_get_filled
        date_formatted = date.strftime(DATE_FORMAT)

        cs = ClassSchedule(
            date,
            filter_id_school_classtype = session.schedule_filter_classtype,
            filter_id_school_location = session.schedule_filter_location,
            filter_id_school_level = session.schedule_filter_level,
            filter_id_teacher = session.schedule_filter_teacher,
            filter_id_status = session.schedule_filter_status,
            sorting = session.classes_schedule_sort,
            trend_medium = trend_medium,
            trend_high = trend_high)

        return cs.get_day_table()


        ####### new query end

    session.classes_attendance_back = 'classes_schedule'
    session.schedule_staff_holidays_back = 'classes_schedule'

    ## sort form begin ##
    if 'sort' in request.vars:
        session.classes_schedule_sort = request.vars['sort']
        #cache_clear_classschedule()
    elif session.classes_schedule_sort is None or \
         session.classes_schedule_sort == '':
        # check if we have a default value
        sys_property = 'classes_schedule_default_sort'
        row = db.sys_properties(Property=sys_property)
        if row:
            session.classes_schedule_sort = row.PropertyValue
        else:
            # default to sorting by location
            session.classes_schedule_sort = 'location'

    ## sort form end ##

    if 'schedule_year' in request.vars:
        year = int(request.vars['schedule_year'])
    elif not session.schedule_year is None:
        year = session.schedule_year
    else:
        today = datetime.date.today()
        year = datetime.date.today().year
    session.schedule_year = year
    if 'schedule_week' in request.vars:
        week = int(request.vars['schedule_week'])
    elif not session.schedule_week is None:
        week = session.schedule_week
    else:
        week = today.isocalendar()[1]
        if week == 0:
            week = 1
    session.schedule_week = week
    if not session.schedule_show_date is None:
        date = datestr_to_python(DATE_FORMAT, session.schedule_show_date)
        year = date.year
        week = date.isocalendar()[1]
        session.schedule_show_date = None

    # if we used the jump date box to select a week
    if 'schedule_jump_date' in request.vars:
        jump_date = datestr_to_python(DATE_FORMAT,
                                      request.vars['schedule_jump_date'])
        jump_date_iso = jump_date.isocalendar()
        year = jump_date_iso[0]
        week = jump_date_iso[1]
        session.schedule_year = year
        session.schedule_week = week
    else:
        jump_date = datetime.date.today()

    #lastweek = get_lastweek_year(year)

    ## week selection
    form_week = schedule_get_form_week(year, week)
    # year selection
    form_year = schedule_get_form_year(year)

    # jump to date
    form_jump = schedule_get_form_jump(jump_date)

    current_week = A(T('Current week'),
                     _href=URL('schedule_current_week'),
                     _class='btn btn-default full-width input-margins',
                     _id='schedule_current_week')

    modals = DIV()
    # schedule weekly status
    schedule_status = schedule_get_status(modals)

    # show schedule status

    week_chooser = schedule_get_week_chooser()
    day_chooser = schedule_get_day_chooser()

    ## week selection end ##

    # filter
    if 'classtype' in request.vars:
        # Clear cache
        #cache_clear_classschedule()
        # Set new values
        location = request.vars['location']
        teacher = request.vars['teacher']
        classtype = request.vars['classtype']
        level = request.vars['level']
        status = request.vars['status']
        filter_form = schedule_get_filter_form(request.vars['location'],
                                                request.vars['teacher'],
                                                request.vars['classtype'],
                                                request.vars['level'],
                                                request.vars['status'])
        session.schedule_filter_location = location
        session.schedule_filter_teacher = teacher
        session.schedule_filter_classtype = classtype
        session.schedule_filter_level = level
        session.schedule_filter_status = status
    elif not session.schedule_filter_location is None:
        filter_form = schedule_get_filter_form(session.schedule_filter_location,
                                               session.schedule_filter_teacher,
                                               session.schedule_filter_classtype,
                                               session.schedule_filter_level,
                                               session.schedule_filter_status)
    else:
        filter_form = schedule_get_filter_form()
        session.schedule_filter_location = None
        session.schedule_filter_teacher = None
        session.schedule_filter_classtype = None
        session.schedule_filter_level = None
        session.schedule_filter_status = None

    title = T('Classes')
    response.title = T('Classes')
    response.subtitle = schedule_get_subtitle(year, week)

    db.classes.id.readable = False
    db.classes.Week_day.readable = False
    db.classes.Maxstudents.readable = False

    # Get trend percentages here, so they don't have to be loaded for each day, if permission is granted
    trend_medium = ''
    trend_high = ''
    if (auth.has_membership(group_id='Admins') or
        auth.has_permission('read', 'classes_schedule_set_trend_precentages')):
        trend_medium = get_sys_property('classes_schedule_trend_medium', int)
        trend_high = get_sys_property('classes_schedule_trend_high', int)

    # Get classes for days
    days = dict()
    for day in range(1, 8):
        days[day] = get_day(day, trend_medium, trend_high)

    overlapping_workshops = ''
    if ( auth.has_membership(group_id='Admins') or
         auth.has_permission('update', 'workshops_activities') ):
        overlapping_workshops = DIV(
            LOAD('classes', 'schedule_get_overlapping_workshops.load',
                 vars={'year':year, 'week':week},
                 ajax=True,
                 content=''),
        _class='inline-block pull-right')

    schedule_tools = schedule_get_schedule_tools()
    export = schedule_get_export()

    # add new class
    add = ''
    permission = auth.has_membership(group_id='Admins') or \
                 auth.has_permission('create', 'classes')

    if permission:
        add = os_gui.get_button('add',
                URL('class_add'),
                tooltip=T("Add a new class"),
                _class='pull-right')


    return dict(filter_form=filter_form,
                year=year,
                title=title,
                add=add,
                schedule_tools=schedule_tools,
                export=export,
                overlapping_workshops=overlapping_workshops,
                days=days,
                schedule_status=schedule_status,
                modals=modals,
                week_chooser=week_chooser,
                current_week=current_week,
                day_chooser=day_chooser,
                form_jump=form_jump,
                form_week=form_week,
                form_year=form_year,
                )


def schedule_get_query(weekday, date):
    '''
        Returns the default query for the schedule workshops check
    '''
    return (db.classes.Week_day == weekday) & \
           (db.classes.Startdate <= date) & \
           ((db.classes.Enddate >= date) |
            (db.classes.Enddate == None)) # can't use is here, query doesn't work


def schedule_get_subtitle(year, week):
    '''
        Returns subtitle for schedule
    '''
    return unicode(year) + " " + T("week") + " " + unicode(week)



def schedule_get_form_week(year, week):
    '''
        Returns week selection form for schedule
    '''
    lastweek = get_lastweek_year(year)

    input_week = INPUT(_name='schedule_week',
                       requires=IS_INT_IN_RANGE(1, lastweek),
                       _type='number',
                       _value=week,
                       _min=1,
                       _max=lastweek,
                       _id='no_table_schedule_week')

    submit = INPUT(_type='submit',
                   _value=T('Week'),
                   _class='full-width',
                   _id='no_table_schedule_submit'),

    form = FORM(DIV(DIV(input_week,
                        DIV(submit,
                            _class='input-group-btn'),
                        _class='input-group')),
                _class='form_inline')

    return form


def schedule_get_form_year(year):
    '''
        Returns week selection form for schedule
    '''
    lastweek = get_lastweek_year(year)

    input_year = INPUT(_name='schedule_year',
                       requires=IS_INT_IN_RANGE(1900, 3000),
                       _type='number',
                       _value=year,
                       _min=datetime.date.today().year-25,
                       _max=datetime.date.today().year+25,
                       _id='no_table_schedule_year')

    submit = INPUT(_type='submit',
                   _value=T('Year'),
                   _class='full-width',
                   _id='no_table_schedule_submit'),

    form = FORM(DIV(DIV(input_year,
                        DIV(submit,
                            _class='input-group-btn'),
                        _class='input-group')),
                _class='form_inline')

    return form

# def schedule_get_form_week(year, week):
#     '''
#         Returns week/year selection form for schedule
#     '''
#     lastweek = get_lastweek_year(year)
#
#     input_year = INPUT(_name='schedule_year',
#                        requires=IS_INT_IN_RANGE(1900, 3000),
#                        _type='number',
#                        _value=year,
#                        _min=datetime.date.today().year-25,
#                        _max=datetime.date.today().year+25,
#                        _id='no_table_schedule_year')
#

#
#     form = FORM(DIV(SPAN(T('Y'), _class="input-group-addon"),
#                     input_year,
#                     _class='input-group'),
#                 DIV(SPAN(T('W'), _class="input-group-addon"),
#                     input_week,
#                     _class="input-group"),
#                 INPUT(_type='submit',
#                       _value=T('Go'),
#                       _id='no_table_schedule_submit'),
#                 _id='schedule_form_week')
#
#     # form = FORM(DIV(SPAN(T('Year'), _class='input-group-addon'),
#     #                 input_year,
#     #                 _class='input-group input-margins full-width'),
#     #             DIV(SPAN(T('Week'), _class='input-group-addon'),
#     #                 input_week,
#     #                 _class='input-group input-margins full-width'),
#     #             INPUT(_type='submit',
#     #                   _value=T('Go to week'),
#     #                   _class='full-width',
#     #                   _id='no_table_schedule_submit'),
#     #             _id='schedule_form_week')
#
#     return form


def schedule_get_form_jump(jump_date):
    '''
        Returns a form to jump to a date
    '''
    form_jump = SQLFORM.factory(
                Field('schedule_jump_date', 'date',
                      requires=IS_DATE_IN_RANGE(
                                format=DATE_FORMAT,
                                minimum=datetime.date(1900,1,1),
                                maximum=datetime.date(2999,1,1)),
                      default=jump_date,
                      label=T(""),
                      widget=os_datepicker_widget_small),
                submit_button=T('Go'),
                )

    submit_jump = form_jump.element('input[type=submit]')
    submit_jump['_class'] = 'full-width'

    form_jump = DIV(form_jump.custom.begin,
                    DIV(form_jump.custom.widget.schedule_jump_date,
                        DIV(form_jump.custom.submit,
                            _class='input-group-btn'),
                        _class='input-group'),
                    form_jump.custom.end,
                    _class='form_inline',
                    _id='form_jump')

    return form_jump


def schedule_get_week_chooser(var=None):
    '''
        Returns a week chooser for the schedule
    '''
    year = session.schedule_year
    week = session.schedule_week

    lastweek = get_lastweek_year(year)

    if week == 1:
        prev_week  = lastweek
        prev_year  = year - 1
    else:
        prev_week  = week - 1
        prev_year  = year

    if week == lastweek:
        next_week  = 1
        next_year  = year + 1
    else:
        next_week  = week + 1
        next_year  = year


    previous = A(I(_class='fa fa-angle-left'),
                 _href=URL('schedule_set_week', vars={ 'week' : prev_week,
                                                       'year' : prev_year }),
                 _class='btn btn-default')
    nxt = A(I(_class='fa fa-angle-right'),
            _href=URL('schedule_set_week', vars={ 'week' : next_week,
                                                  'year' : next_year }),
            _class='btn btn-default')

    return DIV(previous, nxt, _class='btn-group', _id='schedule_week_chooser')


def schedule_get_day_chooser(var=None):
    '''
        Links to quickly go to a selected day
    '''
    day_chooser = DIV(_class='schedule_day_chooser text-center grey')

    for i in range (1, 8):
        weekday = NRtoDay(i)
        day = A(weekday[:1], _href='#' + weekday, _title=T('Scroll to') + ' ' + T(weekday))
        day_chooser.append(day)
        day_chooser.append(' ')

    return day_chooser


def schedule_get_schedule_tools(var=None):
    '''
        Returns tools for schedule
    '''
    schedule_tools = []

    # teacher holidays
    permission = auth.has_membership(group_id='Admins') or \
                 auth.has_permission('read', 'teacher_holidays')

    if permission:
        teacher_holidays = A(os_gui.get_fa_icon('fa-sun-o'),
                             T("Staff holidays"),
                             _href=URL('schedule', 'staff_holidays'),
                             _title=T('Open classes & shifts in a range of dates'))
        schedule_tools.append(teacher_holidays)

    # List open classes
    permission = auth.has_membership(group_id='Admins') or \
                 auth.has_permission('read', 'classes_open')

    if permission:
        classes_open = A(os_gui.get_fa_icon('fa-bell-o'),
                              T('All open classes'),
                        _href=URL('classes_open'),
                        _title=T('List open classes'))
        schedule_tools.append(classes_open)


    # Set default sorting
    permission = auth.has_membership(group_id='Admins') or \
                 auth.has_permission('update', 'schedule_set_default_sort')
    if permission:
        set_default_sorting = A(os_gui.get_fa_icon('fa-sort-amount-asc'), ' ',
                                T('Set default sorting'),
                                _href=URL('schedule_set_sort_default'),
                                _title=T('Set default sorting for classes'))
        schedule_tools.append(set_default_sorting)

    # Set trend percentages
    permission = auth.has_membership(group_id='Admins') or \
                 auth.has_permission('update', 'classes_schedule_set_trend_precentages')
    if permission:
        set_trend_percentages = A(os_gui.get_fa_icon('fa-percent'), ' ',
                                  T('Set trend colors'),
                                  _href=URL('schedule_set_trend_percentages'),
                                  _title=T('Set percentage colors for trend column'))
        schedule_tools.append(set_trend_percentages)




    # get menu
    tools = os_gui.get_dropdown_menu(schedule_tools,
                                     '',
                                     btn_size='btn-sm',
                                     btn_icon='wrench',
                                     menu_class='pull-right'
                                     )

    return tools


def schedule_get_export(var=None):
    '''
        Returns export drop down for schedule
    '''
    export_locations = A(os_gui.get_fa_icon('fa-calendar'), T("Schedule"),
                          _href=URL('schedule_export_excel',
                                    vars=dict(year=session.schedule_year,
                                              week=session.schedule_week,
                                              export='locations')))

    links = [ export_locations ]

    export = os_gui.get_dropdown_menu(
        links = links,
        btn_text = '',
        btn_size = 'btn-sm',
        btn_icon = 'download',
        menu_class='pull-right')

    return export


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'workshops_activities'))
def schedule_get_overlapping_workshops():
    '''
        Returns overlapping workhshops for a week in the schedule
    '''
    year = request.vars['year']
    week = request.vars['week']

    overlapping_workshops = ''

    permission = auth.requires(auth.has_membership(group_id='Admins') or \
                 auth.has_permission('update', 'workshops_activities'))
    if permission:
        ow_count = overlapping_workshops_count()
        if ow_count == 0:
            ow_badge = get_badge('success', ow_count)
        else:
            ow_badge = get_badge('default', ow_count)

        overlapping_workshops = DIV(A(T('Overlapping events'),
                                    _href=URL('overlapping_workshops',
                                              vars={'year':year,
                                                    'week':week},
                                              extension='')),
                                    ' ',
                                    ow_badge,
                                    _class='center')

    return dict(content=overlapping_workshops)


def schedule_get_status(modals):
    '''
        Returns status for schedule
    '''
    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('read', 'schedule_classes_status'):

        schedule_status = DIV(DIV(T("Status"), ": ",
                                  SPAN(T("Open"), _class='bold'),
                                  ' ',
                                  _id="classes_schedule_week_status_text"),
                               _class='pull-right')
        row = db.schedule_classes_status(ScheduleYear=session.schedule_year,
                                         ScheduleWeek=session.schedule_week)
        if row:
            if row.Status == 'final':
                status = SPAN(T('Final'), _class='green bold')
            schedule_status = DIV(T('Status') + ': ', status + ' ',
                                  _class='pull-right')

        if auth.has_membership(group_id='Admins') or \
           auth.has_permission('update', 'schedule_classes_status'):
            btn_icon = SPAN(_class='glyphicon glyphicon-pencil')

            modal_content = UL(LI(A(T('Open'),
                      _href=URL('schedule_set_week_status',
                                vars={'status':''})),
                      BR(),
                      T('Teachers are being scheduled')),
                   LI(A(SPAN(T('Final'), _class='green'),
                      _href=URL('schedule_set_week_status',
                                vars={'status':'final'})),
                      BR(),
                      T("Schedule for this week has been set")),
                   _class='ul_liststyle_none')
            result = os_gui.get_modal(button_text=XML(btn_icon),
                                      modal_title=T('Schedule week status'),
                                      modal_content=modal_content,
                                      modal_class='classes_schedule_status',
                                      button_class='btn-sm')
            modals.append(result['modal'])
            btn_status = result['button']

            schedule_status.append(btn_status)
    else:
        schedule_status = ""

    return schedule_status


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'classes'))
def schedule_set_week():
    '''
        Set the session variables for schedule week and year
    '''
    year  = request.vars['year']
    week  = request.vars['week']

    session.schedule_year = int(year)
    session.schedule_week = int(week)

    redirect(URL('schedule'))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'schedule_classes_status'))
def schedule_set_week_status():
    '''
        Function to set the weekly status of the schedule
    '''
    status = request.vars['status']
    row = db.schedule_classes_status(ScheduleYear=session.schedule_year,
                                     ScheduleWeek=session.schedule_week)
    if row:
        if not status == '':
            # update status
            row.Status = status
            row.update_record()
        else:
            # remove record
            query = (db.schedule_classes_status.id == row.id)
            db(query).delete()
    elif not status == '':
        # add status
        db.schedule_classes_status.insert(ScheduleYear=session.schedule_year,
                                          ScheduleWeek=session.schedule_week,
                                          Status=status)

    redirect(URL('schedule'))


@auth.requires_login()
def schedule_get_sort_options():
    '''
        returns a list of sorting options for classes
    '''
    sort_options = [ ['location', T('Location')],
                     ['starttime', T('Start time')] ]

    return sort_options


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'schedule_set_default_sort'))
def schedule_set_sort_default():
    '''
        Displays a page to edit the default sorting order for the schedule
    '''
    response.title = T('Schedule')
    response.subtitle = T('Set default sort order for classes')
    response.view = 'general/only_content.html'

    sort_options = schedule_get_sort_options()
    sort = session.classes_schedule_sort

    form = SQLFORM.factory(
        Field('sort',
            requires=IS_IN_SET(sort_options,
                              zero=None),
            default=sort,
            label=""),
        formstyle='divs',
        submit_button=T('Save'),
        )

    result = set_form_id_and_get_submit_button(form, 'MainForm')
    form = result['form']
    submit = result['submit']


    title = H4(T('Sort classes by'))
    content = form

    return_url = URL('schedule')

    if form.process().accepted:
        session.flash = T('Updated default sort order')
        sort = request.vars['sort']
        session.classes_schedule_sort = sort

        sys_property = 'classes_schedule_default_sort'
        row = db.sys_properties(Property=sys_property)
        if not row:
            db.sys_properties.insert(Property=sys_property,
                                     PropertyValue=sort)
        else:
            row.PropertyValue=sort
            row.update_record()

        # Clear cache
        cache_clear_sys_properties()


        redirect(return_url)

    back = schedule_get_back()
    return dict(content=content,
                save=submit,
                back=back)


def schedule_get_back(var=None):
    '''
        Return to class schedule
    '''
    return os_gui.get_button('back', URL('schedule'))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'classes_schedule_set_trend_precentages'))
def schedule_set_trend_percentages():
    '''
        Set percentages for trend colors
    '''
    response.title = T('Schedule')
    response.subtitle = T('Set trend color percentages')
    response.view = 'general/only_content.html'

    trend_medium = get_sys_property('classes_schedule_trend_medium')
    trend_high = get_sys_property('classes_schedule_trend_high')

    form = SQLFORM.factory(
        Field('medium', 'integer',
              default=trend_medium,
              requires=IS_INT_IN_RANGE(0, 100),
              label=T('Medium %'),
              comment=T('Shows in red if below this percentage and orange if equal or above')),
        Field('high', 'integer',
              default=trend_high,
              requires=IS_INT_IN_RANGE(0, 101),
              label=T('High %'),
              comment = T('Shows trend in green if above this percentage')),
        submit_button=T("Save"),
        separator=' ',
        formstyle='bootstrap3_stacked'
    )

    result = set_form_id_and_get_submit_button(form, 'MainForm')
    form = result['form']
    submit = result['submit']


    if form.accepts(request.vars, session):
        # check medium
        trend_medium = request.vars['medium'].strip()
        row = db.sys_properties(Property='classes_schedule_trend_medium')
        if not row:
            db.sys_properties.insert(Property='classes_schedule_trend_medium',
                                     PropertyValue=trend_medium)
        else:
            row.PropertyValue = trend_medium
            row.update_record()

        # check high
        trend_high = request.vars['high'].strip()
        row = db.sys_properties(Property='classes_schedule_trend_high')
        if not row:
            db.sys_properties.insert(Property='classes_schedule_trend_high',
                                     PropertyValue=trend_high)
        else:
            row.PropertyValue = trend_high
            row.update_record()

        # clear schedule cache
        cache_clear_classschedule()
        # clear schedule trend cache
        cache_clear_classschedule_trend()
        # Clear system properties cache
        cache_clear_sys_properties()
        # User feedback
        session.flash = T('Saved')
        # reload so the user sees how the values are stored in the db now
        redirect(URL())

    content = DIV(DIV(form, _class='col-md-6'),
                  _class="row")

    back = schedule_get_back()


    clear = ''
    if trend_medium or trend_high:
        clear = os_gui.get_button('noicon',
                                  URL('schedule_set_trend_percentages_clear'),
                                  title=T('Clear'),
                                  btn_size='')

    header_tools = SPAN(clear)

    return dict(content=content,
                back=back,
                header_tools=header_tools,
                save=submit)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'classes_schedule_set_trend_precentages'))
def schedule_set_trend_percentages_clear():
    '''
        Clear trend percentages
    '''
    # Clear values in DB
    query = (db.sys_properties.Property == 'classes_schedule_trend_medium') | \
            (db.sys_properties.Property == 'classes_schedule_trend_high')

    db(query).delete()

    # clear schedule cache
    cache_clear_classschedule()
    # clear schedule trend cache
    cache_clear_classschedule_trend()

    # Show message to user
    session.flash = T('Cleared trend color percentages')

    # Go back to where we came from
    redirect(URL('schedule_set_trend_percentages'))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'classes'))
def schedule_current_week():
    session.schedule_week = None
    session.schedule_year = None

    redirect(URL('schedule'))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'classes'))
def _schedule_clear_filter():
    session.schedule_filter_location = None
    session.schedule_filter_teacher = None
    session.schedule_filter_classtype = None
    session.schedule_filter_level = None
    session.schedule_filter_status = None

    #cache_clear_classschedule()

    redirect(URL('schedule'))


def schedule_get_filter_form(school_locations_id='',
                             teachers_id='',
                             school_classtypes_id='',
                             school_levels_id='',
                             status=''):

    ct_query = (db.school_classtypes.Archived == False)
    slo_query = (db.school_locations.Archived == False)
    sle_query = (db.school_levels.Archived == False)

    au_query = (db.auth_user.teacher == True) & \
               (db.auth_user.trashed == False)

    form = SQLFORM.factory(
        Field('location',
            requires=IS_IN_DB(db(slo_query),'school_locations.id', '%(Name)s',
                              zero=T('All locations')),
            default=school_locations_id,
            label=""),
        Field('teacher',
            requires=IS_IN_DB(db(au_query),'auth_user.id',
                              '%(full_name)s',
                              zero=T('All teachers')),
            default=teachers_id,
            label=""),
        Field('classtype',
            requires=IS_IN_DB(db(ct_query),'school_classtypes.id', '%(Name)s',
                              zero=T('All classtypes')),
            default=school_classtypes_id,
            label=""),
        Field('level',
            requires=IS_IN_DB(db(sle_query),'school_levels.id', '%(Name)s',
                              zero=T('All levels')),
            default=school_levels_id,
            label=""),
        Field('status',
            requires=IS_IN_SET(session.class_status,
                               zero=T('All statuses')),
            default=status,
            label=''),
        submit_button=T('Filter'),
        formstyle='divs',
        )

    # sumbit form on change
    selects = form.elements('select')
    for select in selects:
        select.attributes['_onchange'] = "this.form.submit();"

    clear = A(T("Clear"), _class="btn btn-default",
                _href=URL('_schedule_clear_filter'),
                _title=T("Reset filter to default"))

    div = DIV(
        form.custom.begin,
        DIV(form.custom.widget.location,
            _class='col-md-2'),
        DIV(form.custom.widget.teacher,
            _class='col-md-2'),
        DIV(form.custom.widget.classtype,
            _class='col-md-2'),
        DIV(form.custom.widget.level,
            _class='col-md-2'),
        DIV(form.custom.widget.status,
            _class='col-md-2'),
        DIV(DIV(form.custom.submit,
                clear,
                _class="pull-right"),
            _class='col-md-2'),
        form.custom.end,
        _id="schedule_filter_form")

    return div


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('read', 'classes'))
def schedule_export_excel():
    iso_year = request.vars['year']
    iso_week = request.vars['week']
    export = request.vars['export']

    def get_cell_id(col, row):
        '''
            Returns cell id for colums / row
        '''
        col_letter = openpyxl.utils.get_column_letter(col)
        return col_letter + unicode(row)

    def writer_location(locID=None):
        """
        This helper function writes the locations schedule
        """
        teachers_dict = create_teachers_dict()
        locations_dict = create_locations_dict()
        classtypes_dict = create_classtypes_dict()

        for day in range(1,8):
            date = iso_to_gregorian(iso_year, iso_week, day)
            col = day
            dayname = NRtoDay(day)
            c_id = get_cell_id(col, 2)
            ws[c_id] = dayname + " \n" + unicode(date)
            ws[c_id].alignment = alignment

            cs = ClassSchedule(
                date,
                filter_id_school_location = locID
            )

            rows = cs.get_day_rows()

            r = 3
            for i, row in enumerate(rows):
                repr_row = list(rows[i:i + 1].render())[0]

                clsID = row.classes.id

                location = repr_row.classes.school_locations_id
                classtype = repr_row.classes.school_classtypes_id
                teacher = repr_row.classes_teachers.auth_teacher_id
                teacher2 = repr_row.classes_teachers.auth_teacher_id2
                class_data = ''
                if row.classes_otc.Status == 'cancelled':
                    class_data += T("CANCELLED") + "\n"

                time = repr_row.classes.Starttime + " - " + repr_row.classes.Endtime
                if locID is None:
                    class_data += location + "\n" + \
                                  time + " \n" + \
                                  classtype + " \n" + \
                                  teacher.decode('utf-8') + " \n" + \
                                  teacher2.decode('utf-8') + " \n"
                else:
                    class_data += time + " \n" + \
                                  classtype + " \n" + \
                                  teacher.decode('utf-8') + " \n" + \
                                  teacher2.decode('utf-8') + " \n"
                c_id = get_cell_id(col, r)
                ws[c_id] = class_data
                ws[c_id].font = font
                ws[c_id].alignment = alignment
                r += 1

    # Set some general values
    font = openpyxl.styles.Font(size=10)
    alignment = openpyxl.styles.Alignment(wrap_text=True,
                                          shrink_to_fit=True,
                                          vertical='top')

    if export == 'locations':
        # Full schedule
        wb = openpyxl.workbook.Workbook()
        title = "ALL"
        ws = wb.active
        ws.title = title
        ws['A1'] = T("Schedule week") + " " + unicode(iso_week)
        writer_location()
        # schedule by location
        rows = db().select(db.school_locations.id,
                           db.school_locations.Name,
                           orderby=db.school_locations.Name)
        for row in rows:
            location = locations_dict.get(row.id, None)
            title = location.decode('utf-8')[0:30]
            ws = wb.create_sheet(title=title)
            ws['A1'] = "Schedule" + " " + title + " " + \
                       "week " + unicode(iso_week)
            writer_location(row.id)

        # create filestream
        stream = cStringIO.StringIO()

        fname = T("Schedule") + '.xlsx'
        wb.save(stream)
        response.headers['Content-Type']='application/vnd.ms-excel'
        response.headers['Content-disposition']='attachment; filename=' + fname

        return stream.getvalue()


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('create', 'classes_attendance_override'))
def attendance_override():
    '''
        This function shows a page that allows the customer count of attendance to be overridden
        request.args[0] is expected to be the classes_id
        request.args[1] is expected to be the data formatted using DATE_FORMAT
    '''
    clsID = request.vars['clsID']
    date_formatted = request.vars['date']
    date = datestr_to_python(DATE_FORMAT, date_formatted)
    response.title = T("Class")
    response.subtitle = get_classname(clsID) + ": " + date_formatted
    response.view = 'general/tabs_menu.html'

    return_url = URL('attendance_override', vars={'clsID':clsID,
                                                  'date':date_formatted})

    db.classes_attendance_override.classes_id.default = clsID
    db.classes_attendance_override.ClassDate.default = date

    fields = ['Amount']
    row = db.classes_attendance_override(classes_id=clsID, ClassDate=date)
    if not row: # Create new, else update current
        form = SQLFORM(db.classes_attendance_override,
                         fields=fields,
                         submit_button=T("Save"),
                         formstyle='table3cols')
    else:
        form = SQLFORM(db.classes_attendance_override,
                       row,
                       fields=fields,
                       showid=False,
                       submit_button=T("Save"),
                       formstyle='table3cols')

    result = set_form_id_and_get_submit_button(form, 'MainForm')
    form = result['form']
    submit = result['submit']

    if form.accepts(request.vars, session):
        if form.vars['Amount'] == None:
            query = (db.classes_attendance_override.classes_id==clsID) & \
                    (db.classes_attendance_override.ClassDate==date)
            db(query).delete()
        session.flash = T("Saved attendance count")
        redirect(return_url)

    description = DIV( T("Here the attendance count from the attendance list can be overridden."), " ",
        T("By adding a value below, you can set the number of attending customers manually."), " ",
        T("This number will be used instead of the count from the attendance list for generating attendance statistics."))

    content = DIV(DIV(description, BR(), form,
                      _class="col-md-12"),
                  _class='row')

    back = DIV(class_get_back(), submit, classes_get_week_chooser(request.function, clsID, date))

    menu = classes_get_menu(request.function, clsID, date_formatted)

    return dict(content=content, menu=menu, back=back)


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('update', 'classes_waitinglist'))
def waitinglist():
    '''
        lists waitinglist for a class
        request.vars['clsID'] is expected to be classes.id
        request.vars['date']  is expected to be classdate formatted using
            DATE_FORMAT
    '''
    clsID = request.vars['clsID']
    date_formatted = request.vars['date']
    date = datestr_to_python(DATE_FORMAT, date_formatted)
    response.title = T("Class")
    response.subtitle = get_classname(clsID) + ": " + date_formatted
    response.view = 'general/tabs_menu.html'

    title = ''

    query = (db.classes_waitinglist.classes_id==clsID)
    rows = db(query).select(db.auth_user.id,
                            db.auth_user.trashed,
                            db.auth_user.thumbsmall,
                            db.auth_user.birthday,
                            db.auth_user.display_name,
        left = [(db.auth_user.on(db.classes_waitinglist.auth_customer_id==db.auth_user.id))],
        orderby=db.classes_waitinglist.id
        )

    table = TABLE(_class="os-customer_table")
    for row in rows.render():
        paused = ''
        customer = Customer(row.id)
        subscriptions = customer.get_subscriptions_on_date(date)
        if subscriptions:
            cs = subscriptions.first()
            csID = cs.customers_subscriptions.id
            subscription = SPAN(cs.school_subscriptions.Name)

            csh = CustomerSubscriptions(csID)
            paused = csh.get_paused(date)
            if paused:
                paused = T("Yes")
            else:
                paused = ''

        table.append(TR(TD(row.thumbsmall,
                           _class='os-customer_image_td'),
                        TD(SPAN(row.display_name,
                                _class="os-customer_name"),
                            BR(), paused)))

    edit = A(T("Edit waitinglist"), _href=URL('waitinglist_edit', vars=dict(clsID=clsID, date=date_formatted)))

    waitinglist = DIV(DIV(title, table, edit,
                          _class='col-md-12'),
                      _class='row')

    back = DIV(class_get_back(), classes_get_week_chooser(request.function, clsID, date))
    menu = classes_get_menu(request.function, clsID, date_formatted)

    return dict(content=waitinglist,
                back=back,
                menu=menu)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'classes_reservation'))
def reservations():
    '''
        Manage reservations for a class
    '''
    response.title = T("Class")
    session.classes_reserve_back = None
    session.classes_reservation_cancel_next = 'reservations'
    session.customers_load_list_search_name = None
    clsID = request.vars['clsID']
    date = datestr_to_python(DATE_FORMAT, request.vars['date'])
    date_formatted = date.strftime(DATE_FORMAT)
    session.clsID = clsID
    session.register_classdate = date
    session.classes_reserve_back = 'classes'
    response.subtitle = get_classname(clsID) + ": " + date_formatted


    if 'filter' in request.vars:
        session.classes_reservations_filter = request.vars['filter']
    elif session.classes_reservations_filter:
        pass
    else:
        session.classes_reservations_filter = 'this'

    buttons = [ [ 'this', T('This class') ],
                [ 'recurring', T('History') ]]
                # [ 'single', T('Drop in') ],
                # [ 'trial', T('Trial') ] ]
    filter_form = os_gui.get_radio_buttons_form(
        session.classes_reservations_filter,
        buttons)

    name = ''
    form = get_customers_searchform(clsID,
                                    date,
                                    name,
                                    request.function,
                                    showall=False,
                                    placeholder=T('Search to add an enrollment...'))

    content = DIV(LOAD('customers', 'load_list.load',
                       target='reservation_list_customers_list',
                       content=os_gui.get_ajax_loader(message=T("Searching...")),
                       vars={'list_type':'classes_manage_reservation',
                             'items_per_page':10,
                             'clsID':clsID,
                             'date':date_formatted},
                        ajax=True),
                   _id="reservation_list_customers_list",
                   _class="load_list_customers clear")

    # list of reserved customers
    query = (db.classes_reservation.classes_id==clsID)

    spaces = ''
    orderby = ~db.classes_reservation.Startdate
    if session.classes_reservations_filter == 'this':
        query &= (db.classes_reservation.Startdate <= date) & \
                 ((db.classes_reservation.Enddate >= date) |
                  (db.classes_reservation.Enddate == None))
        orderby = db.auth_user.display_name
        spaces = reservations_get_spaces(clsID, date)
    elif session.classes_reservations_filter == 'recurring':
        query &= (db.classes_reservation.SingleClass == False)
        query &= (db.classes_reservation.TrialClass == False)
        orderby = (db.classes_reservation.Enddate |
                   ~db.classes_reservation.Startdate)
    # elif session.classes_reservations_filter == 'single':
    #     query &= (db.classes_reservation.SingleClass == True)
    #     query &= (db.classes_reservation.TrialClass == False)
    # elif session.classes_reservations_filter == 'trial':
    #     query &= (db.classes_reservation.TrialClass == True)

    rows = db(query).select(db.auth_user.id,
                            db.auth_user.trashed,
                            db.auth_user.thumbsmall,
                            db.auth_user.birthday,
                            db.auth_user.display_name,
                            db.classes_reservation.ALL,
                            left=[(db.auth_user.on(db.classes_reservation.auth_customer_id==db.auth_user.id))],
                            orderby=orderby)
    table = TABLE(_class='table table-hover')
    for index, row in enumerate(rows):
        repr_row = list(rows[index:index+1].render())[0]
        crID = row.classes_reservation.id

        buttons = DIV(_class='btn-group pull-right')

        recurring = not row.classes_reservation.SingleClass and \
                    not row.classes_reservation.TrialClass
        if recurring:
            edit = os_gui.get_button('edit_notext',
                                     URL('reservation_edit',
                                         vars={'crID'  : crID,
                                               'clsID' : clsID,
                                               'date'  : date_formatted}))
            buttons.append(edit)

        # cancel button
        cancelled_class = ''
        cancelled = False
        if session.classes_reservations_filter == 'this':
            cancelled = reservation_get_cancelled(crID, date)


        if session.classes_reservations_filter == 'single':
            class_date = row.classes_reservation.Startdate
            cancelled = reservation_get_cancelled(crID, date)

        if cancelled:
                cancelled_class = 'line-through'

        # delete button
        if auth.has_membership(group_id='Admins') or \
           auth.has_permission('delete', 'classes_reservation'):
            delete = os_gui.get_button('delete_notext',
                                          URL('reservation_remove',
                                              vars={'crID'  : crID,
                                                    'clsID' : clsID,
                                                    'date'  : date_formatted}),
                                        tooltip=T('Delete'))
            buttons.append(delete)


        # put everything together
        table.append(TR(TD(repr_row.auth_user.thumbsmall,
                          _class='os-customer_image_td'),
                        TD(row.auth_user.display_name,
                           BR(),
                           SPAN(repr_row.classes_reservation.SingleClass,
                                _class='small_font grey'),
                           _class=cancelled_class),
                        TD(buttons)))

    reserved_header = DIV(DIV(H3(T('Enrollments'), _class='os-no_margin_top'),
                              _class='col-md-4'),
                          DIV(filter_form,
                              _class='col-md-8 no_padding-right'))
    reserved = DIV(spaces, BR(), table)

    export = reservations_get_export(clsID, date_formatted)

    menu = classes_get_menu(request.function, clsID, date_formatted)
    back = DIV(class_get_back(), classes_get_week_chooser(request.function, clsID, date), export)

    return dict(content=content,
                form=form,
                reserved_header=reserved_header,
                reserved=reserved,
                back=back,
                menu=menu,
                left_sidebar_enabled=True)


def reservations_get_export(clsID, date_formatted):
    '''
        Returns dict with export button and bs3 modal containing the links
        to different export options.
    '''
    mailinglist = A((os_gui.get_fa_icon('fa-envelope-o'),
                     T("Mailing list")),
                    _href=URL('reservations_export_mailinglist', vars={'clsID':clsID, 'date':date_formatted}),
                    _class='textalign_left')

    links = [ mailinglist ]

    export = os_gui.get_dropdown_menu(
            links = links,
            btn_text = '',
            btn_icon = 'download',
            btn_size = '',
            menu_class='pull-right' )

    return export


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'classes_reservation'))
def reservations_export_mailinglist():
    '''
        Excel export mailing list
    '''
    clsID = request.vars['clsID']
    date_formatted = request.vars['date']
    date = datestr_to_python(DATE_FORMAT, request.vars['date'])

    ##
    # Create filestream
    ##
    stream = cStringIO.StringIO()

    ##
    # Create Excel workbook
    ##
    wb = openpyxl.workbook.Workbook(write_only=True)
    title = 'Reservations ' + date_formatted
    ws = wb.create_sheet(title=title)

    ##
    # Set header
    ##
    header = [
        'First Name',
        'Last Name',
        'Email'
    ]
    ws.append(header)

    ##
    # Add reservations to worksheet
    ##
    left = [ db.auth_user.on(db.classes_reservation.auth_customer_id == db.auth_user.id) ]
    query = (db.classes_reservation.classes_id == clsID) & \
            (db.classes_reservation.ResType == 'recurring') & \
            (db.classes_reservation.Startdate <= date) & \
            ((db.classes_reservation.Enddate >= date) |
             (db.classes_reservation.Enddate == None))
    rows = db(query).select(db.auth_user.first_name,
                            db.auth_user.last_name,
                            db.auth_user.email,
                            left=left,
                            orderby=db.auth_user.first_name)
    for row in rows:
        ws.append([
            row.first_name,
            row.last_name,
            row.email
        ])

    wb.save(stream)

    fname = T('MailingList.xlsx')

    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-disposition'] = 'attachment; filename=' + fname

    return stream.getvalue()




def reservations_get_spaces(clsID, date):
    '''
        Returns number of spaces for a class based on reservation
    '''
    # get cancelled
    cancelled = reservation_get_spaces_cancelled(clsID, date)
    # get reserved
    reservations = reservation_get_spaces_reserved(clsID, date)

    # get all numbers together
    cls = db.classes(clsID)
    spaces = cls.Maxstudents

    spaces_recur = cls.MaxReservationsRecurring or T('Not set')

    res_recur = reservations['recurring']
    available = spaces - res_recur

    table_summary = TABLE(
        THEAD(TR(TH(T('Spaces')),
           TH(T('Enrollments')),
           TH(T('Available')))),
        TR(TD(spaces_recur),
           TD(res_recur),
           TD(available)),
        _class='table table-condensed'
    )

    warning = ''
    if res_recur > spaces_recur and not cls.MaxReservationsRecurring is None:
        warning = os_gui.get_alert('danger',
            SPAN(B(T('Warning')), BR(),
                 T('Enrollments exceed available spaces')))

    spaces = DIV(warning,
                 table_summary,
                 _class="col-md-12")

    return spaces


def reservation_get_spaces_cancelled(clsID, date):
    '''
        Returns count of cancelled reservations for a class
    '''
    reservations_cancelled = {}

    res_types = ['recurring', 'single', 'trial']
    for res_type in res_types:
        query = (db.classes_reservation.ResType == res_type) & \
                (db.classes_reservation.id ==
                 db.classes_reservation_cancelled.classes_reservation_id) & \
                (db.classes_reservation.classes_id == clsID) & \
                (db.classes_reservation_cancelled.ClassDate == date)
        reservations_cancelled[res_type] = db(query).count()

    total = 0
    for rc in reservations_cancelled:
        total += reservations_cancelled[rc]

    reservations_cancelled['total'] = total

    return reservations_cancelled


def reservation_get_spaces_reserved(clsID, date):
    '''
        Returns count of reservations for a class
    '''
    reservations = {}

    res_types = ['recurring', 'single', 'trial']
    for res_type in res_types:
        query = (db.classes_reservation.ResType == res_type) & \
                (db.classes_reservation.classes_id==clsID) & \
                (db.classes_reservation.Startdate <= date) & \
                ((db.classes_reservation.Enddate >= date) |
                 (db.classes_reservation.Enddate == None))
        reservations[res_type] = db(query).count()

    total = 0
    for reservation in reservations:
        total += reservations[reservation]

    reservations['total'] = total

    return reservations


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('update', 'classes_waitinglist'))
def waitinglist_edit():
    response.title = T("Waitinglist")
    clsID = request.vars['clsID']
    date = datestr_to_python(DATE_FORMAT, request.vars['date'])
    date_formatted = date.strftime(DATE_FORMAT)
    session.waitinglist_classdate = date
    session.clsID = clsID
    response.subtitle = get_classname(clsID)
    response.view = 'general/only_content.html'

    name = ''
    query = (db.auth_user.trashed == False) & \
            (db.auth_user.id > 1) # generic show all active customers query
    if 'name' in request.vars: # check whether a search filter is in use
        if request.vars['name'] != '':
            name = request.vars['name']
            search_name = '%' + request.vars['name'] + '%'
            query &= ((db.auth_user.display_name.like(search_name)))

    form = get_customers_searchform(clsID, date, name, request.function)

    fields = [ db.auth_user.display_name ]

    if session.show_location:
        db.auth_user.school_locations_id.readable = True
        db.auth_user.school_locations_id.writable = True
        if request.user_agent()['is_mobile']:
            pass
        else:
            fields.append(db.auth_user.school_locations_id)
    else:
        db.auth_user.school_locations_id.readable = False
        db.auth_user.school_locations_id.writable = False

    links = [ dict(header=T(''), body=_check_waitinglist)]
    grid = SQLFORM.grid(query,
        fields=fields,
        editable=False,
        details=False,
        deletable=False,
        create=False,
        csv=False,
        searchable=False,
        links=links,
        ui = grid_ui)
    grid.element('.web2py_counter', replace=None) # remove the counter
    grid.elements('span[title=Delete]', replace=None) # remove text from delete button

    # list of waiting customers
    query = (db.classes_waitinglist.classes_id==clsID)
    rows = db(query).select(db.auth_user.id,
                            db.auth_user.trashed,
                            db.auth_user.thumbsmall,
                            db.auth_user.birthday,
                            db.auth_user.display_name,
                            left=[(db.auth_user.on(db.classes_waitinglist.auth_customer_id== db.auth_user.id))],
                            orderby=db.classes_waitinglist.id)

    table = TABLE(_class='table table-hover')
    for row in rows.render():

        # check delete button permissions
        btn_delete = ''

        if auth.has_membership(group_id='Admins') or \
           auth.has_permission('delete', 'classes_waitinglist'):
            btn_delete = os_gui.get_button('delete_notext',
              URL('waitinglist_process',
                  vars=dict(customers_id=row.id,
                            clsID=clsID,
                            date=date_formatted)),
              _class='pull-right')

        table.append(TR(TD(row.thumbsmall,
                          _class='os-customer_image_td'),
                        TD(row.display_name),
                        TD(btn_delete)))

    sidebar_title = H3(T("Waitinglist"))
    onlist = DIV(sidebar_title, table)

    content = DIV(DIV(form, grid, _class='col-md-9'),
                  DIV(onlist, _class='col-md-3'),
                  _class='row')


    back = os_gui.get_button('back', URL('waitinglist',
                                         vars={'clsID':clsID,
                                               'date': date_formatted}),
                      _class='left')

    return dict(back=back, content=content)


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('update', 'classes_attendance'))
def attendance_get_chart_title():
    '''
        This function returns the title for the attendance chart
    '''
    return dict(title=T("Attendance chart") + " " +
                 unicode(session.stats_attendance_year))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'classes_attendance'))
def attendance_set_chart_year():
    '''
        This function is calles as json and sets the attendance year + or - 1
        year. It's part of the back-end for the attendance chart buttons on the
        attendance page.
    '''
    response.view = 'generic.json'
    try:
        session.stats_attendance_year += int(request.vars['year'])
        status = 'success'
        message = T('')
    except:
        status = 'fail'
        message = T('Error changing years')

    chart_title = attendance_get_chart_title()

    return dict(status=status,
                message=message,
                chart_title=chart_title)


def classes_get_week_chooser(function, clsID, date):
    '''
        Returns a week chooser for the schedule
    '''
    delta = datetime.timedelta(days=7)
    date_prev = (date-delta).strftime(DATE_FORMAT)
    date_next = (date+delta).strftime(DATE_FORMAT)

    url_prev = URL(function, vars={'clsID' : clsID,
                                       'date'  : date_prev})
    url_next = URL(function, vars={'clsID' : clsID,
                                       'date'  : date_next})


    previous = A(I(_class='fa fa-angle-left'),
                 _href=url_prev,
                 _class='btn btn-default')
    nxt = A(I(_class='fa fa-angle-right'),
            _href=url_next,
            _class='btn btn-default')

    return DIV(previous, nxt, _class='btn-group pull-right')

    # previous = A(SPAN(_class='glyphicon glyphicon-chevron-left'), ' ',
    #              T("Previous"),
    #              _href=url_prev,
    #              _class='btn_date_chooser')
    # nxt = A(T("Next"), ' ',
    #         SPAN(_class='glyphicon glyphicon-chevron-right'),
    #         _href=url_next,
    #         _class='btn_date_chooser right')
    #
    # return DIV(previous, nxt, _class='overview_date_chooser')


def attendance_get_notice(message):
    '''
        Returns info message to inform user that the selected date isn't
        valid for this class
    '''
    return os_gui.get_alert('info', message, dismissable=False)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'classes_attendance'))
def attendance():
    '''
        This function shows a page of people expected to attend a class
        request.vars['clsID'] is expected to be the classes_id
        request.vars['date']  is expected to be the class date
            formatted using DATE_FORMAT
    '''
    # set classes ID
    if 'clsID' in request.vars:
        clsID = request.vars['clsID']
        session.classes_attendance_clsID = clsID
    elif session.classes_attendance_clsID:
        clsID = session.classes_attendance_clsID

    # set date
    if 'date' in request.vars:
        date_formatted = request.vars['date']
        session.classes_attendance_date_formatted = date_formatted
    elif session.classes_attendance_date_formatted:
        date_formatted = session.classes_attendance_date_formatted

    session.invoices_edit_back = 'classes_attendance'
    session.invoices_payment_add_back = 'classes_attendance'


    date = datestr_to_python(DATE_FORMAT, date_formatted)

    session.classes_attendance_signin_back = 'attendance'
    session.classes_reservation_cancel_next = 'attendance'
    session.classes_attendance_remove_back = None # prevent redirecting back to customers after removing attendance
    session.customers_add_back = 'classes_attendance'

    response.search_available = True
    try:
        response.q = session.customers_load_list_search_name.replace('%', '')
    except AttributeError:
        response.q = ''

    response.title = T("Class")
    response.subtitle = get_classname(clsID) + ": " + date_formatted

    modals = DIV()

    # session variables for stats/attendance
    session.stats_attendance_year = date.year
    session.stats_attendance_clsID = clsID


    ah = AttendanceHelper()

    # attendance = ah.get_checkin_list_customers(clsID,
    #                                            date,
    #                                            pictures=True,
    #                                            manual_enabled=True,
    #                                            reservations_cancel=True,
    #                                            invoices=True,
    #                                            show_notes=True)

    attendance = ah.get_checkin_list_customers_booked(clsID, date)

    ch = CustomersHelper()
    result = ch.get_add_modal(
        button_text   = "Customer",
        button_class  = '',
        redirect_vars = {'clsID' : clsID,
                         'date'  : date_formatted})
    add_customer = result['button']
    modals.append(result['modal'])


    chart_buttons = DIV(SPAN(I(_class='fa fa-angle-left'),
                             _class='btn btn-default btn-sm',
                             _id='previous'),
                        SPAN(I(_class='fa fa-angle-right'),
                             _class='btn btn-default btn-sm',
                             _id='next'),
                        _class='os-attendance_buttons btn-group pull-right')
    chart_header = DIV(chart_buttons, _class='clear os-attendance_chart_title')

    button_text = XML(SPAN(SPAN(_class='glyphicon glyphicon-stats'), ' ',
                           T('Attendance')))
    modal_content = DIV(chart_header,
        XML('<div id="attendance_barchart"> \
        <canvas id="attendance-chart-area" width="870" height="290"></canvas> \
        </div>'))

    result = os_gui.get_modal(button_text=button_text,
                              modal_title=attendance_get_chart_title()['title'],
                              modal_content=modal_content,
                              modal_class='modal_attendance_chart',
                              modal_size='lg',
                              button_class='')

    btn_attendance_chart = result['button']

    modals.append(result['modal'])

    # search
    name = request.vars['name']
    search_results = DIV(LOAD('customers', 'load_list.load',
                              target='attendance_list_customers_list',
                              content=os_gui.get_ajax_loader(message=T("Searching...")),
                              vars={'list_type':'classes_attendance_list',
                                    'items_per_page':10,
                                    'clsID':clsID,
                                    'date':date_formatted},
                              ajax=True),
                         _id="attendance_list_customers_list",
                         _class="load_list_customers clear",
                         _style="display:none;")


    content = DIV(search_results,
                  attendance,
                  modals)

    cls = db.classes(clsID)
    date_error = False

    if not cls.Week_day == date.isoweekday():
        date_error = True
        content = attendance_get_notice(
            T("This class isn't held on this day of the week")
        )
    elif cls.Startdate > date:
        date_error = True
        content = attendance_get_notice(
            T("This class hasn't started yet on this date")
        )
    elif cls.Enddate:
        if cls.Enddate < date:
            date_error = True
            content = attendance_get_notice(
                T("This class has ended on this date")
            )


    session.schedule_show_date = date_formatted
    back = class_get_back()
    week_chooser = DIV(classes_get_week_chooser(request.function, clsID, date), _class='pull-right')

    # session variable to control redirect / back behaviour
    session.classes_attendance_back = None

    attendance_count = attendance_get_count(clsID, date)

    menu = classes_get_menu(request.function, clsID, date_formatted)
    export = attendance_get_export(clsID, date_formatted)
    header_tools = DIV(add_customer, btn_attendance_chart, export,
                _class='pull-right')

    tools = DIV(attendance_count, _class='pull-right')

    if date_error:
        menu = ''
        tools = ''


    return dict(content=content,
                back=back,
                week_chooser=week_chooser,
                menu=menu,
                tools=tools,
                header_tools=header_tools)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'customers_contact'))
def attendance_export_excel_mailinglist():
    '''
        :return: Mailing list for a class
    '''
    clsID = request.vars['clsID']
    date_formatted  = request.vars['date']
    date = datestr_to_python(DATE_FORMAT, date_formatted)

    ah = AttendanceHelper()
    stream = ah.get_checkin_list_customers_email_excel(clsID, date, days=15)

    fname = 'MailingList.xlsx'

    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-disposition'] = 'attachment; filename=' + fname

    return stream.getvalue()


def attendance_get_export(clsID, date_formatted):
    '''
        Returns dict with export button and bs3 modal containing the links
        to different export options.
    '''
    mailinglist = A(SPAN(os_gui.get_fa_icon('fa-envelope-o'), ' ',
                         T("Mailing list")),
                    _href=URL('attendance_export_excel_mailinglist', vars={'clsID': clsID,
                                                                           'date' : date_formatted}),
                    _class='textalign_left')

    links = [ mailinglist ]

    export = os_gui.get_dropdown_menu(
            links = links,
            btn_text = '',
            btn_size = '',
            btn_icon = 'download',
            menu_class='pull-right' )

    return export


def attendance_get_count(clsID, date):
    '''
        :param clsID: db.classes.id
        :param date: date of class
        :return: SPAN with count of attending customers
    '''
    query = (db.classes_attendance.classes_id == clsID) & \
            (db.classes_attendance.ClassDate == date)
    count = db(query).count()

    count_text = 'Customers attending'
    if count == 1:
        count_text = 'Customer attending'

    return SPAN(count, ' ', count_text, _class='grey pull-right')



@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'classes_attendance'))
def attendance_booking_options():
    """
        Page to list booking options for a customer
    """
    cuID = request.vars['cuID']
    clsID = request.vars['clsID']
    date_formatted = request.vars['date']
    date = datestr_to_python(DATE_FORMAT, date_formatted)

    response.title = T("Class")
    response.subtitle = get_classname(clsID) + ": " + date_formatted
    response.view = 'general/tabs_menu.html'

    return_url = URL('attendance', vars={'clsID':clsID, 'date':date_formatted})

    customer = Customer(cuID)
    title = H4(T('Check in options for'), ' ', customer.row.display_name, _class='center')

    complementary_permission = (auth.has_membership(group_id='Admins') or
                                auth.has_permission('complementary', 'classes_attendance'))

    ah = AttendanceHelper()
    content = ah.get_customer_class_booking_options(clsID,
                                                    date,
                                                    customer,
                                                    trial=True,
                                                    complementary=complementary_permission,
                                                    list_type='attendance',
                                                    controller='classes')
    cancel = os_gui.get_button('noicon',
                               URL('attendance', vars={'clsID': clsID, 'date': date_formatted}),
                               title=T('Cancel'),
                               btn_size='')


    menu = classes_get_menu('attendance', clsID, date_formatted)
    back = os_gui.get_button('back', return_url)

    return dict(content=DIV(DIV(title, BR(), content,
                                _class='row'),
                            DIV(DIV(BR(), BR(), cancel, _class='col-md-12 center'),
                                _class='row')),
                menu=menu,
                back=back)


def reservation_get_return_url(clsID, date):
    """
        Returns the redirect URL for a reservation
    """
    if session.classes_reserve_back == 'customers_reservations':
        cuID = session.customers_classes_reservation_add_vars['cuID']
        url = URL('customers', 'classes_reservations',
                  vars={'cuID' : cuID})
    else:
        url = URL('reservations', vars={'clsID' : clsID,
                                        'date'  : date})

    return url


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('update', 'classes_reservation'))
def reservation_remove():
    '''
        Remove reservation
    '''
    crID  = request.vars['crID']
    clsID = request.vars['clsID']
    date_formatted  = request.vars['date']

    query = (db.classes_reservation.id == crID)
    db(query).delete()

    redirect(reservation_get_return_url(clsID, date_formatted))


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('create', 'classes_reservation'))
def reservation_add():
    """
        Add reservation for a customer

        if startdate is not set, a one time reservation for 'date' is assumed
    """
    '''
        Edit page for recurring reservations
    '''
    cuID  = request.vars['cuID']
    clsID = request.vars['clsID']
    date_formatted = request.vars['date']
    date = datestr_to_python(DATE_FORMAT, date_formatted)
    response.title = T("Class")
    response.subtitle = get_classname(clsID) + ": " + date_formatted
    response.view = 'general/only_content.html'

    customer = Customer(cuID)

    db.classes_reservation.classes_id.default = clsID
    db.classes_reservation.auth_customer_id.default = cuID

    # db.classes_reservation.auth_customer_id.readable = False
    # db.classes_reservation.auth_customer_id.writable = False
    # db.classes_reservation.classes_id.readable = False
    # db.classes_reservation.classes_id.writable = False
    # db.classes_reservation.SingleClass.readable = False
    # db.classes_reservation.SingleClass.writable = False
    # db.classes_reservation.TrialClass.readable = False
    # db.classes_reservation.TrialClass.writable = False

    return_url = reservation_get_return_url(clsID, date_formatted)

    crud.messages.submit_button = T("Save")
    crud.messages.record_updated = T("Saved reservation")
    crud.settings.create_onaccept = [cache_clear_classschedule]
    crud.settings.create_next = return_url
    crud.settings.formstyle='bootstrap3_stacked'
    form = crud.create(db.classes_reservation)

    form_element = form.element('form')
    form['_id'] = 'MainForm'

    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = "MainForm"

    submit = form.element('input[type=submit]')

    content = DIV(H4(T('Enroll'), ' ', customer.row.display_name), form)


    back = os_gui.get_button("back", return_url)

    return dict(content=content, back=back, save=submit)



@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'classes_reservation'))
def reservation_edit():
    """
        Edit page for recurring reservations
    """
    crID = request.vars['crID']
    clsID = request.vars['clsID']
    response.subtitle = get_classname(clsID)
    date_formatted = None # when redirected from customers controller we don't have a date
    if not session.classes_reserve_back == 'customers_reservations':
        date_formatted = request.vars['date']
        date = datestr_to_python(DATE_FORMAT, date_formatted)
        response.subtitle += ": " + date_formatted

    response.title = T("Class")
    response.view = 'general/only_content.html'

    reservation = db.classes_reservation(crID)
    customer = Customer(reservation.auth_customer_id)

    return_url = reservation_get_return_url(clsID, date_formatted)

    crud.messages.submit_button = T("Save")
    crud.messages.record_updated = T("Saved reservation")
    crud.settings.update_onaccept = [cache_clear_classschedule]
    crud.settings.update_next = return_url
    crud.settings.update_deletable = False
    crud.settings.formstyle='bootstrap3_stacked'
    form = crud.update(db.classes_reservation, crID)

    form_element = form.element('form')
    form['_id'] = 'MainForm'

    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = "MainForm"

    submit = form.element('input[type=submit]')

    back = os_gui.get_button("back", return_url)

    return dict(content=form, back=back, save=submit)


def reservation_get_cancelled(crID, date):
    '''
        Checks if a reservation is cancelled
    '''
    cancelled = False
    # check if the reservation exists in the database
    query = (db.classes_reservation_cancelled.classes_reservation_id == crID) & \
            (db.classes_reservation_cancelled.ClassDate == date)

    count = db(query).count()
    if count:
        cancelled = True

    return cancelled


def reservation_get_cancel_button(crID, date):
    '''
        Returns cancel button for a reservation
    '''
    date_formatted = date.strftime(DATE_FORMAT)

    permission = auth.has_membership(group_id='Admins') or \
                 auth.has_permission('update', 'classes_reservation_cancelled')

    if not permission:
        cancel_res = ''
    else:
        cancelled = reservation_get_cancelled(crID, date)
        if cancelled:
            cancel_res = os_gui.get_button('cancel_notext',
                           URL('reservation_cancel',
                               vars={'date' : date_formatted,
                                     'crID' : crID}),
                           tooltip=T("Uncancel reservation"))
        else:
            cancel_res = os_gui.get_button('cancel_notext',
                           URL('reservation_cancel',
                               vars={'date' : date_formatted,
                                     'crID' : crID}),
                           tooltip=T("Cancel reservation"))

    return cancel_res


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('update', 'classes_waitinglist'))
def waitinglist_process():
    clsID = request.vars['clsID']
    customers_id = request.vars['customers_id']
    date_formatted = request.vars['date']

    list_check = db.classes_waitinglist(auth_customer_id=customers_id,
                                        classes_id=clsID)
    if list_check is None:
        db.classes_waitinglist.insert(auth_customer_id=customers_id,
                                      classes_id=clsID)
    else:
        query = (db.classes_waitinglist.auth_customer_id==customers_id) & \
                (db.classes_waitinglist.classes_id==clsID)
        db(query).delete()
    redirect(URL('waitinglist_edit', vars=dict(clsID=clsID,
                                               date=date_formatted)))



@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('update', 'classes_attendance'))
def attendance_remove_ajaj():
    '''
        Called as JSON, used to remove the attendance to a class for a customer
    '''
    response.view = 'generic.json'
    customers_id = request.vars['cuID']
    clattID = request.vars['clattID']
    status = 'success'
    message = T('Successfully removed')

    # Get variabled
    clatt = db.classes_attendance(clattID)
    cuID = clatt.auth_customer_id
    clsID = clatt.classes_id
    date = clatt.ClassDate

    attendance_remove_classcard_decrease_classes_taken(cuID, clsID, date)

    query = (db.classes_attendance.id == clattID)
    if not db(query).delete(): # returns 0 when it fails
        status = 'fail'
        message = T("Remove failed")

    return dict(status=status, message=message)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('delete', 'classes_attendance'))
def attendance_remove():
    '''
        Removes a customer from a class
    '''
    clattID = request.vars['clattID']

    clatt = db.classes_attendance(clattID)
    cuID = clatt.auth_customer_id
    clsID = clatt.classes_id
    date_formatted = clatt.ClassDate.strftime(DATE_FORMAT)

    ##
    # Change invoice status to cancelled
    ##
    query = (db.invoices_classes_attendance.classes_attendance_id == clattID)
    rows = db(query).select(db.invoices_classes_attendance.ALL)
    for row in rows:
        invoice = Invoice(row.invoices_id)
        invoice.set_status('cancelled')

    ##
    # Delete attendance record
    ##
    query = (db.classes_attendance.id == clattID)
    db(query).delete()

    # Clear cache to refresh subscription credit count
    cache_clear_customers_subscriptions(cuID)

    # Clear api cache to refresh available spaces
    cache_clear_classschedule_api()


    if clatt.customers_classcards_id:
        # update class count on classcard
        ccdh = ClasscardsHelper()
        ccdh.set_classes_taken(clatt.customers_classcards_id)


    if session.classes_attendance_remove_back == 'customers':
        redirect(URL('customers', 'classes_attendance',vars={'cuID':cuID}))
    else:
        redirect(attendance_sign_in_get_returl_url(clsID, date_formatted, cuID))


@auth.requires(auth.has_membership(group_id='Admins') or
               auth.has_permission('update', 'classes_attendance'))
def attendance_set_status():
    """
         Set status of class booking
    """
    clattID = request.vars['clattID']
    status = request.vars['status']

    clatt = db.classes_attendance(clattID)
    clatt.BookingStatus = status
    clatt.update_record()

    cuID = clatt.auth_customer_id
    clsID = clatt.classes_id
    date_formatted = clatt.ClassDate.strftime(DATE_FORMAT)

    # Clear api cache to refresh available spaces
    cache_clear_classschedule_api()

    redirect(attendance_sign_in_get_returl_url(clsID, date_formatted, cuID))


@auth.requires(auth.has_membership(group_id='Admins') or
               auth.has_permission('update', 'classes_attendance'))
def attendance_sign_in_ajaj():
    """
        Sign customer in to a class (change booking status from 'booked' to 'attending'
    """
    clattID = request.vars['clattID']

    clatt = db.classes_attendance(clattID)
    clatt.BookingStatus = 'attending'
    clatt.update_record()

    #TODO: Do an actual status check

    return dict(status = 'ok')


def class_book():
    '''
        Actually book class
    '''
    cuID = request.vars['cuID']
    csID = request.vars['csID']
    ccdID = request.vars['ccdID']
    clsID = request.vars['clsID']
    dropin = request.vars['dropin']
    trial = request.vars['trial']
    date_formatted = request.vars['date']
    date  = datestr_to_python(DATE_FORMAT, request.vars['date'])

    url_booking_options = URL('classes', 'attendance_booking_options', vars={'clsID':clsID,
                                                                             'date':date_formatted})
    ah = AttendanceHelper()

    if csID:
        result = ah.attendance_sign_in_subscription(cuID,
                                                    clsID,
                                                    csID,
                                                    date,
                                                    credits_hard_limit=True,
                                                    booking_status='attending')

    if ccdID:
        result = ah.attendance_sign_in_classcard(cuID, clsID, ccdID, date, booking_status='attending')

    if request.vars['dropin'] == 'true':
        # Add drop in class to shopping cart
        result = ah.attendance_sign_in_dropin(cuID, clsID, date, booking_status='attending')

    if request.vars['trial'] == 'true':
        # Add drop in class to shopping cart
        result = ah.attendance_sign_in_trialclass(cuID, clsID, date, booking_status='attending')

    if request.vars['complementary'] == 'true':
        result = ah.attendance_sign_in_complementary(cuID, clsID, date, booking_status='attending')

    if result['status'] == 'fail':
        session.flash = result['message']

    redirect(attendance_sign_in_get_returl_url(clsID, date_formatted, cuID))


def attendance_sign_in_get_returl_url(clsID, date_formatted, cuID):
    '''
        Return url for sign in functions
    '''
    if session.classes_attendance_signin_back == 'attendance_list':
        url = URL('attendance_list',
                   vars={'clsID': clsID,
                         'date' : date_formatted},
                   extension='html')
    elif session.classes_attendance_signin_back == 'attendance':
        url = URL('attendance',
                   vars={'clsID': clsID,
                         'date' : date_formatted},
                   extension='html')
    elif session.classes_attendance_signin_back == 'cu_classes_attendance':
        url = URL('customers', 'classes_attendance',
                   vars={'cuID':cuID},
                   extension='html')
    elif session.classes_attendance_signin_back == 'self_checkin':
        url = URL('selfcheckin', 'checkin',
                  vars={'clsID': clsID,
                        'date' : date_formatted},
                  extension='')

    return url


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'classes_attendance'))
def attendance_list_classcards():
    customers_id = request.vars['cuID']
    clsID = request.vars['clsID']
    date_formatted = request.vars['date']
    date = datestr_to_python(DATE_FORMAT, date_formatted)
    customer = Customer(customers_id)
    response.title = customer.get_name()
    response.subtitle = T('Which card would you like to use to check-in?')

    if session.classes_attendance_signin_back == 'self_checkin':
        response.view = 'selfcheckin/default.html'

    db.customers_classcards.auth_customer_id.readable=False
    db.customers_classcards.Note.readable=False
    headers = {'customers_classcards.id':T("Card")}

    left = [db.school_classcards.on( db.customers_classcards.school_classcards_id == db.school_classcards.id)]

    fields = [
        db.customers_classcards.id,
        db.customers_classcards.school_classcards_id,
        db.customers_classcards.Startdate,
        db.customers_classcards.Enddate
    ]


    links = [dict(header=T('Classes remaining'),
                  body=attendance_list_classcards_count_classes),
             lambda row: A(SPAN(_class="glyphicon glyphicon-ok"),
                           " " + T("This one"),
                           _class="btn btn-default btn-sm",
                           _href=URL('attendance_sign_in_classcard',
                                     vars={'clsID' : clsID,
                                           'ccdID' : row.id,
                                           'cuID'  : customers_id,
                                           'date'  : date_formatted}))]

    delete_permission = auth.has_membership(group_id='Admins') or \
                        auth.has_permission('delete', 'customers_classcards')

    query = (db.customers_classcards.auth_customer_id == customers_id) & \
            (db.customers_classcards.Startdate <= date) & \
            ((db.customers_classcards.Enddate >= date) |
             (db.customers_classcards.Enddate == None)) & \
            ((db.school_classcards.Classes > db.customers_classcards.ClassesTaken) |
             (db.school_classcards.Classes == 0))
    #db.customers_classcards.auth_customer_id.default = customers_id
    grid = SQLFORM.grid(query,
        headers=headers,
        links=links,
        fields=fields,
        create=False,
        editable=False,
        details=False,
        csv=False,
        searchable=False,
        deletable=delete_permission,
        left=left,
        field_id=db.customers_classcards.id,
        ui = grid_ui)
    grid.element('.web2py_counter', replace=None) # remove the counter
    grid.elements('span[title=Delete]', replace=None) # remove text from delete button

    add = ''
    add_perm = auth.has_membership(group_id='Admins') or \
               auth.has_permission('create', 'customers_classcards')
    if add_perm:
        add = os_gui.get_button('add',
                                URL('customers', 'classcard_add',
                                    vars={'cuID' : customers_id,
                                          'clsID': clsID,
                                          'date' : date_formatted}))

    back_url = attendance_sign_in_get_returl_url(clsID,
                                                 date_formatted,
                                                 customers_id)

    back = os_gui.get_button('back', back_url)

    return dict(grid=grid, back=back, header_tools=add, content=grid)


@auth.requires(auth.has_membership(group_id='Admins') or
               auth.has_permission('read', 'customers_notes_teachers'))
def attendance_teacher_notes():
    '''
        List teacher notes for a customer
    '''
    cuID = request.vars['cuID']
    clsID = request.vars['clsID']
    date_formatted = request.vars['date']
    date = datestr_to_python(DATE_FORMAT, date_formatted)

    response.title = T("Class")
    response.subtitle = get_classname(clsID) + ": " + date_formatted
    response.view = 'general/tabs_menu.html'


    return_url = URL('classes', 'attendance', vars={'clsID':clsID, 'date':date_formatted})
    next_url = URL('attendance_teacher_notes', vars={'clsID':clsID,
                                                     'cuID': cuID,
                                                     'date':date_formatted})

    customer = Customer(cuID)

    content = DIV(
        H4(T('Notes for'), ' ', customer.row.display_name),
        attendance_teacher_notes_get_content(cuID, clsID, date_formatted))

    db.customers_notes.auth_customer_id.default = cuID
    db.customers_notes.auth_user_id.default = auth.user.id
    db.customers_notes.TeacherNote.default = True

    crud.messages.submit_button = T("Add note")
    crud.messages.record_created = T("")
    crud.settings.create_next = next_url
    form = crud.create(db.customers_notes)

    form = DIV(
        form.custom.begin,
        form.custom.widget.Injury, LABEL(form.custom.label.Injury),
        form.custom.widget.Note,
        form.custom.submit,
        A(T('Cancel'),
          _href=return_url,
          _class='btn btn-link'),
        form.custom.end,
        _class='direct-chat-input-form'
    )

    content.append(form)

    menu = classes_get_menu('attendance', clsID, date_formatted)
    back = os_gui.get_button('back', URL('classes', 'attendance', vars={'clsID':clsID,
                                                                        'date':date_formatted}))

    return dict(content=content,
                menu=menu,
                back=back)


def attendance_teacher_notes_get_content(cuID, clsID, date_formatted):
    '''
        :param cuID: db.auth_user.id
        :return:
    '''
    query = (db.customers_notes.auth_customer_id == cuID) & \
            (db.customers_notes.TeacherNote == True)
    rows = db(query).select(db.customers_notes.ALL,
                            orderby=~db.customers_notes.NoteDate|~db.customers_notes.NoteTime)

    edit_permission = (auth.has_membership(group_id='Admins') or
                       auth.has_permission('edit', 'customers_notes'))
    delete_permission = (auth.has_membership(group_id='Admins') or
                         auth.has_permission('delete', 'customers_notes'))
    delete_onclick = "return confirm('" + T('Are you sure you want to delete this note?') + "');"

    messages = DIV(_class='direct-chat-messages direct-chat-messages-high')
    for i, row in enumerate(rows):
        repr_row = list(rows[i:i + 1].render())[0]


        edit = ''
        if edit_permission:
            edit = A(T('Edit'),
                     _href=URL('attendance_teacher_note_edit', vars={'cnID': row.id,
                                                                     'cuID': cuID,
                                                                     'clsID': clsID,
                                                                     'date': date_formatted}))

        delete = ''
        if delete_permission:
            delete = A(T('Delete'),
                       _href=URL('attendance_teacher_note_delete', vars={'cnID':row.id,
                                                                         'cuID':cuID,
                                                                         'clsID':clsID,
                                                                         'date':date_formatted}),
                       _onclick=delete_onclick)


        injury = ''
        if row.Injury:
            injury = SPAN(SPAN(T("Injury"), _class="text-red bold"), ' (',
                          SPAN(A(T('Set healed'),
                                 _class='text-green',
                                 _href=URL('attendance_teacher_notes_injury_status',
                                           vars={'cnID':row.id,
                                                 'clsID':clsID,
                                                 'date':date_formatted}))), ') ',
                          _class='direct-chat-scope pull-right grey')


        msg = DIV(
            DIV(SPAN(repr_row.auth_user_id,
                     _class="direct-chat-name pull-left"),
                SPAN(delete,
                     _class="direct-chat-scope pull-right"),
                SPAN(edit,
                     _class="direct-chat-scope pull-right"),
                injury,
                SPAN(repr_row.NoteDate, ' ', repr_row.NoteTime, ' ',
                     _class="direct-chat-timestamp pull-right"),
                _class="direct-chat-info clearfix"
            ),
            IMG(_src=URL('static', 'images/person_inverted_small.png'), _class="direct-chat-img"),
            DIV(XML(row.Note.replace('\n','<br>')), _class="direct-chat-text"),
        _class="direct-chat-msg"
        )

        messages.append(msg)

    return messages


@auth.requires(auth.has_membership(group_id='Admins') or
               auth.has_permission('update', 'customers_notes'))
def attendance_teacher_notes_injury_status():
    '''
        Set or unset the Injury boolean for a customer note
    '''
    cnID = request.vars['cnID']
    clsID = request.vars['clsID']
    date_formatted = request.vars['date']

    cn = db.customers_notes(cnID)
    cn.Injury = not cn.Injury
    cn.update_record()

    redirect(URL('classes', 'attendance_teacher_notes', vars={'clsID':clsID,
                                                              'cuID':cn.auth_customer_id,
                                                              'date':date_formatted}))


@auth.requires(auth.has_membership(group_id='Admins') or
               auth.has_permission('edit', 'customers_notes_teachers'))
def attendance_teacher_note_edit():
    '''
        List teacher notes for a customer
    '''
    cnID = request.vars['cnID']
    cuID = request.vars['cuID']
    clsID = request.vars['clsID']
    date_formatted = request.vars['date']
    date = datestr_to_python(DATE_FORMAT, date_formatted)

    response.title = T("Class")
    response.subtitle = get_classname(clsID) + ": " + date_formatted
    response.view = 'general/tabs_menu.html'


    return_url = URL('classes', 'attendance_teacher_notes', vars={'clsID':clsID,
                                                                  'cuID': cuID,
                                                                  'date':date_formatted})
    next_url = URL('attendance_teacher_notes', vars={'clsID':clsID,
                                                     'cuID': cuID,
                                                     'date':date_formatted})

    customer = Customer(cuID)

    content = DIV(
        H4(T('Notes for'), ' ', customer.row.display_name),
        attendance_teacher_notes_get_content(cuID, clsID, date_formatted),
        SPAN(T("Edit note"), _class='bold text-red'))

    db.customers_notes.auth_customer_id.default = cuID
    db.customers_notes.auth_user_id.default = auth.user.id
    db.customers_notes.TeacherNote.default = True

    crud.messages.submit_button = T("Save")
    crud.messages.record_updated = T("")
    crud.settings.update_next = next_url
    form = crud.update(db.customers_notes, cnID)

    form = DIV(
        form.custom.begin,
        form.custom.widget.Injury, LABEL(form.custom.label.Injury),
        form.custom.widget.Note,
        form.custom.submit,
        A(T('Cancel'),
          _href=return_url,
          _class='btn btn-link'),
        form.custom.end,
        _class='direct-chat-input-form'
    )

    content.append(form)

    menu = classes_get_menu('attendance', clsID, date_formatted)
    back = os_gui.get_button('back', return_url)

    return dict(content=content,
                menu=menu,
                back=back)


@auth.requires(auth.has_membership(group_id='Admins') or
               auth.has_permission('delete', 'customers_notes'))
def attendance_teacher_note_delete():
    '''
        Delete a teacher note
    '''
    cnID = request.vars['cnID']
    cuID = request.vars['cuID']
    clsID = request.vars['clsID']
    date_formatted = request.vars['date']

    query = (db.customers_notes.id == cnID)
    db(query).delete()

    redirect(URL('attendance_teacher_notes', vars={'cuID':cuID,
                                                   'clsID':clsID,
                                                   'date':date_formatted}))


@auth.requires(auth.has_membership(group_id='Admins') or
               auth.has_permission('read', 'classes_open'))
def classes_open():
    '''
        List all open classes. Initially list 25, list 25 more each time
        more is clicked
    '''
    response.title = T('Open classes')
    response.subtitle = ''
    response.view = 'general/only_content.html'

    table = TABLE(TR(TH(''), # status marker
                     TH(T('Date')),
                     TH(T('Location')),
                     TH(T('Class type')),
                     TH(T('Time')),
                     TH(),
                     _class='os-table_header'),
                  _class='os-schedule')

    rows = classes_open_get_rows(TODAY_LOCAL)

    for i, row in enumerate(rows):
        repr_row = list(rows[i:i + 1].render())[0]

        clsID = row.classes.id
        date = row.classes_otc.ClassDate
        date_formatted = repr_row.classes_otc.ClassDate
        result = classes_get_status(clsID, date)
        status = result['status']
        status_marker = TD(result['status_marker'],
                           _class='td_status_marker')

        location = max_string_length(repr_row.classes.school_locations_id, 15)
        classtype = max_string_length(repr_row.classes.school_classtypes_id, 24)
        time = SPAN(repr_row.classes.Starttime, ' - ', repr_row.classes.Endtime)
        teachers = class_get_teachers(clsID, date)

        vars = {'clsID':clsID,
                'date':date_formatted}

        reservations = A(T('Reservations'),
                         _href=URL('reservations', vars=vars))

        status = A(T('Status'), _href=URL('status', vars=vars))


        tools = DIV(_class='os-schedule_links')

        edit = ''
        if auth.has_membership(group_id='Admins') or \
            auth.has_permission('update', 'classes_otc'):
            edit = os_gui.get_button('edit',
                                     URL('class_edit_on_date',
                                         vars={'clsID':clsID,
                                               'date' :date_formatted}),
                                     _class='pull-right')

        row_class = TR(status_marker,
                       TD(date),
                       TD(location),
                       TD(classtype),
                       TD(time),
                       TD(edit),
                       _class='os-schedule_class')


        table.append(row_class)

    back = os_gui.get_button('back', URL('schedule'))

    return dict(content=DIV(BR(), table),
                back=back)


def classes_open_get_rows(from_date):
    '''
        Return rows for classes_open
    '''
    fields = [
        db.classes_otc.id,
        db.classes_otc.ClassDate,
        db.classes_otc.Status,
        db.classes.id,
        db.classes.school_locations_id,
        db.classes.school_classtypes_id,
        db.classes.Starttime,
        db.classes.Endtime
    ]

    query = '''
    SELECT cotc.id,
           cotc.ClassDate,
           cotc.Status,
           cla.id,
           CASE WHEN cotc.school_locations_id IS NOT NULL
                THEN cotc.school_locations_id
                ELSE cla.school_locations_id
                END AS school_locations_id,
           CASE WHEN cotc.school_classtypes_id IS NOT NULL
                THEN cotc.school_classtypes_id
                ELSE cla.school_classtypes_id
                END AS school_classtypes_id,
           CASE WHEN cotc.Starttime IS NOT NULL
                THEN cotc.Starttime
                ELSE cla.Starttime
                END AS Starttime,
           CASE WHEN cotc.Endtime IS NOT NULL
                THEN cotc.Endtime
                ELSE cla.Endtime
                END AS Endtime
    FROM classes_otc cotc
    LEFT JOIN classes cla on cla.id = cotc.classes_id
    WHERE cotc.ClassDate >= '{date}' AND cotc.Status = 'open'
    ORDER BY cotc.ClassDate, Starttime
    '''.format(date=from_date)

    rows = db.executesql(query, fields=fields)

    return rows


def overlapping_workshops_count():
    '''
        Returns a count for overlapping workshops
    '''
    year = session.schedule_year
    week = session.schedule_week

    wsa_ids = []
    for day in range(1, 8):
        date = iso_to_gregorian(year, week, day)
        # get a list of classes
        query = schedule_get_query(date.isoweekday(), date)
        rows = db(query).select(db.classes.ALL)
        for row in rows:
            # check each class for overlapping workshop activities
            waquery = overlapping_workshops_get_query(row, date)
            wsarows = db(waquery).select(db.workshops_activities.id)
            if wsarows:
                for wsarow in wsarows:
                    if not wsarow.id in wsa_ids:
                        wsa_ids.append(wsarow.id)

    return len(wsa_ids)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'workshops_activities'))
def overlapping_workshops():
    '''
        Shows a list of overlapping workshop activities
    '''
    response.title = T("Overlapping events")
    year = request.vars['year']
    week = request.vars['week']
    response.subtitle = schedule_get_subtitle(year, week)

    days = dict()
    for day in range(1, 8):
        date = iso_to_gregorian(year, week, day)

        overlapping = overlapping_workshops_get_overlapping(date)
        days[day] = dict()
        if overlapping:
            days[day]['title'] = NRtoDay(day)
            days[day]['date'] = date.strftime(DATE_FORMAT)
            days[day]['table'] = overlapping
        else:
            days[day]['title'] = ''
            days[day]['date'] = ''
            days[day]['table'] = ''


    back = os_gui.get_button('back', URL('schedule'))

    return dict(days=days, back=back)


def overlapping_workshops_get_overlapping(date):
    '''
        Return a list of overlapping workshops
    '''
    return_value = None

    query = schedule_get_query(date.isoweekday(), date)
    # get classes for a day
    count = 0
    header = TR(TH(T('Event')),
                TH(T('Activity')),
                TH(T('Location')),
                TH(T('Start')),
                TH(T('End')),
                TH())
    table = TABLE(header, _class='table table-hover')
    rows = db(query).select(db.classes.ALL)
    wsa_ids = []
    for row in rows:
        # check each class for overlapping workshop activities
        waquery = overlapping_workshops_get_query(row, date)
        warows = db(waquery).select(db.workshops_activities.ALL)
        for i, warow in enumerate(warows):
            wsaID = warow.id
            if not wsaID in wsa_ids:
                repr_warow = list(rows[i:i + 1].render())[0]
                wsa_ids.append(wsaID)
                count += 1

                wsID = warow.workshops_id
                buttons = overlapping_workshops_get_overlapping_buttons(wsID)
                ws_name = db.workshops(wsID).Name
                tr = TR(TD(ws_name),
                        TD(warow.Activity),
                        TD(repr_warow.school_locations_id),
                        TD(warow.Starttime),
                        TD(warow.Endtime),
                        TD(buttons))
                table.append(tr)

    if count > 0:
        return_value = table

    return return_value


def overlapping_workshops_get_overlapping_buttons(wsID):
    '''
        Returns edit button to link to manage page of workshops
    '''
    btn_group = ''
    permission = auth.has_membership(group_id='Admins') or \
                 auth.has_permission('update', 'workshops')
    if permission:
        btn_group = DIV(_class='btn-group pull-right')
        edit = os_gui.get_button('edit_notext', URL('events', 'activities',
                                             vars={'wsID':wsID}))
        btn_group.append(edit)

    return btn_group


def overlapping_workshops_get_query(row, class_date):
    '''
        Returns a query to check whether a class overlaps with workshop
        activities
        The row parameter is expected to be a record from db.classes
    '''
    starttime = row.Starttime
    endtime = row.Endtime

    # check for location
    query = (row.school_locations_id ==
             db.workshops_activities.school_locations_id)
    # check dates
    query &= (db.workshops_activities.Activitydate == class_date)
    # check times
    query &= (((db.workshops_activities.Starttime <= starttime) &
               (db.workshops_activities.Endtime >= starttime)) |
              ((db.workshops_activities.Starttime <= endtime) &
               (db.workshops_activities.Endtime >= endtime)) |
              ((db.workshops_activities.Starttime == starttime) &
               (db.workshops_activities.Endtime == endtime)) |
              ((db.workshops_activities.Starttime >= starttime) &
               (db.workshops_activities.Endtime <= endtime))
              )

    return query


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'classes'))
def class_prices():
    '''
        List prices for a class
    '''
    clsID = request.vars['clsID']
    date_formatted = request.vars['date']
    response.title = T("Edit class")
    classname = get_classname(clsID)
    response.subtitle = classname
    response.view = 'general/tabs_menu.html'

    links = [lambda row: os_gui.get_button('edit',
                                           URL('class_price_edit',
                                               vars={'clpID':row.id,
                                                     'clsID':clsID,
                                                     'date' :date_formatted}))]

    query = (db.classes_price.classes_id == clsID)

    fields = [ db.classes_price.Startdate,
               db.classes_price.Enddate,
               db.classes_price.Dropin,
               db.classes_price.tax_rates_id_dropin,
               db.classes_price.Trial,
               db.classes_price.tax_rates_id_trial ]

    delete_permission = auth.has_membership(group_id='Admins') or \
                        auth.has_permission('delete', 'classes_price')

    grid = SQLFORM.grid(query,
                        fields=fields,
                        links=links,
                        details=False,
                        searchable=False,
                        deletable=delete_permission,
                        csv=False,
                        create=False,
                        editable=False,
                        orderby=~db.classes_price.Startdate,
                        field_id=db.classes_price.id,
                        ui = grid_ui)
    grid.element('.web2py_counter', replace=None) # remove the counter
    grid.elements('span[title=Delete]', replace=None) # remove text from delete button

    alert_msg = T("Please make sure the new price starts on the first day of a month and the previous price ends on the last day of the month before. ")
    alert_msg += T("Otherwise you might see unexpected results in the stats.")
    alert_icon = SPAN(_class='glyphicon glyphicon-info-sign')
    alert = os_gui.get_alert('info', SPAN(alert_icon, ' ', alert_msg))


    add_permission = (auth.has_membership(group_id='Admins') or
                      auth.has_permission('create', 'classes_price'))
    if add_permission:
        add = os_gui.get_button('add', URL('class_price_add',
                                           vars={'clsID':clsID,
                                                 'date' :date_formatted}))
    else:
        add = ''

    menu = class_edit_get_menu(request.function, clsID)
    back = class_get_back()

    content = DIV(alert, grid)

    return dict(content=content,
                menu=menu,
                back=back,
                add=add)


@auth.requires_login()
def class_price_add():
    """
        This function shows an add page for classes_teachers
    """
    clsID = request.vars['clsID']
    date_formatted = request.vars['date']

    response.title = T("Add price")
    classname = get_classname(clsID)
    response.subtitle = classname

    response.view = 'general/only_content.html'

    # set some default values
    query = (db.classes_price.classes_id == clsID)
    teachers_count = db(query).count()
    if teachers_count == 0:
        startdate = db.classes(clsID).Startdate
        enddate = db.classes(clsID).Enddate
        db.classes_teachers.Startdate.default = startdate
        db.classes_teachers.Enddate.default = enddate

    db.classes_price.classes_id.default = clsID

    return_url = class_prices_add_edit_get_return_url(clsID, date_formatted)

    crud.messages.submit_button = T("Save")
    crud.messages.record_created = T("Saved price")
    crud.settings.create_next = return_url
    form = crud.create(db.classes_price)

    form_id = "MainForm"
    form_element = form.element('form')
    form['_id'] = form_id

    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = form_id

    submit = form.element('input[type=submit]')

    back = os_gui.get_button('back', return_url)

    return dict(content=form, back=back, save=submit)


@auth.requires_login()
def class_price_edit():
    """
        This function shows an edit page for a teacher of a class
        request.vars[clsID] is expected to be the classes_id
    """
    response.title = T("Edit price")
    clpID = request.vars['clpID']
    clsID = request.vars['clsID']
    date_formatted = request.vars['date']
    classname = get_classname(clsID)
    response.subtitle = classname
    response.view = 'general/only_content.html'

    return_url = class_prices_add_edit_get_return_url(clsID, date_formatted)

    crud.messages.submit_button = T("Save")
    crud.messages.record_updated = T("Saved price")
    crud.messages.record_deleted = T('Deleted price for') + ': ' + classname
    crud.settings.update_next = return_url
    crud.settings.update_deletable = False
    form = crud.update(db.classes_price, clpID)

    form_id = "MainForm"
    form_element = form.element('form')
    form['_id'] = form_id

    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = form_id

    submit = form.element('input[type=submit]')

    back = os_gui.get_button('back', return_url)

    return dict(content=form,
                back=back,
                save=submit)


def class_prices_add_edit_get_return_url(clsID, date_formatted):
    '''
        Returns return url for adding or editing a teacher
    '''
    return URL('class_prices', vars ={'clsID':clsID,
                                        'date' :date_formatted})


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'classes_notes'))
def notes():
    '''
        Add notes to a class
    '''
    clsID = request.vars['clsID']
    note_type = request.vars['note_type']
    date_formatted = request.vars['date']
    date = datestr_to_python(DATE_FORMAT, date_formatted)
    response.title = T("Class")
    response.subtitle = get_classname(clsID) + ": " + date_formatted

    response.view = 'general/tabs_menu.html'

    backoffice_class = ''
    teachers_class = ''
    all_class = ''

    active_class = 'web2py-menu-active'

    query = (db.classes_notes.classes_id == clsID) & \
            (db.classes_notes.ClassDate == date)

    if note_type is None:
        if session.classes_notes_filter:
            note_type = session.classes_notes_filter
        else:
            note_type = 'teachers'

    if note_type == 'backoffice':
        db.classes_notes.BackofficeNote.default = True
        query &= (db.classes_notes.BackofficeNote == True)

    if note_type == 'teachers':
        db.classes_notes.TeacherNote.default = True
        query &= (db.classes_notes.TeacherNote == True)

    session.classes_notes_filter = note_type


    notes = UL(_id='os-customers_notes')
    rows = db(query).select(db.classes_notes.ALL,
                            orderby=~db.classes_notes.ClassDate|\
                                    ~db.classes_notes.id)
    for row in rows.render():
        row_note_type = ''
        if row.BackofficeNote:
            row_note_type = T('Back office')
        elif row.TeacherNote:
            row_note_type = T('Teachers')

        vars = {'cnID':row.id,
                'date':date_formatted,
                'clsID':clsID}

        buttons = DIV(_class='btn-group pull-right')
        if auth.has_membership(group_id='Admins') or \
           auth.has_permission('update', 'classes_notes'):
            edit = os_gui.get_button('edit_notext',
                              URL('note_edit', vars=vars),
                              cid=request.cid)
            buttons.append(edit)

        if auth.has_membership(group_id='Admins') or \
           auth.has_permission('delete', 'classes_notes'):

            onclick = "return confirm('" + \
                     T('Do you really want to remove this note?') + "');"

            remove = os_gui.get_button('delete_notext',
                                       URL('note_delete', vars=vars),
                                       onclick=onclick)
            buttons.append(remove)

        notes.append(LI(buttons,
                        SPAN(row.ClassDate,
                             _class='bold'),
                        SPAN(' - ',
                             row.auth_user_id,
                             _class='grey'),
                        BR(),
                        XML(row.Note.replace('\n','<br>')),
                        _id='note_' + unicode(row.id)))


    notes_filter = notes_get_filter_form(session.classes_notes_filter)

    perm = auth.has_membership(group_id='Admins') or \
           auth.has_permission('create', 'classes_notes')
    if perm:
        add = notes_get_add()
        add_title = H4(T('Add a new note'))
    else:
        add = ''
        add_title = ''

    content = DIV(notes_filter, BR())

    # check permissions
    if note_type == 'teachers':
        perm = auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'classes_notes_teachers')
    elif note_type == 'backoffice':
        perm = auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'classes_notes_backoffice')


    if perm:
        content.append(add)
        content.append(notes)

    back = DIV(class_get_back(), classes_get_week_chooser(request.function, clsID, date))
    menu = classes_get_menu(request.function, clsID, date_formatted)

    return dict(content=content,
                back=back,
                menu=menu)


def notes_get_filter_form(page, _class='pull-right'):
    '''
        Backoffice / Teachers view selection form
    '''
    pages = [('backoffice', T('Back office')), ('teachers', T('Teachers'))]

    if page == 'backoffice':
        value = True
        bo_class = 'btn-info'
    else:
        value = False
        bo_class = 'btn-default'
    input_backoffice = INPUT(_type='radio',
                             _name='note_type',
                             _value='backoffice',
                             _id='radio_bo',
                             _onchange="this.form.submit();",
                             value=value)

    if page == 'teachers':
        value = True
        te_class = 'btn-info'
    else:
        value = False
        te_class = 'btn-default'
    input_teachers = INPUT(_type='radio',
                           _name='note_type',
                           _value='teachers',
                           _id='radio_te',
                           _onchange="this.form.submit();",
                           value=value)

    backoffice_text = T('Back office')
    teachers_text = T('Teachers')

    backoffice = LABEL(backoffice_text, input_backoffice,
                    _class='btn btn-sm ' + bo_class)
    teachers = LABEL(teachers_text, input_teachers,
                     _class='btn btn-sm ' + te_class)

    return FORM(DIV(teachers,
                    backoffice,
                    _class='btn-group',
                    **{'_data-toggle': 'buttons'}),
                _class=_class)


@auth.requires_login()
def note_edit():
    '''
        Provides an edit page for a note.
    '''
    clsID = request.vars['clsID']
    date_formatted = request.vars['date']
    date = datestr_to_python(DATE_FORMAT, date_formatted)
    response.title = T("Class")
    response.subtitle = get_classname(clsID) + ": " + date_formatted

    response.view = 'general/tabs_menu.html'

    cnID = request.vars['cnID']
    request.vars.pop('cnID', None)

    note = db.classes_notes(cnID)
    classes_id = note.classes_id
    date = note.ClassDate

    if note.BackofficeNote:
        note_type = 'backoffice'
    elif note.TeacherNote:
        note_type = 'teachers'

    return_url = URL('notes', vars=request.vars)

    crud.messages.submit_button = T("Save")
    crud.messages.record_updated = T('Saved')
    crud.settings.update_next = return_url
    form = crud.update(db.classes_notes, cnID)

    cancel = A(T('Cancel'),
               _href=return_url,
               _class='btn btn-default')


    form = DIV(form.custom.begin,
               form.custom.widget.Note,
               form.custom.submit,
               cancel,
               form.custom.end,
               _class='os-customers_notes_edit clear')

    back =  os_gui.get_button('back', return_url, _class='left')

    content = DIV(B(T('You are now editing the note below')), form)

    back = DIV(class_get_back(), BR(), BR(),
               classes_get_week_chooser(request.function, clsID, date))
    menu = classes_get_menu('notes', clsID, date_formatted)

    return dict(content=content,
                back=back,
                menu=menu,
                left_sidebar_enabled=True)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('delete', 'classes_notes'))
def note_delete():
    '''
        Used to remove a note
    '''
    response.view = 'generic.json'
    cnID = request.vars['cnID']
    date_formatted = request.vars['date']
    clsID = request.vars['clsID']

    request.vars.pop('cnID', None)

    query = (db.classes_notes.id == cnID)
    result = db(query).delete()

    redirect(URL('notes', vars=request.vars))


def notes_get_add(var=None):
    '''
        Provides a page to add a note
        request.vars['note_type'] can be 2 values
            'backoffice' for a backoffice note
            'teachers' for a teachers note
    '''
    note_type = request.vars['note_type']
    clsID = request.vars['clsID']
    date_formatted = request.vars['date']
    date = datestr_to_python(DATE_FORMAT, date_formatted)

    vars = {'clsID':clsID,
            'date':date_formatted}

    return_url = URL('notes', vars=vars)

    db.classes_notes.classes_id.default = clsID
    db.classes_notes.ClassDate.default = date

    crud.messages.submit_button = T("Save")
    crud.messages.record_created = T('')
    crud.settings.create_next = return_url
    form = crud.create(db.classes_notes)

    # form.custom.widget.Note['_class'] += ' form-control'

    form_id = 'add_note'
    form = DIV(form.custom.begin,
               form.custom.widget.Note,
               form.custom.submit,
               form.custom.end,
               _class='os-customers_notes_edit collapse',
               _id=form_id)

    add = A(SPAN(SPAN(_class=os_gui.get_icon('plus')), ' ', T('Add')),
            _href='#' + form_id,
            _role='button',
            _class='btn btn-default btn-sm',
            **{'_data-toggle':'collapse',
               '_aria-expanded':'false',
               '_aria-controls':form_id})

    return DIV(add, form)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'classes_school_subscriptions_groups'))
def class_copy_subscriptions_classcards():
    '''
        :return: page to copy subscription and classcard settings from another class
    '''
    clsID = request.vars['clsID']
    response.title = T("Edit class")
    classname = get_classname(clsID)
    response.subtitle = classname
    response.view = 'general/tabs_menu.html'

    date = TODAY_LOCAL
    classes = ''
    if 'class_date' in request.vars:
        date = datestr_to_python(DATE_FORMAT, request.vars['class_date'])
        classes = class_copy_subscriptions_classcards_get_classes(clsID, date)

    form = class_copy_subscriptions_classcards_get_form(date)

    result = set_form_id_and_get_submit_button(form, 'MainForm')
    form = result['form']
    submit = result['submit']

    content = DIV(
        DIV(H4(T('Copy subscription and class card group from other class')),
            BR(),
            _class="col-md-12"),
        DIV(form, _class="col-md-4"),
        DIV(classes, _class='col-md-12'),
        _class="row",
    )

    menu = class_edit_get_menu('', clsID)
    back = os_gui.get_button('back', URL('class_subscriptions', vars={'clsID':clsID}))

    return dict(content=content,
                menu=menu,
                save=submit,
                back=back)


def class_copy_subscriptions_classcards_get_form(date):
    '''
        :return: copy subscriptins and classcards groups from another class
    '''
    form = SQLFORM.factory(
        Field('class_date', 'date',
              default=date,
              requires=IS_DATE_IN_RANGE(format=DATE_FORMAT,
                                        minimum=datetime.date(1900, 1, 1),
                                        maximum=datetime.date(2999, 1, 1)),
              label=T('List classes on date')
              ),
        submit_button=T('List classes'),
        formstyle='bootstrap3_stacked'
    )

    return form


def class_copy_subscriptions_classcards_get_classes(clsID, date):
    '''
        :return: List of classes to copy settings from
    '''
    header = THEAD(TR(TH(T('Location')),
                      TH(T('Time')),
                      TH(T('Class')),
                      TH(),
                      ))

    table = TABLE(header, _class='table table-striped table-hover')
    cs = ClassSchedule(date, sorting='location')
    for c in cs.get_day_list():
        if c['ClassesID'] == int(clsID):
            continue # Don't list the current class

        copy = os_gui.get_button('noicon',
                                 URL('class_copy_subscriptions_classcards_execute',
                                     vars={'clsID_from': c['ClassesID'],
                                           'clsID_to': clsID}),
                                 title=T('Copy'),
                                 _class='pull-right')

        tr = TR(
            TD(c['Location']),
            TD(c['Starttime']),
            TD(c['ClassType']),
            TD(copy)
        )

        table.append(tr)

    return table


def class_subscriptions_get_tools(clsID):
    '''
        Returns tools for schedule
    '''
    links = []

    vars = {'clsID':clsID}
    #
    # permission = auth.has_membership(group_id='Admins') or \
    #              auth.has_permission('read', 'teacher_holidays')
    #
    # if permission:
    link = A(os_gui.get_fa_icon('fa-copy'),
             T("Copy from other class"),
             _href=URL('class_copy_subscriptions_classcards', vars=vars),
             _title=T('Copy settings from other class'))
    links.append(link)


    # get menu
    tools = os_gui.get_dropdown_menu(links,
                                     '',
                                     btn_size='btn-sm',
                                     btn_icon='wrench',
                                     menu_class='pull-right'
                                     )

    return tools


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'classes_school_subscriptions_groups'))
def class_copy_subscriptions_classcards_execute():
    '''
        Copy subscription and classcard settings from another class
    '''
    clsID_from = request.vars['clsID_from']
    clsID_to = request.vars['clsID_to']

    # Delete all current entries for target
    query = (db.classes_school_subscriptions_groups.classes_id == clsID_to)
    db(query).delete()

    query = (db.classes_school_classcards_groups.classes_id == clsID_to)
    db(query).delete()

    # Get all current entries and copy to selected class
    # Subscriptions
    query = (db.classes_school_subscriptions_groups.classes_id == clsID_from)
    rows = db(query).select(db.classes_school_subscriptions_groups.ALL)
    for row in rows:
        db.classes_school_subscriptions_groups.insert(
            classes_id = clsID_to,
            school_subscriptions_groups_id = row.school_subscriptions_groups_id,
            Enroll = row.Enroll,
            ShopBook = row.ShopBook,
            Attend = row.Attend
        )

    # Class cards
    query = (db.classes_school_classcards_groups.classes_id == clsID_from)
    rows = db(query).select(db.classes_school_classcards_groups.ALL)
    for row in rows:
        db.classes_school_classcards_groups.insert(
            classes_id = clsID_to,
            school_classcards_groups_id = row.school_classcards_groups_id,
            Enroll = row.Enroll,
            ShopBook = row.ShopBook,
            Attend = row.Attend
    )


    session.flash = T('Copied subscription and class card access')
    redirect(URL('class_subscriptions', vars={'clsID':clsID_to}))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'classes_school_subscriptions_groups'))
def class_subscriptions():
    '''
        List subscriptions allowed for this class
    '''
    clsID = request.vars['clsID']
    response.title = T("Edit class")
    classname = get_classname(clsID)
    response.subtitle = classname
    response.view = 'general/tabs_menu.html'

    content = class_subscriptions_get_content(clsID)

    tools = class_subscriptions_get_tools(clsID)

    add = os_gui.get_button('add',
                            URL('class_subscription_group_add', vars={'clsID':clsID}),
                            btn_size='btn-sm')

    back = class_get_back()
    menu = class_edit_get_menu(request.function, clsID)

    return dict(content=content,
                tools=SPAN(tools, add),
                back=back,
                menu=menu)


def class_subscriptions_get_content(clsID):
    '''
        List subscription groups for a class
    '''
    left = [ db.school_subscriptions_groups.on(db.classes_school_subscriptions_groups.school_subscriptions_groups_id ==
                                               db.school_subscriptions_groups.id)]
    query = (db.classes_school_subscriptions_groups.classes_id == clsID)

    rows = db(query).select(db.classes_school_subscriptions_groups.ALL,
                            left=left,
                            orderby=db.school_subscriptions_groups.Name)

    header = THEAD(TR(TH(T('Subscription Group')),
                      TH(T('Enrollment')),
                      TH(T('Book in advance')),
                      TH(T('Attend class')),
                      TH(),
                      ))

    table = TABLE(header, _class='table table-striped table-hover')

    edit_permission = (auth.has_membership(group_id='Admins') or
                       auth.has_permission('update', 'classes_school_subscriptions_groups'))
    delete_permission = (auth.has_membership(group_id='Admins') or
                         auth.has_permission('delete', 'classes_school_subscriptions_groups'))

    delete_onclick = "return confirm('" + \
        T('Are you sure you want to delete this subscription group?') + "');"

    for i, row in enumerate(rows):
        repr_row = list(rows[i:i + 1].render())[0]

        # Edit
        edit = ''
        if edit_permission:
            edit = os_gui.get_button('edit',
                                     URL('class_subscription_group_edit', vars={'cssgID':row.id,
                                                                                  'clsID':clsID}),
                                     _class='pull-right')

        # Delete
        delete = ''
        if delete_permission:
            delete = os_gui.get_button('delete_notext',
                                       URL('class_subscription_group_delete', vars={'cssgID':row.id,
                                                                                    'clsID':clsID}),
                                       onclick=delete_onclick,
                                       _class='pull-right')

        tr = TR(
            TD(repr_row.school_subscriptions_groups_id),
            TD(repr_row.Enroll),
            TD(repr_row.ShopBook),
            TD(repr_row.Attend),
            TD(delete, edit) # Buttons
        )

        table.append(tr)

    return table


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('delete', 'classes_school_subscriptions_groups'))
def class_subscription_group_delete():
    '''
        Delete a subscription group from this class
    '''
    clsID = request.vars['clsID']
    cssgID = request.vars['cssgID']


    query = (db.classes_school_subscriptions_groups.id == cssgID)
    db(query).delete()

    session.flash = T('Deleted')

    redirect(URL('class_subscriptions', vars={'clsID':clsID}))


def class_subscription_add_edit_onaccept(form):
    '''
        Set the "Attend" permission if either "Enroll" or "ShopBook" is set
    '''
    csspID = form.vars.id

    cssp = db.classes_school_subscriptions_groups(csspID)
    if cssp.Enroll or cssp.ShopBook:
        cssp.Attend = True
        cssp.update_record()


@auth.requires_login()
def class_subscription_group_add():
    '''
        Add subscription group for a class
    '''
    clsID = request.vars['clsID']
    response.title = T("Edit class")
    classname = get_classname(clsID)
    response.subtitle = classname
    response.view = 'general/tabs_menu.html'

    # Requires to prevent adding the same group twice
    ids = class_subscription_group_add_get_already_added(clsID)
    query = (~db.school_subscriptions_groups.id.belongs(ids))


    db.classes_school_subscriptions_groups.school_subscriptions_groups_id.requires = IS_IN_DB(
        db(query), 'school_subscriptions_groups.id', '%(Name)s'
    )

    db.classes_school_subscriptions_groups.classes_id.default = clsID
    return_url = URL('class_subscriptions', vars={'clsID':clsID})

    crud.messages.submit_button = T("Save")
    crud.messages.record_created = T("Added subscription")
    crud.settings.create_onaccept = [class_subscription_add_edit_onaccept]
    crud.settings.formstyle = "bootstrap3_stacked"
    crud.settings.create_next = return_url
    form = crud.create(db.classes_school_subscriptions_groups)

    result = set_form_id_and_get_submit_button(form, 'MainForm')
    form = result['form']
    submit = result['submit']

    content = DIV(H4(T("Add subscription group")), form)

    back = os_gui.get_button('back', return_url)
    menu = class_edit_get_menu('class_subscriptions', clsID)

    return dict(content=content,
                back=back,
                save=submit,
                menu=menu)


@auth.requires_login()
def class_subscription_group_edit():
    '''
        Add subscription group for a class
    '''
    clsID = request.vars['clsID']
    cssgID = request.vars['cssgID']
    response.title = T("Edit class")
    classname = get_classname(clsID)
    response.subtitle = classname
    response.view = 'general/tabs_menu.html'

    db.classes_school_subscriptions_groups.school_subscriptions_groups_id.writable = False

    return_url = URL('class_subscriptions', vars={'clsID':clsID})

    crud.messages.submit_button = T("Save")
    crud.messages.record_updated = T("Saved")
    crud.settings.update_onaccept = [class_subscription_add_edit_onaccept]
    crud.settings.formstyle = "bootstrap3_stacked"
    crud.settings.update_next = return_url
    form = crud.update(db.classes_school_subscriptions_groups, cssgID)

    result = set_form_id_and_get_submit_button(form, 'MainForm')
    form = result['form']
    submit = result['submit']

    content = DIV(H4(T("Edit subscription group permissions")), form)

    back = os_gui.get_button('back', return_url)
    menu = class_edit_get_menu('class_subscriptions', clsID)

    return dict(content=content,
                back=back,
                save=submit,
                menu=menu)


def class_subscription_group_add_get_already_added(clsID):
    '''
        :param clsID: db.classes.id
        :return: list of db.school_subscriptions_group already added to this class
    '''
    query = (db.classes_school_subscriptions_groups.classes_id == clsID)
    rows = db(query).select(db.classes_school_subscriptions_groups.school_subscriptions_groups_id)

    ids = []
    for row in rows:
        ids.append(row.school_subscriptions_groups_id)

    return ids


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'classes_school_classcards_groups'))
def class_classcards():
    '''
        List classcards allowed for this class
    '''
    clsID = request.vars['clsID']
    response.title = T("Edit class")
    classname = get_classname(clsID)
    response.subtitle = classname
    response.view = 'general/tabs_menu.html'

    content = class_classcards_get_content(clsID)

    tools = class_subscriptions_get_tools(clsID)

    add = os_gui.get_button('add',
                            URL('class_classcard_group_add', vars={'clsID':clsID}),
                            btn_size='btn-sm')

    back = class_get_back()
    menu = class_edit_get_menu(request.function, clsID)

    return dict(content=content,
                tools=SPAN(tools, add),
                back=back,
                menu=menu)


def class_classcards_get_content(clsID):
    '''
        List classcard groups for a class
    '''
    left = [ db.school_classcards_groups.on(db.classes_school_classcards_groups.school_classcards_groups_id ==
                                               db.school_classcards_groups.id)]
    query = (db.classes_school_classcards_groups.classes_id == clsID)

    rows = db(query).select(db.classes_school_classcards_groups.ALL,
                            left=left,
                            orderby=db.school_classcards_groups.Name)

    header = THEAD(TR(TH(T('classcard Group')),
                      #TH(T('Enrollment')),
                      TH(T('Book in advance')),
                      TH(T('Attend class')),
                      TH(),
                      ))

    table = TABLE(header, _class='table table-striped table-hover')

    edit_permission = (auth.has_membership(group_id='Admins') or
                       auth.has_permission('update', 'classes_school_classcards_groups'))
    delete_permission = (auth.has_membership(group_id='Admins') or
                         auth.has_permission('delete', 'classes_school_classcards_groups'))

    delete_onclick = "return confirm('" + \
        T('Are you sure you want to delete this classcard group?') + "');"

    for i, row in enumerate(rows):
        repr_row = list(rows[i:i + 1].render())[0]

        # Edit
        edit = ''
        if edit_permission:
            edit = os_gui.get_button('edit',
                                     URL('class_classcard_group_edit', vars={'cssgID':row.id,
                                                                                  'clsID':clsID}),
                                     _class='pull-right')

        # Delete
        delete = ''
        if delete_permission:
            delete = os_gui.get_button('delete_notext',
                                       URL('class_classcard_group_delete', vars={'cssgID':row.id,
                                                                                    'clsID':clsID}),
                                       onclick=delete_onclick,
                                       _class='pull-right')

        tr = TR(
            TD(repr_row.school_classcards_groups_id),
            #TD(repr_row.Enroll),
            TD(repr_row.ShopBook),
            TD(repr_row.Attend),
            TD(delete, edit) # Buttons
        )

        table.append(tr)

    return table


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('delete', 'classes_school_classcards_groups'))
def class_classcard_group_delete():
    '''
        Delete a classcard group from this class
    '''
    clsID = request.vars['clsID']
    cssgID = request.vars['cssgID']


    query = (db.classes_school_classcards_groups.id == cssgID)
    db(query).delete()

    session.flash = T('Deleted')

    redirect(URL('class_classcards', vars={'clsID':clsID}))


def class_classcard_add_edit_onaccept(form):
    '''
        Set the "Attend" permission if either "Enroll" or "ShopBook" is set
    '''
    csspID = form.vars.id

    cssp = db.classes_school_classcards_groups(csspID)
    if cssp.Enroll or cssp.ShopBook:
        cssp.Attend = True
        cssp.update_record()


@auth.requires_login()
def class_classcard_group_add():
    '''
        Add classcard group for a class
    '''
    clsID = request.vars['clsID']
    response.title = T("Edit class")
    classname = get_classname(clsID)
    response.subtitle = classname
    response.view = 'general/tabs_menu.html'

    # Requires to prevent adding the same group twice
    ids = class_classcard_group_add_get_already_added(clsID)
    query = (~db.school_classcards_groups.id.belongs(ids))


    db.classes_school_classcards_groups.school_classcards_groups_id.requires = IS_IN_DB(
        db(query), 'school_classcards_groups.id', '%(Name)s'
    )

    db.classes_school_classcards_groups.classes_id.default = clsID
    return_url = URL('class_classcards', vars={'clsID':clsID})

    crud.messages.submit_button = T("Save")
    crud.messages.record_created = T("Added classcard")
    crud.settings.create_onaccept = [class_classcard_add_edit_onaccept]
    crud.settings.formstyle = "bootstrap3_stacked"
    crud.settings.create_next = return_url
    form = crud.create(db.classes_school_classcards_groups)

    result = set_form_id_and_get_submit_button(form, 'MainForm')
    form = result['form']
    submit = result['submit']

    content = DIV(H4(T("Add classcard group")), form)

    back = os_gui.get_button('back', return_url)
    menu = class_edit_get_menu('class_classcards', clsID)

    return dict(content=content,
                back=back,
                save=submit,
                menu=menu)


@auth.requires_login()
def class_classcard_group_edit():
    '''
        Add classcard group for a class
    '''
    clsID = request.vars['clsID']
    cssgID = request.vars['cssgID']
    response.title = T("Edit class")
    classname = get_classname(clsID)
    response.subtitle = classname
    response.view = 'general/tabs_menu.html'

    db.classes_school_classcards_groups.school_classcards_groups_id.writable = False

    return_url = URL('class_classcards', vars={'clsID':clsID})

    crud.messages.submit_button = T("Save")
    crud.messages.record_updated = T("Saved")
    crud.settings.update_onaccept = [class_classcard_add_edit_onaccept]
    crud.settings.formstyle = "bootstrap3_stacked"
    crud.settings.update_next = return_url
    form = crud.update(db.classes_school_classcards_groups, cssgID)

    result = set_form_id_and_get_submit_button(form, 'MainForm')
    form = result['form']
    submit = result['submit']

    content = DIV(H4(T("Edit classcard group permissions")), form)

    back = os_gui.get_button('back', return_url)
    menu = class_edit_get_menu('class_classcards', clsID)

    return dict(content=content,
                back=back,
                save=submit,
                menu=menu)


def class_classcard_group_add_get_already_added(clsID):
    '''
        :param clsID: db.classes.id
        :return: list of db.school_classcards_group already added to this class
    '''
    query = (db.classes_school_classcards_groups.classes_id == clsID)
    rows = db(query).select(db.classes_school_classcards_groups.school_classcards_groups_id)

    ids = []
    for row in rows:
        ids.append(row.school_classcards_groups_id)

    return ids