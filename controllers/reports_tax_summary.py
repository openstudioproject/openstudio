# -*- coding: utf-8 -*-

import datetime
# import operator
import cStringIO
import openpyxl


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'reports_tax_summary'))
def index():
    """
        Main page for reports tax summary controller
    """
    response.title = T("Reports")
    response.subtitle = T("Tax summary")


    from openstudio.os_reports import Reports

    index_process_request_vars()

    reports = Reports()

    data = reports.get_tax_summary_rows(
        session.reports_tax_summary_index_date_from,
        session.reports_tax_summary_index_date_until
    )

    rows = data['rows']
    sum_subtotal = data['sum_subtotal']
    sum_vat = data['sum_vat']
    sum_total = data['sum_total']

    header = THEAD(TR(
        TH("Tax rate"),
        TH(SPAN("Revenue", _class='pull-right')),
        TH(SPAN("Taxes", _class='pull-right')),
        TH()
    ))

    content = TABLE(
        header,
        _class='table table-striped table-hover'
    )

    for i, row in enumerate(rows):
        repr_row = list(rows[i:i + 1].render())[0]

        vars = {}
        if row.invoices_items.tax_rates_id:
            vars = {'tID': row.invoices_items.tax_rates_id}

        details = os_gui.get_button(
            'noicon',
            URL('details', vars=vars),
            title=T("Details"),
            _class='pull-right'
        )

        content.append(TR(
            TD(repr_row.invoices_items.tax_rates_id or "Not specified"),
            TD(SPAN(represent_float_as_amount(row[sum_total]), _class='pull-right')),
            TD(SPAN(represent_float_as_amount(row[sum_vat]), _class='pull-right')),
            TD(details)
        ))


    result = index_get_form(
        session.reports_tax_summary_index_date_from,
        session.reports_tax_summary_index_date_until,
    )
    form = result['form']

    show_current_month = A(
        T("Current month"),
        _href=URL('index_show_current_month'),
        _class='btn btn-default pull-right'
    )

    header_tools = SPAN(
        show_current_month
    )

    export = index_get_export()

    return dict(
        form=result['form_display'],
        content=content,
        submit=result['submit'],
        header_tools=header_tools,
        export=export,
    )


def index_process_request_vars(var=None):
    """
        This function takes the request.vars as a argument and
    """
    from general_helpers import get_last_day_month
    from general_helpers import datestr_to_python

    today = TODAY_LOCAL
    if 'date_from' in request.vars:
        date_from = datestr_to_python(DATE_FORMAT, request.vars['date_from'])
    elif not session.reports_tax_summary_index_date_from is None:
        date_from = session.reports_tax_summary_index_date_from
    else:
        date_from = datetime.date(
            today.year,
            today.month,
            1
        )
    session.reports_tax_summary_index_date_from = date_from

    if 'date_until' in request.vars:
        date_until = datestr_to_python(DATE_FORMAT, request.vars['date_until'])
    elif not session.reports_tax_summary_index_date_until is None:
        date_until = session.reports_tax_summary_index_date_until
    else:
        date_until = get_last_day_month(today)
    session.reports_tax_summary_index_date_until = date_until

    # if 'school_locations_id' in request.vars:
    #     slID = request.vars['school_locations_id']
    # elif not session.reports_tax_summary_index_school_locations_id is None:
    #     slID = session.reports_tax_summary_index_school_locations_id
    # else:
    #     slID = None
    # session.reports_tax_summary_index_school_locations_id = slID

    # session.reports_tax_summary_index = request.function


def index_get_export(var=None):
    """
        Returns dict with export button and bs3 modal containing the links
        to different export options.
    """
    export = ''
    if auth.has_membership(group_id='Admins') or auth.has_permission('read', 'reports_tax_summary'):
        summary = A((os_gui.get_fa_icon('fa-check'),
                     T("Export tax summary")),
                     _href=URL('export_summary'),
                     _class='textalign_left')

        links = [
            summary
        ]

        export = os_gui.get_dropdown_menu(
            links=links,
            btn_text='',
            btn_icon='download',
            btn_size='btn',
            menu_class='pull-right')

    return export


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'reports_tax_summary'))
def index_show_current_month():
    """
    Reset date from & until to this month
    :return: Redirect back to index
    """
    from general_helpers import get_last_day_month

    today = TODAY_LOCAL

    date_from = datetime.date(
        today.year,
        today.month,
        1
    )
    session.reports_tax_summary_index_date_from = date_from

    date_until = get_last_day_month(today)
    session.reports_tax_summary_index_date_until = date_until

    redirect(URL('index'))


def index_get_form(date_from, date_until):
    """
    Get month chooser form for index
    """
    from general_helpers import get_months_list
    from general_helpers import set_form_id_and_get_submit_button

    months = get_months_list()

    form = SQLFORM.factory(
        Field('date_from', 'date', required=True,
            default=date_from,
            requires=IS_DATE_IN_RANGE(format=DATE_FORMAT,
                                      minimum=datetime.date(1900,1,1),
                                      maximum=datetime.date(2999,1,1)),
            represent=represent_date,
            label=T("From date"),
            widget=os_datepicker_widget),
        Field('date_until', 'date', required=True,
            default=date_until,
            requires=IS_DATE_IN_RANGE(format=DATE_FORMAT,
                                      minimum=datetime.date(1900,1,1),
                                      maximum=datetime.date(2999,1,1)),
            represent=represent_date,
            label=T("Until date"),
            widget=os_datepicker_widget),
        # Field('school_locations_id', db.school_locations,
        #       requires=IS_IN_DB(db(loc_query),
        #                         'school_locations.id',
        #                         '%(Name)s',
        #                         zero=T("All locations")),
        #       default=session.reports_tax_summary_index_school_locations_id,
        #       represent=lambda value, row: locations_dict.get(value, T("No location")),
        #       label=T("Location")),
        formstyle='bootstrap3_stacked',
        submit_button=T("Run report")
    )

    result = set_form_id_and_get_submit_button(form, 'MainForm')
    form = result['form']
    submit = result['submit']

    form_display = DIV(
        XML('<form id="MainForm" action="#" enctype="multipart/form-data" method="post">'),
        DIV(LABEL(form.custom.label.date_from),
            form.custom.widget.date_from,
            _class='col-md-6'
        ),
        DIV(LABEL(form.custom.label.date_until),
            form.custom.widget.date_until,
            _class='col-md-6'
        ),
        form.custom.end,
        _class='row'
    )

    return dict(
        form=result['form'],
        submit=result['submit'],
        form_display=form_display
    )


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'reports_tax_summary'))
def export_summary():
    """
    Export summary of tax rates
    :return: Excel worksheet
    """
    from openstudio.os_reports import Reports
    reports = Reports()

    # create filestream
    stream = cStringIO.StringIO()

    wb = openpyxl.workbook.Workbook(write_only=True)
    # Create worksheet
    ws = wb.create_sheet(title="Tax rate summary ")

    data = reports.get_tax_summary_rows(
        session.reports_tax_summary_index_date_from,
        session.reports_tax_summary_index_date_until
    )

    rows = data['rows']
    sum_subtotal = data['sum_subtotal']
    sum_vat = data['sum_vat']
    sum_total = data['sum_total']

    ws.append([
        "Date from",
        "Date until",
        "Tax rate",
        "Revenue",
        "VAT"
    ])

    for i, row in enumerate(rows):
        repr_row = list(rows[i:i + 1].render())[0]

        ws.append([
            session.reports_tax_summary_index_date_from,
            session.reports_tax_summary_index_date_until,
            repr_row.invoices_items.tax_rates_id or "Not specified",
            row[sum_total],
            row[sum_vat]
        ])

    fname = T("TaxSummary") + '.xlsx'
    wb.save(stream)

    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-disposition'] = 'attachment; filename=' + fname

    return stream.getvalue()


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'reports_tax_summary'))
def details():
    """
    Details page to show full list of invoice items for tax rate in period
    :return:
    """
    from general_helpers import max_string_length
    from openstudio.os_reports import Reports

    response.title = T("Reports")
    tID = request.vars['tID']
    response.subtitle = details_subtitle(tID)
    response.view = 'reports_tax_summary/index.html'

    reports = Reports()
    rows = reports.get_tax_summary_detail_rows(
        tID,
        session.reports_tax_summary_index_date_from,
        session.reports_tax_summary_index_date_until
    )
    # Counters
    subtotal = 0
    vat = 0
    total = 0

    # populate rows
    header = THEAD(TR(
        TH(T("Invoice ID")),
        TH(T("Customer")),
        TH(T("Product name")),
        TH(T("Description")),
        TH(T("Quantity")),
        TH(T("Item price")),
        TH(T("Sub total")),
        TH(T("Tax rate")),
        TH(T("VAT")),
        TH(T("Total")),
    ))

    table = TABLE(header, _class='table table-striped table-hover table-condensed small_font')

    for i, row in enumerate(rows):
        repr_row = list(rows[i:i + 1].render())[0]

        subtotal += row.invoices_items.TotalPrice or 0
        vat += row.invoices_items.VAT or 0
        total += row.invoices_items.TotalPriceVAT or 0

        table.append(TR(
            TD(A(row.invoices.InvoiceID,
                 _href=URL('invoices', 'edit', vars={'iID': row.invoices.id}))),
            TD(A(max_string_length(row.auth_user.display_name, 20),
                 _href=URL('customers', 'edit', args=[row.auth_user.id]))),
            TD(max_string_length(row.invoices_items.ProductName, 20)),
            TD(max_string_length(row.invoices_items.Description, 40)),
            TD(row.invoices_items.Quantity),
            TD(repr_row.invoices_items.Price),
            TD(repr_row.invoices_items.TotalPrice),
            TD(repr_row.invoices_items.tax_rates_id),
            TD(repr_row.invoices_items.VAT),
            TD(repr_row.invoices_items.TotalPriceVAT),
        ))


    totals = DIV(
        DIV(LABEL(T("Sub total")), BR(),
            represent_float_as_amount(subtotal),
            _class='col-md-2'),
        DIV(LABEL(T("VAT")), BR(),
            represent_float_as_amount(vat),
            _class='col-md-2'),
        DIV(LABEL(T("Total")), BR(),
            represent_float_as_amount(total),
            _class='col-md-2'),
        _class='row'
    )

    back = os_gui.get_button(
        'back',
        URL('index')
    )

    return dict(
        form = totals,
        content = table,
        back = back,
        export = details_get_export(tID)
    )


def details_subtitle(tID):
    """
    Return details subtitle
    :param tID: db.tax_rates.id
    :return: Text
    """
    tax_rate = db.tax_rates(tID)
    if tax_rate:
        name = tax_rate.Name
    else:
        name = T("Not specified")

    subtitle = SPAN(
        T("Tax summary"), ' ', XML('&bull;'), ' ',
        name, ' ', XML('&bull;'), ' ',
        session.reports_tax_summary_index_date_from.strftime(DATE_FORMAT), ' - ',
        session.reports_tax_summary_index_date_until.strftime(DATE_FORMAT),
    )

    return subtitle


def details_get_export(tID):
    """
        Returns dict with export button and bs3 modal containing the links
        to different export options.
    """
    export = ''
    vars = {}
    if tID:
        vars = {'tID': tID}

    if auth.has_membership(group_id='Admins') or auth.has_permission('read', 'reports_tax_summary'):
        summary = A((os_gui.get_fa_icon('fa-check'),
                     T("Export")),
                     _href=URL('export_details', vars=vars),
                     _class='textalign_left')

        links = [
            summary
        ]

        export = os_gui.get_dropdown_menu(
            links=links,
            btn_text='',
            btn_icon='download',
            btn_size='btn',
            menu_class='pull-right')

    return export


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'reports_tax_summary'))
def export_details():
    """
    Export summary of tax rates
    :return: Excel worksheet
    """
    from openstudio.os_reports import Reports
    reports = Reports()

    tID = request.vars['tID']
    tax_rate = db.tax_rates(tID)
    name = "Not specified"
    if tax_rate:
        name = tax_rate.Name

    # create filestream
    stream = cStringIO.StringIO()

    wb = openpyxl.workbook.Workbook(write_only=True)
    # write the sheet for all mail addresses
    ws = wb.create_sheet(title="Tax rate details")
    ws.append([
        name,
        'from:',
        session.reports_tax_summary_index_date_from,
        'until:',
        session.reports_tax_summary_index_date_until
    ])

    reports = Reports()
    rows = reports.get_tax_summary_detail_rows(
        tID,
        session.reports_tax_summary_index_date_from,
        session.reports_tax_summary_index_date_until
    )

    ws.append([
        "Invoice ID",
        "Customer ID",
        "Customer",
        "Product name",
        "Description",
        "Quantity",
        "Item price",
        "Sub total",
        "Tax rate",
        "VAT",
        "Total",
    ])

    for i, row in enumerate(rows):
        repr_row = list(rows[i:i + 1].render())[0]

        ws.append([
            row.invoices.InvoiceID,
            row.auth_user.id,
            row.auth_user.display_name,
            row.invoices_items.ProductName,
            row.invoices_items.Description,
            row.invoices_items.Quantity,
            row.invoices_items.Price,
            row.invoices_items.TotalPrice,
            repr_row.invoices_items.tax_rates_id,
            row.invoices_items.VAT,
            row.invoices_items.TotalPriceVAT,
        ])

    fname = T("TaxRateDetails") + '.xlsx'
    wb.save(stream)

    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-disposition'] = 'attachment; filename=' + fname

    return stream.getvalue()


# helpers start

# def subscriptions_get_menu(page=None):
#     pages = [
#         (['subscriptions_overview', T('Subscriptions overview'), URL('reports',"subscriptions_overview")]),
#         (['subscriptions_new', T('New subscriptions'), URL('reports',"subscriptions_new")]),
#         (['subscriptions_stopped', T('Stopped subscriptions'), URL('reports',"subscriptions_stopped")]),
#         (['subscriptions_paused', T('Paused subscriptions'), URL('reports',"subscriptions_paused")]),
#         (['subscriptions_alt_prices', T('Alt. prices'), URL('reports',"subscriptions_alt_prices")]),
#         ]
#
#     horizontal = True
#     if request.user_agent()['is_mobile']:
#         horizontal = False
#
#     return os_gui.get_submenu(pages,
#                               page,
#                               horizontal=horizontal,
#                               htype='tabs')
#
