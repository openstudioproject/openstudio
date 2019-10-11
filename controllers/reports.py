# -*- coding: utf-8 -*-

from general_helpers import datestr_to_python
from general_helpers import get_label
from general_helpers import get_submenu
from general_helpers import get_months_list
from general_helpers import get_last_day_month
from general_helpers import get_classname
from general_helpers import User_helpers
from general_helpers import Memo_links
from general_helpers import class_get_teachers
from general_helpers import max_string_length
from general_helpers import iso_to_gregorian
from general_helpers import classes_get_status
from general_helpers import set_form_id_and_get_submit_button

from gluon.tools import prettydate

from openstudio.os_classcards_helper import ClasscardsHelper
from openstudio.os_class import Class
from openstudio.os_class_schedule import ClassSchedule
from openstudio.os_attendance_helper import AttendanceHelper
from openstudio.os_reports import Reports
from openstudio.os_invoice import Invoice
from openstudio.os_invoices import Invoices
from openstudio.os_school_subscription import SchoolSubscription
from openstudio.os_customer_classcard import CustomerClasscard

import datetime
import operator
import io
import openpyxl
from decimal import Decimal


def index():
    """
        Main page for reports controller
    """
    pass


# helpers start

def subscriptions_get_menu(page=None):
    pages = [
        (['subscriptions_overview', T('Overview'), URL('reports',"subscriptions_overview")]),
        (['subscriptions_new', T('New'), URL('reports',"subscriptions_new")]),
        (['subscriptions_online', T('Online'), URL('reports',"subscriptions_online")]),
        (['subscriptions_stopped', T('Stopped'), URL('reports',"subscriptions_stopped")]),
        (['subscriptions_paused', T('Paused'), URL('reports',"subscriptions_paused")]),
        (['subscriptions_alt_prices', T('Alt. prices'), URL('reports',"subscriptions_alt_prices")]),
        ]

    horizontal = True
    if request.user_agent()['is_mobile']:
        horizontal = False

    return os_gui.get_submenu(pages,
                              page,
                              horizontal=horizontal,
                              htype='tabs')


def attendance_get_menu(page=None):
    pages = [
        (['attendance_classes', T('Class Revenue'), URL('reports','attendance_classes')]),
        (['attendance_classtypes', T('Classtypes'), URL('reports','attendance_classtypes')]),
        (['attendance_organizations',
          T('Organizations'),
          URL('reports',"attendance_organizations")]),
        (['attendance_review_requested',
          T('Review requested'),
          URL('reports','attendance_review_requested')]),
        (['attendance_reconcile_later',
          T('Reconcile later'),
          URL('reports','attendance_reconcile_later')]),
    ]


    return get_submenu(pages,
                       page,
                       horizontal=True,
                       htype='tabs')


def get_month_subtitle(month, year):
    """
    :param month: int 1 - 12
    :return: subtitle
    """
    months = get_months_list()
    subtitle = ''
    if year and month:
        for m in months:
            if m[0] == month:
                month_title = m[1]
        subtitle = month_title + " " + str(year)

    return subtitle


def get_form_subtitle(month=None,
                      year=None,
                      function=None,
                      location_filter=False,
                      _class='col-md-4'):
    months = get_months_list()
    subtitle = ''
    if year and month:
        for m in months:
            if m[0] == month:
                month_title = m[1]
        subtitle = month_title + " " + str(year)
    else:
        year = TODAY_LOCAL.year
        month = TODAY_LOCAL.month

    if not location_filter:
        form = SQLFORM.factory(
            Field('month',
                   requires=IS_IN_SET(months, zero=None),
                   default=month,
                   label=T("")),
            Field('year', 'integer',
                  default=year,
                  label=T("")),
            submit_button=T("Run report")
            )
    else:
        loc_query = (db.school_locations.Archived == False)

        form = SQLFORM.factory(
            Field('month',
                   requires=IS_IN_SET(months, zero=None),
                   default=month,
                   label=T("")),
            Field('year', 'integer',
                  default=year,
                  label=T("")),
            Field('school_locations_id', db.school_locations,
                  requires=IS_IN_DB(db(loc_query),
                                    'school_locations.id',
                                    '%(Name)s',
                                    zero=T("All locations")),
                  default=session.reports_subscriptions_school_locations_id,
                  represent=lambda value, row: locations_dict.get(value, T("No location")),
                  label=T("Location")),
            submit_button=T("Run report")
            )
    form.attributes['_name']  = 'form_select_date'
    form.attributes['_class'] = 'overview_form_select_date'

    input_month = form.element('select[name=month]')
    #input_month.attributes['_onchange'] = "this.form.submit();"

    input_year = form.element('input[name=year]')
    #input_year.attributes['_onchange'] = "this.form.submit();"
    input_year.attributes['_type']     = 'number'
    #input_year.attributes['_class']    = 'input_margins'

    form.element('input[name=year]')

    result = set_form_id_and_get_submit_button(form, 'MainForm')
    form = result['form']
    submit = result['submit']

    ## Show current

    if function == 'subscriptions_new' or \
            function == 'subscriptions_online' or \
            function == 'subscriptions_paused' or \
            function == 'subscriptions_stopped' or \
            function == 'subscriptions_overview' or \
            function == 'subscriptions_overview_customers' or \
            function == 'subscriptions_alt_prices':
        url_current_month = URL('subscriptions_show_current')
    elif function == 'dropinclasses':
        url_current_month = URL('dropinclasses_show_current')
    elif function == 'trialclasses':
        url_current_month = URL('trialclasses_show_current')
    elif function == 'trialcards':
        url_current_month = URL('trialcards_show_current')
    elif function == 'classcards':
        url_current_month =  URL('classcards_show_current')
    elif function == 'attendance_classtypes' or \
            function == 'attendance_classes' or \
            function == 'attendance_subcription_exceeded' or \
            function == 'attendance_organizations':
        url_current_month = URL('attendance_show_current')
    elif function == 'direct_debit_extra':
        url_current_month = URL('direct_debit_extra_show_current')
    elif function == 'teacher_classes':
        url_current_month = URL('teacher_classes_show_current')


    show_current_month = A(T("Current month"),
                     _href=url_current_month,
                     _class='btn btn-default')
    month_chooser = ''
    if not function == 'attendance_classes':
        month_chooser = overview_get_month_chooser(function)

    location = ''
    if location_filter:
        location = form.custom.widget.school_locations_id

    form = DIV(XML('<form id="MainForm" action="#" enctype="multipart/form-data" method="post">'),
               DIV(form.custom.widget.month,
                   form.custom.widget.year,
                   location,
                   _class=_class),
               form.custom.end,
               _class='row')

    return dict(
        form=form,
        subtitle=subtitle,
        month_chooser=month_chooser,
        current_month=show_current_month,
        submit=submit
    )

# helpers end


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('read', 'reports_trial'))
def trialclasses_export():
    """
        Export list of trialclasses to Excel sheet
    """
    year = session.reports_tc_year
    month = session.reports_tc_month


    # create filestream
    stream = io.BytesIO()

    # init workbook & sheet
    wb = openpyxl.workbook.Workbook(write_only=True)
    title = "All"
    ws = wb.create_sheet(title=title)
    ws_no_subscription = wb.create_sheet(title="Without subscription")

    header = [
        'Class date',
        'Class type',
        'Class start',
        'Customer ID',
        'Customer Name',
        'Email',
        'Telephone',
        'Mobile'
    ]
    ws.append(header)
    ws_no_subscription.append(header)

    date = datetime.date(year,month,1)
    firstdaythismonth = date
    lastdaythismonth = get_last_day_month(date)

    query = (db.classes_attendance.AttendanceType == 1) &\
            (db.classes_attendance.ClassDate >= firstdaythismonth) &\
            (db.classes_attendance.ClassDate <= lastdaythismonth)

    left = [ db.auth_user.on(db.classes_attendance.auth_customer_id == \
                             db.auth_user.id),
             db.classes.on(db.classes_attendance.classes_id ==
                           db.classes.id),
             db.school_classtypes.on(db.classes.school_classtypes_id ==
                                     db.school_classtypes.id) ]

    rows = db(query).select(db.classes_attendance.ALL,
                            db.auth_user.id,
                            db.auth_user.trashed,
                            db.auth_user.display_name,
                            db.auth_user.email,
                            db.auth_user.thumbsmall,
                            db.auth_user.birthday,
                            db.auth_user.phone,
                            db.auth_user.mobile,
                            db.classes.ALL,
                            db.school_classtypes.Name,
                            left=left,
                            orderby=db.classes_attendance.ClassDate|\
                                    db.auth_user.display_name)

    for row in rows:
        fields = [
            row.classes_attendance.ClassDate,
            row.school_classtypes.Name,
            row.classes.Starttime.strftime(TIME_FORMAT),
            row.auth_user.id,
            row.auth_user.display_name,
            row.auth_user.email,
            row.auth_user.phone,
            row.auth_user.mobile
        ]

        ws.append(fields)

        has_subscription = trial_get_subscription(row.auth_user.id)
        if not has_subscription:
            ws_no_subscription.append(fields)

    fname = T("Trialclasses") + '.xlsx'
    wb.save(stream)

    response.headers['Content-Type']='application/vnd.ms-excel'
    response.headers['Content-disposition']='attachment; filename=' + fname

    return stream.getvalue()






@auth.requires_login()
def trialclasses_show_current():
    session.reports_tc_year = None
    session.reports_tc_month = None

    redirect(URL('trialclasses'))



@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('read', 'reports_trial'))
def trialclasses():
    response.title = T('Reports')
    response.view = 'reports/subscriptions.html'
    session.customers_back = 'trialclasses'

    today = datetime.date.today()
    if 'year' in request.vars:
        year = int(request.vars['year'])
    elif not session.reports_tc_year is None:
        year = session.reports_tc_year
    else:
        year = today.year
    session.reports_tc_year = year
    if 'month' in request.vars:
        month = int(request.vars['month'])
    elif not session.reports_tc_month is None:
        month = session.reports_tc_month
    else:
        month = today.month
    session.reports_tc_month = month

    date = datetime.date(year,month,1)

    form_subtitle = get_form_subtitle(month, year, request.function)
    response.subtitle =  T("Trial classes") + ' - ' + form_subtitle['subtitle']
    form = form_subtitle['form']
    month_chooser = form_subtitle['month_chooser']
    current_month = form_subtitle['current_month']
    submit = form_subtitle['submit']

    rows = dropin_trial_classes_get_rows(date, 'trial')

    table = TABLE(_class="table table-hover")
    table.append(THEAD(TR(TH(), # image
                    TH(), # name,
                    TH(T("Subscription")),
                    TH(T("Class card")),
                    TH(T("Price")),
                    TH())))

    subscription_total = 0
    classcard_total    = 0
    trialprice_total   = 0
    trial_total        = len(rows)

    green_check = SPAN(_class='fa fa-check green big_check')
    for i, row in enumerate(rows):
        repr_row = list(rows[i:i+1].render())[0]
        # get price
        price = 0
        clsID = row.classes_attendance.classes_id
        date = row.classes_attendance.ClassDate
        cls = Class(clsID, date)

        prices = cls.get_prices()
        if prices['trial']:
            price = prices['trial']
            trialprice_total += price
            price_display = SPAN(CURRSYM, ' ', format(price, '.2f'))
        else:
            price_display = ''

        # check for subscription
        subscription = trial_get_subscription(row.auth_user.id)
        if subscription:
            subscription = TABLE(TR(TD(green_check, ' '),
                                    TD(subscription, _class='smaller_font grey')))
            subscription_total += 1
        else:
            subscription = ''
        # check for classcard
        classcard = trial_get_classcard(row.auth_user.id)
        if classcard:
            classcard = TABLE(TR(TD(green_check, ' '),
                                 TD(classcard, _class='smaller_font grey')))
            classcard_total += 1
        else:
            classcard = ''

        # get edit button
        edit = os_gui.get_button('edit_notext', URL('customers', 'classes_attendance',
                                             vars={'cuID' : row.auth_user.id}))
        # format class name
        classtype = max_string_length(repr_row.classes.school_classtypes_id, 18)
        classname = SPAN(repr_row.classes.school_locations_id, ' ',
                         repr_row.classes.Starttime, ' ',
                         classtype,
                         _title = repr_row.classes.school_classtypes_id)
        # add table row
        table.append(TR(TD(repr_row.auth_user.thumbsmall,
                           _class='os-customer_image_td'),
                        TD(DIV(repr_row.auth_user.display_name),
                            SPAN(repr_row.classes_attendance.ClassDate,
                                 ' ',
                                 classname,
                                 _class='small_font grey')),
                        TD(subscription),
                        TD(classcard),
                        TD(price_display),
                        TD(DIV(DIV(edit),
                               _class='btn-group pull-right'),
                           _class='td-icons table-vertical-align-middle'),
                        _id=row.auth_user.id))

    trialclasses = table

    totals_table = TABLE(
        THEAD(TR(TH(T("Counter")),
                 TH(T("Value")))),
        TR(TD(T('Total revenue')),
           TD(CURRSYM, ' ', format(trialprice_total, '.2f'))),
        TR(TD(T('Total trial classes')), TD(trial_total)),
        TR(TD(T('Total with subscription')), TD(subscription_total)),
        TR(TD(T('Total with classcard')), TD(classcard_total)),
        _class='table')

    button_text = XML(T("Totals"))
    result = os_gui.get_modal(button_text=button_text,
                              modal_title=T('Totals'),
                              modal_content=totals_table,
                              modal_class=generate_password())

    total = result['button']
    modals = DIV(result['modal'])

    link_trialclasses = A(SPAN(os_gui.get_fa_icon('fa-compass'), ' ',
                               T('Trialclasses')),
                          _href=URL('trialclasses_export'))

    links = [ link_trialclasses ]

    export = os_gui.get_dropdown_menu(
            links = links,
            btn_text = '',
            btn_icon = 'download',
            menu_class='pull-right' )

    menu = trial_get_menu(request.function)

    return dict(form=form,
                menu=menu,
                total=total,
                content=table,
                modals=modals,
                export=export,
                current_month=current_month,
                month_chooser=month_chooser,
                submit=submit)


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('read', 'reports_trial'))
def trialcards():
    response.title = T('Reports')
    response.view = 'reports/subscriptions.html'
    session.customers_back = 'trialcards'

    today = datetime.date.today()
    if 'year' in request.vars:
        year = int(request.vars['year'])
    elif not session.reports_tcd_year is None:
        year = session.reports_tcd_year
    else:
        year = today.year
    session.reports_tcd_year = year
    if 'month' in request.vars:
        month = int(request.vars['month'])
    elif not session.reports_tcd_month is None:
        month = session.reports_tcd_month
    else:
        month = today.month
    session.reports_tcd_month = month

    date = datetime.date(year,month,1)
    firstdaythismonth = date
    next_month = date.replace(day=28) + datetime.timedelta(days=4)  # this will never fail
    lastdaythismonth = next_month - datetime.timedelta(days=next_month.day)

    form_subtitle = get_form_subtitle(month, year, request.function)
    response.subtitle = T("Trial cards") + ' - ' + form_subtitle['subtitle']
    form = form_subtitle['form']
    month_chooser = form_subtitle['month_chooser']
    current_month = form_subtitle['current_month']
    submit = form_subtitle['submit']

    query = (db.customers_classcards.Startdate >= firstdaythismonth) &\
            (db.customers_classcards.Startdate <= lastdaythismonth) &\
            (db.school_classcards.Trialcard == True)

    left = [ db.auth_user.on(db.customers_classcards.auth_customer_id == \
                             db.auth_user.id),
             db.school_classcards.on(
                 db.customers_classcards.school_classcards_id == \
                 db.school_classcards.id) ]

    rows = db(query).select(db.customers_classcards.ALL,
                            db.auth_user.id,
                            db.auth_user.display_name,
                            db.auth_user.trashed,
                            db.auth_user.thumbsmall,
                            db.auth_user.birthday,
                            db.school_classcards.ALL,
                            left=left,
                            orderby=db.customers_classcards.Startdate|\
                                    db.auth_user.display_name)

    table = TABLE(_class="table table-hover")
    table.append(THEAD(TR(TH(), # image
                    TH(), # name + small grey card name
                    TH(T("Trial card")),
                    TH(T("Subscription")),
                    TH(T("Class card")),
                    TH(T("Price")),
                    TH())))

    subscription_total = 0
    classcard_total    = 0
    trialprice_total   = 0
    trial_total        = len(rows)

    green_check = SPAN(_class='fa fa-check green')
    for i, row in enumerate(rows):
        repr_row = list(rows[i:i+1].render())[0]
        # check for subscription
        subscription = trial_get_subscription(row.auth_user.id)
        if subscription:
            subscription = green_check
            subscription_total += 1
        else:
            subscription = ''
        # check for classcard
        classcard = trial_get_classcard(row.auth_user.id)
        if classcard:
            classcard = green_check
            classcard_total += 1
        else:
            classcard = ''
        # process trialprice total
        if row.school_classcards.Price:
            trialprice_total += row.school_classcards.Price
        # get edit button
        edit = os_gui.get_button('edit_notext', URL('customers', 'classcards',
                                             vars={'cuID' : row.auth_user.id}))
        # add table row
        table.append(TR(TD(repr_row.auth_user.thumbsmall,
                           _class='os-customer_image_td'),
                        TD(DIV(repr_row.auth_user.display_name)),
                        TD(row.school_classcards.Name + \
                          T(' - ') + \
                           repr_row.customers_classcards.Startdate),
                        TD(subscription),
                        TD(classcard),
                        TD(repr_row.school_classcards.Price),
                        TD(DIV(DIV(edit),
                               _class='btn-group pull-right'),
                           _class='td-icons table-vertical-align-middle'),
                        _id=row.auth_user.id))

    trialclasses = table

    totals_table = TABLE(
        THEAD(TR(TH(T("Counter")),
                 TH(T("Value")))),
        TR(TD(T('Total revenue')),
           TD(CURRSYM, ' ', format(trialprice_total, '.2f'))),
        TR(TD(T('Total trial cards')), TD(trial_total)),
        TR(TD(T('Total with subscription')), TD(subscription_total)),
        TR(TD(T('Total with classcard')), TD(classcard_total)),
        _class='table')

    button_text = XML(T("Totals"))
    result = os_gui.get_modal(button_text=button_text,
                              modal_title=T('Totals'),
                              modal_content=totals_table,
                              modal_class=generate_password())

    total = result['button']
    modals = DIV(result['modal'])

    menu = trial_get_menu(request.function)

    return dict(form=form,
                menu=menu,
                total=total,
                content=table,
                modals=modals,
                current_month=current_month,
                month_chooser=month_chooser,
                submit=submit)


def trial_get_subscription(cuID):
    """
        Check if a customer has/had a subscription, ever
    """
    left = [ db.school_subscriptions.on(
        db.customers_subscriptions.school_subscriptions_id == \
        db.school_subscriptions.id) ]

    query = (db.customers_subscriptions.auth_customer_id == cuID)
    rows = db(query).select(db.customers_subscriptions.ALL,
                            db.school_subscriptions.Name,
                            left=left,
                            orderby=db.customers_subscriptions.Startdate)
    if rows:
        subscription = rows.first()
        name = max_string_length(subscription.school_subscriptions.Name, 26)
        return_value = SPAN(subscription.customers_subscriptions.Startdate,
                            BR(),
                            name,
                            _title = subscription.school_subscriptions.Name)
    else:
        return_value = False

    return return_value


def trial_get_classcard(cuID):
    """
        Check if a customer has/had a subscription, ever
    """
    left = [ db.school_classcards.on(
        db.customers_classcards.school_classcards_id == \
        db.school_classcards.id) ]

    query = (db.customers_classcards.auth_customer_id == cuID) & \
            (db.school_classcards.Trialcard == False) & \
            (db.school_classcards.id ==
             db.customers_classcards.school_classcards_id)
    left = [ db.school_classcards.on(
        db.customers_classcards.school_classcards_id ==
        db.school_classcards.id) ]

    rows = db(query).select(db.customers_classcards.ALL,
                            db.school_classcards.Name,
                            left=left,
                            orderby=db.customers_classcards.Startdate)

    if rows:
        card = rows.first()
        name = max_string_length(card.school_classcards.Name, 26)
        return_value = SPAN(card.customers_classcards.Startdate, BR(),
                            name,
                            _title = card.school_classcards.Name)
    else:
        return_value = False

    return return_value


def trial_get_menu(page=None):
    pages = [
        (['trialclasses', T('Trial classes'), URL('reports',"trialclasses")]),
        (['trialcards', T('Trial cards'), URL('reports',"trialcards")]),
        ]

    return get_submenu(pages,
                       page,
                       horizontal=True,
                       htype='tabs')


@auth.requires_login()
def trialcards_show_current():
    session.reports_tcd_year = None
    session.reports_tcd_month = None

    redirect(URL('trialcards'))


@auth.requires_login()
def dropinclasses_show_current():
    session.reports_dic_year = None
    session.reports_dic_month = None

    redirect(URL('dropinclasses'))


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('read', 'reports_dropinclasses'))
def dropinclasses():
    response.title = T('Reports')
    session.customers_back = 'dropinclasses'

    response.view = 'reports/general.html'

    today = datetime.date.today()
    if 'year' in request.vars:
        year = int(request.vars['year'])
    elif not session.reports_dic_year is None:
        year = session.reports_dic_year
    else:
        year = today.year
    session.reports_dic_year = year
    if 'month' in request.vars:
        month = int(request.vars['month'])
    elif not session.reports_dic_month is None:
        month = session.reports_dic_month
    else:
        month = today.month
    session.reports_dic_month = month

    date = datetime.date(year,month,1)

    result = get_form_subtitle(month, year, request.function)
    response.subtitle = T("Drop in classes") + ' - ' + result['subtitle']
    form = result['form']
    month_chooser = result['month_chooser']
    current_month = result['current_month']
    submit = result['submit']

    rows = dropin_trial_classes_get_rows(date, 'dropin')

    classes_total = len(rows)
    price_total = 0

    table = TABLE(_class="table table-striped table-hover")
    table.append(THEAD(TR(TH(), # image
                    TH(), # name,
                    TH(db.classes_attendance.ClassDate.label),
                    TH(db.classes.school_locations_id.label),
                    TH(db.classes.school_classtypes_id.label),
                    TH(T('Time')),
                    TH(T('Price')),
                    TH())))
    for i, row in enumerate(rows):
        repr_row = list(rows[i:i+1].render())[0]
        # get price
        price = 0
        clsID = row.classes_attendance.classes_id
        date = row.classes_attendance.ClassDate
        cls = Class(clsID, date)

        prices = cls.get_prices()
        if prices['dropin']:
            price = prices['dropin']
            price_total += price
            price_display = SPAN(CURRSYM, ' ', format(price, '.2f'))
        else:
            price_display = ''

        edit = os_gui.get_button('edit_notext', URL('customers', 'classes_attendance',
                                             vars={'cuID' : row.auth_user.id}))

        table.append(TR(TD(repr_row.auth_user.thumbsmall,
                           _class='os-customer_image_td'),
                        TD(DIV(repr_row.auth_user.display_name,
                                _class="os-customer_name")),
                        TD(repr_row.classes_attendance.ClassDate),
                        TD(repr_row.classes.school_locations_id),
                        TD(repr_row.classes.school_classtypes_id),
                        TD(repr_row.classes.Starttime, ' - ',
                           repr_row.classes.Endtime),
                        TD(price_display),
                        TD(DIV(DIV(edit),
                               _class='btn-group pull-right'),
                           _class='td-icons table-vertical-align-middle'),
                        _id=row.auth_user.id))

    totals_table = TABLE(
        THEAD(TR(TH(T("Counter")),
                 TH(T("Value")))),
        TR(TD(T('Total revenue')),
           TD(CURRSYM, ' ', format(price_total, '.2f'))),
        TR(TD(T('Total drop in classes')), TD(classes_total)),
        _class='table')

    button_text = XML(T("Totals"))
    result = os_gui.get_modal(button_text=button_text,
                              modal_title=T('Totals'),
                              modal_content=totals_table,
                              modal_class=generate_password())

    total = result['button']
    modals = DIV(result['modal'])

    return dict(form=form,
                total=total,
                content=table,
                modals=modals,
                month_chooser=month_chooser,
                current_month=current_month,
                submit=submit)



def dropin_trial_classes_get_rows(date, att_type):
    """
    @param date: datetime.date - first day of month for the selected month
    @param attendance_type - can be 'dropin' or 'trial'
    @return: gluon.dal.rows object containing all drop in classes for a given month
    """
    if att_type == 'trial':
        att_type = 1
    elif att_type == 'dropin':
        att_type = 2

    firstdaythismonth = date
    next_month = date.replace(day=28) + datetime.timedelta(days=4)  # this will never fail
    lastdaythismonth = next_month - datetime.timedelta(days=next_month.day)

    fields = [
        db.auth_user.id,
        db.auth_user.trashed,
        db.auth_user.thumbsmall,
        db.auth_user.birthday,
        db.auth_user.display_name,
        db.auth_user.date_of_birth,
        db.classes_attendance.classes_id,
        db.classes_attendance.ClassDate,
        db.classes.school_locations_id,
        db.classes.school_classtypes_id,
        db.classes.Starttime,
        db.classes.Endtime
    ]

    query = """
        SELECT au.id,
               au.trashed,
               au.thumbsmall,
               au.birthday,
               au.display_name,
               au.date_of_birth,
               clatt.classes_id,
               clatt.ClassDate,
               CASE WHEN cotc.school_locations_id IS NOT NULL
                    THEN cotc.school_locations_id
                    ELSE cla.school_locations_id
               END AS school_locations_id,
               CASE WHEN cotc.school_classtypes_id IS NOT NULL
                    THEN cotc.school_classtypes_id
                    ELSE cla.school_classtypes_id
               END AS school_classtypes_id,
               CASE WHEN cotc.Starttime IS NOT NULL
                    THEN cotc.Starttime
                    ELSE cla.Starttime
               END AS Starttime,
               CASE WHEN cotc.Endtime IS NOT NULL
                    THEN cotc.Endtime
                    ELSE cla.Endtime
               END AS Endtime
        FROM classes_attendance clatt
        LEFT JOIN classes cla ON cla.id = clatt.classes_id
        LEFT JOIN auth_user au ON au.id = clatt.auth_customer_id
        LEFT JOIN
            ( SELECT id,
                     classes_id,
                     ClassDate,
                     Status,
                     school_locations_id,
                     school_classtypes_id,
                     Starttime,
                     Endtime,
                     auth_teacher_id,
                     teacher_role,
                     auth_teacher_id2,
                     teacher_role2
              FROM classes_otc ) cotc
            ON clatt.classes_id = cotc.classes_id AND clatt.ClassDate = cotc.ClassDate
        WHERE clatt.ClassDate >= '{firstdaythismonth}' AND
              clatt.ClassDate <= '{lastdaythismonth}' AND
              clatt.AttendanceType = {att_type}
        ORDER BY clatt.ClassDate, clatt.classes_id, au.display_name
    """.format(firstdaythismonth=firstdaythismonth,
               lastdaythismonth=lastdaythismonth,
               att_type=att_type)

    rows = db.executesql(query, fields=fields)

    return rows


@auth.requires_login()
def classcards_show_current():
    session.reports_cc_year = None
    session.reports_cc_month = None

    redirect(URL('classcards'))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'reports_classcards'))
def classcards_set_month():
    """
        Sets the session variables for classcards year and month
    """
    year  = request.vars['year']
    month = request.vars['month']
    back  = request.vars['back']

    session.reports_cc_year = int(year)
    session.reports_cc_month = int(month)

    redirect(URL(back))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'reports_dropinclasses'))
def dropinclasses_set_month():
    """
        Sets the session variables for dropin classes year and month
    """
    year  = request.vars['year']
    month = request.vars['month']
    back  = request.vars['back']

    session.reports_dic_year  = int(year)
    session.reports_dic_month = int(month)

    redirect(URL(back))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'reports_attendance'))
def attendance_set_month():
    """
        Sets the session variables for classtypes year and month
    """
    year  = request.vars['year']
    month = request.vars['month']
    back  = request.vars['back']

    session.reports_att_year = int(year)
    session.reports_att_month = int(month)

    redirect(URL(back))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'reports_teacherclasses'))
def teacher_classes_set_month():
    """
        Sets the session variables for teacher_classes year and month
    """
    year  = request.vars['year']
    month = request.vars['month']
    back  = request.vars['back']

    session.reports_te_classes_year = int(year)
    session.reports_te_classes_month = int(month)

    redirect(URL(back))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'reports_trial'))
def trialclasses_set_month():
    """
        Sets the session variables for trial classes year and month
    """
    year  = request.vars['year']
    month = request.vars['month']
    back  = request.vars['back']

    session.reports_tc_year = int(year)
    session.reports_tc_month = int(month)

    redirect(URL(back))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'reports_trial'))
def trialcards_set_month():
    """
        Sets the session variables for trial cards year and month
    """
    year  = request.vars['year']
    month = request.vars['month']
    back  = request.vars['back']

    session.reports_tcd_year = int(year)
    session.reports_tcd_month = int(month)

    redirect(URL(back))


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('read', 'reports_subscriptions'))
def subscriptions_set_month():
    """
        Sets the session variables for classcards year and month
    """
    year  = request.vars['year']
    month = request.vars['month']
    back  = request.vars['back']

    session.reports_subscriptions_year  = int(year)
    session.reports_subscriptions_month = int(month)

    redirect(URL(back))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'reports_direct_debit_extra'))
def direct_debit_extra_set_month():
    """
        Sets the session variables for classcards year and month
    """
    year  = request.vars['year']
    month = request.vars['month']
    back  = request.vars['back']

    session.reports_ap_year  = int(year)
    session.reports_ap_month = int(month)

    redirect(URL(back))


def overview_get_month_chooser(page):
    """
        Returns month chooser for overview
    """
    if page == 'classcards':
        year  = session.reports_cc_year
        month = session.reports_cc_month

        link = 'classcards_set_month'

    if ( page == 'subscriptions_new' or
         page == 'subscriptions_online' or
         page == 'subscriptions_stopped' or
         page == 'subscriptions_paused' or
         page == 'subscriptions_overview' or
         page == 'subscriptions_overview_customers' or
         page == 'subscriptions_alt_prices'):
         year  = session.reports_subscriptions_year
         month = session.reports_subscriptions_month

         link = 'subscriptions_set_month'

    if page == 'dropinclasses':
        year  = session.reports_dic_year
        month = session.reports_dic_month

        link = 'dropinclasses_set_month'

    if page == 'trialclasses':
        year  = session.reports_tc_year
        month = session.reports_tc_month

        link = 'trialclasses_set_month'

    if page == 'trialcards':
        year  = session.reports_tcd_year
        month = session.reports_tcd_month

        link = 'trialcards_set_month'

    if (page == 'attendance_classtypes' or
        page == 'attendance_classes' or
        page == 'attendance_subcription_exceeded' or
        page == 'attendance_organizations'):
        year  = session.reports_att_year
        month = session.reports_att_month

        link = 'attendance_set_month'

    if page == 'direct_debit_extra':
        year  = session.reports_ap_year
        month = session.reports_ap_month

        link = 'direct_debit_extra_set_month'

    if page == 'teacher_classes':
        year  = session.reports_te_classes_year
        month = session.reports_te_classes_month

        link = 'teacher_classes_set_month'


    if month == 1:
        prev_month = 12
        prev_year  = year - 1
    else:
        prev_month = month - 1
        prev_year  = year

    if month == 12:
        next_month = 1
        next_year  = year + 1
    else:
        next_month = month + 1
        next_year  = year

    url_prev = URL(link, vars={'month':prev_month,
                               'year' :prev_year,
                               'back' :page})

    url_next = URL(link, vars={'month':next_month,
                                  'year' :next_year,
                                  'back' :page})

    previous = A(I(_class='fa fa-angle-left'),
                 _href=url_prev,
                 _class='btn btn-default')
    nxt = A(I(_class='fa fa-angle-right'),
            _href=url_next,
            _class='btn btn-default')

    return DIV(previous, nxt, _class='btn-group pull-right')


    # previous = A(SPAN(_class='glyphicon glyphicon-chevron-left'), ' ',
    #              T("Previous"),
    #              _href=URL(link, vars={'month':prev_month,
    #                                    'year' :prev_year,
    #                                    'back' :page}),
    #              _class='btn_date_chooser')
    # nxt = A(T("Next"), ' ',
    #         SPAN(_class='glyphicon glyphicon-chevron-right'),
    #         _href=URL(link, vars={'month':next_month,
    #                               'year' :next_year,
    #                               'back' :page}),
    #         _class='btn_date_chooser pull-right')
    #
    # return DIV(previous, nxt, _class='overview_date_chooser')


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('read', 'reports_classcards'))
def classcards():
    response.title = T("Reports")
    session.customers_back = 'classcards'
    response.view = 'reports/subscriptions.html'

    today = datetime.date.today()
    if 'year' in request.vars:
        year = int(request.vars['year'])
    elif not session.reports_cc_year is None:
        year = session.reports_cc_year
    else:
        year = today.year
    session.reports_cc_year = year
    if 'month' in request.vars:
        month = int(request.vars['month'])
    elif not session.reports_cc_month is None:
        month = session.reports_cc_month
    else:
        month = today.month
    session.reports_cc_month = month

    date = datetime.date(year,month,1)
    firstdaythismonth = date
    next_month = date.replace(day=28) + datetime.timedelta(days=4)  # this will never fail
    lastdaythismonth = next_month - datetime.timedelta(days=next_month.day)

    result = get_form_subtitle(month, year, request.function)
    response.subtitle = T('Class cards') + ' - ' + result['subtitle']
    form = result['form']
    month_chooser = result['month_chooser']
    current_month = result['current_month']
    submit = result['submit']


    query = (db.customers_classcards.Startdate >= firstdaythismonth) & \
            (db.customers_classcards.Startdate <= lastdaythismonth) & \
            (db.school_classcards.Trialcard == False)

    rows = db(query).select(db.auth_user.id,
                            db.auth_user.trashed,
                            db.auth_user.thumbsmall,
                            db.auth_user.birthday,
                            db.auth_user.display_name,
                            db.auth_user.date_of_birth,
                            db.customers_classcards.id,
                            db.customers_classcards.Startdate,
                            db.customers_classcards.Enddate,
                            db.customers_classcards.school_classcards_id,
                            db.school_classcards.Name,
                            db.school_classcards.Classes,
                            db.school_classcards.Price,
                            db.school_classcards.Unlimited,
        left=[db.auth_user.on(db.auth_user.id==\
                              db.customers_classcards.auth_customer_id),
              db.school_classcards.on(
                db.customers_classcards.school_classcards_id ==\
                db.school_classcards.id)],
        orderby=db.customers_classcards.Startdate|~db.auth_user.display_name)

    total_price = 0
    for row in rows:
        total_price += row.school_classcards.Price


    table = TABLE(_class="table table-striped table-hover")
    table.append(THEAD(TR(TH(), # image
                    TH(T('Customer')), # name,
                    TH(T("Card")),
                    TH(db.customers_classcards.Startdate.label),
                    TH(db.customers_classcards.Enddate.label),
                    TH(db.school_classcards.Price.label),
                    TH(),
                    _class='header')))

    for row in rows.render():
        edit = os_gui.get_button('edit_notext', URL('customers', 'classcards',
                                             vars={'cuID' : row.auth_user.id}))
        table.append(TR(TD(row.auth_user.thumbsmall,
                           _class='os-customer_image_td'),
                        TD(DIV(row.auth_user.display_name),
                           SPAN(T("Card"), ' ',
                                 str(row.customers_classcards.id),
                                 _class='small_font grey'),
                           _class="os-customer_name"),
                        TD(row.customers_classcards.school_classcards_id),
                        TD(row.customers_classcards.Startdate),
                        TD(row.customers_classcards.Enddate),
                        TD(row.school_classcards.Price),
                        TD(DIV(DIV(edit),
                               _class='btn-group pull-right'),
                           _class='td-icons table-vertical-align-middle'),
                     _id=row.auth_user.id))

    classcards = table

    totals_table = TABLE(
        THEAD(TR(TH(T("Counter")),
                 TH(T("Value")))),
        TR(TD(T('Total revenue')),
           TD(CURRSYM, ' ', format(total_price, '.2f'))),
        TR(TD(T('Cards sold')), TD(len(rows))),
        _class='table')

    button_text = XML(T("Totals"))
    result = os_gui.get_modal(button_text=button_text,
                              modal_title=T('Totals'),
                              modal_content=totals_table,
                              modal_class=generate_password())

    total = result['button']
    modals = DIV(result['modal'])

    menu = classcards_get_menu(request.function)

    return dict(form=form,
                menu=menu,
                total=total,
                content=classcards,
                month_chooser=month_chooser,
                current_month=current_month,
                modals=modals,
                submit=submit)


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('read', 'reports_classcards'))
def classcards_current():
    response.title = T('Reports')
    response.subtitle = T("Class cards - Current")
    session.customers_back = 'classcards'
    response.view = 'reports/subscriptions.html'

    from openstudio.os_customer_classcard import CustomerClasscard

    sort_by = request.vars['sort_by']
    sort_by = classcards_current_get_parameter_or_session('sort_by',
                                                          'startdate',
                                                          session_parameter='reports_classcards_current_sort_by')

    today = TODAY_LOCAL

    classes_remaining = (db.school_classcards.Classes - db.customers_classcards.ClassesTaken)

    if sort_by == 'startdate' or sort_by is None:
        orderby = db.customers_classcards.Startdate|~db.auth_user.display_name
    elif sort_by == 'expiration':
        orderby = db.customers_classcards.Enddate|~db.auth_user.display_name
    elif sort_by == 'classes_remaining':
        orderby = classes_remaining|~db.auth_user.display_name

    query = ((db.customers_classcards.Enddate >= today) |
             (db.customers_classcards.Enddate == None)) & \
            ((db.customers_classcards.ClassesTaken <
              db.school_classcards.Classes) |
             (db.school_classcards.Unlimited == True))


    rows = db(query).select(db.auth_user.id,
                            db.auth_user.trashed,
                            db.auth_user.thumbsmall,
                            db.auth_user.birthday,
                            db.auth_user.display_name,
                            db.auth_user.date_of_birth,
                            db.auth_user.email,
                            db.customers_classcards.id,
                            db.customers_classcards.Startdate,
                            db.customers_classcards.Enddate,
                            db.customers_classcards.school_classcards_id,
                            db.school_classcards.Name,
                            db.school_classcards.Classes,
                            db.school_classcards.Price,
                            db.school_classcards.Unlimited,
                            db.school_classcards.Trialcard,
                            classes_remaining,
        left=[db.auth_user.on(db.auth_user.id==
                              db.customers_classcards.auth_customer_id),
              db.school_classcards.on(
                db.customers_classcards.school_classcards_id ==
                db.school_classcards.id)],
        orderby=orderby)

    total_price = 0
    for row in rows:
        total_price += row.school_classcards.Price


    table = TABLE(_class="table table-hover table-striped")
    table.append(THEAD(TR(TH(), # image
                    TH(T('Customer')), # name,
                    TH(T("Card")),
                    TH(db.customers_classcards.Startdate.label),
                    TH(db.customers_classcards.Enddate.label),
                    TH(db.school_classcards.Price.label),
                    TH(),
                    _class='header')))

    ccdh = ClasscardsHelper()
    for i, row in enumerate(rows):
        repr_row = list(rows[i:i + 1].render())[0]

        edit = os_gui.get_button('edit_notext', URL('customers', 'classcards',
                                             vars={'cuID' : row.auth_user.id}))

        mail = A(SPAN(_class='glyphicon glyphicon-envelope'),
                      _class="btn btn-default btn-sm",
                      _href='mailto:' + row.auth_user.email)

        trial_label = ''
        if row.school_classcards.Trialcard:
            trial_label = SPAN(
                BR(),
                os_gui.get_label(
                    'success',
                    T("Trial card")
                )
            )


        ccdID = row.customers_classcards.id
        cc = CustomerClasscard(ccdID)
        classes_remaining = cc.get_classes_remaining_formatted()

        table.append(TR(TD(repr_row.auth_user.thumbsmall,
                           _class='os-customer_image_td'),
                        TD(DIV(row.auth_user.display_name),
                           SPAN(T("Card"), ' ',
                                 str(ccdID), ' - ',
                                 B(classes_remaining),
                                 _class='small_font grey'),
                           _class="os-customer_name"),
                        TD(repr_row.customers_classcards.school_classcards_id,
                           trial_label),
                        TD(repr_row.customers_classcards.Startdate),
                        TD(repr_row.customers_classcards.Enddate),
                        TD(repr_row.school_classcards.Price),
                        TD(DIV(mail, edit,
                               _class='btn-group pull-right'),
                           _class='td-icons table-vertical-align-middle'),
                     _id=row.auth_user.id))

    classcards = table

    totals_table = TABLE(
        THEAD(TR(TH(T("Counter")),
                 TH(T("Value")))),
        TR(TD(T('Current cards')), TD(len(rows))),
        _class='table')

    button_text = XML(T("Totals"))
    result = os_gui.get_modal(button_text=button_text,
                              modal_title=T('Totals'),
                              modal_content=totals_table,
                              modal_class=generate_password(),
                              button_class='full-width')

    total = result['button']
    modals = DIV(result['modal'])

    menu = classcards_get_menu(request.function)

    return dict(form=classcards_current_get_form_sort(sort_by),
                menu=menu,
                total=total,
                content=classcards,
                modals=modals)


def classcards_current_get_form_sort(sort_by):
    """
        :return: SQLFORM to choose sorting for list of current class cards
    """
    sort_options = [
        [ 'startdate', T('Start date') ],
        [ 'expiration', T('Expiration date') ],
        [ 'classes_remaining', T('Classes remaining') ]
    ]

    form = SQLFORM.factory(
        Field('sort_by',
              default=sort_by,
              requires=IS_IN_SET(sort_options, zero=None),
              label=T('Sort by')),
        submit_button = T('Go'),
        formstyle = 'divs'
    )

    select = form.element('#no_table_sort_by')
    select['_onchange'] = 'this.form.submit();'

    form = DIV(
        form.custom.begin,
        LABEL(form.custom.label.sort_by),
        form.custom.widget.sort_by,
        form.custom.end
    )

    return form


def classcards_current_get_parameter_or_session(parameter, default_value, session_parameter=None):
    """
        Gets a value from the request or the session.

        If the parameter is not found in the request the it tries to get it from the session.
        Otherwise returns the default value.
        If the session_parameter is not None then stores the value to the session.

        :param parameter: the name of the request parameter
        :param default_value: the value to return when the parameter is not found
        :param session_parameter: the session parameter name
    """
    if parameter in request.vars:
        value = request.vars[parameter]
    elif session_parameter and not session[session_parameter] is None:
        value = session[session_parameter]
    else:
        value = default_value

    if session_parameter:
        session[session_parameter] = value

    return value


def classcards_get_menu(page=None):
    pages = [
        (['classcards', T('Sold cards'), URL("reports","classcards")]),
        (['classcards_current', T('Current cards'),
          URL("reports","classcards_current")]),
        ]

    horizontal = True

    return get_submenu(pages,
                       page,
                       horizontal=horizontal,
                       htype='tabs')


def subscriptions_count_totals(rows):
    """
        Returns panel of counts for each subscription type
    """
    # count subscriptions
    counts = {}
    for row in rows.render():
        ssuID = row.customers_subscriptions.school_subscriptions_id
        if ssuID in counts:
            counts[ssuID] += 1
        else:
            counts[ssuID] = 1

    # sort counts
    import operator
    sorted_subscriptions = sorted(list(counts.items()),
                                  key=operator.itemgetter(1),
                                  reverse=True)

    totals_table = TABLE(THEAD(TR(TH(T("Subscription")),
                                  TH(T("Count")))),
                         _class='table table-striped table-hover')

    for subscription in sorted_subscriptions:
        totals_table.append(TR(TD(subscription[0]), TD(subscription[1])))

    totals_table.append(TR(TD(B("Total")), TD(B(len(rows)))))

    # panel = os_gui.get_panel_table(T("Subscription count"),
    #                                totals_table)

    button_text = XML(T("Totals"))
    result = os_gui.get_modal(button_text=button_text,
                              modal_title=T('Totals'),
                              modal_content=totals_table,
                              modal_class=generate_password())

    return result


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'reports_subscriptions'))
def subscriptions_show_current():
    """
        Reset the session variables for month and year to show the current
        month.
    """
    session.reports_subscriptions_year = None
    session.reports_subscriptions_month = None

    redirect(URL(session.reports_subscriptions))


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('read', 'reports_subscriptions'))
def subscriptions_new():
    from openstudio.os_reports import Reports

    response.title = T("Reports")
    session.customers_back = 'subscriptions_new'
    response.view = 'reports/subscriptions.html'

    # set the session vars for year/month
    subscriptions_process_request_vars()

    date = datetime.date(
        session.reports_subscriptions_year,
        session.reports_subscriptions_month,
        1
    )

    reports = Reports()

    location_filter = False
    filter_school_locations_id = None
    if session.show_location:
        location_filter = True
        filter_school_locations_id = session.reports_subscriptions_school_locations_id

    query = reports.get_query_subscriptions_new_in_month(
        date,
        filter_school_locations_id = filter_school_locations_id
    )

    result = get_form_subtitle(
        session.reports_subscriptions_month,
        session.reports_subscriptions_year,
        request.function,
        location_filter=location_filter
    )
    response.subtitle = T('Subscriptions') + ' - ' + result['subtitle']
    form = result['form']
    month_chooser = result['month_chooser']
    current_month = result['current_month']
    submit = result['submit']

    fields = [
        db.auth_user.id,
        db.auth_user.trashed,
        db.auth_user.thumbsmall,
        db.auth_user.birthday,
        db.auth_user.first_name,
        db.auth_user.last_name,
        db.auth_user.display_name,
        db.auth_user.date_of_birth,
        db.auth_user.email,
        db.customers_subscriptions.school_subscriptions_id,
        db.customers_subscriptions.Startdate,
        db.customers_subscriptions.payment_methods_id
    ]
    rows = db.executesql(query, fields=fields)

    total = T("Total: " + str(len(rows)))

    table = TABLE(_class="table table-hover")
    table.append(THEAD(TR(TH(), # image
                    TH(), # name,
                    TH(db.customers_subscriptions.Startdate.label),
                    TH(db.customers_subscriptions.school_subscriptions_id.label),
                    TH(db.customers_subscriptions.payment_methods_id.label),
                    TH(),
                    _class='header')))

    for row in rows.render():
        edit = os_gui.get_button('edit_notext', URL('customers', 'subscriptions',
                                             vars={'cuID' : row.auth_user.id}))

        table.append(TR(TD(row.auth_user.thumbsmall,
                           _class='os-customer_image_td'),
                        TD(DIV(row.auth_user.display_name),
                           _class="os-customer_name"),
                        TD(row.customers_subscriptions.Startdate),
                        TD(row.customers_subscriptions.school_subscriptions_id),
                        TD(row.customers_subscriptions.payment_methods_id),
                        TD(DIV(DIV(edit),
                               _class='btn-group pull-right'),
                           _class='td-icons table-vertical-align-middle'),
                        _id=row.auth_user.id))

    # totals table begin
    result = subscriptions_count_totals(rows)
    total = result['button']
    modals = DIV(result['modal'])
    # totals table end

    menu = subscriptions_get_menu(request.function)

    link_all_customers = A(
        SPAN(os_gui.get_fa_icon('fa-envelope-o'), ' ', T("Mailinglist")),
        _href=URL('subscriptions_new_export_mailinglist'),
        _class='textalign_left',
        _title=T('All customers with new subscription this month'))

    links = [ link_all_customers ]

    export = os_gui.get_dropdown_menu(
            links = links,
            btn_text = '',
            btn_icon = 'download',
            menu_class='pull-right' )

    return dict(content=table,
                total=total,
                form=form,
                menu=menu,
                modals=modals,
                export=export,
                month_chooser=month_chooser,
                current_month=current_month,
                submit=submit)


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('read', 'reports_subscriptions'))
def subscriptions_new_export_mailinglist():
    """
    Mailchimp compatible excel export
    :return:
    """
    from openstudio.os_reports import Reports

    reports = Reports()

    date = datetime.date(
        session.reports_subscriptions_year,
        session.reports_subscriptions_month,
        1
    )

    location_filter = False
    filter_school_locations_id = None
    if session.show_location:
        location_filter = True
        filter_school_locations_id = session.reports_subscriptions_school_locations_id

    query = reports.get_query_subscriptions_new_in_month(
        date,
        filter_school_locations_id = filter_school_locations_id
    )


    fields = [
        db.auth_user.id,
        db.auth_user.trashed,
        db.auth_user.thumbsmall,
        db.auth_user.birthday,
        db.auth_user.first_name,
        db.auth_user.last_name,
        db.auth_user.display_name,
        db.auth_user.date_of_birth,
        db.auth_user.email,
        db.customers_subscriptions.school_subscriptions_id,
        db.customers_subscriptions.Startdate,
        db.customers_subscriptions.payment_methods_id
     ]
    rows = db.executesql(query, fields=fields)

    # create filestream
    stream = io.BytesIO()

    # Create the workbook
    wb = openpyxl.workbook.Workbook(write_only=True)
    ws = wb.create_sheet(title='Mailinglist')


    for row in rows:
        ws.append([row.auth_user.first_name,
                   row.auth_user.last_name,
                   row.auth_user.email])

    fname = T("Mailinglist") + '.xlsx'
    wb.save(stream)

    response.headers['Content-Type']='application/vnd.ms-excel'
    response.headers['Content-disposition']='attachment; filename=' + fname

    return stream.getvalue()


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('read', 'reports_subscriptions'))
def subscriptions_online():
    """
    List of online subscriptions for a given month
    """
    from openstudio.os_reports import Reports

    response.title = T("Reports")
    session.customers_back = 'subscriptions_online'
    response.view = 'reports/subscriptions.html'

    # Set the session vars for year/month
    subscriptions_process_request_vars()

    result = get_form_subtitle(
        session.reports_subscriptions_month,
        session.reports_subscriptions_year,
        request.function
    )
    response.subtitle = T('Subscriptions') + ' - ' + result['subtitle']
    form = result['form']
    month_chooser = result['month_chooser']
    current_month = result['current_month']
    submit = result['submit']

    reports = Reports()
    rows = reports.get_subscriptions_online_in_month_rows(datetime.date(
        session.reports_subscriptions_year,
        session.reports_subscriptions_month,
        1
    ))

    header = THEAD(TR(
        TH(), # Image
        TH(), # Name
        TH(T("Start")),
        TH(T("Subscription")),
        TH(T("Payment method")),
        TH(T("Verified")),
    ))

    table = TABLE(header, _class="table table-hover")
    permission_edit = (
        auth.has_membership(group_id='Admins') or
        auth.has_permission('update', 'customers_subscriptions')
    )
    permission_edit_pm = (
        auth.has_membership(group_id='Admins') or
        auth.has_permission('update', 'customers_payment_info')
    )

    for i, row in enumerate(rows):
        repr_row = list(rows[i:i + 1].render())[0]

        btn_vars = {"cuID": row.auth_user.id}
        subscription = A(
            repr_row.customers_subscriptions.school_subscriptions_id,
            _href=URL('customers', 'subscriptions', vars=btn_vars)
        ) if permission_edit else repr_row.customers_subscriptions.school_subscriptions_id

        pm = A(
            repr_row.customers_subscriptions.payment_methods_id,
            _href=URL("customers", "bankaccount", vars=btn_vars)
        ) if permission_edit_pm else repr_row.customers_subscriptions.payment_methods_id

        table.append(TR(
            TD(repr_row.auth_user.thumbsmall, _class="os-customer_image_td"),
            TD(row.auth_user.display_name, _class="os-customer_name"),
            TD(repr_row.customers_subscriptions.Startdate),
            TD(subscription),
            TD(pm),
            TD(repr_row.customers_subscriptions.Verified),
        ))


    menu = subscriptions_get_menu(request.function)

    return dict(
        content=table,
        form=form,
        menu=menu,
        month_chooser=month_chooser,
        current_month=current_month,
        submit=submit
    )


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('read', 'reports_subscriptions'))
def subscriptions_stopped():
    response.title = T("Reports")
    session.customers_back = 'subscriptions_stopped'
    response.view = 'reports/subscriptions.html'

    # set the session vars for year/month
    subscriptions_process_request_vars()

    date = datetime.date(
        session.reports_subscriptions_year,
        session.reports_subscriptions_month,
        1
    )

    firstdaythismonth = date
    next_month = date.replace(day=28) + datetime.timedelta(days=4)  # this will never fail
    lastdaythismonth = next_month - datetime.timedelta(days=next_month.day)

    reports = Reports()

    location_filter = False
    filter_school_locations_id = None
    if session.show_location:
        location_filter = True
        filter_school_locations_id = session.reports_subscriptions_school_locations_id

    query = reports.get_query_subscriptions_stopped_in_month(
        date,
        filter_school_locations_id = filter_school_locations_id
    )

    form_subtitle = get_form_subtitle(
        session.reports_subscriptions_month,
        session.reports_subscriptions_year,
        request.function,
        location_filter=location_filter
    )
    response.subtitle = T('Subscriptions') + ' - ' + form_subtitle['subtitle']
    form = form_subtitle['form']
    month_chooser = form_subtitle['month_chooser']
    current_month = form_subtitle['current_month']
    submit = form_subtitle['submit']

    fields = [ db.auth_user.id,
               db.auth_user.trashed,
               db.auth_user.thumbsmall,
               db.auth_user.birthday,
               db.auth_user.display_name,
               db.auth_user.date_of_birth,
               db.customers_subscriptions.school_subscriptions_id,
               db.customers_subscriptions.Enddate ]

    rows = db.executesql(query, fields=fields)

    total = T("Total: " + str(len(rows)))

    table = TABLE(_class="table table-hover")
    table.append(THEAD(TR(TH(), # image
                    TH(), # name,
                    TH(db.customers_subscriptions.Enddate.label),
                    TH(db.customers_subscriptions.school_subscriptions_id.label),
                    TH(),
                    _class='header')))
    for row in rows.render():
        edit = os_gui.get_button('edit_notext', URL('customers', 'subscriptions',
                                             vars={'cuID' : row.auth_user.id}))

        table.append(TR(TD(row.auth_user.thumbsmall,
                           _class='os-customer_image_td'),
                        TD(DIV(row.auth_user.display_name),
                           _class="os-customer_name"),
                        TD(row.customers_subscriptions.Enddate),
                        TD(row.customers_subscriptions.school_subscriptions_id),
                        TD(DIV(DIV(edit),
                               _class='btn-group pull-right'),
                           _class='td-icons table-vertical-align-middle'),
                        _id=row.auth_user.id))

    # totals table begin
    result = subscriptions_count_totals(rows)
    total = result['button']
    modals = DIV(result['modal'])
    # totals table end

    menu = subscriptions_get_menu(request.function)

    return dict(content=table,
                total=total,
                form=form,
                menu=menu,
                modals=modals,
                month_chooser=month_chooser,
                current_month=current_month,
                submit=submit)


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('read', 'reports_subscriptions'))
def subscriptions_paused():
    response.title = T("Reports")
    session.customers_back = 'subscriptions_paused'
    response.view = 'reports/subscriptions.html'

    # set the session vars for year/month
    subscriptions_process_request_vars()

    date = datetime.date(session.reports_subscriptions_year,
                         session.reports_subscriptions_month,
                         1)
    firstdaythismonth = date
    next_month = date.replace(day=28) + datetime.timedelta(days=4)  # this will never fail
    lastdaythismonth = next_month - datetime.timedelta(days=next_month.day)

    form_subtitle = get_form_subtitle(session.reports_subscriptions_month,
                                            session.reports_subscriptions_year,
                                            request.function)
    response.subtitle = T('Subscriptions') + ' - ' + form_subtitle['subtitle']
    form = form_subtitle['form']
    month_chooser = form_subtitle['month_chooser']
    current_month = form_subtitle['current_month']
    submit = form_subtitle['submit']

    # Paused subscriptions
    query = (db.customers_subscriptions_paused.Startdate <=
                                               lastdaythismonth) & \
            ((db.customers_subscriptions_paused.Enddate >=
                                                firstdaythismonth) |
             (db.customers_subscriptions_paused.Enddate == None))
    rows = db(query).select(db.auth_user.id,
                            db.auth_user.trashed,
                            db.auth_user.thumbsmall,
                            db.auth_user.birthday,
                            db.auth_user.display_name,
                            db.customers_subscriptions_paused.Description,
        left=[ db.customers_subscriptions.on(
                    db.customers_subscriptions.id==
                    db.customers_subscriptions_paused.customers_subscriptions_id),
               db.auth_user.on(db.auth_user.id==\
                               db.customers_subscriptions.auth_customer_id)])
    total = T("Total: " + str(len(rows)))

    table = TABLE(_class="table table-hover")
    table.append(THEAD(TR(TH(), # image
                    TH(), # name,
                    TH(db.customers_subscriptions_paused.Description.label),
                    _class='header')))
    for row in rows.render():
        edit = os_gui.get_button('edit_notext', URL('customers', 'subscriptions',
                                             vars={'cuID' : row.auth_user.id}))
        table.append(TR(TD(row.auth_user.thumbsmall,
                           _class='os-customer_image_td'),
                        TD(DIV(row.auth_user.display_name),
                           _class="os-customer_name"),
                        TD(row.customers_subscriptions_paused.Description),
                        TD(DIV(DIV(edit),
                               _class='btn-group pull-right'),
                           _class='td-icons table-vertical-align-middle'),
                        _id=row.auth_user.id))
    paused_subscriptions = table

    content = DIV(table, DIV(total, _class='right'))
    menu = subscriptions_get_menu(request.function)

    return dict(content=content,
                form=form,
                menu=menu,
                month_chooser=month_chooser,
                current_month=current_month,
                submit=submit)


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('read', 'reports_subscriptions'))
def subscriptions_overview():
    """
        Lists current subscriptions by type
    """
    response.title = T("Reports")
    response.view = 'reports/subscriptions.html'

    # set the session vars for year/month & optional location filter
    subscriptions_process_request_vars()

    # get first and last day of this month
    date = datetime.date(session.reports_subscriptions_year,
                         session.reports_subscriptions_month,
                         1)
    firstdaythismonth = date
    lastdaythismonth = get_last_day_month(date)

    # get month/year selection form and subtitle
    location_filter = False
    if session.show_location:
        location_filter = True

    form_subtitle = get_form_subtitle(session.reports_subscriptions_month,
                                      session.reports_subscriptions_year,
                                      request.function,
                                      location_filter=location_filter)
    response.subtitle = T('Subscriptions') + ' - ' + form_subtitle['subtitle']
    form = form_subtitle['form']
    month_chooser = form_subtitle['month_chooser']
    current_month = form_subtitle['current_month']
    submit = form_subtitle['submit']


    count = db.customers_subscriptions.school_subscriptions_id.count()
    query = (db.customers_subscriptions.Startdate <= lastdaythismonth) & \
            ((db.customers_subscriptions.Enddate >= firstdaythismonth) |
             (db.customers_subscriptions.Enddate == None))

    if session.show_location:
        if session.reports_subscriptions_school_locations_id:
            query &= (db.auth_user.school_locations_id ==
                      session.reports_subscriptions_school_locations_id)

    rows = db(query).select(db.customers_subscriptions.ALL,
                            db.school_subscriptions.ALL,
                            count,
                            left=[db.school_subscriptions.on(
                                db.customers_subscriptions.school_subscriptions_id == \
                                db.school_subscriptions.id),
                               db.auth_user.on(
                                db.customers_subscriptions.auth_customer_id == \
                                db.auth_user.id)],
                            groupby=db.customers_subscriptions.school_subscriptions_id,
                            orderby=~count)

    revenue_total = 0
    subscriptions_total = 0
    table = TABLE(THEAD(TR(TH(T('Subscription')),
                     TH(T('Monthly fee'), _class='text-right'),
                     TH(T('Customers'), _class='text-right'),
                     TH(T('Revenue'), _class='text-right'),
                     TH(),
                     _class='header')),
                  _class='table table-hover table-striped')
    for row in rows:
        ssuID = row.school_subscriptions.id
        ssu = SchoolSubscription(ssuID)
        monthly_fee = ssu.get_price_on_date(firstdaythismonth, False)
        revenue = monthly_fee * row[count]

        if not revenue:
            revenue = 0

        links = [ A(SPAN(os_gui.get_fa_icon('fa-envelope-o'), ' ',
                         T('Mailinglist')),
                    _href=URL('subscriptions_overview_export_mailinglist',
                              vars={'ssuID':row.school_subscriptions.id})),
                  A(SPAN(os_gui.get_fa_icon('fa-check'), ' ', T('Attendance')),
                    _href=URL('subscriptions_overview_export_attendance',
                              vars={'ssuID':row.school_subscriptions.id})) ]

        export_row = os_gui.get_dropdown_menu(
                links = links,
                btn_text = '',
                btn_icon = 'download',
                btn_size = 'btn-sm',
                menu_class='btn-group pull-right')

        table.append(TR(TD(row.school_subscriptions.Name),
                        TD(SPAN(CURRSYM, ' ', format(monthly_fee, '.2f')),
                           _class='text-right'),
                        TD(A(row[count],
                             _href=URL('subscriptions_overview_customers',
                                       vars={'ssuID':ssuID})),
                           _class='text-right'),
                        TD(SPAN(CURRSYM, ' ', format(revenue, '.2f')),
                           _class='text-right'),
                        TD(export_row)))

        subscriptions_total += row[count]
        revenue_total += revenue

    table.append(TR(TD(T('Total')),
                    TD(),
                    TD(subscriptions_total,
                       _class='text-right'),
                    TD(SPAN(CURRSYM, ' ', format(revenue_total, '.2f')),
                       _class='text-right'),
                    TD(),
                 _class='total'))


    link_all_customers = A(
        SPAN(os_gui.get_fa_icon('fa-users'), ' ', T("All customers")),
        _href=URL('subscriptions_overview_export_all_customers'),
        _class='textalign_left',
        _title=T('All customers with a subscription this month'))

    links = [ link_all_customers ]

    export = os_gui.get_dropdown_menu(
            links = links,
            btn_text = '',
            btn_icon = 'download',
            menu_class='pull-right' )

    menu = subscriptions_get_menu(request.function)

    return dict(content=table,
                form=form,
                export=export,
                menu=menu,
                month_chooser=month_chooser,
                current_month=current_month,
                submit=submit)


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('read', 'reports_subscriptions'))
def subscriptions_overview_export_all_customers():
    """
        Export all customers with a subscription this month to Excel
    """
    # create filestream
    stream = io.BytesIO()
    # create dictionary to lookup latest subscription values
    date = datetime.date(session.reports_subscriptions_year,
                         session.reports_subscriptions_month,
                         1)
    firstdaythismonth = date
    lastdaythismonth = get_last_day_month(date)
    result = db.executesql("""SELECT cs.auth_customer_id,
                                     ssu.name,
                                     cs.startdate,
                                     cs.enddate,
                                     pm.Name
                               FROM customers_subscriptions cs
        LEFT JOIN
        (SELECT id, name FROM school_subscriptions) ssu
        ON ssu.id = cs.school_subscriptions_id
        LEFT JOIN payment_methods pm ON cs.payment_methods_id = pm.id
        WHERE cs.startdate <= %s AND
              (cs.enddate > %s OR cs.enddate IS NULL)
        GROUP BY cs.auth_customer_id """, (lastdaythismonth, firstdaythismonth))
    fname = T("Subscriptions overview") + '.xlsx'
    title = 'Subscriptions'
    cu_mem_dict = dict()
    for record in result:
        subscription = record[1] or ""
        start = record[2] or ""
        end = record[3] or ""
        pm = record[4] or ""
        cu_mem_dict[record[0]] = [subscription, start, end, pm]

    bd_dict = dict()

    rows = db().select(db.customers_payment_info.ALL)
    for row in rows:
        bd_dict[row.auth_customer_id] = [
                                      row.AccountNumber,
                                      row.AccountHolder,
                                      row.BankName,
                                      row.BankLocation ]

    # Create the workbook
    wb = openpyxl.workbook.Workbook(write_only=True)
    ws = wb.create_sheet(title=title)
    headers = [ "id",
                "First name",
                "Last name",
                "Date of birth",
                "Gender",
                "Company",
                "Address",
                "Postal code",
                "City",
                "Country",
                "Email",
                "Newsletter",
                "Telephone",
                "Mobile",
                "Key",
                "Location",
                "Subscription",
                "Startdate",
                "Enddate",
                "Payment method",
                "AccountNR",
                "AccountHolder",
                "BankName",
                "BankLocation"]
    ws.append(headers)

    query = (db.auth_user.trashed == False)
    rows = db(query).select(db.auth_user.ALL,
                            db.school_locations.Name,
            left=[db.school_locations.on(db.auth_user.school_locations_id==\
                                         db.school_locations.id)])
    for row in rows:
        customers_id = row.auth_user.id
        if customers_id not in cu_mem_dict:
            # subscription list, if no subscription --> check the next customer.
            continue
        else:
            data = [ row.auth_user.id,
                     row.auth_user.first_name,
                     row.auth_user.last_name,
                     row.auth_user.date_of_birth,
                     row.auth_user.gender,
                     row.auth_user.company,
                     row.auth_user.address,
                     row.auth_user.postcode,
                     row.auth_user.city,
                     row.auth_user.country,
                     row.auth_user.email,
                     row.auth_user.newsletter,
                     row.auth_user.phone,
                     row.auth_user.mobile,
                     row.auth_user.keynr,
                     row.school_locations.Name,
                     cu_mem_dict[customers_id][0],
                     cu_mem_dict[customers_id][1],
                     cu_mem_dict[customers_id][2],
                     cu_mem_dict[customers_id][3]]
            if not bd_dict.get(customers_id, None) is None:
                data.append(bd_dict[customers_id][0])
                data.append(bd_dict[customers_id][1])
                data.append(bd_dict[customers_id][2])
                data.append(bd_dict[customers_id][3])

            ws.append(data)

    wb.save(stream)

    response.headers['Content-Type']='application/vnd.ms-excel'
    response.headers['Content-disposition']='attachment; filename=' + fname

    return stream.getvalue()


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('read', 'reports_subscriptions'))
def subscriptions_overview_export_mailinglist():
    """
        Export mailinglist based on subscriptions overview
    """
    ssuID = request.vars['ssuID']
    year  = session.reports_subscriptions_year
    month = session.reports_subscriptions_month

    # create filestream
    stream = io.BytesIO()

    # Create the workbook
    wb = openpyxl.workbook.Workbook(write_only=True)
    ws = wb.create_sheet(title='Mailinglist')

    rows = subscriptions_overview_customers_get_rows(ssuID)

    for row in rows:
        ws.append([row.auth_user.first_name,
                   row.auth_user.last_name,
                   row.auth_user.email])

    fname = T("Mailinglist") + '.xlsx'
    wb.save(stream)

    response.headers['Content-Type']='application/vnd.ms-excel'
    response.headers['Content-disposition']='attachment; filename=' + fname

    return stream.getvalue()


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('read', 'reports_subscriptions'))
def subscriptions_overview_export_attendance():
    """
        Export mailinglist based on subscriptions overview
    """
    ssuID = request.vars['ssuID']
    year  = session.reports_subscriptions_year
    month = session.reports_subscriptions_month

    subscription = db.school_subscriptions(ssuID)

    # create filestream
    stream = io.BytesIO()

    # Create the workbook
    wb = openpyxl.workbook.Workbook(write_only=True)
    ws = wb.create_sheet(title=subscription.Name[0:30])

    header = [
        'Class date',
        'Class type',
        'Class start',
        'First name',
        'Last name'
    ]

    ws.append(header)

    # get first and last day of this month
    date = datetime.date(session.reports_subscriptions_year,
                         session.reports_subscriptions_month,
                         1)
    firstdaythismonth = date
    lastdaythismonth = get_last_day_month(date)

    query = (db.school_subscriptions.id == ssuID) & \
            (db.classes_attendance.ClassDate >= firstdaythismonth) & \
            (db.classes_attendance.ClassDate <= lastdaythismonth)
    left  = [ db.customers_subscriptions.on(
                db.classes_attendance.customers_subscriptions_id ==
                db.customers_subscriptions.id),
              db.school_subscriptions.on(
                db.customers_subscriptions.school_subscriptions_id ==
                db.school_subscriptions.id),
              db.auth_user.on(
                db.customers_subscriptions.auth_customer_id ==
                db.auth_user.id),
              db.classes.on(
                db.classes_attendance.classes_id ==
                db.classes.id),
              db.school_classtypes.on(
                db.classes.school_classtypes_id ==
                db.school_classtypes.id)
             ]

    rows = db(query).select(db.auth_user.first_name,
                            db.auth_user.last_name,
                            db.school_subscriptions.Name,
                            db.classes_attendance.ClassDate,
                            db.school_classtypes.Name,
                            db.classes.Starttime,
                            left=left)

    for row in rows:
        ws.append([row.classes_attendance.ClassDate,
                   row.school_classtypes.Name,
                   row.classes.Starttime,
                   row.auth_user.first_name,
                   row.auth_user.last_name])

    fname = T("Subscription attendance") + '.xlsx'
    wb.save(stream)

    response.headers['Content-Type']='application/vnd.ms-excel'
    response.headers['Content-disposition']='attachment; filename=' + fname

    return stream.getvalue()


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'reports_subscriptions'))
def subscriptions_overview_customers():
    """
        Lists customers for a selected subscription in a month
    """
    # set the session vars for year/month & optional location filter
    subscriptions_process_request_vars()

    if 'ssuID' in request.vars:
        ssuID = request.vars['ssuID']
        session.reports_subscriptions_overview_customers_ssuID = ssuID
        # prevent duplicate ssuID (it could be in URL or from the ssuID form)
        redirect(URL())
    else:
        ssuID = session.reports_subscriptions_overview_customers_ssuID


    # Redirect back from customers edit pages
    session.customers_back = request.function

    response.title = T("Subscriptions")
    response.view = 'reports/general.html'

    ssu = SchoolSubscription(ssuID)

    # get month/year selection form and subtitle
    location_filter = False
    if session.show_location:
        location_filter = True

    form_subtitle = get_form_subtitle(session.reports_subscriptions_month,
                                      session.reports_subscriptions_year,
                                      request.function,
                                      location_filter=location_filter)
    response.subtitle = SPAN(form_subtitle['subtitle'], ' - ',
                             T('Customers'), ' - ',
                             ssu.get_name())

    form = form_subtitle['form']
    month_chooser = form_subtitle['month_chooser']
    current_month = form_subtitle['current_month']
    submit = form_subtitle['submit']

    # fill the table

    table = TABLE(_class='table table-hover')
    header = THEAD(TR(TH(),
                      TH(T('Customer')),
                      TH(T('Subscription')),
                      TH(T('Payment Method')),
                      TH(T('Start')),
                      TH(),
                      TH(),
                      ))

    table.append(header)

    rows = subscriptions_overview_customers_get_rows(ssuID, session.reports_subscriptions_school_locations_id)
    for i, row in enumerate(rows):
        repr_row = list(rows[i:i+1].render())[0]

        buttons = DIV(_class='btn-group pull-right')

        edit = os_gui.get_button('edit_notext',
                                 URL('customers', 'edit',
                                     args=row.auth_user.id))

        buttons.append(edit)

        location = ''
        if session.show_location:
            location = TD(os_gui.get_label('info', repr_row.auth_user.school_locations_id))

        tr = TR(TD(repr_row.auth_user.thumbsmall, _class='os-customer_image_td'),
                TD(SPAN(row.auth_user.display_name)),
                TD(repr_row.customers_subscriptions.school_subscriptions_id),
                TD(repr_row.customers_subscriptions.payment_methods_id),
                TD(repr_row.customers_subscriptions.Startdate),
                location,
                TD(buttons)
                        )

        table.append(tr)


    back = os_gui.get_button('back', URL('subscriptions_overview'))


    form_ssuID = subscriptions_overview_customers_get_ssu_form(ssuID)


    form = DIV(form, HR(),
               LABEL(T('Subscription')),
               form_ssuID)

    return dict(content=table,
                form=form,
                back=back,
                month_chooser=month_chooser,
                current_month=current_month,
                run_report=submit)


def subscriptions_overview_customers_get_rows(ssuID, locationID=None):
    """
        Returns rows for customers
    """
    # get first and last day of this month
    date = datetime.date(session.reports_subscriptions_year,
                         session.reports_subscriptions_month,
                         1)
    firstdaythismonth = date
    lastdaythismonth = get_last_day_month(date)

    left = [ db.auth_user.on(db.customers_subscriptions.auth_customer_id == \
                             db.auth_user.id)]

    query = (db.customers_subscriptions.school_subscriptions_id == ssuID) & \
            (db.customers_subscriptions.Startdate <= lastdaythismonth) & \
            ((db.customers_subscriptions.Enddate >= firstdaythismonth) |
             (db.customers_subscriptions.Enddate == None))

    if session.show_location:
        if locationID:
            query &= (db.auth_user.school_locations_id == locationID)

    rows = db(query).select(db.customers_subscriptions.ALL,
                            db.auth_user.id,
                            db.auth_user.thumbsmall,
                            db.auth_user.birthday,
                            db.auth_user.trashed,
                            db.auth_user.first_name,
                            db.auth_user.last_name,
                            db.auth_user.display_name,
                            db.auth_user.email,
                            db.auth_user.school_locations_id,
                            left=left,
                            orderby=db.auth_user.display_name)

    return rows


def subscriptions_overview_customers_get_ssu_form(ssuID):
    """
        Returns form to select school subscription
    """
    subsc_query = (db.school_subscriptions.Archived == False)
    form = SQLFORM.factory(
        Field('ssuID',
              db.school_subscriptions,
              default=ssuID,
              requires=IS_IN_DB(db(subsc_query),
                                            'school_subscriptions.id',
                                            '%(Name)s',
                                            zero=None),
              label=T("Subscription filter")),
        )

    select = form.element('select')
    select['_onchange'] = 'this.form.submit();'

    form = DIV(DIV(form.custom.begin,
                   form.custom.widget.ssuID,
                   form.custom.end,
                   _class='col-md-4'),
               _class='row')

    return form



@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'reports_subscriptions'))
def subscriptions_alt_prices():
    """
        Alternative prices for all subscriptions in a month
    """
    response.title = T("Reports")
    response.view = 'reports/subscriptions.html'

    session.customers_back = request.function
    session.invoices_edit_back = 'reports_' + request.function

    # set the session vars for year/month
    subscriptions_process_request_vars()

    # get first and last day of this month
    date = datetime.date(session.reports_subscriptions_year,
                         session.reports_subscriptions_month,
                         1)

    # get month/year selection form and subtitle
    form_subtitle = get_form_subtitle(session.reports_subscriptions_month,
                                            session.reports_subscriptions_year,
                                            request.function)
    response.subtitle = T('Subscriptions') + ' - ' + form_subtitle['subtitle']
    form = form_subtitle['form']
    month_chooser = form_subtitle['month_chooser']
    current_month = form_subtitle['current_month']
    submit = form_subtitle['submit']

    csap = db.customers_subscriptions_alt_prices

    fields = [csap.Description,
              csap.Amount,
              db.customers_subscriptions.id,
              db.customers_subscriptions.auth_customer_id,
              db.customers_subscriptions.school_subscriptions_id,
              db.auth_user.id,
              db.auth_user.trashed,
              db.auth_user.birthday,
              db.auth_user.display_name,
              db.auth_user.thumbsmall,
              db.invoices.id,
              db.invoices.InvoiceID,
              db.invoices_amounts.TotalPriceVAT]

    rows = db.executesql("""SELECT csap.Description,
                                   csap.Amount,
                                   cs.id,
                                   cs.auth_customer_id,
                                   cs.school_subscriptions_id,
                                   au.id,
                                   au.trashed,
                                   au.birthday,
                                   au.display_name,
                                   au.thumbsmall,
                                   i.id,
                                   i.InvoiceID,
                                   inva.TotalPriceVAT
                            FROM customers_subscriptions_alt_prices csap
                            LEFT JOIN customers_subscriptions cs
                                   ON cs.id = csap.customers_subscriptions_id
                            LEFT JOIN auth_user au
                                   ON au.id = cs.auth_customer_id
							LEFT JOIN 
								(SELECT i.id,
										i.InvoiceID,
										i.SubscriptionYear,
                                        i.SubscriptionMonth,
                                        iics.customers_subscriptions_id
									FROM invoices i
									LEFT JOIN invoices_items ii ON ii.invoices_id = i.id 
                                    LEFT JOIN invoices_items_customers_subscriptions iics ON iics.invoices_items_id = ii.id
                                    WHERE i.SubscriptionYear = '{year}' AND i.SubscriptionMonth = '{month}'
								) i ON i.customers_subscriptions_id = cs.id
							LEFT JOIN invoices_amounts inva 
								   ON inva.invoices_id = i.id
                            WHERE csap.SubscriptionYear = '{year}' AND
                                  csap.SubscriptionMonth = '{month}'
                            ORDER BY au.display_name DESC """.format(
                                year=session.reports_subscriptions_year,
                                month=session.reports_subscriptions_month
                            ), fields=fields)

    table = TABLE(_class='table table-striped')

    header = THEAD(TR(TH(),
                      TH(T('Customer')),
                      TH(T('Subscription')),
                      TH(T('Amount')),
                      TH(T('Description')),
                      TH(T('Invoice')),
                      TH(T('Invoice amt.')),
                      TH(), # Warning
                      TH(), # Actions
                      ))

    table.append(header)


    for i, row in enumerate(rows):
        repr_row = list(rows[i:i+1].render())[0]
        repr_casp = repr_row.customers_subscriptions_alt_prices

        edit = os_gui.get_button('edit_notext',
            URL('customers', 'subscription_alt_prices',
                vars={'cuID':row.auth_user.id,
                      'csID':row.customers_subscriptions.id}),
            _class='pull-right')

        invoice_link = A(repr_row.invoices.InvoiceID,
                         _href=URL('invoices', 'edit',
                                   vars={'iID':row.invoices.id}))

        color_class = ''
        warning = ''
        if not (row.customers_subscriptions_alt_prices.Amount ==
                row.invoices_amounts.TotalPriceVAT):
            color_class = 'red'
            msg = T("Alt. price amount and invoice amount doesn't match")
            warning = os_gui.get_label('warning', '!', title=msg)

        tr = TR(TD(repr_row.auth_user.thumbsmall,
                   _class='os-customer_image_td'),
                TD(repr_row.auth_user.display_name),
                TD(repr_row.customers_subscriptions.school_subscriptions_id),
                TD(repr_casp.Amount),
                TD(repr_casp.Description),
                TD(invoice_link),
                TD(repr_row.invoices_amounts.TotalPriceVAT),
                TD(warning),
                TD(edit),
                _class=color_class)

        table.append(tr)

    menu = subscriptions_get_menu(request.function)

    return dict(content=table,
                form=form,
                menu=menu,
                month_chooser=month_chooser,
                current_month=current_month,
                submit=submit)


@auth.requires_login()
def subscriptions_process_request_vars(var=None):
    """
        This function takes the request.vars as a argument and
    """
    today = TODAY_LOCAL
    if 'year' in request.vars:
        year = int(request.vars['year'])
    elif not session.reports_subscriptions_year is None:
        year = session.reports_subscriptions_year
    else:
        year = today.year
    session.reports_subscriptions_year = year
    if 'month' in request.vars:
        month = int(request.vars['month'])
    elif not session.reports_subscriptions_month is None:
        month = session.reports_subscriptions_month
    else:
        month = today.month
    session.reports_subscriptions_month = month

    if 'school_locations_id' in request.vars:
        slID = request.vars['school_locations_id']
    elif not session.reports_subscriptions_school_locations_id is None:
        slID = session.reports_subscriptions_school_locations_id
    else:
        slID = None
    session.reports_subscriptions_school_locations_id = slID

    session.reports_subscriptions = request.function


@auth.requires_login()
def direct_debit_extra_show_current():
    session.reports_ap_year = None
    session.reports_ap_month = None

    redirect(URL('direct_debit_extra'))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'reports_direct_debit_extra'))
def direct_debit_extra():
    response.title = T("Direct debit extra")
    session.customers_back = 'direct_debit_extra'

    response.view = 'reports/general.html'

    today = datetime.date.today()
    if 'year' in request.vars:
        year = int(request.vars['year'])
    elif not session.reports_ap_year is None:
        year = session.reports_ap_year
    else:
        year = today.year
    session.reports_ap_year = year
    if 'month' in request.vars:
        month = int(request.vars['month'])
    elif not session.reports_ap_month is None:
        month = session.reports_ap_month
    else:
        month = today.month
    session.reports_ap_month = month

    date = datetime.date(year,month,1)

    form_subtitle = get_form_subtitle(month, year, request.function)
    response.subtitle = form_subtitle['subtitle']
    form = form_subtitle['form']
    month_chooser = form_subtitle['month_chooser']
    current_month = form_subtitle['current_month']
    submit = form_subtitle['submit']

    # collection
    query = (db.alternativepayments.PaymentYear==date.year) & \
            (db.alternativepayments.PaymentMonth==date.month) & \
            ((db.payment_categories.CategoryType != 1) |
             (db.payment_categories.CategoryType == None))
    rows = db(query).select(db.auth_user.id,
                            db.auth_user.trashed,
                            db.auth_user.thumbsmall,
                            db.auth_user.birthday,
                            db.auth_user.display_name,
                            db.auth_user.date_of_birth,
                            db.alternativepayments.Amount,
                            db.alternativepayments.payment_categories_id,
                            db.alternativepayments.Description,
        left=[db.auth_user.on(db.auth_user.id==\
                              db.alternativepayments.auth_customer_id),
            db.payment_categories.on(db.payment_categories.id==\
                db.alternativepayments.payment_categories_id)],
        orderby=db.alternativepayments.payment_categories_id|\
                db.auth_user.display_name)

    col_total = DIV(T("Total: " + str(len(rows))), _class='right')

    table = TABLE(_class="table table-hover")
    table.append(THEAD(TR(TH(), # image
                    TH(), # name,
                    TH(db.alternativepayments.Amount.label),
                    TH(db.alternativepayments.payment_categories_id.label),
                    TH(db.alternativepayments.Description.label),
                    TH())))

    for row in rows.render():
        edit = os_gui.get_button('edit_notext', URL('customers', 'payments',
                                             vars={'cuID' : row.auth_user.id}))
        table.append(TR(TD(row.auth_user.thumbsmall,
                           _class='os-customer_image_td'),
                        TD(DIV(row.auth_user.display_name),
                           _class="os-customer_name"),
                        TD(row.alternativepayments.Amount),
                        TD(row.alternativepayments.payment_categories_id),
                        TD(row.alternativepayments.Description),
                        TD(DIV(DIV(edit),
                               _class='btn-group pull-right'),
                           _class='td-icons table-vertical-align-middle'),
                        _id=row.auth_user.id))

    collection = table
    # payments
    query = (db.alternativepayments.PaymentYear==date.year) & \
            (db.alternativepayments.PaymentMonth==date.month) & \
            (db.payment_categories.CategoryType == 1)
    rows = db(query).select(db.auth_user.id,
                            db.auth_user.trashed,
                            db.auth_user.thumbsmall,
                            db.auth_user.birthday,
                            db.auth_user.display_name,
                            db.auth_user.date_of_birth,
                            db.alternativepayments.Amount,
                            db.alternativepayments.payment_categories_id,
                            db.alternativepayments.Description,
        left=[db.auth_user.on(db.auth_user.id==\
                              db.alternativepayments.auth_customer_id),
            db.payment_categories.on(db.payment_categories.id==\
                db.alternativepayments.payment_categories_id)],
        orderby=db.alternativepayments.payment_categories_id|\
            db.auth_user.display_name)
    pay_total = DIV(T("Total: " + str(len(rows))), _class='right')

    table = TABLE(_class="table table-hover")
    table.append(THEAD(TR(TH(), # image
                    TH(), # name,
                    TH(db.alternativepayments.Amount.label),
                    TH(db.alternativepayments.payment_categories_id.label),
                    TH(db.alternativepayments.Description.label))))

    for row in rows.render():
        edit = os_gui.get_button('edit_notext', URL('customers', 'payments',
                                             vars={'cuID' : row.auth_user.id}))
        table.append(TR(TD(row.auth_user.thumbsmall,
                           _class='os-customer_image_td'),
                        TD(DIV(row.auth_user.display_name,
                                _class="os-customer_name"),
                            BR()),
                        TD(row.alternativepayments.Amount),
                        TD(row.alternativepayments.payment_categories_id),
                        TD(row.alternativepayments.Description),
                        TD(DIV(DIV(edit),
                               _class='btn-group pull-right'),
                           _class='td-icons table-vertical-align-middle'),
                        _id=row.auth_user.id))

    payments = table

    col_title = H3(T("Collection"))
    pay_title = H3(T('Payments'))

    content = DIV(col_title, collection, col_total,
                  pay_title, payments, pay_total)

    return dict(form=form,
                content=content,
                month_chooser=month_chooser,
                current_month=current_month,
                submit=submit)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'reports_attendance'))
def attendance_classes():
    """
        List classes for a selected month with revenue
    """
    response.title = T("Reports")
    response.view = 'reports/subscriptions.html'

    session.reports_attendance = request.function
    session.reports_teacher_classes_class_revenue_back = 'reports_attendance_classes'
    session.classes_attendance_back = 'reports_attendance_classes'

    form_subtitle = get_form_subtitle(function=request.function)
    response.subtitle = T('Classes attendance')

    year = TODAY_LOCAL.year
    month = TODAY_LOCAL.month
    soID = None
    slID = None
    content = T("Please click 'Run Report' to generate a report. This might take a little while depending on the number of classes in the schedule for the selected month.")
    month_chooser = ''
    if 'year' in request.vars:
        year = int(request.vars['year'])
        month = int(request.vars['month'])
        slID = request.vars['slID']
        soID = request.vars['soID']

        session.reports_att_year = year
        session.reports_att_month = month

        date = datetime.date(year, month, 1)
        firstdaythismonth = date
        next_month = date.replace(day=28) + datetime.timedelta(days=4)  # this will never fail
        lastdaythismonth = next_month - datetime.timedelta(days=next_month.day)
        content = attendance_classes_get_content(firstdaythismonth, lastdaythismonth, slID, soID)

        subtitle = get_month_subtitle(month, year)
        response.subtitle = T('Classes attendance') + ' - ' + subtitle


    form = attendance_classes_get_form(year, month, slID, soID)

    current_month = form_subtitle['current_month']
    submit = form_subtitle['submit']

    menu = attendance_get_menu(request.function)

    return dict(form=form,
                menu=menu,
                content=content,
                current_month='',
                month_chooser='', # Month chooser doesn't work here as we require the form the be submitted before anything happens
                submit=submit)


def attendance_classes_get_content(date_start, date_end, slID, soID):
    """
        Return list of classes with revenue for a selected period
    """
    one_day = datetime.timedelta(days=1)
    current_date = date_start

    if len(ORGANIZATIONS) > 2:
        header = THEAD(TR(TH('Date'),
                          TH('Start'),
                          TH('Class Type'),
                          TH('Location'),
                          TH('Organization'),
                          TH('Teacher'),
                          TH('Teacher2'),
                          TH('Revenue'),
                          TH()))
    else:
        header = THEAD(TR(TH('Date'),
                          TH('Start'),
                          TH('Class Type'),
                          TH('Location'),
                          TH(),
                          TH('Teacher'),
                          TH('Teacher2'),
                          TH('Revenue'),
                          TH()))
    table = TABLE(header, _class='table table-hover table-striped')
    # Find all classes starting on the first day of the month
    while current_date <= date_end:
        # get list of today's classes.
        class_schedule = ClassSchedule(
            date=current_date,
            filter_id_school_location=slID,
            filter_id_sys_organization=soID,
        )

        date_formatted = current_date.strftime(DATE_FORMAT)

        rows = class_schedule.get_day_rows()
        for i, row in enumerate(rows):
            repr_row = list(rows[i:i + 1].render())[0]
            revenue = teacher_classes_get_class_revenue_total(row.classes.id, current_date)

            if row.classes_otc.Status == 'cancelled':
                if row.classes_otc.Description:
                    amount = max_string_length(row.classes_otc.Description, 12)
                else:
                    amount = T('Cancelled')
            elif row.school_holidays.id:
                amount = max_string_length(row.school_holidays.Description, 12)
            else:
                amount = A(represent_decimal_as_amount(revenue['revenue_in_vat']),
                           _href=URL('teacher_classes_class_revenue', vars={'clsID':row.classes.id,
                                                                          'date':date_formatted}),
                           _target='_blank')

            organization = ''
            if len(ORGANIZATIONS) > 2:
                organization = repr_row.classes.sys_organizations_id or ''

            tr = TR(
                TD(current_date.strftime(DATE_FORMAT)),
                TD(repr_row.classes.Starttime),
                TD(repr_row.classes.school_classtypes_id),
                TD(repr_row.classes.school_locations_id),
                TD(organization),
                TD(repr_row.classes_teachers.auth_teacher_id),
                TD(repr_row.classes_teachers.auth_teacher_id2),
                TD(amount),
                TD(os_gui.get_button('next_no_text',
                                     URL('classes', 'attendance', vars={'clsID': row.classes.id,
                                                                        'date': date_formatted}),
                                     _class='pull-right',
                                     _target='_blank'))
            )

            table.append(tr)

        current_date += one_day

    return table


def attendance_classes_get_form(year=TODAY_LOCAL.year,
                                month=TODAY_LOCAL.month,
                                slID=None,
                                soID=None):
    """
    :param month: int 1 - 12
    :param year: int 1900 - 2999
    :param slID: db.school_locations.id
    :param soID: db.school_organizations.id
    :return: classes attendance filter form
    """
    loc_query = (db.school_locations.Archived == False)
    so_query = (db.sys_organizations.Archived == False)

    months = get_months_list()

    form = SQLFORM.factory(
        Field('month',
               requires=IS_IN_SET(months, zero=None),
               default=month,
               label=T("")),
        Field('year', 'integer',
              default=year,
              label=T("")),
        Field('slID',
              requires=IS_EMPTY_OR(IS_IN_DB(db(loc_query),
                                            'school_locations.id',
                                            '%(Name)s',
                                            zero=T('All locations'))),
              default=slID,
              label=T("")),
        Field('soID',
              requires=IS_EMPTY_OR(IS_IN_DB(db(so_query),
                                            'sys_organizations.id',
                                            '%(Name)s',
                                            zero=T('All organizations'))),
              default=soID,
              label=T("")),
        submit_button=T("Run report")
        )
    form.attributes['_name']  = 'form_select_date'
    form.attributes['_class'] = 'overview_form_select_date'

    input_month = form.element('select[name=month]')
    input_year = form.element('input[name=year]')
    input_year.attributes['_type']     = 'number'

    result = set_form_id_and_get_submit_button(form, 'MainForm')
    form = result['form']
    submit = result['submit']

    location = ''
    if session.show_location:
        location = form.custom.widget.slID

    organization = ''
    if len(ORGANIZATIONS) > 2:
        organization = form.custom.widget.soID

    form = DIV(DIV(XML('<form id="MainForm" action="#" enctype="multipart/form-data" method="post">'),
                   form.custom.widget.month,
                   form.custom.widget.year,
                   location,
                   organization,
                   form.custom.end,
                   _class='col-md-4'),
               _class = 'row')

    return form


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'classes_attendance'))
def attendance_review_requested():
    """
        List Problem Check-ins that occured
    """
    response.title = T("Reports")
    response.subtitle = T("Classes attendance")
    response.view = 'general/tabs_menu.html'

    # Redirect back here after reviewing a check-in
    session.classes_attendance_signin_back = 'reports_attendance_review_requested'

    header = THEAD(TR(TH(''),
                     TH("Customer"),
                     TH(db.classes.school_classtypes_id.label),
                     TH(db.classes_attendance.ClassDate.label),
                     TH(db.classes.Starttime.label),
                     TH(db.classes.school_locations_id.label),
                     TH())  # buttons
                  )

    table = TABLE(header, _class='table table-hover table-striped')

    left = [
        db.classes.on(db.classes_attendance.classes_id == db.classes.id),
        db.auth_user.on(db.classes_attendance.auth_customer_id == db.auth_user.id)
    ]

    query = (db.classes_attendance.AttendanceType == 5)
    rows = db(query).select(
        db.classes_attendance.ALL,
        db.classes.ALL,
        db.auth_user.ALL,
        left=left
    )
    # print rows
    for i, row in enumerate(rows):
        repr_row = list(rows[i:i + 1].render())[0]

        resolve = os_gui.get_button(
            'noicon',
             URL('classes', 'attendance_booking_options',
                 vars = {'clsID': row.classes.id,
                         'cuID': row.classes_attendance.auth_customer_id,
                         'date': row.classes_attendance.ClassDate.strftime(DATE_FORMAT)}),
             title= T("Review"),
             _class= 'pull-right'
        )

        tr = TR(TD(repr_row.auth_user.thumbsmall),
                TD(repr_row.classes_attendance.auth_customer_id),
                TD(repr_row.classes.school_classtypes_id),
                TD(repr_row.classes_attendance.ClassDate),
                TD(repr_row.classes.Starttime),
                TD(repr_row.classes.school_locations_id),
                TD(resolve))

        table.append(tr)

    menu = attendance_get_menu(request.function)

    return dict(
        content=table,
        menu=menu
    )


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'classes_attendance'))
def attendance_reconcile_later():
    """
        List Problem Check-ins that occured
    """
    response.title = T("Reports")
    response.subtitle = T("Classes attendance")
    response.view = 'general/tabs_menu.html'

    # Redirect back to class after reconciling a check-in
    session.classes_attendance_signin_back = 'attendance_list'

    header = THEAD(TR(TH(''),
                     TH("Customer"),
                     TH(db.classes.school_classtypes_id.label),
                     TH(db.classes_attendance.ClassDate.label),
                     TH(db.classes.Starttime.label),
                     TH(db.classes.school_locations_id.label),
                     TH())  # buttons
                  )

    table = TABLE(header, _class='table table-hover table-striped')

    left = [
        db.classes.on(db.classes_attendance.classes_id == db.classes.id),
        db.auth_user.on(db.classes_attendance.auth_customer_id == db.auth_user.id)
    ]

    query = (db.classes_attendance.AttendanceType == 6)
    rows = db(query).select(
        db.classes_attendance.ALL,
        db.classes.ALL,
        db.auth_user.ALL,
        left=left
    )
    # print rows
    for i, row in enumerate(rows):
        repr_row = list(rows[i:i + 1].render())[0]


        #TODO: Add correct button

        resolve = os_gui.get_button(
            'noicon',
             URL('classes', 'attendance_reconcile_later_to_dropin',
                 vars = {'clattID': row.classes_attendance.id }),
             title= T("Reconcile"),
             _class= 'pull-right',
            _target="_blank"
        )

        tr = TR(TD(repr_row.auth_user.thumbsmall),
                TD(repr_row.classes_attendance.auth_customer_id),
                TD(repr_row.classes.school_classtypes_id),
                TD(repr_row.classes_attendance.ClassDate),
                TD(repr_row.classes.Starttime),
                TD(repr_row.classes.school_locations_id),
                TD(resolve))

        table.append(tr)

    menu = attendance_get_menu(request.function)

    return dict(
        content=table,
        menu=menu
    )


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'reports_attendance'))
def attendance_classtypes():
    """
        Shows how many customers are attending a class type in a given month
    """
    response.title = T("Reports")
    response.view = 'reports/subscriptions.html'

    session.reports_attendance = request.function

    today = TODAY_LOCAL
    if 'year' in request.vars:
        year = int(request.vars['year'])
    elif not session.reports_att_year is None:
        year = session.reports_att_year
    else:
        year = today.year
    session.reports_att_year = year
    if 'month' in request.vars:
        month = int(request.vars['month'])
    elif not session.reports_att_month is None:
        month = session.reports_att_month
    else:
        month = today.month
    session.reports_att_month = month

    date = datetime.date(year,month,1)
    firstdaythismonth = date
    next_month = date.replace(day=28) + datetime.timedelta(days=4)  # this will never fail
    lastdaythismonth = next_month - datetime.timedelta(days=next_month.day)

    form_subtitle = get_form_subtitle(month, year, request.function)
    response.subtitle = T('Attendance classtypes') + ' - ' + form_subtitle['subtitle']
    form = form_subtitle['form']
    month_chooser = form_subtitle['month_chooser']
    current_month = form_subtitle['current_month']
    submit = form_subtitle['submit']

    count = db.classes.school_classtypes_id.count()
    query = (db.classes_attendance.ClassDate >= firstdaythismonth) & \
            (db.classes_attendance.ClassDate <= lastdaythismonth)

    rows = db(query).select(
        db.school_classtypes.id,
        db.school_classtypes.Name,
        count,
        left=[db.classes.on(db.classes.id== \
                            db.classes_attendance.classes_id),
              db.school_classtypes.on(
                db.school_classtypes.id == \
                db.classes.school_classtypes_id)],
        groupby=db.school_classtypes.id,
        orderby=~count)

    header = THEAD(TR(TH(T('Class type')),
                      TH(T('Attendance count')),
                      TH()))
    table = TABLE(header,
                  _class='table table-striped table-hover')

    total = 0
    for row in rows:
        link_mailinglist = A(
            SPAN(os_gui.get_fa_icon('fa-envelope-o'), ' ', T("Mailinglist")),
            _href=URL('attendance_classtypes_export_excel_mailinglist', vars={'cltID':row.school_classtypes.id}),
            #_class='textalign_left',
            _title=T('Mailinglist for customers who attended this type of class this month'))

        links = [link_mailinglist]

        export_row = os_gui.get_dropdown_menu(
                links = links,
                btn_text = '',
                btn_icon = 'download',
                btn_size = 'btn-sm',
                menu_class='btn-group pull-right')



        tr = TR(TD(row.school_classtypes.Name),
                TD(row[count]),
                TD(export_row))
        table.append(tr)

        total += row[count]

    footer = TFOOT(TR(TD(T("Total"), _class='bold'),
                      TD(total, _class='bold'),
                      TD()))

    table.append(footer)

    menu = attendance_get_menu(request.function)

    return dict(form=form,
                menu=menu,
                content=table,
                current_month=current_month,
                month_chooser=month_chooser,
                submit=submit)


def attendance_classtypes_export_excel_mailinglist():
    """
        :return: Openpyxl.worbook with email addresses of customers attending a certain class type in a chosen month
    """
    cltID = request.vars['cltID']
    month = session.reports_att_month
    year  = session.reports_att_year


    month_begin = datetime.date(int(year), int(month), 1)
    month_end    = get_last_day_month(month_begin)

    # create filestream
    stream = io.BytesIO()

    classtype = db.school_classtypes(cltID)
    title = max_string_length(classtype.Name, 30)

    wb = openpyxl.workbook.Workbook(write_only=True)
    ws = wb.create_sheet(title=title)
    headers = ["CustomerID",
               "First name",
               "Last name",
               "Email",
               "Telephone",
               "Mobile"]
    ws.append(headers)

    # write the sheets sorted by classtype

    left = [ db.classes.on(db.classes_attendance.classes_id == db.classes.id),
             db.auth_user.on(db.classes_attendance.auth_customer_id == db.auth_user.id) ]


    query = (db.classes.school_classtypes_id == cltID) & \
            (db.classes_attendance.ClassDate >= month_begin) & \
            (db.classes_attendance.ClassDate <= month_end)

    rows = db(query).select(db.auth_user.id,
                            db.auth_user.first_name,
                            db.auth_user.last_name,
                            db.auth_user.email,
                            db.auth_user.phone,
                            db.auth_user.mobile,
                            left=left,
                            distinct=True)

    for row in rows:
        ws.append([
            row.id,
            row.first_name,
            row.last_name,
            row.email,
            row.phone,
            row.mobile
        ])

    wb.save(stream)

    fname = T("Mailinglist") + '.xlsx'
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-disposition'] = 'attachment; filename=' + fname

    return stream.getvalue()


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'reports_attendance'))
def attendance_show_current():
    session.reports_att_year = None
    session.reports_att_month = None

    redirect(URL(session.reports_attendance))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'reports_attendance'))
def attendance_organizations_list_attendance():
    """
        List attendance for a selected month and organization
    """
    response.view = 'general/only_content.html'
    response.title = T('Reports')

    year = int(request.vars['year'])
    month = int(request.vars['month'])
    soID = request.vars['soID']

    so = db.sys_organizations(soID)

    months = get_months_list()

    for m in months:
        if m[0] == month:
            month_title = m[1]

    response.subtitle = SPAN(T('Attendance'), ' ', so.Name, ' - ', month_title, ' ', year)


    rows = attendance_organizations_list_attendance_get_rows(year, month, soID, customers=True)

    header = THEAD(TR(TH(),
                      TH(T('Customer')),
                      TH(T('Date')),
                      TH(T('Time')),
                      TH(T('Location')),
                      TH(T('Class')),
                      TH(T('Card/Subscription')),
                   ))
    table = TABLE(header, _class='table table-hover')
    for i, row in enumerate(rows):
        repr_row = list(rows[i:i + 1].render())[0]

        if row.school_classcards.Name:
            attended_using = repr_row.school_classcards.Name
            rsoID = row.school_classcards.sys_organizations_id
        else:
            attended_using = repr_row.classes_attendance.customers_subscriptions_id
            rsoID = row.school_subscriptions.sys_organizations_id


        if rsoID and int(rsoID) != int(soID):
            attended_using = SPAN(attended_using, _class='orange bold')


        time = SPAN(repr_row.classes.Starttime, ' - ', repr_row.classes.Endtime)

        tr = TR(TD(repr_row.auth_user.thumbsmall,
                   _class='os-customer_image_td'),
                TD(repr_row.auth_user.display_name),
                TD(repr_row.classes_attendance.ClassDate),
                TD(time),
                TD(repr_row.classes.school_locations_id),
                TD(repr_row.classes.school_classtypes_id),
                TD(attended_using),
                )

        table.append(tr)

    link_all_customers = A(
        SPAN(os_gui.get_fa_icon('fa-check-square-o'), ' ', T("Attendance list")),
        _href=URL('attendance_organizations_list_attendance_export', vars={'soID':soID,
                                                                           'year':year,
                                                                           'month':month}),
        _class='textalign_left',
        _title=T('Export Attendance listed on this page'))

    links = [ link_all_customers ]

    export = os_gui.get_dropdown_menu(
            links = links,
            btn_text = '',
            btn_icon = 'download',
            menu_class='pull-right' )


    back = os_gui.get_button('back', URL('attendance_organizations', vars={'year':year,
                                                                           'month':month}))

    return dict(content=table, header_tools=export, back=back)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'reports_attendance'))
def attendance_organizations_list_attendance_export():
    """
        Export attendance for an organization
    """
    soID = request.vars['soID']
    year = int(request.vars['year'])
    month = int(request.vars['month'])

    stream = io.BytesIO()

    so = db.sys_organizations(soID)

    wb = openpyxl.workbook.Workbook(write_only=True)
    title = str(so.Name) + ' ' + str(year) + '-' + str(month)
    ws = wb.create_sheet(title=title)

    header = [
        "Customer",
        "Date",
        "Time",
        "Location",
        "Class",
        "Card/Subscription",
        "Other Organization"
    ]

    ws.append(header)

    rows = attendance_organizations_list_attendance_get_rows(year, month, soID, customers=True)
    for i, row in enumerate(rows):
        repr_row = list(rows[i:i + 1].render())[0]

        if row.school_classcards.Name:
            attended_using = row.school_classcards.Name
            rsoID = row.school_classcards.sys_organizations_id
        else:
            attended_using = row.school_subscriptions.Name
            rsoID = row.school_subscriptions.sys_organizations_id

        other_organization = 'No'
        if rsoID and int(rsoID) != int(soID):
            other_organization = 'Yes'

        data = [
            repr_row.auth_user.display_name,
            row.classes_attendance.ClassDate,
            repr_row.classes.Starttime,
            repr_row.classes.school_locations_id,
            repr_row.classes.school_classtypes_id,
            attended_using,
            other_organization
        ]

        ws.append(data)

    fname = 'Attendance.xlsx'
    wb.save(stream)

    response.headers['Content-Type']='application/vnd.ms-excel'
    response.headers['Content-disposition']='attachment; filename=' + fname

    return stream.getvalue()


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'reports_attendance'))
def attendance_organizations():
    """
        Shows how many customers are attending a class type in a given month
    """
    response.title = T("Reports")
    response.view = 'reports/subscriptions.html'

    session.reports_attendance = request.function

    today = TODAY_LOCAL
    if 'year' in request.vars:
        year = int(request.vars['year'])
    elif not session.reports_att_year is None:
        year = session.reports_att_year
    else:
        year = today.year
    session.reports_att_year = year
    if 'month' in request.vars:
        month = int(request.vars['month'])
    elif not session.reports_att_month is None:
        month = session.reports_att_month
    else:
        month = today.month
    session.reports_att_month = month

    date = datetime.date(year,month,1)
    firstdaythismonth = date
    next_month = date.replace(day=28) + datetime.timedelta(days=4)  # this will never fail
    lastdaythismonth = next_month - datetime.timedelta(days=next_month.day)

    form_subtitle = get_form_subtitle(month, year, request.function)
    response.subtitle = T('Attendance - Organizations') + ' - ' + form_subtitle['subtitle']
    form = form_subtitle['form']
    month_chooser = form_subtitle['month_chooser']
    current_month = form_subtitle['current_month']
    submit = form_subtitle['submit']

    table = attendance_organizations_get_content(year, month)

    tools = attendance_organizations_get_tools()

    menu = attendance_get_menu(request.function)

    return dict(form=form,
                menu=menu,
                content=table,
                current_month=current_month,
                header_tools=tools,
                month_chooser=month_chooser,
                submit=submit)


def attendance_organizations_get_tools(var=None):
    """
        Returns tools for schedule
    """
    att_tools = []

    # teacher holidays
    permission = auth.has_membership(group_id='Admins') or \
                 auth.has_permission('read', 'teacher_holidays')

    if permission:
        resolve_prices = A(os_gui.get_fa_icon('fa-dollar'),
                           T("Resolve class prices"),
                           _href=URL('reports', 'attendance_organizations_res_prices'),
                           _title=T('Set organization prices to be used for resolve in this report'))
        att_tools.append(resolve_prices)

    # get menu
    tools = os_gui.get_dropdown_menu(att_tools,
                                     '',
                                     btn_size='',
                                     btn_icon='wrench',
                                     menu_class='pull-right'
                                     )

    return tools


def attendance_organizations_get_content(year, month):
    """
        Get and display data for attendance_organizations
    """
    content = DIV()

    query = (db.sys_organizations.Archived == False)
    rows = db(query).select(db.sys_organizations.ALL,
                            orderby=db.sys_organizations.Name)

    data = DIV(H4(T('Data')), _class='col-md-6')

    # get class prices
    class_prices = {}
    for organization in ORGANIZATIONS:
        if organization == 'default':
            continue

        class_prices[organization] = ORGANIZATIONS[organization]['ReportsClassPrice']

    for i, row in enumerate(rows):
        soID = row.id

        data.append(B(row.Name))
        data.append(BR())
        data.append(DIV(DIV(T('Resolve class price'),
                            _class='col-md-5'),
                        DIV(represent_decimal_as_amount(class_prices[soID]),
                            _class='col-md-5'),
                        _class='row'))
        data.append(attendance_organizations_get_content_classes_taught(year, month, soID, formatted=True))
        data.append(attendance_organizations_get_content_attendance_total(year, month, soID, formatted=True))
        result = attendance_organizations_get_content_attendance_other_organizations(year, month, soID, formatted=True)
        for oa in result:
            data.append(oa)
        result = attendance_organizations_get_content_organization_resolve(year, month, soID, rows, formatted=True)
        for resolve in result:
            data.append(resolve)

        if i + 1 < len(rows):
            data.append(HR())
        else:
            data.append(BR())

    content.append(data)

    resolve = DIV(_class='col-md-6')
    resolve.append(H4(T('Resolve')))
    resolve.append(attendance_organizations_get_content_resolve_total(year,
                                                                      month,
                                                                      rows,
                                                                      formatted=True))

    content.append(resolve)

    return content


def attendance_organizations_get_content_classes_taught(year, month, soID, formatted=False):
    """
        Returns total classes taught for an organization
    """
    from general_helpers import get_number_weekdays_in_month

    # get all classes for organization
    query = (db.classes.sys_organizations_id == soID)
    rows = db(query).select(db.classes.id, db.classes.Week_day)

    total_classes = 0
    # for each class get weekday and total of weekdays in month
    for row in rows:
        total_classes += get_number_weekdays_in_month(year, month, row.Week_day)

    if not formatted:
        return total_classes
    else:
        return DIV(DIV(T('Classes taught'),
                       _class='col-md-5'),
                   DIV(total_classes,
                       _class='col-md-5'),
                   _class='row')


def attendance_organizations_list_attendance_get_rows(year, month, soID, customers=False):
    """
        :param year: int
        :param month: int
        :param soID: db.sys_organizations.id
        :return: rows
    """
    first_day = datetime.date(year, month, 1)
    last_day = get_last_day_month(first_day)

    left = [ db.classes.on(db.classes_attendance.classes_id == db.classes.id),
             db.customers_classcards.on(db.classes_attendance.customers_classcards_id ==
                                        db.customers_classcards.id),
             db.school_classcards.on(db.customers_classcards.school_classcards_id ==
                                     db.school_classcards.id),
             db.customers_subscriptions.on(db.classes_attendance.customers_subscriptions_id ==
                                           db.customers_subscriptions.id),
             db.school_subscriptions.on(db.customers_subscriptions.school_subscriptions_id ==
                                        db.school_subscriptions.id) ]


    if customers:
        left.append( db.auth_user.on(db.classes_attendance.auth_customer_id == db.auth_user.id))

    query = (db.classes.sys_organizations_id == soID) & \
            (db.classes_attendance.ClassDate >= first_day) & \
            (db.classes_attendance.ClassDate <= last_day)

    if customers:
        rows = db(query).select(db.classes_attendance.id,
                                db.classes_attendance.ClassDate,
                                db.classes_attendance.customers_classcards_id,
                                db.classes_attendance.customers_subscriptions_id,
                                db.classes.Starttime,
                                db.classes.Endtime,
                                db.classes.school_locations_id,
                                db.classes.school_classtypes_id,
                                db.auth_user.id,
                                db.auth_user.thumbsmall,
                                db.auth_user.trashed,
                                db.auth_user.birthday,
                                db.auth_user.display_name,
                                db.school_classcards.Name,
                                db.school_classcards.sys_organizations_id,
                                db.school_subscriptions.Name,
                                db.school_subscriptions.sys_organizations_id,
                                left=left,
                                orderby=db.classes_attendance.ClassDate|\
                                        db.auth_user.display_name)
    else:
        rows = db(query).select(db.classes_attendance.id,
                                left=left)

    return rows


def attendance_organizations_get_content_attendance_total(year, month, soID, formatted=False):
    """
        Returns total attendance for an organization
    """
    rows = attendance_organizations_list_attendance_get_rows(year, month, soID)

    if not formatted:
        return len(rows)
    else:
        return DIV(DIV(T('Total attendance'),
                       _class='col-md-5'),
                   DIV(A(len(rows),
                         _href=URL('attendance_organizations_list_attendance', vars={'year':year,
                                                                                     'month':month,
                                                                                     'soID':soID})),
                       _class='col-md-5'),
                   _class='row')


def attendance_organizations_get_content_attendance_other_organizations(year, month, soID, formatted=False):
    """
        Returns attendance from other organizations
    """
    first_day = datetime.date(year, month, 1)
    last_day = get_last_day_month(first_day)

    ##
    # Count classcards
    ##
    count = db.school_classcards.sys_organizations_id.count()

    left = [db.classes.on(db.classes_attendance.classes_id == db.classes.id),
            db.customers_classcards.on(db.classes_attendance.customers_classcards_id == db.customers_classcards.id),
            db.school_classcards.on(db.customers_classcards.school_classcards_id == db.school_classcards.id),
            db.sys_organizations.on(db.school_classcards.sys_organizations_id == db.sys_organizations.id)]

    query = (db.classes.sys_organizations_id == soID) & \
            (db.classes_attendance.ClassDate >= first_day) & \
            (db.classes_attendance.ClassDate <= last_day) & \
            ((db.school_classcards.sys_organizations_id != soID) &
             (db.school_classcards.sys_organizations_id != None))

    rows = db(query).select(db.sys_organizations.Name,
                            db.school_classcards.sys_organizations_id,
                            count,
                            left=left,
                            groupby=db.school_classcards.sys_organizations_id,
                            orderby=db.sys_organizations.Name)

    data = []
    for row in rows:
        data.append({'Name': row.sys_organizations.Name,
                     'sys_organizations_id': row.school_classcards.sys_organizations_id,
                     'count': row[count]})

    ##
    # Count subscriptions
    ##
    count = db.school_subscriptions.sys_organizations_id.count()
    left = [db.classes.on(db.classes_attendance.classes_id == db.classes.id),
            db.customers_subscriptions.on(db.classes_attendance.customers_subscriptions_id ==
                                          db.customers_subscriptions.id),
            db.school_subscriptions.on(db.customers_subscriptions.school_subscriptions_id ==
                                       db.school_subscriptions.id),
            db.sys_organizations.on(db.school_subscriptions.sys_organizations_id == db.sys_organizations.id)]

    query = (db.classes.sys_organizations_id == soID) & \
            (db.classes_attendance.ClassDate >= first_day) & \
            (db.classes_attendance.ClassDate <= last_day) & \
            ((db.school_subscriptions.sys_organizations_id != soID) &
             (db.school_subscriptions.sys_organizations_id != None))

    rows = db(query).select(db.sys_organizations.Name,
                            db.school_subscriptions.sys_organizations_id,
                            count,
                            left=left,
                            groupby=db.school_subscriptions.sys_organizations_id,
                            orderby=db.sys_organizations.Name)

    for row in rows:
        create_new_record = True
        # Loop through class card data, if found, add subscription count to class card count
        for oa in data:
            if oa['sys_organizations_id'] == row.school_subscriptions.sys_organizations_id:
                oa['count'] += row[count]
                create_new_record = False
                break

        # If not found, add entry
        if create_new_record:
            data.append({'Name': row.sys_organizations.Name,
                         'sys_organizations_id': row.school_subscriptions.sys_organizations_id,
                         'count': row[count]})

    if not formatted:
        return data
    else:
        formatted_data = []
        for oa in data:
            # oa = other attendance
            formatted_data.append(DIV(DIV(T('Attendance from'), ' ',
                                          oa['Name'],
                                          _class='col-md-5'),
                                      DIV(oa['count'],
                                          _class='col-md-5'),
                                      _class='row'))
        return formatted_data


def attendance_organizations_get_content_organization_resolve(year, month, soID, organizations_rows, formatted=False):
    """
        Return resolve for organization
        eg. StudioX owns...
            StudioY owns...
    """
    # get class prices
    class_prices = {}
    for organization in ORGANIZATIONS:
        if organization == 'default':
            continue

        class_prices[organization] = ORGANIZATIONS[organization]['ReportsClassPrice'] or 0


    first_day = datetime.date(year, month, 1)
    last_day = get_last_day_month(first_day)

    count = db.school_classcards.sys_organizations_id.count()

    left = [ db.classes.on(db.classes_attendance.classes_id == db.classes.id),
             db.customers_classcards.on(db.classes_attendance.customers_classcards_id == db.customers_classcards.id),
             db.school_classcards.on(db.customers_classcards.school_classcards_id == db.school_classcards.id),
             db.customers_subscriptions.on(db.classes_attendance.customers_subscriptions_id ==
                                           db.customers_subscriptions.id),
             db.school_subscriptions.on(db.customers_subscriptions.school_subscriptions_id ==
                                        db.school_subscriptions.id),
             db.sys_organizations.on(db.school_classcards.sys_organizations_id == db.sys_organizations.id) ]

    query = (db.classes.sys_organizations_id == soID) & \
            (db.classes_attendance.ClassDate >= first_day) & \
            (db.classes_attendance.ClassDate <= last_day) & \
            (((db.school_classcards.sys_organizations_id != soID) &
              (db.school_classcards.sys_organizations_id != None)) |
             (db.school_subscriptions.sys_organizations_id != soID) &
              (db.school_subscriptions.sys_organizations_id != None))
            #TODO: update query to include sys_organizations_id from subscriptions

    rows = db(query).select(db.classes_attendance.id,
                            db.sys_organizations.Name,
                            db.classes.sys_organizations_id,
                            db.school_classcards.sys_organizations_id,
                            db.school_subscriptions.sys_organizations_id,
                            left=left,
                            orderby=db.sys_organizations.Name)

    resolve = {}

    for row in rows:
        class_price = class_prices[row.classes.sys_organizations_id]
        if row.school_classcards.sys_organizations_id:
            key = int(row.school_classcards.sys_organizations_id)
        else:
            key = int(row.school_subscriptions.sys_organizations_id)
        try:
            resolve[key] += class_price
        except KeyError:
            resolve[key] = class_price

    if not formatted:
        return resolve
    else:
        rows = organizations_rows
        resolve_formatted = []
        for row in rows:
            if row.id  == soID:
                # Don't show owed amounts for own ID
                continue

            try:
                amount = resolve[int(row.id)]

                if amount > 0:
                    amount = represent_decimal_as_amount(amount)
                    resolve_formatted.append(DIV(DIV(row.Name, ' ', T('owes'),
                                                       _class='col-md-5'),
                                                 DIV(amount,
                                                     _class='col-md-5'),
                                                 _class='row'))
            except KeyError:
                pass

        return resolve_formatted


def attendance_organizations_get_content_resolve_total(year, month, organizations_rows, formatted=False):
    """
        Total resolve between all studios
    """
    resolve = {}
    for row in organizations_rows:
        resolve[row.id] = {'owed': attendance_organizations_get_content_organization_resolve(year,
                                                                                             month,
                                                                                             row.id,
                                                                                             organizations_rows,
                                                                                             formatted=False),
                           'Name':row.Name}

    if not formatted:
        return resolve
    else:
        resolve_formatted = DIV()
        for row in organizations_rows:
            # row.id is current organization
            resolve_formatted.append(B(row.Name))
            resolve_formatted.append(BR())

            for org in organizations_rows: # go over organizations again to calculate resolve
                try:
                    owed = resolve[row.id]['owed'][org.id]
                except KeyError:
                    owed = 0

                try:
                    owes = resolve[org.id]['owed'][row.id]
                except KeyError:
                    owes = 0

                total = owes - owed

                if total > 0:
                    amount = represent_decimal_as_amount(total)
                    resolve_formatted.append(DIV(DIV(row.Name, ' ', T('owes'), ' ', org.Name,
                                                     _class='col-md-5'),
                                                 DIV(amount,
                                                     _class='col-md-5'),
                                                 _class='row'))

        return resolve_formatted


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'sys_organizations'))
def attendance_organizations_res_prices():
    """
        List organizations and resolve class prices
    """
    response.title = T("Reports")
    response.subtitle = T("Attendance organizations - Resolve prices")
    response.view = 'general/only_content.html'


    db.sys_organizations.ReportsClassPrice.readable = True
    db.sys_organizations.ReportsClassPrice.writable = True

    query = (db.sys_organizations.Archived == False)

    # TODO: check if there's a default organization, if not display a message

    fields = [db.sys_organizations.Archived,
              db.sys_organizations.Name,
              db.sys_organizations.ReportsClassPrice]
    links = [lambda row: os_gui.get_button('edit',
                                           URL('attendance_organizations_res_price_edit', vars={'soID': row.id}),
                                           T("Edit resolve price this organization"))]
    grid = SQLFORM.grid(query,
                        fields=fields,
                        links=links,
                        create=False,
                        editable=False,
                        deletable=False,
                        details=False,
                        searchable=False,
                        csv=False,
                        ondelete=cache_clear_sys_organizations,
                        orderby=db.sys_organizations.Name,
                        ui=grid_ui)
    grid.element('.web2py_counter', replace=None)  # remove the counter
    grid.elements('span[title=Delete]', replace=None)  # remove text from delete button


    back = os_gui.get_button('back', URL('reports', 'attendance_organizations'))

    return dict(content=grid, back=back)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'sys_organizations'))
def attendance_organizations_res_price_edit():
    """
        Set resolve class price for an organization
    """
    soID = request.vars['soID']

    so = db.sys_organizations(soID)

    response.title = T("Reports")
    response.subtitle = T('Attendance organizations - Set revolve price for ') + so.Name
    response.view = 'general/only_content.html'

    for field in db.sys_organizations:
        field.readable = False
        field.writable = False

    db.sys_organizations.ReportsClassPrice.readable = True
    db.sys_organizations.ReportsClassPrice.writable = True

    return_url = URL('attendance_organizations_res_prices')

    crud.messages.submit_button = T("Save")
    crud.messages.record_updated = T("Saved")
    crud.settings.update_next = return_url
    crud.settings.update_onaccept = cache_clear_sys_organizations
    crud.settings.formstyle = 'bootstrap3_stacked'
    form = crud.update(db.sys_organizations, soID)

    form_id = "MainForm"
    form_element = form.element('form')
    form['_id'] = form_id

    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = form_id

    textareas = form.elements('textarea')
    for textarea in textareas:
        textarea['_class'] += ' tmced'

    submit = form.element('input[type=submit]')

    back = os_gui.get_button('back', return_url)

    return dict(content=form, back=back, save=submit)


# @auth.requires(auth.has_membership(group_id='Admins') or \
#                auth.has_permission('read', 'reports_attendance'))
# def attendance_subcription_exceeded():
#     """
#         Show list of customers who exceeded the number of classes allowed by
#         their subscription
#     """
#     response.title = T("Reports")
#     response.view = 'reports/subscriptions.html'
#
#     session.reports_attendance = request.function
#     session.customers_back = 'reports_attendance_subscription_exceeded'
#
#     today = datetime.date.today()
#     if 'year' in request.vars:
#         year = int(request.vars['year'])
#     elif not session.reports_att_year is None:
#         year = session.reports_att_year
#     else:
#         year = today.year
#     session.reports_att_year = year
#     if 'month' in request.vars:
#         month = int(request.vars['month'])
#     elif not session.reports_att_month is None:
#         month = session.reports_att_month
#     else:
#         month = today.month
#     session.reports_att_month = month
#
#     date = datetime.date(year,month,1)
#     firstdaythismonth = date
#     next_month = date.replace(day=28) + datetime.timedelta(days=4)  # this will never fail
#     lastdaythismonth = next_month - datetime.timedelta(days=next_month.day)
#
#     form_subtitle = get_form_subtitle(month, year, request.function)
#     response.subtitle = T('Attendance - subscription exceeded') + ' - ' + form_subtitle['subtitle']
#     form = form_subtitle['form']
#     month_chooser = form_subtitle['month_chooser']
#     current_month = form_subtitle['current_month']
#     submit = form_subtitle['submit']
#
#     header = THEAD(TR(TH(),
#                       TH(T('Customer')),
#                       TH(T('Class Date')),
#                       TH(T('Location')),
#                       TH(T('Class type')),
#                       TH(T('Time')),
#                       TH(T('Subscription')),
#                       TH(T('Classes allowed')),
#                       TH(T('Classes taken')),
#                       ))
#     table = TABLE(header, _class='table table-hover table-striped')
#
#
#     # Fill the main page
#     left = [
#         db.customers_subscriptions.on(
#             db.customers_subscriptions_exceeded.customers_subscriptions_id ==
#             db.customers_subscriptions.id ),
#         db.school_subscriptions.on(
#             db.customers_subscriptions.school_subscriptions_id ==
#             db.school_subscriptions.id ),
#         db.classes_attendance.on(
#             db.customers_subscriptions_exceeded.classes_attendance_id ==
#             db.classes_attendance.id ),
#          db.classes.on(
#             db.classes_attendance.classes_id ==
#             db.classes.id),
#          db.auth_user.on(
#             db.classes_attendance.auth_customer_id ==
#             db.auth_user.id)]
#
#     query = (db.classes_attendance.ClassDate >= firstdaythismonth) & \
#             (db.classes_attendance.ClassDate <= lastdaythismonth)
#
#     rows = db(query).select(db.school_subscriptions.Classes,
#                             db.school_subscriptions.SubscriptionUnit,
#                             db.school_subscriptions.Unlimited,
#                             db.classes_attendance.ClassDate,
#                             db.classes.school_locations_id,
#                             db.classes.school_classtypes_id,
#                             db.classes.Starttime,
#                             db.auth_user.id,
#                             db.auth_user.thumbsmall,
#                             db.auth_user.trashed,
#                             db.auth_user.birthday,
#                             db.auth_user.display_name,
#                             db.customers_subscriptions.school_subscriptions_id,
#                             db.customers_subscriptions_exceeded.ClassCount,
#                             left=left,
#                             orderby=db.auth_user.display_name|
#                                     ~db.classes_attendance.ClassDate)
#
#     for i, row in enumerate(rows):
#         repr_row = list(rows[i:i+1].render())[0]
#
#         cust_name = SPAN(row.auth_user.display_name)
#
#         allowed = attendance_subcription_exceeded_get_classes_allowed(
#             classes   = row.school_subscriptions.Classes,
#             unit      = row.school_subscriptions.SubscriptionUnit,
#             unlimited = row.school_subscriptions.Unlimited)
#
#         taken = A(row.customers_subscriptions_exceeded.ClassCount,
#                   _href=URL('customers', 'classes_attendance',
#                             vars={'cuID':row.auth_user.id}))
#
#         tr = TR(TD(repr_row.auth_user.thumbsmall, _class='os-customer_image_td'),
#                 TD(cust_name),
#                 TD(repr_row.classes_attendance.ClassDate),
#                 TD(repr_row.classes.school_locations_id),
#                 TD(repr_row.classes.school_classtypes_id),
#                 TD(repr_row.classes.Starttime),
#                 TD(repr_row.customers_subscriptions.school_subscriptions_id),
#                 TD(allowed),
#                 TD(taken)
#                 )
#
#         table.append(tr)
#
#     menu = attendance_get_menu(request.function)
#
#     return dict(form=form,
#                 menu=menu,
#                 content=table,
#                 current_month=current_month,
#                 month_chooser=month_chooser,
#                 submit=submit)


def attendance_subcription_exceeded_get_classes_allowed(classes=None,
                                                        unit=None,
                                                        unlimited=False):
    """
        Returns a friendly representation of classes allowed for a subscription
    """
    allowed = ''
    allowed_week  = SPAN(str(classes) + ' / ' + T('Week'))
    allowed_month = SPAN(str(classes) + ' / ' + T('Month'))
    if unlimited:
        allowed = T('unlimited')
    elif unit == 'week':
        allowed = allowed_week
    elif unit == 'month':
        allowed = allowed_month

    return allowed


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('read', 'reports_discovery'))
def discovery():
    """
        Show a doughnut chart of discovery field from auth_user
    """
    response.title = T("Reports")
    response.subtitle = T("Discovery")

    result = discovery_get_data()

    return dict(data_table = result['data_table'])



@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('read', 'reports_discovery'))
def discovery_get_data():
    """
        Return chart data discovery field from auth_user
    """
    if request.extension == 'json':
        response.view = 'generic.json'

    count = db.auth_user.school_discovery_id.count()
    data = list()
    json_data = list()
    labels = list()
    color = dict(red=82, green=136, blue=154)
    rows = db().select(db.school_discovery.id,
                       db.school_discovery.Name,
                       db.auth_user.school_discovery_id,
                       count,
                       orderby=~count,
                       groupby=db.auth_user.school_discovery_id,
                       left=db.school_discovery.on(
                                            db.auth_user.school_discovery_id==\
                                            db.school_discovery.id))
    for row in rows:
        if not row.school_discovery.Name is None and row.school_discovery != '':
            label = row.school_discovery.Name
            value = row[count]
            current_color = 'rgb(' + str(color['red']) + ',' + \
                             str(color['green']) + ',' + \
                             str(color['blue']) + ')'

            data.append(value)
            labels.append(label)
            json_data.append(dict(label=label,
                                   value=value,
                                   color=current_color,
                                   highlight="rgb(17,160,94)")
                            )

            color['red'] += 20
            color['green'] += 20
            color['blue'] += 20

    data_table = TABLE(TR(TH(T("Discovery")), TH(T("Number"))),
        _class="table table-condensed")
    i = 0
    for value in data:
        data_table.append(TR(labels[i], value))
        i += 1

    data_table = os_gui.get_box_table(T("Discovery data"),
                                      data_table)


    return dict(data_table=data_table,
                json_data=json_data)


def geography_get_submenu(page):
    """
        This function returns a submenu for the geography reports pages
    """
    pages = [['geography', T('Top 10 cities & countries'),
               URL('geography')],
             ['postcodes', T('Postcodes'),
               URL('postcodes')] ]
    horizontal = True

    return get_submenu(pages, page, horizontal=horizontal, htype='tabs')


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('read', 'reports_geography'))
def geography():
    """
        Lists top 10 cities and countries for people
    """
    response.title = T("Reports")
    response.subtitle = T("Geography - Top 10 cities and countries")

    result = geography_get_data()

    menu = geography_get_submenu(request.function)

    return dict(menu=menu,
                city_data_table=result['city_data_table'],
                city_title=result['city_title'],
                country_data_table=result['country_data_table'],
                country_title=result['country_title'])


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('read', 'reports_geography'))
def geography_get_data():
    """
        Returns JSON data for geography page
    """
    if request.extension == 'json':
        response.view = 'generic.json'
    # chart datasets for chart.js
    fillColor = "rgba(151,187,205,0.5)"
    strokeColor = "rgba(151,187,205,0.8)"
    highlightFill = "rgba(151,187,205,0.75)"
    highlightStroke = "rgba(151,187,205,1)"
    json_data = dict(city=dict(labels=[],
                                datasets=[dict(label="city_data",
                                           fillColor=fillColor,
                                           strokeColor=strokeColor,
                                           highlightFill=highlightFill,
                                           highlightStroke= highlightStroke,
                                           data=[])]),
                     country=dict(labels=[],
                                  datasets=[dict(label="country_data",
                                           fillColor=fillColor,
                                           strokeColor=strokeColor,
                                           highlightFill=highlightFill,
                                           highlightStroke= highlightStroke,
                                           data=[])])
                     )

    # City
    count = db.auth_user.city.count()
    query = (db.auth_user.customer == True)
    rows = db(query).select(db.auth_user.city, count,
        groupby=db.auth_user.city,
        orderby=~count|db.auth_user.city,
        limitby=(0,10))

    city_data = list()
    city_labels = list()

    for row in rows:
        value = row[count]
        label = row.auth_user.city

        city_data.append(value)
        city_labels.append(label)

        json_data['city']['labels'].append(label)
        json_data['city']['datasets'][0]['data'].append(value)


    city_title = H2(T("Top 10 Cities"))
    city_data_table = TABLE(TR(TH(T("City")), TH(T("Customers"))),
                            _class="table table-condensed")
    i = 0
    for value in city_data:
        city_data_table.append(TR(city_labels[i], value))
        i += 1

    city_data_table = os_gui.get_box_table(T("City data"),
                                                city_data_table)

    # Country
    count = db.auth_user.country.count()
    query = (db.auth_user.customer == True)
    rows = db(query).select(db.auth_user.country, count,
        groupby=db.auth_user.country,
        orderby=~count|db.auth_user.country,
        limitby=(0,10))

    country_data = list()
    country_labels = list()

    for row in rows:
        value = row[count]
        label = row.auth_user.country

        country_data.append(value)
        country_labels.append(label)

        json_data['country']['labels'].append(label)
        json_data['country']['datasets'][0]['data'].append(value)


    country_headers = [ T("Country"), T("Customers") ]
    country_title = H2(T("Top 10 Countries"))

    country_data_table = TABLE(TR(TH(T("Country")), TH(T("Customers"))),
                               _class='table table-condensed')
    i = 0
    for value in country_data:
        country_data_table.append(TR(country_labels[i], value))
        i += 1

    country_data_table = os_gui.get_box_table(T("Country data"),
                                                country_data_table)


    return dict(city_data_table=city_data_table,
                city_title=city_title,
                country_data_table=country_data_table,
                country_title=country_title,
                json_data=json_data)



@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'postcode_groups'))
def postcodes():
    """
        Overview page for postcode groups
    """
    response.title = T('Reports')
    response.subtitle = T('Postcodes')

    edit_groups = os_gui.get_button('edit_custom',
                                    URL('postcodes_groups_list'),
                                    title="Define postcode groups",
                                    _class='btn-block',
                                    btn_size='')

    data = postcodes_get_data()
    data_table = TABLE(TR(TH(T("Postcode Group")), TH(T("Number"))),
                       _class="table table-condensed")
    for label, count in data:
        data_table.append(TR(label, count))

    data_table = os_gui.get_box_table(T("Postcode groups"),
                                        data_table)

    menu = geography_get_submenu(request.function)

    return dict(edit_groups = edit_groups,
                data_table  = data_table,
                chart_title = T("Postcode groups chart"),
                menu        = menu)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'postcode_groups'))
def postcodes_get_data():
    """
        Returns sorted list of items
    """
    # get groups
    data = {}
    query = (db.postcode_groups.Archived == False)
    rows = db(query).select(db.postcode_groups.ALL,
                            orderby=db.postcode_groups.Name)
    for row in rows:
        query = (db.auth_user.postcode_asint >= row.PostcodeStart_AsInt) & \
                (db.auth_user.postcode_asint <= row.PostcodeEnd_AsInt) & \
                (db.auth_user.customer == True)
        count = db(query).count()

        data[row.Name] = count

    sorted_data = sorted(list(data.items()), key=operator.itemgetter(1), reverse=True)

    # what to do now?
    if request.extension != 'json':
        return sorted_data
    elif request.extension == 'json':
        response.view = 'generic.json'
        # make a nice json file for chart.js
        #fillColor = "rgba(151,187,205,0.5)"
        #strokeColor = "rgba(151,187,205,0.8)"
        #highlightFill = "rgba(151,187,205,0.75)"
        #highlightStroke = "rgba(151,187,205,1)"
        #data = dict(labels=[],
                     #datasets=[dict(label="postcodes_data",
                                    #fillColor=fillColor,
                                    #strokeColor=strokeColor,
                                    #highlightFill=highlightFill,
                                    #highlightStroke= highlightStroke,
                                    #data=[])])

        #for label, value in sorted_data:
            #data['labels'].append(label)
            #data['datasets'][0]['data'].append(value)
        color = dict(red=82, green=136, blue=154)
        data  = []
        for label, value in sorted_data:
            current_color = 'rgb(' + str(color['red']) + ',' + \
                                     str(color['green']) + ',' + \
                                     str(color['blue']) + ')'

            data.append(dict(label=label,
                             value=value,
                             color=current_color,
                             highlight="rgb(17,160,94)"))

            color['red'] += 20
            color['green'] += 20
            color['blue'] += 20

        return dict(data=data)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'postcode_groups'))
def postcodes_groups_list():
    """
        List postcode groups
    """
    response.title = T("Reports")
    response.subtitle = T("Edit postcode groups")

    response.view = 'general/only_content.html'

    show = 'current'
    query = (db.postcode_groups.Archived == False)

    if 'show_archive' in request.vars:
        show = request.vars['show_archive']
        session.reports_postcode_groups_show = show
        if show == 'current':
            query = (db.postcode_groups.Archived == False)
        elif show == 'archive':
            query = (db.postcode_groups.Archived == True)
    elif session.reports_postcode_groups_show == 'archive':
            query = (db.postcode_groups.Archived == True)
    else:
        session.reports_postcode_groups_show = show

    db.postcode_groups.id.readable=False
    fields = [db.postcode_groups.Name,
              db.postcode_groups.PostcodeStart,
              db.postcode_groups.PostcodeEnd]
    links = [ lambda row: os_gui.get_button('edit',
                                     URL('postcode_group_edit',
                                         vars={'pgID':row.id}),
                                     T("Edit the name of this group")),
              postcode_groups_get_link_archive ]
    maxtextlengths = {'postcode_groups.Name' : 50}

    delete_permission = auth.has_membership(group_id='Admins') or \
                        auth.has_permission('delete', 'postcode_groups')

    grid = SQLFORM.grid(query,
                        maxtextlengths=maxtextlengths,
                        fields=fields,
                        links=links,
                        create=False,
                        editable=False,
                        details=False,
                        searchable=False,
                        csv=False,
                        deletable=delete_permission,
                        orderby=db.postcode_groups.Name,
                        field_id=db.postcode_groups.id,
                        ui = grid_ui)
    grid.element('.web2py_counter', replace=None) # remove the counter
    grid.elements('span[title=Delete]', replace=None) # remove text from delete button

    add_url = URL('postcode_group_add')
    add = os_gui.get_button('add', add_url, T("Add a postcode_group"), _class="pull-right")

    archive_buttons = os_gui.get_archived_radio_buttons(
        session.reports_postcode_groups_show)

    back = os_gui.get_button('back', URL('postcodes'))
    back = DIV(add, archive_buttons, back)

    return dict(content=grid, back=back)


def postcode_groups_get_link_archive(row):
    """
        Called from the index function. Changes title of archive button
        depending on whether a postcode group is archived or not
    """
    row = db.postcode_groups(row.id)

    if row.Archived:
        tt = T("Move to current")
    else:
        tt = T("Archive")

    return os_gui.get_button('archive',
                             URL('postcode_groups_archive', vars={'pgID':row.id}),
                             tooltip=tt)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'postcode_groups'))
def postcode_groups_archive():
    """
        This function archives a subscription
        request.vars[pgID] is expected to be db.postcode_groups.id
    """
    pgID = request.vars['pgID']
    if not pgID:
        session.flash = T('Unable to (un)archive group')
    else:
        row = db.postcode_groups(pgID)

        if row.Archived:
            session.flash = T('Moved to current')
        else:
            session.flash = T('Archived')

        row.Archived = not row.Archived
        row.update_record()

    redirect(URL('postcodes_groups_list'))


@auth.requires_login()
def postcode_group_add():
    """
        This function shows an add page for a postcode group
    """
    response.title = T("New postcode group")
    response.subtitle = T('')
    response.view = 'general/only_content.html'

    return_url = URL('postcodes_groups_list')

    crud.messages.submit_button = T("Save")
    crud.messages.record_created = T("Saved")
    crud.settings.create_next = return_url
    form = crud.create(db.postcode_groups)

    form_id = "MainForm"
    form_element = form.element('form')
    form['_id'] = form_id

    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = form_id

    submit = form.element('input[type=submit]')


    back = os_gui.get_button('back', return_url)

    return dict(content=form, back=back, save=submit)


@auth.requires_login()
def postcode_group_edit():
    """
        This function shows an edit page for a postcode group
        request.vars['pgID'] is expected to be db.postcode_groups.id
    """
    pgID = request.vars['pgID']
    response.title = T("Edit postcode group")
    response.subtitle = T('')
    response.view = 'general/only_content.html'

    return_url = URL('postcodes_groups_list')

    crud.messages.submit_button = T("Save")
    crud.messages.record_updated = T("Saved")
    crud.settings.update_next = return_url
    crud.settings.update_deletable = False
    form = crud.update(db.postcode_groups, pgID)

    form_id = "MainForm"
    form_element = form.element('form')
    form['_id'] = form_id

    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = form_id

    submit = form.element('input[type=submit]')


    back = os_gui.get_button('back', return_url)

    return dict(content=form, back=back, save=submit)


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('read', 'reports_revenue'))
def revenue():
    if request.extension == 'json':
        response.view = 'generic.json'

    try:
        year = int(request.vars['year'])
        session.reports_revenue_year = year
    except:
        if session.reports_revenue_year:
            year = session.reports_revenue_year
        else:
            year = datetime.date.today().year
            session.reports_revenue_year = year

    response.title = T("Reports")
    response.subtitle=T("Revenue") + " " + str(year)

    today = datetime.date.today()

    result = revenue_get_data()

    tabs = UL(LI(A(T("Total"),
                   _href='#total',
                   _role='tab'),
                 _role='presentation',
                 _class='active'),
              LI(A(T("Subscriptions"),
                   _href='#subscriptions',
                   _role='tab'),
                 _role='presentation'),
              LI(A(T("Class cards"),
                   _href='#classcards',
                   _role='tab'),
                 _role='presentation'),
              LI(A(T("Events"),
                   _href='#workshops',
                   _role='tab'),
                 _role='presentation'),
              LI(A(T("Drop-in classes"),
                   _href='#dropin',
                   _role='tab'),
                 _role='presentation'),
              LI(A(T("Trial classes"),
                   _href='#trial',
                   _role='tab'),
                 _role='presentation'),
              _class='nav nav-tabs',
              _role='tablist',
              _id='reports_revenue_tabs')


    return dict(tabs=tabs,
        total_chart_title=result['total_chart_title'],
        total_avg_table=result['total_avg_table'],
        total_data_table=result['total_data_table'],
        subscriptions_chart_title=result['subscriptions_chart_title'],
        subscriptions_avg_table=result['subscriptions_avg_table'],
        subscriptions_data_table=result['subscriptions_data_table'],
        classcards_chart_title=result['classcards_chart_title'],
        classcards_avg_table=result['classcards_avg_table'],
        classcards_data_table=result['classcards_data_table'],
        workshops_chart_title=result['workshops_chart_title'],
        workshops_avg_table=result['workshops_avg_table'],
        workshops_data_table=result['workshops_data_table'],
        dropin_chart_title=result['dropin_chart_title'],
        dropin_avg_table=result['dropin_avg_table'],
        dropin_data_table=result['dropin_data_table'],
        trialclasses_chart_title=result['trialclasses_chart_title'],
        trialclasses_avg_table=result['trialclasses_avg_table'],
        trialclasses_data_table=result['trialclasses_data_table'],
        year_chooser=revenue_get_year_chooser(request.function)
        )


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('read', 'reports_revenue'))
def revenue_get_data():
    """
        Returns data for revenue graphs / data tables
    """
    ### Common useful values
    today = datetime.date.today()
    year = session.reports_revenue_year

    labels = [T('Jan'), T('Feb'), T('Mar'), T('Apr'), T('May'), T('Jun'),
              T('Jul'), T('Aug'), T('Sep'), T('Oct'), T('Nov'), T('Dec')]
    months = [(1,T("January")),
              (2,T("February")),
              (3,T("March")),
              (4,T("April")),
              (5,T("May")),
              (6,T("June")),
              (7,T("July")),
              (8,T("August")),
              (9,T("September")),
              (10,T("October")),
              (11,T("November")),
              (12,T("December"))]
    avg_title = H4(T("General statistics"))
    data_table_titles = [ T("Month"), T("Revenue") ]


    # chart datasets for chart.js
    fillColor = "rgba(151,187,205,0.5)"
    strokeColor = "rgba(151,187,205,0.8)"
    highlightFill = "rgba(151,187,205,0.75)"
    highlightStroke = "rgba(151,187,205,1)"
    json_data = dict(subscriptions=dict(labels=labels,
                            datasets=[dict(label="subscriptions_data",
                                       fillColor=fillColor,
                                       strokeColor=strokeColor,
                                       highlightFill=highlightFill,
                                       highlightStroke=highlightStroke,
                                       data=[])]),
                      classcards=dict(labels=labels,
                              datasets=[dict(label="classcards_data",
                                       fillColor=fillColor,
                                       strokeColor=strokeColor,
                                       highlightFill=highlightFill,
                                       highlightStroke=highlightStroke,
                                       data=[])]),
                      workshops=dict(labels=labels,
                              datasets=[dict(label="workshops_data",
                                       fillColor=fillColor,
                                       strokeColor=strokeColor,
                                       highlightFill=highlightFill,
                                       highlightStroke=highlightStroke,
                                       data=[])]),
                      dropin=dict(labels=labels,
                              datasets=[dict(label="dropin_data",
                                       fillColor=fillColor,
                                       strokeColor=strokeColor,
                                       highlightFill=highlightFill,
                                       highlightStroke=highlightStroke,
                                       data=[])]),
                      trialclasses=dict(labels=labels,
                              datasets=[dict(label="trialclasses_data",
                                       fillColor=fillColor,
                                       strokeColor=strokeColor,
                                       highlightFill=highlightFill,
                                       highlightStroke=highlightStroke,
                                       data=[])]),
                      total=dict(labels=labels,
                              datasets=[dict(label="total_data",
                                       fillColor=fillColor,
                                       strokeColor=strokeColor,
                                       highlightFill=highlightFill,
                                       highlightStroke=highlightStroke,
                                       data=[])])
                 )

    def get_average_table(totals):
        """
            Returns average table
        """
        average = calculate_average(totals[1])

        return TABLE(THEAD(TR(TH(total[0]), TH(average[0]))),
                           TR(TD(CURRSYM, ' ', format(total[1], '.2f')),
                              TD(CURRSYM, ' ', average[1])),
                     _class='table')


    def get_data_table(data, title):
        """
            Formats data into a nice table
        """
        table = TABLE(_class='table table-condensed dataTable')

        for i in range(0, 12):
            amount = format(data[i], '.2f')
            table.append(TR(TD(months[i][1]),
                            TD(CURRSYM, ' ',
                               SPAN(amount, _class='pull-right'))))

        title = title + ' ' + T('data')
        return os_gui.get_box_table(title, table)


    def get_month_subscriptions(date):
        # helper function to get monthly membership revenue
        year = str(date.year)
        month = date.month

        left = [
            db.invoices.on(
                db.invoices_amounts.invoices_id ==
                db.invoices.id
            ),
            db.invoices_items.on(
                db.invoices_items.invoices_id ==
                db.invoices.id
            ),
            db.invoices_items_customers_subscriptions.on(
                db.invoices_items_customers_subscriptions.invoices_items_id ==
                db.invoices_items.id
            ),
        ]

        query = (db.invoices_items_customers_subscriptions.id != None) & \
                (db.invoices.SubscriptionMonth == date.month) & \
                (db.invoices.SubscriptionYear == date.year)
        rows = db(query).select(db.invoices_amounts.ALL,
                                left=left)


        total = 0
        for row in rows:
            total += row.TotalPriceVAT

        if (int(year) == today.year and month > today.month) or \
            int(year) > today.year:
            return_value = 0
        else:
            return_value = round(total, 2)

        return return_value


    def get_month_classcards(date):
        # helper function to get monthly class card revenue
        year = date.year
        month = date.month
        firstdaythismonth = date
        lastdaythismonth = get_last_day_month(date)

        left = [
            db.invoices.on(
                db.invoices_amounts.invoices_id ==
                db.invoices.id
            ),
            db.invoices_items.on(
                db.invoices_items.invoices_id ==
                db.invoices.id
            ),
            db.invoices_items_customers_classcards.on(
                db.invoices_items_customers_classcards.invoices_items_id ==
                db.invoices_items.id
            )
        ]

        query = (db.invoices_items_customers_classcards.id != None) & \
                (db.invoices.DateCreated >= firstdaythismonth) & \
                (db.invoices.DateCreated <= lastdaythismonth)
        sum = db.invoices_amounts.TotalPriceVAT.sum()
        value = db(query).select(sum, left=left).first()[sum]


        if (year == today.year and month > today.month) or \
            year > today.year:
            return_value = 0
        else:
            if value == None:
                return_value = 0
            else:
                return_value = round(value, 2)

        return return_value


    def get_month_workshops(date):
        # helper function to get monthly workshops revenue
        year = date.year
        month = date.month
        firstdaythismonth = date
        lastdaythismonth = get_last_day_month(date)

        left = [
            db.invoices.on(
                db.invoices_amounts.invoices_id ==
                db.invoices.id
            ),
            db.invoices_items.on(
                db.invoices_items.invoices_id ==
                db.invoices.id
            ),
            db.invoices_items_workshops_products_customers.on(
                db.invoices_items_workshops_products_customers.invoices_items_id ==
                db.invoices_items.id
            )
        ]

        query = (db.invoices_items_workshops_products_customers.id != None) & \
                (db.invoices.DateCreated >= firstdaythismonth) & \
                (db.invoices.DateCreated <= lastdaythismonth)
        sum = db.invoices_amounts.TotalPriceVAT.sum()
        value = db(query).select(sum, left=left).first()[sum]

        if (year == today.year and month > today.month) or \
            year > today.year:
            return_value = 0
        else:
            if value == None:
                return_value = 0
            else:
                return_value = round(value, 2)

        return return_value


    def get_month_classes(date, attendance_type):
        # helper function to get monthly drop in class revenue
        year = date.year
        month = date.month
        firstdaythismonth = date
        lastdaythismonth = get_last_day_month(date)

        left = [
            db.invoices_amounts.on(
                db.invoices_amounts.invoices_id ==
                db.invoices.id
            ),
            db.invoices_items.on(
                db.invoices_items.invoices_id ==
                db.invoices.id
            ),
            db.invoices_items_classes_attendance.on(
                db.invoices_items_classes_attendance.invoices_items_id ==
                db.invoices_items.id
            ),
            db.classes_attendance.on(
                db.invoices_items_classes_attendance.classes_attendance_id ==
                db.classes_attendance.id
            ),
        ]

        if attendance_type == 'dropin':
            query = (db.classes_attendance.AttendanceType==2)
        elif attendance_type == 'trial':
            query = (db.classes_attendance.AttendanceType==1)

        query &= (db.invoices.DateCreated >= firstdaythismonth)
        query &= (db.invoices.DateCreated <= lastdaythismonth)

        sum = db.invoices_amounts.TotalPriceVAT.sum()
        value = db(query).select(sum, left=left).first()[sum]

        if (year == today.year and month > today.month) or \
            year > today.year:
            # Don't make future predictions
            return_value = 0
        else:
            if value == None:
                return_value = 0
            else:
                return_value = round(value, 2)

        return return_value

    def calculate_average(total):
        # calculate average
        if month > today.month and int(year) >= today.year:
            return [T("Monthly average") ,"{0:.2f}".format(total/today.month)]
            # only show the last two decimal numbers
        else:
            return [T("Monthly average") ,"{0:.2f}".format(total/12)]
            # only show the last two decimal numbers

    ### subscriptions
    total = 0
    for month in range(1,13):
        revenue = get_month_subscriptions(datetime.date(year, month, 1))
        json_data['subscriptions']['datasets'][0]['data'].append(revenue)
        total += revenue

    total = [ T("Total"), total ]

    subscriptions_chart_title = H3(T("Subscriptions revenue"))
    subscriptions_avg_table = get_average_table(total)
    subscriptions_data_table = TABLE(_class="table table-condensed")

    # fill the data table
    table_data = json_data['subscriptions']['datasets'][0]['data']
    subscriptions_data_table = get_data_table(table_data,
                                              T('Subscriptions'))
    ### subscriptions end

    ### classcards begin
    classcards_chart_data = list()
    total = 0
    for month in range(1,13):
        revenue = get_month_classcards(datetime.date(year, month, 1))
        json_data['classcards']['datasets'][0]['data'].append(revenue)
        total += revenue

    total = [ T("Total"), total ]

    classcards_chart_title = H3(T("Class cards revenue"))

    classcards_avg_table = get_average_table(total)
    classcards_data_table = TABLE(_class="table table-condensed")

    # fill the details table
    table_data = json_data['classcards']['datasets'][0]['data']
    classcards_data_table = get_data_table(table_data, T('Class cards'))

    ### classcards end

    ### workshops begin
    classcards_chart_data = list()
    total = 0
    for month in range(1,13):
        revenue = get_month_workshops(datetime.date(year, month, 1))
        json_data['workshops']['datasets'][0]['data'].append(revenue)
        total += revenue

    total = [ T("Total"), total ]

    workshops_chart_title = H3(T("Event revenue"))

    workshops_avg_table = get_average_table(total)
    workshops_data_table = TABLE(_class="table table-condensed")

    # fill the details table

    table_data = json_data['workshops']['datasets'][0]['data']
    workshops_data_table = get_data_table(table_data, T('Workshops'))

    ### workshops end


    ### drop in classes begin
    total = 0
    for month in range(1,13):
        revenue = get_month_classes(datetime.date(year, month, 1), 'dropin')
        json_data['dropin']['datasets'][0]['data'].append(revenue)
        total += revenue

    dropin_title = T('Drop in classes revenue for ') + str(year)

    total = [ T("Total"), total ]

    dropin_chart_title = H3(T("Drop in classes revenue"))

    dropin_avg_table = get_average_table(total)
    dropin_data_table = TABLE(_class="table table-condensed")

    # fill the details table
    table_data = json_data['dropin']['datasets'][0]['data']
    dropin_data_table = get_data_table(table_data, T('Drop-in classes'))

    ### drop in classes end

    ### trialclasses begin
    total = 0
    for month in range(1,13):
        revenue = get_month_classes(datetime.date(year, month, 1), 'trial')
        json_data['trialclasses']['datasets'][0]['data'].append(revenue)
        total += revenue

    total = [ T("Total"), total ]

    trialclasses_chart_title = H3(T("Trialclasses revenue"))
    trialclasses_avg_table = get_average_table(total)
    trialclasses_data_table = TABLE(_class="table table-condensed")

    # fill the details table
    table_data = json_data['trialclasses']['datasets'][0]['data']
    trialclasses_data_table = get_data_table(table_data, T('Trial classes'))
    ### trialclasses end

    ### total begin
    total = 0
    for month in range(0,12):
        revenue = json_data['subscriptions']['datasets'][0]['data'][month] + \
                  json_data['classcards']['datasets'][0]['data'][month] + \
                  json_data['workshops']['datasets'][0]['data'][month] + \
                  json_data['dropin']['datasets'][0]['data'][month] + \
                  json_data['trialclasses']['datasets'][0]['data'][month]
        json_data['total']['datasets'][0]['data'].append(round(revenue, 2))
        total += round(revenue, 2)

    total = [ T("Total"), total ]

    total_chart_title = H3(T("Total revenue"))

    total_avg_table = get_average_table(total)
    total_data_table = TABLE(_class="table")

    # fill the details table
    table_data = json_data['total']['datasets'][0]['data']
    total_data_table = get_data_table(table_data, T('Total revenue'))

    ### total end


    return dict(
        json_data=json_data,
        total_chart_title=total_chart_title,
        total_avg_table=total_avg_table,
        total_data_table=total_data_table,
        subscriptions_chart_title=subscriptions_chart_title,
        subscriptions_avg_table=subscriptions_avg_table,
        subscriptions_data_table=subscriptions_data_table,
        classcards_chart_title=classcards_chart_title,
        classcards_avg_table=classcards_avg_table,
        classcards_data_table=classcards_data_table,
        workshops_chart_title=workshops_chart_title,
        workshops_avg_table=workshops_avg_table,
        workshops_data_table=workshops_data_table,
        dropin_chart_title=dropin_chart_title,
        dropin_avg_table=dropin_avg_table,
        dropin_data_table=dropin_data_table,
        trialclasses_chart_title=trialclasses_chart_title,
        trialclasses_avg_table=trialclasses_avg_table,
        trialclasses_data_table=trialclasses_data_table)



def revenue_get_year_chooser(page='revenue'):
    """
        Returns month chooser for overview
    """

    if page == 'revenue':
        year  = session.reports_revenue_year

        link = 'revenue_set_year'

    prev_year = year - 1
    next_year = year + 1

    url_prev = URL(link, vars={'year' :prev_year})
    url_next = URL(link, vars={'year' :next_year})

    previous = A(I(_class='fa fa-angle-left'),
                 _href=url_prev,
                 _class='btn btn-default')
    nxt = A(I(_class='fa fa-angle-right'),
            _href=url_next,
            _class='btn btn-default')

    return DIV(previous, nxt, _class='btn-group pull-right')


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'reports_revenue'))
def revenue_set_year():
    """
        Sets the year session variable for revenue overview
    """
    year  = request.vars['year']

    session.reports_revenue_year = int(year)

    redirect(URL('revenue'))


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('read', 'classes_attendance'))
def class_attendance_get_json():
    """
        This function returns the attendance data for a class in a year
        specified by session.stats_attendance_year
        and should be calles as .json
    """
    response.view = 'generic.json'

    def count_attendance(clsID, date):
        query = (db.classes_attendance.classes_id==clsID) & \
                (db.classes_attendance.ClassDate==date) & \
                (db.classes_attendance.BookingStatus == 'attending')
        attendancecount = int(db(query).count())

        query = (db.classes_attendance_override.classes_id==clsID) & \
                (db.classes_attendance_override.ClassDate==date)
        override_check = db(query).isempty()

        if not override_check: # not empty
            row = db(query).select(db.classes_attendance_override.Amount,
                                   cache=(cache.ram, 30))[0]
            return row.Amount
        else:
            return attendancecount

    year = session.stats_attendance_year
    clsID = session.stats_attendance_clsID

    # chart datasets for chart.js
    fillColor = "rgba(151,187,205,0.5)"
    strokeColor = "rgba(151,187,205,0.8)"
    highlightFill = "rgba(151,187,205,0.75)"
    highlightStroke = "rgba(151,187,205,1)"
    json_data = dict(attendance=dict(labels=[],
                                datasets=[dict(label="attendance_data",
                                           fillColor=fillColor,
                                           strokeColor=strokeColor,
                                           highlightFill=highlightFill,
                                           highlightStroke= highlightStroke,
                                           data=[])]))

    # get the data
    rows = db(db.classes.id==clsID).select(db.classes.Week_day,
                                           db.classes.Startdate,
                                           db.classes.Starttime,
                                           db.school_locations.Name,
                                           db.school_classtypes.Name,
                                           db.classes.school_classtypes_id,
        left = [(db.school_locations.on(db.classes.school_locations_id==db.school_locations.id)),
                (db.school_classtypes.on(db.classes.school_classtypes_id==db.school_classtypes.id))],
        cache=(cache.ram, 30)
        )
    row = rows.first()
    dayofweek = int(row.classes.Week_day)
    dayofweek_name = NRtoDay(dayofweek)
    startdate = row.classes.Startdate
    starttime = str(row.classes.Starttime.strftime('%H:%M'))
    location = str(row.school_locations.Name)
    classtype = str(row.school_classtypes.Name)
    firstday = iso_to_gregorian(year, 1, dayofweek)

    date = firstday
    delta = datetime.timedelta(days=7)
    next_year = datetime.date(year+1, 1, 1)
    maximum = 0
    total = 0
    week = 1
    weeks_with_data = 0
    attendance_data = []
    attendance_labels = []
    while date < next_year:
        if date < startdate:
            attendance = 0
        else:
            attendance = count_attendance(clsID, date)

        if attendance > 0:
            weeks_with_data += 1
            total += attendance
            if attendance > maximum:
                maximum = attendance

        label = str(week)

        attendance_labels.append(label)
        attendance_data.append(attendance)

        json_data['attendance']['labels'].append(label)
        json_data['attendance']['datasets'][0]['data'].append(attendance)

        date += delta
        week += 1

    if weeks_with_data != 0:
        average = total/weeks_with_data
    else:
        average = 0

    title = location + " " + dayofweek_name + " " + starttime + " " + classtype + " " + str(year)
    response.subtitle = title

    # set other data
    chart_title = H4(T("Attendance barchart"))
    avg_title = H4(T("General statistics"))

    return dict(json_data=json_data)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'reports_teacherclasses'))
def teacher_classes():
    """
        show a list with all classes of a teacher by month
    """
    response.title = T('Reports')
    response.subtitle = T('')
    response.view = 'reports/general.html'

    session.classes_attendance_back = 'reports_teacher_classes'
    session.reports_teacher_classes_class_revenue_back = 'reports_teacher_classes'


    if 'teachers_id' in request.vars:
        session.reports_te_classes_teID = request.vars['teachers_id']

    if 'month' in request.vars:
        session.reports_te_classes_month = int(request.vars['month'])
        session.reports_te_classes_year = int(request.vars['year'])
    elif session.reports_te_classes_month is None or \
         session.reports_te_classes_year is None:
        today = datetime.date.today()
        session.reports_te_classes_year = today.year
        session.reports_te_classes_month = today.month


    revenue_totals = ''
    revenue_permission = (auth.has_membership(group_id='Admins') or
                          auth.has_permission('read', 'reports_teacherclasses_revenue'))

    if not session.reports_te_classes_teID:
        table = ''
    else:
        table = TABLE(_class='table table-hover')

        revenue_header = ''
        if revenue_permission:
            revenue_header = T('Revenue')

        table.append(TR(TH(),
                        TH(T('Date')),
                        TH(T('Start')),
                        TH(T('Location')),
                        TH(T('Class Type')),
                        TH(T('Teacher')),
                        TH(T('Teacher 2')),
                        TH(revenue_header),
                        TH(), # actions
                        _class='header'))

        date = datetime.date(session.reports_te_classes_year,
                             session.reports_te_classes_month, 5)
        last_day = get_last_day_month(date)

        revenue_in_vat = 0
        revenue_ex_vat = 0
        revenue_vat = 0

        for each_day in range(1,last_day.day+1):
            # list days
            day = datetime.date(session.reports_te_classes_year,
                                session.reports_te_classes_month, each_day)
            weekday = day.isoweekday()

            class_schedule = ClassSchedule(
                date = day,
                filter_id_teacher = session.reports_te_classes_teID
            )

            rows = class_schedule.get_day_rows()

            for i, row in enumerate(rows):
                repr_row = list(rows[i:i+1].render())[0]

                result = class_schedule._get_day_row_status(row)
                status_marker = result['marker']

                class_revenue = ''
                if revenue_permission:
                    result = class_schedule._get_day_row_teacher_roles(row, repr_row)
                    revenue = teacher_classes_get_class_revenue_total(row.classes.id, day)
                    revenue_in_vat += revenue['revenue_in_vat']
                    revenue_ex_vat += revenue['revenue_ex_vat']
                    revenue_vat += revenue['revenue_vat']

                    class_revenue = A(SPAN(CURRSYM, ' ', format(revenue['revenue_in_vat'], '.2f')),
                                      _href=URL('teacher_classes_class_revenue',
                                                vars={'clsID': row.classes.id,
                                                      'date': day.strftime(DATE_FORMAT)}))

                date_formatted = day.strftime(DATE_FORMAT)

                tr = TR(
                    TD(status_marker,
                       _class='td_status_marker'),
                    TD(date_formatted),
                    TD(repr_row.classes.Starttime),
                    TD(repr_row.classes.school_locations_id),
                    TD(repr_row.classes.school_classtypes_id),
                    TD(result['teacher_role']),
                    TD(result['teacher_role2']),
                    TD(class_revenue),
                    TD(os_gui.get_button('next_no_text',
                                         URL('classes', 'attendance', vars={'clsID':row.classes.id,
                                                                            'date':date_formatted}),
                                         _class='pull-right'))
                )

                table.append(tr)

        if revenue_permission:
            revenue_totals = DIV(
                H4(T('Revenue totals')),
                TABLE(TR(TH(T('Revenue excl VAT')),
                         TD(SPAN(CURRSYM, ' ', format(revenue_ex_vat, '.2f'), _class='pull-right'))),
                      TR(TH(T('Revenue VAT')),
                         TD(SPAN(CURRSYM, ' ', format(revenue_vat, '.2f'), _class='pull-right'))),
                      TR(TH(T('Revenue incl VAT')),
                         TD(SPAN(CURRSYM, ' ', format(revenue_in_vat, '.2f'), _class='pull-right'))),
                      _class='table')
            )

    form_subtitle = get_form_subtitle(session.reports_te_classes_month,
                                      session.reports_te_classes_year,
                                      request.function,
                                      _class='col-md-8')
    response.subtitle = form_subtitle['subtitle']
    form_month = form_subtitle['form']
    month_chooser = form_subtitle['month_chooser']
    current_month = form_subtitle['current_month']
    submit = form_subtitle['submit']

    response.subtitle = SPAN(T('Teacher classes'), ' ',
                             form_subtitle['subtitle'])

    form = teacher_classes_get_form_teachers(session.reports_te_classes_teID)


    return dict(form=DIV(DIV(form_month, form, _class="col-md-6"),
                         DIV(revenue_totals, _class='col-md-6')),
                menu='',
                content=table,
                month_chooser=month_chooser,
                current_month=current_month,
                run_report=submit)


def teacher_classes_get_form_teachers(teachers_id):
    """
        returns list teachers select form
    """
    au_query = (db.auth_user.trashed == False) &\
               (db.auth_user.teacher == True)
    form = SQLFORM.factory(
        Field('teachers_id', db.auth_user,
                requires=IS_IN_DB(db(au_query),
                                  'auth_user.id',
                                  '%(full_name)s',
                                  zero=(T('Select teacher...'))),
                represent=lambda value, row: teachers_dict.get(value, None),
                default = teachers_id,
                label=T("Teacher")),
        submit_button = T('Go'))

    select = form.element('select[name=teachers_id]')
    select['_onchange'] = "this.form.submit();"

    form = DIV(DIV(HR(),
                   form.custom.begin,
                   LABEL(T('Teacher')),
                   form.custom.widget.teachers_id,
                   form.custom.end,
                   _class='col-md-8'),
               _class='row')

    return form


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'reports_teacherclasses'))
def teacher_classes_show_current():
    """
        Resets some session variables to show the current month for
        teacher_classes
    """
    session.reports_te_classes_year = None
    session.reports_te_classes_month = None

    redirect(URL('teacher_classes'))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'reports_teacherclasses_revenue'))
def teacher_classes_class_revenue():
    """
        Page to show revenue breakdown for a class
    """
    clsID = request.vars['clsID']
    date_formatted = request.vars['date']
    date = datestr_to_python(DATE_FORMAT, date_formatted)

    response.title = T('Reports')
    response.subtitle = T('Class revenue') + ' - ' + get_classname(clsID) + ': ' + date_formatted
    response.view = 'general/only_content.html'

    header = THEAD(TR(TH(),
                      TH(T('Customer')),
                      TH(T('Attendance')),
                      TH(T('Description')),
                      TH(T('Revenue excl VAT')),
                      TH(T('Revenue VAT')),
                      TH(T('Revenue incl VAT')),
                      ))

    table = TABLE(header, _class='table table-hover teacher-classes-revenue')

    rows = teacher_classes_get_class_revenue_rows(clsID, date)
    for i, row in enumerate(rows):
        repr_row = list(rows[i:i + 1].render())[0]

        ex_vat = 0
        vat = 0
        in_vat = 0
        description = ''
        if row.classes_attendance.AttendanceType is None:
            # Subscription
            revenue = teacher_classes_get_class_revenue_subscription(row, date)
            ex_vat = revenue['revenue_ex_vat']
            vat = revenue['revenue_vat']
            in_vat = revenue['revenue_in_vat']
            description = SPAN(repr_row.customers_subscriptions.school_subscriptions_id,
                               _title=row.customers_subscriptions.id)
        elif row.classes_attendance.AttendanceType == 1:
            # Trial
            revenue = teacher_classes_get_class_revenue_dropin_trial(row, date)
            ex_vat = revenue['revenue_ex_vat']
            vat = revenue['revenue_vat']
            in_vat = revenue['revenue_in_vat']
        elif row.classes_attendance.AttendanceType == 2:
            # Drop in
            revenue = teacher_classes_get_class_revenue_dropin_trial(row, date)
            ex_vat = revenue['revenue_ex_vat']
            vat = revenue['revenue_vat']
            in_vat = revenue['revenue_in_vat']
        elif row.classes_attendance.AttendanceType == 3:
            # Class pass
            revenue = teacher_classes_get_class_revenue_classcard(row)
            ex_vat = revenue['revenue_ex_vat']
            vat = revenue['revenue_vat']
            in_vat = revenue['revenue_in_vat']
            description = SPAN(repr_row.customers_classcards.school_classcards_id,
                               _title=T('Card') +  ' ' + str(row.customers_classcards.id))

        tr = TR(TD(repr_row.auth_user.thumbsmall),
                TD(repr_row.auth_user.display_name),
                TD(repr_row.classes_attendance.AttendanceType),
                TD(description),
                TD(SPAN(CURRSYM, ' ', format(ex_vat, '.2f'))),
                TD(SPAN(CURRSYM, ' ', format(vat, '.2f'))),
                TD(SPAN(CURRSYM, ' ', format(in_vat, '.2f'))),
                )

        table.append(tr)



    # totals
    result = teacher_classes_get_class_revenue_total(clsID, date)
    tfoot = TFOOT(TR(TD(),
                     TD(),
                     TD(),
                     TH(T('Total')),
                     TH(SPAN(CURRSYM, ' ', format(result['revenue_ex_vat'], '.2f'))),
                     TH(SPAN(CURRSYM, ' ', format(result['revenue_vat'], '.2f'))),
                     TH(SPAN(CURRSYM, ' ', format(result['revenue_in_vat'], '.2f'))),
                     ))

    table.append(tfoot)

    if session.reports_teacher_classes_class_revenue_back == 'reports_attendance_classes':
        back = ''
    else:
        back_url = URL('teacher_classes')
        back = os_gui.get_button('back', back_url, _class='full-width')


    return dict(content = table,
                back=back)


def teacher_classes_get_class_revenue_rows(clsID, date):
    """
    :param clsID: db.classes.id
    :param date: Class date
    :return: All customers attending a class (db.customers_attendance.ALL & db.customers_subscriptions.ALL)
    """
    left = [db.customers_classcards.on(
                db.customers_classcards.id == db.classes_attendance.customers_classcards_id),
            db.customers_subscriptions.on(
                db.customers_subscriptions.id == db.classes_attendance.customers_subscriptions_id),
            db.auth_user.on(db.classes_attendance.auth_customer_id == db.auth_user.id)]
    query = (db.classes_attendance.classes_id == clsID) & \
            (db.classes_attendance.ClassDate == date)
    rows = db(query).select(db.auth_user.ALL,
                            db.classes_attendance.ALL,
                            db.customers_subscriptions.ALL,
                            db.customers_classcards.ALL,
                            left=left,
                            orderby=db.auth_user.display_name)

    return rows


def teacher_classes_get_class_revenue_total(clsID, date):
    """
    :param clsID: db.classes.id
    :param date: class date
    :return: total revenue
    """
    rows = teacher_classes_get_class_revenue_rows(clsID, date)

    total_revenue_in_vat = Decimal(0)
    total_revenue_ex_vat = Decimal(0)
    total_revenue_vat = Decimal(0)

    for row in rows:
        if row.classes_attendance.AttendanceType is None:
            # Subscription
            result = teacher_classes_get_class_revenue_subscription(row, date)
            total_revenue_in_vat += result['revenue_in_vat']
            total_revenue_ex_vat += result['revenue_ex_vat']
            total_revenue_vat += result['revenue_vat']
        elif row.classes_attendance.AttendanceType == 1:
            # Trial
            result = teacher_classes_get_class_revenue_dropin_trial(row, date)
            total_revenue_in_vat += result['revenue_in_vat']
            total_revenue_ex_vat += result['revenue_ex_vat']
            total_revenue_vat += result['revenue_vat']
        elif row.classes_attendance.AttendanceType == 2:
            # Drop in
            result = teacher_classes_get_class_revenue_dropin_trial(row, date)
            total_revenue_in_vat += result['revenue_in_vat']
            total_revenue_ex_vat += result['revenue_ex_vat']
            total_revenue_vat += result['revenue_vat']
        elif row.classes_attendance.AttendanceType == 3:
            # Class card
            result = teacher_classes_get_class_revenue_classcard(row)
            total_revenue_in_vat += result['revenue_in_vat']
            total_revenue_ex_vat += result['revenue_ex_vat']
            total_revenue_vat += result['revenue_vat']

    return dict(revenue_in_vat = total_revenue_in_vat,
                revenue_ex_vat = total_revenue_ex_vat,
                revenue_vat = total_revenue_vat)


def teacher_classes_get_class_revenue_classcard(row):
    """
        :param row: row from db.classes_attendance with left join on db.customers_subscriptions
        :return: Revenue for class taken on a card
    """
    ccdID = row.classes_attendance.customers_classcards_id
    classcard = CustomerClasscard(ccdID)

    query = (db.invoices_items_customers_classcards.customers_classcards_id == ccdID)
    left = [
        db.invoices_items.on(
            db.invoices_items.invoices_id ==
            db.invoices.id
        ),
        db.invoices_items_customers_classcards.on(
            db.invoices_items_customers_classcards.invoices_items_id ==
            db.invoices_items.id
        )
    ]

    rows = db(query).select(
        db.invoices_items.invoices_id,
        db.invoices_items_customers_classcards.ALL,
        left=left
    )

    if not rows:
        revenue_in_vat = 0
        revenue_ex_vat = 0
        revenue_vat = 0
    else:
        row = rows.first()
        invoice = Invoice(row.invoices_items.invoices_id)

        amounts = invoice.get_amounts()

        price_in_vat = amounts.TotalPriceVAT
        price_ex_vat = amounts.TotalPrice

        # Divide by classes taken on card
        if classcard.unlimited:
            # Count all classes taken on card
            query = (db.classes_attendance.customers_classcards_id == ccdID)
            count_classes = db(query).count()

            revenue_in_vat = price_in_vat / count_classes
            revenue_ex_vat = price_ex_vat / count_classes
            revenue_vat = revenue_in_vat - revenue_ex_vat
        else:
            revenue_in_vat = price_in_vat / classcard.classes
            revenue_ex_vat = price_ex_vat / classcard.classes
            revenue_vat = revenue_in_vat - revenue_ex_vat

    return dict(revenue_in_vat=revenue_in_vat,
                revenue_ex_vat=revenue_ex_vat,
                revenue_vat=revenue_vat)


def teacher_classes_get_class_revenue_dropin_trial(row, date):
    """
        :param row: row from db.classes_attendance with left join on db.customers_subscriptions
        :param date: date of class
        :param product_type: 'dropin' or 'trial'
        :return : revenue for a drop in or trial class
    """
    query = (db.invoices_items_classes_attendance.classes_attendance_id == row.classes_attendance.id)
    left = [
        db.invoices_items.on(
            db.invoices_items_classes_attendance.invoices_items_id ==
            db.invoices_items.id
        )
    ]
    rows = db(query).select(
        db.invoices_items.ALL,
        left=left
    )

    if not rows:
        price_in_vat = 0
        price_ex_vat = 0
        vat = 0
    else:
        invoice_item = rows.first()

        price_in_vat = invoice_item.TotalPriceVAT
        price_ex_vat = invoice_item.TotalPrice
        vat = invoice_item.VAT



    # cls = Class(row.classes_attendance.classes_id, date)
    # prices = cls.get_price()
    #
    # if product_type == 'dropin':
    #     price_in_vat = prices['dropin']
    #     tax_percentage = prices['dropin_tax_percentage']
    #     tax_rates_id = prices['dropin_tax_rates_id']
    #
    # elif product_type == 'trial':
    #     price_in_vat = prices['trial']
    #     tax_percentage = prices['trial_tax_percentage']
    #     tax_rates_id = prices['dropin_tax_rates_id']
    #
    # if not price_in_vat:
    #     # Check if a price is defined
    #     price_in_vat = 0
    #     price_ex_vat = 0
    #     vat = 0
    # else:
    #     if tax_percentage:
    #         # Check if a vat percentage is defined
    #         price_ex_vat = price_in_vat / (1 + (tax_percentage / 100))
    #         vat = price_in_vat - price_ex_vat
    #     else:
    #         price_ex_vat = price_in_vat
    #         vat = 0

    return dict(revenue_in_vat=price_in_vat or 0,
                revenue_ex_vat=price_ex_vat or 0,
                revenue_vat=vat or 0)


def teacher_classes_get_class_revenue_subscription(row, date):
    """
        :param row: row from db.classes_attendance with left join on db.customers_subscriptions
        :return: Revenue by subscription
    """
    first_day = datetime.date(date.year, date.month, 1)
    last_day  = get_last_day_month(date)

    # Get number of classes on subscription this month
    query = (db.classes_attendance.customers_subscriptions_id == row.customers_subscriptions.id) & \
            (db.classes_attendance.ClassDate >= first_day) & \
            (db.classes_attendance.ClassDate <= last_day)
    classes_taken = db(query).count()
    # Get price for subscription invoice

    csID = row.customers_subscriptions.id
    subscr_month = date.month
    subscr_year  = date.year

    query = (db.invoices_items_customers_subscriptions.customers_subscriptions_id == csID) & \
            (db.invoices.SubscriptionMonth == subscr_month) & \
            (db.invoices.SubscriptionYear == subscr_year)

    left = [
        db.invoices_amounts.on(
            db.invoices_amounts.invoices_id ==
            db.invoices.id
        ),
        db.invoices_items.on(
            db.invoices_items.invoices_id ==
            db.invoices.id
        ),
        db.invoices_items_customers_subscriptions.on(
             db.invoices_items_customers_subscriptions.invoices_items_id ==
             db.invoices_items.id
        )
    ]

    rows = db(query).select(db.invoices.ALL,
                            db.invoices_amounts.ALL,
                            left=left)

    if not rows:
        revenue_in_vat = 0
        revenue_ex_vat = 0
        revenue_vat = 0
    else:
        row = rows.first()
        price_in_vat = row.invoices_amounts.TotalPriceVAT
        price_ex_vat = row.invoices_amounts.TotalPrice

        # Divide by classes taken on subscription in month
        revenue_in_vat = price_in_vat / classes_taken
        revenue_ex_vat = price_ex_vat / classes_taken
        revenue_vat = revenue_in_vat - revenue_ex_vat


    return dict(revenue_in_vat = revenue_in_vat,
                revenue_ex_vat = revenue_ex_vat,
                revenue_vat = revenue_vat)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'reports_retention'))
def retention_rate():
    """
        Renders the retention rate page
    """
    return retention_or_dropoff_rate(retention=True)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'reports_retention'))
def dropoff_rate():
    """
        Renders the dropoff rate page
    """
    return retention_or_dropoff_rate(retention=False)


def retention_or_dropoff_rate(retention=True):
    """
        Builds the retention rate form.

        The retention rate is computed in a simple manner: 
        in a given period of time (say one year) takes the list of customers that attended at
        least one class in the first window (ie. first month), and computes the same list for the last window, too.
        Takes the intersection of the two lists (this is the list of customers that attended
        both ath the beginning and the end of the period).
        The retention rate is the size of the intersection divided by the number of attendees
        in the first window. 

        :param retention: if True then the retention page is returnded. Otherwise returns the dropoff page.
    """
    response.title = T('Reports')
    # set the subtitle
    if retention:
        response.subtitle = T('Retention rate')
    else:
        response.subtitle = T('Drop off rate')

    response.view = 'reports/subscriptions.html'
    session.customers_back = 'reports_retention_rate'

    today = datetime.date.today()

    # get the parameters from the request or use defaults
    first_day_this_month = datetime.date(TODAY_LOCAL.year, TODAY_LOCAL.month, 1)
    p1_end_default = first_day_this_month - datetime.timedelta(days=1)
    p1_start_default = datetime.date(TODAY_LOCAL.year - 1, TODAY_LOCAL.month, 1)
    p2_start_default = first_day_this_month
    p2_end_default = get_last_day_month(first_day_this_month)

    p1_start = retention_get_parameter_or_session('p1_start', p1_start_default, 'reports_rr_p1_start')
    p1_end = retention_get_parameter_or_session('p1_end', p1_end_default, 'reports_rr_p1_end')
    p2_start = retention_get_parameter_or_session('p2_start', p2_start_default, 'reports_rr_p2_start')
    p2_end = retention_get_parameter_or_session('p2_end', p2_end_default, 'reports_rr_p2_end')

    # create a form with two mont and two year fields
    form = SQLFORM.factory(
        Field('p1_start', 'date',
              default=p1_start,
              requires=IS_DATE_IN_RANGE(format=DATE_FORMAT,
                                        minimum=datetime.date(1900, 1, 1),
                                        maximum=datetime.date(2999, 1, 1)),
              represent=represent_date,
              label=T('First period start')),
        Field('p1_end', 'date',
              default=p1_end,
              requires=IS_DATE_IN_RANGE(format=DATE_FORMAT,
                                        minimum=datetime.date(1900, 1, 1),
                                        maximum=datetime.date(2999, 1, 1)),
              represent=represent_date,
              label=T('First period end')),
        Field('p2_start', 'date',
              default=p2_start,
              requires=IS_DATE_IN_RANGE(format=DATE_FORMAT,
                                        minimum=datetime.date(1900, 1, 1),
                                        maximum=datetime.date(2999, 1, 1)),
              represent=represent_date,
              label=T('Second period start')),
        Field('p2_end', 'date',
              default=p2_end,
              requires=IS_DATE_IN_RANGE(format=DATE_FORMAT,
                                        minimum=datetime.date(1900, 1, 1),
                                        maximum=datetime.date(2999, 1, 1)),
              represent=represent_date,
              label=T('Second period end')),
        submit_button=T("Run report"),
        formstyle='divs'
        )

    submit = form.element('input[type=submit]')
    submit['_class'] = ' pull-right'

    form = DIV(DIV(DIV(LABEL(form.custom.label.p1_start),
                       form.custom.widget.p1_start,
                       _class='col-md-3'),
                   DIV(LABEL(form.custom.label.p1_end),
                       form.custom.widget.p1_end,
                       _class="col-md-3")),
                   DIV(LABEL(form.custom.label.p2_start),
                       form.custom.widget.p2_start,
                       _class='col-md-3'),
                   DIV(LABEL(form.custom.label.p2_end),
                       form.custom.widget.p2_end,
                       _class="col-md-3"),
               _class='row')



    # compute the retention based on the parameters
    helper = AttendanceHelper()

    # attendee lists for both windows
    attendees_first_window, attendees_second_window = retention_get_windows(p1_start, p1_end, p2_start, p2_end)

    intersection = set(attendees_first_window) & set(attendees_second_window)

    first_window_size = len(attendees_first_window)
    intersect_size = len(intersection)

    if first_window_size:
        retention_rate = float(intersect_size) / first_window_size
        dropoff_rate = float((first_window_size - intersect_size)) / first_window_size
    else:
        retention_rate = 0
        dropoff_rate = 0    

    if retention:
        stats_box = H3(DIV(format(retention_rate * 100, '.2f'), "% ",
                           _class="label label-success"))
        # get the loyal customers table
        table = retention_get_customer_table(intersection)
    else:
        stats_box = H3(DIV(format(dropoff_rate * 100, '.2f'), "% ",
                           _class="label label-danger"))
        # select the dropped off customers and render the dropoff customer table
        dropped_off_customers = set(attendees_first_window) - set(intersection)
        table = retention_get_customer_table(dropped_off_customers)

    info = SPAN(retention_get_info(p1_start, p1_end, p2_start, p2_end, retention), _class='grey')
    content = DIV(info, stats_box, table)

    menu = retention_get_menu(request.function)

    link_export_retention = A(SPAN(os_gui.get_fa_icon('fa-magnet'), ' ',
                               T('Retention rate')),
                          _href=URL('retention_rate_export'))

    links = [link_export_retention]

    export = os_gui.get_dropdown_menu(
        links=links,
        btn_text='',
        btn_icon='download',
        menu_class='pull-right')

    return dict(form=form,
                menu=menu,
                content=content,
                export=export,
                run_report=submit)


def retention_get_customer_table(customer_ids):
    """
        Builds a table of customers that is used on the retention rate page.
    """
    rows = db(db.auth_user.id.belongs(customer_ids)).select(db.auth_user.ALL,
                                                            orderby=db.auth_user.display_name)

    helper = AttendanceHelper()
    last_attendances = helper.get_last_attendance(customer_ids) # TODO: combine the two queries

    table = TABLE(_class="table table-hover")
    table.append(THEAD(TR(
                    TH(),
                    TH(T('Customer')),
                    TH(T('Last class date')))))

    for row in rows.render():                        
        table.append(TR(
            TD(row.thumbsmall,
                           _class='os-customer_image_td'),
            TD(DIV(row.display_name)),
            TD(last_attendances[row.id])))

    return table


def retention_get_info(p1_start, p1_end, p2_start, p2_end, retention):
    """
        Returns the subtitle string for the  retention rate repport
    """

    if retention:
        retention_text = T('also')
    else:
        retention_text = T('not')

    subtitle = SPAN(
        T('Customers attending classes using a subscription or class card between'), ' ',
        p1_start.strftime(DATE_FORMAT), ' ', T('and'), ' ',
        p1_end.strftime(DATE_FORMAT), ', ',
        retention_text, ' ', T('attending between'), ' ',
        p2_start.strftime(DATE_FORMAT), ' ', T('and'), ' ',
        p2_end.strftime(DATE_FORMAT)
    )

    return subtitle


def retention_get_menu(page=None):
    """
        Builds the top menu for the retention and dropoff report pages
    """
    pages = [
        (['retention_rate', T('Retention rate'), URL('reports',"retention_rate")]),
        (['dropoff_rate', T('Drop off rate'), URL('reports',"dropoff_rate")]),
        ]

    return get_submenu(pages,
                       page,
                       horizontal=True,
                       htype='tabs')


@auth.requires(auth.has_membership(group_id='Admins') or \
                   auth.has_permission('read', 'reports_retention'))
def retention_rate_export():
    """
        Exports the retention or the dropoff based on the session parameters
    """
    # create filestream
    stream = io.BytesIO()

    # init the workbook and the two sheets
    wb = openpyxl.workbook.Workbook(write_only=True)

    # compute the retention based on the parameters
    helper = AttendanceHelper()

    # attendee lists for both windows
    attendees_first_window, attendees_second_window = retention_get_windows(session.reports_rr_p1_start,
                                                                            session.reports_rr_p1_end,
                                                                            session.reports_rr_p2_start,
                                                                            session.reports_rr_p2_end)

    intersection = set(attendees_first_window) & set(attendees_second_window)

    # export the loyal customers
    retention_export_table("Retention", intersection, wb)

    # export the dropped off customers
    dropped_off_customers = set(attendees_first_window) - set(intersection)
    retention_export_table("Drop off", dropped_off_customers, wb)

    # save the workbook to the stream
    wb.save(stream)

    # set the response headers
    fname = T("Retention rate") + '.xlsx'
    response.headers['Content-Type']='application/vnd.ms-excel'
    response.headers['Content-disposition']='attachment; filename=' + fname

    return stream.getvalue()


def retention_get_windows(p1_start, p1_end, p2_start, p2_end):
    """
        Computes the first and the last window customer ids for the retention.
        Returns two customer id lists: attendees_first_window, attendees_second_window
    """
    helper = AttendanceHelper()

    attendees_first_window = helper.get_attending_list_between(p1_start, p1_end)
    attendees_second_window = helper.get_attending_list_between(p2_start, p2_end)

    return attendees_first_window, attendees_second_window


def retention_export_table(title, customer_ids, workbook):
    """
        Exports the retention data

        :param title: the title of the sheet where the data is exported
        :param customer_ids the customers to export
        :param workbook: the workbook that will contain the sheet
    """
    sheet = workbook.create_sheet(title=title)

    headers = [
        'Customer ID',
        'Customer Name',
        'Email',
        'Last Date'
    ]

    sheet.append(headers)

    helper = AttendanceHelper()

    rows = db(db.auth_user.id.belongs(customer_ids)).select(db.auth_user.ALL,
                                                            orderby=db.auth_user.display_name)

    last_attendances = helper.get_last_attendance(customer_ids)

    for row in rows:
        fields = [
            row.id,
            row.display_name,
            row.email,
            last_attendances[row.id]
        ]

        sheet.append(fields)


def retention_get_parameter_or_session(parameter, default_value, session_parameter=None):
    """
        Gets a value from the request or the session.

        If the parameter is not found in the request the it tries to get it from the session.
        Otherwise returns the default value.
        If the session_parameter is not None then stores the value to the session.

        :param parameter: the name of the request parameter
        :param default_value: the value to return when the parameter is not found
        :param session_parameter: the session parameter name
    """
    if parameter in request.vars:
        value = datestr_to_python(DATE_FORMAT, request.vars[parameter])
    elif session_parameter and not session[session_parameter] is None:
        value = session[session_parameter]
    else:
        value = default_value

    if session_parameter:
        session[session_parameter] = value

    return value


@auth.requires(auth.has_membership(group_id='Admins') or
               auth.has_permission('read', 'reports_customers'))
def customers_inactive():
    """
        List customers inactive since date and offer option to clean up
    """
    from openstudio.os_customers import Customers

    response.title = T('Reports')
    response.subtitle = T('Inactive customers')
    response.view = 'reports/general.html'

    session.customers_back = 'reports_customers_inactive'

    if 'date' in request.vars:
        date = datestr_to_python(DATE_FORMAT, request.vars['date'])
        customers = Customers()
        result = customers.list_inactive_after_date_formatted(date)
        content = DIV(
            customers_inactive_get_button_delete(date),
            H4(T('Found'), ' ', result['count'], ' ',
               T('customers without activity after'), ' ',
               request.vars['date']),
            HR(),
            result['table']
        )
    else:
        date = TODAY_LOCAL
        content = DIV(
            T('Please select a date and click "Run report".'),
            _class="col-md-12"
        )

    form = customers_inactive_get_form(date)
    result = set_form_id_and_get_submit_button(form, 'MainForm')
    form = result['form']
    submit = result['submit']

    return dict(content=content,
                form=DIV(form, _class='col-md-4'),
                run_report=submit)


def customers_inactive_get_button_delete(date):
    """
        :return: delete button if the user has permissions to delete users
    """
    delete = ''

    onclick = "return confirm('" + \
     T('Do you really want to remove all customers on this list and all data associated with these customers?')\
     + "');"
    permission = (auth.has_membership(group_id='Admins') or
                  auth.has_permission('delete', 'auth_user'))
    if permission:
        date_formatted = date.strftime(DATE_FORMAT)
        delete = os_gui.get_button(
            'noicon',
            URL('customers_inactive_delete', vars={'date':date_formatted}),
            title = T('Delete customers'),
            tooltip = T('Delete customers on list'),
            onclick=onclick,
            btn_class='btn-danger',
            _class="pull-right"
        )


    return delete


def customers_inactive_get_form(date):
    """
        Return date form for clean_up
    """
    form = SQLFORM.factory(
        Field('date', 'date',
              default=date,
              requires=IS_DATE_IN_RANGE(format=DATE_FORMAT,
                                        minimum=datetime.date(1900, 1, 1),
                                        maximum=datetime.date(2999, 1, 1)),
              label=T('Date'),
              widget=os_datepicker_widget,
              comment=T('List customers without activity after this date')),
        formstyle='bootstrap3_stacked',
        submit_button=T('Run report')
    )

    return form


@auth.requires(auth.has_membership(group_id='Admins') or
               auth.has_permission('delete', 'auth_user'))
def customers_inactive_delete():
    """
        :return: None
    """
    from openstudio.os_customers import Customers

    date_formatted = request.vars['date']
    date = datestr_to_python(DATE_FORMAT, date_formatted)

    customers = Customers()
    nr_deleted = customers.delete_inactive_after_date(date)

    session.flash = SPAN(T("Deleted"), ' ', nr_deleted, ' ', T('customers'))
    redirect(URL('customers_inactive'))
