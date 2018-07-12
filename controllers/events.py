# coding: utf8

from openstudio.os_workshops_helper import WorkshopsHelper

from openstudio.os_workshop_product import WorkshopProduct
from openstudio.os_workshop import Workshop
from openstudio.os_invoice import Invoice
from openstudio.os_invoices import Invoices
from openstudio.os_tasks import Tasks
from openstudio.os_customer import Customer
from openstudio.os_mail import OsMail

from general_helpers import get_weekday
from general_helpers import get_badge
from general_helpers import get_label
from general_helpers import get_ajax_loader
from general_helpers import get_submenu
from general_helpers import get_input_search
from general_helpers import datestr_to_python
from general_helpers import class_get_teachers
from general_helpers import max_string_length
from general_helpers import Memo_links
from general_helpers import workshops_get_full_workshop_product_id
from general_helpers import classes_get_status
from general_helpers import set_form_id_and_get_submit_button

import html2text
import cStringIO
import openpyxl


### helpers begin
def activity_count_reservations(wsaID, fwsID):
    """
        Returns count of reservations for an activity
    """
    # count full workshop customers
    query = (db.workshops_products_customers.workshops_products_id == fwsID) & \
            (db.workshops_products_customers.Cancelled == False) & \
            (db.workshops_products_customers.Waitinglist == False)
    count_full_ws = db(query).count()
    # count activity customers
    query = activity_list_customers_get_list_activity_query(wsaID)
    count_activity = db(query).count()

    return count_full_ws + count_activity


def activity_get_filled(row, fwsID):
    """
        Formats count of reservations for use in manage page activity list
    """
    total = unicode(row.Spaces)
    wsaID = row.id

    count = activity_count_reservations(wsaID, fwsID)
    used = unicode(count)

    filled = used + "/" + total
    if used == total:
        filled = SPAN(filled, _class='green')
    return filled


def get_workshops_menu(page, wsID):
    vars = {'wsID': wsID}

    pages = [['event_edit',
              T('General'),
              URL('event_edit', vars=vars)],
             ['tickets',
              T('Tickets'),
              URL('tickets', vars=vars)],
             ['activities',
              T('Activities'),
              URL('activities', vars=vars)],
             ['tasks',
              T('Tasks'),
              URL('tasks', vars=vars)],
             ['stats',
              T('Quick stats'),
              URL('stats', vars=vars)],
             ['info_mail',
              T('Info mail'),
              URL('info_mail', vars=vars)]]

    return get_submenu(pages, page, horizontal=True, htype='tabs')


def get_subtitle(wsID):
    """
        Returns a subtitle for a workshop
    """
    workshop = db.workshops(wsID)
    st = SPAN(workshop.Name + ' ')
    if workshop.Enddate:
        if workshop.Startdate == workshop.Enddate:
            st.append(SPAN(workshop.Startdate.strftime(DATE_FORMAT),
                           _class='small_font'))
        else:
            st.append(SPAN(workshop.Startdate.strftime(DATE_FORMAT) + ' - ' + \
                           workshop.Enddate.strftime(DATE_FORMAT),
                           _class='small_font'))
    elif workshop.Startdate:
        st.append(SPAN(workshop.Startdate.strftime(DATE_FORMAT),
                       _class='small_font'))

    return st


### helpers end

def index_get_filter_form(filter_option):
    filter_options = [('current', T('Current')), ('archived', T('Archived'))]

    form = FORM(_action='#',
                _method='post',
                _class='right select-narrow',
                _id='workshops_show_filter')
    select = SELECT(value=filter_option,
                    _name='filter',
                    _class='generic-widget',
                    _onchange="this.form.submit();",
                    _style='vertical-align:text;')
    for option in filter_options:
        selected = False
        if filter_option == option[0]:
            selected = True
        select.append(OPTION(option[1], _value=option[0], _selected=selected))

    submit = INPUT(_type='submit', _value=T("Go"))
    form.insert(0, select)
    form.append(submit)

    return DIV(form, _class='form_inline')


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'workshops'))
def index():
    response.title = T("Events")
    response.subtitle = T("")
    response.view = 'general/only_content_no_box.html'

    # Make sure we're not going back to workshop payments page from here
    session.workshops_manage_back = None

    if not session.workshops_show:
        session.workshops_show = 'current'

    if 'show_archive' in request.vars:
        show = request.vars['show_archive']
        session.workshops_show = show

    if session.workshops_show == 'current':
        query = (db.workshops.Archived == False)
    elif session.workshops_show == 'archive':
        query = (db.workshops.Archived == True)
        response.subtitle = T("Archived")

    if session.workshops_show == 'archive':
        archive_class = 'active'
        current_class = ''
    else:
        current_class = 'active'
        archive_class = ''

    if ('all' in request.vars and session.workshops_show == 'archive') or session.workshops_show == 'current':
        limit = False
        show_all = ''
    else:
        # Only show initial limit for archive
        limit = 15
        show_all = DIV(A(T("Show all"),
                         _href=URL('events', 'index', vars={'all': True}),
                         _title=T('Show the full archive, might take a little while to load')))

    delete_permission = auth.has_membership(group_id='Admins') or \
                        auth.has_permission('delete', 'workshops')

    add_url = URL('event_add')
    add = os_gui.get_button('add', add_url, T("Add a new event"))

    archive_buttons = os_gui.get_archived_radio_buttons(
        session.workshops_show)

    tools = DIV(archive_buttons, add)

    content = DIV(
        UL(LI(A(T('Current'),
                _href=URL(vars={'show_archive': 'current'})),
              _class=current_class),
           LI(A(T('Archive'),
                _href=URL(vars={'show_archive': 'archive'})),
              _class=archive_class),
           # LI(I(_class='fa fa-object-group'),
           #    _class='pull-left header'),
           _class='nav nav-tabs pull-right'),
        DIV(DIV(index_get_content(query, limit),
                show_all,
                _class='tab-pane active'),
            _class='tab-content'),
        _class='nav-tabs-custom')

    return dict(content=content, add=add)


def index_get_content(query, limit=False):
    """
        Returns table with workshop content
    """
    delete_onclick = "return confirm('" + \
                     T('Are you sure you want to delete this event?') + "');"

    header = THEAD(TR(TH(T('Image')),
                      TH(T('Name')),
                      TH(T('Start')),
                      TH(T('Teacher')),
                      TH(T('Location')),
                      TH(T('Shop')),
                      TH(),  # Actions
                      ))

    table = TABLE(header, _class='table table-striped table-hover')

    orderby = db.workshops.Startdate
    if session.workshops_show == 'archive':
        # Sort by latest first in archive
        orderby = ~db.workshops.Startdate


    if limit:
        rows = db(query).select(db.workshops.ALL, limitby=(0, limit), orderby=orderby)
    else:
        rows = db(query).select(db.workshops.ALL, orderby=orderby)

    for i, row in enumerate(rows):
        repr_row = list(rows[i:i + 1].render())[0]

        # Shop (API)
        public = INPUT(value=row.PublicWorkshop,
                       _type='checkbox',
                       _value='public',
                       _disabled='disabled')

        # Buttons
        edit = index_get_link_edit(row)

        pdf = os_gui.get_button('print',
                                URL('events', 'pdf', vars={'wsID': row.id}),
                                tooltip=T('PDF'))

        archive = ''
        archive_permission = auth.has_membership(group_id='Admins') or \
                             auth.has_permission('update', 'workshops')
        if archive_permission:
            archive = index_get_link_archive(row)

        delete = ''
        delete_permission = auth.has_membership(group_id='Admins') or \
                            auth.has_permission('delete', 'workshops')
        if delete_permission:
            delete = os_gui.get_button('delete_notext',
                                       URL('event_delete', vars={'wsID': row.id}),
                                       onclick=delete_onclick)

        buttons = DIV(edit, pdf, archive, delete, _class='pull-right')

        tr = TR(
            TD(repr_row.thumbsmall),
            TD(max_string_length(repr_row.Name, 34),
               _title=repr_row.Name),
            TD(repr_row.Startdate),
            TD(max_string_length(repr_row.auth_teacher_id, 20),
               _title=repr_row.auth_teacher_id),
            TD(max_string_length(repr_row.school_locations_id, 23),
               _title=repr_row.school_locations_id),
            TD(public),
            TD(buttons)
        )

        table.append(tr)

    return table


def index_get_link_edit(row):
    """
        Returns drop down link for index edit
    """
    vars = {'wsID': row.id}

    links = []

    permission = (auth.has_membership(group_id='Admins') or
                  auth.has_permission('update', 'workshops'))
    if permission:
        edit = A((os_gui.get_fa_icon('fa-pencil'), T('Edit')),
                 _href=URL('event_edit', vars=vars))
        links.append(edit)

    permission = (auth.has_membership(group_id='Admins') or
                  auth.has_permission('read', 'workshops_products'))
    if permission:
        link = A((os_gui.get_fa_icon('fa-ticket'), T('Tickets')),
                 _href=URL('tickets', vars=vars))
        links.append(link)

    permission = (auth.has_membership(group_id='Admins') or
                  auth.has_permission('read', 'workshops_activities'))
    if permission:
        link = A((os_gui.get_fa_icon('fa-clock-o'), T('Activities')),
                 _href=URL('activities', vars=vars))
        links.append(link)

    permission = (auth.has_membership(group_id='Admins') or
                  auth.has_permission('read', 'tasks'))
    if permission:
        link = A((os_gui.get_fa_icon('fa-check-square-o'), T('Tasks')),
                 _href=URL('tasks', vars=vars))
        links.append(link)

    permission = (auth.has_membership(group_id='Admins') or
                  auth.has_permission('read', 'workshops'))
    if permission:
        link = A((os_gui.get_fa_icon('fa-pie-chart'), ' ', T("Quick Stats")),
                 _href=URL('stats', vars=vars))
        links.append(link)

    duplicate_permission = auth.has_membership(group_id='Admins') or \
                           auth.has_permission('create', 'workshops')
    if duplicate_permission:
        links.append(LI(_role='separator', _class='divider'))
        duplicate_onclick = "return confirm('" + \
                            T('Are you sure you want to duplicate this event?') + "');"

        link = A((os_gui.get_fa_icon('fa-clone'), ' ', T('Duplicate')),
                 _href=URL('event_duplicate',
                           vars={'wsID': row.id}),
                 _onclick=duplicate_onclick)
        links.append(link)


    menu = os_gui.get_dropdown_menu(
        links=links,
        btn_text='',
        btn_size='btn-sm',
        btn_icon='pencil',
        menu_class='btn-group pull-left')

    return menu


def index_get_link_archive(row):
    """
        Called from the index function. Changes title of archive button
        depending on whether a workshop is archived or not
    """
    workshop = db.workshops(row.id)

    if workshop.Archived:
        tt = T("Move to current")
    else:
        tt = T("Archive")

    return os_gui.get_button('archive',
                             URL('event_archive', vars={'wsID': row.id}),
                             tooltip=tt)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'workshops'))
def event_archive():
    """
        This function archives a workshop
        request.vars[wsID] is expected to be the workshop ID
    """
    wsID = request.vars['wsID']
    if not wsID:
        session.flash = T('Unable to (un)archive event')
    else:
        workshop = db.workshops(wsID)

        if workshop.Archived:
            session.flash = T('Moved event to current')
        else:
            session.flash = T('Archived event')

        workshop.Archived = not workshop.Archived
        workshop.update_record()

        # Clear cache
        cache_clear_workshops()

    redirect(URL('index'))


def event_add_get_menu(page):
    """
        Returns submenu for adding a workshop
    """
    pages = [['event_add', T('1. Event'), "#"],
             ['event_add_set_price', T('2. Price'), "#"]]

    return os_gui.get_submenu(pages, page, horizontal=True, htype='tabs')


@auth.requires_login()
def event_add():
    """
        This function shows an add page for a workshop
    """
    response.title = T('Add event')
    response.subtitle = T("Event info")
    response.view = 'general/tabs_menu.html'

    return_url = URL('index')
    db.workshops.Archived.readable = False
    db.workshops.Archived.writable = False

    event_hide_teacher_fields()

    next_url = '/events/event_add_set_price?wsID=[id]'
    crud.messages.submit_button = T("Next")
    crud.messages.record_created = T("")
    crud.settings.create_onaccept = [event_add_onaccept, cache_clear_workshops]
    crud.settings.create_next = next_url
    crud.settings.formstyle = 'bootstrap3_stacked'
    form = crud.create(db.workshops)

    textareas = form.elements('textarea')
    for textarea in textareas:
        textarea['_class'] += ' tmced'

    result = set_form_id_and_get_submit_button(form, 'MainForm')
    form = result['form']
    submit = result['submit']

    menu = event_add_get_menu(request.function)
    back = os_gui.get_button('back', return_url)

    return dict(content=form,
                save=submit,
                back=back,
                menu=menu)


def event_add_onaccept(form):
    """
        Insert full workshop product after a workshop is created
    """
    wsID = form.vars.id
    ws_row = db.workshops(wsID)
    db.workshops_products.insert(workshops_id=wsID,
                                 FullWorkshop=True,
                                 Deletable=False,
                                 Name=T('Full event'),
                                 Price=0,
                                 Description=T('Full event'))


@auth.requires_login()
def event_add_set_price():
    """
        Set the price of the full workshop product for a workshop
        Uses the permissions for workshops_products
    """
    wsID = request.vars['wsID']
    fwspID = workshops_get_full_workshop_product_id(wsID)

    response.title = T('Add event')
    response.subtitle = T("Full event price")
    response.view = 'general/tabs_menu.html'

    db.workshops_products.Name.readable = False
    db.workshops_products.Name.writable = False
    db.workshops_products.Description.readable = False
    db.workshops_products.Description.writable = False
    db.workshops_products.ExternalShopURL.readable = False
    db.workshops_products.ExternalShopURL.writable = False
    db.workshops_products.AddToCartText.readable = False
    db.workshops_products.AddToCartText.writable = False
    db.workshops_products.PublicProduct.readable = False
    db.workshops_products.PublicProduct.writable = False

    crud.messages.submit_button = T("Save")
    crud.messages.record_updated = T("Saved")
    crud.settings.update_deletable = False
    crud.settings.update_next = URL('tickets', vars={'wsID': wsID})
    form = crud.update(db.workshops_products, fwspID)

    result = set_form_id_and_get_submit_button(form, 'MainForm')
    form = result['form']
    submit = result['submit']

    menu = event_add_get_menu(request.function)

    return dict(content=form,
                save=submit,
                menu=menu)


@auth.requires(auth.has_membership(group_id='Admins') or
               auth.has_permission('delete', 'workshops'))
def event_delete():
    """
        Delete a workshop
    """
    wsID = request.vars['wsID']

    query = (db.workshops.id == wsID)
    db(query).delete()

    session.flash = T('Event deleted')

    # clear cache
    cache_clear_workshops()

    redirect(URL('index'))


@auth.requires_login()
def event_edit():
    """
        This function shows an edit page for a workshop
        request.args[0] is expected to be the workshop ID (wsID)
    """
    wsID = request.vars['wsID']
    return_url = URL('index')

    response.title = T("Edit event")
    response.subtitle = get_subtitle(wsID)
    response.view = 'general/tabs_menu.html'

    event_hide_teacher_fields()

    crud.messages.submit_button = T("Save")
    crud.messages.record_updated = T("Saved")
    crud.settings.update_next = URL(vars={'wsID': wsID})
    crud.settings.update_onaccept = [cache_clear_workshops]
    crud.settings.update_deletable = False
    crud.settings.formstyle = 'bootstrap3_stacked'
    form = crud.update(db.workshops, wsID)

    form_id = "MainForm"
    form_element = form.element('form')
    form['_id'] = form_id

    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = form_id

    submit = form.element('input[type=submit]')

    textareas = form.elements('textarea')
    for textarea in textareas:
        textarea['_class'] += ' tmced'

    menu = get_workshops_menu(request.function, wsID)
    back = os_gui.get_button('back', return_url)
    content = DIV(event_get_alert_no_activities(wsID),
                  form)

    return dict(content=content,
                back=back,
                save=submit,
                menu=menu)


# @auth.requires_login()
# def event_edit_teachers():
#     """
#         Edit teachers for a workshop
#         request.vars['wsID'] is expected to be the workshops_id
#         request.vars['wiz'] is expected to be 'True' in case the wizzard is
#         used
#     """
#     wsID = request.vars['wsID']
#     wizzard = True if 'wiz' in request.vars else False
#     return_url = URL('index')
#     response.title = T("Edit workshop")
#     response.subtitle = get_subtitle(wsID)
#
#     response.view = 'general/tabs_menu.html'
#
#     if wizzard:
#         # call js for styling the form
#         crud.messages.submit_button = T("Next")
#         crud.messages.record_updated = T("")
#         crud.settings.update_next = URL('events',
#                                         'event_add_set_price',
#                                         vars={'wsID': wsID})
#         back = ''
#         menu = event_add_get_menu(request.function)
#         response.title = T('Add workshop')
#         response.subtitle = T("Teachers")
#
#     else:
#         crud.messages.submit_button = T("Save")
#         crud.messages.record_updated = T("Saved teachers")
#         # crud.settings.update_next = return_url
#         back = os_gui.get_button('back', return_url)
#         menu = workshop_edit_get_submenu(request.function, wsID)
#
#     db.workshops.Name.readable = False
#     db.workshops.Name.writable = False
#     db.workshops.Startdate.readable = False
#     db.workshops.Startdate.writable = False
#     db.workshops.Enddate.readable = False
#     db.workshops.Enddate.writable = False
#     db.workshops.picture.readable = False
#     db.workshops.picture.writable = False
#     db.workshops.PublicWorkshop.readable = False
#     db.workshops.PublicWorkshop.writable = False
#     db.workshops.Description.readable = False
#     db.workshops.Description.writable = False
#     db.workshops.school_locations_id.readable = False
#     db.workshops.school_locations_id.writable = False
#
#     crud.settings.update_onaccept = [cache_clear_workshops]
#     crud.settings.update_deletable = False
#     form = crud.update(db.workshops, wsID)
#
#     result = set_form_id_and_get_submit_button(form, 'MainForm')
#     form = result['form']
#     submit = result['submit']
#
#     return dict(content=form,
#                 back=back,
#                 save=submit,
#                 menu=menu)


def event_hide_teacher_fields(var=None):
    """
        Sets readable and writable to False for teacher fields in
        db.workshops
    """
    db.workshops.Teacher.readable = False
    db.workshops.Teacher.writable = False
    db.workshops.TeacherEmail.readable = False
    db.workshops.TeacherEmail.writable = False
    db.workshops.Teacher2.readable = False
    db.workshops.Teacher2.writable = False
    db.workshops.Teacher2Email.readable = False
    db.workshops.Teacher2Email.writable = False
    db.workshops.Startdate.readable = False
    db.workshops.Startdate.writable = False


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('create', 'workshops'))
def event_duplicate():
    """
        Duplicate a workshop including products & activities
    """
    wsID = request.vars['wsID']
    # workshop
    workshop = db.workshops(wsID)
    new_wsID = db.workshops.insert(
        Archived=workshop.Archived,
        PublicWorkshop=False,
        Name=workshop.Name + ' ' + T('(Copy)'),
        Startdate=workshop.Startdate,
        Enddate=workshop.Enddate,
        Starttime=workshop.Starttime,
        Endtime=workshop.Endtime,
        auth_teacher_id=workshop.auth_teacher_id,
        auth_teacher_id2=workshop.auth_teacher_id2,
        Description=workshop.Description,
        school_locations_id=workshop.school_locations_id,
        picture=workshop.picture,
        thumbsmall=workshop.thumbsmall,
        thumblarge=workshop.thumblarge,
        picture_2=workshop.picture_2,
        thumbsmall_2=workshop.thumbsmall_2,
        thumblarge_2=workshop.thumblarge_2,
        picture_3=workshop.picture_3,
        thumbsmall_3=workshop.thumbsmall_3,
        thumblarge_3=workshop.thumblarge_3,
        picture_4=workshop.picture_4,
        thumbsmall_4=workshop.thumbsmall_4,
        thumblarge_4=workshop.thumblarge_4,
        picture_5=workshop.picture_5,
        thumbsmall_5=workshop.thumbsmall_5,
        thumblarge_5=workshop.thumblarge_5,
    )

    # activities
    activity_ids_old = []
    activity_ids_map = []
    query = (db.workshops_activities.workshops_id == wsID)
    rows = db(query).select(db.workshops_activities.ALL)
    for row in rows:
        new_wsaID = db.workshops_activities.insert(
            workshops_id=new_wsID,
            Activity=row.Activity,
            Activitydate=row.Activitydate,
            school_locations_id=row.school_locations_id,
            Starttime=row.Starttime,
            Endtime=row.Endtime,
            Spaces=row.Spaces,
            auth_teacher_id=row.auth_teacher_id,
            auth_teacher_id2=row.auth_teacher_id2,
        )
        activity_ids_old.append(row.id)
        activity_ids_map.append({'old': row.id,
                                 'new': new_wsaID})

    # products
    products_ids_old = []
    products_ids_map = []
    query = (db.workshops_products.workshops_id == wsID)
    rows = db(query).select(db.workshops_products.ALL)
    for row in rows:
        new_wspID = db.workshops_products.insert(
            workshops_id=new_wsID,
            FullWorkshop=row.FullWorkshop,
            Deletable=row.Deletable,
            PublicProduct=row.PublicProduct,
            Name=row.Name,
            Price=row.Price,
            PriceSubscription=row.PriceSubscription,
            PriceEarlybird=row.PriceEarlybird,
            PriceSubscriptionEarlybird=row.PriceSubscriptionEarlybird,
            EarlybirdUntil=row.EarlybirdUntil,
            tax_rates_id=row.tax_rates_id,
            Description=row.Description,
            ExternalShopURL=row.ExternalShopURL,
            AddToCartText=row.AddToCartText,
            Donation=row.Donation
        )
        products_ids_old.append(row.id)
        products_ids_map.append({'old': row.id,
                                 'new': new_wspID})

    # products activities
    query = db.workshops_products_activities.workshops_products_id.belongs(products_ids_old)
    rows = db(query).select(db.workshops_products_activities.ALL)
    for row in rows:
        for pim in products_ids_map:
            if pim['old'] == row.workshops_products_id:
                new_wspID = pim['new']

        for aim in activity_ids_map:
            if aim['old'] == row.workshops_activities_id:
                new_wsaID = aim['new']

        db.workshops_products_activities.insert(
            workshops_products_id=new_wspID,
            workshops_activities_id=new_wsaID
        )

    # info mail
    query = (db.workshops_mail.workshops_id == wsID)
    rows = db(query).select(db.workshops_mail.ALL)
    for row in rows:
        db.workshops_mail.insert(
            workshops_id = new_wsID,
            MailContent = row.MailContent
        )

    # Clear cache
    cache_clear_workshops()

    session.flash = T('You are now editing the duplicated event')

    redirect(URL('event_edit', vars={'wsID': new_wsID}))


def activity_add_edit_update_workshop_datetime_info(form):
    """
        :param form: db.workshops_activities add or edit form
        Set dates & times for a workshop based on activities
    """
    activity = db.workshops_activities(form.vars.id)
    workshop = Workshop(activity.workshops_id)

    workshop.update_dates_times()


@auth.requires_login()
def activity_add():
    """
        This function shows an add page for a workshop activity
    """
    wsID = request.args[0]
    response.title = T('Event')
    response.subtitle = get_subtitle(wsID)
    response.view = 'general/tabs_menu.html'

    workshop = db.workshops(wsID)

    db.workshops_activities.workshops_id.default = wsID
    db.workshops_activities.Teacher.default = workshop.Teacher
    db.workshops_activities.Teacher2.default = workshop.Teacher2
    db.workshops_activities.auth_teacher_id.default = workshop.auth_teacher_id
    db.workshops_activities.auth_teacher_id2.default = workshop.auth_teacher_id2

    return_url = URL('activities', vars={'wsID': wsID})

    crud.messages.submit_button = T("Save")
    crud.messages.record_created = T("Added activity")
    crud.settings.create_onaccept = [activity_add_edit_update_workshop_datetime_info, cache_clear_workshops]
    crud.settings.create_next = return_url
    crud.settings.formstyle = 'table3cols'
    form = crud.create(db.workshops_activities)

    form_id = "MainForm"
    form_element = form.element('form')
    form['_id'] = form_id

    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = form_id

    submit = form.element('input[type=submit]')

    content = DIV(H4(T('New activity')), BR(), form)

    back = os_gui.get_button('back', return_url)
    menu = get_workshops_menu('activities', wsID)

    return dict(content=content, back=back, menu=menu, save=submit)


@auth.requires_login()
def activity_edit():
    """
        This function shows an edit page for an activity
        request.args[0] is expected to be the workshops_activities_id (wsaID)
    """
    wsaID = request.args[0]
    response.title = T('Event')
    wsID = db.workshops_activities(wsaID).workshops_id
    response.subtitle = get_subtitle(wsID)
    response.view = 'general/tabs_menu.html'

    return_url = URL('activities', vars={'wsID': wsID})

    crud.messages.submit_button = T("Save")
    crud.messages.record_updated = T("Saved")
    crud.messages.record_deleted = T("Deleted activity")
    crud.settings.update_onaccept = [activity_add_edit_update_workshop_datetime_info, cache_clear_workshops]
    crud.settings.update_next = return_url
    crud.settings.update_deletable = False

    form = crud.update(db.workshops_activities, wsaID)

    form_id = "MainForm"
    form_element = form.element('form')
    form['_id'] = form_id

    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = form_id

    submit = form.element('input[type=submit]')

    back = os_gui.get_button('back', return_url)
    content = DIV(H4(T('Edit activity')), BR(), form)

    menu = get_workshops_menu('activities', wsID)

    return dict(content=content, back=back, menu=menu, save=submit)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('delete', 'workshops_activities'))
def activity_delete():
    """
        Delete the activity specified by request.args[1] (wsaID)
    """
    wsID = request.args[0]
    wsaID = request.args[1]

    # Delete activity
    query = (db.workshops_activities.id == wsaID)
    db(query).delete()

    # Update dates & times for workshop
    workshop = Workshop(wsID)
    workshop.update_dates_times()

    # Clear cache
    cache_clear_workshops()

    redirect(URL('activities', vars={'wsID': wsID}))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'workshops'))
def mail_activity_attendance():
    """
        This function sends the attendance list for all activities
        to the workshop teacher
    """
    wsID = request.vars['wsID']

    workshop = db.workshops(wsID)

    # start the message.
    message = "<html>\n<head><STYLE TYPE='text/css'> \
            <!-- TD{font-size: 9pt; align:left;} \
            TD, TH { padding:0 20px 0 0; } \
            TH{font-size: 9pt; text-align:left;} --->  </STYLE></head> \
            <body style=font-size:10pt; font-style:normal;'>"
    message += T("Dear teacher(s)")
    message += ",<br><br>"
    message += 'Below are the attendance lists for the activities in your '
    message += 'workshop, ' + workshop.Name + '.<br><br>'

    fws_rows = activity_list_customers_get_fullws_rows(wsID)

    query = (db.workshops_activities.workshops_id == wsID)
    rows = db(query).select(db.workshops_activities.ALL,
                            orderby=db.workshops_activities.Activitydate | \
                                    db.workshops_activities.Starttime)
    for row in rows:
        title = row.Activity
        subtitle = T("On ") + row.Activitydate.strftime(DATE_FORMAT) + \
                   T(' at ') + \
                   row.Starttime.strftime(TIME_FORMAT)
        message += unicode(H3(title))
        message += unicode(subtitle)
        message += '<br><br>'
        table = TABLE()
        for fws_row in fws_rows.render():
            name = fws_row.auth_user.display_name
            table.append(TR(TD(name),
                            TD(unicode(T('Full event')))))
        wsa_rows = activity_list_customers_get_activity_rows(row.id)
        for wsa_row in wsa_rows:
            name = wsa_row.auth_user.display_name,
            table.append(TR(TD(name), TD()))

        message += unicode(table)
        message += '<br>'

        subject = T("Reservations for") + ' ' + workshop.Name + ', ' + \
                  T("starting on") + ' ' + \
                  workshop.Startdate.strftime(DATE_FORMAT) + '.'

    if db.sys_properties(Property='smtp_signature'):
        smtp_signature = db.sys_properties(Property='smtp_signature').PropertyValue
    else:
        smtp_signature = ""
    message += smtp_signature
    message += "<br><br><br><font size=2>" + T("This message is autogenerated by OpenStudio")
    message += " - <a href='http://www.openstudioproject.com'>www.openstudioproject.com</a></font>"
    message += "</body></html>"

    # now send the mail
    to = []
    if workshop.TeacherEmail:
        to.append(workshop.TeacherEmail.strip())
    if workshop.Teacher2Email:
        to.append(workshop.Teacher2Email.strip())
    if workshop.auth_teacher_id:
        teacher = db.auth_user(workshop.auth_teacher_id)
        to.append(teacher.email.strip())
    if workshop.auth_teacher_id2:
        teacher = db.auth_user(workshop.auth_teacher_id2)
        to.append(teacher.email.strip())

    if len(to) < 1:
        session.flash = T("Please check the teachers' email address(es).")
    else:
        check = MAIL.send(
            to=to,
            subject=subject,
            # If reply_to is omitted, then mail.settings.sender is used
            reply_to=None,
            message=message)

        if check:
            session.flash = T("Successfully sent mail")
        else:
            session.flash = T("Unable to send mail")

    redirect(URL('activities', vars={'wsID': wsID}))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'workshops_activities'))
def attendance_delete():
    """
        This function removed the specified attendance record
        request.vars['id'] is expected to be the
            workshops_reservation id (wsr_id)
    """
    response.view = 'generic.json'
    wsr_id = request.vars['id']

    query = (db.workshops_reservation.id == wsr_id)
    result = db(query).delete()

    # result > 0 means something was removed
    if result:
        message = T("Successfully removed")
        status = "success"
    else:
        message = T("Uh oh... something went wrong...")
        status = "fail"

    return dict(message=message, status=status)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'workshops_activities'))
def activity_update_attendance():
    """
        Called as JSON
        This function is meant to be called with the json extension.
        It takes id and Info as variables.
    """
    if not request.extension == 'json':
        return T("Error, please call as JSON")

    response.view = 'generic.json'
    status = "fail"

    if request.vars:
        cuID = request.vars['cuID']
        wsaID = request.vars['wsaID']

        # info
        if 'attending' in request.vars:
            result = db.workshops_activities_customers.insert(
                auth_customer_id=cuID,
                workshops_activities_id=wsaID)
        else:
            query = \
                (db.workshops_activities_customers.auth_customer_id == cuID) & \
                (db.workshops_activities_customers.workshops_activities_id == wsaID)
            result = db(query).delete()

        if result:
            status = "success"
            message = T("Updated attendance")
        else:
            message = T("Uh oh... something went wrong...")
    else:
        message = T("Error: no data received")

    return dict(status=status, message=message)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'workshops_activities'))
def activity_duplicate():
    """
        This function duplicates an activity and
        redirects to the duplicated activity's edit page
    """
    wsaID = request.args[0]
    row = db.workshops_activities[wsaID]

    id = db.workshops_activities.insert(
        workshops_id=row.workshops_id,
        Activity=row.Activity + u' (Copy)',
        Activitydate=row.Activitydate,
        school_locations_id=row.school_locations_id,
        auth_teacher_id=row.auth_teacher_id,
        auth_teacher_id2=row.auth_teacher_id2,
        Starttime=row.Starttime,
        Endtime=row.Endtime,
        Spaces=row.Spaces)

    # Clear cache
    cache_clear_workshops()

    session.flash = T("You are now editing the duplicated activity")

    redirect(URL('activity_edit', args=[id]))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'workshops'))
def stats():
    """
        This function shows a page with a quick view of the revenue and
        a table listing the top 10 cities of attending customers.
        request.vars['wsID'] is expected to be the workshopID
    """
    wsID = request.vars['wsID']
    workshop = db.workshops(wsID)
    response.title = T('Event')
    response.subtitle = get_subtitle(wsID)
    response.view = 'general/tabs_menu.html'

    ### top 10 cities begin
    cities = DIV(stats_get_top10cities(wsID),
                 _class='col-md-6')

    ### top 10 cities end

    ### Revenue begin ###
    revenue = DIV(stats_get_revenue(wsID), _class='col-md-6')

    ### Revenue end ###

    content = DIV(DIV(revenue, cities,
                      _class='col-md-12'),
                  _class='row')

    menu = get_workshops_menu(request.function, wsID)
    back = manage_get_back()

    return dict(content=content,
                menu=menu,
                back=back,
                left_sidebar_enabled=True)


def stats_get_top10cities(wsID):
    """
        Returns overview of top10 cities for a workshop
    """
    count = db.auth_user.city.count()
    query = (db.workshops_products.workshops_id == wsID) & \
            (db.workshops_products_customers.workshops_products_id ==
             db.workshops_products.id) & \
            (db.workshops_products_customers.auth_customer_id ==
             db.auth_user.id) & \
            ((db.auth_user.city != None) & (db.auth_user.city != ''))

    rows = db(query).select(db.auth_user.city, count,
                            groupby=db.auth_user.city,
                            orderby=~count | db.auth_user.city, limitby=(0, 10))

    l = []
    for row in rows:
        l.append([row.auth_user.city, row[count]])

    title = T("Top 10 cities")
    table = TABLE(TR(TH(T("City")), TH(T("Customers"))), _class='table')
    for item in l:
        table.append(TR(*item))

    panel = os_gui.get_panel_table(title, table)

    return panel


def stats_get_revenue(wsID):
    """
        Returns revenue of a workshop, specified by product
    """
    # first get all products
    query = (db.workshops_products.workshops_id == wsID)
    rows = db(query).select(db.workshops_products.ALL,
                            orderby=~db.workshops_products.FullWorkshop | \
                                    db.workshops_products.Name)

    products_ids = []
    for row in rows:
        products_ids.append(row.id)

    # Get all workshops_products_customers rows
    query = (db.workshops_products_customers.workshops_products_id.belongs(products_ids))
    rows = db(query).select(db.workshops_products_customers.ALL)
    wspc_ids = []

    for row in rows:
        wspc_ids.append(row.id)


    # Get invoices
    left = [
        db.invoices_amounts.on(
            db.invoices_workshops_products_customers.invoices_id ==
            db.invoices_amounts.invoices_id
        )
    ]
    query = (db.invoices_workshops_products_customers.workshops_products_customers_id.belongs(wspc_ids))
    rows = db(query).select(db.invoices_workshops_products_customers.ALL,
                            db.invoices_amounts.ALL,
                            left=left)

    revenue_total = 0
    for row in rows:
        try:
            revenue_total += row.invoices_amounts.TotalPriceVAT
        except TypeError:
            pass


    title = T("Revenue")
    table = TABLE(
        TR(
            TH(T("Tickets sold")),
            TD(len(rows))
        ),
        TR(
            TH(T("Revenue")),
            TD(SPAN(CURRSYM, ' ', format(revenue_total, '.2f')))
        ),
        _class='table table-hover table-striped'
    )

    panel = os_gui.get_panel_table(title, table)

    return panel


def manage_get_back(var=None):
    """
        Generate back button for workshops manage pages
    """
    if session.workshops_manage_back == 'default_workshoppayments_open':
        url = URL('default', 'workshop_payments_open')
    elif session.workshops_manage_back == 'pinboard':
        url = URL('pinboard', 'index')
    elif session.workshops_manage_back == 'tasks_index':
        url = URL('tasks', 'index')
    else:
        url = URL('index')

    return os_gui.get_button('back', url)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'workshops_products'))
def tickets():
    """
        Lists products for a workshop
        request.vars['wsID'] is expected to be the workshops_id
    """
    wsID = request.vars['wsID']
    response.title = T('Event')
    response.subtitle = get_subtitle(wsID)
    response.view = 'events/manage.html'

    ## Products begin ##
    products = DIV()
    products.append(event_get_alert_no_activities(wsID))

    add = ''
    perm = auth.has_membership(group_id='Admins') or \
           auth.has_permission('create', 'workshops_products')
    if perm:
        add_url = URL('ticket_add', args=[wsID])
        add = os_gui.get_button('add', add_url, T("Add new ticket"), _class='pull-right')

    table = TABLE(THEAD(TR(TH(T('Name')),
                           TH(T('Description')),
                           TH(T('Price')),
                           TH(T('Shop')),
                           TH(),
                           TH(),
                           _class='header')),
                  _class='table table-hover')

    query = (db.workshops_products.workshops_id == wsID)
    rows = db(query).select(db.workshops_products.ALL)
    for row in rows.render():
        edit = ''
        duplicate = ''
        delete = ''
        activities = ''

        # shop checkbox
        shop = INPUT(_type="checkbox",
                     _disabled="disabled")
        if row.PublicProduct:
            shop['_checked'] = "checked"

        if row.FullWorkshop:
            fws_label = get_fullws_label()
        else:
            fws_label = ''

        # check permission for adding activities to products
        perm = auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'workshops_products_activities')
        if perm and not row.FullWorkshop:
            activities = os_gui.get_button('list_notext',
                                           URL('ticket_activities',
                                               vars={'wsID': wsID,
                                                     'wspID': row.id}),
                                           tooltip=T("Activities"))

        # check permission to update workshops (edit and delete)
        perm = auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'workshops')
        if perm:
            edit = os_gui.get_button('edit_notext',
                                     URL('ticket_edit', args=[row.id]),
                                     tooltip=T('Edit ticket'))

        # check permission to create workshop products
        perm = auth.has_membership(group_id='Admins') or \
               auth.has_permission('create', 'workshops_products')
        if perm:
            duplicate = os_gui.get_button('duplicate',
                                          URL('ticket_duplicate', vars={'wspID':row.id}),
                                          tooltip=T("Duplicate product"))

        # check delete permission
        if row.Deletable:
            confirm_msg = T("Really remove this ticket?")
            onclick = "return confirm('" + confirm_msg + "');"
            delete = ''
            if auth.has_membership(group_id='Admins') or \
                    auth.has_permission('delete', 'workshops_products'):
                delete = os_gui.get_button('delete_notext',
                                           URL('ticket_delete', vars={'wsID': wsID,
                                                                       'wspID': row.id}),
                                           onclick=onclick,
                                           tooltip=T('Delete ticket'),
                                           _class='pull-right')

        # check permission to view customers for a product
        perm = auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'workshops_products_customers')
        if perm:
            customers = os_gui.get_button(
                'user_notext',
                URL('tickets_list_customers', vars={'wsID': wsID,
                                                     'wspID': row.id}),
                tooltip=T('Customers')
            )
        else:
            customers = ''

        buttons = DIV(delete,
                      DIV(customers,
                          activities,
                          edit,
                          duplicate,
                          _class='btn-group pull-right'))

        table.append(TR(TD(row.Name),
                        TD(max_string_length(row.Description, 40)),
                        TD(row.Price),
                        TD(shop),
                        TD(fws_label),
                        TD(buttons)))

    products.append(table)

    ## Products end ##

    export = tickets_get_export(wsID)

    content = products

    menu = get_workshops_menu(request.function, wsID)
    back = manage_get_back()

    return dict(content=content,
                menu=menu,
                back=back,
                header_tools=export,
                add=add)


def tickets_get_export(wsID):
    """
        Returns export drop down for schedule
    """
    mailinglist = A(os_gui.get_fa_icon('fa-envelope-o'), T("Mailing list"),
                    _href=URL('tickets_export_excel',
                              vars={'wsID': wsID, 'export_type':'mailinglist'}))
    attendancelist = A(os_gui.get_fa_icon('fa-check-square-o'), T("Attendance list"),
                    _href=URL('tickets_export_excel',
                              vars={'wsID': wsID, 'export_type':'attendancelist'}))

    links = [mailinglist, attendancelist]

    export = os_gui.get_dropdown_menu(
        links=links,
        btn_text='',
        btn_size='btn-sm',
        btn_icon='download',
        menu_class='pull-right')

    return export


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'workshops_products'))
def tickets_export_excel():
    """
        Export mailinglist for a workshop product.
        Have one worksheet for each product and one for all products
    """
    def add_ticket_sheet(wspID):
        wsp = WorkshopProduct(wspID)
        # add sheet
        ws = wb.create_sheet(title=wsp.name)
        # get db info
        left = [ db.auth_user.on(db.auth_user.id == db.workshops_products_customers.auth_customer_id),
                 db.workshops_products.on(
                     db.workshops_products_customers.workshops_products_id == db.workshops_products.id),
                 db.invoices_workshops_products_customers.on(
                     db.invoices_workshops_products_customers.workshops_products_customers_id ==
                     db.workshops_products_customers.id),
                 db.invoices.on(db.invoices_workshops_products_customers.invoices_id == db.invoices.id)
                 ]
        query = ((db.workshops_products_customers.workshops_products_id == wspID) &
                 (db.workshops_products_customers.Cancelled == False))
        rows = db(query).select(db.auth_user.first_name,
                                db.auth_user.last_name,
                                db.auth_user.email,
                                db.invoices.InvoiceID,
                                db.workshops_products.Name,
                                left=left)

        # add info to sheet
        for row in rows:
            if export_type == 'mailinglist':
                ws.append([row.auth_user.first_name,
                           row.auth_user.last_name,
                           row.auth_user.email])
            elif export_type == 'attendancelist':
                ws.append([row.auth_user.first_name,
                           row.auth_user.last_name,
                           row.auth_user.email,
                           row.invoices.InvoiceID,
                           row.workshops_products.Name])

    wsID = request.vars['wsID']
    export_type = request.vars['export_type']
    workshop = Workshop(wsID)

    # create filestream
    stream = cStringIO.StringIO()

    wb = openpyxl.workbook.Workbook(write_only=True)
    # write the sheet for all mail addresses
    ws = wb.create_sheet(title="All products")
    # get all products for a workshop

    products = workshop.get_products()
    product_ids = []
    for product in products:
        product_ids.append(product.id)


    left = [ db.auth_user.on(db.auth_user.id == db.workshops_products_customers.auth_customer_id),
             db.workshops_products.on(db.workshops_products_customers.workshops_products_id == db.workshops_products.id),
             db.invoices_workshops_products_customers.on(
                 db.invoices_workshops_products_customers.workshops_products_customers_id ==
                 db.workshops_products_customers.id),
             db.invoices.on(db.invoices_workshops_products_customers.invoices_id == db.invoices.id),
             ]
    query = ((db.workshops_products_customers.workshops_products_id.belongs(product_ids)) &
             (db.workshops_products_customers.Cancelled == False))
    rows = db(query).select(db.auth_user.first_name,
                            db.auth_user.last_name,
                            db.auth_user.email,
                            db.invoices.InvoiceID,
                            db.workshops_products.Name,
                            left=left)

    emails = []

    for row in rows:
        # distinct isn't working here in w2p 2.15.4, this is a workaround
        if row.auth_user.email not in emails:
            if export_type == 'mailinglist':
                ws.append([row.auth_user.first_name,
                           row.auth_user.last_name,
                           row.auth_user.email])
            elif export_type == 'attendancelist':
                ws.append([row.auth_user.first_name,
                           row.auth_user.last_name,
                           row.auth_user.email,
                           row.invoices.InvoiceID,
                           row.workshops_products.Name])
        emails.append(row.auth_user.email)

    # Add all products
    for wspID in product_ids:
        add_ticket_sheet(wspID)

    # Return to user
    if export_type == 'mailinglist':
        fname = T("Mailinglist") + '.xlsx'
    elif export_type == 'attendancelist':
        fname = T('AttendanceList') + '.xlsx'
    wb.save(stream)

    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-disposition'] = 'attachment; filename=' + fname

    return stream.getvalue()


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('create', 'workshops_products'))
def ticket_duplicate():
    """
        Duplicate workshop product & linked activities
    """
    wspID = request.vars['wspID']

    wsp = db.workshops_products(wspID)

    new_wspID = db.workshops_products.insert(
        workshops_id = wsp.workshops_id,
        FullWorkshop = False,
        Deletable = True,
        PublicProduct = False,
        Name = wsp.Name + ' (Copy)',
        Price = wsp.Price,
        PriceSubscription = wsp.PriceSubscription,
        PriceEarlybird = wsp.PriceEarlybird,
        PriceSubscriptionEarlybird = wsp.PriceSubscriptionEarlybird,
        EarlybirdUntil = wsp.EarlybirdUntil,
        tax_rates_id = wsp.tax_rates_id,
        Description = wsp.Description,
        ExternalShopURL = wsp.ExternalShopURL,
        AddToCartText = wsp.AddToCartText,
        Donation = wsp.Donation
    )

    if wsp.FullWorkshop:
        query = (db.workshops_activities.workshops_id == wsp.workshops_id)
        rows = db(query).select(db.workshops_activities.id)
        for row in rows:
            db.workshops_products_activities.insert(
                workshops_products_id=new_wspID,
                workshops_activities_id=row.id
            )
    else:
        query = (db.workshops_products_activities.workshops_products_id == wspID)
        rows = db(query).select(db.workshops_products_activities.ALL)
        for row in rows:
            db.workshops_products_activities.insert(
                workshops_products_id = new_wspID,
                workshops_activities_id = row.workshops_activities_id
            )

    session.flash = T('You are now editing the duplicated ticket')

    redirect(URL('events', 'ticket_edit', args=new_wspID))


def event_get_alert_no_activities(wsID):
    """
        Display an information banner to notify user when no activities have been found
    """
    alert = ''

    query = (db.workshops_activities.workshops_id == wsID)
    count = db(query).count()
    if not count:
        add = os_gui.get_button('add', URL('activity_add', args=wsID),
                                btn_class='btn-link',
                                title=T('Add activity'))
        alert = os_gui.get_alert('info', SPAN(
            SPAN(T('No activities have been found for this event'), _class="bold"), ' ',
            add))

    return alert


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'workshops_activities'))
def activities():
    """
        Manage agenda for workshops
    """
    wsID = request.vars['wsID']
    response.title = T('Event')
    response.subtitle = get_subtitle(wsID)
    response.view = 'events/manage.html'

    ## Modals container
    modals = DIV()

    ## Agenda begin  ##
    agenda = DIV()

    add = ''
    perm = auth.has_membership(group_id='Admins') or \
           auth.has_permission('create', 'workshops_activities')
    if perm:
        add_url = URL('activity_add', args=[wsID])
        add = os_gui.get_button('add', add_url, T("Add new activity"), _class='pull-right')

    agenda_items = TABLE(THEAD(TR(TH(T("Date")),
                                  TH(T("Time")),
                                  TH(T("Activity")),
                                  TH(T("Location")),
                                  TH(T('Teacher')),
                                  TH(T("Filled")),
                                  TH())),
                         _class='table table-hover')
    query = (db.workshops_activities.workshops_id == wsID)
    rows = db(query).select(db.workshops_activities.ALL,
                            orderby=db.workshops_activities.Activitydate | \
                                    db.workshops_activities.Starttime)
    for row in rows.render():
        wsaID = row.id

        internal_location = not row.school_locations_id is None and \
                            not row.school_locations_id == ''
        if internal_location:
            location = row.school_locations_id
        else:
            location = row.LocationExternal

        fwsID = workshops_get_full_workshop_product_id(wsID)
        activity = TR(TD(row.Activitydate),
                      TD(row.Starttime, ' - ', row.Endtime),
                      TD(row.Activity),
                      TD(location),
                      TD(row.auth_teacher_id),
                      TD(activity_get_filled(row, fwsID)),
                      )

        # check permissions to manage customers attendance
        customers = ''

        perm = auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'workshops_activities_customers')
        if perm:
            customers = os_gui.get_button('user',
                                          URL('activity_list_customers', vars={'wsID':wsID, 'wsaID':wsaID}),
                                          tooltip=T('List customers for this activity'))
            # modal_title = row.Activity + ' ' + \
            #               row.Activitydate + ' ' + \
            #               T('Customers')
            # load_content = os_gui.get_ajax_loader(message=T("Loading..."))
            # modal_content = LOAD('workshops', 'activity_list_customers.load',
            #                      ajax=True,
            #                      content=load_content,
            #                      vars={'wsID': wsID, 'wsaID': row.id})
            #
            # customers_button_id = 'wsa_cust_' + unicode(row.id)
            # btn_icon = SPAN(_class='glyphicon glyphicon-user')
            #
            # result = os_gui.get_modal(button_text=XML(btn_icon),
            #                           modal_title=modal_title,
            #                           modal_content=modal_content,
            #                           modal_class=customers_button_id,
            #                           modal_size='',
            #                           button_class='btn-sm workshops_show_customers',
            #                           button_id=customers_button_id)
            # modals.append(result['modal'])
            # customers = result['button']

        # check permission to edit activities
        edit = ''
        duplicate = ''
        delete = ''

        perm = auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'workshops_activities')
        if perm:
            edit = os_gui.get_button('edit_notext',
                                     URL('activity_edit', args=[wsaID]),
                                     tooltip=T("Edit activity"))

            duplicate = os_gui.get_button('duplicate',
                                          URL('activity_duplicate', args=[wsaID]),
                                          tooltip=T("Duplicate activity"))

        if auth.has_membership(group_id='Admins') or \
                auth.has_permission('delete', 'workshops_activities'):
            confirm_msg = T("Are you sure you want to delete this activity?")
            delete = os_gui.get_button('delete_notext',
                                       URL('activity_delete', args=[wsID, wsaID]),
                                       tooltip=T('Delete activity'),
                                       onclick="return confirm('" + confirm_msg + "');")

        buttons = DIV(customers,
                      edit,
                      duplicate,
                      delete,
                      _class='btn-group pull-right')
        activity.append(TD(buttons))

        agenda_items.append(activity)

    agenda.append(agenda_items)

    permission = auth.has_membership(group_id='Admins') or \
                 auth.has_permission('update', 'class_status')
    if permission:
        oc_count = overlapping_classes_get_count_all(wsID)
        if oc_count == 0:
            badge_type = 'success'
        else:
            badge_type = 'default'
        oc_badge = get_badge(badge_type, oc_count)
        oc = DIV(A(T("Overlapping classes"), ' ', oc_badge,
                   _href=URL('overlapping_classes', vars={'wsID': wsID}),
                   _class='btn btn-primary btn-sm'),
                 _class='pull-right ')
        agenda.append(oc)
    mail = A(SPAN(_class='glyphicon glyphicon-envelope'), ' ',
             T("Mail attendance lists to teacher"),
             _href=URL('mail_activity_attendance', vars={'wsID': wsID}),
             _class='btn btn-primary btn-sm')
    agenda.append(mail)

    ## Agenda end ##

    export_activities = A(SPAN(os_gui.get_fa_icon('fa-check'), ' ',
                               T("Attendance")),
                          _href=URL('activities_export_attendance',
                                    vars={'wsID': wsID}))

    links = [export_activities]

    export = os_gui.get_dropdown_menu(
        links=links,
        btn_text='',
        btn_size='btn-sm',
        btn_icon='download',
        menu_class='pull-right')

    content = DIV(agenda, modals)

    menu = get_workshops_menu(request.function, wsID)
    back = manage_get_back()

    return dict(content=content,
                menu=menu,
                back=back,
                header_tools=export,
                add=add)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'workshops_activities'))
def activities_export_attendance():
    """
        Create excel export of workshop attendance
        request.vars['wsID'] is expected to be workshops.id
    """
    wsID = request.vars['wsID']
    workshop = db.workshops(wsID)

    query = (db.workshops_activities.workshops_id == wsID)
    rows = db(query).select(db.workshops_activities.id,
                            orderby=db.workshops_activities.Activitydate)
    wsa_ids = []
    for row in rows:
        wsa_ids.append(row.id)

    fw_rows = activity_list_customers_get_fullws_rows(wsID)

    # create filestream
    stream = cStringIO.StringIO()

    # Create the workbook
    title = 'test'
    wb = openpyxl.workbook.Workbook(write_only=True)

    # process activities
    for wsaID in wsa_ids:
        activity = db.workshops_activities(wsaID)
        title = activity.Activity[0:30]
        date = activity.Activitydate.strftime(DATE_FORMAT)
        start = activity.Starttime.strftime(TIME_FORMAT)
        ws = wb.create_sheet(title=title)
        desc = [workshop.Name + ' ' + date + ' ' + title + ' ' + start]
        ws.append(desc)
        ws.append([' '])
        header = ['Customer', 'Product', 'Attending']
        ws.append(header)

        # add full workshop customers
        for row in fw_rows.render():
            cuID = row.auth_user.id

            # check attendance
            attending = ''
            check = db.workshops_activities_customers(
                auth_customer_id=cuID,
                workshops_activities_id=wsaID)
            if check:
                attending = 'X'

            custname = row.auth_user.display_name
            excel_row = [custname, 'Full event', attending]
            ws.append(excel_row)

        # add customers with another product
        a_rows = activity_list_customers_get_activity_rows(wsaID)
        for row in a_rows.render():
            cuID = row.auth_user.id
            product = row.workshops_products_customers.workshops_products_id

            # check attendance
            attending = ''
            check = db.workshops_activities_customers(auth_customer_id=cuID,
                                                      workshops_activities_id=wsaID)
            if check:
                attending = 'X'

            custname = row.auth_user.display_name
            excel_row = [custname, product, attending]
            ws.append(excel_row)

    wb.save(stream)
    fname = 'Attendance.xlsx'
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-disposition'] = 'attachment; filename=' + fname

    return stream.getvalue()


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'workshops_activities'))
def activity_list_customers():
    """
        List customers for an activity
        request.args[0] is expected to be the workshops_id (wsID)
        request.args[1] is expected to be the workshops_activities_id (wsaID)
    """
    wsID = request.vars['wsID']
    wsaID = request.vars['wsaID']
    response.title = T('Event')
    response.subtitle = get_subtitle(wsID)
    response.view = 'events/manage.html'

    # get activity attendance
    wsa = db.workshops_activities(wsaID)

    table = activity_list_customers_get_list(wsID, wsaID)
    content = DIV(
        H4(T('Customers for '), wsa.Activity),
        table
    )

    menu = get_workshops_menu('activities', wsID)
    back = os_gui.get_button('back', URL('activities', vars={'wsID':wsID}))

    return dict(content=content, back=back, menu=menu)


def activity_list_customers_get_list(wsID,
                                     wsaID):
    """
        Lists customers for activity
    """

    def process_rows(table, rows, fullWS=False):
        """
            Helper to add rows to table
        """
        for row in rows.render():
            # get customer name
            cust_name = TD(row.auth_user.display_name, BR(),
                           _class='os-customer_name')
            if fullWS:
                cust_name.append(get_fullws_label())
            else:
                product = row.workshops_products_customers.workshops_products_id
                cust_name.append(get_label('default', product))

            # get attendance checkbox
            cuID = row.auth_user.id
            attendance = activity_list_customers_get_list_get_attendance(cuID,
                                                                         wsaID,
                                                                         wsID)
            table.append(TR(TD(row.auth_user.thumbsmall,
                               _class='os-customer_image_td'),
                            cust_name,
                            TD(attendance)))
        return table


    table = TABLE(_class='table')

    # Add full workshop customers
    rows = activity_list_customers_get_fullws_rows(wsID)
    table = process_rows(table, rows, fullWS=True)

    # Add activity customers
    rows = activity_list_customers_get_activity_rows(wsaID)
    table = process_rows(table, rows)

    return table


def activity_list_customers_get_fullws_rows(wsID):
    """
        Returns rows object for full workshops customers of a workshop
    """
    orderby = ~db.auth_user.display_name
    fwsID = workshops_get_full_workshop_product_id(wsID)
    query = (db.workshops_products_customers.Cancelled == False) & \
            (db.workshops_products_customers.Waitinglist == False) & \
            (db.workshops_products_customers.workshops_products_id == fwsID)
    rows = db(query).select(
        db.auth_user.id,
        db.auth_user.trashed,
        db.auth_user.thumbsmall,
        db.auth_user.birthday,
        db.auth_user.display_name,
        db.workshops_products_customers.workshops_products_id,
        left=[db.auth_user.on(db.workshops_products_customers.auth_customer_id == \
                              db.auth_user.id)],
        orderby=orderby)

    return rows


def activity_list_customers_get_activity_rows(wsaID):
    """
        Returns rows for a workshop activity
    """
    orderby = ~db.auth_user.display_name
    left = [db.auth_user.on(db.auth_user.id == \
                            db.workshops_products_customers.auth_customer_id),
            db.workshops_products.on(db.workshops_products.id == \
                                     db.workshops_products_customers.workshops_products_id)]

    query = activity_list_customers_get_list_activity_query(wsaID)
    rows = db(query).select(
        db.auth_user.id,
        db.auth_user.trashed,
        db.auth_user.thumbsmall,
        db.auth_user.birthday,
        db.auth_user.display_name,
        db.workshops_products_activities.workshops_activities_id,
        db.workshops_products_customers.workshops_products_id,
        left=left,
        orderby=orderby)

    return rows


def activity_list_customers_get_list_activity_query(wsaID):
    """
        Returns a query that returns a set of all customers in a specific
        workshop activity, without the full workshop customers
    """
    query = (db.workshops_activities.id ==
             db.workshops_products_activities.workshops_activities_id) & \
            (db.workshops_products_customers.workshops_products_id ==
             db.workshops_products_activities.workshops_products_id) & \
            (db.workshops_products_activities.workshops_activities_id ==
             wsaID) & \
            (db.workshops_products_customers.Waitinglist == False) & \
            (db.workshops_products_customers.Cancelled == False)

    return query


def activity_list_customers_get_list_get_attendance(cuID, wsaID, wsID):
    """
        Checks whether a customer is attending a class
    """
    check = db.workshops_activities_customers(auth_customer_id=cuID,
                                              workshops_activities_id=wsaID)

    vars = {'wsaID':wsaID,
            'wsID': wsID,
            'cuID': cuID}
    if check:
        vars['wsacID'] = check.id

        onclick = "return confirm('" + \
                  T('Do you really want to check out this customer?') + "');"

        attending = DIV(
            os_gui.get_label(
                'success',
                T("attending")
            ), ' ',
            os_gui.get_button(
                'delete_notext',
                URL('activity_check_out_customer', vars=vars),
                onclick=onclick,
            ),
            _class='pull-right'

        )
    else:
        attending = os_gui.get_button(
            'noicon',
            URL('activity_check_in_customer', vars=vars),
            title=T('check-in'),
            _class='pull-right'
        )

    return attending


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'workshops_activities_customers'))
def activity_check_in_customer():
    """
        Check in customer
    """
    cuID = request.vars['cuID']
    wsID = request.vars['wsID']
    wsaID = request.vars['wsaID']

    db.workshops_activities_customers.insert(
        auth_customer_id = cuID,
        workshops_activities_id = wsaID
    )

    redirect(URL('activity_list_customers', vars={'wsID':wsID,
                                                  'wsaID':wsaID}))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'workshops_activities_customers'))
def activity_check_out_customer():
    """
        Check in customer
    """
    cuID = request.vars['cuID']
    wsID = request.vars['wsID']
    wsaID = request.vars['wsaID']
    wsacID = request.vars['wsacID']

    query = (db.workshops_activities_customers.id == wsacID)
    db(query).delete()

    redirect(URL('activity_list_customers', vars={'wsID':wsID,
                                                  'wsaID':wsaID}))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'workshops_products_customers'))
def tickets_list_customers():
    """
        Lists customers for a product
    """

    # set workshop ID
    if 'wsID' in request.vars:
        wsID = request.vars['wsID']
        session.events_tickets_list_customers_wsID = wsID
    elif session.events_tickets_list_customers_wsID:
        wsID = session.events_tickets_list_customers_wsID

    # set workshop product ID
    if 'wspID' in request.vars:
        wspID = request.vars['wspID']
        session.events_tickets_list_customers_wspID = wspID
    elif session.events_tickets_list_customers_wspID:
        wspID = session.events_tickets_list_customers_wspID

    session.invoices_edit_back = 'events_tickets_list_customers'
    session.invoices_payment_add_back = 'events_tickets_list_customers'
    session.workshops_product_resend_info_mail = 'events_tickets_list_customers'

    wsp = WorkshopProduct(wspID)

    response.title = T('Event')
    subtitle = SPAN(get_subtitle(wsID), ' > ', T('Customers'), ': ', wsp.name)
    response.subtitle = subtitle
    #response.view = 'events/ticket_list_customers.html'

    session.customers_payment_back = 'workshops'
    session.events_ticket_sell_back = None

    add = ''
    perm = auth.has_membership(group_id='Admins') or \
           auth.has_permission('update', 'workshops_products_customers')
    if perm:
        sold_out = ticket_check_sold_out(wsID, wspID)
        if not sold_out:
            response.search_available = True
            add = ''
            #add = SPAN(T('Search to add customers...'), _class='grey')
        else:
            response.search_available = False
            add = DIV(  # BR(),
                T("Sold out,"),
                ' ',
                T("one or more activities listed in this "),
                T("product is fully booked."),
                _class='green')

    full_wspID = workshops_get_full_workshop_product_id(wsID)
    table = TABLE(_class='table')
    ## sold products
    # add full workshop customers for other activities
    # if not int(wspID) == full_wspID:
    #     table = products_list_customers_get_list(table,
    #                                              wsID,
    #                                              full_wspID,
    #                                              fullWS=True,
    #                                              deletable=False)
    # add customers for selected  product
    if int(wspID) == full_wspID:  # show "Full workshop for Full workhshop act."
        fullWS = True
    else:
        fullWS = False
    table = tickets_list_customers_get_list(table,
                                            wsID,
                                            wspID,
                                            fullWS,
                                            deletable=True)

    name = request.vars['name']
    search_results = DIV(LOAD('customers', 'load_list.load',
                              target='customers_list',
                              content=os_gui.get_ajax_loader(message=T("Searching...")),
                              vars={'list_type':'events_ticket_sell',
                                    'items_per_page':10,
                                    'wsID':wsID,
                                    'wspID':wspID},
                              ajax=True),
                         _id="customers_list",
                         _class="load_list_customers clear",
                         _style="display:none;")


    content = DIV(search_results,
                  DIV(add, _class='pull-right'),
                  H4(T('Customers for'), ' ', wsp.name), BR(), table)

    menu = get_workshops_menu('tickets', wsID)
    back = os_gui.get_button('back', URL('events', 'tickets', vars={'wsID':wsID}))
    loader = os_gui.get_ajax_loader(message=T("Refreshing list..."))

    return dict(content=content,
                menu=menu,
                back=back,
                loader=loader)


def tickets_list_customers_get_list(table,
                                    wsID,
                                    wspID,
                                    fullWS=False,
                                    deletable=False):
    """
        Append customers to table
    """
    wh = WorkshopsHelper()
    db_icwspc = db.invoices_workshops_products_customers

    query = (db.workshops_products_customers.workshops_products_id == wspID)
    rows = db(query).select(
        db.workshops_products_customers.ALL,
        db.auth_user.id,
        db.auth_user.trashed,
        db.auth_user.thumbsmall,
        db.auth_user.birthday,
        db.auth_user.display_name,
        db.invoices.id,
        db.invoices.InvoiceID,
        db.invoices.Status,
        db.invoices.payment_methods_id,
        left=[db.auth_user.on(db.workshops_products_customers.auth_customer_id == \
                              db.auth_user.id),
              # db.invoices.on(db.invoices.workshops_products_customers_id ==
              #                db.workshops_products_customers.id)
              db.invoices_workshops_products_customers.on(
                  db_icwspc.workshops_products_customers_id == db.workshops_products_customers.id),
              db.invoices.on(db_icwspc.invoices_id == db.invoices.id)],
        orderby=db.workshops_products_customers.Cancelled | \
                db.workshops_products_customers.Waitinglist | \
                db.auth_user.display_name)

    left = [db.workshops_products.on(
        db.workshops_products.id == \
        db.workshops_products_customers.workshops_products_id),
               db.workshops.on(db.workshops_products.workshops_id == \
                               db.workshops.id),
               db.invoices_workshops_products_customers.on(
                   db_icwspc.workshops_products_customers_id ==
                   db.workshops_products_customers.id),
               db.invoices.on(db_icwspc.invoices_id == db.invoices.id)
           ],

    invoices = Invoices()
    for i, row in enumerate(rows):
        repr_row = list(rows[i:i + 1].render())[0]

        wsp_cuID = row.workshops_products_customers.id

        # Event info link
        link_text = T('Send')
        if row.workshops_products_customers.WorkshopInfo:
            link_text = T('Resend')
        resend_link = A(link_text, ' ', T('info mail'),
                        _href=URL('product_resend_info_mail', vars={'wspcID':wsp_cuID}))
        event_info = wh.get_customer_info(wsp_cuID,
                                          wsID,
                                          row.workshops_products_customers.WorkshopInfo,
                                          resend_link)

        cust_name = TD(SPAN(row.auth_user.display_name, _class='bold'), BR())
        if fullWS:
            cust_name.append(get_fullws_label())
        else:
            product = repr_row.workshops_products_customers.workshops_products_id
            cust_name.append(get_label('default', product))

        tr_class = ''
        if row.workshops_products_customers.Cancelled:
            if fullWS:
                cust_name.append(' ')
            cust_name.append(get_cancelled_label())

        if row.workshops_products_customers.Waitinglist:
            wai_url = URL('product_remove_customer_from_waitinglist',
                          vars={'wsp_cuID': wsp_cuID,
                                'wsID': wsID})
            if fullWS:
                cust_name.append(' ')
            else:
                cust_name.append(BR())
            cust_name.append(SPAN(T("Waitinglist"),
                                  _class='label label-danger'))
            cust_name.append(BR())
            if not ticket_check_sold_out(wsID, wspID):
                msg = T("Remove from waitinglist, add to list")
                cust_name.append(SPAN(A(msg,
                                        _href=wai_url,
                                        cid=request.cid),
                                      _class='small_font'))

        # invoice
        if row.invoices.id:
            invoice = invoices.represent_invoice_for_list(
                row.invoices.id,
                repr_row.invoices.InvoiceID,
                repr_row.invoices.Status,
                row.invoices.Status,
                row.invoices.payment_methods_id
            )
        else:
            invoice = ''

        buttons = tickets_list_customers_get_buttons(wsID, wsp_cuID)

        table.append(TR(TD(repr_row.auth_user.thumbsmall,
                           _class='os-customer_image_td'),
                        cust_name,
                        TD(invoice),
                        TD(event_info),
                        TD(' ', buttons, _class='td-icons'),
                        _class=tr_class))

    return table


def tickets_list_customers_get_cancelled(wsp_cuID):
    """
        Returns whether or not a customer has cancelled a workshop product
    """
    row = db.workshops_products_customers(id=wsp_cuID)
    return row.Cancelled


def tickets_list_customers_get_buttons(wsID, wsp_cuID):
    """
        returns buttons for produtcs_list_customers
    """
    confirm_remove_msg = T("Really remove this customer from the list?")
    cid = request.cid

    cancelled = tickets_list_customers_get_cancelled(wsp_cuID)
    if cancelled:
        title_cancel = T('Undo cancellation')
    else:
        title_cancel = T('Cancel customer')

    btn_delete = ''
    if auth.has_membership(group_id='Admins') or \
            auth.has_permission('delete', 'workshops_products_customers'):
        btn_delete = os_gui.get_button(
            'delete_notext',
            URL('ticket_delete_customer', vars={'wsID': wsID,
                                                 'wsp_cuID': wsp_cuID}),
            cid=cid,
            tooltip=T('Remove customer from list'),
            onclick="return confirm('" + confirm_remove_msg + "');",
        )

    btn_cancel = os_gui.get_button(
        'cancel_notext',
        URL('ticket_cancel_customer',
            vars={'wsID': wsID,
                  'wsp_cuID': wsp_cuID}),
        cid=cid,
        tooltip=title_cancel)

    return DIV(btn_cancel,
               btn_delete,
               _class='btn-group pull-right')


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'workshops_activities'))
def ticket_customer_update_info():
    """
        Called as JSON
        This function is meant to be called with the json extension.
        It takes id and WorkshopInfo as variables.
    """
    if not request.extension == 'json':
        return T("Error, please call as JSON")

    response.view = 'generic.json'
    status = "fail"

    if request.vars:
        wsp_cuID = request.vars['id']
        row = db.workshops_products_customers(wsp_cuID)

        # info
        if 'WorkshopInfo' in request.vars:
            row.WorkshopInfo = 'T'
        else:
            row.WorkshopInfo = 'F'

        result = row.update_record()

        if result:
            status = "success"
            message = T("Saved")
        else:
            message = T("Uh oh... something went wrong...")
    else:
        message = T("Error: no data received")

    return dict(status=status, message=message)


def ticket_count_sold(wspID):
    """
        Returns count of customers who bought a product
    """
    query = (db.workshops_products_customers.workshops_products_id == wspID)
    return db(query).count()


def tickets_get_return_url(wsID):
    """
        return the return URL for products add and edit pages
    """
    return URL('tickets', vars={'wsID': wsID})


@auth.requires_login()
def ticket_add():
    """
        Add a new ticket
    """
    wsID = request.args[0]

    response.title = T('Event')
    response.subtitle = get_subtitle(wsID)
    response.view = 'general/tabs_menu.html'

    return_url = tickets_get_return_url(wsID)
    next_url = '/events/ticket_activities?wspID=[id]&wsID=' + wsID

    db.workshops_products.workshops_id.default = wsID
    crud.messages.submit_button = T("Save")
    crud.messages.record_created = T("Saved product")
    crud.settings.create_next = next_url
    crud.settings.create_onaccept = [cache_clear_workshops]
    crud.settings.formstyle = 'bootstrap3_stacked'
    form = crud.create(db.workshops_products)

    form_id = "MainForm"
    form_element = form.element('form')
    form['_id'] = form_id

    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = form_id

    submit = form.element('input[type=submit]')

    back = os_gui.get_button('back', return_url)
    content = DIV(H4(T('New ticket')), BR(), form)

    menu = get_workshops_menu('tickets', wsID)


    return dict(content=content, back=back, menu=menu, save=submit)


@auth.requires_login()
def ticket_edit():
    """
        Edit ticket
    """
    wspID = request.args[0]
    wsID = db.workshops_products(wspID).workshops_id

    response.title = T('Event')
    response.subtitle = get_subtitle(wsID)
    response.view = 'general/tabs_menu.html'

    return_url = URL('tickets',
                     vars={'wsID': wsID})

    crud.messages.submit_button = T("Save")
    crud.messages.record_updated = T("Saved product")
    crud.settings.update_next = return_url
    crud.settings.update_onaccept = [cache_clear_workshops]
    crud.settings.update_deletable = False
    crud.settings.formstyle = 'bootstrap3_stacked'
    form = crud.update(db.workshops_products, wspID)

    form_id = "MainForm"
    form_element = form.element('form')
    form['_id'] = form_id

    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = form_id

    submit = form.element('input[type=submit]')

    content = DIV(H4(T('Edit ticket')), BR(), form)

    menu = get_workshops_menu('tickets', wsID)

    back = os_gui.get_button('back', URL('tickets', vars={'wsID':wsID}))

    return dict(content=content, back=back, menu=menu, save=submit)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'workshops_products'))
def ticket_delete():
    """
        Delete a selected product
    """
    wsID = request.vars['wsID']
    wspID = request.vars['wspID']

    if db.workshops_products(id=wspID).Deletable:
        query = (db.workshops_products.id == wspID)
        db(query).delete()

        # Clear cache
        cache_clear_workshops()
    else:
        session.flash = T("Can't delete this product")

    redirect(URL('tickets', vars={'wsID': wsID}))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'workshops_products'))
def ticket_activities():
    """
        Add or remove activities from a workshop
    """
    wsID = request.vars['wsID']
    wspID = request.vars['wspID']

    product = db.workshops_products(wspID)

    response.title = T('Event')
    response.subtitle = get_subtitle(wsID)
    response.view = 'general/tabs_menu.html'

    pquery = (db.workshops_products_activities.workshops_products_id == wspID)
    field = db.workshops_products_activities.workshops_activities_id
    selected = db(pquery).select(field)

    selected_ids = []
    for row in selected:
        selected_ids.append(row.workshops_activities_id)

    query = (db.workshops_activities.workshops_id == wsID)
    rows = db(query).select(db.workshops_activities.ALL)

    form = FORM(_id="MainForm")
    table = TABLE(THEAD(TR(TH(),
                           TH(T('Date')),
                           TH(T('Time')),
                           TH(T('Activity')),
                           TH(T('Location')))),
                  _class='table table-hover')

    for row in rows.render():
        if row.id in selected_ids:
            checkbox = INPUT(_name=row.id,
                             _value='on',
                             _checked='checked',
                             _type='checkbox')
        else:
            checkbox = INPUT(_name=row.id,
                             _value='on',
                             _type='checkbox')

        internal_location = not row.school_locations_id is None and \
                            not row.school_locations_id == ''
        if internal_location:
            location = row.school_locations_id
        else:
            location = row.LocationExternal

        table.append(TR(TD(checkbox),
                        TD(row.Activitydate),
                        TD(row.Starttime + ' - ' + row.Endtime),
                        TD(row.Activity),
                        TD(location),
                        ))

    form.append(table)

    content = DIV(DIV(H4(T('Activities included in'), ' ', product.Name), form, _class='col-md-8 clear'),
                  _class='row')

    if form.process().accepted:
        db(pquery).delete()
        for activity_id in form.vars:
            if form.vars[activity_id] == 'on':
                db.workshops_products_activities.insert(
                    workshops_activities_id=activity_id,
                    workshops_products_id=wspID)
        session.flash = T('Saved')
        redirect(URL('tickets', vars={'wsID': wsID}))

    menu = get_workshops_menu('tickets', wsID)

    back = os_gui.get_button('back', URL('tickets', vars={'wsID':wsID}))

    return dict(content=content,
                menu=menu,
                save=os_gui.get_submit_button('MainForm'),
                back=back)


# @auth.requires(auth.has_membership(group_id='Admins') or \
#                auth.has_permission('update', 'workshops_products'))
# def ticket_sell():
#     """
#         Page to select which customer to sell a product to
#     """
#     wsID = request.vars['wsID']
#     wspID = request.vars['wspID']
#
#     wsp = WorkshopProduct(wspID)
#
#     response.title = T('Manage')
#     subtitle = SPAN(get_subtitle(wsID), ' > ',
#                     T('Customers'), ': ', wsp.name, ' > ',
#                     T('Sell product'))
#     response.subtitle = subtitle
#
#     # reset session variable for search
#     session.customers_load_list_search_name = None
#
#     _class = 'clear' + ' '
#     _class += 'load_list_customers' + ' '
#
#     back = os_gui.get_button('back_bs',
#                              URL('tickets_list_customers', vars={'wsID': wsID,
#                                                                   'wspID': wspID}),
#                              _class='left',
#                              btn_size='btn')
#
#     search_results = DIV(
#         LOAD('customers', 'load_list.load',
#              content=os_gui.get_ajax_loader(message=T("Loading...")),
#              target='ticket_sell_customers_list',
#              vars={'list_type': 'workshops_product_sell',
#                    'items_per_page': 10,
#                    'wspID': wspID,
#                    'wsID': wsID},
#              ajax=True),
#         _class=_class,
#         _id='ticket_sell_customers_list')
#
#     loader = os_gui.get_ajax_loader(message=T("Searching..."))
#
#     return dict(content=search_results, loader=loader)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'workshops_products_customers'))
def ticket_sell_to_customer():
    """
        Sell a product to a customer
    """
    cuID = request.vars['cuID']
    wsID = request.vars['wsID']
    wspID = request.vars['wspID']

    waitinglist = False
    if 'waiting' in request.vars:
        waitinglist = True
    # register customer

    session.flash = T("Added customer")

    wsp = WorkshopProduct(wspID)
    wsp.sell_to_customer(cuID, waitinglist=waitinglist)

    next_url = ticket_sell_get_return_url(cuID, wsID, wspID)
    if session.events_ticket_sell_back == 'customers':
        redirect(next_url, client_side=True)
    else:
        redirect(next_url)


def ticket_sell_get_return_url(cuID, wsID, wspID, remove=False):
    """
        Return URL for adding a customer to a product or removing a customer
        from a product
    """
    if session.events_ticket_sell_back == 'customers':
        url = URL('customers', 'events', vars={'cuID': cuID})
    else:
        url = URL('tickets_list_customers', vars={'wsID': wsID,
                                                  'wspID': wspID})

    return url


# TODO: move to os_workshops.WorkshopProduct so it can also be used in the shop or os_workshops.WorkshopHelper
def ticket_check_sold_out(wsID, wspID):
    """
        This function checks if a product is sold out
        It's sold out when any of the activities it contains is completely full
    """
    check = False

    fwsID = workshops_get_full_workshop_product_id(wsID)
    if int(wspID) == fwsID:
        # Full workshops check, check if any activity is full
        query = (db.workshops_activities.workshops_id == wsID)
        rows = db(query).select(db.workshops_activities.ALL)
        for row in rows:
            reserved = activity_count_reservations(row.id, fwsID)
            if reserved >= row.Spaces:
                check = True
                break
    else:
        # Product check, check if any a activity is full
        query = (db.workshops_products_activities.workshops_products_id ==
                 wspID)
        rows = db(query).select(db.workshops_products_activities.ALL)
        for row in rows:
            activity = db.workshops_activities(row.workshops_activities_id)
            reserved = activity_count_reservations(activity.id, fwsID)
            if reserved >= activity.Spaces:
                check = True
                break

    return check


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'workshops_products_customers'))
def ticket_delete_customer():
    """
        Remove a customer from the sold products list
    """
    wsID = request.vars['wsID']
    wsp_cuID = request.vars['wsp_cuID']

    # Cancel invoice (if any)
    row = db.invoices_workshops_products_customers(workshops_products_customers_id = wsp_cuID)
    if row:
        iID = row.invoices_id
        if iID:
            invoice = Invoice(iID)
            invoice.set_status('cancelled')

    # get database record for workshop_customers
    row = db.workshops_products_customers(
        db.workshops_products_customers.id == wsp_cuID)

    cuID = row.auth_customer_id
    wspID = row.workshops_products_id

    # remove sold product entry for customer
    query = (db.workshops_products_customers.id == wsp_cuID)
    db(query).delete()

    session.flash = T("Removed")

    next_url = ticket_sell_get_return_url(cuID, wsID, wspID, True)
    redirect(next_url)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'workshops_products_customers'))
def ticket_cancel_customer():
    """
        Change status to cancelled or from cancelled to not cancelled
    """
    wsID = request.vars['wsID']
    wsp_cuID = request.vars['wsp_cuID']

    # Cancel invoice (if any)
    row = db.invoices_workshops_products_customers(workshops_products_customers_id = wsp_cuID)
    if row:
        iID = row.invoices_id
        if iID:
            invoice = Invoice(iID)
            invoice.set_status('cancelled')

    # get database record for workshop_customers
    row = db.workshops_products_customers(
        db.workshops_products_customers.id == wsp_cuID)

    cuID = row.auth_customer_id
    wspID = row.workshops_products_id

    row.Cancelled = not row.Cancelled
    row.update_record()

    # update invoice status to cancelled if status == sent
    query = (db.invoices_workshops_products_customers.workshops_products_customers_id == wsp_cuID)
    rows = db(query).select(db.invoices_workshops_products_customers.ALL)
    if len(rows):
        row = rows.first()
        iID = row.invoices_id

        invoice = Invoice(iID)
        if invoice.invoice.Status == 'sent':
            invoice.set_status('cancelled')

    session.flash = T("Saved")

    next_url = ticket_sell_get_return_url(cuID, wsID, wspID, True)
    redirect(next_url)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'workshops_products_customers'))
def ticket_remove_customer_from_waitinglist():
    """
        Removes a customer from the waitinglist and add customer to list
    """
    wsp_cuID = request.vars['wsp_cuID']
    wsID = request.vars['wsID']

    row = db.workshops_products_customers(wsp_cuID)
    wspID = row.workshops_products_id
    cuID = row.auth_customer_id

    query = (db.workshops_products_customers.id == wsp_cuID)
    db(query).delete()

    wsp = WorkshopProduct(wspID)
    wsp.sell_to_customer(cuID)

    redirect(URL('tickets_list_customers', vars={'wsID': wsID,
                                                  'wspID': wspID}))

@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'workshops_products_customers'))
def ticket_resend_info_mail():
    """
        Resend info mail for customer
    """
    wspcID = request.vars['wspcID']
    wspc = db.workshops_products_customers(wspcID)
    cuID = wspc.auth_customer_id
    customer = Customer(cuID)
    ##
    # Send mail
    ##
    osmail = OsMail()
    msgID = osmail.render_email_template('workshops_info_mail', workshops_products_customers_id=wspcID)
    sent = osmail.send(msgID, cuID)

    ##
    # Check the "Event info" checkbox
    ##
    if sent:
        wspc.WorkshopInfo = True
        wspc.update_record()
        msg = T('Sent event info mail to ')
    else:
        msg = T('Unable to send event info mail to ')

    ##
    # Notify user
    ##
    session.flash = msg + customer.row.display_name

    wsp = db.workshops_products(wspc.workshops_products_id)


    if session.workshops_product_resend_info_mail == 'customers_events':
        redirect(URL('customers', 'events', vars={'cuID':cuID}))
    else:
        redirect(URL('events', 'tickets_list_customers', vars={'wsID':wsp.workshops_id,
                                                                   'wspID':wsp.id}))


def get_fullws_label(value=None):
    """
        Returns label for full Workshop
    """
    return get_label('primary', T("Full event"))


def get_cancelled_label(value=None):
    """
        Returns label for cancelled product
    """
    return SPAN(T("Cancelled"), _class='label label-warning')


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'class_status'))
def overlapping_classes():
    """
        Shows a page of overlapping classes for a given workshop
    """
    wsID = request.vars['wsID']
    workshop = db.workshops(wsID)
    response.title = T('Overlapping classes')
    try:
        response.subtitle = workshop.Name + ' ' + \
                            workshop.Startdate.strftime(DATE_FORMAT)
    except AttributeError:
        response.subtitle = workshop.Name

    session.classes_status_set_cancelled_back = 'workshops_oc'

    return_url = URL('activities', vars={'wsID': wsID})

    content = []
    query = (db.workshops_activities.workshops_id == wsID)
    rows = db(query).select(db.workshops_activities.ALL,
                            orderby=db.workshops_activities.Activitydate)
    for row in rows:
        content.append(activity_get_overlapping_classes(row.id))

    cancel_oc = A(T('Cancel all overlapping classes'),
                  _href=URL('overlapping_classes_cancel_all',
                            vars={'wsID': wsID}),
                  _class='btn btn-default btn-sm right')
    back = os_gui.get_button('back', return_url)
    back = DIV(back, cancel_oc, _class='pull-right')

    return dict(content=content,
                back=back)


def activity_get_overlapping_classes(wsaID):
    """
        Check for overlapping classes
    """
    activity = db.workshops_activities(wsaID)

    loc_check = not activity.school_locations_id is None or \
                not activity.school_locations_id == ''
    if loc_check:
        classes = TABLE(_class='table table-hover')
        classes.append(THEAD(TR(TH(),
                                TH(T('Location')),
                                TH(T('Class type')),
                                TH(T('Class date')),
                                TH(T('Start')),
                                TH(T('End')),
                                TH(T('Teacher')),
                                TH(),
                                _class='os-table_header')))
        activity_weekday = get_weekday(activity.Activitydate)

        query = overlapping_classes_get_query(activity_weekday,
                                              activity.school_locations_id,
                                              activity.Activitydate,
                                              activity.Starttime,
                                              activity.Endtime)

        rows = db(query).select(db.classes.ALL)
        for row in rows.render():
            clsID = row.id
            wsID = activity.workshops_id
            result = classes_get_status(row.id, activity.Activitydate)
            status = result['status_marker']
            buttons = activity_get_overlapping_classes_buttons(
                wsID,
                clsID,
                activity.Activitydate,
                result['status'])

            location = max_string_length(row.school_locations_id, 15)
            classtype = max_string_length(row.school_classtypes_id, 24)
            start = row.Starttime
            end = row.Endtime
            teachers = class_get_teachers(clsID,
                                          activity.Activitydate)

            teacher = teachers['teacher']

            tr = TR(TD(status, _class='td_status_marker'),
                    TD(location, _class='location'),
                    TD(classtype, _class='classtype'),
                    TD(activity.Activitydate, _class='class_date'),
                    TD(start, _class='class_time'),
                    TD(end, _class='class_time'),
                    TD(teacher, _class='class_teacher'),
                    TD(buttons, _class='show_buttons'))
            classes.append(tr)

        if len(classes) == 1:  # only header, so no content
            classes = ''

    else:
        classes = T("No overlapping classes found.")

    return dict(activity=activity,
                classes=classes)


def overlapping_classes_set_status_cancelled_execute(clsID, date):
    """
        Actually cancels the class
    """
    cotc = db.classes_otc(classes_id=clsID, ClassDate=date)
    if not cotc:
        db.classes_otc.insert(
            classes_id=clsID,
            ClassDate=date,
            Status='cancelled'
        )
    else:
        cotc.Status = 'cancelled'
        cotc.update_record()


def overlapping_classes_set_status_cancelled():
    """
        Cancels an overlapping class
    """
    wsID = request.vars['wsID']
    clsID = request.vars['clsID']
    date_formatted = request.vars['date']
    date = datestr_to_python(DATE_FORMAT, date_formatted)

    overlapping_classes_set_status_cancelled_execute(clsID, date)

    redirect(URL('overlapping_classes', vars={'wsID': wsID}))


def overlapping_classes_set_status_normal():
    """
        Sets status to normal
    """
    wsID = request.vars['wsID']
    clsID = request.vars['clsID']
    date_formatted = request.vars['date']
    date = datestr_to_python(DATE_FORMAT, date_formatted)

    cotc = db.classes_otc(classes_id=clsID, ClassDate=date)
    if cotc:  # No need to do anything when there are no changes
        cotc.Status = 'normal'
        cotc.update_record()

        """
         if the status is normal:
            check if there are any other changes
            otherwise just delete. """
        change = False
        fields = [
            cotc.school_locations_id,
            cotc.school_classtypes_id,
            cotc.Starttime,
            cotc.Endtime,
            cotc.auth_teacher_id,
            cotc.auth_teacher_id2,
            cotc.Description
        ]
        for field in fields:
            if not field is None or field == '':
                change = True

        if not change:
            query = (db.classes_otc.id == cotc.id)
            result = db(query).delete()

    redirect(URL('overlapping_classes', vars={'wsID': wsID}))


def overlapping_classes_cancel_all():
    """
        Function called which sets all overlapping classes to cancelled
    """
    wsID = request.vars['wsID']
    # Get all activities for a workshop
    query = (db.workshops_activities.workshops_id == wsID)
    rows = db(query).select(db.workshops_activities.ALL)
    for row in rows:
        # Get overlapping classes for activity
        date = row.Activitydate
        weekday = get_weekday(date)
        oquery = overlapping_classes_get_query(weekday,
                                               row.school_locations_id,
                                               date,
                                               row.Starttime,
                                               row.Endtime)
        crows = db(oquery).select(db.classes.ALL)
        for crow in crows:
            clsID = crow.id
            overlapping_classes_set_status_cancelled_execute(clsID, date)

    redirect(URL('overlapping_classes', vars={'wsID': wsID}))


def activity_get_overlapping_classes_buttons(wsID, clsID, class_date, status):
    """
        Returns (un)cancel buttons for a class
    """
    btn_group = DIV(_class='btn-group pull-right')
    class_date = class_date.strftime(DATE_FORMAT)
    if status == 'cancelled':
        btn = A(SPAN(_class='glyphicon glyphicon-ok'),
                _href=URL('overlapping_classes_set_status_normal',
                          vars={'clsID': clsID,
                                'date': class_date,
                                'wsID': wsID}),
                _title=T('Set status Normal'),
                _class='btn btn-default btn-sm green')
    else:
        btn = A(SPAN(_class='glyphicon glyphicon-ban-circle'),
                _href=URL('overlapping_classes_set_status_cancelled',
                          vars={'clsID': clsID,
                                'date': class_date,
                                'wsID': wsID}),
                _title=T('Set status Cancelled'),
                _class='btn btn-default btn-sm red')

    btn_group.append(btn)

    return btn_group


def overlapping_classes_get_query(activity_weekday,
                                  activity_locations_id,
                                  activity_date,
                                  starttime,
                                  endtime):
    """
        Returns query to check for overlapping classes
    """
    return (db.classes.school_locations_id == activity_locations_id) & \
           (db.classes.Startdate <= activity_date) & \
           ((db.classes.Enddate >= activity_date) |
            (db.classes.Enddate == None)) & \
           (db.classes.Week_day == activity_weekday) & \
           (((db.classes.Starttime <= starttime) &
             (db.classes.Endtime <= endtime) &
             (db.classes.Endtime >= endtime)) |
            ((db.classes.Starttime >= starttime) &
             (db.classes.Starttime <= endtime) &
             (db.classes.Endtime <= endtime) &
             (db.classes.Endtime >= starttime)) |
            ((db.classes.Starttime >= starttime) &
             (db.classes.Starttime <= endtime) &
             (db.classes.Endtime >= endtime)) |
            ((db.classes.Starttime <= starttime) &
             (db.classes.Endtime >= endtime))
            )


def overlapping_classes_get_count(wsaID):
    """
        Returns count for overlapping classes
    """
    activity = db.workshops_activities(wsaID)
    activity_weekday = get_weekday(activity.Activitydate)
    query = overlapping_classes_get_query(activity_weekday,
                                          activity.school_locations_id,
                                          activity.Activitydate,
                                          activity.Starttime,
                                          activity.Endtime)
    return db(query).count()


def overlapping_classes_get_count_all(wsID):
    """
        Returns count of all overlapping classes for a workshop
    """
    count = 0
    query = (db.workshops_activities.workshops_id == wsID)
    rows = db(query).select(db.workshops_activities.ALL)
    for row in rows:
        count += overlapping_classes_get_count(row.id)

    return count


# @auth.requires(auth.has_membership(group_id='Admins') or \
#                auth.has_permission('update', 'workshops_mail_customers'))
# def mail_customers():
#     """
#         Send an email to all customers for a workshop
#         Checks db.workshops_products_customers to see who should get the mail
#     """
#     # Fire up TinyMCE
#     response.js = 'tinymce_init_default();'
#
#     wsID = request.vars['wsID']
#
#     form = mail_customers_get_form()
#
#     form_msg = ''
#     form_info = ''
#     if form.process().accepted:
#         mail_customers_send(wsID,
#                             form.vars.subject,
#                             '<html><body>' + XML(form.vars.message) + '</body></html>')
#
#         response.flash = T("Message sent")
#         redirect(URL('messages', vars={'wsID': wsID}, extension=False),
#                  client_side=True)
#
#     elif form.errors:
#         msg = SPAN(B(T('Oh snap!')), ' ',
#                    T("Change a few things up and try sending again..."))
#         form_msg = os_gui.get_alert('danger', msg)
#
#     response.flash = None
#
#     content = DIV(
#         form.custom.begin,
#         form.custom.widget.subject,
#         form.custom.widget.message,
#         form.custom.submit,
#         form.custom.end
#     )
#
#     return dict(content=content)


# def mail_customers_send(wsID, subject, message, msgID=None):
#     """
#         send mail to customers of a workshop
#     """
#     # Save message to database
#     if msgID is None:
#         msgID = db.messages.insert(msg_subject=subject,
#                                    msg_content=message)
#         db.workshops_messages.insert(workshops_id=wsID,
#                                      messages_id=msgID)
#
#     # To send, first get list of all customers with email for a workshop
#     wh = WorkshopsHelper()
#     customers_rows = wh.get_all_workshop_customers(wsID)
#
#     osmail = OsMail()
#     for row in customers_rows:
#         osmail.send(msgID, row.auth_user.id)
#
#     # Get customers for those products
#
#
# def mail_customers_get_form(value=None):
#     """
#         Returns a form to mail customers
#     """
#     signature = '<br><br>'
#     if db.sys_properties(Property='smtp_signature'):
#         signature += db.sys_properties(Property='smtp_signature').PropertyValue
#
#     form = SQLFORM.factory(
#         Field('subject',
#               requires=IS_NOT_EMPTY()),
#         Field('message', 'text',
#               default=signature,
#               requires=IS_NOT_EMPTY()),
#         submit_button=T("Send"))
#
#     confirm_send_msg = T("Are you sure you want to send this message?")
#     submit_onclick = "return confirm('" + confirm_send_msg + "');"
#
#     form.element('#no_table_subject').attributes['_placeholder'] = T("Subject...")
#     form.element('#no_table_message').attributes['_placeholder'] = T("Message...")
#     form.element('input[type=submit]').attributes['_onclick'] = submit_onclick
#
#     # add class for BS3
#     form.custom.widget.subject['_class'] += ' form-control'
#     # form.custom.widget.message['_class'] += ' form-control'
#     form.custom.widget.message['_class'] += ' form-control tmced'
#     form.custom.submit['_class'] = ' btn-primary'
#
#     return form
#
#
# @auth.requires(auth.has_membership(group_id='Admins') or \
#                auth.has_permission('read', 'workshops_messages'))
# def messages():
#     """
#         List messages sent to customers for a workshop.
#     """
#     wsID = request.vars['wsID']
#     response.title = T('Event')
#     response.subtitle = get_subtitle(wsID)
#     session.workshops_msgID = None
#     # response.view = 'events/manage.html'
#
#     ## Modals container
#     modals = DIV()
#
#     ## Mail button & modal begin ##
#     result = messages_get_mail(wsID)
#     btn_mail = result['button']
#     modals.append(result['modal'])
#
#     ## Mail button & modal end ##
#
#     ## Message list begin ##
#     content_msg = DIV(_class='container_messages')
#
#     h = html2text.HTML2Text()
#     h.ignore_links = True
#     h.images_to_alt = True
#
#     messages = UL(_class='ul_liststyle_none list_messages col-md-3')
#     left = [db.messages.on(db.workshops_messages.messages_id == db.messages.id)]
#     query = (db.workshops_messages.workshops_id == wsID)
#     rows = db(query).select(db.messages.ALL,
#                             db.workshops_messages.ALL,
#                             left=left,
#                             orderby=~db.workshops_messages.Created_at)
#     for i, row in enumerate(rows):
#         li_class = 'os-clickable'
#         if i == 0:
#             session.workshops_msgID = row.messages.id
#             li_class += ' active'
#
#         created_at = row.workshops_messages.Created_at
#         created_at = created_at.strftime(DATE_FORMAT)
#         msg_preview = h.handle(row.messages.msg_content)
#         msg_preview = msg_preview.replace('*', '')
#         msg_preview = msg_preview.replace('_', '')
#         msg_preview = SPAN(XML(msg_preview[0:60]),
#                            _class='vsmall_font grey')
#
#         subject = max_string_length(row.messages.msg_subject, 25)
#
#         messages.append(LI(B(subject), ' ',
#                            SPAN(created_at,
#                                 _class='right vsmall_font grey'),
#                            BR(),
#                            msg_preview,
#                            _class=li_class,
#                            _id=row.messages.id))
#
#     content_msg.append(messages)
#
#     ## Message list end ##
#
#     ## Message display begin ##
#     message_display = LOAD('messages', 'message.load',
#                            ajax=True,
#                            target='message_display',
#                            content=os_gui.get_ajax_loader(),
#                            vars={'category': 'workshops',
#                                  'wsID': wsID})
#
#     content_msg.append(DIV(message_display,
#                            _class='col-md-9 message_display'))
#
#     ## Message display end ##
#
#     content = DIV(DIV(content_msg, modals, _class='col-md-12'),
#                   _class='row')
#
#     menu = get_workshops_menu(request.function, wsID)
#     back = manage_get_back()
#
#     return dict(back=back,
#                 menu=menu,
#                 content=content,
#                 btn_mail=DIV(btn_mail, _class='pull-right'),
#                 left_sidebar_enabled=True)
#
#
# def messages_get_mail(wsID):
#     """
#         Returns mail button and modal for all workshop customers
#     """
#     workshop = db.workshops(wsID)
#     modal_title = T("Mail all customers: ") + workshop.Name
#     modal_content = LOAD('workshops', 'mail_customers.load',
#                          ajax=True,
#                          vars={'wsID': wsID})
#
#     btn_icon = SPAN(SPAN(_class='glyphicon glyphicon-envelope'), ' ',
#                     T('Send message'))
#
#     result = os_gui.get_modal(button_text=XML(btn_icon),
#                               modal_title=modal_title,
#                               modal_content=modal_content,
#                               modal_class='mail_' + unicode(wsID),
#                               modal_size='lg',
#                               button_class='btn-sm')
#
#     return result
#
#
# @auth.requires(auth.has_membership(group_id='Admins') or \
#                auth.has_permission('read', 'workshops_messages'))
# def messages_set_id():
#     """
#         Set session ID used to show message in main window of messages pane
#         This function should be called as JSON
#     """
#     if request.extension == 'json':
#         session.workshops_msgID = request.vars['msgID']
#         status = 'success'
#         message = T('OK')
#     else:
#         status = 'fail'
#         message = T('Please call this function as "json"')
#
#     return dict(status=status,
#                 message=message)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'tasks'))
def tasks():
    """
        Display list of tasks for a workshop
    """
    wsID = request.vars['wsID']
    # now continue settings things as usual
    workshop = db.workshops(wsID)
    response.title = T('Event')
    response.subtitle = get_subtitle(wsID)
    response.view = 'general/tabs_menu.html'

    tasks = DIV(LOAD('tasks', 'list_tasks.load',
                     vars=request.vars,
                     content=os_gui.get_ajax_loader()))

    content = tasks

    # Add permission
    add = ''
    permission = auth.has_membership(group_id='Admins') or \
                 auth.has_permission('create', 'tasks')
    if permission:
        # add = os_gui.get_button('add', url_add)
        tasks = Tasks()
        add = tasks.add_get_modal({'wsID': wsID})

    back = manage_get_back()
    menu = get_workshops_menu(request.function, wsID)

    return dict(content=content,
                menu=menu,
                back=back,
                add=add)


# No decorator here, permissions are checked inside the function
def pdf():
    """
        Converts a invoice to PDF
    """
    import weasyprint

    wsID = request.vars['wsID']
    workshop = Workshop(wsID)

    permission = (auth.has_membership(group_id='Admins') or
                  auth.has_permission('read', 'workshops'))

    if not permission:
        return T("Not authorized")

    html = pdf_template(wsID)

    fname = workshop.Startdate.strftime(DATE_FORMAT) + u'_' + workshop.Name + u'.pdf'
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-disposition'] = 'attachment; filename=' + fname
    # return pyfpdf_from_html(html)

    stream = cStringIO.StringIO()
    workshop = weasyprint.HTML(string=html).write_pdf(stream)

    return stream.getvalue()


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'workshops'))
def pdf_template_show():
    """
        Preview PDF template
    """
    return pdf_template(request.vars['wsID'])


def pdf_template_qrcode(wsID):
    """
    :param wsID: db.workshops.id
    :return: QRcode as div
    """
    import pyqrcode

    url = URL('shop', 'event', vars={'wsID':wsID},
              host=True,
              scheme=True)
    qr = DIV(_class="qrcode")
    row = DIV()
    numbers = pyqrcode.create(url).text()
    for i in numbers + '\n':
        if i == '\n':
            qr.append(row)
            row = DIV()

        if i == "0":
            row.append(SPAN(' '))
        elif i == '1':
            row.append(SPAN(' ', _class="qrdata"))

    return qr


def pdf_template(wsID):
    """
        Print friendly display of a Workshop
    """
    template = get_sys_property('branding_default_template_events') or 'events/default.html'
    template_file = 'templates/' + template

    workshop = Workshop(wsID)
    activities = workshop.get_activities()
    products = workshop.get_products()

    if len(activities) == 0:
        session.flash = T('No activities found, unble to export to PDF')
        redirect(URL('index'))
    elif len(products) == 0:
        session.fash = T('No products found, unable to export to PDF')
        redirect(URL('index'))

    price = format(products[0].Price or 0, '.2f')

    workshop_image_url = URL('default', 'download', args=workshop.picture, host=True, scheme=True)
    shop_url = URL('shop', 'event', vars={'wsID': wsID}, host=True, scheme=True)
    qr = pdf_template_qrcode(wsID)

    html = response.render(template_file,
                           dict(workshop=workshop,
                                workshop_image_url=workshop_image_url,
                                dates=pdf_template_get_display_dates(workshop, activities),
                                times=pdf_template_get_display_times(workshop, activities),
                                activities=activities,
                                products=products,
                                price=price,
                                qr=qr,
                                logo=pdf_template_get_logo()))

    return html


def pdf_template_get_logo(var=None):
    """
        Returns logo for pdf template
    """
    branding_logo = os.path.join(request.folder,
                                 'static',
                                 'plugin_os-branding',
                                 'logos',
                                 'branding_logo_invoices.png')
    if os.path.isfile(branding_logo):
        abs_url = '%s://%s/%s/%s' % (request.env.wsgi_url_scheme,
                                     request.env.http_host,
                                     'static',
                                     'plugin_os-branding/logos/branding_logo_invoices.png')
        logo_img = IMG(_src=abs_url)

    else:
        logo_img = ''

    return logo_img


def pdf_template_get_display_dates(workshop, activities):
    """
        :param workshop: Workshop object
        :param activities: workshop activities rows
        :return: formatted date for workshop
    """
    date_until = ''
    if len(activities) > 0:
        date_from = activities[0].Activitydate

    if len(activities) > 1:
        date_until = activities[len(activities) - 1].Activitydate

    if len(activities) == 0:  # no activities
        date_from = T("No activities found...")
        date_until = T("No activities found...")

    return dict(date_from=date_from,
                date_until=date_until)


def pdf_template_get_display_times(workshop, activities):
    """
        :param workshop: Workshop object
        :param activities: workshop activities rows
        :return: formatted date for workshop
    """
    time_until = ''
    if len(activities) > 0:
        time_from = activities[0].Starttime

    if len(activities) > 1:
        time_until = activities[len(activities) - 1].Endtime

    if len(activities) == 0:  # no activities
        date_from = T("No activities found...")
        date_until = T("No activities found...")

    return dict(time_from=time_from,
                time_until=time_until)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'workshops_mail'))
def info_mail():
    """
        Information mail for workshops
    """
    wsID = request.vars['wsID']
    workshop = db.workshops(wsID)
    response.title = T('Event')
    response.subtitle = get_subtitle(wsID)
    response.view = 'general/tabs_menu.html'

    ###
    # Get ID
    ###
    row = db.workshops_mail(workshops_id = wsID)
    if not row:
        # create record
        wsmID = db.workshops_mail.insert(
            workshops_id = wsID,
            MailContent = None
        )
    else:
        # we have an id
        wsmID = row.id

    crud.messages.submit_button = T("Save")
    crud.messages.record_updated = T("Saved")
    crud.settings.formstyle = 'bootstrap3_stacked'
    crud.settings.update_next = URL('info_mail', vars={'wsID':wsID})
    form = crud.update(db.workshops_mail, wsmID)

    result = set_form_id_and_get_submit_button(form, 'MainForm')
    form = result['form']
    submit = result['submit']

    textareas = form.elements('textarea')
    for textarea in textareas:
        textarea['_class'] += ' tmced'

    content = form

    # preview = os_gui.get_button('notext',
    #                             URL('events', 'info_mail_preview', vars={'wsmID':wsmID}),
    #                             title=T('Preview'),
    #                             _class='pull-right',
    #                             btn_size='',
    #                             _target='_blank')

    menu = get_workshops_menu(request.function, wsID)
    back = manage_get_back()

    return dict(content=content,
                menu=menu,
                back=back,
                tools='',
                save=submit)