# -*- coding: utf-8 -*-
# plugin (local modules) import

from general_helpers import datestr_to_python
from general_helpers import highlight_submenu
from general_helpers import get_badge
from general_helpers import get_label
from general_helpers import get_last_day_month
from general_helpers import get_months_list
from general_helpers import class_get_teachers
from general_helpers import get_paused_subscriptions
from general_helpers import max_string_length
from general_helpers import Memo_links
from general_helpers import set_form_id_and_get_submit_button
#from general_helpers import workshops_get_full_workshop_product_id

from os_storage import uploads_available_space

from openstudio.os_class_attendance import ClassAttendance
from openstudio.os_customers import Customers
from openstudio.os_customer import Customer
from openstudio.os_customer_classcard import CustomerClasscard
from openstudio.os_classcards_helper import ClasscardsHelper
from openstudio.os_school_classcard import SchoolClasscard
from openstudio.os_school_subscription import SchoolSubscription
from openstudio.os_invoices import Invoices
from openstudio.os_workshops_helper import WorkshopsHelper

from openstudio.os_customers_subscriptions_credits import CustomersSubscriptionsCredits

# python general modules import
import cStringIO
import os
import openpyxl
import calendar
import codecs

# helper functions

def _edit_check_picture(form):
    if form.vars['Picture'] == '':
        row = db.auth_user[form.vars['id']]
        row.picture = None
        row.thumbsmall = None
        row.thumblarge = None
        row.update_record()


def classes_check_reservation(row):
    """
        Check if a customer is already reserved for a class.
        If not, return an add button to use in the grid.
    """
    customers_id = request.vars['cuID']
    clsID = row.id
    check = db.classes_reservation(customers_id=customers_id, classes_id=clsID)
    if check is None:
        return os_gui.get_button('add', URL("classes", "reserve_class",
                                     vars={'cuID':customers_id,
                                           'clsID':clsID}))
                                           #'back'="customers_classes")
    else:
        return ""


#TODO: Rename & reuse for subscriptions_alt_prices repeat
@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('update', 'alternativepayments'))
def alternativepayment_repeat():
    apID = request.vars['apID']
    row = db.alternativepayments[apID]

    year = row.PaymentYear
    month = row.PaymentMonth

    if month == 12:
        month = 1
        year += 1
    else:
        month += 1
    db.alternativepayments.insert(auth_customer_id=row.auth_customer_id,
                                  PaymentYear=year,
                                  PaymentMonth=month,
                                  payment_categories_id=row.payment_categories_id,
        Amount=row.Amount, Description=row.Description)

    session.customers_payments_tab = '#ap'

    redirect(URL("payments",vars={'cuID':row.auth_customer_id}))


def index_get_export(val=None):
    """
        Returns dict with export button and bs3 modal containing the links
        to different export options.
    """
    export = ''
    if auth.has_membership(group_id='Admins') or auth.has_permission('update', 'auth_user'):
        mailinglist = A((os_gui.get_fa_icon('fa-envelope-o'),
                         T("Mailing list")),
                        _href=URL('export_excel', vars=dict(export='mailing_list')),
                        _class='textalign_left')
        active_customers = A((os_gui.get_fa_icon('fa-check'),
                             T("Active customers")),
                             _href=URL('export_excel',
                                       vars=dict(export='customers_list')),
                             _class='textalign_left')

        links = [ mailinglist, active_customers ]

        export = os_gui.get_dropdown_menu(
                links = links,
                btn_text = '',
                btn_icon = 'download',
                btn_size = 'btn-sm',
                menu_class='pull-right' )

    return export


def _edit_clear_old_thumbs(form):
    """
        This function cleans up generated thumbnails of the old picture when a new picture is uploaded.
        It does nothing when no new picture is uploaded.
    """
    customers_id = session.customers_edit_clear_thumbs_customers_id
    row = db.customers[customers_id]

    if not form.vars['Picture'] is None:
        if not row.ThumbSmall is None:
            os.remove(request.folder + 'uploads/' + row.ThumbSmall)
        if not row.ThumbLarge is None:
            os.remove(request.folder + 'uploads/' + row.ThumbLarge)


def edit_remodel_form(form,
                      picture,
                      change_picture,
                      label_id,
                      customers_id,
                      bo_button='',
                      te_button='',
                      mail_button='',
                      merged=False,
                      row=None):
    """
        This function takes the default form and makes a nice looking desktop browser friendly form out of it.
        row = db.auth_user(customers_id)
    """

    # check for contact permissions
    contact_permission = auth.has_membership(group_id='Admins') or \
                         auth.has_permission('update', 'customers_contact')

    email_widget = DIV(
            DIV(LABEL(form.custom.label.email),
                DIV(form.custom.widget.email,
                    DIV(mail_button, _class='input-group-btn'),
                    _class='input-group'),
                _class="col-md-12"
            ),
            _class="row"
        )
    if not contact_permission:
        # hide contact info
        mail_button = ''
        form.custom.label.phone = ''
        form.custom.widget.phone = ''
        form.custom.label.mobile = ''
        form.custom.widget.mobile = ''
        form.custom.label.email = ''
        form.custom.widget.email = ''
        form.custom.label.emergency = ''
        form.custom.widget.emergency = ''
        email_widget = ''

    div_picture = DIV(
        H3(T("Picture")),
        picture,
        change_picture, ' ',
        SPAN(label_id, ' ', customers_id,
           _id='customers_id'),
        _class='col-md-6'
    )

    # basic info
    div_basic_info = DIV(
        H3(T("Contact")),
        DIV(
            DIV(LABEL(form.custom.label.first_name),
                form.custom.widget.first_name,
                _class="col-md-6"
            ),
            DIV(LABEL(form.custom.label.last_name),
                form.custom.widget.last_name,
                _class="col-md-6"
            ),
            _class="row"
        ),
        DIV(
            DIV(LABEL(form.custom.label.date_of_birth),
                form.custom.widget.date_of_birth,
                _class="col-md-6"
            ),
            DIV(LABEL(form.custom.label.gender),
                form.custom.widget.gender,
                _class="col-md-6"
            ),
            _class="row"
        ),
        email_widget,
        DIV(
            DIV(LABEL(form.custom.label.phone),
                form.custom.widget.phone,
                _class="col-md-6"
                ),
            DIV(LABEL(form.custom.label.mobile),
                form.custom.widget.mobile,
                _class="col-md-6"
                ),
            _class="row"
        ),
        DIV(
            DIV(LABEL(form.custom.label.school_languages_id),
                form.custom.widget.school_languages_id,
                _class="col-md-12"
                ),
            DIV(LABEL(form.custom.label.emergency),
                form.custom.widget.emergency,
                _class="col-md-12"
                ),
            _class="row"
        ),
        _class='col-md-6 customers_edit_basic_info'
    )


    # check if we have to separate customers by location
    if session.show_location:
        location_label = LABEL(form.custom.label.school_locations_id)
        location_widget = form.custom.widget.school_locations_id
    else:
        location_label = ''
        location_widget = ''

    # address info
    div_address = ''
    business_info = ''
    address_permission = auth.has_membership(group_id='Admins') or \
                         auth.has_permission('update', 'customers_address')

    if address_permission:
        div_address = DIV(
                DIV(H3(T('Address')), _class='col-md-12'),
                DIV(DIV(LABEL(form.custom.label.address),
                        form.custom.widget.address,
                        _class='col-md-3'),
                    DIV(LABEL(form.custom.label.postcode),
                        form.custom.widget.postcode,
                        _class='col-md-3'),
                    DIV(LABEL(form.custom.label.city),
                        form.custom.widget.city,
                        _class='col-md-3'),
                    DIV(LABEL(form.custom.label.country),
                        form.custom.widget.country,
                        _class='col-md-3'),

                ),
            _class='col-md-12 customers_edit_address_info'
        )

        business_info = DIV(
            H3(form.custom.widget.business, ' ', T("Business"), _class='col-md-12'),
            DIV(DIV(LABEL(T("Company name")),
                    form.custom.widget.company,
                    _class='col-md-12'),
                DIV(LABEL(T("Company registration ID")),
                    form.custom.widget.company_registration,
                    _class='col-md-6'),
                DIV(LABEL(T("Company tax registration ID")),
                    form.custom.widget.company_tax_registration,
                    _class='col-md-6'),
                ),
            _class='col-md-6 no-padding-left no-padding-right'
        )

    # business info
    div_studio_and_business = DIV(
        DIV(
            H3(T("Studio"), _class='col-md-12'),
            DIV(DIV(LABEL(form.custom.label.school_levels_id),
                    form.custom.widget.school_levels_id,
                    _class='col-md-6'),
                DIV(LABEL(form.custom.label.school_discovery_id),
                    form.custom.widget.school_discovery_id,
                    _class='col-md-6'),
                DIV(LABEL(form.custom.label.keynr),
                    form.custom.widget.keynr,
                    _class='col-md-6'),
                DIV(LABEL(location_label),
                    location_widget,
                    _class='col-md-6'),
            ),
            _class='col-md-6 no-padding-left no-padding-right'),
        business_info,
        _class='col-md-12'
    )


    # notes
    bo_label = ''
    bo_note = ''
    bo_notes = ''
    bo_all_notes = ''
    te_label = ''
    te_note = ''
    te_notes = ''
    te_all_notes = ''
    if not customers_id is None and not customers_id == '':
        if auth.has_membership(group_id='Admins') or \
           auth.has_permission('read', 'customers_notes_backoffice'):
            bo_label = LABEL(T("Back office"))
            bo_note = DIV(LOAD('customers', 'note_latest.load', ajax=True,
                                        target='os-bonote_latest',
                                        vars={'cuID':customers_id,
                                              'note_type':'backoffice',
                                              'latest':True,
                                              'latest_length':140}),
                                  _class='os-customers_note_latest')

            bo_notes = DIV(
                DIV(LABEL(bo_label),
                    bo_note, bo_button,
                    _class='col-md-12 no-padding-left no-padding-right'),
                _class='col-md-6'
            )

        if auth.has_membership(group_id='Admins') or \
           auth.has_permission('read', 'customers_notes_teachers'):
            te_label = LABEL(T("Teachers"))
            te_note = DIV(LOAD('customers', 'note_latest.load', ajax=True,
                                     target='os-tenote_latest',
                                   vars={'cuID':customers_id,
                                         'note_type':'teachers',
                                         'latest':True,
                                         'latest_length':140}),
                                _class='os-customers_note_latest')

            te_notes = DIV(
                DIV(LABEL(te_label),
                    te_note, te_button,
                    _class='col-md-12 no-padding-left no-padding-right'),
                _class='col-md-6'
            )

    # address info
    # check for address permissions
    address_permission = auth.has_membership(group_id='Admins') or \
                         auth.has_permission('update', 'customers_address')

    address_header = DIV(H3(T('Address')), _class='col-md-12')
    if not address_permission:
        # hide contact info
        address_header = ''
        form.custom.label.address = ''
        form.custom.widget.address = ''
        form.custom.label.postcode = ''
        form.custom.widget.postcode = ''
        form.custom.label.city = ''
        form.custom.widget.city = ''
        form.custom.label.country = ''
        form.custom.widget.country = ''

    div_address = DIV(
            address_header,
            DIV(DIV(LABEL(form.custom.label.address),
                    form.custom.widget.address,
                    _class='col-md-3'),
                DIV(LABEL(form.custom.label.postcode),
                    form.custom.widget.postcode,
                    _class='col-md-3'),
                DIV(LABEL(form.custom.label.city),
                    form.custom.widget.city,
                    _class='col-md-3'),
                DIV(LABEL(form.custom.label.country),
                    form.custom.widget.country,
                    _class='col-md-3'),

            ),
        _class='col-md-12 customers_edit_address_info')

    # business info
    div_business = DIV(
        DIV(
            H3(T("Studio"), _class='col-md-12'),
            DIV(DIV(LABEL(form.custom.label.school_levels_id),
                    form.custom.widget.school_levels_id,
                    _class='col-md-6'),
                DIV(LABEL(form.custom.label.school_discovery_id),
                    form.custom.widget.school_discovery_id,
                    _class='col-md-6'),
                DIV(LABEL(form.custom.label.keynr),
                    form.custom.widget.keynr,
                    _class='col-md-6'),
                DIV(LABEL(form.custom.label.barcode_id),
                    form.custom.widget.barcode_id,
                    _class='col-md-6'),
                DIV(LABEL(location_label),
                    location_widget,
                    _class='col-md-12'),
            ),
            _class='col-md-6 no-padding-left no-padding-right'),
        DIV(
            H3(form.custom.widget.business, ' ', T("Business"), _class='col-md-12'),
            DIV(DIV(LABEL(T("Company name")),
                    form.custom.widget.company,
                    _class='col-md-12'),
                DIV(LABEL(T("Company registration ID")),
                    form.custom.widget.company_registration,
                    _class='col-md-6'),
                DIV(LABEL(T("Company tax registration ID")),
                    form.custom.widget.company_tax_registration,
                    _class='col-md-6'),
                ),
            _class='col-md-6 no-padding-left no-padding-right'),
        _class='col-md-12'
    )

    notes = DIV(
        DIV(H3(T('Notes')), _class='col-md-12'),
        bo_notes,
        te_notes,
        _class='col-md-12 customers_edit_notes')

    return DIV(XML('<form id="MainForm" action="#" enctype="multipart/form-data" method="post">'),
               DIV(div_picture,
                   div_basic_info,
                   _class='col-md-12'),
               div_address,
               div_studio_and_business,
               notes,
               form.custom.end,
               _class='customers_edit_container row')


def subscriptions_get_link_membership_check(row):
    """
    :param row: gluon.dal.row object of db.customers_subscriptions
    :return: Warning if membership required for subscription but not found
    """
    ssu = SchoolSubscription(row.school_subscriptions_id, set_db_info=True)
    required_membership = ssu.school_memberships_id

    customer = Customer(row.auth_customer_id)

    memberships = customer.get_memberships_on_date(TODAY_LOCAL)
    ids = []
    for row in memberships:
        ids.append(row.id)

    if required_membership and not required_membership in ids:
        return os_gui.get_label(
            'warning',
            T('No membership'),
            title=T("This customer doesn't have the required membership for this subscription.")
        )
    else:
        return ''


def subscriptions_get_link_latest_pauses(row):
    """
        Returns latest pauses for a subscription
    """
    csID = row.id
    cuID = row.auth_customer_id
    query = (db.customers_subscriptions_paused.customers_subscriptions_id == row.id)
    rows = db(query).select(db.customers_subscriptions_paused.ALL,
                            orderby=~db.customers_subscriptions_paused.Startdate,
                            limitby=(0,3))

    pause_list = DIV()
    for row in rows:
        item = SPAN(row.Startdate, ' - ', row.Enddate, ' ',
                    _title=row.Description,
                    _class='grey small_font')
        pause_list.append(item)
        pause_list.append(BR())

    pause_list.append(A(SPAN(T("Edit pauses"), _class='small_font'),
                        _href=URL('subscription_pauses', vars={'cuID':cuID,
                                                               'csID':csID}),
                        _title=T("View all pauses and add new")))

    return pause_list


def subscriptions_get_link_latest_blocks(row):
    """
        Returns latest pauses for a subscription
    """
    csID = row.id
    cuID = row.auth_customer_id
    query = (db.customers_subscriptions_blocked.customers_subscriptions_id == row.id)
    rows = db(query).select(db.customers_subscriptions_blocked.ALL,
                            orderby=~db.customers_subscriptions_blocked.Startdate,
                            limitby=(0,3))

    blocked_list = DIV()
    for row in rows:
        item = SPAN(row.Startdate,
                    _title=row.Description,
                    _class='grey small_font')
        if row.Enddate:
            item.append(' - ')
            item.append(row.Enddate)
            item.append(' ')
        blocked_list.append(item)
        blocked_list.append(BR())

    blocked_list.append(A(SPAN(T("Edit blocks"), _class='small_font'),
                   _href=URL('subscription_blocks', vars={'cuID':cuID,
                                                          'csID':csID}),
                   _title=T("View all blocks and add new")))

    return blocked_list


def subscriptions_get_link_credits(row):
    """
        Returns total number of credits for a subscription
    """
    cs = CustomerSubscription(row.id)

    credits = cs.get_credits_balance()

    return A(credits,
             _href=URL('subscription_credits', vars={'cuID':row.auth_customer_id,
                                                     'csID':row.id}))


### helpers end ###

@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'auth_user'))
def index():
    """
        List customers
    """
    response.title = T("Customers")
    # make sure we're redirected back to the list from the edit page
    session.customers_back = None
    # Redirect back to edit page after adding
    session.customers_add_back = None

    # deleted filter
    show = 'current'

    if 'show' in request.vars:
        show = request.vars['show']
        session.customers_show = show
    if not session.customers_show:
        session.customers_show = 'current'

    if session.customers_show == 'deleted':
        deleted_class = 'active'
        current_class = ''
        list_type = 'customers_index_deleted'
    else:
        deleted_class = ''
        current_class = 'active'
        list_type = 'customers_index'

    response.search_available = True
    try:
        response.q = session.customers_load_list_search_name.replace('%', '')
    except AttributeError:
        response.q = ''

    if 'nr_items' in request.vars:
        session.customers_index_items_per_page = int(request.vars['nr_items'])

    show_location = index_get_show_location()
    show_email = index_get_show_email()
    search_results = DIV(LOAD('customers', 'load_list.load',
                              target='customers_load_list',
                              content=os_gui.get_ajax_loader(message=T("Searching...")),
                              vars={'list_type':list_type,
                                    'items_per_page':session.customers_index_items_per_page,
                                    'initial_list':True,
                                    'show_location':show_location,
                                    'show_email': show_email,
                                    'show_deleted': session.customers_show},
                              ajax=True),
                         _id="customers_load_list",
                         _class="load_list_customers clear")

    content = DIV(
        index_get_menu(session.customers_show),
        DIV(DIV(search_results,
                _class='tab-pane active'),
            _class='tab-content'),
        _class='nav-tabs-custom')


    export = index_get_export()
    add = index_get_add()
    # tools = index_get_tools()

    return dict(add=add,
                export=export,
                content=content,
                nr_items=index_get_select_nr_items(),
                header_tools='')


def index_get_show_location():
    """
        Should we show customer locations in the list?
    """
    show_location = False
    if session.show_location:
        show_location = 'True'

    return show_location


def index_get_show_email():
    """
        Returns show contact info permissions
    """
    show_email = False
    if ( auth.has_membership(group_id='Admins') or
         auth.has_permission('update', 'customer-contact') ):
        show_email = True

    return show_email


def index_get_add():
    """
        Add button for index page
    """
    add = ''
    if ( auth.has_membership(group_id='Admins') or
         auth.has_permission('create', 'auth_user') ):
        customers = Customers()
        add = customers.get_add()

    return add


def index_get_menu(page):
    """
        Tabs Menu for index page
    """
    active = 'active'
    current_class = ''
    deleted_class = ''

    if page == 'current':
        current_class = active
    elif page == 'deleted':
        deleted_class = active

    tabs = UL(LI(A(T('Current'),
                    _href=URL('index', vars={'show':'current'})),
                  _class=current_class),
              LI(A(T('Deleted'),
                    _href=URL('index', vars={'show':'deleted'})),
                  _class=deleted_class),
               # LI(I(_class='fa fa-users'),
               #    _class='pull-left header'),
              _class='nav nav-tabs pull-right')

    # print session.customers_load_list_search_name
    #
    # if session.customers_load_list_search_name and session.customers_load_list_search_name != '%%':
    #     title = SPAN(os_gui.get_fa_icon('fa-search'), ' ', T("Search results"))
    # else:
    #     title = SPAN(os_gui.get_fa_icon('fa-circle-o'), ' ', T("All customers"))
    # tabs.append(LI(title, _class='header os-header pull-left'))

    return tabs


def index_get_select_nr_items(var=None):
    """
        Returns a form to select number of items to show on a page
    """
    view_set = [10, 15, 25]
    form = SQLFORM.factory(
        Field('nr_items',
              requires=IS_IN_SET(view_set, zero=None),
              default=session.customers_index_items_per_page or 10,
              label=T('Customers per page')),
        formstyle='divs',
    )

    select = form.element('#no_table_nr_items')
    select['_onchange'] = 'this.form.submit();'
    select['_class'] += ' inline-block'

    form = DIV(DIV(form.custom.begin,
                   form.custom.label.nr_items, ' ',
                   form.custom.widget.nr_items,
                   form.custom.end,
                   _class='well pull-right'),
               _id='customers_index_settings',
               _class='collapse')

    return form


# def index_get_tools(var=None):
#     """
#         Returns tools menu for customers list
#     """
#     tools = []
#
#     # teacher holidays
#     permission = auth.has_membership(group_id='Admins') or \
#                  auth.has_permission('read', 'customers_subscriptions_credits')
#
#     if permission:
#         subscription_credits = A(os_gui.get_fa_icon('fa-check-square-o'),
#                                  T("Subscription credits"),
#                                  _href=URL('customers', 'subscription_credits_month'),
#                                  _title=T('List subscription credits for a selected month'))
#         tools.append(subscription_credits)
#
#     # get menu
#     tools = os_gui.get_dropdown_menu(tools,
#                                      '',
#                                      btn_size='btn-sm',
#                                      btn_icon='wrench',
#                                      menu_class='pull-right'
#                                      )
#
#     return tools


# @auth.requires(auth.has_membership(group_id='Admins') or
#                auth.has_permission('read', 'auth_user'))
# def index_deleted():
#     """
#         List deleted customers
#     """
#     from openstudio import Customers
#
#     response.title = T('Customers')
#     response.subtitle = T('Deleted')
#     response.view = 'general/tabs_menu.html'
#
#     customers = Customers()
#     list = customers.list_deleted_formatted()
#
#     return dict(menu=index_get_menu('deleted'),
#                 content=list)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('delete', 'auth_user'))
def restore():
    """
        Restore from trash
    """
    cuID = request.vars['cuID']

    row = db.auth_user(cuID)

    query = (db.auth_user.id == cuID)
    db(query).update(trashed = False)

    session.flash = SPAN(
        T('Moved'), ' ',
        row.display_name, ' ',
        T('to current')
    )
    redirect(URL('index'))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('delete', 'auth_user'))
def trash():
    """
        Delete a customer
    """
    cuID = request.vars['cuID']

    # set enddate for recurring reservations
    query = (db.classes_reservation.auth_customer_id == cuID) & \
            (db.classes_reservation.ResType == 'recurring') & \
            ((db.classes_reservation.Enddate == None) |
             (db.classes_reservation.Enddate > TODAY_LOCAL))
    db(query).update(Enddate=TODAY_LOCAL)
    # remove all waitinglist entries
    query = (db.classes_waitinglist.auth_customer_id == cuID)
    db(query).delete()
    # Cancel all class bookings >= today
    query = (db.classes_attendance.auth_customer_id == cuID) & \
            (db.classes_attendance.ClassDate > TODAY_LOCAL)
    db(query).update(BookingStatus = 'cancelled')

    # Move customer to deleted
    query = (db.auth_user.id == cuID)
    db(query).update(trashed=True)

    session.flash = T('Moved to deleted')

    redirect(URL('index'))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('delete', 'auth_user'))
def delete():
    """
        Delete a customer
    """
    cuID = request.vars['cuID']

    query = (db.auth_user.id == cuID)
    db(query).delete()

    session.flash = T('Deleted')

    redirect(URL('index'))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('create', 'auth_user'))
def add():
    """
        Page to add a new customer, only show the required field and after
        adding redirect to the edit page
    """
    response.view = 'general/only_content.html'
    response.title = T("New account")
    response.subtitle = T("Add a customer")

    # # call js for styling the form
    # response.js = 'set_form_classes();'

    # enable only required fields
    for field in db.auth_user:
        field.readable=False
        field.writable=False

    db.auth_user.first_name.readable=True
    db.auth_user.first_name.writable=True
    db.auth_user.last_name.readable=True
    db.auth_user.last_name.writable=True
    db.auth_user.email.readable=True
    db.auth_user.email.writable=True

    if session.show_location:
        loc_query = (db.school_locations.Archived == False)

        db.auth_user.school_locations_id.readable = True
        db.auth_user.school_locations_id.writable = True

        db.auth_user.school_locations_id.requires = \
            IS_IN_DB(db(loc_query),
                     'school_locations.id',
                     '%(Name)s',
                     zero=T("Please select..."))

    db.auth_user.password.default = generate_password(30)
    db.auth_user.customer.default = True

    if request.vars['teacher'] == 'True':
        response.subtitle = T("Add a teacher")
        db.auth_user.teacher.default = True
        db.auth_user.login_start.default = 'backend'
        crud.settings.create_onaccept = [cache_clear_school_teachers]

    if request.vars['employee'] == 'True':
        response.subtitle = T("Add an employee")
        db.auth_user.employee.default = True
        db.auth_user.login_start.default  = 'backend'

    if 'clsID' in request.vars:
        # if we're comming from classes/attendance_list
        clsID = request.vars['clsID']
        date_formatted  = request.vars['date']
        next_url = '/customers/add_redirect_on_create?cuID=[id]'
        next_url += '&clsID=' + clsID
        next_url += '&date='  + date_formatted
    else:
        next_url = '/customers/add_redirect_on_create?cuID=[id]'

    crud.messages.submit_button = T("Save")
    crud.messages.record_created = T("Saved")
    crud.settings.formstyle = "bootstrap3_stacked"
    crud.settings.create_onaccept = [add_oncreate]
    crud.settings.create_next = next_url
    form = crud.create(db.auth_user)

    form_id = "customer_add"
    form_element = form.element('form')
    form['_id'] = form_id

    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = form_id

    submit = form.element('input[type=submit]')

    back = add_get_back()

    return dict(content=form,
                back=back,
                save=submit)


def add_oncreate(form):
    from openstudio.os_customer import Customer

    customer = Customer(form.vars.id)
    customer.on_create()



def add_get_back(var=None):
    if 'teacher' in request.vars:
        url = URL('teachers', 'index')
    elif 'employee' in request.vars:
        url = URL('school_properties', 'employees')
    elif 'clsID' in request.vars:
        url = URL('classes', 'attendance', vars={
            'clsID': request.vars['clsID'],
            'date': request.vars['date'],
        })
    else:
        url = URL('customers', 'index')

    return os_gui.get_button(
        'back',
        url
    )



@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('create', 'auth_user'))
def add_redirect_on_create():
    """
        Redirect to edit, from the client side, to leave the add modal
    """
    cuID = request.vars['cuID']

    if session.customers_add_back == 'classes_attendance':
        session.flash = T("Added customer")
        redirect(URL('classes', 'attendance', vars=request.vars),
                 client_side=True)
    elif session.customers_add_back == 'school_employees':
        session.flash = T("Added employee")
        redirect(URL('school_properties', 'employees'),
                 client_side=True)
    elif session.customers_add_back == 'teachers':
        session.flash = T("Added teacher")
        redirect(URL('teachers' , 'index'),
                 client_side=True)
    else:
        # to edit page
        redirect(URL('edit', args=[cuID], extension=''), client_side=True)


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('update', 'auth_user'))
def edit():
    today = TODAY_LOCAL

    if not session.show_location:
        db.auth_user.school_locations_id.readable = False
        db.auth_user.school_locations_id.writable = False
    else:
        loc_query = (db.school_locations.Archived == False)
        db.auth_user.school_locations_id.requires = \
            IS_IN_DB(db(loc_query),
                     'school_locations.id',
                     '%(Name)s',
                     zero=T("Please select..."))

    db.auth_user.picture.readable=False
    db.auth_user.picture.writable=False

    # we're not entering password, so don't include it in the form
    db.auth_user.password.readable=False
    db.auth_user.password.writable=False

    # enable business checkbox
    db.auth_user.business.readable=True
    db.auth_user.business.writable=True
    # db.auth_user.business.widget=SQLFORM.widgets.checkboxes.widget


    picture_class = 'customer_image_edit'
    picture = DIV(IMG(_src=URL('static', 'images/person.png'), _alt="Person picture"),
                      _class=picture_class)

    # update customer
    customers_id = request.args[0]
    db.auth_user.id.label = T('Customer ID')

    if customers_id == '1' and auth.user.id != 1:
        redirect(URL('default', 'user', args=['not_authorized']))

    row = db.auth_user[customers_id]
    response.title = row.display_name
    response.subtitle = T("Profile")
    if row.merged:
        menu = ''
    else:
        menu = customers_get_menu(customers_id, 'general')

    left_sidebar_enabled = True
    label_id = T('ID') + ": "

    if not row.picture:
        change_picture_title = T("Add picture")
        picture = DIV(IMG(_src=URL('static', 'images/person.png'),
                          _alt=row.display_name),
                          _class=picture_class)
    else:
        picture = DIV(IMG(_src=URL('default', 'download',
                                   args=row.thumblarge),
                          _alt=row.display_name),
                          _class=picture_class)
        change_picture_title = T("Change picture")
    change_picture = A(change_picture_title,
                       _href=URL('edit_picture', args=[customers_id]))

    crud.settings.update_onaccept = [edit_onaccept]
    crud.messages.submit_button = T('Save')
    crud.messages.record_updated = T('Saved')

    # Clear teachers cache if we're updating a teacher
    if row.teacher:
        crud.settings.update_onaccept.append(cache_clear_school_teachers)

    form = crud.update(db.auth_user, customers_id)
    # Tie the elements together using the form html5 attribute.
    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = "MainForm"

    submit = os_gui.get_submit_button('MainForm')

    back = edit_get_back()

    # set mail button
    email = row.email
    mail_button = ''
    if row.email:
        mail_button = A(I(_class='fa fa-envelope-o'),
                          _class="btn btn-default",
                          _href='mailto:' + email)

    # add notes
    te_button = ''
    bo_button = ''

    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('read', 'customers_notes_backoffice'):
        bo_button = os_gui.get_button(
            'noicon',
            URL('notes', vars={'cuID': customers_id,
                               'note_type': 'backoffice'}),
            title=T("All notes"),
            btn_class='btn-link',
            btn_size=''
        )


    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('read', 'customers_notes_teachers'):
        te_button = os_gui.get_button(
            'noicon',
            URL('notes', vars={'cuID': customers_id,
                               'note_type': 'teachers'}),
            title=T("All notes"),
            btn_class='btn-link',
            btn_size=''
        )

    # get styles form
    form = edit_remodel_form(
        form,
        picture,
        change_picture,
        label_id,
        customers_id,
        bo_button=bo_button,
        te_button=te_button,
        mail_button=mail_button,
        merged=row.merged,
        row=row
    )

    alert = ''
    if row.merged:
        merged_into = db.auth_user(row.merged_into)
        merged_link = A(SPAN(merged_into.display_name, ' ',
                             T('ID'), ': ',
                             row.merged_into),
                        _title=T("link to account merged into"),
                        _href=URL('edit', args=[row.merged_into]))
        alert = os_gui.get_alert(
            'success',
            SPAN(B(T('Note')), ' ',
                 T("This account has been merged into"), ' ',
                 merged_link, ' ', T('on'), ' ',
                 row.merged_on.strftime(DATETIME_FORMAT), '.'),
            dismissable=False)
        # Don't show a submit button for merged customers
        submit = ''

    content = DIV(
        alert,
        form,
    _class = 'tab-pane active')


    return dict(content=content,
                back=back,
                menu=menu,
                save=submit)


def edit_onaccept(form):
    """
    :param form: crud form for db.auth_user
    :return: None
    """
    from openstudio.os_customer import Customer

    customer = Customer(form.vars.id)
    customer.on_update()


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'auth_user'))
def edit_picture():
    """
        This function is called to show a page to change a customers' picture
        After successfully changing the picture, the user is redirected to the edit page.
    """
    response.title = T("Picture")
    response.view = 'general/only_content.html'
    customers_id = request.args[0]
    row = db.auth_user[customers_id]
    response.subtitle = row.display_name

    session.customers_edit_clear_thumbs_customers_id = customers_id

    for field in db.auth_user:
        field.writable = False
        field.readable = False

    db.auth_user.picture.readable = True
    db.auth_user.picture.writable = True
    db.auth_user.picture.label = T("Picture")

    submit = ''
    space = uploads_available_space(request.folder)
    if space['available'] > 1 or not row.Picture is None:
        crud.settings.label_separator = ''
        crud.settings.update_deletable = False
        crud.settings.update_onvalidation.customers.append(_edit_clear_old_thumbs)
        crud.settings.update_onaccept.customers.extend(
            [ _edit_check_picture,
              cache_clear_school_teachers,
              edit_onaccept ]
        )

        crud.messages.submit_button = T('Save')
        crud.messages.record_updated = T('Saved')

        form = crud.update(db.auth_user, customers_id, next='edit/[id]')

        result = set_form_id_and_get_submit_button(form, 'MainForm')
        form = result['form']
        submit = result['submit']

        msg_list = UL(T("Maximum filesize: 4MB"))

        content = DIV(BR(), msg_list, form)
    else:
        content = space['full_message']

    back = os_gui.get_button('back', URL('edit', args=[customers_id]))

    return dict(content=content, save=submit, back=back)


def edit_get_back(_class=''):
    """
        This function looks at the session variable
            session.customers_back
        to determine where the back button for customers edit pages should link to.
    """
    if session.customers_back == 'keys':
        # check if we came from the edit button on the keys list page in school properties
        url = URL('school_properties', 'list_keys')
    elif session.customers_back == 'teachers':
        # check if we came from the school / teachers page
        url = URL('teachers', 'index')
    elif session.customers_back == 'school_employees':
        # check if we came fromthe school/employees page
        url = URL('school_properties', 'employees')
    elif session.customers_back == 'pinboard':
        # check if the birthday notification or a memo on the pinboard
        # referred to this page
        url = URL('pinboard', 'index')
    elif session.customers_back == 'subscriptions_new' or \
       session.customers_back == 'subscriptions_stopped' or \
       session.customers_back == 'subscriptions_paused' or \
       session.customers_back == 'subscriptions_overview_customers' or \
       session.customers_back == 'subscriptions_alt_prices':
        # check if the we came from the default/subscriptions page
        url = URL('reports', session.customers_back)
    elif session.customers_back == 'reports_customers_inactive':
        url = URL('reports', 'customers_inactive')
    elif session.customers_back == 'trialclasses':
        # check if the we came from the default/trialclasses page
        url = URL('reports', 'trialclasses')
    elif session.customers_back == 'trialcards':
        # check if the we came from the default/trialcards page
        url = URL('reports', 'trialcards')
    elif session.customers_back == 'dropinclasses':
        # check if the we came from the default/dropinclasses page
        url = URL('reports', 'dropinclasses')
    elif session.customers_back == 'classcards':
        # check if the we came from the default/classcards page
        url = URL('reports', 'classcards')
    elif session.customers_back == 'direct_debit_extra':
        # check if the we came from the default/alternative_payments page
        url = URL('reports', 'direct_debit_extra')
    elif session.customers_back == 'reports_attendance_subscription_exceeded':
        # check if we came from the reports/attendance_subcription_exceeded page
        url = URL('reports', 'attendance_subcription_exceeded')
    elif session.customers_back == 'reports_retention_rate':
        # check if we came from the reports/attendance_subcription_exceeded page
        url = URL('reports', 'retention_rate')
    # elif session.customers_back == 'classes_manage':
    #     # check if the classes/manage page referred to this page
    #     url = URL('classes', 'manage')
    # elif session.customers_back == 'workshops_attendance_add':
    #     # check if the workshops manage page referred to this page
    #     url = URL('workshops', 'attendance_add')
    # elif session.customers_back == 'workshops_attendance':
    #     # check if the workshops attendance page referred to this page
    #     url = URL('workshops', 'attendance')
    # elif session.customers_back == 'workshops_activity_attendance':
    #     # check if the workshops manage page referred to this page
    #     url = URL('workshops', 'activity_attendance')
    # elif session.customers_back == 'workshops_activity_attendance_add':
    #     # check if the workshops manage page referred to this page
    #     url = URL('workshops', 'activity_attendance_add')
    elif session.customers_back == 'finance_orders':
        # check if the finance orders page referred to this page
        url = URL('finance', 'orders')
    elif session.customers_back == 'tasks_index':
        # check if the tasks index page referred to this page
        url = URL('tasks', 'index')
    elif session.customers_back == 'finance_batch_content':
        # check if we're coming from a batch content page
        url = URL('finance', 'batch_content')
    else:
        url = URL('index')

    return os_gui.get_button('back', url, _class=_class)


def customers_get_menu(customers_id, page=None):
    pages = []

    pages.append(['general',
                   T('Profile'),
                  URL("customers","edit", args=[customers_id])])

    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('read', 'customers_memberships'):
        pages.append(['memberships',
                      T("Memberships"),
                      URL("customers","memberships",
                          vars={'cuID':customers_id})])
    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('read', 'customers_subscriptions'):
        pages.append(['subscriptions',
                      T("Subscriptions"),
                      URL("customers","subscriptions",
                          vars={'cuID':customers_id})])
    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('read', 'customers_classcards'):
        pages.append(['classcards',
                      T("Class cards"),
                      URL("customers","classcards", vars={'cuID':customers_id})])
    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('read', 'classes_reservation'):
        pages.append(['classes',
                      T("Classes"),
                      URL("customers", "classes_attendance",
                          vars={'cuID':customers_id})])
    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('read', 'workshops'):
        pages.append(['events',
                      T("Events"),
                      URL("customers","events", vars={'cuID':customers_id})])
    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('read', 'customers_documents'):
        pages.append(['documents',
                      T("Documents"),
                      URL("customers","documents", vars={'cuID':customers_id})])
    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('read', 'tasks'):
        pages.append(['tasks',
                      T("Tasks"),
                      URL("customers","tasks", vars={'cuID':customers_id})])
    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('read', 'orders'):
        pages.append(['orders',
                      T('Orders'),
                      URL("customers","orders", vars={'cuID':customers_id})])
    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('read', 'invoices'):
        pages.append(['invoices',
                      T("Invoices"),
                      URL("customers","invoices", vars={'cuID':customers_id})])


    more = []

    customer = Customer(customers_id)
    if customer.row.teacher:
        more.append([
            'edit_teacher',
            (os_gui.get_fa_icon('fa-graduation-cap'), ' ', T('Teacher profile')),
            URL('edit_teacher', vars={'cuID':customers_id})
        ])

    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('read', 'customers_payments_info'):
        more.append([
            'bankaccount',
            (os_gui.get_fa_icon('fa-university'), ' ', T("Finance")),
            URL("customers", "bankaccount", vars={'cuID':customers_id})])

    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('read', 'auth_user_account'):
        more.append([
            'account',
            (os_gui.get_fa_icon('fa-user-circle'), ' ', T('Account')),
            URL('customers', 'account', vars={'cuID':customers_id})
        ])

    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('read', 'auth_user'):
        more.append([
            'barcode_label',
            (os_gui.get_fa_icon('fa-barcode'), ' ', T('Barcode label')),
            URL('customers', 'barcode_label', vars={'cuID':customers_id}),
            'blank'
        ])

    pages.append([
        'more',
        T('More'),
        more
    ])


    return os_gui.get_submenu(pages, page, _id='os-customers_edit_menu', horizontal=True, htype='tabs')


def classcards_count_classes(row):
    ccd = CustomerClasscard(row.customers_classcards.id)
    link_text = ccd.get_classes_remaining_formatted()

    link = A(link_text,
             _href=URL('classcard_classes', vars=dict(ccdID=row.customers_classcards.id)))

    return link



@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'auth_user'))
def export_excel():
    """
        valid export_types include: payment_summary, mailinglist, attendance_list, customers_list, subscription_list
    """
    export_type = request.vars['export']

    # create filestream
    stream = cStringIO.StringIO()

    if export_type.lower() == "customers_list":
        # create dictionary to lookup latest subscription values
        if export_type.lower() == 'customers_list':
            result = db.executesql("""SELECT cu.id,
                                             ssu.name,
                                             cs.startdate,
                                             cs.enddate,
                                             pm.Name
                                       FROM auth_user cu
                LEFT JOIN customers_subscriptions cs
                ON cs.auth_customer_id = cu.id
                LEFT JOIN
                (SELECT auth_customer_id, school_subscriptions_id, max(startdate) as startdate, enddate
                FROM customers_subscriptions GROUP BY auth_customer_id) chk
                ON cu.id = chk.auth_customer_id
                LEFT JOIN
                (SELECT id, name FROM school_subscriptions) ssu
                ON ssu.id = cs.school_subscriptions_id
                LEFT JOIN payment_methods pm ON cs.payment_methods_id = pm.id
                WHERE cs.startdate = chk.startdate OR cs.startdate IS NULL """)

            fname = T("Customers list") + '.xlsx'
            title = 'Customers list'
        cu_mem_dict = dict()
        for record in result:
            subscription = record[1] or ""
            start = record[2] or ""
            end = record[3] or ""
            payment_method = record[4] or ""
            cu_mem_dict[record[0]] = [subscription, start, end, payment_method]

        bd_dict = dict()
        rows = db().select(db.customers_payment_info.ALL)
        for row in rows:
            bd_dict[row.auth_customer_id] = [
                                          row.AccountNumber,
                                          row.AccountHolder,
                                          row.BankName,
                                          row.BankLocation ]

        # Create the workbook
        wb = openpyxl.workbook.Workbook(write_only=True)
        ws = wb.create_sheet(title=title)
        headers = [ "id",
                    "First name",
                    "Last name",
                    "Date of birth",
                    "Gender",
                    "Address",
                    "Postal code",
                    "City",
                    "Country",
                    "Email",
                    "Newsletter",
                    "Telephone",
                    "Mobile",
                    "Key",
                    "Location",
                    "Subscription",
                    "Startdate",
                    "Enddate",
                    "Payment",
                    "AccountNR",
                    "AccountHolder",
                    "BankName",
                    "BankLocation"]
        ws.append(headers)

        query = (db.auth_user.trashed == False)
        rows = db(query).select(db.auth_user.ALL,
                                db.school_locations.Name,
                left=[db.school_locations.on(db.auth_user.school_locations_id==\
                                             db.school_locations.id)])
        for row in rows:
            customers_id = row.auth_user.id
            if export_type.lower() == 'subscription_list' and \
                not cu_mem_dict.has_key(customers_id):
                # subscription list, if no subscription --> check the next customer.
                continue
            else:
                data = [ row.auth_user.id,
                         row.auth_user.first_name,
                         row.auth_user.last_name,
                         row.auth_user.date_of_birth,
                         row.auth_user.gender,
                         row.auth_user.address,
                         row.auth_user.postcode,
                         row.auth_user.city,
                         row.auth_user.country,
                         row.auth_user.email,
                         row.auth_user.newsletter,
                         row.auth_user.phone,
                         row.auth_user.mobile,
                         row.auth_user.keynr,
                         row.school_locations.Name,
                         cu_mem_dict[customers_id][0],
                         cu_mem_dict[customers_id][1],
                         cu_mem_dict[customers_id][2],
                         cu_mem_dict[customers_id][3]]
                if not bd_dict.get(customers_id, None) is None:
                    data.append(bd_dict[customers_id][0])
                    data.append(bd_dict[customers_id][1])
                    data.append(bd_dict[customers_id][2])
                    data.append(bd_dict[customers_id][3])

                ws.append(data)

        wb.save(stream)

        response.headers['Content-Type']='application/vnd.ms-excel'
        response.headers['Content-disposition']='attachment; filename=' + fname

        return stream.getvalue()


    elif export_type.lower() == "mailing_list":
        wb = openpyxl.workbook.Workbook(write_only=True)
        # write the sheet for all mail addresses
        ws = wb.create_sheet(title="All customers")
        today = datetime.date.today()
        query = ((db.auth_user.trashed == False) &
                 (db.auth_user.id > 1))
        rows = db(query).select(db.auth_user.first_name,
                                db.auth_user.last_name,
                                db.auth_user.email)
        for row in rows:
            ws.append([row.first_name,
                       row.last_name,
                       row.email])
        # All newsletter
        ws = wb.create_sheet(title="Newsletter")
        today = datetime.date.today()
        query = ((db.auth_user.trashed == False) &
                 (db.auth_user.newsletter == True) &
                 (db.auth_user.id > 1))
        rows = db(query).select(db.auth_user.first_name,
                                db.auth_user.last_name,
                                db.auth_user.email)
        for row in rows:
            ws.append([row.first_name,
                       row.last_name,
                       row.email])

        fname = T("Mailinglist") + '.xlsx'
        wb.save(stream)

        response.headers['Content-Type']='application/vnd.ms-excel'
        response.headers['Content-disposition']='attachment; filename=' + fname

        return stream.getvalue()


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'customers_classcards'))
def classcards():
    """
        List class cards for a customer
    """
    response.view = 'customers/edit_general.html'
    customers_id = request.vars['cuID']

    customer = Customer(customers_id)
    response.title = customer.get_name()
    response.subtitle = T("Class cards")

    session.invoices_payment_add_back = 'customers_classcards'
    session.invoices_edit_back = 'customers_classcards'

    query = (db.customers_classcards.auth_customer_id == customers_id)
    db.customers_classcards.id.label = T("Pass number")
    db.customers_classcards.Enddate.readable = True

    maxtextlengths = {'customers_classcards.school_classcards_id' : 32,
                      'customers_classcards.Note' : 20}

    left = [
        db.invoices_items_customers_classcards.on(
            db.invoices_items_customers_classcards.customers_classcards_id ==
            db.customers_classcards.id),
        db.invoices_items.on(
            db.invoices_items_customers_classcards.invoices_items_id ==
            db.invoices_items.id),
        db.invoices.on(db.invoices_items.invoices_id ==
                       db.invoices.id) ]

    links = [ dict(header=T('Classes'), body=classcards_count_classes),
              dict(header=T('Invoices'), body=classcards_get_link_invoice),
              classcards_get_link_membership_check,
              lambda row: os_gui.get_button('edit',
                    URL('classcard_edit',
                        vars={'cuID':customers_id,
                              'ccdID':row.customers_classcards.id}))]
    fields = [ db.customers_classcards.id,
               db.customers_classcards.Startdate,
               db.customers_classcards.Enddate,
               db.customers_classcards.school_classcards_id,
               db.customers_classcards.auth_customer_id,
               db.invoices.id,
               db.invoices.Status,
               db.invoices.payment_methods_id ]
    headers = { 'customers_classcards.id':T("Card") }

    delete_permission = auth.has_membership(group_id='Admins') or \
                        auth.has_permission('delete', 'customers_classcards')

    db.invoices.id.readable = False
    db.invoices.Status.readable = False

    grid = SQLFORM.grid(query,
                        fields=fields,
                        headers=headers,
                        field_id=db.customers_classcards.id,
                        left=left,
                        links=links,
                        maxtextlengths=maxtextlengths,
                        create=False,
                        details=False,
                        csv=False,
                        searchable=False,
                        editable=False,
                        deletable = delete_permission,
                        ondelete=classcards_ondelete,
                        orderby=~db.customers_classcards.id,
                        ui=grid_ui)
    grid.element('.web2py_counter', replace=None) # remove the counter
    grid.elements('span[title=Delete]', replace=None) # remove text from delete button

    add_url = URL('classcard_add', vars={'cuID':customers_id})
    add = os_gui.get_button('add', add_url, T("Add a new class card"), btn_size='btn-sm', _class='pull-right')

    back = edit_get_back()

    menu = customers_get_menu(customers_id, request.function)

    return dict(content=grid, menu=menu, back=back, add=add)


def classcards_ondelete(table, record_id):
    """
        Function to clear cache for customer after deleting a card 
    """
    # Find customer id
    ccd = db.customers_classcards(record_id)
    cuID = ccd.auth_customer_id

    # Cancel invoice(s) for this classcard
    query = (db.invoices_items_customers_classcards.customers_classcards_id == record_id)
    rows = db(query).select(db.invoices_items_customers_classcards.ALL)
    for row in rows:
        invoice_item = db.invoices_items(row.invoices_items_id)

        invoice_query = (db.invoices.id == invoice_item.invoices_id)
        db(invoice_query).update(Status='cancelled')

    # Clear cache
    cache_clear_customers_classcards(cuID)


def classcards_get_link_membership_check(row):
    """
    :param row: gluon.dal.row object of db.customers_classcards
    :return: Warning if membership required for card but not found
    """
    scd = SchoolClasscard(row.customers_classcards.school_classcards_id, set_db_info=True)
    required_membership = scd.row.school_memberships_id

    customer = Customer(row.customers_classcards.auth_customer_id)

    memberships = customer.get_memberships_on_date(TODAY_LOCAL)
    ids = []
    for row in memberships:
        ids.append(row.id)

    if required_membership and not required_membership in ids:
        return os_gui.get_label(
            'warning',
            T('No membership'),
            title=T("This customer doesn't have the required membership for this card.")
        )
    else:
        return ''


def classcards_get_link_invoice(row):
    """
        Returns invoice for classcard in list
    """
    if row.invoices.id:
        invoices = Invoices()

        query = (db.invoices.id == row.invoices.id)
        rows = db(query).select(db.invoices.ALL)
        repr_row = rows.render(0)

        invoice_link = invoices.represent_invoice_for_list(
            row.invoices.id,
            repr_row.InvoiceID,
            repr_row.Status,
            row.invoices.Status,
            row.invoices.payment_methods_id
        )

    else:
        invoice_link = ''

    return invoice_link


def classcards_get_return_url(cuID, clsID=None, date_formatted=None):
    """
        returns the return url for class cards
    """
    if clsID:
        url = URL('classes', 'attendance_booking_options',
                     vars={'cuID' : cuID,
                           'clsID': clsID,
                           'date' : date_formatted})
    else:
        url = URL('classcards', vars={'cuID':cuID})

    return url


def classcards_clear_cache(form):
    """
        Clear the subscriptions cache for customer 
    """
    ccdID = form.vars.id
    ccd = db.customers_classcards(ccdID)
    cuID = ccd.auth_customer_id

    cache_clear_customers_classcards(cuID)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('create', 'customers_classcards'))
def classcard_add():
    """
        Add a new classcard for a customer in more graphic way than
        a drop down menu
        request.vars['cuID'] is expected to be the customers_id
    """
    def over_times_bought(row):
        if not row.Trialcard:
            return False

        query = (db.customers_classcards.auth_customer_id == customers_id) & \
                (db.customers_classcards.school_classcards_id == row.id)
        times_bought = db(query).count()

        if times_bought >= row.TrialTimes:
            return True
        else:
            return False

    customers_id   = request.vars['cuID']
    clsID          = request.vars['clsID'] # for redirect to classes attendance_list_classcards
    date_formatted = request.vars['date'] # for redirect to classes attendance_list_classcards
    customer = Customer(customers_id)
    response.title = customer.get_name()
    response.subtitle = T("New Class card")
    response.view = 'general/tabs_menu.html'

    cuID = request.vars['cuID']
    return_url = classcards_get_return_url(customers_id, clsID, date_formatted)

    query = (db.school_classcards.Archived == False)
    rows = db(query).select(db.school_classcards.ALL,
                            orderby=db.school_classcards.Trialcard|\
                                    db.school_classcards.Name)

    # check for no class cards
    count_cards = len(rows)

    back = DIV(os_gui.get_button('back_bs', return_url), BR(), BR(), _class='col-md-12')
    if count_cards < 1:
        content = DIV(back, T("No class cards found, please add one under school."), _class='row')
    else:
        content = DIV(back, _class='row')
    # populate as usual
    modals = DIV()
    for i, row in enumerate(rows):
        repr_row = list(rows[i:i + 1].render())[0]
        card_name = max_string_length(row.Name, 33)
        validity = classcard_get_validity(row)

        if clsID:
            vars = {'cuID' : cuID,
                    'scdID': row.id,
                    'clsID': clsID,
                    'date' : date_formatted}
        else:
            vars = {'cuID' : cuID,
                    'scdID': row.id}

        form_id = 'form_' + unicode(row.id)

        modal_content = LOAD('customers', 'classcard_add_modern_add_card.load',
                             ajax_trap=True,
                             vars=vars)
        result =  os_gui.get_modal(button_text=T('This one'),
                                   button_class='btn-block btn-link',
                                   modal_title=card_name,
                                   modal_content=modal_content,
                                   modal_footer_content=os_gui.get_submit_button(form_id),
                                   modal_class='modal_card_' + unicode(row.id))
        modals.append(result['modal'])
        max_bought = ''
        if over_times_bought(row):
            max_bought = DIV(
                os_gui.get_fa_icon('fa-info-circle'), ' ',
                T("Maximum cards bought"),
                _class='text-muted center'
            )


        card_content = DIV(
            TABLE(TR(TD(T("Validity"), TD(validity))),
                  TR(TD(T("Classes"), TD(repr_row.Classes))),
                  TR(TD(T("Price"), TD(repr_row.Price))),
                  _class='table'),
            max_bought,
            result['button'],
            )


        if row.Trialcard:
            panel_class = 'panel-success'
        else:
            panel_class = 'panel-primary'
        card = DIV(os_gui.get_panel_table(card_name, card_content, panel_class),
                   _class='col-md-3')

        content.append(card)


        if i == 4:
            content.append(BR())

    content.append(modals)

    back = edit_get_back()

    menu = customers_get_menu(customers_id, 'classcards')

    return dict(content=content, back=back, menu=menu)


def classcard_get_validity(row):
    """
        takes a db.school_classcards() row as argument
    """
    validity = SPAN(unicode(row.Validity), ' ')

    validity_in = represent_validity_units(row.ValidityUnit, row)
    if row.Validity == 1:  # Cut the last 's"
        validity_in = validity_in[:-1]

    validity.append(validity_in)

    return validity


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('create', 'customers_classcards'))
def classcard_add_modern_add_card():
    """
        This function shows an add page for a classcard
        request.vars['cuID'] is expected to be the customers_id
        request.vars['scdID'] is expected to be the school_classcards.id
    """
    cuID = request.vars['cuID']
    scdID = request.vars['scdID']

    response.title = T("New Class card")
    customer = Customer(cuID)
    response.subtitle = customer.get_name()

    return_url = URL('classcard_add_modern_add_card_redirect_classcards',
                     vars=request.vars, extension='')

    db.customers_classcards.school_classcards_id.default = scdID
    db.customers_classcards.school_classcards_id.readable = False
    db.customers_classcards.school_classcards_id.writable = False
    db.customers_classcards.Enddate.readable = False
    db.customers_classcards.Enddate.writable = False

    db.customers_classcards.auth_customer_id.default = cuID

    functions_onadd = classcard_add_get_functions()
    crud.messages.submit_button = T("Save")
    crud.messages.record_created = T("Added card")
    crud.settings.formstyle = 'bootstrap3_stacked'
    crud.settings.create_next = return_url
    crud.settings.create_onaccept = functions_onadd
    form = crud.create(db.customers_classcards)

    form_id = "form_" + unicode(scdID)
    form_element = form.element('form')
    form['_id'] = form_id

    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = form_id

    return dict(content=form)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('create', 'customers_classcards'))
def classcard_add_modern_add_card_redirect_classcards():
    """
        Redirect back to classcards list after adding a new card
        We need this extra function, because otherwise we're stuck in the modal
        This way we can call redirect() with client_side=True
    """
    session.flash = T("Added card")
    cuID  = request.vars['cuID']
    clsID = request.vars['clsID']
    date_formatted  = request.vars['date']

    redirect_url = classcards_get_return_url(cuID, clsID, date_formatted)
    redirect(redirect_url, client_side=True)

#
# @auth.requires(auth.has_membership(group_id='Admins') or \
#                auth.has_permission('create', 'customers_classcards'))
# def classcard_add_classic():
#     """
#         This function shows an add page for a classcard
#         request.vars['cuID'] is expected to be the customers_id
#     """
#     customers_id   = request.vars['cuID']
#     clsID          = request.vars['clsID'] # for redirect to classes attendance_list_classcards
#     date_formatted = request.vars['date'] # for redirect to classes attendance_list_classcards
#     customer = Customer(customers_id)
#     response.title = customer.get_name()
#     response.subtitle = T("New Class card")
#     response.view = 'general/tabs_menu.html'
#
#     db.customers_classcards.auth_customer_id.default = customers_id
#
#     classcard_add_check_trialcard_set_query(customers_id)
#
#     return_url = classcards_get_return_url(customers_id, clsID, date_formatted)
#
#     functions_onadd = classcard_add_get_functions()
#
#     db.customers_classcards.Enddate.readable = False
#     db.customers_classcards.Enddate.writable = False
#
#     crud.messages.submit_button = T("Save")
#     crud.messages.record_created = T("Added card")
#     crud.settings.create_next = return_url
#     crud.settings.create_onaccept = functions_onadd
#     crud.settings.formstyle = "bootstrap3_stacked"
#     form = crud.create(db.customers_classcards)
#
#     form_element = form.element('form')
#     form['_id'] = 'MainForm'
#
#     elements = form.elements('input, select, textarea')
#     for element in elements:
#         element['_form'] = "MainForm"
#
#     submit = form.element('input[type=submit]')
#
#     content = DIV(
#         H4(T("Add new card")),
#         BR(),
#         form
#     )
#
#     back = os_gui.get_button('back', return_url)
#
#
#     menu = customers_get_menu(customers_id, 'classcards')
#
#     return dict(content=content, back=back, menu=menu, save=submit)
#
#


def classcard_add_get_functions(var=None):
    """
        Functions to execute after adding a classcard
    """
    functions = [ classcard_add_set_enddate,
                  classcard_add_create_invoice,
                  classcards_clear_cache ]

    return functions


def classcard_add_set_enddate(form):
    """
        Calculate and set enddate when adding a classcard
    """
    # get info
    if 'school_classcards_id' in form.vars:
        # used for classic classcards
        scdID = form.vars.school_classcards_id
    else:
        # used for modern classcards
        scdID = db.customers_classcards.school_classcards_id.default

    scd = SchoolClasscard(scdID)
    enddate = scd.sell_to_customer_get_enddate(form.vars.Startdate)

    # set enddate
    row = db.customers_classcards(form.vars.id)
    row.Enddate = enddate
    row.update_record()


def classcard_add_create_invoice(form):
    """
        Add an invoice after adding a classcard
    """
    ccdID   = form.vars.id
    scdID   = form.vars.school_classcards_id

    scd = SchoolClasscard(scdID)
    scd.sell_to_customer_create_invoice(ccdID)


@auth.requires_login()
def classcard_edit():
    """
        This function shows an edit page for a classcard
        request.vars['cuID'] is expected to be the customers_id
        request.vars['ccdID'] is expected to be the classcardID
    """
    customers_id = request.vars['cuID']
    classcardID = request.vars['ccdID']
    response.title = T("Edit Class card") + " " + unicode(classcardID)
    customer = Customer(customers_id)
    classcard = CustomerClasscard(classcardID)
    response.subtitle = customer.get_name()
    response.view = 'general/tabs_menu.html'

    db.customers_classcards.school_classcards_id.writable = False
    db.customers_classcards.school_classcards_id.readable = False

    # permission check to prevent editing of enddate
    permission = auth.has_membership(group_id='Admins') or \
                 auth.has_permission('update', 'customers_classcards_enddate')
    if not permission:
        db.customers_classcards.Enddate.readable = False
        db.customers_classcards.Enddate.writable = False

    return_url = classcards_get_return_url(customers_id)

    crud.messages.submit_button = T("Save")
    crud.messages.record_updated = T("Updated class card")
    crud.settings.update_onaccept = [ classcard_edit_update_classes_taken, classcards_clear_cache ]
    crud.settings.update_next = return_url
    crud.settings.update_deletable = False
    crud.settings.formstyle = "bootstrap3_stacked"
    form = crud.update(db.customers_classcards, classcardID)

    form_element = form.element('form')
    form['_id'] = 'MainForm'

    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = "MainForm"

    submit = form.element('input[type=submit]')

    back = os_gui.get_button('back', return_url)

    content = DIV(
        H4(T("Edit class card")),
        BR(),
        form
   )


    menu = customers_get_menu(customers_id, 'classcards')

    return dict(content=content, back=back, menu=menu, save=submit)


def classcard_edit_update_classes_taken(form):
    """
        Updates number of classes taken when saving a classcard
    """
    ccdh = ClasscardsHelper()
    ccdh.set_classes_taken(form.vars.id)


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('read', 'customers_classcards'))
def classcard_classes():
    response.title = T("Class cards")
    ccdID = request.vars['ccdID']
    back = request.vars['back']
    response.subtitle = T("Classes taken using card") + " " + ccdID
    response.view = 'general/only_content.html'
    row = db.customers_classcards(ccdID)
    customers_id = row.auth_customer_id

    classcard = CustomerClasscard(ccdID)
    rows = classcard.get_rows_classes_taken()

    table = TABLE(TR(TH(T("Class date")),
                     TH(T("Location")),
                     TH(T("Type")),
                     TH(T("Start")),
                     _class='header'),
                  _class='table')
    for i, row in enumerate(rows):
        repr_row = list(rows[i:i+1].render())[0]

        clsID = row.classes.id
        date = row.classes_attendance.ClassDate
        teachers = class_get_teachers(clsID, date)

        table.append(TR(TD(repr_row.classes_attendance.ClassDate),
                        TD(repr_row.classes.school_locations_id),
                        TD(repr_row.classes.school_classtypes_id),
                        TD(repr_row.classes.Starttime)))

    back = os_gui.get_button('back',
                             URL('classcards', vars={'cuID':customers_id}))

    return dict(content=table, back=back)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'classes_reservation'))
def classes_reservations():
    """
        Show reservations for a customer
    """
    response.view = 'customers/edit_general.html'

    cuID = request.vars['cuID']
    session.customers_classes_reservation_add_vars = {}
    session.customers_classes_reservation_add_vars['cuID'] = cuID
    session.classes_reserve_back = 'customers_reservations'
    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = T("Classes")

    submenu = classes_get_submenu(request.function, cuID)
    content = DIV(submenu)

    add = os_gui.get_button('add',
                            URL('classes_reservation_add', vars={'cuID':cuID}),
                            btn_size='btn-sm',
                            _class='pull-right')

    if 'filter' in request.vars:
        session.customers_reservations_filter = request.vars['filter']
    elif session.customers_reservations_filter:
        pass
    else:
        session.customers_reservations_filter = 'recurring'

    # buttons = [ [ 'recurring', T('Enrollments') ],
    #             [ 'single', T('Drop in') ],
    #             [ 'trial', T('Trial') ] ]
    # filter_form = os_gui.get_radio_buttons_form(
    #     session.customers_reservations_filter,
    #     buttons)

    content.append(BR())
    #tools = DIV(filter_form, _class='pull-right')
    #content.append(tools)
    #content.append(BR())
    #content.append(BR())

    # list of reservations
    db.classes.id.readable = False
    db.classes_reservation.id.readable = False
    db.classes_reservation.auth_customer_id.readable = False
    links = ''
    headers = {'classes.Starttime' : T('Time'),
               'classes_reservation.Startdate' : T('Class date')}
    fields = [
          db.classes.id,
          db.classes.Week_day,
          db.classes.Starttime,
          db.classes.school_locations_id,
          db.classes.school_classtypes_id,
          db.classes_reservation.Startdate,
    ]

    query = (db.classes_reservation.auth_customer_id == cuID)
    if session.customers_reservations_filter == 'recurring':
        query &= (db.classes_reservation.SingleClass == False)
        query &= (db.classes_reservation.TrialClass == False)

        fields.append(db.classes_reservation.Enddate)
        headers.pop('classes_reservation.Startdate')

        links = [ lambda row: os_gui.get_button(
                        'edit', URL('classes', 'reservation_edit',
                                    vars={'crID':row.classes_reservation.id,
                                          'clsID':row.classes.id})) ]
    elif session.customers_reservations_filter == 'single':
        query &= (db.classes_reservation.SingleClass == True)
        query &= (db.classes_reservation.TrialClass == False)
    elif session.customers_reservations_filter == 'trial':
        query &= (db.classes_reservation.SingleClass == True)
        query &= (db.classes_reservation.TrialClass == True)


    maxtextlengths = {'classes_reservation.classes_id' : 50}
    left = db.classes.on(db.classes_reservation.classes_id == db.classes.id)

    delete_permission = auth.has_membership(group_id='Admins') or \
                        auth.has_permission('delete', 'classes_reservation')

    grid = SQLFORM.grid(
        query,
        fields=fields,
        headers=headers,
        links=links,
        details=False,
        searchable=False,
        deletable=delete_permission,
        ondelete=classes_reservations_ondelete,
        csv=False,
        create=False,
        editable=False,
        maxtextlengths=maxtextlengths,
        left=left,
        orderby=~db.classes_reservation.Startdate,
        field_id=db.classes_reservation.id,
        ui = grid_ui
    )
    grid.element('.web2py_counter', replace=None) # remove the counter
    grid.elements('span[title=Delete]', replace=None) # remove text from delete button

    content.append(grid)

    # Back button
    back = edit_get_back()
    menu = customers_get_menu(cuID, 'classes')

    return dict(content=content,
                back=back,
                add=add,
                menu=menu,
                left_sidebar_enabled=True)


def classes_reservations_ondelete(table, id):
    """
    :param table: db table
    :param id: classes_reservation record
    :return: none
    """
    from openstudio.os_classes_reservation import ClassesReservation
    ##
    # Remove booked classes after date
    ##
    reservation = ClassesReservation(id)
    bookings_removed = reservation.remove_attendance_booked_classes(TODAY_LOCAL)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('create', 'classes_reservation'))
def classes_reservation_add():
    """
        Add a new reservation for a customer on the selected date
        request.vars['cuID'] is expected to be db.customers.id
    """
    response.view = 'general/only_content.html'

    cuID = request.vars['cuID']
    session.customers_classes_reservation_add_vars = {}
    session.customers_classes_reservation_add_vars['cuID'] = cuID
    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = T("Enroll in a class")

    session.classes_reserve_back = 'customers_reservations'

    if 'date' in request.vars:
        # response.subtitle = SPAN(T('for'), ' ',
        #                          customer.get_name(), ' ',
        #                          request.vars['date'])
        default_date = datestr_to_python(DATE_FORMAT, request.vars['date'])
    else:
        default_date = datetime.date.today()

    session.customers_classes_reservation_add_vars['date'] = default_date

    customers = Customers()
    result = customers.classes_add_get_form_date(cuID, default_date)
    form = result['form']
    form_date = result['form_styled']


    db.classes.id.readable = False

    # list of classes
    grid = customers.classes_add_get_list(default_date, 'reservations')

    back = os_gui.get_button('back', URL('classes_reservations',
                                         vars={'cuID':cuID}),
                             _class='left')

    return dict(content=DIV(form_date, grid),
                back=back)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('create', 'classes_attendance'))
def classes_attendance_add():
    """
        Add customers to attendance for a class
    """
    response.view = 'general/only_content.html'

    cuID = request.vars['cuID']
    session.customers_classes_attendance_add_vars = {}
    session.customers_classes_attendance_add_vars['cuID'] = cuID
    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = T("Add attendance")

    if 'date' in request.vars:
        date = datestr_to_python(DATE_FORMAT, request.vars['date'])
    else:
        date = datetime.date.today()

    session.customers_classes_attendance_add_vars['date'] = date

    customers = Customers()
    result = customers.classes_add_get_form_date(cuID, date)
    form = result['form']
    form_date = result['form_styled']

    db.classes.id.readable = False
    # list of classes
    grid = customers.classes_add_get_list(date, 'attendance', cuID)

    back = os_gui.get_button('back', URL('classes_attendance',
                                         vars={'cuID':cuID}),
                             _class='left')

    return dict(content=DIV(form_date, grid),
                back=back)


@auth.requires(auth.has_membership(group_id='Admins') or
               auth.has_permission('create', 'classes_attendance'))
def classes_attendance_add_booking_options():
    """
        Page to show listing of booking options for customer
    """
    from openstudio.os_attendance_helper import AttendanceHelper
    from openstudio.os_class import Class

    cuID = request.vars['cuID']
    clsID = request.vars['clsID']
    date_formatted = request.vars['date']
    date = datestr_to_python(DATE_FORMAT, date_formatted)

    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = T("Add attendance")
    response.view = 'general/only_content.html'

    return_url = URL('customers', 'classes_attendance_add', vars={'cuID':cuID})

    cls = Class(clsID, date)

    ah = AttendanceHelper()
    options = ah.get_customer_class_booking_options_formatted(clsID,
                                                              date,
                                                              customer,
                                                              trial=True,
                                                              complementary=True,
                                                              list_type='attendance',
                                                              controller='classes')
    cancel = os_gui.get_button('noicon',
                               return_url,
                               title=T('Cancel'),
                               btn_size='')

    content = DIV(
        H4(T('Booking options for class'), ' ', cls.get_name(), _class='center'), BR(),
        options,
        DIV(BR(), cancel, _class="col-md-12 center"),
        _class="row"
    )

    back = os_gui.get_button('back', return_url)

    return dict(content=content,
                back=back)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'classes_waitinglist'))
def classes_waitinglist():
    """
        Show waitinglist for a customer
    """
    response.view = 'customers/edit_general.html'

    cuID = request.vars['cuID']
    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = T("Classes")

    submenu = classes_get_submenu(request.function, cuID)
    content = DIV(submenu)

    delete_permission = auth.has_membership(group_id='Admins') or \
                        auth.has_permission('delete', 'classes_waitinglist')

    # fill content area
    query = (db.classes_waitinglist.auth_customer_id == cuID)
    fields = [ db.classes.Week_day,
               db.classes.school_locations_id,
               db.classes.school_classtypes_id,
               db.classes.Starttime ]
    wai_grid = SQLFORM.grid(query, fields,
                            create=False,
                            details=False,
                            editable=False,
                            csv=False,
                            searchable=False,
                            deletable=delete_permission,
                            user_signature=False,
                            field_id=db.classes_waitinglist.id,
                            left=db.classes.on(
                                    db.classes_waitinglist.classes_id==
                                    db.classes.id),
                            ui = grid_ui)
    wai_grid.element('.web2py_counter', replace=None) # remove the counter
    # remove text from delete button
    wai_grid.elements('span[title=Delete]', replace=None)
    wai_title = H3(T("Waitinglist"))

    content.append(wai_grid)

    # Back button
    back = edit_get_back()
    menu = customers_get_menu(cuID, 'classes')

    return dict(content=content,
                back=back,
                menu=menu,
                left_sidebar_enabled=True)


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('read', 'classes_attendance'))
def classes_attendance():
    """
        Show waitinglist for a customer
    """
    response.view = 'customers/edit_general.html'

    session.classes_attendance_remove_back = 'customers'
    session.invoices_payment_add_back = 'customers_classes_attendance'

    cuID = request.vars['cuID']
    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = T("Classes")

    if 'all' in request.vars:
        limit = False
        limit_by = False
        link_all = ''
    else:
        limit = 25
        limit_by = limit + 1
        link_all = A(T('Show all'), _href=URL(vars={'cuID':cuID,
                                                    'all': True}))

    submenu = classes_get_submenu(request.function, cuID)
    content = DIV(submenu)
    add = os_gui.get_button('add',
                            URL('classes_attendance_add', vars={'cuID':cuID}),
                            btn_size='btn-sm',
                            _class='pull-right')
    content.append(BR())

    header = THEAD(TR(TH(T('Date')),
                      TH(T('Time')),
                      TH(T('Class')),
                      TH(T('Location')),
                      TH(T('Used')),
                      TH(T('Status')),
                      TH(),
                      TH())) # Actions

    table = TABLE(header, _class='table table-striped table-hover')

    edit_permission = (auth.has_membership(group_id='Admins') or
                       auth.has_permission('update', 'classes_attendance'))
    delete_permission = (auth.has_membership(group_id='Admins') or
                         auth.has_permission('delete', 'classes_attendance'))


    rows = customer.get_classes_attendance_rows(limit_by)


    for i, row in enumerate(rows):
        if limit:
            if i + 1 > limit:
                break

        repr_row = list(rows[i:i + 1].render())[0]

        cancel = ''
        if edit_permission and not row.classes_attendance.BookingStatus == 'cancelled':
            cancel = classes_attendance_get_link_cancel(row)

        remove = ''
        if delete_permission:
            remove = classes_attendance_get_link_delete(row)

        ##
        # Attendance type labels
        ##
        ct = ''
        if row.classes_attendance.AttendanceType == 1:
            # trial class
            ct = os_gui.get_label('success', repr_row.classes_attendance.AttendanceType)
        elif row.classes_attendance.AttendanceType == 2:
            # drop in class
            ct = os_gui.get_label('primary', repr_row.classes_attendance.AttendanceType)
        elif row.classes_attendance.AttendanceType == 4:
            # Complementary class
            ct = os_gui.get_label('default', repr_row.classes_attendance.AttendanceType)
        elif row.classes_attendance.AttendanceType == 5:
            # Review requested
            ct = os_gui.get_label('warning', repr_row.classes_attendance.AttendanceType)

        ##
        # Invoice
        ##
        invoice = ''
        if row.invoices.id:
            invoices = Invoices()
            invoice = invoices.represent_invoice_for_list(
                row.invoices.id,
                repr_row.invoices.InvoiceID,
                repr_row.invoices.Status,
                row.invoices.Status,
                row.invoices.payment_methods_id
            )


        att_type = ''
        if row.classes_attendance.customers_subscriptions_id:
            att_type = repr_row.classes_attendance.customers_subscriptions_id
        elif row.classes_attendance.customers_classcards_id:
            att_type = SPAN(row.school_classcards.Name,
                            _title=T('Class card') + ' ' + unicode(row.classes_attendance.customers_classcards_id))

        tr = TR(TD(repr_row.classes_attendance.ClassDate),
                TD(SPAN(repr_row.classes.Starttime, ' - ',
                        repr_row.classes.Endtime)),
                TD(repr_row.classes.school_classtypes_id),
                TD(repr_row.classes.school_locations_id),
                TD(att_type),
                TD(repr_row.classes_attendance.BookingStatus, ' ', ct),
                TD(invoice),
                TD(remove, cancel)
                )
        table.append(tr)

    content.append(table)

    # determine whether to show show all link
    if limit:
        if len(rows) <= limit:
            link_all = ''
    content.append(link_all)

    # Back button
    back = edit_get_back()
    menu = customers_get_menu(cuID, 'classes')

    return dict(content=content,
                back=back,
                add=add,
                menu=menu,
                left_sidebar_enabled=True)


def classes_attendance_get_link_cancel(row):
    """
        Returns cancel button for classes_attendance
    """
    button = ''

    onclick_cancel = "return confirm('" + \
     T('Do you really want to cancel this booking and refund any credits associated with it?')\
     + "');"

    button = os_gui.get_button('cancel_notext',
       URL('customers', 'classes_attendance_cancel',
           vars={'caID':row.classes_attendance.id}),
       onclick=onclick_cancel,
       tooltip=T('Cancel booking'),
       _class='pull-right')

    return button


def classes_attendance_get_link_delete(row):
    """
        Checks delete permissions and returns button if granted
    """
    button = ''

    onclick_delete = "return confirm('" + \
     T('Do you really want to remove this class from the attendance list?')\
     + "');"

    date = row.classes_attendance.ClassDate.strftime(DATE_FORMAT)
    cuID = row.classes_attendance.auth_customer_id

    button = os_gui.get_button('delete_notext',
       URL('classes', 'attendance_remove',
           vars={'clattID':row.classes_attendance.id}),
       onclick=onclick_delete,
       tooltip=T('Delete'),
       _class='pull-right')

    return button


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('update', 'classes_attendance'))
def classes_attendance_cancel():
    """
        Actually cancel a booking
        request.vars['caID'] is expected to be from db.classes_attendance.id
    """
    caID = request.vars['caID']
    clatt = ClassAttendance(caID)

    clatt.set_status_cancelled(force=True)

    redirect(URL('customers', 'classes_attendance', vars={'cuID':clatt.row.auth_customer_id}))


def classes_get_submenu(page, cuID):
    """
        This function returns a submenu for the classes edit pages
    """
    vars = {'cuID':cuID}
    pages = [['classes_attendance', T('Attendance'),
              URL('classes_attendance', vars=vars)],
             ['classes_reservations', T('Enrollments'),
               URL('classes_reservations', vars=vars)],
             ['classes_waitinglist', T('Waitinglist'),
               URL('classes_waitinglist', vars=vars)]]

    horizontal = True

    return os_gui.get_submenu(pages, page, horizontal=horizontal, htype='tabs')


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'classes_reservation'))
def classes_edit():
    customers_id = request.vars['cuID']
    session.classes_reserve_back = 'customers'
    response.view = 'general/one_grid_with_sidebar.html'
    response.title = T("Edit class reservations")
    row = db.customers[customers_id]
    response.subtitle = row.display_name

    query = (db.classes)
    links = [classes_check_reservation]
    fields = [ db.classes.Week_day,
               db.classes.school_locations_id,
               db.classes.school_classtypes_id,
               db.classes.Starttime ]
    grid = SQLFORM.grid(query,
                        fields,
                        create=False,
                        details=False,
                        deletable=False,
                        editable=False,
                        searchable=False,
                        csv=False,
                        links=links,
                        paginate=50,
                        orderby=db.classes.Week_day|\
                                db.classes.school_locations_id|\
                                db.classes.Starttime,
                        ui=grid_ui)
    grid.element('.web2py_counter', replace=None) # remove the counter

    back = os_gui.get_button('back', URL('classes', vars={'cuID':customers_id}))

    return dict(grid=grid, back=back)


def subscription_credits_clear_cache(form):
    """
        Clear the subscriptions cache for customer
    """
    cscID = form.vars.id
    csc = db.customers_subscriptions_credits(cscID)
    cs = db.customers_subscriptions(csc.customers_subscriptions_id)
    cuID = cs.auth_customer_id

    cache_clear_customers_subscriptions(cuID)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'customers_subscriptions_credits'))
def subscription_credits():
    """
        This function shows a page listing the credits total for a subscription
        request.vars['csID'] is expected to be db.customers_subscriptions.id
    """
    response.view = 'general/tabs_menu.html'
    cuID = request.vars['cuID']
    csID = request.vars['csID']
    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = subscription_edit_get_subtitle(csID)

    cs = CustomerSubscription(csID)

    total = H4(T('Balance:'), ' ', cs.get_credits_balance(), _class='pull-right')
    mutations = cs.get_credits_mutations_rows(formatted=True,
                                              editable=True,
                                              deletable=True,
                                              delete_controller='customers',
                                              delete_function='subscription_credits_delete')

    add = os_gui.get_button('add',
                            URL('subscription_credits_add', vars={'csID':csID,
                                                                  'cuID':cuID}),
                            btn_size='btn-sm')

    back = subscription_edit_get_back(cuID)
    menu = subscription_edit_get_menu(cuID, csID, request.function)

    return dict(content=DIV(total, mutations), menu=menu, back=back, add=add)


def subscription_credits_get_return_url(cuID, csID):
    """
        Return url for subscription credits
    """
    return URL('customers', 'subscription_credits', vars={'cuID':cuID,
                                                          'csID':csID})


@auth.requires_login()
def subscription_credits_add():
    """
        This function shows an add page for subscription credits
        request.vars['cuID'] is expected to be auth.user.id
        request.vars['csID'] is expected to be customers_subscriptions_id
    """
    response.view = 'general/tabs_menu.html'
    cuID = request.vars['cuID']
    csID = request.vars['csID']
    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = subscription_edit_get_subtitle(csID)

    db.customers_subscriptions_credits.customers_subscriptions_id.default = csID

    return_url = subscription_credits_get_return_url(cuID, csID)

    crud.messages.submit_button = T("Save")
    crud.messages.record_created = T("Added subscription credit(s)")
    crud.settings.create_next = return_url
    crud.settings.create_onaccept = [subscription_credits_clear_cache]
    form = crud.create(db.customers_subscriptions_credits)

    element_form = form.element('form')
    element_form['_id'] = "MainForm"

    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = "MainForm"

    submit = form.element('input[type=submit]')


    subscr_back = os_gui.get_button('back_bs', URL('subscription_credits', vars={'cuID':cuID,
                                                                                 'csID':csID}))
    content = DIV(subscr_back, form)

    back = os_gui.get_button("back", return_url, _class='')

    menu = customers_get_menu(cuID, 'subscriptions_credits')

    return dict(content=content, back=back, menu=menu, save=submit)


@auth.requires_login()
def subscription_credits_edit():
    """
        This function shows an edit page for  subscription credits
        request.vars['cuID'] is expected to be auth.user.id
        request.vars['csID'] is expected to be customers_subscriptions_id
    """
    response.view = 'general/tabs_menu.html'
    cuID = request.vars['cuID']
    csID = request.vars['csID']
    cscID = request.vars['cscID']
    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = subscription_edit_get_subtitle(csID)

    db.customers_subscriptions_credits.customers_subscriptions_id.default = csID

    return_url = subscription_credits_get_return_url(cuID, csID)

    crud.messages.submit_button = T("Save")
    crud.messages.record_updated = T("Saved")
    crud.settings.update_next = return_url
    crud.settings.update_onaccept = [subscription_credits_clear_cache]
    form = crud.update(db.customers_subscriptions_credits, cscID)

    element_form = form.element('form')
    element_form['_id'] = "MainForm"

    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = "MainForm"

    submit = form.element('input[type=submit]')


    subscr_back = os_gui.get_button('back_bs', URL('subscription_credits', vars={'cuID':cuID,
                                                                                 'csID':csID}))
    content = DIV(subscr_back, form)

    back = os_gui.get_button("back", return_url, _class='')

    menu = customers_get_menu(cuID, 'subscriptions_credits')

    return dict(content=content, back=back, menu=menu, save=submit)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('delete', 'customers_subscriptions_credits'))
def subscription_credits_delete():
    """
        Delete subscription credits
    """
    cuID = request.vars['cuID']
    csID = request.vars['csID']
    cscID = request.vars['cscID']

    query = (db.customers_subscriptions_credits.id == cscID)
    db(query).delete()

    cache_clear_customers_subscriptions(cuID)

    redirect(URL('subscription_credits', vars={'cuID':cuID, 'csID':csID}))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'customers_subscriptions_pauses'))
def subscription_pauses():
    """
        This function shows a page which lists all pauses for a subscription
        request.vars['csID'] is expected to be the subscription ID
    """
    response.view = 'general/tabs_menu.html'
    cuID = request.vars['cuID']
    csID = request.vars['csID']
    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = subscription_edit_get_subtitle(csID)

    row = db.customers_subscriptions(csID)
    db.customers_subscriptions_paused.id.readable = False

    query = (db.customers_subscriptions_paused.customers_subscriptions_id == csID)
    if db(query).count() == 0:
        grid = DIV(BR(), T("This subscription hasn't been paused before."))
    else:
        grid = SQLFORM.grid(query,
            create=False,
            details=False,
            editable=False,
            searchable=False,
            csv=False,
            paginate=50,
            orderby=db.customers_subscriptions_paused.Startdate,
            field_id=db.customers_subscriptions_paused.id,
            ui = grid_ui)
        grid.element('.web2py_counter', replace=None) # remove the counter
        grid.elements('span[title=Delete]', replace=None) # remove text from delete button

    add = os_gui.get_button('add', URL('subscription_pause_add', vars={'csID':csID}), btn_size='btn-sm')

    back = subscription_edit_get_back(cuID)
    menu = subscription_edit_get_menu(cuID, csID, request.function)

    return dict(content=grid, menu=menu, back=back, add=add)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'customers_subscriptions_paused'))
def subscription_pause_add():
    """
        This function shows a page to allow a user to pause a subscription for multiple months
        request.vars['csID'] is expected to be the customers_subscriptions.id
    """
    response.view = 'general/tabs_menu.html'
    csID = request.vars['csID']
    cs   = CustomerSubscription(csID)
    cuID = cs.auth_customer_id
    customer = Customer(cuID)

    response.title = customer.get_name()
    response.subtitle = SPAN(T("Pause subscription"), ' ',
                             cs.name)

    today = datetime.date.today()
    return_url = URL('subscription_pauses', vars={'cuID':cuID,
                                                  'csID':csID})

    months = get_months_list()

    form = SQLFORM.factory(
        Field('from_month', 'integer',
            requires=IS_IN_SET(months),
            default=today.month),
        Field('from_year', 'integer',
            default=today.year),
        Field('until_month', 'integer',
            requires=IS_IN_SET(months),
            default=today.month),
        Field('until_year', 'integer',
            default=today.year),
        Field('description'),
        separator = '',
        submit_button = T("Save")
        )

    result = set_form_id_and_get_submit_button(form, 'MainForm')
    form = result['form']
    submit = result['submit']

    # change type to number from the default text
    form.element('#no_table_from_year').attributes['_type'] = 'number'
    form.element('#no_table_from_year').attributes['_class'] += ' os-input_year'
    form.element('#no_table_from_month').attributes['_class'] += ' os-input_month'
    form.element('#no_table_until_year').attributes['_type'] = 'number'
    form.element('#no_table_until_year').attributes['_class'] += ' os-input_year'
    form.element('#no_table_until_month').attributes['_class'] += ' os-input_month'
    form.element('#no_table_description').attributes['_placeholder'] = T("Description")


    if form.process().accepted:
        start = datetime.date(int(form.vars.from_year), int(form.vars.from_month), 1)
        end = get_last_day_month(datetime.date(int(form.vars.until_year), int(form.vars.until_month), 1))
        description = form.vars.description
        db.customers_subscriptions_paused.insert(customers_subscriptions_id = csID,
                                                 Startdate = start,
                                                 Enddate = end,
                                                 Description = description)
        redirect(return_url)


    back = os_gui.get_button('back_bs', return_url)
    content = DIV(
        DIV(back,
            XML('<form action="#" enctype="multipart/form-data" id="MainForm" method="post">'),
            TABLE(
                TR(TD(LABEL(T("From the start of "))),
                   TD(form.custom.widget.from_month),
                   TD(form.custom.widget.from_year)),
                TR(TD(LABEL(T("Until the end of "))),
                   TD(form.custom.widget.until_month),
                   TD(form.custom.widget.until_year)),
                TR(TD(),
                   TD(form.custom.widget.description, _colspan='2')),
                _id='customer_subscription_pause_add'),
            BR(),
            form.custom.end,
            _class="col-md-12"),
        _class='row')


    back = subscription_edit_get_back(cuID)
    menu = subscription_edit_get_menu(cuID, csID, 'subscription_pauses')

    return dict(content=content, menu=menu, back=back, save=submit)




@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'customers_subscriptions_blocked'))
def subscription_blocks():
    """
        This function shows a page which lists all pauses for a subscription
        request.vars['csID'] is expected to be the subscription ID
    """
    response.view = 'general/tabs_menu.html'
    cuID = request.vars['cuID']
    csID = request.vars['csID']
    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = subscription_edit_get_subtitle(csID)

    row = db.customers_subscriptions(csID)
    db.customers_subscriptions_blocked.id.readable = False

    query = (db.customers_subscriptions_blocked.customers_subscriptions_id == csID)
    if db(query).count() == 0:
        grid = DIV(BR(), T("This subscription hasn't been blocked before."))
    else:
        links = [
            subscription_blocks_get_link_edit
        ]
        maxtextlengths = {'customers_subscriptions_blocked.Description': 60}
        grid = SQLFORM.grid(query,
            links=links,
            create=False,
            details=False,
            editable=False,
            searchable=False,
            csv=False,
            paginate=50,
            maxtextlengths=maxtextlengths,
            orderby=db.customers_subscriptions_blocked.Startdate,
            field_id=db.customers_subscriptions_blocked.id,
            ui = grid_ui)
        grid.element('.web2py_counter', replace=None) # remove the counter
        grid.elements('span[title=Delete]', replace=None) # remove text from delete button

    add = os_gui.get_button(
        'add',
        URL('subscription_block_add', vars={'csID':csID, 'cuID': cuID}),
        btn_size='btn-sm'
    )

    back = subscription_edit_get_back(cuID)
    menu = subscription_edit_get_menu(cuID, csID, request.function)

    return dict(content=grid, menu=menu, back=back, add=add)


def subscription_blocks_get_link_edit(row):
    """
    Return HTML edit button for subscription blocks if the user
    has update permissions
    """
    if not (auth.has_membership(group_id='Admins') or
            auth.has_permission('update', 'customers_subscriptions_blocked')):
        return ''

    cs = db.customers_subscriptions(row.customers_subscriptions_id)
    cuID = cs.auth_customer_id

    return os_gui.get_button(
        'edit',
        URL('subscription_block_edit', vars={
            'cuID': cuID,
            'csID': row.customers_subscriptions_id,
            'csbID': row.id
        })
    )



@auth.requires_login()
def subscription_block_add():
    """
        Add a new product
    """
    from openstudio.os_forms import OsForms

    response.view = 'general/tabs_menu.html'
    cuID = request.vars['cuID']
    csID = request.vars['csID']
    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = subscription_edit_get_subtitle(csID)

    return_url = subscription_block_get_return_url(cuID, csID)

    db.customers_subscriptions_blocked.customers_subscriptions_id.default = csID

    os_forms = OsForms()
    result = os_forms.get_crud_form_create(
        db.customers_subscriptions_blocked,
        return_url,
    )

    form = result['form']
    back = os_gui.get_button('back', return_url)

    content = DIV(
        H4(T('Add check-in block')),
        form
    )

    menu = subscription_edit_get_menu(cuID, csID, 'subscription_blocks')

    return dict(content=content,
                save=result['submit'],
                back=back,
                menu=menu)


@auth.requires_login()
def subscription_block_edit():
    """
        Add a new product
    """
    from openstudio.os_forms import OsForms

    response.view = 'general/tabs_menu.html'
    cuID = request.vars['cuID']
    csID = request.vars['csID']
    csbID = request.vars['csbID']
    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = subscription_edit_get_subtitle(csID)

    return_url = subscription_block_get_return_url(cuID, csID)

    db.customers_subscriptions_blocked.customers_subscriptions_id.default = csID

    os_forms = OsForms()
    result = os_forms.get_crud_form_update(
        db.customers_subscriptions_blocked,
        return_url,
        csbID
    )

    form = result['form']
    back = os_gui.get_button('back', return_url)

    content = DIV(
        H4(T('Edit check-in block')),
        form
    )

    menu = subscription_edit_get_menu(cuID, csID, 'subscription_blocks')

    return dict(content=content,
                save=result['submit'],
                back=back,
                menu=menu)


def subscription_block_get_return_url(cuID, csID):
    return URL('subscription_blocks', vars={'cuID':cuID, 'csID': csID})



def subscriptions_get_return_url(customers_id):
    """
        Returns return URL for subscriptions
    """
    return URL('subscriptions', vars={'cuID':customers_id})


def subscriptions_clear_cache(form):
    """
        Clear the subscriptions cache for customer 
    """
    csID = form.vars.id
    cs = db.customers_subscriptions(csID)
    cuID = cs.auth_customer_id

    cache_clear_customers_subscriptions(cuID)


@auth.requires_login()
def subscription_add():
    """
        This function shows an add page for a subscription
        request.args[0] is expected to be the customers_id
    """
    customers_id = request.vars['cuID']
    customer = Customer(customers_id)
    response.view = 'general/tabs_menu.html'
    response.title = customer.get_name()
    response.subtitle = T("New subscription")

    db.customers_subscriptions.auth_customer_id.default = customers_id

    return_url = subscriptions_get_return_url(customers_id)

    crud.messages.submit_button = T("Save")
    crud.messages.record_created = T("Added subscription")
    crud.settings.formstyle = "bootstrap3_stacked"
    crud.settings.create_next = return_url
    crud.settings.create_onaccept = [
        subscriptions_clear_cache,
        subscription_add_add_credits,
        subscription_add_create_invoice

    ]
    form = crud.create(db.customers_subscriptions)

    element_form = form.element('form')
    element_form['_id'] = "MainForm"

    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = "MainForm"

    submit = form.element('input[type=submit]')

    back = os_gui.get_button('back', URL('subscriptions', vars={'cuID':customers_id}))
    content = DIV(
        H4(T("Add subscription")), BR(),
        form
    )

    menu = customers_get_menu(customers_id, 'subscriptions')

    return dict(content=content, back=back, menu=menu, save=submit)


def subscription_add_add_credits(form):
    """
        Add credits when adding a subscription
    """
    csID = form.vars.id
    date = form.vars.Startdate

    cs = CustomerSubscription(csID)
    cs.add_credits_month(date.year, date.month)


def subscription_add_create_invoice(form):
    """
    Create invoice when adding a subscription
    """
    csID = form.vars.id
    date = form.vars.Startdate
    year = date.year
    month = date.month

    igpt = db.invoices_groups_product_types(ProductType='subscription')
    igID = igpt.invoices_groups_id

    iID = db.invoices.insert(
        invoices_groups_id=igID,
        payment_methods_id=form.vars.payment_methods_id,
        customers_subscriptions_id=csID,
        SubscriptionYear=year,
        SubscriptionMonth=month,
        Description=T("Subscription"),
        Status='sent'
    )

    invoice = Invoice(iID)
    invoice.link_to_customer(request.vars['cuID'])
    iiID = invoice.item_add_subscription(csID, year, month)
    invoice.link_item_to_customer_subscription(csID, iiID)


@auth.requires_login()
def subscription_edit():
    """
        This function shows an edit page for a subscription
        request.args[0] is expected to be the customers_id
        request.args[1] is expected to be the subscriptionID
    """
    response.view = 'general/tabs_menu.html'
    cuID = request.vars['cuID']
    csID = request.vars['csID']
    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = subscription_edit_get_subtitle(csID)

    return_url = subscriptions_get_return_url(cuID)

    crud.messages.submit_button = T("Save")
    crud.messages.record_updated = T("Updated subscription")
    crud.settings.formstyle = "bootstrap3_stacked"
    crud.settings.update_next = return_url
    crud.settings.update_onaccept = [subscriptions_clear_cache, subscription_edit_onaccept]
    crud.settings.update_deletable = False
    form = crud.update(db.customers_subscriptions, csID)

    element_form = form.element('form')
    element_form['_id'] = "MainForm"

    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = "MainForm"

    submit = form.element('input[type=submit]')


    back = subscription_edit_get_back(cuID)
    menu = subscription_edit_get_menu(cuID, csID, request.function)

    return dict(content=form,
                menu=menu,
                save=submit,
                back=back)


def subscription_edit_onaccept(form):
    """
        :param form: CRUD form from db.customers_subscriptions
        :return: None
    """
    if not form.vars.Enddate is None:
        enddate = form.vars.Enddate

        query = (db.classes_attendance.customers_subscriptions_id == form.vars.id) & \
                ((db.classes_attendance.ClassDate > enddate) & (db.classes_attendance.ClassDate >= TODAY_LOCAL))
        rows = db(query).select(db.classes_attendance.id)
        ids = []
        for row in rows:
            ids.append(row.id)

        query = (db.classes_attendance.id.belongs(ids))
        db(query).update(BookingStatus = 'cancelled')



@auth.requires(auth.has_membership(group_id='Admins') or
               auth.has_permission('delete', 'customers_subscriptions'))
def subscription_delete():
    """
        Delete a subscription if no invoices are linked to it
        Otherwise display a message saying to have a look at those first
    """
    response.view = 'general/only_content.html'
    cuID = request.vars['cuID']
    csID = request.vars['csID']
    customer = Customer(cuID)
    response.title = customer.get_name()
    cs = CustomerSubscription(csID)
    response.subtitle = SPAN(T("Delete subscription"), ': ', cs.name)

    query = (db.invoices_items_customers_subscriptions.customers_subscriptions_id == csID)
    invoice_count = db(query).count()

    if invoice_count:
        content = DIV(
            H3(T("Unable to delete")),
            T("One or more invoices is linked to this subscription.")
        )
    else:
        session.flash = T('Deleted subscription')
        cache_clear_customers_subscriptions(cuID)

        query = (db.customers_subscriptions.id == csID)
        db(query).delete()

        cache_clear_customers_subscriptions(cuID)

        redirect(URL('subscriptions', vars={'cuID':cuID}))


    back = subscription_edit_get_back(cuID)

    return dict(content = content,
                back=back)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'invoices'))
def subscription_invoices():
    """
        Page to list invoices for a subscription
    """
    if 'csID' in request.vars:
        csID = request.vars['csID']
        session.customers_subscription_invoices_csID = csID
    elif session.customers_subscription_invoices_csID:
        csID = session.customers_subscription_invoices_csID


    cuID  = request.vars['cuID']
    response.view = 'general/tabs_menu.html'

    session.invoices_edit_back = 'customers_subscription_invoices'
    session.invoices_payment_add_back = 'customers_subscription_invoices'

    # Always reset filter
    session.invoices_list_status = None

    customer = Customer(cuID)
    cs = CustomerSubscription(csID)
    response.title = customer.get_name()
    response.subtitle = SPAN(T("Edit subscription"), ' ',
                             cs.name)

    # add button
    invoices = Invoices()
    form = invoices.add_get_form(cuID, csID)
    result = invoices.add_get_modal(form)
    add = result['button']
    modal_class = result['modal_class']

    status_filter = invoices.list_get_status_filter()

    if len(form.errors):
        response.js = "show_add_modal();"

    list = invoices.list_invoices(cuID=cuID, csID=csID)

    # main list
    content = DIV(DIV(status_filter,list))

    menu = subscription_edit_get_menu(cuID, csID, request.function)
    back = subscription_edit_get_back(cuID)

    return dict(content=content,
                menu=menu,
                add=add,
                back=back,
                form_add=form,
                modal_class=modal_class)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'customers_subscriptions_alt_prices'))
def subscription_alt_prices():
    """
        Specify a different price for a given month for a subscription
    """
    csID = request.vars['csID']
    cuID  = request.vars['cuID']
    response.view = 'general/tabs_menu.html'

    customer = Customer(cuID)
    cs = CustomerSubscription(csID)
    response.title = customer.get_name()
    response.subtitle = subscription_edit_get_subtitle(csID)

    csap = db.customers_subscriptions_alt_prices
    ## main listing
    # set the default value in the grid
    links = [ lambda row: os_gui.get_button('edit',
                URL('subscription_alt_prices_edit', vars={'cuID':cuID,
                                                          'csID':csID,
                                                          'csapID':row.id})),
              lambda row: os_gui.get_button('repeat',
                                            URL('subscription_alt_price_repeat', vars={'csapID':row.id}),
                                            tooltip=T("Repeat")) ]

    query = (csap.customers_subscriptions_id == csID)

    fields = [ csap.SubscriptionMonth,
               csap.SubscriptionYear,
               csap.Amount,
               csap.Description,
               csap.Note ]

    maxtextlengths = {'customers_subscriptions_alt_prices.Description' : 36,
                      'customers_subscriptions_alt_prices.Note' : 36}
    delete_permission = auth.has_membership(group_id='Admins') or \
                        auth.has_permission('delete', 'customers_subscriptions_alt_prices')

    grid = SQLFORM.grid(query,
                        links=links,
                        fields=fields,
                        details=False,
                        searchable=False,
                        deletable=delete_permission,
                        csv=False,
                        create=False,
                        editable=False,
                        maxtextlengths=maxtextlengths,
                        orderby=~csap.SubscriptionYear | ~csap.SubscriptionMonth,
                        field_id=db.customers_subscriptions_alt_prices.id,
                        ui = grid_ui)
    grid.element('.web2py_counter', replace=None) # remove the counter
    grid.elements('span[title=Delete]', replace=None) # remove text from delete button

    ## add form
    add = ''
    add_permission = auth.has_membership(group_id='Admins') or \
                     auth.has_permission('create',
                                         'customers_subscriptions_alt_prices')
    if add_permission:
        result = subscription_alt_prices_get_add_form(cuID, csID)
        form = result['form']
        submit = result['submit']
        modal_class = generate_password()
        button_text = XML(SPAN(SPAN(_class='fa fa-plus'), ' ',
                          T("Add")))
        result = os_gui.get_modal(button_text=button_text,
                                  button_title=T("Add alternative price"),
                                  modal_title=T("Add alternative price"),
                                  modal_content=form,
                                  modal_footer_content=submit,
                                  modal_class=modal_class,
                                  button_class='btn-sm')

        add = SPAN(result['button'], result['modal'])

    menu = subscription_edit_get_menu(cuID, csID, request.function)

    back = subscription_edit_get_back(cuID)

    return dict(content=grid,
                menu=menu,
                back=back,
                add=add,
                form_add=form,
                modal_class=modal_class,
                left_sidebar_enabled=True)


def subscription_alt_prices_get_add_form(cuID, csID, full_width=True):
    """
        Returns add form for an invoice
    """
    csap = db.customers_subscriptions_alt_prices
    csap.customers_subscriptions_id.default = csID

    crud.messages.submit_button = T("Save")
    crud.messages.record_created = T("Saved")
    crud.settings.create_next = URL('customers', 'subscription_alt_prices',
                                    vars={'cuID':cuID,
                                          'csID':csID})
    form = crud.create(csap)

    form_element = form.element('form')
    form['_id'] = 'MainForm'

    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = "MainForm"

    submit = form.element('input[type=submit]')

    if full_width:
        # make the inputs in the table full width
        table = form.element('table')
        table['_class'] = 'full-width'

    return dict(form=form, submit=submit)


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('update', 'customers_suscriptions_alt_prices'))
def subscription_alt_price_repeat():
    csapID = request.vars['csapID']
    row = db.customers_subscriptions_alt_prices(csapID)

    cs = db.customers_subscriptions(row.customers_subscriptions_id)

    year = row.SubscriptionYear
    month = row.SubscriptionMonth

    if month == 12:
        month = 1
        year += 1
    else:
        month += 1
    db.customers_subscriptions_alt_prices.insert(customers_subscriptions_id=row.customers_subscriptions_id,
                                                 SubscriptionYear=year,
                                                 SubscriptionMonth=month,
                                                 Amount=row.Amount,
                                                 Description=row.Description,
                                                 Note=row.Note)

    redirect(URL("customers", 'subscription_alt_prices', vars={'csID':row.customers_subscriptions_id,
                                                               'cuID':cs.auth_customer_id}))


@auth.requires_login()
def subscription_alt_prices_edit():
    """
        Edit page for subscription alternative price
    """
    csID = request.vars['csID']
    cuID  = request.vars['cuID']
    csapID = request.vars['csapID']
    response.view = 'general/tabs_menu.html'

    customer = Customer(cuID)
    cs = CustomerSubscription(csID)
    response.title = customer.get_name()
    response.subtitle = subscription_edit_get_subtitle(csID)

    csap = db.customers_subscriptions_alt_prices

    return_url = URL('subscription_alt_prices', vars={'cuID':cuID,
                                                      'csID':csID})

    crud.messages.submit_button = T("Save")
    crud.messages.record_updated = T("Saved")
    crud.settings.update_next = return_url
    crud.settings.update_deletable = False
    form = crud.update(csap, csapID)

    content_back = os_gui.get_button('back_bs', return_url)
    content = DIV(content_back, form)

    menu = subscription_edit_get_menu(cuID, csID, 'subscription_alt_prices')

    back = subscription_edit_get_back(cuID)

    return dict(content=content,
                menu=menu,
                back=back,
                left_sidebar_enabled=True)


def subscription_edit_get_subtitle(csID):
    """
        Returns subtitle for subscription edit pages
    """
    cs = CustomerSubscription(csID)

    return SPAN(T("Edit subscription"), ' ', cs.name)


def subscription_edit_get_back(cuID):
    """
        Returns back button for customer subscription edit pages
    """
    return os_gui.get_button('back',
        URL('subscriptions', vars={'cuID':cuID}),
        _class='')


def subscription_edit_get_menu(cuID, csID, page):
    """
        Returns submenu for subscription edit pages
    """
    vars = { 'cuID':cuID,
             'csID':csID }

    pages = []

    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('update', 'customers_subscriptions'):
        pages.append(['subscription_edit',
                      SPAN(os_gui.get_fa_icon('fa-edit'), ' ', T("Edit")),
                      URL('subscription_edit', vars=vars)])

    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('update', 'customers_subscriptions_paused'):
        pages.append(['subscription_pauses',
                      SPAN(os_gui.get_fa_icon('fa-pause'), ' ', T("Pauses")),
                      URL('subscription_pauses', vars=vars)])

    if auth.has_membership(group_id='Admins') or \
            auth.has_permission('update', 'customers_subscriptions_blocked'):
        pages.append(['subscription_blocks',
                      SPAN(os_gui.get_fa_icon('fa-ban'), ' ', T("Blocks")),
                      URL('subscription_blocks', vars=vars)])

    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('read', 'invoices'):
        pages.append(['subscription_invoices',
                      SPAN(os_gui.get_fa_icon('fa-file-o'), ' ', T("Invoices")),
                      URL('subscription_invoices', vars=vars)])

    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('update', 'customers_subscriptions_alt_prices'):
        pages.append(['subscription_alt_prices',
                      SPAN(os_gui.get_fa_icon('fa-random'), ' ', T("Alt. Prices")),
                      URL('subscription_alt_prices', vars=vars)])

    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('read', 'customers_subscriptions_credits'):
        pages.append(['subscription_credits',
                      SPAN(os_gui.get_fa_icon('fa-check-square-o'), ' ', T("Credits")),
                      URL('subscription_credits', vars=vars)])


    return os_gui.get_submenu(pages, page, horizontal=True, htype='tabs')


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'customers_subscriptions'))
def subscriptions():
    """
        Lists subscriptions for a customer
        request.vars['cuID'] is expected to be the customersID
    """
    customers_id = request.vars['cuID']
    response.view = 'customers/edit_general.html'

    row = db.auth_user(customers_id)
    response.title = row.display_name
    response.subtitle = T("Subscriptions")

    header = THEAD(TR(TH('#'),
                      TH(db.customers_subscriptions.school_subscriptions_id.label),
                      TH(db.customers_subscriptions.Startdate.label),
                      TH(db.customers_subscriptions.Enddate.label),
                      TH(db.customers_subscriptions.payment_methods_id.label),
                      TH(db.customers_subscriptions.Note.label),
                      TH(T('Pauses')),
                      TH(T('Blocks')),
                      TH(T('Credits')),
                      TH(), # membership warning (if any)
                      TH()) # buttons
                   )

    table = TABLE(header, _class='table table-hover table-striped')

    query = (db.customers_subscriptions.auth_customer_id == customers_id)
    rows = db(query).select(db.customers_subscriptions.id,
                            db.customers_subscriptions.auth_customer_id,
                            db.customers_subscriptions.school_subscriptions_id,
                            db.customers_subscriptions.Startdate,
                            db.customers_subscriptions.Enddate,
                            db.customers_subscriptions.payment_methods_id,
                            db.customers_subscriptions.Note,
                            orderby=~db.customers_subscriptions.Startdate)

    for i, row in enumerate(rows):
        repr_row = list(rows[i:i + 1].render())[0]

        delete_permission = auth.has_membership(group_id='Admins') or \
                            auth.has_permission('delete', 'customers_subscriptions')

        delete = ''
        if delete_permission:
            confirm_msg = T("Really delete this subscription?")
            onclick_del = "return confirm('" + confirm_msg + "');"
            delete = os_gui.get_button('delete_notext',
                                        URL('subscription_delete', vars={'cuID': customers_id,
                                                                         'csID': row.id}),
                                        onclick=onclick_del,
                                       _class='pull-right')

        edit = subscriptions_get_link_edit(row)


        tr = TR(TD(row.id),
                TD(repr_row.school_subscriptions_id),
                TD(repr_row.Startdate),
                TD(repr_row.Enddate),
                TD(repr_row.payment_methods_id),
                TD(repr_row.Note),
                TD(subscriptions_get_link_latest_pauses(row)),
                TD(subscriptions_get_link_latest_blocks(row)),
                TD(subscriptions_get_link_credits(row)),
                TD(subscriptions_get_link_membership_check(row)),
                TD(delete, edit))

        table.append(tr)

    add = ''
    if ( auth.has_membership(group_id='Admins') or
         auth.has_permission('create', 'customers_subscriptions') ):
        add_url = URL('subscription_add', vars={'cuID':customers_id})
        add = os_gui.get_button('add', add_url, T("Add a new subscription"), btn_size='btn-sm', _class='pull-right')

    content = DIV(table)

    back = edit_get_back()
    menu = customers_get_menu(customers_id, request.function)

    return dict(content=content, menu=menu, back=back, tools=add)


def subscriptions_get_link_edit(row):
    """
        Returns drop down link for subscriptions
    """
    vars = {'cuID': row.auth_customer_id,
            'csID': row.id}

    links = []

    permission = ( auth.has_membership(group_id='Admins') or
                   auth.has_permission('update', 'customers_subscriptions') )
    if permission:
        link_edit = A((os_gui.get_fa_icon('fa-pencil'), T('Edit')),
                      _href=URL('subscription_edit', vars=vars))
        links.append(link_edit)

    permission = ( auth.has_membership(group_id='Admins') or
                   auth.has_permission('update', 'customers_subscriptions_paused') )
    if permission:
        link_pauses = A((os_gui.get_fa_icon('fa-pause'), T('Pauses')),
                        _href=URL('subscription_pauses', vars=vars))
        links.append(link_pauses)

    # Blocked subscriptions
    permission = (auth.has_membership(group_id='Admins') or
                  auth.has_permission('update', 'customers_subscriptions_blocked'))

    if permission:
        link_blocks = A((os_gui.get_fa_icon('fa-ban'), T('Blocks')),
                        _href=URL('subscription_blocks', vars=vars))
        links.append(link_blocks)

    # Invoices
    permission = ( auth.has_membership(group_id='Admins') or
                   auth.has_permission('read', 'invoices') )
    if permission:
        link_invoices = A((os_gui.get_fa_icon('fa-file-o'), ' ', T('Invoices')),
                          _href=URL('subscription_invoices', vars=vars))

        links.append(link_invoices)

    permission = ( auth.has_membership(group_id='Admins') or
                   auth.has_permission('update', 'customers_subscriptions_alt_prices') )

    if permission:
        link = A((os_gui.get_fa_icon('fa-random'), ' ', T('Alt. prices')),
                        _href=URL('subscription_alt_prices', vars=vars))
        links.append(link)


    permission = ( auth.has_membership(group_id='Admins') or
                   auth.has_permission('read', 'customers_subscriptions_credits') )
    if permission:
        link_credits = A((os_gui.get_fa_icon('fa-check-square-o'), ' ', T("Credits")),
                         _href=URL('subscription_credits', vars=vars))
        links.append(link_credits)

    menu = os_gui.get_dropdown_menu(
        links=links,
        btn_text='',
        btn_size='btn-sm',
        btn_icon='pencil',
        menu_class='btn-group pull-right')

    return menu


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'customers_subscriptions_credits'))
def subscription_credits_month_expired():
    """
        Page to list given credits for a selected month for all customers
    """
    # process request.vars
    subscription_credits_month_set_date()

    # title and subtitle
    response.title = T('Customers')
    response.subtitle = T('Subscription credits')
    response.view = 'customers/subscription_credits_month.html'

    # get from
    result = subscription_credits_month_get_form(
        session.customers_subscription_credits_month,
        session.customers_subscription_credits_year,
        current_url = request.function
    )
    response.subtitle += ' - ' + result['subtitle']

    form = result['form']
    month_chooser = os_gui.get_month_chooser(
        request.function,
        'subscription_credits_set_month',
        session.customers_subscription_credits_year,
        session.customers_subscription_credits_month
    )

    content = subscription_credits_month_get_content(expired=True)

    expire = ''
    permission = (auth.has_membership(group_id='Admins') or
                  auth.has_permission('update', 'customers_subscriptions_credits'))
    if permission:
        expire = os_gui.get_button('noicon',
                                   URL('subscription_credits_month_expire_credits'),
                                   title=T('Expire'),
                                   btn_class='btn-primary')

    back = os_gui.get_button('back', URL('customers', 'index'))
    menu = subscription_credits_month_get_menu(request.function)

    return dict(content=content,
                form=form,
                current=result['current'],
                header_tools=expire,
                month_chooser=month_chooser,
                menu=menu,
                back=back)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'customers_subscriptions_credits'))
def subscription_credits_month_expire_credits():
    """
        Expire credits on the current day
    """
    from openstudio.os_customers_subscriptions_credits import CustomersSubscriptionsCredits

    csch = CustomersSubscriptionsCredits()
    sub_credits_expired = csch.expire_credits(TODAY_LOCAL)

    session.flash = T('Expired credits for') + ' ' + unicode(sub_credits_expired) + ' ' + T('subscriptions')

    redirect(URL('subscription_credits_month_expired'))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'customers_subscriptions_credits'))
def subscription_credits_month():
    """
        Page to list given credits for a selected month for all customers
    """
    # process request.vars
    subscription_credits_month_set_date()

    # title and subtitle
    response.title = T('Customers')
    response.subtitle = T('Subscription credits')

    # get from
    result = subscription_credits_month_get_form(
        session.customers_subscription_credits_month,
        session.customers_subscription_credits_year,
        request.function
    )
    response.subtitle += ' - ' + result['subtitle']

    form = result['form']
    month_chooser = os_gui.get_month_chooser(
        request.function,
        'subscription_credits_set_month',
        session.customers_subscription_credits_year,
        session.customers_subscription_credits_month
    )

    add = os_gui.get_button('add',
                            URL('automation_customer_subscriptions', 'index'),
                            btn_class='btn-primary')
    content = subscription_credits_month_get_content()
    back = os_gui.get_button('back', URL('customers', 'index'))
    menu = subscription_credits_month_get_menu(request.function)

    return dict(content=content,
                form=form,
                current=result['current'],
                header_tools=add,
                month_chooser=month_chooser,
                menu=menu,
                back=back)


def subscription_credits_month_get_content(expired=False):
    """
        :param expired: Boolean
        Get list of credits of this month, default is added credits, but expired credits are returned when
        expired boolean is True
    """
    first_day = datetime.date(session.customers_subscription_credits_year,
                              session.customers_subscription_credits_month,
                              1)
    last_day = get_last_day_month(first_day)

    left = [
        db.customers_subscriptions.on(db.customers_subscriptions_credits.customers_subscriptions_id ==
                                      db.customers_subscriptions.id),
        db.auth_user.on(db.customers_subscriptions.auth_customer_id == db.auth_user.id),
        db.school_subscriptions.on(db.customers_subscriptions.school_subscriptions_id ==
                                   db.school_subscriptions.id)
    ]

    if expired:
        query = (db.customers_subscriptions_credits.MutationType == 'sub') & \
                (db.customers_subscriptions_credits.Expiration == True) & \
                (db.customers_subscriptions_credits.MutationDateTime <= last_day) & \
                (db.customers_subscriptions_credits.MutationDateTime >= first_day)
    else:
        query = (db.customers_subscriptions_credits.MutationType == 'add') & \
                (db.customers_subscriptions_credits.MutationDateTime <= last_day) & \
                (db.customers_subscriptions_credits.MutationDateTime >= first_day)

    rows = db(query).select(db.auth_user.id,
                            db.auth_user.display_name,
                            db.auth_user.thumbsmall,
                            db.auth_user.birthday,
                            db.auth_user.trashed,
                            db.school_subscriptions.Name,
                            db.customers_subscriptions.id,
                            db.customers_subscriptions.Startdate,
                            db.customers_subscriptions.school_subscriptions_id,
                            db.customers_subscriptions_credits.MutationDateTime,
                            db.customers_subscriptions_credits.MutationType,
                            db.customers_subscriptions_credits.MutationAmount,
                            db.customers_subscriptions_credits.SubscriptionYear,
                            db.customers_subscriptions_credits.SubscriptionMonth,
                            left=left,
                            orderby=db.auth_user.display_name
                            )

    header = THEAD(TR(
        TH(),
        TH(T('Customer')),
        TH(T('Subscription')),
        TH(T('Subscription start')),
        TH(T('Mutation Date')),
        TH(T('For')),
        TH(T('Credits')),
        TH(),
        TH(),
    ))

    table = TABLE(header, _class='table table-striped')
    for i, row in enumerate(rows):
        repr_row = list(rows[i:i + 1].render())[0]

        tr = TR(
            TD(repr_row.auth_user.thumbsmall, _class='os-customer_image_td'),
            TD(repr_row.auth_user.display_name),
            TD(repr_row.customers_subscriptions.school_subscriptions_id),
            TD(repr_row.customers_subscriptions.Startdate),
            TD(repr_row.customers_subscriptions_credits.MutationDateTime),
            TD(repr_row.customers_subscriptions_credits.SubscriptionMonth, ' ',
               repr_row.customers_subscriptions_credits.SubscriptionYear),
            TD(repr_row.customers_subscriptions_credits.MutationAmount),
            TD(repr_row.customers_subscriptions_credits.MutationType),
            TD(os_gui.get_button('edit',
                                 URL('subscription_credits', vars={'cuID':row.auth_user.id,
                                                                   'csID':row.customers_subscriptions.id}),
                                 _class="pull-right"))
        )

        table.append(tr)

    return table


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'customers_subscriptions_credits'))
def subscription_credits_set_month():
    """
        Sets the session variables for customers_subscriptions_credits year and month
    """
    year  = request.vars['year']
    month = request.vars['month']
    back  = request.vars['back']

    session.customers_subscription_credits_year  = int(year)
    session.customers_subscription_credits_month = int(month)


    if back:
        redirect(back)
    else:
        redirect(URL('subscription_credits_month'))


def subscription_credits_month_set_date(var=None):
    """
        Set session variable for date vars
    """
    today = TODAY_LOCAL
    if 'year' in request.vars:
        year = int(request.vars['year'])
    elif not session.customers_subscription_credits_year is None:
        year = session.customers_subscription_credits_year
    else:
        year = today.year
    session.customers_subscription_credits_year = year
    if 'month' in request.vars:
        month = int(request.vars['month'])
    elif not session.customers_subscription_credits_month is None:
        month = session.customers_subscription_credits_month
    else:
        month = today.month
    session.customers_subscription_credits_month = month


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'customers_subscriptions_credits'))
def subscription_credits_month_show_current():
    """
        Show current
    """
    session.customers_subscription_credits_year  = int(TODAY_LOCAL.year)
    session.customers_subscription_credits_month = int(TODAY_LOCAL.month)

    redirect(URL('subscription_credits_month'))


def subscription_credits_month_get_form(month, year, current_url, _class='col-md-4'):
    """
        Get month chooser form for subscription_credits_month
    """
    months = get_months_list()

    for m in months:
        if m[0] == month:
            month_title = m[1]
    subtitle = month_title + " " + unicode(year)

    form = SQLFORM.factory(
        Field('month',
              requires=IS_IN_SET(months, zero=None),
              default=month,
              label=T("")),
        Field('year', 'integer',
              default=year,
              label=T("")),
        submit_button=T("Run report")
    )
    form.attributes['_name'] = 'form_select_date'
    form.attributes['_class'] = 'overview_form_select_date'

    input_month = form.element('select[name=month]')
    input_month.attributes['_onchange'] = "this.form.submit();"

    input_year = form.element('input[name=year]')
    input_year.attributes['_onchange'] = "this.form.submit();"
    input_year.attributes['_type'] = 'number'

    form.element('input[name=year]')

    result = set_form_id_and_get_submit_button(form, 'MainForm')
    form = result['form']
    submit = result['submit']

    ## Show current
    show_current = A(T("Current month"),
                     _href=URL(current_url),
                     _class='btn btn-default')

    form = DIV(XML('<form id="MainForm" action="#" enctype="multipart/form-data" method="post">'),
               DIV(form.custom.widget.month,
                   form.custom.widget.year,
                   _class=_class),
               form.custom.end,
               _class='row')

    return dict(form=form, subtitle=subtitle, current=show_current, submit=submit)


def subscription_credits_month_get_menu(page=None):
    pages = [
        (['subscription_credits_month', T('Added credits'), URL('customers', 'subscription_credits_month')]),
        (['subscription_credits_month_expired',
          T('Expired credits'),
          URL('customers', 'subscription_credits_month_expired')]),
    ]

    return os_gui.get_submenu(pages,
                              page,
                              horizontal=True,
                              htype='tabs')


# @auth.requires(auth.has_membership(group_id='Admins') or \
#                auth.has_permission('create', 'customers_subscriptions_credits'))
# def subscription_credits_month_add_confirm():
#     """
#         Show confirmation page before adding credits for a month
#     """
#     response.title = T('Customers')
#     response.subtitle = T('Subscription credits - Add confirmation')
#     response.view = 'general/only_content_header_footer.html'
#
#     # Set default values for session vars if not set
#     subscription_credits_month_set_date()
#
#     year = session.customers_subscription_credits_year
#     month = session.customers_subscription_credits_month
#     date = datetime.date(year, month, 1)
#
#     box_title = SPAN(T('Add subscription credits for'), ' ', date.strftime('%B %Y'))
#
#     content = DIV(
#
#         P(T("This operation will only add credits for subscriptions where credits haven't been granted yet for this month.")),
#         P(T("Paused subscriptions and subscriptions where credits are already granted for this month will be skipped."))
#     )
#
#     buttons = SPAN(
#         os_gui.get_button('noicon',
#                           URL('customers', 'subscription_credits_month_add'),
#                           title=T('Continue'),
#                           btn_class='btn-primary'),
#         os_gui.get_button('noicon',
#                           URL('customers', 'subscription_credits_month'),
#                           title=T('Cancel'),
#                           btn_class='btn-link')
#     )
#
#
#     back = os_gui.get_button('back', URL('customers', 'subscription_credits_month'))
#
#     return dict(content=content,
#                 box_title=box_title,
#                 footer=buttons,
#                 back=back)

#
# @auth.requires(auth.has_membership(group_id='Admins') or \
#                auth.has_permission('create', 'customers_subscriptions_credits'))
# def subscription_credits_month_add():
#     """
#         Add credits for subscriptions in selected month
#     """
#     from openstudio.os_customers_subscriptions_credits import CustomersSubscriptionsCredits
#
#     year = session.customers_subscription_credits_year
#     month = session.customers_subscription_credits_month
#
#     csch = CustomersSubscriptionsCredits()
#     added = csch.add_credits(year, month)
#
#     session.flash = T("Added subscription credits for") + ' ' + unicode(added) + ' ' + T('customers') + '.'
#
#     redirect(URL('subscription_credits_month'))


def payments_get_submenu(page, cuID):
    """
        Returns submenu for account pages
    """
    vars = {'cuID':cuID}

    pages = []

    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('read', 'customers_payments_info'):
        pages.append(
            [
                'bankaccount',
                T('Bank account'),
                URL('bankaccount', vars=vars)
            ]
        )

        pages.append(
            [
                'direct_debit_extra',
                T('Direct debit extra'),
                URL('direct_debit_extra', vars=vars)
            ]
        )

        pages.append(
            [
                'mollie_mandates',
                T('Mollie mandates'),
                URL('mollie_mandates', vars=vars)
            ]
        )

    horizontal = True

    return os_gui.get_submenu(pages, page, horizontal=horizontal, htype='tabs')



def payments_delete_payment_info(form):
    page = redirect(URL('payments', vars={'cuID':customers_id}))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'customers_payments_info'))
def mollie_mandates():
    """
        Lists mollie mandates for customer
    """
    cuID = request.vars['cuID']
    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = T("Finance")
    response.view = 'general/tabs_menu.html'

    # back button
    back = edit_get_back()

    customer = Customer(cuID)
    mollie_mandates = customer.get_mollie_mandates_formatted()
   
   
    menu = customers_get_menu(cuID, request.function)
    submenu = payments_get_submenu(request.function, cuID)

    content = DIV(submenu, BR(), mollie_mandates)


    return dict(content=content,
                menu=menu,
                back=back)


@auth.requires_login()
def bankaccount():
    """
        Lists bank account info
    """
    cuID = request.vars['cuID']
    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = T("Finance")
    response.view = 'general/tabs_menu.html'

    # back button
    back = edit_get_back()

    # payment_info
    db.customers_payment_info.id.readable=False
    db.customers_payment_info.auth_customer_id.readable=False

    query = (db.customers_payment_info.auth_customer_id == cuID)
    count = db(query).count()

    if not count:
        db.customers_payment_info.insert(
            auth_customer_id = cuID
        )

    rows = db(query).select(db.customers_payment_info.ALL)
    row = rows.first()

    return_url = bankaccount_get_returl_url(cuID)

    crud.messages.submit_button = T("Save")
    crud.messages.record_updated = T("Saved")
    crud.settings.formstyle = "bootstrap3_stacked"
    crud.settings.update_next = return_url
    crud.settings.update_deletable = False
    crud.settings.update_onaccept = [bankaccount_onaccept]
    form = crud.update(db.customers_payment_info, row.id)

    result = set_form_id_and_get_submit_button(form, 'MainForm')
    form = result['form']
    submit = result['submit']

    menu = customers_get_menu(cuID, request.function)
    submenu = payments_get_submenu(request.function, cuID)

    mandates = DIV(
        customer.get_payment_info_mandates(formatted=True),
        _class="col-md-12"
    )
    content = DIV(submenu, BR(), form)

    eo_authorized = get_sys_property('exact_online_authorized')
    if auth.has_membership(group_id='Admins') and eo_authorized:
        if row.exact_online_bankaccount_id:
            eo_message = SPAN(
                os_gui.get_fa_icon('fa-check'), ' ',
                T("This bank account is linked to Exact Online"),
                _class='text-green'
            )
        else:
            eo_message = SPAN(
                os_gui.get_fa_icon('fa-ban'), ' ',
                T("This bank account is not linked to Exact Online"),
                _class='text-red'
            )

        content.append(DIV(
            A(os_gui.get_fa_icon('fa-pencil'), ' ',
              T("Edit Exact Online link"),
              _href=URL('bankaccount_exact_online', vars={'cuID':cuID, 'cpiID':row.id}),
              _class='pull-right'),
            eo_message
        ))


    add_mandate = ''
    query = (db.customers_payment_info_mandates.customers_payment_info_id == row.id)
    if not db(query).count():
        add_mandate = os_gui.get_button(
            'noicon',
            URL('bankaccount_mandate_add', vars={'cuID':cuID, 'cpiID':row.id}),
            title=T("Add mandate"),
            btn_size='',
        )


    return dict(content=content,
                content_extra=mandates,
                menu=menu,
                back=back,
                tools=SPAN(add_mandate, submit))


def bankaccount_onaccept(form):
    """
    :param form: crud form for db.customers_payment_info
    :return:
    """
    from openstudio.os_customers_payment_info import OsCustomersPaymentInfo

    cpiID = form.vars.id
    cpi = OsCustomersPaymentInfo(cpiID)
    cpi.on_update()


def bankaccount_get_returl_url(customers_id):
    """
        Returns the return url for payment_info_add and payment_info_edit
    """
    return URL('bankaccount', vars={'cuID':customers_id})


@auth.requires(auth.has_membership(group_id='Admins'))
def bankaccount_exact_online():
    """
    Update Exact Online link for Payment info
    """
    cpiID = request.vars['cpiID']
    cuID = request.vars['cuID']
    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = T("Finance")
    response.view = 'general/tabs_menu.html'

    # back button
    back = os_gui.get_button(
        'back',
        URL('bankaccount', vars={'cuID': cuID})
    )

    # payment_info
    menu = customers_get_menu(cuID, request.function)
    submenu = payments_get_submenu('bankaccount', cuID)

    # Customer EO account code

    eo_account_ID = customer.row.exact_online_relation_id
    if not eo_account_ID:
        content = T("Unable to link bank account, this customer isn't linked to an Exact Online relation")
    else:
        from openstudio.os_customers_payment_info import OsCustomersPaymentInfo

        accounts = customer.exact_online_get_bankaccounts()

        search_result = ''
        if not len(accounts):
            search_result = T("No bank accounts found for this relation in Exact Online")
        else:
            header = THEAD(TR(
                TH(T('Exact Online Relation Name')),
                TH(T('Exact Online Bank Account')),
                TH()
            ))
            search_result = TABLE(header, _class="table table-striped table-hover")

            for account in accounts:
                btn_link = os_gui.get_button(
                    'noicon',
                    URL('bankaccount_exact_online_link_bankaccount',
                        vars={'cuID': cuID,
                              'cpiID': cpiID,
                              'eoID': account['ID']}),
                    title=T("Link to this bank account"),
                    _class='pull-right'
                )

                search_result.append(TR(
                    TD(account['AccountName']),
                    TD(account['BankAccount']),
                    TD(btn_link)
                ))


        current_link = T("Please select a bank account listed on the right.")

        cpi = OsCustomersPaymentInfo(cpiID)
        linked_account = cpi.exact_online_get_bankaccount()

        if linked_account:
            current_link = DIV(
                T("This customer is linked to the following Exact Online Bank Account"), BR(), BR(),
                T("OpenStudio bank account: %s" % (cpi.row.AccountNumber)), BR(), BR(),
                T("Exact Online bank account: %s" % (linked_account[0]['BankAccount'])),  BR(),
                T("Exact Online relation name: %s" % (linked_account[0]['AccountName'])), BR(), BR(),
                T("To link this bank account to another Exact Online bank account, please select one from the list on the right.")
            )

        display = DIV(
            DIV(
                H4(T("Current link")),
                current_link,
                _class='col-md-6'
            ),
            DIV(
                H4(T('Bank accounts for this relation in Exact Online')),
                search_result,
                _class='col-md-6'
            ),
            _class='row'
        )

        content = DIV(submenu, BR(), display)

    return dict(
        content=content,
        menu=menu,
        back=back,
        tools=''
    )


@auth.requires(auth.has_membership(group_id='Admins'))
def bankaccount_exact_online_link_bankaccount():
    """
    Link exact online relation to OpenStudio customer
    """
    from openstudio.os_customers_payment_info import OsCustomersPaymentInfo

    cuID = request.vars['cuID']
    cpiID = request.vars['cpiID']
    eoID = request.vars['eoID']

    cpi = OsCustomersPaymentInfo(cpiID)
    message = cpi.exact_online_link_to_bankaccount(eoID)

    session.flash = message
    redirect(URL('bankaccount_exact_online', vars={'cuID': cuID,
                                                   'cpiID': cpiID}))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('create', 'customers_payments_info_mandates'))
def bankaccount_mandate_add():
    """
    Page to add a mandate
    :return:
    """
    import uuid

    cuID = request.vars['cuID']
    cpiID = request.vars['cpiID']

    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = T("Finance")
    response.view = 'general/tabs_menu.html'

    db.customers_payment_info_mandates.customers_payment_info_id.default = cpiID

    return_url = bankaccount_get_returl_url(cuID)

    crud.messages.submit_button = T("Save")
    crud.messages.record_updated = T("Saved")
    crud.settings.formstyle = "bootstrap3_stacked"
    crud.settings.create_next = return_url
    crud.settings.create_onaccept = [bankaccount_mandate_on_create]
    form = crud.create(db.customers_payment_info_mandates)

    result = set_form_id_and_get_submit_button(form, 'MainForm')
    form = result['form']
    submit = result['submit']

    menu = customers_get_menu(cuID, request.function)
    submenu = payments_get_submenu(request.function, cuID)

    content = DIV(
        submenu, BR(),
        H3(T("New mandate")),
        form
    )

    back = os_gui.get_button(
        'back',
        return_url
    )

    return dict(
        content=content,
        menu=menu,
        back=back,
        save=submit
    )


def bankaccount_mandate_on_create(form):
    """
    :param form: crud form for db.customers_payment_info_mandates
    :return:
    """
    from openstudio.os_customers_payment_info_mandate import OsCustomersPaymentInfoMandate

    cpimID = form.vars.id
    cpim = OsCustomersPaymentInfoMandate(cpimID)
    cpim.on_create()


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('create', 'customers_payments_info_mandates'))
def bankaccount_mandate_delete():
    """
    Delete bankaccount mandate
    """
    from openstudio.os_customers_payment_info_mandate import OsCustomersPaymentInfoMandate

    cuID = request.vars['cuID']
    cpimID = request.vars['cpimID']

    cpim = OsCustomersPaymentInfoMandate(cpimID)
    cpim.on_delete()

    query = (db.customers_payment_info_mandates.id == cpimID)
    db(query).delete()

    redirect(bankaccount_get_returl_url(cuID))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'customers_payments_info'))
def direct_debit_extra():
    """
        List direct debit extra lines
    """
    customers_id = request.vars['cuID']
    customer = Customer(customers_id)
    response.title = customer.get_name()
    response.subtitle = T("Finance")
    response.view = 'general/tabs_menu.html'

    # back button
    back = edit_get_back()

    # alternative payments
    db.alternativepayments.id.readable=False
    db.alternativepayments.auth_customer_id.readable=False
    query = (db.alternativepayments.auth_customer_id == customers_id)
    db.alternativepayments.auth_customer_id.default = customers_id

    fields = [ db.alternativepayments.PaymentYear,
               db.alternativepayments.PaymentMonth,
               db.alternativepayments.Amount,
               db.alternativepayments.payment_categories_id,
               db.alternativepayments.Description ]

    links = [lambda row: os_gui.get_button('edit',
                                    URL('alternativepayment_edit',
                                        args=[customers_id, row.id])),
            lambda row: A(SPAN(_class="buttontext button", _title=T("Repeat")),
                          SPAN(_class="glyphicon glyphicon-repeat"),
                          " ", T("Repeat"),
                          _class="btn btn-default",
                          _href=URL('alternativepayment_repeat',
                                    vars=dict(apID=row.id)),
                          _title=T("Repeat")) ]
    maxtextlengths = {'alternativepayments.Description' : 30}

    delete_permission = auth.has_membership(group_id='Admins') or \
                        auth.has_permission('delete', 'alternativepayments')

    ap_grid = SQLFORM.grid(query,
                           fields=fields,
                           links=links,
                           maxtextlengths=maxtextlengths,
                           create=False,
                           details=False,
                           editable=False,
                           csv=False,
                           searchable=False,
                           deletable=delete_permission,
                           ui=grid_ui,
                           orderby=~db.alternativepayments.PaymentYear|\
                                   ~db.alternativepayments.PaymentMonth,
                           field_id=db.alternativepayments.id)
    add = os_gui.get_button('add', URL('alternativepayment_add', args=[customers_id]))
    ap_grid.element('.web2py_counter', replace=None) # remove the counter
    ap_grid.elements('span[title=Delete]', replace=None) # remove text from delete button

    menu = customers_get_menu(customers_id, 'bankaccount')
    submenu = payments_get_submenu(request.function, customers_id)

    content = DIV(submenu, BR(), ap_grid)

    return dict(content=content,
                menu=menu,
                tools=add,
                back=back)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'customers_notes_backoffice') or \
               auth.has_permission('read', 'customers_notes_teachers'))
def notes():
    """
        Lists all notes for the backoffice
        request.vars['note_type'] can be 2 values
            'backoffice' for a backoffice note
            'teacher' for a teacher note
    """
    cnID = request.vars['cnID']
    cuID = request.vars['cuID']
    note_type = request.vars['note_type']

    customer = Customer(cuID)
    response.title = customer.row.display_name
    response.subtitle = T("Profile")
    response.view = 'general/tabs_menu.html'

    sub_subtitle = SPAN(T("Notes"), XML(' &bull; '))
    if note_type is None:
        db.customers_notes.BackofficeNote.default = True

    if note_type == 'backoffice':
        sub_subtitle.append(T("Back office"))
        db.customers_notes.BackofficeNote.default = True

    if note_type == 'teachers':
        sub_subtitle.append(T("Teachers"))
        db.customers_notes.TeacherNote.default = True

    permission_edit = (
            auth.has_membership(group_id='Admins') or
            auth.has_permission('update', 'customers_notes')
    )
    permission_delete = (
            auth.has_membership(group_id='Admins') or
            auth.has_permission('delete', 'customers_notes')
    )


    notes = customer.get_notes_formatted(
        note_type,
        permission_edit,
        permission_delete
    )

    # form
    form = ''
    form_title = ''

    if not cnID:
        perm = auth.has_membership(group_id='Admins') or \
               auth.has_permission('create', 'customers_notes')
        if perm:
            form = notes_get_add()
            form_title = H4(T('Add a new note'))
    else:
        perm = auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'customers_notes')
        if perm:
            form = notes_get_edit(cnID, cuID, note_type)
            form_title = H4(T('Edit note'))

    content = DIV(
        H4(sub_subtitle),
        notes,
        form_title,
        form,
    )

    menu = customers_get_menu(cuID, 'general')
    back = os_gui.get_button(
        'back',
        URL('edit', args=[cuID])
    )

    return dict(
        content=content,
        menu=menu,
        back=back,
    )


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'customers_notes_backoffice') or \
               auth.has_permission('read', 'customers_notes_teachers'))
def note_latest():
    """
        Lists all notes for the backoffice
        request.vars['note_type'] can be 2 values
            'backoffice' for a backoffice note
            'teacher' for a teacher note
    """
    from openstudio.os_customer import Customer

    customers_id = request.vars['cuID']
    note_type = request.vars['note_type']
    latest = request.vars['latest']
    latest_length = request.vars['latest_length']
    try:
        latest_length = int(latest_length)
    except:
        latest_length = 50 # set default

    customer = Customer(customers_id)
    rows = customer.get_notes(note_type=note_type)

    if not rows:
        return ''

    latest_note = rows.first()

    return DIV(XML(max_string_length(latest_note.Note.replace('\n','<br>'),
                                     latest_length)))


# @auth.requires_login()
# def note_edit():
#     """
#         Provides an edit page for a note.
#         request.args[0] is expected to be the customers_note_id (cn_id)
#     """
#     cn_id = request.args[0]
#
#     note = db.customers_notes(cn_id)
#     customers_id = note.auth_customer_id
#
#     if note.BackofficeNote:
#         note_type = 'backoffice'
#     elif note.TeacherNote:
#         note_type = 'teachers'
#
#     crud.messages.submit_button = T("Save")
#     crud.messages.record_updated = T('Saved')
#     form = crud.update(db.customers_notes, cn_id)
#
#     form.custom.widget.Note['_class'] += ' form-control'
#
#     form = DIV(form.custom.begin,
#                form.custom.widget.Note,
#                form.custom.submit,
#                form.custom.end,
#                _class='os-customers_notes_edit clear')
#
#     back =  os_gui.get_button('back',
#                       URL('notes', vars={'cuID':customers_id,
#                                          'note_type':note_type}),
#                       _class='left',
#                       cid=request.cid)
#
#     content = DIV(BR(),
#                   back,
#                   BR(),BR(),
#                   form)
#
#     response.js = "setTimeout(function() { $('div.flash').fadeOut(); }, 2500 );"
#
#     return dict(content=content)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('delete', 'customers_notes'))
def note_delete():
    """
        Called as JSON, used to remove a note
    """
    cuID = request.vars['cuID']
    cnID = request.vars['cnID']

    note = db.customers_notes(cnID)
    if note.TeacherNote:
        note_type = 'teachers'
    else:
        note_type = 'backoffice'

    query = (db.customers_notes.id == cnID)
    db(query).delete()

    session.flash = T("Deleted note")
    redirect(URL('customers', 'notes', vars={'cuID': cuID,
                                             'note_type': note_type}))

def notes_get_add(var=None):
    """
        Provides a page to add a note
        request.vars['note_type'] can be 2 values
            'backoffice' for a backoffice note
            'teachers' for a teachers note
    """
    note_type = request.vars['note_type']
    customers_id = request.vars['cuID']

    if note_type is None:
        vars = {'cuID':customers_id}
    else:
        vars = {'cuID':customers_id,
                'note_type':note_type}

    return_url = URL('notes', vars=vars)

    db.customers_notes.auth_customer_id.default = customers_id
    db.customers_notes.auth_user_id.default = auth.user.id

    crud.messages.submit_button = T("Save")
    crud.messages.record_created = T('')
    crud.settings.create_next = return_url
    form = crud.create(db.customers_notes)

    form.custom.widget.Note['_class'] += ' form-control'

    form = DIV(form.custom.begin,
               form.custom.widget.Note,
               form.custom.submit,
               form.custom.end,
               _class='os-customers_notes_edit')

    return form


def notes_get_edit(cnID, cuID, note_type):
    """
        Provides a page to add a note
        request.vars['note_type'] can be 2 values
            'backoffice' for a backoffice note
            'teachers' for a teachers note
    """
    if note_type is None:
        vars = {'cuID': cuID}
    else:
        vars = {'cuID': cuID,
                'note_type': note_type}

    return_url = URL('notes', vars=vars)

    db.customers_notes.auth_customer_id.default = cuID
    db.customers_notes.auth_user_id.default = auth.user.id

    crud.messages.submit_button = T("Save")
    crud.messages.record_updated = T("Saved")
    crud.settings.update_next = return_url
    form = crud.update(db.customers_notes, cnID)

    form.custom.widget.Note['_class'] += ' form-control'

    form = DIV(form.custom.begin,
               form.custom.widget.Note,
               form.custom.submit,
               A(T('Cancel'),
                 _href=return_url,
                 _class='btn btn-link'),
               form.custom.end,
               _class='os-customers_notes_edit')

    return form


def alternativepayment_get_return_url(customers_id):
    """
        Returns return url for alternative payments
    """
    return URL('direct_debit_extra', vars={'cuID':customers_id})


@auth.requires_login()
def alternativepayment_add():
    """
        This function shows an add page for alternative payments
        request.args[0] is expected to be the customers_id
    """
    customers_id = request.args[0]
    customer = Customer(customers_id)
    response.title = customer.get_name()
    response.subtitle = SPAN(T('Finance - Direct debit extra'))
    response.view = 'general/only_content.html'

    session.customers_payments_tab = '#ap'

    db.alternativepayments.auth_customer_id.default = customers_id

    return_url = alternativepayment_get_return_url(customers_id)

    crud.messages.submit_button = T("Save")
    crud.messages.record_created = T("Saved")
    crud.settings.create_next = return_url
    form = crud.create(db.alternativepayments)

    result = set_form_id_and_get_submit_button(form, 'MainForm')
    form = result['form']
    submit = result['submit']

    back = os_gui.get_button('back', return_url)

    return dict(content=form, save=submit, back=back)


@auth.requires_login()
def alternativepayment_edit():
    """
        This function shows an add page for alternative payments
        request.args[0] is expected to be the customers_id
        request.args[1] is expected to be the alternativepaymentsID
    """
    customers_id = request.args[0]
    apID = request.args[1]
    customer = Customer(customers_id)
    response.title = customer.get_name()
    response.subtitle = SPAN(T("Finance - Direct debit extra"))
    response.view = 'general/only_content.html'

    session.customers_payments_tab = '#ap'

    db.alternativepayments.auth_customer_id.default = customers_id

    return_url = alternativepayment_get_return_url(customers_id)

    crud.messages.submit_button = T("Save")
    crud.messages.record_updated = T("Saved")
    crud.settings.update_next = return_url
    crud.settings.update_deletable = False
    form = crud.update(db.alternativepayments, apID)

    result = set_form_id_and_get_submit_button(form, 'MainForm')
    form = result['form']
    submit = result['submit']

    back = os_gui.get_button('back', return_url)

    return dict(content=form, save=submit, back=back)


def alternativepayments_check(form):
    month = form.vars['PaymentMonth']
    year = form.vars['PaymentYear']
    categoryID = form.vars['PaymentCategoryID']

    query = (db.alternativepayments.auth_customer_id==customers_id) & \
            (db.alternativepayments.PaymentYear==year) & \
            (db.alternativepayments.PaymentMonth==month) & \
            (db.alternativepayments.payment_categories_id==categoryID)
    count = db(query).count()
    if count > 1:
        session.flash = T("Only one alternative payment per category per month is allowed.")
        db.rollback() # undo changes, we don't want this data in the db to prevent export issues.


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'workshops'))
def events():
    """
        This function shows a page which lists the workshop attendance for a
        customer.
    """
    cuID = request.vars['cuID']

    session.customers_payment_back = None
    # To redirect back after removing a product
    session.events_ticket_sell_back = 'customers'
    # Invoice back redirects
    session.invoices_edit_back = 'customers_events'
    session.invoices_payment_add_back = 'customers_events'
    # To redirect back here after sending info mail
    session.workshops_ticket_resend_info_mail = 'customers_events'

    session.workshops_payment_back = 'customer'
    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = T("Events")

    rows = customer.get_workshops_rows()

    wh = WorkshopsHelper()
    header = THEAD(TR(TH(_class='workshop_image_th'),
                      TH(T('Event'),
                      TH(T('Invoice')),
                      TH(T('Event info')),
                      TH(),

                      )))
    table = TABLE(header, _class="table table-hover table-striped")
    invoices = Invoices()

    for i, row in enumerate(rows):
        repr_row = list(rows[i:i+1].render())[0]

        wsID     = row.workshops.id
        wsp_cuID = row.workshops_products_customers.id

        # invoice
        if row.invoices.id:
            invoice = invoices.represent_invoice_for_list(
                row.invoices.id,
                repr_row.invoices.InvoiceID,
                repr_row.invoices.Status,
                row.invoices.Status,
                row.invoices.payment_methods_id
            )
        else:
            invoice = ''


        # Event info link
        link_text = T('Send')
        if row.workshops_products_customers.WorkshopInfo:
            link_text = T('Resend')
        resend_link = A(link_text, ' ', T('info mail'),
                        _href=URL('events', 'ticket_resend_info_mail', vars={'wspcID':wsp_cuID}))
        event_info = wh.get_customer_info(wsp_cuID,
                                          wsID,
                                          row.workshops_products_customers.WorkshopInfo,
                                          resend_link)

        permission = auth.has_membership(group_id='Admins') or \
                     auth.has_permission('delete', 'workshops_products_customers')
        if permission:
            confirm_remove_msg = T("Really remove this workshop product?")
            btn_delete = os_gui.get_button(
                'delete_notext',
                URL('events', 'ticket_delete_customer',
                    vars={'wsID'     : wsID,
                          'wsp_cuID' : wsp_cuID}),
                tooltip=T('Remove customer from list'),
                onclick="return confirm('" + confirm_remove_msg + "');" )
        else:
            btn_delete = ''

        # check waitinglist
        if row.workshops_products_customers.Waitinglist:
            waitinglist = os_gui.get_label('danger', T('Waitinglist'))
        else:
            waitinglist = ''

        # check full workshop
        if row.workshops_products.FullWorkshop:
            label_class = 'primary'
        else:
            label_class = 'default'

        # check cancelled
        through_class= ''
        cancelled_label = ''
        title_cancel = T('Cancel customer')
        if row.workshops_products_customers.Cancelled:
            title_cancel = T('Undo cancellation')
            cancelled_label = SPAN(' ', os_gui.get_label('warning',
                                                         T("Cancelled")))
            through_class = 'line-through'

        btn_cancel = os_gui.get_button(
            'cancel_notext',
            URL('events', 'ticket_cancel_customer',
                               vars={'wsID'     : wsID,
                                     'wsp_cuID' : wsp_cuID}),
            tooltip=title_cancel )


        tr = TR(TD(repr_row.workshops.thumbsmall),
                TD(SPAN(A(row.workshops.Name,
                          _href=URL('events', 'tickets',
                                    vars={'wsID':wsID})),
                        _class=' ' + through_class),
                   SPAN(' - ', row.workshops.Startdate,
                        _class='small_font grey'),
                   BR(),
                   os_gui.get_label(label_class, row.workshops_products.Name),
                   ' ', waitinglist, cancelled_label),
                TD(invoice),
                TD(event_info),
                TD(DIV(btn_cancel, btn_delete, _class='btn-group pull-right')))

        table.append(tr)

    # back button
    back = edit_get_back()
    add  = os_gui.get_button('add',
                             URL('event_add', vars={'cuID':cuID}),
                             btn_size='btn-sm',
                             tooltip=T("Add a workshop"))
    menu = customers_get_menu(cuID, request.function)
    loader = os_gui.get_ajax_loader(message=T("Refreshing list..."))

    return dict(content=table,
                menu=menu,
                add=add,
                back=back,
                loader=loader,
                left_sidebar_enabled=True)


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('update', 'workshops_products_customers'))
def event_add():
    """
        Select a workshop to list products from
    """
    response.view = 'general/only_content.html'

    cuID = request.vars['cuID']
    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = SPAN(T("Add event product"), B(' > '),
                             T("List events"))


    # list of workhsops
    grid = events_add_get_list(cuID)

    back = os_gui.get_button('back', URL('events',
                                         vars={'cuID':cuID}))


    return dict(content=grid,
                back=back)


def events_add_get_list(cuID):
    """
        Returns grid of workshops with select button
    """
    query = (db.workshops.Archived == False)

    links = [ lambda row: os_gui.get_button(
                        'noicon',
                        URL('events_add_list_tickets',
                            vars={'wsID' : row.id,
                            'cuID' : cuID}),
                        title=T('Tickets'),
                        tooltip=T('Show products for workshop')) ]

    fields = [ db.workshops.thumbsmall,
               db.workshops.Name,
               db.workshops.Startdate,
               db.workshops.Teacher,
               db.workshops.auth_teacher_id ]

    db.workshops.id.readable = False

    maxtextlengths = {'workshops.Name' : 40}
    grid = SQLFORM.grid(query,
                        fields=fields,
                        links=links,
                        details=False,
                        searchable=False,
                        csv=False,
                        create=False,
                        editable=False,
                        deletable=False,
                        maxtextlengths=maxtextlengths,
                        orderby=db.workshops.Startdate|db.workshops.Name,
                        ui = grid_ui)
    grid.element('.web2py_counter', replace=None) # remove the counter
    grid.elements('span[title=Delete]', replace=None) # remove text from delete button

    return grid


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'workshops_products_customers'))
def events_add_list_tickets():
    """
        List products for a workshop
    """
    response.title = T("Add workshop product")
    response.view = 'general/only_content.html'

    # To redirect back when adding a workshop product
    session.events_ticket_sell_back = 'customers'

    cuID = request.vars['cuID']
    wsID = request.vars['wsID']
    workshop = db.workshops(wsID)
    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = SPAN(T("Add event product"), B(' > '),
                             T("Products for "),
                             workshop.Name)

    # list of workhsops
    grid = events_add_list_products_get_list(wsID, cuID)

    back = os_gui.get_button('back', URL('event_add',
                                         vars={'cuID':cuID}))

    return dict(content=grid,
                back=back)


def events_add_list_products_get_list(wsID, cuID):
    """
        Returns list of products for a workshop
    """
    query = (db.workshops_products.workshops_id == wsID)

    links = [event_add_list_products_get_list_get_link_add]

    fields = [ db.workshops_products.id,
               db.workshops_products.Name,
               db.workshops_products.Description,
               db.workshops_products.Price ]

    left = [ db.workshops.on(db.workshops_products.workshops_id == \
                             db.workshops.id) ]

    db.workshops_products.id.readable = False

    maxtextlengths = {'workshops.Name' : 40}
    grid = SQLFORM.grid(query,
                        fields=fields,
                        left=left,
                        links=links,
                        details=False,
                        searchable=False,
                        csv=False,
                        create=False,
                        editable=False,
                        deletable=False,
                        maxtextlengths=maxtextlengths,
                        orderby=db.workshops_products.FullWorkshop | \
                                db.workshops_products.Name,
                        ui = grid_ui)
    grid.element('.web2py_counter', replace=None) # remove the counter
    grid.elements('span[title=Delete]', replace=None) # remove text from delete button

    return grid


def event_add_list_products_get_list_get_link_add(row):
    """
        Returns an add button if a customer isn't already added,
        otherwise returns ''
    """
    cuID  = request.vars['cuID']
    wspID = row.id
    row  = db.workshops_products(wspID)
    wsID = row.workshops_id

    wh = WorkshopsHelper()

    buttons = wh.get_product_sell_buttons(cuID, wsID, wspID, request.cid)

    return buttons


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'customers_documents'))
def documents():
    """
        This function shows a list of documents uploaded for a customer
    """
    customers_id = request.vars['cuID']
    response.view = 'customers/edit_general.html'
    customer = Customer(customers_id)
    response.title = customer.get_name()
    response.subtitle = T("Documents")

    db.customers_documents.id.readable=False
    db.customers_documents.UploadDateTime.readable=True
    # set the default value in the grid
    links = [lambda row: os_gui.get_button('edit',
                                     URL('document_edit',
                                         args=[customers_id, row.id]),
                                     T("Edit document description"))]

    query = (db.customers_documents.auth_customer_id == customers_id)
    maxtextlengths = {'customers_documents.Description' : 30}
    headers = {'customers_documents.DocumentFile': T("File")}

    fields = [db.customers_documents.UploadDateTime,
              db.customers_documents.Description,
              db.customers_documents.DocumentFile]

    delete_permission = auth.has_membership(group_id='Admins') or \
                        auth.has_permission('delete', 'customers_documents')
    grid = SQLFORM.grid(query, fields=fields, links=links,
        headers=headers,
        details=False,
        searchable=False,
        deletable=delete_permission,
        csv=False,
        create=False,
        editable=False,
        maxtextlengths=maxtextlengths,
        orderby=~db.customers_documents.UploadDateTime,
        field_id=db.customers_documents.id,
        ui = grid_ui)
    grid.element('.web2py_counter', replace=None) # remove the counter
    grid.elements('span[title=Delete]', replace=None) # remove text from delete button

    add_url = URL('document_add', args=[customers_id])
    add = os_gui.get_button('add', add_url, T("Add a new document"),  btn_size='btn-sm', _class='pull-right')

    back = edit_get_back()

    menu = customers_get_menu(customers_id, request.function)

    return dict(content=grid, menu=menu, back=back, add=add)


def documents_get_return_url(customers_id):
    """
        Returns the return url for documents add and edit
    """
    return URL('documents', vars={'cuID':customers_id})


@auth.requires_login()
def document_add():
    """
        This function shows an add page for a document
        request.args[0] is expected to be the customers_id
    """
    customers_id = request.args[0]
    response.view = 'general/only_content.html'
    customer = Customer(customers_id)
    response.title = customer.get_name()
    response.subtitle = T("Upload document")

    db.customers_documents.auth_customer_id.default = customers_id

    return_url = documents_get_return_url(customers_id)

    space = uploads_available_space(request.folder)
    if space['available'] > 4:
        crud.messages.submit_button = T("Save")
        crud.messages.record_created = T("Uploaded document")
        crud.settings.create_next = return_url
        form = crud.create(db.customers_documents)

        form_element = form.element('form')
        form['_id'] = 'MainForm'

        elements = form.elements('input, select, textarea')
        for element in elements:
            element['_form'] = "MainForm"

        submit = form.element('input[type=submit]')

        content = form
    else:
        content = space['full_message']

    back = os_gui.get_button("back", return_url)

    return dict(content=content, back=back, save=submit)


@auth.requires_login()
def document_edit():
    """
        This function shows an edit page for a document
        request.args[0] is expected to be the customers_id
        request.args[1] is expected to be the customers_documentID (cudID)
    """
    response.view = 'general/only_content.html'

    customers_id = request.args[0]
    cudID = request.args[1]
    customer = Customer(customers_id)
    response.title = customer.get_name()
    response.subtitle = T("Edit Document")

    db.customers_documents.DocumentFile.readable = False
    db.customers_documents.DocumentFile.writable = False

    return_url = documents_get_return_url(customers_id)

    crud.messages.submit_button = T("Save")
    crud.messages.record_updated = T("Updated document")
    crud.settings.update_next = return_url
    crud.settings.update_deletable = False
    form = crud.update(db.customers_documents, cudID)

    form_element = form.element('form')
    form['_id'] = 'MainForm'

    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = "MainForm"

    submit = form.element('input[type=submit]')

    back = os_gui.get_button("back", return_url)

    return dict(content=form, back=back, save=submit)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'auth_user'))
def load_list_set_search():
    """
        Expected to be called as JSON

        variables that can be set are:
        - name
    """
    name = request.vars['name']
    session.customers_load_list_search_name = '%' + name.strip() + '%'
    try:
        search_id = int(name.strip())
        session.customers_load_list_search_name_int = search_id
    except ValueError:
        session.customers_load_list_search_name_int = None

    message = T("Updated search requirements")
    status = 'success'

    return dict(message=message, status=status)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'auth_user'))
def load_list():
    """
        Returns a list of customers, to be used as LOAD
        request.vars['items_per_page'] sets the items shown on each page
        request.vars['list_type'] can be 'classes_attendance_list'
        request.vars['show_location'] can be 'True' or 'False'
    """
    items_per_page = request.vars['items_per_page']
    list_type = request.vars['list_type']
    initial_list = request.vars['initial_list']
    clsID = request.vars['clsID']
    date_formatted = request.vars['date']
    wsID = request.vars['wsID']
    wspID = request.vars['wspID']

    show_location = request.vars['show_location']
    if show_location == 'True':
        show_location = True
    else:
        show_location = False

    pictures = request.vars['pictures']
    if pictures == 'True' or pictures is None:
        pictures = True
    else:
        pictures = False

    contact_permission = ( auth.has_membership(group_id='Admins') or
                           auth.has_permission('update', 'customers_contact') )

    show_email = request.vars['show_email']
    if show_email == 'True' and contact_permission:
        show_email = True
    else:
        show_email = False

    show_deleted = request.vars['show_deleted']
    if show_deleted == 'deleted':
        delete_permission = (auth.has_membership(group_id='Admins') or
                             auth.has_permission('delete', 'auth_user'))
        trashed = True
    else:
        trashed = False

    if date_formatted:
        date = datestr_to_python(DATE_FORMAT, date_formatted)
    else:
        date = datetime.date.today()

    # general settings
    try:
        items_per_page = int(items_per_page)
    except ValueError:
        items_per_page = 10

    if len(request.args):
        page=int(request.args[0])
    else:
        page=0
    limitby=(page*items_per_page,(page+1)*items_per_page+1)

    # query and select
    search_name = session.customers_load_list_search_name

    if (search_name and search_name != '%%') or (initial_list):
        query = (db.auth_user.id > 1)
    else:
        query = (db.auth_user.id < 1)

    title = ''

    query &= (db.auth_user.trashed == trashed)

    if list_type == 'classes_attendance_list':
        title = H4(T('Search results'))

    if search_name:
        if list_type == 'classes_attendance_list' and session.customers_load_list_search_name_int:
            date = datestr_to_python(DATE_FORMAT, date_formatted)
            cuID = session.customers_load_list_search_name_int
            vars = {'cuID': cuID,
                    'clsID': clsID,
                    'date': date_formatted}

            session.customers_load_list_search_name = None
            session.customers_load_list_search_name_int = None

            check_customer_exists = db.auth_user(id = cuID)
            if not check_customer_exists:
                session.flash = SPAN(
                    T("No customer registered with id"), ' ',
                    unicode(cuID), ', ',
                    "please try again."
                )
                redirect(URL('classes', 'attendance',
                             vars=vars,
                             extension=''),
                         client_side=True)

            check = db.classes_attendance(auth_customer_id=cuID,
                                          classes_id=clsID,
                                          ClassDate=date)

            if check:
                if check.BookingStatus == "booked":
                    session.flash = None
                    redirect(URL('classes', 'attendance_set_status',
                                 vars={'clattID':check.id,
                                       'status':'attending'},
                                 extension=''
                                 ),
                             client_side=True)
                elif check.BookingStatus == "cancelled":
                    session.flash = T("Customer booking for this class has status 'Cancelled'")
                else:
                    session.flash = T("Customer is already checked-in")

                redirect(URL('classes', 'attendance',
                             vars=vars,
                             extension=''),
                         client_side=True)
            else:
                redirect(URL('classes', 'attendance_booking_options',
                             vars=vars,
                             extension=''),
                         client_side=True)

        query &= ((db.auth_user.display_name.like(search_name)) |
                  (db.auth_user.email == search_name.replace('%', '')) |
                  (db.auth_user.id == session.customers_load_list_search_name_int))


    if initial_list and (search_name == '%%' or not search_name):
        orderby = ~db.auth_user.id
    else:
        orderby = db.auth_user.display_name

    rows = db(query).select(db.auth_user.id,
                            db.auth_user.trashed,
                            db.auth_user.thumbsmall,
                            db.auth_user.birthday,
                            db.auth_user.display_name,
                            db.auth_user.email,
                            db.auth_user.school_locations_id,
                            limitby=limitby,
                            orderby=orderby)

    table_class = 'table table-hover'
    table = TABLE(_class=table_class)
    for i, row in enumerate(rows.render()):
        if i == items_per_page:
            break

        cuID = row.id

        # get subscription for customer
        customer = Customer(cuID)
        if list_type == 'selfcheckin_checkin':
            # For self check-in display of subscriptions can be configured by the user.
            show_subscriptions_prop = 'selfcheckin_show_subscriptions'
            show_subscriptions = get_sys_property(show_subscriptions_prop)
            if show_subscriptions:
                subscriptions = True
            else:
                subscriptions = False

            subscr_cards = customer.get_subscriptions_and_classcards_formatted(date,
                                                                               show_subscriptions=subscriptions)
        else:
            subscr_cards = customer.get_subscriptions_and_classcards_formatted(date)


        # get email if requested
        email = ''
        if show_email and row.email:
            email = TD(os_gui.get_fa_icon('fa-envelope-o'), ' ',
                       A(row.email,
                         _href='mailto:' + row.email,
                         _class='grey'),
                       _class='grey small_font')

        # get location if requested
        location = ''
        if show_location:
            repr_row = list(rows[i:i + 1].render())[0]
            location = TD(os_gui.get_label('primary', repr_row.school_locations_id))


        # add everything to the table cell
        customer_info = TD(SPAN(row.display_name,
                                _class='customer_name bold'),
                           BR(),
                           subscr_cards)
        if pictures:
            table_row = TR(TD(row.thumbsmall, _class='os-customer_image_td'),
                           customer_info,
                           email,
                           location)

        else:
            table_row = TR(customer_info,
                           email,
                           location)

        if list_type == 'customers_index':
            buttons = TD(load_list_get_customer_index_buttons(row))
        elif list_type == 'customers_index_deleted':
            buttons = TD(load_list_get_customer_index_deleted_buttons(
                row,
                delete_permission
            ))
        elif list_type == 'classes_attendance_list':
            buttons = TD(load_list_get_attendance_list_buttons(row,
                                                               clsID,
                                                               date_formatted),
                         _class='table-vertical-align-middle')
        elif list_type == 'selfcheckin_checkin':
            buttons = TD(load_list_get_selfcheckin_checkin_buttons(
                row,
                clsID,
                date_formatted),
                _class='table-vertical-align-middle')
        elif list_type == 'classes_manage_reservation':
            buttons = TD(load_list_get_reservation_list_buttons(
                row,
                clsID,
                date_formatted),
                _class='table-vertical-align-middle')
        elif list_type == 'events_ticket_sell':
            buttons = TD(load_list_get_events_ticket_sell_buttons(row,
                                                                  wsID,
                                                                  wspID),
                         _class='table-vertical-align-middle')

        try:
            table_row.append(buttons)
        except:
            pass

        table.append(table_row)


    # Navigation
    previous = ''
    url_previous = None
    if page:
        url_previous = URL(args=[page-1], vars=request.vars)
        previous = A(SPAN(_class='glyphicon glyphicon-chevron-left'),
                     _href=url_previous,
                     cid=request.cid)

    nxt = ''
    url_next = None
    if len(rows) > items_per_page:
        url_next = URL(args=[page+1], vars=request.vars)
        nxt = A(SPAN(_class='glyphicon glyphicon-chevron-right'),
                _href=url_next,
                cid=request.cid)

    navigation = os_gui.get_page_navigation_simple(url_previous, url_next, page+1, request.cid)

    if previous or nxt:
        pass
    else:
        navigation = ''

    content = DIV(title, table, navigation)

    if len(rows) == 0:
        content = DIV(DIV(BR(), T("No results..."), BR(),
                          _class='grey col-md-12'),
                      _class='row')

    return dict(content=content)


def load_list_get_reservation_list_buttons(row,
                                           clsID,
                                           date_formatted):
    """
        Returns buttons for the manage_reservation list type
    """
    cuID = row.id
    date = datestr_to_python(DATE_FORMAT, date_formatted)
    customer = Customer(cuID)
    customer_name = customer.get_name()

    recurring_url = URL('classes', 'reservation_add_choose',
                        vars={'cuID'  : cuID,
                              'clsID' : clsID,
                              'date'  : date_formatted},
                        extension='')

    button = os_gui.get_button('add', recurring_url,
                               tooltip=T('Enroll this customer in this class'),
                               _class='pull-right')

    return button


def load_list_get_customer_index_buttons(row):
    """
        Returns buttons for customers.py/index
    """
    buttons = DIV(_class='btn-group')

    btn_mail = ''
    contact_permission = ( auth.has_membership(group_id='Admins') or
                           auth.has_permission('update', 'customers_contact') )

    if contact_permission and row.email:
        btn_mail = A(I(_class="fa fa-envelope"), " ",
                     _class="btn btn-default btn-sm",
                     _href='mailto:' + row.email or '',
                     _title=T("Mail customer"))

    buttons.append(btn_mail)

    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('update', 'auth_user'):

       btn_edit =  os_gui.get_button('edit', URL('edit',
                                                  args=[row.id],
                                                  extension=''))
       buttons.append(btn_edit)

    delete = ''
    if (auth.has_membership(group_id='Admins') or
        auth.has_permission('delete', 'auth_user')):

        if auth.user.id == row.id:
            onclick = "alert('Unable to delete, you are currently logged in using this account.');"
            url = '#'
        else:
            onclick = "return confirm('" + \
                 T('Move this customer to deleted?') + "');"
            url = URL('trash', vars={'cuID':row.id}, extension='')

        delete = os_gui.get_button('delete_notext',
                                   url,
                                   onclick=onclick)

    return DIV(buttons, delete, _class='pull-right')


def load_list_get_customer_index_deleted_buttons(row, permission):
    """
        Return customer index deleted buttons
    """
    onclick_delete = "return confirm('" + \
                     T('Do you really want to delete this customer and all associated data?') \
                     + "');"
    onclick_restore = "return confirm('" + \
                      T('Restore customer to current?') \
                      + "');"
    restore = ''
    if permission:
        restore = os_gui.get_button(
            'noicon',
            URL('customers', 'restore',
                vars={'cuID': row.id},
                extension=''),
            title=T('Restore'),
            onclick=onclick_restore,
            _class="pull-right"
        )

    delete = ''
    if permission:
        delete = os_gui.get_button(
            'delete_notext',
            URL('customers', 'delete',
                vars={'cuID': row.id},
                extension=''),
            onclick=onclick_delete,
            _class="pull-right"
        )

    return DIV(delete, restore)


def load_list_get_attendance_list_buttons(row,
                                          clsID,
                                          date_formatted):
    """
        Returns buttons for the attendance_list list type
    """
    date = datestr_to_python(DATE_FORMAT, date_formatted)
    check = db.classes_attendance(auth_customer_id=row.id,
                                  classes_id=clsID,
                                  ClassDate=date)
    if not check:
        url = URL('classes', 'attendance_booking_options',
                  vars={'cuID':row.id,
                        'clsID':clsID,
                        'date':date_formatted},
                  extension='')

        return A(T('Check in'),
                 _href=url,
                 _class='btn btn-default btn-sm pull-right')
    else:
        return ''


def load_list_get_selfcheckin_checkin_buttons(row,
                                              clsID,
                                              date_formatted):
    """
        Returns buttons for the selfcheckin_checkin list type
    """
    date = datestr_to_python(DATE_FORMAT, date_formatted)
    check = db.classes_attendance(auth_customer_id=row.id,
                                  classes_id=clsID,
                                  ClassDate=date)
    if not check:
        url = URL('selfcheckin', 'checkin_booking_options',
                  vars={'cuID':row.id,
                        'clsID':clsID,
                        'date':date_formatted},
                  extension='')

        return A(T('Check in'),
                 _href=url,
                 _class='btn btn-default btn-sm pull-right')
    else:
        return ''


def load_list_get_events_ticket_sell_buttons(row,
                                             wsID,
                                             wspID):
    """
        Returns buttons for workshop_product_sell list type
        This is a select button to select a customer to sell a product to
    """
    cuID = row.id

    wh = WorkshopsHelper()

    buttons = wh.get_product_sell_buttons(cuID, wsID, wspID, request.cid)

    return buttons


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'tasks'))
def tasks():
    """
        Display list of tasks for a customer
    """
    response.view = 'customers/edit_general.html'
    cuID = request.vars['cuID']
    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = T("Tasks")
    #session.tasks_index_filter = 'open'
    cuID = request.vars['cuID']

    content = DIV(LOAD('tasks', 'list_tasks.load',
                       vars=request.vars,
                       content=os_gui.get_ajax_loader()))

    # Add permission
    add = ''
    permission = auth.has_membership(group_id='Admins') or \
                 auth.has_permission('create', 'tasks')
    if permission:
        #add = os_gui.get_button('add', url_add)
        from openstudio.os_tasks import Tasks
        tasks = Tasks()
        add = tasks.add_get_modal({'cuID':cuID})

    back = edit_get_back()
    menu = customers_get_menu(cuID, request.function)

    #save = os_gui.get_submit_button('task_edit')

    return dict(content=content,
                menu=menu,
                back=back,
                add=add)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'auth_user_account'))
def account():
    """
        Account options for an account
    """
    response.view = 'customers/edit_general.html'
    cuID = request.vars['cuID']
    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = T("Account")

    for field in db.auth_user:
        field.readable = False
        field.writable = False

    db.auth_user.trashed.readable = True
    db.auth_user.trashed.writable = True
    db.auth_user.customer.readable = True
    db.auth_user.customer.readable = False
    db.auth_user.customer.writable = False
    db.auth_user.enabled.readable = True
    db.auth_user.enabled.writable = True
    db.auth_user.teacher.readable = True
    db.auth_user.teacher.writable = True
    db.auth_user.employee.readable = True
    db.auth_user.employee.writable = True
    db.auth_user.business.readable = True
    db.auth_user.business.writable = True
    db.auth_user.login_start.readable = True
    db.auth_user.login_start.writable = True

    crud.messages.submit_button = T("Save")
    crud.messages.record_updated = T("Saved")
    crud.settings.update_onaccept.auth_user.append(edit_onaccept)
    crud.settings.update_deletable = False
    form = crud.update(db.auth_user, cuID)

    form_id = "MainForm"
    form_element = form.element('form')
    form['_id'] = form_id

    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = form_id

    submit = form.element('input[type=submit]')


    submenu = account_get_submenu(request.function, cuID)
    verify_email = account_get_verify_email(cuID)
    content = DIV(
        DIV(
            submenu, BR(),
            DIV(form, _class='col-md-6'),
            DIV(verify_email, _class='col-md-6'),
            _class='col-md-12'
        ),
        _class='row'
    )

    back = edit_get_back()
    menu = customers_get_menu(cuID, 'account')

    return dict(content=content,
                menu=menu,
                back=back,
                save=submit)


def account_get_verify_email(cuID):
    """
    Returns a button to verify the customer's email address
    if not verified
    """
    permission = (auth.has_membership(group_id='Admins') or
                  auth.has_permission('update', 'auth_user'))

    au = db.auth_user(cuID)

    if not permission:
        return ''
    else:
        if not au.registration_key:
            return DIV(
                os_gui.get_fa_icon('fa-check'), ' ',
                T("The email address of this account has been verified"),
                _class='text-green'
            )
        else:
            verify = os_gui.get_button(
                'noicon',
                URL('account_verify_email', vars={'cuID': cuID}),
                title=T("Verify email address"),
                btn_class='btn-success'
            )

            return DIV(
                B(T("Manually verify {email} as valid".format(email=au.email))), BR(),
                T("This account hasn't verified it's email address yet using the verification email."), BR(), BR(),
                verify
            )


@auth.requires(auth.has_membership(group_id='Admins') or
               auth.has_permission('update', 'auth_user'))
def account_verify_email():
    """
    Manually verify email address
    """
    cuID = request.vars['cuID']

    au = db.auth_user(cuID)
    au.registration_key = None
    au.update_record()

    redirect(URL('account', vars={'cuID': cuID}))


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('merge', 'auth_user'))
def account_merge():
    """
        Page to allow merging of 2 entries in auth_user
    """
    response.view = 'customers/edit_general.html'
    cuID = request.vars['cuID']
    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = T("Account")

    warning = ''
    if 'auth_merge_id' in request.vars:
        auth_merge_id = request.vars['auth_merge_id']
        btn_onclick = "return confirm('" + \
                      T('Absolutely sure you want to merge the selected account into this one?') + \
                      "');"
        btn_title = XML(SPAN(T("Click here")))
        btn = A(btn_title,
                _href=URL('account_merge_execute',
                          vars={'cuID':cuID,
                                'auth_merge_id':auth_merge_id}),
                _title=T('Execute merge'),
                _class='btn btn-link',
                _id   ='btn_merge_execute',
                _onclick = btn_onclick)
        msg_extra = SPAN(T('to merge all info except the profile of account with ID'), ' ',
                         auth_merge_id, ' ',
                         T("into this account."))
        warning_text = DIV(B(T("Warning")), ' ',
                           T("Merging cannot be undone!"), ' ',
                           btn,
                           msg_extra)
        warning = DIV(HR(),
                      os_gui.get_alert('warning',
                                       warning_text,
                                       dismissable=False))


    submenu = account_get_submenu(request.function, cuID)
    description = DIV(B(T("Select an account to merge into this one")))
    form = account_merge_get_input_form(request.vars['auth_merge_id'])

    result = set_form_id_and_get_submit_button(form, 'MainForm')
    form = result['form']
    submit = result['submit']

    content = DIV(submenu, BR(),
                  description,
                  form,
                  warning)

    back = edit_get_back()
    menu = customers_get_menu(cuID, 'account')

    return dict(content=content,
                menu=menu,
                back=back,
                save=submit)


def account_merge_get_input_form(auth_merge_id):
    """
        Simple input form to enter an auth_user_id to merge with
    """
    db.auth_user._format = '%(id)s - %(display_name)s'

    merge_query = (db.auth_user.merged == False) & \
                  (db.auth_user.id > 1) # exclude admin user
    form = SQLFORM.factory(
        Field('auth_merge_id', db.auth_user,
              default  = auth_merge_id,
              requires = IS_IN_DB(db(merge_query),
                                  'auth_user.id',
                                  '%(id)s - %(display_name)s - %(email)s - Trashed: %(trashed)s - Teacher: %(teacher)s',
                              zero=T("Please select...")),
              label    = T('')),
        submit_button = T('Select'),
        formstyle='table3cols'
        )

    select = form.element('#no_table_auth_merge_id')


    return form


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('merge', 'auth_user'))
def account_merge_execute():
    """
        Actually merge account
    """
    cuID          = request.vars['cuID'] # merge into
    auth_merge_id = request.vars['auth_merge_id'] # merge from

    merge_into = db.auth_user(cuID)
    merge_from = db.auth_user(auth_merge_id)
    if not merge_from.merged:
        # loop over all tables in db
        for table in db:
            # auth_user_id
            try:
                query = (table.auth_user_id == auth_merge_id)
                rows = db(query).select(table.ALL)
                for row in rows:
                    row.auth_user_id = cuID
                    row.update_record()
            except AttributeError:
                pass
            # auth_teacher_id
            try:
                query = (table.auth_teacher_id == auth_merge_id)
                rows = db(query).select(table.ALL)
                for row in rows:
                    row.auth_teacher_id = cuID
                    row.update_record()
            except AttributeError:
                pass
            # auth_teacher_id2
            try:
                query = (table.auth_teacher_id2 == auth_merge_id)
                rows = db(query).select(table.ALL)
                for row in rows:
                    row.auth_teacher_id2 = cuID
                    row.update_record()
            except AttributeError:
                pass
            # auth_customer_id
            try:
                query = (table.auth_customer_id == auth_merge_id)
                rows = db(query).select(table.ALL)
                for row in rows:
                    row.auth_customer_id = cuID
                    row.update_record()
            except AttributeError:
                pass

        # mark row as merged
        # set merged for auth_user auth_merge_id
        merge_from.merged = True
        merge_from.trashed = True
        # clear merged email
        merge_from.email = None
        # set merged_with for auth_user auth_merge_id
        merge_from.merged_into = int(cuID)
        merge_from.merged_on = datetime.datetime.now()

        # if the merge_from account is a teacher, make the merge into account a teacher
        if merge_from.teacher:
            merge_into.teacher = True
        # if the merge_from account is enabled, enable the merge into account
        if merge_from.enabled:
            merge_into.enabled = True
        # if the merge_from account is an employee, make the merge into account an employee
        if merge_from.employee:
            merge_into.employee = True

        merge_from.update_record()
        merge_into.update_record()

        session.flash = T("Merge success")
    else:
        session.flash = T('Merge failed - already merged')


    redirect(URL('edit', args=[cuID]))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('set_password', 'auth_user'))
def account_set_password():
    """
        Set a new password for an account
    """
    response.view = 'customers/edit_general.html'
    cuID = request.vars['cuID']
    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = T('Account')

    for field in db.auth_user:
        field.readable = False
        field.writable = False

    db.auth_user.password.readable = True
    db.auth_user.password.writable = True
    # Enforce strong passwords!
    db.auth_user.password.requires.insert(0, IS_STRONG())

    crud.settings.update_onaccept.auth_user.append(edit_onaccept)
    form = crud.update(db.auth_user, cuID)

    form_id = "MainForm"
    form_element = form.element('form')
    form['_id'] = form_id

    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = form_id

    submit = form.element('input[type=submit]')

    submenu = account_get_submenu(request.function, cuID)
    description = DIV(B(T("Enter a new password for this account")))
    description_send_link = DIV(B(
        T("Send a set password link to this customers' email address")
    ))

    content = DIV(
        submenu,
        BR(),
        DIV(
            description,
            form,
            _class="col-md-6"
        ),
        DIV(
            #description_send_link,
            _class="col-md-6"
        ),
        _class='row'
    )


    back = edit_get_back()
    menu = customers_get_menu(cuID, request.function)

    return dict(content=content,
                menu=menu,
                back=back,
                save=submit)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'log_customers_accepted_documents'))
def account_acceptance_log():
    """
        Lists accepted documents for customer
    """
    response.view = 'customers/edit_general.html'
    cuID = request.vars['cuID']
    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = T('Account')

    submenu = account_get_submenu(request.function, cuID)

    header = THEAD(
        TR(
            TH(T('Document')),
            TH(T('Document Description')),
            TH(T('Document Version')),
            TH(T('Document URL')),
            TH(T('OpenStudio Version')),
            TH(T('Accepted On')),
        )
    )
    table = TABLE(header, _class='table table-striped table-hover')
    rows = customer.get_accepted_documents()
    for row in rows.render():
        tr = TR(
            TD(row.DocumentName),
            TD(row.DocumentDescription),
            TD(row.DocumentVersion),
            TD(row.DocumentURL),
            TD(row.OpenStudioVersion),
            TD(row.CreatedOn),
        )

        table.append(tr)

    content = DIV(submenu, table)

    menu = customers_get_menu(cuID, 'account')
    back = edit_get_back()

    return dict(content=content,
                menu=menu,
                back=back)


@auth.requires(auth.has_membership(group_id='Admins'))
def account_exact_online():
    """
    Manage link exact online linked customer
    """
    response.view = 'customers/edit_general.html'
    cuID = request.vars['cuID']
    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = T('Account')

    submenu = account_get_submenu(request.function, cuID)

    form = SQLFORM.factory(
        Field('code',
              defualt=request.vars['code'],
              requires=IS_NOT_EMPTY(),
              label=T("Exact Online relation code")
              ),
        formstyle = 'bootstrap3_stacked',
        submit_button = T('Find Exact relations')
    )

    search_result = ''
    if form.process().accepted:
        from openstudio.os_exact_online import OSExactOnline

        response.flash = T("Successfully submitted search to Exact Online")
        code = request.vars['code']

        os_eo = OSExactOnline()
        api = os_eo.get_api()
        relations = api.relations.filter(relation_code=code)

        if not len(relations):
            search_result = T("No relations found with this code in Exact Online")
        else:
            header = THEAD(TR(
                TH(T('Exact Online Code')),
                TH(T('Exact Online Name')),
                TH()
            ))
            search_result = TABLE(header, _class="table table-striped table-hover")

            for relation in relations:
                btn_link = os_gui.get_button(
                    'noicon',
                    URL('account_exact_online_link_relation', vars={'cuID': cuID,
                                                                    'eoID': relation['ID']}),
                    title=T("Link to this relation"),
                    _class='pull-right'
                )

                search_result.append(TR(
                    TD(relation['Code']),
                    TD(relation['Name']),
                    TD(btn_link)
                ))

    result = set_form_id_and_get_submit_button(form, 'MainForm')
    form = result['form']
    submit = result['submit']


    current_link = T("Search for a relation code to link an Exact Online relation to this customer.")
    linked_relations = customer.exact_online_get_relation()
    if linked_relations:
        current_link = DIV(
            T("This customer is linked to the following Exact Online relation"), BR(), BR(),
            T("Exact Online code: %s" % (linked_relations[0]['Code'])), BR(),
                T("Exact Online name: %s" % (linked_relations[0]['Name'])), BR(), BR(),
            T("To link this customer to another Exact Online relation, search for a relation code.")
        )

    display = DIV(
        DIV(
            H4(T("Current link")),
            current_link,
            _class='col-md-6'
        ),
        DIV(
            H4(T('Find relations in Exact Online')),
            form,
            HR(),
            search_result,
            _class='col-md-6'
        ),
        _class='row'
    )

    content = DIV(submenu, BR(), display)

    menu = customers_get_menu(cuID, 'account')
    back = edit_get_back()

    return dict(content=content,
                menu=menu,
                save=submit,
                back=back)


@auth.requires(auth.has_membership(group_id='Admins'))
def account_exact_online_link_relation():
    """
    Link exact online relation to OpenStudio customer
    """
    from openstudio.os_customer import Customer

    cuID = request.vars['cuID']
    eoID = request.vars['eoID']

    customer = Customer(cuID)
    message = customer.exact_online_link_to_relation(eoID)

    session.flash = message
    redirect(URL('account_exact_online', vars={'cuID': cuID}))


def account_get_submenu(page, cuID):
    """
        Returns submenu for account pages
    """
    vars = {'cuID':cuID}
    pages = [['account', T('Account'), URL('account', vars=vars)]]

    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('merge', 'auth_user'):
        pages.append(['account_merge', T('Merge'), URL('account_merge',
                                                        vars=vars)])
    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('set_password', 'auth_user'):
        pages.append(['account_set_password', T('Reset password'),
                       URL('account_set_password', vars=vars)])

    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('account_acceptance_log', 'auth_user'):
        pages.append(['account_acceptance_log', T('Accepted documents'),
                       URL('account_acceptance_log', vars=vars)])

    eo_authorized = get_sys_property('exact_online_authorized')
    if auth.has_membership(group_id='Admins') and eo_authorized:
        pages.append(['account_exact_online', T('Exact Online'),
                       URL('account_exact_online', vars=vars)])

    horizontal = True

    return os_gui.get_submenu(pages, page, horizontal=horizontal, htype='tabs')


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('read', 'invoices'))
def invoices():
    """
        List invoices for a customer
    """
    cuID = request.vars['cuID']

    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = T("Invoices")
    response.view = 'customers/edit_general.html'

    session.invoices_edit_back = 'customers_invoices'
    session.invoices_payment_add_back = 'customers_invoices'
    # Always reset filter
    session.invoices_list_status = None

    invoices = Invoices()
    form = invoices.add_get_form(cuID)
    result = invoices.add_get_modal(form)
    add = result['button']
    modal_class = result['modal_class']

    status_filter = invoices.list_get_status_filter()

    content = DIV(status_filter, BR())

    list = invoices.list_invoices(cuID=cuID)
    content.append(list)

    back = edit_get_back()
    menu = customers_get_menu(cuID, request.function)

    return dict(content=content,
                menu=menu,
                back=back,
                form_add=form,
                add=add,
                modal_class=modal_class,
                left_sidebar_enabled=True)


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('read', 'customers_orders'))
def orders():
    """
        List orders for a customer
    """
    cuID = request.vars['cuID']

    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = T("Orders")
    response.view = 'customers/edit_general.html'

    session.invoices_edit_back = 'customers_orders'
    session.invoices_payment_add_back = 'customers_orders'
    session.orders_edit_back = 'customers_orders'

    header = THEAD(TR(TH(T('Order #')),
                      TH(T('Time')),
                      TH(T('Total amount incl. VAT')),
                      TH(T('Status')),
                      TH(T('Invoice')),
                      TH())
                     )

    table = TABLE(header, _class='table table-striped table-hover')
    rows = customer.get_orders_rows()
    for i, row in enumerate(rows):
        repr_row = list(rows[i:i + 1].render())[0]

        link_view = os_gui.get_button('edit', URL('orders', 'edit', vars={'cuID': cuID,
                                                                          'coID': row.customers_orders.id}),
                                      _class='pull-right',
                                      btn_size='')

        table.append(TR(TD(repr_row.customers_orders.id),
                        TD(repr_row.customers_orders.DateCreated),
                        TD(repr_row.customers_orders_amounts.TotalPriceVAT),
                        TD(repr_row.customers_orders.Status),
                        TD(orders_get_link_invoice(row)),
                        TD(link_view)
                        ))

    content = table

    back = edit_get_back()
    menu = customers_get_menu(cuID, request.function)

    return dict(content=content,
                menu=menu,
                back=back,
                left_sidebar_enabled=True)


def orders_get_link_invoice(row):
    """
        Returns invoice for an order in list
    """
    if row.invoices.id:
        invoices = Invoices()

        query = (db.invoices.id == row.invoices.id)
        rows = db(query).select(db.invoices.ALL)
        repr_row = rows.render(0)

        invoice_link = invoices.represent_invoice_for_list(
            row.invoices.id,
            repr_row.InvoiceID,
            repr_row.Status,
            row.invoices.Status,
            row.invoices.payment_methods_id
        )

    else:
        invoice_link = ''

    return invoice_link



@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('read', 'customers_orders'))
def order():
    """
        Display order content for a customer
    """
    cuID = request.vars['cuID']
    coID = request.vars['coID']

    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = T("Order #") + coID
    response.view = 'customers/edit_general.html'

    content = DIV(os_gui.get_button('back_bs', URL('orders', vars={'cuID': cuID})),
                  BR(), BR(),
                  DIV(LOAD('orders', 'display_order.load', ajax=False,
                           vars={'coID': coID}))
                  )

    back = edit_get_back()
    menu = customers_get_menu(cuID, 'orders')

    return dict(content=content,
                menu=menu,
                back=back,
                left_sidebar_enabled=True)


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('update', 'auth_user'))
def edit_teacher():
    """
        Teacher profile for a user
    """
    cuID = request.vars['cuID']
    response.view = 'customers/edit_general.html'

    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = T("Teacher profile")

    # disable requires for all fields to allow submitting a limited nr of values
    for field in db.auth_user:
        field.readable=False
        field.writable=False

    fields = [ db.auth_user.teacher_role,
               db.auth_user.teacher_bio,
               db.auth_user.education,
               db.auth_user.teacher_bio_link,
               db.auth_user.teacher_website ]

    for field in fields:
        field.readable=True
        field.writable=True

    crud.messages.submit_button = T('Save')
    crud.messages.record_updated = T('Saved')
    crud.settings.update_onaccept = [
        cache_clear_school_teachers,
        cache_clear_classschedule,
        edit_onaccept
    ]
    form = crud.update(db.auth_user, cuID)

    result = set_form_id_and_get_submit_button(form, 'MainForm')
    form = result['form']
    submit = result['submit']

    textareas = form.elements('textarea')
    for textarea in textareas:
        textarea['_class'] += ' tmced'


    teacher_info = DIV(
        XML('<form action="#" id="MainForm" enctype="multipart/form-data" method="post">'),
        DIV(DIV(LABEL(form.custom.label.teacher_role),
                form.custom.widget.teacher_role,
                _class='col-md-12'),
            DIV(LABEL(form.custom.label.teacher_bio),
                form.custom.widget.teacher_bio,
                _class='col-md-6'),
            DIV(LABEL(form.custom.label.education),
                form.custom.widget.education,
                _class='col-md-6'),
            DIV(LABEL(form.custom.label.teacher_bio_link),
                    form.custom.widget.teacher_bio_link,
                    _class='col-md-6'),
            DIV(LABEL(form.custom.label.teacher_website),
                form.custom.widget.teacher_website,
                _class='col-md-6'),
            _class='row'),
        form.custom.end,
        _class='tab-pane active')

    back = edit_get_back()
    menu = customers_get_menu(cuID, request.function)

    return dict(content=teacher_info,
                menu=menu,
                back=back,
                save=submit)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'customers_memberships'))
def memberships():
    """
        Lists memberships for a customer
        request.vars['cuID'] is expected to be the customersID
    """
    customers_id = request.vars['cuID']
    response.view = 'customers/edit_general.html'

    session.invoices_edit_back = 'customers_memberships'
    session.invoices_payment_add_back = 'customers_memberships'

    row = db.auth_user(customers_id)
    response.title = row.display_name
    response.subtitle = T("Memberships")

    header = THEAD(TR(TH('#'),
                      TH(db.customers_memberships.school_memberships_id.label),
                      TH(db.customers_memberships.Startdate.label),
                      TH(db.customers_memberships.Enddate.label),
                      TH(db.customers_memberships.payment_methods_id.label),
                      TH(db.customers_memberships.Note.label),
                      TH(T("Invoice")),
                      TH())  # buttons
                   )

    table = TABLE(header, _class='table table-hover table-striped')

    left = [
        db.invoices_items_customers_memberships.on(
            db.invoices_items_customers_memberships.customers_memberships_id ==
            db.customers_memberships.id
        ),
        db.invoices_items.on(
            db.invoices_items_customers_memberships.invoices_items_id ==
            db.invoices_items.id
        ),
        db.invoices.on(
            db.invoices_items.invoices_id ==
            db.invoices.id
        )
    ]

    query = (db.customers_memberships.auth_customer_id == customers_id)
    rows = db(query).select(
        db.customers_memberships.ALL,
        db.invoices.id,
        db.invoices.Status,
        db.invoices.InvoiceID,
        db.invoices.payment_methods_id,
        left=left,
        orderby=~db.customers_memberships.Startdate|\
                ~db.customers_memberships.id
    )

    edit_permission = auth.has_membership(group_id='Admins') or \
                        auth.has_permission('update', 'customers_memberships')
    delete_permission = auth.has_membership(group_id='Admins') or \
                        auth.has_permission('delete', 'customers_memberships')

    for i, row in enumerate(rows):
        repr_row = list(rows[i:i + 1].render())[0]

        delete = ''
        if delete_permission:
            confirm_msg = T("Really delete this membership?")
            onclick_del = "return confirm('" + confirm_msg + "');"
            delete = os_gui.get_button('delete_notext',
                                       URL('membership_delete', vars={'cuID': customers_id,
                                                                      'cmID': row.customers_memberships.id}),
                                       onclick=onclick_del,
                                       _class='pull-right')
        edit = ''
        if edit_permission:
            edit = memberships_get_link_edit(row)

        tr = TR(TD(row.customers_memberships.id),
                TD(repr_row.customers_memberships.school_memberships_id),
                TD(repr_row.customers_memberships.Startdate),
                TD(repr_row.customers_memberships.Enddate),
                TD(repr_row.customers_memberships.payment_methods_id),
                TD(repr_row.customers_memberships.Note),
                TD(memberships_get_link_invoice(row)),
                TD(delete, edit))

        table.append(tr)

    add = ''
    if (auth.has_membership(group_id='Admins') or
            auth.has_permission('create', 'customers_memberships')):
        add_url = URL('membership_add', vars={'cuID': customers_id})
        add = os_gui.get_button(
            'add',
            add_url,
            T("Add a new membership"),
            btn_size='btn-sm',
            _class='pull-right'
        )

    content = DIV(table)

    back = edit_get_back()
    menu = customers_get_menu(customers_id, request.function)

    return dict(content=content, menu=menu, back=back, tools=add)


def memberships_get_link_edit(row):
    """
        Returns drop down link for subscriptions
    """
    cmID = row.customers_memberships.id
    vars = {'cuID': row.customers_memberships.auth_customer_id,
            'cmID': cmID}

    # menu = ''
    # permission = ( auth.has_membership(group_id='Admins') or
    #                auth.has_permission('update', 'customers_memberships') )
    # if permission:
    #     links = []
    #     link_edit = A((os_gui.get_fa_icon('fa-pencil'), T('Edit')),
    #                   _href=URL('membership_edit', vars=vars))
    #     links.append(link_edit)
    #
    #
    #     menu = os_gui.get_dropdown_menu(
    #         links=links,
    #         btn_text='',
    #         btn_size='btn-sm',
    #         btn_icon='pencil',
    #         menu_class='btn-group pull-right')

    edit = ''
    permission = ( auth.has_membership(group_id='Admins') or
                   auth.has_permission('update', 'customers_memberships') )

    if permission:
        edit = os_gui.get_button(
            'edit',
            URL('membership_edit', vars=vars),
            _class='pull-right'
        )

    return edit


def memberships_get_link_invoice(row):
    """
        Returns invoice for classcard in list
    """
    if row.invoices.id:
        invoices = Invoices()

        query = (db.invoices.id == row.invoices.id)
        rows = db(query).select(db.invoices.ALL)
        repr_row = rows.render(0)

        invoice_link = invoices.represent_invoice_for_list(
            row.invoices.id,
            repr_row.InvoiceID,
            repr_row.Status,
            row.invoices.Status,
            row.invoices.payment_methods_id
        )

    else:
        invoice_link = ''

    return invoice_link


def memberships_clear_cache(form):
    """
        Clear the memberships cache for customer
    """
    cmID = form.vars.id
    cm = db.customers_memberships(cmID)
    cuID = cm.auth_customer_id

    cache_clear_customers_memberships(cuID)


@auth.requires_login()
def membership_add():
    """
        This function shows an add page for a membership
        request.vars['cuID is expected to be the customers_id
    """
    from openstudio.os_forms import OsForms

    cuID = request.vars['cuID']
    customer = Customer(cuID)
    response.view = 'general/tabs_menu.html'
    response.title = customer.get_name()
    response.subtitle = T("Memberships")

    return_url = memberships_get_return_url(cuID)

    db.customers_memberships.auth_customer_id.default = cuID
    db.customers_memberships.Enddate.readable = False
    db.customers_memberships.Enddate.writable = False

    os_forms = OsForms()
    result = os_forms.get_crud_form_create(
        db.customers_memberships,
        return_url,
        onaccept = [
            membership_add_set_enddate,
            membership_add_create_invoice,
            memberships_clear_cache,
        ]
    )

    form = result['form']
    back = os_gui.get_button('back', return_url)
    menu = customers_get_menu(cuID, 'memberships')

    content = DIV(
        H4(T('Add membership')),
        form
    )

    return dict(content=content,
                save=result['submit'],
                back=back,
                menu=menu)


def membership_add_create_invoice(form):
    """
        Add an invoice after adding a membership
    """
    from openstudio.os_school_membership import SchoolMembership

    cmID = form.vars.id
    smID = form.vars.school_memberships_id

    # create invoice
    sm = SchoolMembership(smID)
    sm.sell_to_customer_create_invoice(cmID)


def membership_add_set_enddate(form):
    """
        Calculate and set enddate when adding a membership
    """
    from openstudio.os_school_membership import SchoolMembership

    cmID = form.vars.id
    smID = form.vars.school_memberships_id
    sm = SchoolMembership(smID)
    enddate = sm.sell_to_customer_get_enddate(form.vars.Startdate)

    # set enddate
    row = db.customers_memberships(cmID)
    row.Enddate = enddate
    row.update_record()


def membership_edit_get_subtitle(cmID):
    """
        Returns subtitle for subscription edit pages
    """
    from openstudio.os_customer_membership import CustomerMembership

    cm = CustomerMembership(cmID)

    return SPAN(T("Edit membership"), ' ', cm.school_membership.Name)


def membership_edit_get_back(cuID):
    """
        Returns back button for customer membership edit pages
    """
    return os_gui.get_button('back',
        URL('memberships', vars={'cuID':cuID}),
        _class='')


@auth.requires_login()
def membership_edit():
    """
        This function shows an edit page for a membership
        request.args[0] is expected to be the customers_id
        request.args[1] is expected to be the membershipID
    """
    from openstudio.os_customer_membership import CustomerMembership
    from openstudio.os_forms import OsForms

    cuID = request.vars['cuID']
    cmID = request.vars['cmID']
    response.view = 'general/only_content.html'

    session.invoices_edit_back = 'customers_membership_invoices'
    session.invoices_payment_add_back = 'customers_membership_invoices'

    # Always reset filter
    session.invoices_list_status = None

    customer = Customer(cuID)
    cm = CustomerMembership(cmID)
    response.title = customer.get_name()
    response.subtitle = SPAN(T("Edit membership"), ' ',
                             cm.get_name())

    return_url = memberships_get_return_url(cuID)

    db.customers_memberships.auth_customer_id.default = cuID

    os_forms = OsForms()
    result = os_forms.get_crud_form_update(
        db.customers_memberships,
        return_url,
        cmID,
        onaccept = [
            memberships_clear_cache
        ]
    )

    form = result['form']
    back = membership_edit_get_back(cuID)

    return dict(content=form,
                save=result['submit'],
                back=back)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'customers_memberships'))
def membership_delete():
    """
        Function to delete a membership
    """
    cmID = request.vars['cmID']
    cuID = request.vars['cuID']

    query = (db.customers_memberships.id == cmID)
    db(query).delete()

    cache_clear_customers_memberships(cuID)

    session.flash = T('Deleted membership')
    redirect(memberships_get_return_url(cuID))


def memberships_get_return_url(customers_id):
    """
        Returns return URL for memberships
    """
    return URL('memberships', vars={'cuID':customers_id})


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'auth_user'))
def barcode_label():
    """
        Preview barcode label
    """
    from openstudio.os_customer import Customer

    cuID = request.vars['cuID']
    customer = Customer(cuID)

    return customer.get_barcode_label()
