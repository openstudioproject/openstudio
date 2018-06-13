# -*- coding: utf-8 -*-

from gluon import *


class CustomerExport:
    def __init__(self, cuID):
        """
            :param cuID: db.auth_user.id
        """
        db = current.db

        self.cuID = cuID
        self.row = db.auth_user(self.cuID)


    def excel(self):
        """
            Customer export all data
        """
        from cStringIO import StringIO
        import openpyxl


        db = current.db

        stream = StringIO()
        # Create the workbook
        wb = openpyxl.workbook.Workbook(write_only=True)

        # Add customer data to workbook
        self._excel_account(db, wb)
        self._excel_customers_notes(db, wb)
        self._excel_alternative_payments(db, wb)
        self._excel_customers_classcards(db, wb)
        self._excel_customers_subscriptions(db, wb)
        self._excel_customers_payment_info(db, wb)
        self._excel_log_customers_accepted_documents(db, wb)
        self._excel_customers_shoppingcart(db, wb)
        self._excel_customers_orders(db, wb)
        self._excel_invoices(db, wb)
        self._excel_classes_attendance(db, wb)
        self._excel_classes_reservation(db, wb)
        self._excel_classes_waitinglist(db, wb)
        self._excel_workshops_products(db, wb)
        self._excel_workshops_activities(db, wb)
        self._excel_messages(db, wb)
        self._excel_payment_batch_items(db, wb)


        wb.save(stream)
        return stream


    def _excel_account(self, db, wb):
        """
            Account info for excel export of customer data
        """
        ws = wb.create_sheet('Account')

        data = []
        header = [
            'business',
            'customer',
            'teacher',
            'teaches_classes',
            'teaches_workshops',
            'employee',
            'first_name',
            'last_name',
            'gender',
            'date_of_birth',
            'address',
            'postcode',
            'city',
            'country',
            'email',
            'phone',
            'mobile',
            'emergency',
            'keynr',
            'company',
            'discovery',
            'level',
            'location',
            'language',
            'teacher_role',
            'teacher_bio',
            'teacher_education',
            'teacher_bio_link',
            'teacher_website',
            'mollie_customer_id',
            'last_login',
            'created_on',
        ]

        ws.append(header)

        discovery = None
        if self.row.school_discovery_id:
            discovery = db.school_discovery(self.row.school_discovery_id).Name

        level = None
        if self.row.school_levels_id:
            level = db.school_levels(self.row.school_levels_id).Name

        location = None
        if self.row.school_locations_id:
            location = db.school_locations(self.row.school_locations_id).Name

        language = None
        if self.row.school_languages_id:
            language = db.school_languages(self.row.school_languages_id).Name

        data = [
            self.row.business,
            self.row.customer,
            self.row.teacher,
            self.row.teaches_classes,
            self.row.teaches_workshops,
            self.row.employee,
            self.row.first_name,
            self.row.last_name,
            self.row.gender,
            self.row.date_of_birth,
            self.row.address,
            self.row.postcode,
            self.row.city,
            self.row.country,
            self.row.email,
            self.row.phone,
            self.row.mobile,
            self.row.emergency,
            self.row.keynr,
            self.row.company,
            discovery,
            level,
            location,
            language,
            self.row.teacher_role,
            self.row.teacher_bio,
            self.row.education,
            self.row.teacher_bio_link,
            self.row.teacher_website,
            self.row.mollie_customer_id,
            self.row.last_login,
            self.row.created_on
        ]

        ws.append(data)


    def _excel_customers_notes(self, db, wb):
        """
            Customers Notes for excel export of customer data
        """
        ws = wb.create_sheet('Notes')

        data = []
        header = [
            'backoffice_note',
            'teacher_note',
            'date',
            'time',
            'note',
            'injury'
        ]

        ws.append(header)

        query = (db.customers_notes.auth_customer_id == self.cuID)
        rows = db(query).select(db.customers_notes.ALL)

        for row in rows:
            data = [
                row.BackofficeNote,
                row.TeacherNote,
                row.NoteDate,
                row.NoteTime,
                row.Note,
                row.Injury,
            ]

            ws.append(data)


    def _excel_alternative_payments(self, db, wb):
        """
            Customers Notes for excel export of customer data
        """
        ws = wb.create_sheet('Subscriptions Alt. payments')

        data = []
        header = [
            'year',
            'month',
            'amount',
            'description'
        ]

        ws.append(header)

        query = (db.alternativepayments.auth_customer_id == self.cuID)
        rows = db(query).select(db.alternativepayments.ALL)

        for row in rows:
            data = [
                row.PaymentYear,
                row.PaymentMonth,
                row.Amount,
                row.Description,
            ]

            ws.append(data)


    def _excel_customers_classcards(self, db, wb):
        """
            Customers classcards for excel export of customer data
        """
        ws = wb.create_sheet('class cards')

        data = []
        header = [
            'card',
            'start',
            'end',
            'note'
        ]

        ws.append(header)

        left = [db.school_classcards.on(
            db.customers_classcards.school_classcards_id ==
            db.school_classcards.id
        )]
        query = (db.customers_classcards.auth_customer_id == self.cuID)
        rows = db(query).select(db.customers_classcards.ALL,
                                db.school_classcards.Name,
                                left=left)

        for row in rows:
            data = [
                row.school_classcards.Name,
                row.customers_classcards.Startdate,
                row.customers_classcards.Enddate,
                row.customers_classcards.Note,
            ]

            ws.append(data)


    def _excel_customers_subscriptions(self, db, wb):
        """
            Customers subscriptions for excel export of customer data
        """
        ws = wb.create_sheet('subscriptions')

        data = []
        header = [
            'subscription',
            'start',
            'end',
            'note',
        ]

        ws.append(header)

        left = [db.school_subscriptions.on(
            db.customers_subscriptions.school_subscriptions_id ==
            db.school_subscriptions.id),
        ]
        query = (db.customers_subscriptions.auth_customer_id == self.cuID)
        rows = db(query).select(db.customers_subscriptions.ALL,
                                db.school_subscriptions.Name,
                                left=left)

        for row in rows:
            data = [
                row.school_subscriptions.Name,
                row.customers_subscriptions.Startdate,
                row.customers_subscriptions.Enddate,
                row.customers_subscriptions.Note,
            ]

            ws.append(data)


    def _excel_customers_payment_info(self, db, wb):
        """
            Customers payment info for excel export of customer data
        """
        ws = wb.create_sheet('payment info')

        data = []
        header = [
            'account_nr',
            'account_holder',
            'bic',
            'mandate_sign_date',
            'bank',
            'bank_locaction'
        ]

        ws.append(header)

        query = (db.customers_payment_info.auth_customer_id == self.cuID)
        rows = db(query).select(db.customers_payment_info.ALL)

        for row in rows:
            data = [
                row.AccountNumber,
                row.AccountHolder,
                row.BIC,
                row.MandateSignatureDate,
                row.BankName,
                row.BankLocation,
            ]

            ws.append(data)


    def _excel_log_customers_accepted_documents(self, db, wb):
        """
            Customers accepted documents for excel export of customer data
        """
        ws = wb.create_sheet('accepted docs')

        data = []
        header = [
            'doc_name',
            'doc_desc',
            'doc_ver',
            'doc_url',
            'os_version',
            'accepted_on',
        ]

        ws.append(header)

        query = (db.log_customers_accepted_documents.auth_customer_id == self.cuID)
        rows = db(query).select(db.log_customers_accepted_documents.ALL)

        for row in rows:
            data = [
                row.DocumentName,
                row.DocumentDescription,
                row.DocumentVersion,
                row.DocumentURL,
                row.OpenStudioVersion,
                row.CreatedOn,
            ]

            ws.append(data)


    def _excel_customers_shoppingcart(self, db, wb):
        """
            Customers accepted documents for excel export of customer data
        """
        ws = wb.create_sheet('shoppingcart')

        data = []
        header = [
            'event_ticket',
            'classcard',
            'class',
            'class_date',
            'att_type',
            'created_on',
        ]

        ws.append(header)

        # left = [
        #     db.workshops_products.on(
        #         db.customers_shoppingcart.workshops_products_id ==
        #         db.workshops_products.id),
        #     db.workshops.on(
        #         db.workshops_products.workshops_id ==
        #         db.workshops.id),
        #     db.school_classcards.on(
        #         db.customers_shoppingcart.school_classcards_id ==
        #         db.school_classcards.id),
        #     db.classes.on(
        #         db.customers_shoppingcart.classes_id ==
        #         db.classes
        #     )
        # ]

        query = (db.customers_shoppingcart.auth_customer_id == self.cuID)
        rows = db(query).select(db.customers_shoppingcart.ALL)

        for row in rows.render():
            data = [
                row.workshops_products_id,
                row.school_classcards_id,
                row.classes_id,
                row.ClassDate,
                row.AttendanceType,
                row.CreatedOn
            ]

            ws.append(data)


    def _excel_customers_orders(self, db, wb):
        """
            Customers orders for excel export of customer data
        """
        ws = wb.create_sheet('orders')

        data = []
        header = [
            'order',
            'status',
            'date_created',
            'classcard',
            'event_ticket',
            'class_id',
            'class_date',
            'att_type',
            'prod_name',
            'desc',
            'qty',
            'price',
            'price_in_vat'
        ]

        ws.append(header)

        left = [
            db.customers_orders.on(
                db.customers_orders_items.customers_orders_id ==
                db.customers_orders.id
            ),
            db.school_classcards.on(
                db.customers_orders_items.school_classcards_id ==
                db.school_classcards.id
            ),
            db.workshops_products.on(
                db.customers_orders_items.workshops_products_id ==
                db.workshops_products.id
            ),
            db.classes.on(
                db.customers_orders_items.classes_id ==
                db.classes.id
            )
        ]

        query = (db.customers_orders.auth_customer_id == self.cuID)
        rows = db(query).select(db.customers_orders.ALL,
                                db.customers_orders_items.ALL,
                                db.school_classcards.Name,
                                db.workshops_products.Name,
                                db.classes.id,
                                left=left)

        for row in rows:
            data = [
                row.customers_orders.id,
                row.customers_orders.Status,
                row.customers_orders.DateCreated,
                row.school_classcards.Name,
                row.workshops_products.Name,
                row.classes.id,
                row.customers_orders_items.ClassDate,
                row.customers_orders_items.AttendanceType,
                row.customers_orders_items.ProductName,
                row.customers_orders_items.Description,
                row.customers_orders_items.Quantity,
                row.customers_orders_items.Price,
                row.customers_orders_items.TotalPriceVAT,
            ]

            ws.append(data)


    def _excel_invoices(self, db, wb):
        """
            Customers invoices for excel export of customer data
        """
        ws = wb.create_sheet('invoices')

        data = []
        header = [
            'invoice',
            'status',
            'date_created',
            'date_due',
            'prod_name',
            'desc',
            'qty',
            'price',
            'price_in_vat'
        ]

        ws.append(header)

        left = [
            db.invoices.on(
                db.invoices_customers.invoices_id ==
                db.invoices.id
            ),
            db.invoices_items.on(
                db.invoices_items.invoices_id ==
                db.invoices.id
            )
        ]

        query = (db.invoices_customers.auth_customer_id == self.cuID)
        rows = db(query).select(db.invoices_customers.ALL,
                                db.invoices.ALL,
                                db.invoices_items.ALL,
                                left=left)

        for row in rows:
            data = [
                row.invoices.InvoiceID,
                row.invoices.Status,
                row.invoices.DateCreated,
                row.invoices.DateDue,
                row.invoices_items.ProductName,
                row.invoices_items.Description,
                row.invoices_items.Quantity,
                row.invoices_items.Price,
                row.invoices_items.TotalPriceVAT,
            ]

            ws.append(data)


    def _excel_classes_attendance(self, db, wb):
        """
            Customers class attendance for excel export of customer data
        """
        ws = wb.create_sheet('class_attendance')

        data = []
        header = [
            'class_id',
            'class_date',
            'att_type',
            'subscription_id',
            'classcard_id',
            'online_booking',
            'booking_status',
            'created_on',
            'created_by',
        ]

        ws.append(header)

        query = (db.classes_attendance.auth_customer_id == self.cuID)
        rows = db(query).select(db.classes_attendance.ALL)

        for row in rows:
            data = [
                row.classes_id,
                row.ClassDate,
                row.AttendanceType,
                row.customers_subscriptions_id,
                row.customers_classcards_id,
                row.online_booking,
                row.BookingStatus,
                row.CreatedOn,
                row.CreatedBy,
            ]

            ws.append(data)


    def _excel_classes_reservation(self, db, wb):
        """
            Customers class enrollment for excel export of customer data
        """
        ws = wb.create_sheet('class_enrollment')

        data = []
        header = [
            'class',
            'start',
            'end',
        ]

        ws.append(header)


        query = (db.classes_reservation.auth_customer_id == self.cuID)
        rows = db(query).select(db.classes_reservation.ALL)

        for row in rows.render():
            data = [
                row.classes_id,
                row.Startdate,
                row.Enddate
            ]

            ws.append(data)


    def _excel_classes_waitinglist(self, db, wb):
        """
            Customers class waitinglist for excel export of customer data
        """
        ws = wb.create_sheet('class_waitinglist')

        data = []
        header = [
            'class',
        ]

        ws.append(header)


        query = (db.classes_waitinglist.auth_customer_id == self.cuID)
        rows = db(query).select(db.classes_waitinglist.ALL)

        for row in rows.render():
            data = [
                row.classes_id,
            ]

            ws.append(data)


    def _excel_workshops_products(self, db, wb):
        """
            Customers event tickets for excel export of customer data
        """
        ws = wb.create_sheet('event_tickets')

        data = []
        header = [
            'event',
            'ticket',
            'cancelled',
            'info_sent',
            'waitinglist',
            'created_on'
        ]

        ws.append(header)

        left = [
            db.workshops_products.on(
                db.workshops_products_customers.workshops_products_id ==
                db.workshops_products.id
            ),
            db.workshops.on(
                db.workshops_products.workshops_id ==
                db.workshops.id
            )
        ]

        query = (db.workshops_products_customers.auth_customer_id == self.cuID)
        rows = db(query).select(db.workshops_products.Name,
                                db.workshops.Name,
                                db.workshops_products_customers.ALL,
                                left=left)

        for row in rows.render():
            data = [
                row.workshops.Name,
                row.workshops_products.Name,
                row.workshops_products_customers.Cancelled,
                row.workshops_products_customers.WorkshopInfo,
                row.workshops_products_customers.Waitinglist,
                row.workshops_products_customers.CreatedOn,
            ]

            ws.append(data)


    def _excel_workshops_activities(self, db, wb):
        """
            Customers event attendance for excel export of customer data
        """
        ws = wb.create_sheet('event_att')

        data = []
        header = [
            'event',
            'ticket',
            'cancelled',
            'info_sent',
            'waitinglist',
            'created_on'
        ]

        ws.append(header)

        left = [
            db.workshops_activities.on(
                db.workshops_activities_customers.workshops_activities_id ==
                db.workshops_activities.id
            ),
            db.workshops.on(
                db.workshops_activities.workshops_id ==
                db.workshops.id
            )
        ]

        query = (db.workshops_activities_customers.auth_customer_id == self.cuID)
        rows = db(query).select(db.workshops_activities.Activity,
                                db.workshops.Name,
                                db.workshops_activities_customers.ALL,
                                left=left)

        for row in rows.render():
            data = [
                row.workshops.Name,
                row.workshops_activities.Activity,
                row.workshops_activities_customers.Cancelled,
                row.workshops_activities_customers.WorkshopInfo,
                row.workshops_activities_customers.Waitinglist,
                row.workshops_activities_customers.CreatedOn,
            ]

            ws.append(data)


    def _excel_messages(self, db, wb):
        """
            Customers messages for excel export of customer data
        """
        ws = wb.create_sheet('messages')

        data = []
        header = [
            'subject',
            'content',
            'sent'
        ]

        ws.append(header)

        left = [
            db.messages.on(
                db.customers_messages.messages_id ==
                db.messages.id),
        ]

        query = (db.customers_messages.auth_customer_id == self.cuID)
        rows = db(query).select(db.messages.ALL,
                                db.customers_messages.ALL,
                                left=left)

        for row in rows.render():
            data = [
                row.messages.msg_subject,
                row.messages.msg_content,
                row.customers_messages.CreatedOn
            ]
            ws.append(data)


    def _excel_payment_batch_items(self, db, wb):
        """
            Customers batch_items for excel export of customer data
        """
        ws = wb.create_sheet('payment_batch_items')

        data = []
        header = [
            'invoices_id',
            'account_holder',
            'bic',
            'account_nr',
            'mandate_sign_date',
            'amount',
            'currency',
            'description',
            'bank',
            'bank_loc',
        ]

        ws.append(header)

        left = [
            db.invoices.on(
                db.payment_batches_items.invoices_id ==
                db.invoices.id
            )
        ]

        query = (db.payment_batches_items.auth_customer_id == self.cuID)
        rows = db(query).select(db.payment_batches_items.ALL,
                                db.invoices.InvoiceID,
                                left=left)

        for row in rows:
            data = [
                row.invoices.InvoiceID,
                row.payment_batches_items.AccountHolder,
                row.payment_batches_items.BIC,
                row.payment_batches_items.AccountNumber,
                row.payment_batches_items.MandateSignatureDate,
                row.payment_batches_items.Amount,
                row.payment_batches_items.Currency,
                row.payment_batches_items.Description,
                row.payment_batches_items.BankName,
                row.payment_batches_items.BankLocation,
            ]

            ws.append(data)