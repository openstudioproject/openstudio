# -*- coding: utf-8 -*-
import datetime
import calendar
import random
import os

from decimal import Decimal, ROUND_HALF_UP

from gluon import *
from general_helpers import get_last_day_month
from general_helpers import workshops_get_full_workshop_product_id
from general_helpers import max_string_length
from general_helpers import NRtoDay
from general_helpers import represent_validity_units


class Customer:
    """
        Class that contains functions for customer
    """
    def __init__(self, cuID):
        """
            Class init function which sets cuID
        """
        db = current.globalenv['db']

        self.cuID = cuID
        self.row = db.auth_user(cuID)


    def get_name(self):
        """
            Returns the name for a customer
        """
        return self.row.display_name


    def get_email_hash(self, hash_type='md5'):
        """

        """
        import hashlib

        md5 = hashlib.md5()
        md5.update(self.row.email.lower())

        return md5.hexdigest()


    def _get_subscriptions_on_date(self, date):
        '''
            Returns subscription for a date
        '''
        db = current.globalenv['db']
        cache = current.globalenv['cache']
        request = current.globalenv['request']
        web2pytest = current.globalenv['web2pytest']

        fields = [
            db.customers_subscriptions.id,
            db.customers_subscriptions.auth_customer_id,
            db.customers_subscriptions.Startdate,
            db.customers_subscriptions.Enddate,
            db.customers_subscriptions.payment_methods_id,
            db.customers_subscriptions.Note,
            db.school_subscriptions.Name,
            db.school_subscriptions.ReconciliationClasses,
            db.school_subscriptions.Unlimited,
            db.customers_subscriptions.CreditsRemaining,
        ]

        sql = '''SELECT cs.id,
                        cs.auth_customer_id,
                        cs.Startdate,
                        cs.Enddate,
                        cs.payment_methods_id,
                        cs.Note,
                        ssu.Name,
                        ssu.ReconciliationClasses,
                        ssu.Unlimited,
(IFNULL(( SELECT SUM(csc.MutationAmount)
 FROM customers_subscriptions_credits csc
 WHERE csc.customers_subscriptions_id = cs.id AND
	   csc.MutationType = 'add'), 0) -
IFNULL(( SELECT SUM(csc.MutationAmount)
 FROM customers_subscriptions_credits csc
 WHERE csc.customers_subscriptions_id = cs.id AND
	   csc.MutationType = 'sub'), 0)) AS credits
FROM customers_subscriptions cs
LEFT JOIN
school_subscriptions ssu ON cs.school_subscriptions_id = ssu.id
WHERE cs.auth_customer_id = {cuID} AND
(cs.Startdate <= '{date}' AND (cs.Enddate >= '{date}' OR cs.Enddate IS NULL))
ORDER BY cs.Startdate'''.format(cuID=self.cuID, date=date)

        rows = db.executesql(sql, fields=fields)

        #print db._lastsql[0]

        if len(rows) > 0:
            return_value = rows
        else:
            return_value = False

        return return_value


    def get_subscriptions_on_date(self, date, from_cache=True):
        '''
            Get day rows with caching
        '''
        web2pytest = current.globalenv['web2pytest']
        request = current.globalenv['request']

        # Don't cache when running tests
        if web2pytest.is_running_under_test(request, request.application) or not from_cache:
            rows = self._get_subscriptions_on_date(date)
            # if rows:
            #     for row in rows:
            #         row.customers_subscriptions.CreditsRemaining = 3456.0
        else:
            cache = current.globalenv['cache']
            DATE_FORMAT = current.globalenv['DATE_FORMAT']
            CACHE_LONG = current.globalenv['CACHE_LONG']
            cache_key = 'openstudio_customer_get_subscriptions_on_date_' + \
                        str(self.cuID) + '_' + \
                        date.strftime(DATE_FORMAT)
            rows = cache.ram(cache_key , lambda: self._get_subscriptions_on_date(date), time_expire=CACHE_LONG)

        return rows


    def has_subscription_on_date(self, date):
        '''
        :param date: datetime.date
        :return: Boolean
        '''
        if len(self.get_subscriptions_on_date(date)) > 0:
            return True
        else:
            return False


    def get_subscription_latest(self):
        '''
            @return: Latest subscription for a customer
        '''
        db = current.globalenv['db']
        os_gui = current.globalenv['os_gui']
        DATE_FORMAT = current.globalenv['DATE_FORMAT']

        fields = [
            db.auth_user.id,
            db.school_subscriptions.Name,
            db.customers_subscriptions.Startdate,
            db.customers_subscriptions.Enddate
        ]

        query = """
            SELECT au.id,
                   ssu.name,
                   cs.startdate,
                   cs.enddate
                   FROM auth_user au
            LEFT JOIN customers_subscriptions cs ON cs.auth_customer_id = au.id
            LEFT JOIN
                (SELECT auth_customer_id,
                        school_subscriptions_id,
                        max(startdate) as startdate,
                        enddate
                FROM customers_subscriptions GROUP BY auth_customer_id) chk
            ON au.id = chk.auth_customer_id
            LEFT JOIN
                (SELECT id, name FROM school_subscriptions) ssu
            ON ssu.id = cs.school_subscriptions_id
            WHERE cs.startdate = chk.startdate
                  AND au.id = {cuID} """.format(cuID=self.cuID)

        # show the latest subscription
        result = db.executesql(query, fields=fields)

        if len(result):
            record = result.first()
            try:
                return SPAN(os_gui.get_fa_icon('fa-clock-o'), ' ',
                            record.school_subscriptions.Name,
                            ' [',
                            record.customers_subscriptions.Startdate.strftime(DATE_FORMAT), ' - ',
                            record.customers_subscriptions.Enddate.strftime(DATE_FORMAT),
                            '] ',
                            _class='small_font',
                            _title='Latest subscription (past)')
            except AttributeError:
                return False

        else:
            return False


    def _get_classcards(self, date):
        '''
            Returns classcards for customer(cuID) on date
        '''
        db = current.globalenv['db']
        cache = current.globalenv['cache']
        request = current.globalenv['request']
        web2pytest = current.globalenv['web2pytest']

        left = [ db.school_classcards.on(
            db.customers_classcards.school_classcards_id==\
            db.school_classcards.id)]
        query = (db.customers_classcards.auth_customer_id == self.cuID) & \
                (db.customers_classcards.Startdate <= date) & \
                ((db.customers_classcards.Enddate >= date) |
                 (db.customers_classcards.Enddate == None)) & \
                ((db.school_classcards.Classes > db.customers_classcards.ClassesTaken) |
                 (db.school_classcards.Classes == 0) |
                 (db.school_classcards.Unlimited == True))

        rows = db(query).select(db.customers_classcards.ALL,
                                db.school_classcards.Name,
                                db.school_classcards.Classes,
                                db.school_classcards.Unlimited,
                                left=left,
                                orderby=db.customers_classcards.Enddate)

        #print rows

        if len(rows) > 0:
            return_value = rows
        else:
            return_value = False

        return return_value


    def get_classcards(self, date, from_cache=True):
        '''
            Get day rows with caching
        '''
        web2pytest = current.globalenv['web2pytest']
        request = current.globalenv['request']

        # Don't cache when running tests
        if web2pytest.is_running_under_test(request, request.application) or not from_cache:
            rows = self._get_classcards(date)
        else:
            cache = current.globalenv['cache']
            DATE_FORMAT = current.globalenv['DATE_FORMAT']
            CACHE_LONG = current.globalenv['CACHE_LONG']
            cache_key = 'openstudio_customer_get_classcards_' + \
                        str(self.cuID) + '_' + \
                        date.strftime(DATE_FORMAT)
            rows = cache.ram(cache_key , lambda: self._get_classcards(date), time_expire=CACHE_LONG)

        return rows


    def get_subscriptions_and_classcards_formatted(self,
                date,
                new_cards=True,
                show_subscriptions=True):
        '''
            Returns a formatted list of subscriptions and classcards for
            a customer
        '''
        DATE_FORMAT = current.globalenv['DATE_FORMAT']
        T = current.globalenv['T']
        os_gui = current.globalenv['os_gui']

        cuID = self.cuID
        subscription = ''
        has_subscription = False
        if show_subscriptions:
            subscriptions = self.get_subscriptions_on_date(date)
            if subscriptions:
                has_subscription = True
                subscription = DIV()
                for cs in subscriptions:
                    csID = cs.customers_subscriptions.id
                    # dates
                    subscr_dates = SPAN(' [', cs.customers_subscriptions.Startdate.strftime(DATE_FORMAT))
                    if cs.customers_subscriptions.Enddate:
                        subscr_dates.append(' - ')
                        subscr_dates.append(cs.customers_subscriptions.Enddate.strftime(DATE_FORMAT))
                    subscr_dates.append('] ')
                    # credits
                    #TODO: Add check for system setting if we should show the credits
                    subscr_credits = ''
                    if cs.customers_subscriptions.CreditsRemaining:
                        subscr_credits = SPAN(XML(' &bull; '), round(cs.customers_subscriptions.CreditsRemaining, 1), ' ',
                                              T('Credits'))
                    subscription.append(SPAN(cs.school_subscriptions.Name, subscr_dates, subscr_credits))

                    csh = CustomerSubscriptionsHelper(csID)
                    paused = csh.get_paused(date)
                    if paused:
                        pause_text = SPAN(' | ', paused, _class='bold')
                        subscription.append(pause_text)
                    subscription.append(BR())


        # get class card for customer
        has_classcard = False
        classcards = self.get_classcards(date)
        if classcards:
            has_classcard = True
            classcard = DIV()
            ccdh  = ClasscardsHelper()
            for ccd in classcards:
                ccdID = ccd.customers_classcards.id
                remaining_classes = ccdh.get_classes_remaining(ccdID)
                if not remaining_classes:
                    continue

                try:
                    enddate = ccd.customers_classcards.Enddate.strftime(DATE_FORMAT)
                except AttributeError:
                    enddate = T('No expiry')

                classcard.append(SPAN(ccd.school_classcards.Name, XML(' &bull; '),
                                 T('expires'), ' ',
                                 enddate, XML(' &bull; '),
                                 remaining_classes))

                if not ccd.school_classcards.Unlimited:
                    classcard.append(SPAN(' ', T("Classes remaining")))

                classcard.append(BR())

            # if remaining_classes == 0 and new_cards:
            #     link_new = A(T("New card"),
            #                  _href=URL('classcard_add',
            #                            vars={'cuID' : cuID},
            #                            extension=''),
            #                  _title=T("Add new card"))
            #     classcard.append(' - ')
            #     classcard.append(link_new)
        else:
            classcard = T("No class card")

        # format data for display
        subscr_cards = TABLE(_class='grey small_font')

        if not has_subscription and not has_classcard:
            if show_subscriptions:
                subscr_cards.append(DIV(T("No subscription or class card"),
                                         _class='red'))
                latest = self.get_subscription_latest()
                subscr_cards.append(latest if latest else '')

        else:
            if subscription and show_subscriptions:
                subscr_cards.append(TR(subscription))
            if classcards:
                subscr_cards.append(TR(classcard))

        return subscr_cards


    def get_had_trialclass(self):
        '''
            Returns True if a customer has had a trialclass and false when not
        '''
        db = current.globalenv['db']

        query = (db.classes_attendance.auth_customer_id == self.cuID) & \
                (db.classes_attendance.AttendanceType == 1)

        count = db(query).count()

        if count > 0:
            had_trial = True
        else:
            had_trial = False

        return had_trial


    def get_workshops_rows(self, upcoming=False):
        """
            Returns workshops for a customer
        """
        db = current.globalenv['db']
        TODAY_LOCAL = current.globalenv['TODAY_LOCAL']

        db_icwspc = db.invoices_workshops_products_customers


        orderby = ~db.workshops.Startdate
        query = (db.workshops_products_customers.auth_customer_id == self.cuID)

        if upcoming:
            query &= (db.workshops.Startdate >= TODAY_LOCAL)
            orderby = db.workshops.Startdate

        rows = db(query).select(
            db.workshops_products_customers.ALL,
            db.workshops.ALL,
            db.workshops_products.Name,
            db.workshops_products.FullWorkshop,
            db.invoices.ALL,
            left=[db.workshops_products.on(
                db.workshops_products.id == \
                db.workshops_products_customers.workshops_products_id),
                db.workshops.on(db.workshops_products.workshops_id == \
                                db.workshops.id),
                db.invoices_workshops_products_customers.on(
                    db_icwspc.workshops_products_customers_id ==
                    db.workshops_products_customers.id),
                db.invoices.on(db_icwspc.invoices_id == db.invoices.id)
            ],
            orderby=~db.workshops.Startdate)

        return rows


    def get_invoices_rows(self):
        """
            Returns invoices records for a customer as gluon.dal.rows object
        """
        db = current.globalenv['db']

        left = [
            db.invoices_amounts.on(
                db.invoices_amounts.invoices_id == db.invoices.id),
            db.invoices_groups.on(
                db.invoices.invoices_groups_id == db.invoices_groups.id),
            db.invoices_customers.on(
                db.invoices_customers.invoices_id ==
                db.invoices.id
            )
        ]
        query = (db.invoices_customers.auth_customer_id == self.cuID) & \
                (db.invoices.Status != 'draft') & \
                (db.invoices_groups.PublicGroup == True)
        rows = db(query).select(db.invoices.ALL,
                                db.invoices_amounts.ALL,
                                left=left,
                                orderby=~db.invoices.DateCreated)

        return rows


    def get_orders_rows(self):
        '''
            Returns orders for a customer
        '''
        db = current.globalenv['db']

        query = (db.customers_orders.auth_customer_id == self.cuID)
        rows = db(query).select(
            db.customers_orders.ALL,
            db.customers_orders_amounts.ALL,
            db.invoices.ALL,
            db.invoices_amounts.ALL,
            left = [ db.customers_orders_amounts.on(db.customers_orders.id ==
                                                    db.customers_orders_amounts.customers_orders_id),
                     db.invoices_customers_orders.on(db.customers_orders.id ==
                                                     db.invoices_customers_orders.customers_orders_id),
                     db.invoices.on(db.invoices.id == db.invoices_customers_orders.invoices_id),
                     db.invoices_amounts.on(db.invoices_amounts.invoices_id == db.invoices.id)],
            orderby = ~db.customers_orders.id
        )

        return rows


    def get_orders_with_items_and_amounts(self):
        '''
            Returns orders info for a customer with additional info
        '''
        db = current.globalenv['db']

        orders = []
        rows = self.get_orders_rows()
        for i, row in enumerate(rows):
            repr_row = list(rows[i:i + 1].render())[0]

            order_obj = Order(row.customers_orders.id)
            order = {}
            order['row'] = row
            order['repr_row'] = repr_row
            order['items'] = order_obj.get_order_items_rows()

            orders.append(order)

        return orders


    def get_documents_rows(self):
        """
        :return: document rows for customer
        """
        db = current.globalenv['db']

        query = (db.customers_documents.auth_customer_id == self.cuID)
        return db(query).select(db.customers_documents.ALL)


    def has_recurring_reservation_for_class(self, clsID, date):
        '''
        :param clsID: db.classes.id
        :param date: datetime.date
        :return: Boolean
        '''
        db = current.globalenv['db']

        query = (db.classes_reservation.auth_customer_id == self.cuID) & \
                (db.classes_reservation.classes_id == clsID) & \
                (db.classes_reservation.Startdate <= date) & \
                ((db.classes_reservation.Enddate >= date) |
                 (db.classes_reservation.Enddate == None)) & \
                (db.classes_reservation.ResType == 'recurring')

        count = db(query).count()

        if count > 0:
            return True
        else:
            return False


    def get_reservations_rows(self, date=None, recurring_only=True):
        '''
            Returns upcoming reservations for this customer
        '''
        db = current.globalenv['db']

        left = [ db.classes.on(db.classes_reservation.classes_id == db.classes.id) ]

        query = (db.classes_reservation.auth_customer_id == self.cuID)
        if date:
            query &= (db.classes_reservation.Startdate <= date) & \
                     ((db.classes_reservation.Enddate >= date) |
                      (db.classes_reservation.Enddate == None))

        if recurring_only:
            query &= (db.classes_reservation.ResType == 'recurring')


        rows = db(query).select(db.classes_reservation.ALL,
                                db.classes.ALL,
                                left=left,
                                orderby=~db.classes_reservation.Startdate)

        return rows


    def get_classes_attendance_rows(self, limit=False, upcoming=False):
        '''
            @param limit: (int) number of attendance records to return
            @return: gluon.dal.rows obj with classes attendance rows for
            customer
        '''
        TODAY_LOCAL = current.globalenv['TODAY_LOCAL']
        db = current.globalenv['db']

        fields = [
            db.classes_attendance.id,
            db.classes_attendance.ClassDate,
            db.classes_attendance.AttendanceType,
            db.classes_attendance.customers_subscriptions_id,
            db.classes_attendance.customers_classcards_id,
            db.classes_attendance.auth_customer_id,
            db.classes_attendance.BookingStatus,
            db.classes.id,
            db.classes.school_locations_id,
            db.classes.school_classtypes_id,
            db.classes.school_levels_id,
            db.classes.Week_day,
            db.classes.Starttime,
            db.classes.Endtime,
            db.classes.Startdate,
            db.classes.Enddate,
            db.invoices.id,
            db.invoices.InvoiceID,
            db.invoices.Status,
            db.invoices.payment_methods_id,
            db.school_classcards.Name
        ]

        where_sql = ''
        if upcoming:
            where_sql = "AND clatt.ClassDate >= '{today}'".format(today=TODAY_LOCAL)
            limit = 20

        limit_sql = ''
        if limit:
            limit_sql = 'LIMIT ' + unicode(limit)

        orderby_sql = 'clatt.ClassDate DESC, cla.Starttime DESC'


        query = '''
        SELECT clatt.id,
               clatt.ClassDate,
               clatt.AttendanceType,
               clatt.customers_subscriptions_id,
               clatt.customers_classcards_id,
               clatt.auth_customer_id,
               clatt.BookingStatus,
               cla.id,
               CASE WHEN cotc.school_locations_id IS NOT NULL
                    THEN cotc.school_locations_id
                    ELSE cla.school_locations_id
                    END AS school_locations_id,
               CASE WHEN cotc.school_classtypes_id IS NOT NULL
                    THEN cotc.school_classtypes_id
                    ELSE cla.school_classtypes_id
                    END AS school_classtypes_id,
               cla.school_levels_id,
               cla.Week_day,
               CASE WHEN cotc.Starttime IS NOT NULL
                    THEN cotc.Starttime
                    ELSE cla.Starttime
                    END AS Starttime,
               CASE WHEN cotc.Endtime IS NOT NULL
                    THEN cotc.Endtime
                    ELSE cla.Endtime
                    END AS Endtime,
               cla.Startdate,
               cla.Enddate,
               inv.id,
               inv.InvoiceID,
               inv.Status,
               inv.payment_methods_id,
               scd.Name
        FROM classes_attendance clatt
        LEFT JOIN classes cla on cla.id = clatt.classes_id
        LEFT JOIN customers_classcards cd ON cd.id = clatt.customers_classcards_id
        LEFT JOIN school_classcards scd ON scd.id = cd.school_classcards_id
        LEFT JOIN
            invoices_classes_attendance ica
            ON ica.classes_attendance_id = clatt.id
        LEFT JOIN
            invoices inv ON ica.invoices_id = inv.id
        LEFT JOIN
            ( SELECT id,
                     classes_id,
                     ClassDate,
                     Status,
                     school_locations_id,
                     school_classtypes_id,
                     Starttime,
                     Endtime,
                     auth_teacher_id,
                     teacher_role,
                     auth_teacher_id2,
                     teacher_role2
              FROM classes_otc ) cotc
            ON clatt.classes_id = cotc.classes_id AND clatt.ClassDate = cotc.ClassDate
        WHERE clatt.auth_customer_id = {cuID}
        {where_sql}
        ORDER BY {orderby_sql}
        {limit_sql}
        '''.format(orderby_sql = orderby_sql,
                   where_sql = where_sql,
                   limit_sql = limit_sql,
                   cuID = self.cuID)

        rows = db.executesql(query, fields=fields)

        # print db._lastsql
        # print rows

        return rows


    def get_shoppingcart_rows(self):
        '''
            Get shopping cart rows for customer
        '''
        db = current.globalenv['db']

        left = [
            db.workshops_products.on(db.workshops_products.id == db.customers_shoppingcart.workshops_products_id),
            db.workshops.on(db.workshops.id == db.workshops_products.workshops_id),
            db.school_classcards.on(db.school_classcards.id == db.customers_shoppingcart.school_classcards_id),
            db.classes.on(db.classes.id == db.customers_shoppingcart.classes_id)
        ]

        query = (db.customers_shoppingcart.auth_customer_id == self.cuID)
        rows = db(query).select(db.customers_shoppingcart.ALL,
                                db.workshops.Name,
                                db.workshops.Startdate,
                                db.workshops_products.id,
                                db.workshops_products.Name,
                                db.workshops_products.Price,
                                db.workshops_products.tax_rates_id,
                                db.school_classcards.id,
                                db.school_classcards.Name,
                                db.school_classcards.Price,
                                db.school_classcards.Classes,
                                db.school_classcards.Unlimited,
                                db.school_classcards.tax_rates_id,
                                db.classes.id,
                                db.classes.school_classtypes_id,
                                db.classes.school_locations_id,
                                db.classes.Starttime,
                                db.classes.Endtime,
                                left=left)

        return rows


    def shoppingcart_maintenance(self):
        '''
            Do some housekeeping to keep things neat and tidy
        '''
        messages = []
        message = self.shoppingcart_remove_past_classes()
        if message:
            messages.append(message)

        return messages


    def shoppingcart_remove_past_classes(self):
        '''
            Check if a class is already past, if so, remove it from the shopping cart.
        '''
        import pytz

        T = current.globalenv['T']
        db = current.globalenv['db']
        now = current.globalenv['NOW_LOCAL']
        TIMEZONE = current.globalenv['TIMEZONE']

        message = False

        query = (db.customers_shoppingcart.auth_customer_id == self.cuID) & \
                (db.customers_shoppingcart.classes_id != None)
        rows = db(query).select(db.customers_shoppingcart.id,
                                db.customers_shoppingcart.classes_id,
                                db.customers_shoppingcart.ClassDate)
        for row in rows:
            cls = Class(row.classes_id, row.ClassDate)

            if cls.is_past():
                del_query = (db.customers_shoppingcart.id == row.id)
                db(query).delete()

                message = T('One past class was removed from your shopping cart')

        return message


    def get_mollie_mandates(self):
        """
            Returns mollie mandates
        """
        get_sys_property = current.globalenv['get_sys_property']

        import Mollie
        # init mollie
        mollie = Mollie.API.Client()
        mollie_api_key = get_sys_property('mollie_website_profile')
        mollie.setApiKey(mollie_api_key)

        # check if we have a mollie customer id
        if self.row.mollie_customer_id:
            mollie_customer_id = self.row.mollie_customer_id
            #print mollie_customer_id
        else:
            # create one
            mollie_customer = mollie.customers.create({
                'name': self.row.display_name,
                'email': self.row.email
            })
            mollie_customer_id = mollie_customer['id']
            self.row.mollie_customer_id = mollie_customer_id
            self.row.update_record()

        return mollie.customer_mandates.withParentId(mollie_customer_id).all()


    def get_accepted_documents(self):
        """
        :return: rows object with rows of accepted documents for this customer
        """
        db = current.globalenv['db']

        query = (db.log_customers_accepted_documents.auth_customer_id == self.cuID)
        rows = db(query).select(db.log_customers_accepted_documents.ALL,
                                orderby=db.log_customers_accepted_documents.CreatedOn)
        return rows


    def log_document_acceptance(self,
                                document_name,
                                document_description='',
                                document_version='',
                                document_url=''):
        """
            :return:
        """
        db = current.globalenv['db']

        version = db.sys_properties(Property='Version').PropertyValue
        release = db.sys_properties(Property='VersionRelease').PropertyValue

        db.log_customers_accepted_documents.insert(
            auth_customer_id = self.cuID,
            DocumentName = document_name,
            DocumentDescription = document_description,
            DocumentVersion = document_version,
            DocumentURL = document_url,
            OpenStudioVersion = '.'.join([version, release])
        )


class CustomerExport:
    def __init__(self, cuID):
        """
            :param cuID: db.auth_user.id
        """
        db = current.globalenv['db']

        self.cuID = cuID
        self.row = db.auth_user(self.cuID)


    def excel(self):
        """
            Customer export all data
        """
        from cStringIO import StringIO
        import openpyxl


        db = current.globalenv['db']

        stream = StringIO()
        # Create the workbook
        wb = openpyxl.workbook.Workbook(write_only=True)

        # Add customer data to workbook
        self._excel_account(db, wb)
        self._excel_customers_notes(db, wb)
        self._excel_alternative_payments(db, wb)
        self._excel_customers_classcards(db, wb)
        self._excel_customers_subscriptions(db, wb)
        self._excel_customers_payment_info(db, wb)
        self._excel_log_customers_accepted_documents(db, wb)
        self._excel_customers_shoppingcart(db, wb)
        self._excel_customers_orders(db, wb)
        self._excel_invoices(db, wb)
        self._excel_classes_attendance(db, wb)
        self._excel_classes_reservation(db, wb)
        self._excel_classes_waitinglist(db, wb)
        self._excel_workshops_products(db, wb)
        self._excel_workshops_activities(db, wb)
        self._excel_messages(db, wb)
        self._excel_payment_batch_items(db, wb)


        wb.save(stream)
        return stream


    def _excel_account(self, db, wb):
        """
            Account info for excel export of customer data
        """
        ws = wb.create_sheet('Account')

        data = []
        header = [
            'business',
            'customer',
            'teacher',
            'teaches_classes',
            'teaches_workshops',
            'employee',
            'first_name',
            'last_name',
            'gender',
            'date_of_birth',
            'address',
            'postcode',
            'city',
            'country',
            'email',
            'phone',
            'mobile',
            'emergency',
            'keynr',
            'company',
            'discovery',
            'level',
            'location',
            'language',
            'teacher_role',
            'teacher_bio',
            'teacher_education',
            'teacher_bio_link',
            'teacher_website',
            'mollie_customer_id',
            'last_login',
            'created_on',
        ]

        ws.append(header)

        discovery = None
        if self.row.school_discovery_id:
            discovery = db.school_discovery(self.row.school_discovery_id).Name

        level = None
        if self.row.school_levels_id:
            level = db.school_levels(self.row.school_levels_id).Name

        location = None
        if self.row.school_locations_id:
            location = db.school_locations(self.row.school_locations_id).Name

        language = None
        if self.row.school_languages_id:
            language = db.school_languages(self.row.school_languages_id).Name

        data = [
            self.row.business,
            self.row.customer,
            self.row.teacher,
            self.row.teaches_classes,
            self.row.teaches_workshops,
            self.row.employee,
            self.row.first_name,
            self.row.last_name,
            self.row.gender,
            self.row.date_of_birth,
            self.row.address,
            self.row.postcode,
            self.row.city,
            self.row.country,
            self.row.email,
            self.row.phone,
            self.row.mobile,
            self.row.emergency,
            self.row.keynr,
            self.row.company,
            discovery,
            level,
            location,
            language,
            self.row.teacher_role,
            self.row.teacher_bio,
            self.row.education,
            self.row.teacher_bio_link,
            self.row.teacher_website,
            self.row.mollie_customer_id,
            self.row.last_login,
            self.row.created_on
        ]

        ws.append(data)


    def _excel_customers_notes(self, db, wb):
        """
            Customers Notes for excel export of customer data
        """
        ws = wb.create_sheet('Notes')

        data = []
        header = [
            'backoffice_note',
            'teacher_note',
            'date',
            'time',
            'note',
            'injury'
        ]

        ws.append(header)

        query = (db.customers_notes.auth_customer_id == self.cuID)
        rows = db(query).select(db.customers_notes.ALL)

        for row in rows:
            data = [
                row.BackofficeNote,
                row.TeacherNote,
                row.NoteDate,
                row.NoteTime,
                row.Note,
                row.Injury,
            ]

            ws.append(data)


    def _excel_alternative_payments(self, db, wb):
        """
            Customers Notes for excel export of customer data
        """
        ws = wb.create_sheet('Subscriptions Alt. payments')

        data = []
        header = [
            'year',
            'month',
            'amount',
            'description'
        ]

        ws.append(header)

        query = (db.alternativepayments.auth_customer_id == self.cuID)
        rows = db(query).select(db.alternativepayments.ALL)

        for row in rows:
            data = [
                row.PaymentYear,
                row.PaymentMonth,
                row.Amount,
                row.Description,
            ]

            ws.append(data)


    def _excel_customers_classcards(self, db, wb):
        """
            Customers classcards for excel export of customer data
        """
        ws = wb.create_sheet('class cards')

        data = []
        header = [
            'card',
            'start',
            'end',
            'note'
        ]

        ws.append(header)

        left = [db.school_classcards.on(
            db.customers_classcards.school_classcards_id ==
            db.school_classcards.id
        )]
        query = (db.customers_classcards.auth_customer_id == self.cuID)
        rows = db(query).select(db.customers_classcards.ALL,
                                db.school_classcards.Name,
                                left=left)

        for row in rows:
            data = [
                row.school_classcards.Name,
                row.customers_classcards.Startdate,
                row.customers_classcards.Enddate,
                row.customers_classcards.Note,
            ]

            ws.append(data)


    def _excel_customers_subscriptions(self, db, wb):
        """
            Customers subscriptions for excel export of customer data
        """
        ws = wb.create_sheet('subscriptions')

        data = []
        header = [
            'subscription',
            'start',
            'end',
            'note',
        ]

        ws.append(header)

        left = [db.school_subscriptions.on(
            db.customers_subscriptions.school_subscriptions_id ==
            db.school_subscriptions.id),
        ]
        query = (db.customers_subscriptions.auth_customer_id == self.cuID)
        rows = db(query).select(db.customers_subscriptions.ALL,
                                db.school_subscriptions.Name,
                                left=left)

        for row in rows:
            data = [
                row.school_subscriptions.Name,
                row.customers_subscriptions.Startdate,
                row.customers_subscriptions.Enddate,
                row.customers_subscriptions.Note,
            ]

            ws.append(data)


    def _excel_customers_payment_info(self, db, wb):
        """
            Customers payment info for excel export of customer data
        """
        ws = wb.create_sheet('payment info')

        data = []
        header = [
            'account_nr',
            'account_holder',
            'bic',
            'mandate_sign_date',
            'bank',
            'bank_locaction'
        ]

        ws.append(header)

        query = (db.customers_payment_info.auth_customer_id == self.cuID)
        rows = db(query).select(db.customers_payment_info.ALL)

        for row in rows:
            data = [
                row.AccountNumber,
                row.AccountHolder,
                row.BIC,
                row.MandateSignatureDate,
                row.BankName,
                row.BankLocation,
            ]

            ws.append(data)


    def _excel_log_customers_accepted_documents(self, db, wb):
        """
            Customers accepted documents for excel export of customer data
        """
        ws = wb.create_sheet('accepted docs')

        data = []
        header = [
            'doc_name',
            'doc_desc',
            'doc_ver',
            'doc_url',
            'os_version',
            'accepted_on',
        ]

        ws.append(header)

        query = (db.log_customers_accepted_documents.auth_customer_id == self.cuID)
        rows = db(query).select(db.log_customers_accepted_documents.ALL)

        for row in rows:
            data = [
                row.DocumentName,
                row.DocumentDescription,
                row.DocumentVersion,
                row.DocumentURL,
                row.OpenStudioVersion,
                row.CreatedOn,
            ]

            ws.append(data)


    def _excel_customers_shoppingcart(self, db, wb):
        """
            Customers accepted documents for excel export of customer data
        """
        ws = wb.create_sheet('shoppingcart')

        data = []
        header = [
            'event_ticket',
            'classcard',
            'class',
            'class_date',
            'att_type',
            'created_on',
        ]

        ws.append(header)

        # left = [
        #     db.workshops_products.on(
        #         db.customers_shoppingcart.workshops_products_id ==
        #         db.workshops_products.id),
        #     db.workshops.on(
        #         db.workshops_products.workshops_id ==
        #         db.workshops.id),
        #     db.school_classcards.on(
        #         db.customers_shoppingcart.school_classcards_id ==
        #         db.school_classcards.id),
        #     db.classes.on(
        #         db.customers_shoppingcart.classes_id ==
        #         db.classes
        #     )
        # ]

        query = (db.customers_shoppingcart.auth_customer_id == self.cuID)
        rows = db(query).select(db.customers_shoppingcart.ALL)

        for row in rows.render():
            data = [
                row.workshops_products_id,
                row.school_classcards_id,
                row.classes_id,
                row.ClassDate,
                row.AttendanceType,
                row.CreatedOn
            ]

            ws.append(data)


    def _excel_customers_orders(self, db, wb):
        """
            Customers orders for excel export of customer data
        """
        ws = wb.create_sheet('orders')

        data = []
        header = [
            'order',
            'status',
            'date_created',
            'classcard',
            'event_ticket',
            'class_id',
            'class_date',
            'att_type',
            'prod_name',
            'desc',
            'qty',
            'price',
            'price_in_vat'
        ]

        ws.append(header)

        left = [
            db.customers_orders.on(
                db.customers_orders_items.customers_orders_id ==
                db.customers_orders.id
            ),
            db.school_classcards.on(
                db.customers_orders_items.school_classcards_id ==
                db.school_classcards.id
            ),
            db.workshops_products.on(
                db.customers_orders_items.workshops_products_id ==
                db.workshops_products.id
            ),
            db.classes.on(
                db.customers_orders_items.classes_id ==
                db.classes.id
            )
        ]

        query = (db.customers_orders.auth_customer_id == self.cuID)
        rows = db(query).select(db.customers_orders.ALL,
                                db.customers_orders_items.ALL,
                                db.school_classcards.Name,
                                db.workshops_products.Name,
                                db.classes.id,
                                left=left)

        for row in rows:
            data = [
                row.customers_orders.id,
                row.customers_orders.Status,
                row.customers_orders.DateCreated,
                row.school_classcards.Name,
                row.workshops_products.Name,
                row.classes.id,
                row.customers_orders_items.ClassDate,
                row.customers_orders_items.AttendanceType,
                row.customers_orders_items.ProductName,
                row.customers_orders_items.Description,
                row.customers_orders_items.Quantity,
                row.customers_orders_items.Price,
                row.customers_orders_items.TotalPriceVAT,
            ]

            ws.append(data)


    def _excel_invoices(self, db, wb):
        """
            Customers invoices for excel export of customer data
        """
        ws = wb.create_sheet('invoices')

        data = []
        header = [
            'invoice',
            'status',
            'date_created',
            'date_due',
            'prod_name',
            'desc',
            'qty',
            'price',
            'price_in_vat'
        ]

        ws.append(header)

        left = [
            db.invoices.on(
                db.invoices_customers.invoices_id ==
                db.invoices.id
            ),
            db.invoices_items.on(
                db.invoices_items.invoices_id ==
                db.invoices.id
            )
        ]

        query = (db.invoices_customers.auth_customer_id == self.cuID)
        rows = db(query).select(db.invoices_customers.ALL,
                                db.invoices.ALL,
                                db.invoices_items.ALL,
                                left=left)

        for row in rows:
            data = [
                row.invoices.InvoiceID,
                row.invoices.Status,
                row.invoices.DateCreated,
                row.invoices.DateDue,
                row.invoices_items.ProductName,
                row.invoices_items.Description,
                row.invoices_items.Quantity,
                row.invoices_items.Price,
                row.invoices_items.TotalPriceVAT,
            ]

            ws.append(data)


    def _excel_classes_attendance(self, db, wb):
        """
            Customers class attendance for excel export of customer data
        """
        ws = wb.create_sheet('class_attendance')

        data = []
        header = [
            'class_id',
            'class_date',
            'att_type',
            'subscription_id',
            'classcard_id',
            'online_booking',
            'booking_status',
            'created_on',
            'created_by',
        ]

        ws.append(header)

        query = (db.classes_attendance.auth_customer_id == self.cuID)
        rows = db(query).select(db.classes_attendance.ALL)

        for row in rows:
            data = [
                row.classes_id,
                row.ClassDate,
                row.AttendanceType,
                row.customers_subscriptions_id,
                row.customers_classcards_id,
                row.online_booking,
                row.BookingStatus,
                row.CreatedOn,
                row.CreatedBy,
            ]

            ws.append(data)


    def _excel_classes_reservation(self, db, wb):
        """
            Customers class enrollment for excel export of customer data
        """
        ws = wb.create_sheet('class_enrollment')

        data = []
        header = [
            'class',
            'start',
            'end',
        ]

        ws.append(header)


        query = (db.classes_reservation.auth_customer_id == self.cuID)
        rows = db(query).select(db.classes_reservation.ALL)

        for row in rows.render():
            data = [
                row.classes_id,
                row.Startdate,
                row.Enddate
            ]

            ws.append(data)


    def _excel_classes_waitinglist(self, db, wb):
        """
            Customers class waitinglist for excel export of customer data
        """
        ws = wb.create_sheet('class_waitinglist')

        data = []
        header = [
            'class',
        ]

        ws.append(header)


        query = (db.classes_waitinglist.auth_customer_id == self.cuID)
        rows = db(query).select(db.classes_waitinglist.ALL)

        for row in rows.render():
            data = [
                row.classes_id,
            ]

            ws.append(data)


    def _excel_workshops_products(self, db, wb):
        """
            Customers event tickets for excel export of customer data
        """
        ws = wb.create_sheet('event_tickets')

        data = []
        header = [
            'event',
            'ticket',
            'cancelled',
            'info_sent',
            'waitinglist',
            'created_on'
        ]

        ws.append(header)

        left = [
            db.workshops_products.on(
                db.workshops_products_customers.workshops_products_id ==
                db.workshops_products.id
            ),
            db.workshops.on(
                db.workshops_products.workshops_id ==
                db.workshops.id
            )
        ]

        query = (db.workshops_products_customers.auth_customer_id == self.cuID)
        rows = db(query).select(db.workshops_products.Name,
                                db.workshops.Name,
                                db.workshops_products_customers.ALL,
                                left=left)

        for row in rows.render():
            data = [
                row.workshops.Name,
                row.workshops_products.Name,
                row.workshops_products_customers.Cancelled,
                row.workshops_products_customers.WorkshopInfo,
                row.workshops_products_customers.Waitinglist,
                row.workshops_products_customers.CreatedOn,
            ]

            ws.append(data)


    def _excel_workshops_activities(self, db, wb):
        """
            Customers event attendance for excel export of customer data
        """
        ws = wb.create_sheet('event_att')

        data = []
        header = [
            'event',
            'ticket',
            'cancelled',
            'info_sent',
            'waitinglist',
            'created_on'
        ]

        ws.append(header)

        left = [
            db.workshops_activities.on(
                db.workshops_activities_customers.workshops_activities_id ==
                db.workshops_activities.id
            ),
            db.workshops.on(
                db.workshops_activities.workshops_id ==
                db.workshops.id
            )
        ]

        query = (db.workshops_activities_customers.auth_customer_id == self.cuID)
        rows = db(query).select(db.workshops_activities.Activity,
                                db.workshops.Name,
                                db.workshops_activities_customers.ALL,
                                left=left)

        for row in rows.render():
            data = [
                row.workshops.Name,
                row.workshops_activities.Activity,
                row.workshops_activities_customers.Cancelled,
                row.workshops_activities_customers.WorkshopInfo,
                row.workshops_activities_customers.Waitinglist,
                row.workshops_activities_customers.CreatedOn,
            ]

            ws.append(data)


    def _excel_messages(self, db, wb):
        """
            Customers messages for excel export of customer data
        """
        ws = wb.create_sheet('messages')

        data = []
        header = [
            'subject',
            'content',
            'sent'
        ]

        ws.append(header)

        left = [
            db.messages.on(
                db.customers_messages.messages_id ==
                db.messages.id),
        ]

        query = (db.customers_messages.auth_customer_id == self.cuID)
        rows = db(query).select(db.messages.ALL,
                                db.customers_messages.ALL,
                                left=left)

        for row in rows.render():
            data = [
                row.messages.msg_subject,
                row.messages.msg_content,
                row.customers_messages.CreatedOn
            ]
            ws.append(data)


    def _excel_payment_batch_items(self, db, wb):
        """
            Customers batch_items for excel export of customer data
        """
        ws = wb.create_sheet('payment_batch_items')

        data = []
        header = [
            'invoices_id',
            'account_holder',
            'bic',
            'account_nr',
            'mandate_sign_date',
            'amount',
            'currency',
            'description',
            'bank',
            'bank_loc',
        ]

        ws.append(header)

        left = [
            db.invoices.on(
                db.payment_batches_items.invoices_id ==
                db.invoices.id
            )
        ]

        query = (db.payment_batches_items.auth_customer_id == self.cuID)
        rows = db(query).select(db.payment_batches_items.ALL,
                                db.invoices.InvoiceID,
                                left=left)

        for row in rows:
            data = [
                row.invoices.InvoiceID,
                row.payment_batches_items.AccountHolder,
                row.payment_batches_items.BIC,
                row.payment_batches_items.AccountNumber,
                row.payment_batches_items.MandateSignatureDate,
                row.payment_batches_items.Amount,
                row.payment_batches_items.Currency,
                row.payment_batches_items.Description,
                row.payment_batches_items.BankName,
                row.payment_batches_items.BankLocation,
            ]

            ws.append(data)


class Customers:
    """
        This clas sontains functions for multiple customers
    """
    def list_activity_after_date(self, date):
        """
            :param: date: datetime.date
            :return: List of all records in auth_user with activity
        """
        db = current.globalenv['db']
        query = """
SELECT au.id, 
	   au.first_name, 
       au.last_name,
       au.display_name, 
       au.email, 
       au.last_login, 
       au.created_on, 
       cs.count_cs,
       ccd.count_ccd,
       wsp.count_event_tickets,
       cn.count_notes,
       clatt.count_classes,
       co.count_orders,
       ic.count_invoices
FROM auth_user au
LEFT JOIN ( SELECT auth_customer_id, COUNT(id) as count_cs
		    FROM customers_subscriptions
		    WHERE Enddate > '{date}' OR Enddate IS NULL
			GROUP BY auth_customer_id ) AS cs ON cs.auth_customer_id = au.id
LEFT JOIN ( SELECT auth_customer_id, COUNT(id) AS count_ccd
		    FROM customers_classcards
		    WHERE Enddate > '{date}' OR Enddate IS NULL 
            GROUP BY auth_customer_id ) AS ccd ON ccd.auth_customer_id = au.id
LEFT JOIN ( SELECT auth_customer_id, COUNT(id) AS count_event_tickets
			FROM workshops_products_customers
			WHERE CreatedOn > '{date}'
            GROUP BY auth_customer_id) AS wsp ON wsp.auth_customer_id = au.id
LEFT JOIN ( SELECT auth_customer_id, COUNT(id) AS count_notes
			FROM customers_notes
			WHERE NoteDate > '{date}'
            GROUP BY auth_customer_id) AS cn ON cn.auth_customer_id = au.id
LEFT JOIN ( SELECT auth_customer_id, COUNT(id) AS count_classes 
			FROM classes_attendance
			WHERE ClassDate > '{date}'
			GROUP BY auth_customer_id) clatt ON clatt.auth_customer_id = au.id
LEFT JOIN ( SELECT auth_customer_id, COUNT(ID) as count_orders
			FROM customers_orders
            WHERE DateCreated > '{date}'
            GROUP BY auth_customer_id) co ON co.auth_customer_id = au.id
LEFT JOIN ( SELECT ic.auth_customer_id, COUNT(ic.id) as count_invoices
			FROM invoices_customers ic
            LEFT JOIN invoices ON ic.invoices_id = invoices.id
            WHERE invoices.DateCreated > '{date}'
            GROUP BY ic.auth_customer_id) ic ON ic.auth_customer_id = au.id
WHERE (au.last_login < '{date}' OR au.last_login IS NULL) AND
      au.created_on < '{date}' AND
      au.employee = 'F' AND 
      au.teacher = 'F' AND
      au.id > 1
        """.format(date=date)

        return db.executesql(query)


    def list_inactive_after_date(self, date):
        """
        :param date: datetime.date
        :return: list of customers inactive after date
        """
        from general_helpers import datestr_to_python

        records = self.list_activity_after_date(date)

        inactive = []
        for record in records:
            last_login = record[5].date() if record[5] else None
            created_on = record[6].date()

            if (created_on < date and
               (last_login is None or last_login < date) and
                    (record[7] is None and
                     record[8] is None and
                     record[9] is None and
                     record[10] is None and
                     record[11] is None and
                     record[12] is None and
                     record[13] is None
                    )):
                inactive.append(record)

        return inactive


    def delete_inactive_after_date(self, date):
        """
        :param date: datetime.date
        :return: Integer - count of customers deleted
        """
        db = current.globalenv['db']

        records = self.list_inactive_after_date(date)
        ids = [record[0] for record in records]

        query = (db.auth_user.id.belongs(ids))
        db(query).delete()

        return len(records)

    
    def list_inactive_after_date_formatted(self, date):
        """
            :param date: datetime.date 
            :return: dict(table=Table listing inactive customers,
                          count=number of inactive customers)
        """
        T = current.globalenv['T']
        DATE_FORMAT = current.globalenv['DATE_FORMAT']

        records = self.list_inactive_after_date(date)

        header = THEAD(TR(
            TH(T("ID")),
            TH(T("Customer")),
            TH(T("Email")),
            TH(T("Last Login")),
            TH(T("Created")),
            TH(T("Subscriptions")),
            TH(T("Classcards")),
            TH(T("Events")),
            TH(T("Notes")),
            TH(T("Classes")),
            TH(T("Orders")),
            TH(T("Invoices")),
        ))

        table = TABLE(header, _class="table table-striped table-hover small_font")
        for record in records:
            cuID = record[0]
            last_login = record[5]
            try:
                record[5].strftime(DATE_FORMAT)
            except AttributeError:
                last_login = 'None'

            customer_link = A(
                record[3],
                _href=URL('customers', 'edit', args=record[0]),
                _target="_blank"
            )

            table.append(TR(
                TD(cuID),
                TD(customer_link),
                TD(record[4]),
                TD(last_login),
                TD(record[6].strftime(DATE_FORMAT)),
                TD(record[7]),
                TD(record[8]),
                TD(record[9]),
                TD(record[10]),
                TD(record[11]),
                TD(record[12]),
                TD(record[13]),
            ))

        return dict(table=table, count=len(records))


class CustomersHelper:
    '''
        This class collects functions for customers that are useful in
        multiple controllers
    '''
    def get_add_modal(self, button_text  = 'Add',
                            button_class = 'btn-sm',
                            redirect_vars= {}):
        '''
            Returns button and modal for an add button
        '''
        os_gui = current.globalenv['os_gui']

        add = LOAD('customers', 'add.load', ajax=True, vars=redirect_vars)

        button_text = XML(SPAN(I(_class='fa fa-plus'), ' ',
                               current.T(button_text)))

        if 'teacher' in redirect_vars:
            modal_title = current.T('Add teacher')
        elif 'employee' in redirect_vars:
            modal_title = current.T('Add employee')
        else:
            modal_title = current.T('Add customer')


        result = os_gui.get_modal(button_text=button_text,
                                  modal_title=modal_title,
                                  modal_content=add,
                                  modal_footer_content=os_gui.get_submit_button('customer_add'),
                                  modal_class='customers_add_new_modal',
                                  button_class=button_class)

        return result


    def get_credits_balance(self, date, include_reconciliation_classes=False):
        '''
        :param date: datetime.date
        :return: Dictionary of customerID's containing current balance and total of reconcilliation credits allowed
        by all subscriptions a customer has on given date
        '''
        db = current.globalenv['db']

        query = '''SELECT cs.id, 
                          cs.auth_customer_id, 
                          ssu.Name, 
                          ssu.ReconciliationClasses, 
                          cs.Startdate, 
                          cs.Enddate,
                          IFNULL((( SELECT SUM(csc.MutationAmount)
                                    FROM customers_subscriptions_credits csc
                                    WHERE csc.customers_subscriptions_id = cs.id AND
	                                csc.MutationType = 'add') - 
                                  ( SELECT SUM(csc.MutationAmount)
                                    FROM customers_subscriptions_credits csc
                                    WHERE csc.customers_subscriptions_id = cs.id AND
	                                csc.MutationType = 'sub')), 0) AS credits
                          FROM customers_subscriptions cs
                          LEFT JOIN 
                            school_subscriptions ssu ON cs.school_subscriptions_id = ssu.id
                          WHERE (cs.Startdate <= '{date}' AND (cs.Enddate >= '{date}' OR cs.Enddate IS NULL))
                          ORDER BY cs.Startdate'''.format(date=date)

        result = db.executesql(query)

        data = {}
        for record in result:
            # add current balance
            try:
                data[record[1]] += record[6]
            except KeyError:
                data[record[1]] = record[6]

            if include_reconciliation_classes:
                # add recon credits
                data[record[1]] += record[3]


        return data


class CustomersSubscriptionsCreditsHelper:
    '''
        Class to group functions related to operations on multiple customer subscriptions credits records
    '''
    def __init__(self):
        '''
            Init functions
        '''
        self.add_credits_balance = {} # Dictionary with customerID as key to hold about of credits added for a customer

    def _get_customers_list_classes_recurring_reservations(self, year, month):
        '''
            Get list of classes a customer has a reservation for in a selected month
        '''
        db = current.globalenv['db']

        ah = AttendanceHelper()
        crh = ClassReservationHelper()
        first_day = datetime.date(year, month, 1)
        last_day = get_last_day_month(first_day)

        data = {}

        date = first_day
        while date <= last_day:
            # get list of classes on date
            #print date

            cs = ClassSchedule(date)
            #print 'getting classes'
            classes = cs.get_day_list()
            reservations = crh.get_recurring_reservations_on_date(date)
            for cls in classes:
                if cls['Cancelled'] or cls['Holiday']:
                    # Class is cancelled or in a holiday, nothing to do
                    continue

                # Get list of bookings with status "attending" or "booked"
                #print 'getting attendance for class'

                attending = []
                rows = ah.get_attendance_rows(cls['ClassesID'], date)
                for row in rows:
                    # print row
                    if row.classes_attendance.BookingStatus == 'booked' or \
                       row.classes_attendance.BookingStatus == 'attending':
                        attending.append(row.auth_user.id)

                # if classes_id found on both lists, add class to reservations list for that customer
                for res in reservations:
                    if res.classes_id == cls['ClassesID']:
                        # add customer to list in case not already attending
                        if not res.auth_customer_id in attending:
                            #print res.auth_customer_id


                            value = {'clsID':cls['ClassesID'],
                                     'date':date}

                            # print value
                            # print '###############'

                            try:
                                data[res.auth_customer_id].append(value)
                            except KeyError:
                                data[res.auth_customer_id] = [value]

            date += datetime.timedelta(days=1)

        return data


    def refund_credits_in_period(self, query):
        '''
            :param query: query containing constraints for period and classes from classes_attendance
            :return: None
        '''
        db = current.globalenv['db']
        rows = db(query).select(db.classes_attendance.id)
        att_ids = []
        for row in rows:
            att_ids.append(row.id)

        # delete rows from customers_subscriptions_credits
        query = (db.customers_subscriptions_credits.classes_attendance_id.belongs(att_ids))
        db(query).delete()


    def add_subscription_credits_month(self,
                                       csID,
                                       cuID,
                                       year,
                                       month,
                                       p_start,
                                       p_end,
                                       classes,
                                       subscription_unit,
                                       batch_add=True,
                                       book_classes=True):
        '''
            Insert subscription credits and clear cache for customer subscription
            :param csID: db.customers_subscriptions.id
            :param cuID: db.auth_user.id
            :param year: int
            :param month: int
            :param p_start: datetime.date (Period start)
            :param p_end: datetime.date (Period end)
            :param classes: int
            :param subscription_unit: string either 'week' or 'month'
            :return: None
        '''
        T = current.globalenv['T']
        db = current.globalenv['db']
        now = current.globalenv['NOW_LOCAL']
        cache_clear_customers_subscriptions = current.globalenv['cache_clear_customers_subscriptions']
        TODAY_LOCAL = current.globalenv['TODAY_LOCAL']

        first_day = datetime.date(year, month, 1)
        last_day = get_last_day_month(first_day)

        t_days = (last_day - first_day) + datetime.timedelta(days=1)  # Total days (Add 1, when subsctraced it's one day less)
        p_days = (p_end - p_start) + datetime.timedelta(days=1)  # Period days

        percent = float(p_days.days) / float(t_days.days)
        if subscription_unit == 'month':
            credits = round(classes * percent, 1)
        else:
            weeks_in_month = round(t_days.days / float(7), 1)
            credits = round((weeks_in_month * (classes or 0)) * percent, 1)

        db.customers_subscriptions_credits.insert(
            customers_subscriptions_id=csID,
            MutationDateTime=now,
            MutationType='add',
            MutationAmount=credits,
            Description=T('Credits') + ' ' + first_day.strftime('%B %Y'),
            SubscriptionYear=year,
            SubscriptionMonth=month
        )

        if batch_add:
            try:
                self.add_credits_balance[cuID] += credits
            except KeyError:
                self.add_credits_balance[cuID] = credits
        else:
            cs = CustomerSubscription(csID)
            self.add_credits_balance[cuID] = cs.get_credits_balance()

        if book_classes:
            # Get list of classes for customer in a given month, based on reservations
            classes_this_month = self.add_credits_reservations.get(cuID, False)
            ah = AttendanceHelper()
            if classes_this_month:
                # Book classess
                while len(self.add_credits_reservations.get(cuID)) > 0 and self.add_credits_balance[cuID] > 0:
                    # Sign in to a class
                    ##
                    # remove this reservation from the list, as we have just booked it, so it won't be booked again using
                    # another subscriptin
                    ##
                    reservation = self.add_credits_reservations[cuID].pop(0) # always get the first in the list, we pop all classes already booked
                    ah.attendance_sign_in_subscription(cuID, reservation['clsID'], csID, reservation['date'])

                    # Subtract one credit from current balance in this object (self.add_credists_balance)
                    self.add_credits_balance[cuID] -= 1




        # Clear cache
        cache_clear_customers_subscriptions(cuID)


    def add_credits_get_subscription_rows_month(self, year, month):
        '''
            return subscription rows for month
        '''
        db = current.globalenv['db']

        first_day = datetime.date(year, month, 1)
        last_day = get_last_day_month(first_day)

        fields = [
            db.customers_subscriptions.id,
            db.customers_subscriptions.Startdate,
            db.customers_subscriptions.Enddate,
            db.customers_subscriptions.auth_customer_id,
            db.customers_subscriptions_credits.id,
            db.customers_subscriptions_paused.id,
            db.school_subscriptions.Name,
            db.school_subscriptions.Classes,
            db.school_subscriptions.SubscriptionUnit,
            db.school_subscriptions.Unlimited,
        ]

        query = '''
            SELECT cs.id,
                   cs.Startdate,
                   cs.Enddate,
                   cs.auth_customer_id,
                   csc.id, 
                   csp.id, 
                   ssu.Name, 
                   ssu.Classes, 
                   ssu.subscriptionunit, 
                   ssu.unlimited
            FROM customers_subscriptions cs
            LEFT JOIN (	SELECT id, SubscriptionYear, SubscriptionMonth, customers_subscriptions_id
                        FROM customers_subscriptions_credits
                        WHERE MutationType = 'add' AND 
                              SubscriptionYear = '{year}' AND 
                              SubscriptionMonth = '{month}' ) csc
                ON csc.customers_subscriptions_id = cs.id
            LEFT JOIN ( SELECT id, customers_subscriptions_id
                        FROM customers_subscriptions_paused 
                        WHERE startdate <= '{last_day}' AND ENDDATE >= '{first_day}' ) csp
                ON csp.customers_subscriptions_id = cs.id
            LEFT JOIN school_subscriptions ssu
                ON cs.school_subscriptions_id = ssu.id
            WHERE cs.Startdate <= '{last_day}' AND (cs.Enddate >= '{first_day}' OR cs.Enddate IS NULL)
        '''.format(year=year,
                   month=month,
                   first_day=first_day,
                   last_day=last_day)

        rows = db.executesql(query, fields=fields)

        return rows


    def add_credits(self, year, month):
        '''
            Add subscription credits for month
        '''
        T = current.globalenv['T']
        db = current.globalenv['db']

        first_day = datetime.date(year, month, 1)
        last_day = get_last_day_month(first_day)

        # Get list of bookable classes for each customer, based on recurring reservations

        self.add_credits_reservations = self._get_customers_list_classes_recurring_reservations(year, month)
        # Get list of total credits balance for each customer
        ch = CustomersHelper()
        self.add_credits_balance = ch.get_credits_balance(first_day, include_reconciliation_classes=True)


        customers_credits_added = 0

        rows = self.add_credits_get_subscription_rows_month(year, month)


        for row in rows:
            if (row.customers_subscriptions_credits.id or
                row.customers_subscriptions_paused.id or
                row.school_subscriptions.Classes is None or
                row.school_subscriptions.Classes == 0 or
                row.school_subscriptions.SubscriptionUnit is None):
                # Don't do anything if this subscription already got credits for this month or is paused
                # or has no classes or subscription unit defined
                continue

            # calculate number of credits
            # only add partial credits if startdate != first day, add full credits if startdate < first day
            if row.customers_subscriptions.Startdate <= first_day:
                p_start = first_day
            else:
                p_start = row.customers_subscriptions.Startdate

            if row.customers_subscriptions.Enddate is None or row.customers_subscriptions.Enddate >= last_day:
                p_end = last_day
            else:
                p_end = row.customers_subscriptions.Enddate


            self.add_subscription_credits_month(
                row.customers_subscriptions.id,
                row.customers_subscriptions.auth_customer_id,
                year,
                month,
                p_start,
                p_end,
                row.school_subscriptions.Classes,
                row.school_subscriptions.SubscriptionUnit,
            )

            # Increase counter
            customers_credits_added += 1

        return customers_credits_added


    def expire_credits(self, date):
        """
        Check if there are any expired credits, if so, add a subtract mutation with the expired amount
        where the 'Expired' field is set to True

        :param date: datetime.date
        :return: number of subscriptions for which credits were expired
        """
        T = current.globalenv['T']
        db = current.globalenv['db']
        NOW_LOCAL = current.globalenv['NOW_LOCAL']
        web2pytest = current.globalenv['web2pytest']
        request = current.globalenv['request']

        # Create dictionary of expiration for school_subscriptions
        subscriptions_count_expired = 0
        query = (db.school_subscriptions.Archived == False)
        rows = db(query).select(db.school_subscriptions.id,
                                db.school_subscriptions.CreditValidity)

        for row in rows:
            if not row.CreditValidity:
                continue

            # Get list of all active subscriptions
            fields = [
                db.customers_subscriptions.id,
                db.customers_subscriptions.auth_customer_id,
                db.customers_subscriptions.Startdate,
                db.customers_subscriptions.Enddate,
                db.customers_subscriptions.payment_methods_id,
                db.school_subscriptions.id,
                db.school_subscriptions.Name,
                db.customers_subscriptions.CreditsRemaining,
                db.customers_subscriptions.PeriodCreditsAdded,
            ]

            if web2pytest.is_running_under_test(request, request.application):
                # the test environment uses sqlite
                mutation_date_sql = "date('{date}', '-{validity} day')".format(
                    date=date,
                    validity=row.CreditValidity
                )
            else:
                # MySQL format
                mutation_date_sql = "DATE_SUB('{date}', INTERVAL {validity} DAY)".format(
                    date=date,
                    validity=row.CreditValidity
                )

            sql = '''SELECT cs.id, 
                            cs.auth_customer_id,
                            cs.Startdate, 
                            cs.Enddate,
                            cs.payment_methods_id,
                            ssu.id,
                            ssu.Name,
                            ( IFNULL((SELECT SUM(csc.MutationAmount)
                               FROM customers_subscriptions_credits csc
                               WHERE csc.customers_subscriptions_id = cs.id AND
                                     csc.MutationType = 'add'), 0) - 
                               IFNULL(( SELECT SUM(csc.MutationAmount)
                               FROM customers_subscriptions_credits csc
                               WHERE csc.customers_subscriptions_id = cs.id AND
                                     csc.MutationType = 'sub'), 0)) AS credits,
                            IFNULL(( SELECT SUM(csc.MutationAmount)
                             FROM customers_subscriptions_credits csc
                             WHERE csc.customers_subscriptions_id = cs.id AND
                                   csc.MutationType = 'add' AND
                                   csc.MutationDateTime >= {mutation_date}), 0) as c_add_in_period
                            FROM customers_subscriptions cs
                            LEFT JOIN 
                            school_subscriptions ssu ON cs.school_subscriptions_id = ssu.id
                            WHERE ssu.id = {ssuID} AND 
                                  (cs.Startdate <= '{date}' AND 
                                  (cs.Enddate >= '{date}' OR cs.Enddate IS NULL))
                            ORDER BY cs.Startdate
                            '''.format(date=date, ssuID=row.id, mutation_date=mutation_date_sql)

            cs_rows = db.executesql(sql, fields=fields)

            for row in cs_rows:

                expired_credits = row.customers_subscriptions.CreditsRemaining - row.customers_subscriptions.PeriodCreditsAdded

                if expired_credits > 0 and row.customers_subscriptions.CreditsRemaining > 0:
                    db.customers_subscriptions_credits.insert(
                        customers_subscriptions_id = row.customers_subscriptions.id,
                        MutationDateTime = NOW_LOCAL,
                        MutationType = 'sub',
                        MutationAmount = expired_credits,
                        Description = T('Credits expiration'),
                        Expiration = True
                    )

                    subscriptions_count_expired += 1

        return subscriptions_count_expired


class CustomerSubscriptionsHelper:
    '''
        Class that contains functions for customer subscriptions
    '''
    def __init__(self, csID):
        '''
            Class init function which sets csID
        '''
        self.csID = csID

    def get_paused(self, date):
        '''
            Returns whether a subscription is paused on provided date
        '''
        db = current.globalenv['db']
        DATE_FORMAT = current.globalenv['DATE_FORMAT']

        query = (db.customers_subscriptions_paused.customers_subscriptions_id ==
                 self.csID) & \
                (db.customers_subscriptions_paused.Startdate <= date) & \
                ((db.customers_subscriptions_paused.Enddate >= date) |
                 (db.customers_subscriptions_paused.Enddate == None))
        row = db(query).select(db.customers_subscriptions_paused.ALL).first()
        if row:
            return_value = SPAN(current.T('Paused until'), ' ',
                                row.Enddate.strftime(DATE_FORMAT))
        else:
            return_value = False

        return return_value


class CustomerSubscription:
    '''
        Class that contains functions for customer subscriptions
    '''
    def __init__(self, csID):
        '''
            Class init function which sets csID
        '''
        db = current.globalenv['db']

        self.csID = csID
        self.cs = db.customers_subscriptions(csID)

        self.ssuID = self.cs.school_subscriptions_id
        self.ssu   = db.school_subscriptions(self.ssuID)

        self.name               = self.ssu.Name
        self.auth_customer_id   = self.cs.auth_customer_id
        self.payment_methods_id = self.cs.payment_methods_id
        self.startdate          = self.cs.Startdate
        self.enddate            = self.cs.Enddate


    def create_invoice_for_month(self, SubscriptionYear, SubscriptionMonth):
        '''
            :param SubscriptionYear: Year of subscription
            :param SubscriptionMonth: Month of subscription
        '''
        db = current.globalenv['db']
        TODAY_LOCAL = current.globalenv['TODAY_LOCAL']
        DATE_FORMAT = current.globalenv['DATE_FORMAT']

        # create invoice linked to subscription for first subscription term to know the right amount.
        SubscriptionYear = TODAY_LOCAL.year
        SubscriptionMonth = TODAY_LOCAL.month

        firstdaythismonth = datetime.date(SubscriptionYear, SubscriptionMonth, 1)
        lastdaythismonth = get_last_day_month(firstdaythismonth)

        # Check if an invoice already exists, if so, return invoice id
        query = (db.invoices.customers_subscriptions_id == self.csID) & \
                (db.invoices.SubscriptionYear == SubscriptionYear) & \
                (db.invoices.SubscriptionMonth == SubscriptionMonth)
        rows = db(query).select(db.invoices.ALL)
        if len(rows):
            return rows.first().id

        # Check if the subscription is paused
        query = (db.customers_subscriptions_paused.customers_subscriptions_id == self.csID) & \
                (db.customers_subscriptions_paused.Startdate <= lastdaythismonth) & \
                ((db.customers_subscriptions_paused.Enddate >= firstdaythismonth) |
                 (db.customers_subscriptions_paused.Enddate == None))
        if db(query).count():
            return

        # Check if an alt. price with amount 0 has been defined
        csap = db.customers_subscriptions_alt_prices
        query = (csap.customers_subscriptions_id == self.csID) & \
                (csap.SubscriptionYear == SubscriptionYear) & \
                (csap.SubscriptionMonth == SubscriptionMonth)
        csap_rows = db(query).select(csap.ALL)
        if csap_rows:
            csap_row = csap_rows.first()
            if csap_row.Amount == 0:
                return

        # Ok we've survived all checks, continue with invoice creation
        igpt = db.invoices_groups_product_types(ProductType='subscription')
        igID = igpt.invoices_groups_id

        if TODAY_LOCAL > firstdaythismonth:
            period_begin = TODAY_LOCAL
        else:
            period_begin = firstdaythismonth

        period_end = lastdaythismonth
        if self.enddate:
            if self.startdate >= firstdaythismonth and \
               self.enddate < lastdaythismonth:
                period_end = self.enddate

        item_description = period_begin.strftime(DATE_FORMAT) + ' - ' + \
                           period_end.strftime(DATE_FORMAT)

        iID = db.invoices.insert(
            invoices_groups_id=igID,
            payment_methods_id=self.payment_methods_id,
            customers_subscriptions_id=self.csID,
            SubscriptionYear=SubscriptionYear,
            SubscriptionMonth=SubscriptionMonth,
            Description='',
            Status='sent'
        )

        # create object to set Invoice# and due date
        invoice = Invoice(iID)
        invoice.link_to_customer(self.auth_customer_id)
        invoice.link_to_customer_subscription(self.csID)
        invoice.item_add_subscription(SubscriptionYear, SubscriptionMonth)

        return iID


    def get_credits_balance(self):
        '''
            Calculate total credits remaining for a subscription
        '''
        db = current.globalenv['db']

        sum = db.customers_subscriptions_credits.MutationAmount.sum()

        query = (db.customers_subscriptions_credits.MutationType == 'add') & \
                (db.customers_subscriptions_credits.customers_subscriptions_id == self.csID)
        add_total = db(query).select(sum).first()[sum] or 0

        query = (db.customers_subscriptions_credits.MutationType == 'sub') & \
                (db.customers_subscriptions_credits.customers_subscriptions_id == self.csID)
        sub_total = db(query).select(sum).first()[sum] or 0

        return round(add_total - sub_total, 1)


    def get_credits_mutations_rows(self,
                                   formatted=False,
                                   editable=False,
                                   deletable=False,
                                   delete_controller='',
                                   delete_function=''):
        '''
            Returns raw rows of credit mutations for a subscription
        '''
        os_gui = current.globalenv['os_gui']
        auth = current.globalenv['auth']
        db = current.globalenv['db']
        T = current.globalenv['T']

        left = [ db.classes_attendance.on(db.customers_subscriptions_credits.classes_attendance_id ==
                                          db.classes_attendance.id),
                 db.classes.on(db.classes_attendance.classes_id ==
                               db.classes.id) ]

        query = (db.customers_subscriptions_credits.customers_subscriptions_id == self.csID)
        rows = db(query).select(db.customers_subscriptions_credits.ALL,
                                db.classes.Starttime,
                                db.classes.Endtime,
                                db.classes.school_locations_id,
                                db.classes.school_classtypes_id,
                                db.classes_attendance.auth_customer_id,
                                db.classes_attendance.ClassDate,
                                left=left,
                                orderby=~db.customers_subscriptions_credits.MutationDateTime)

        if not formatted:
            return rows
        else:
            edit_permission = (auth.has_membership(group_id='Admins') or
                               auth.has_permission('update', 'customers_subscriptions_credits'))

            delete_permission = (auth.has_membership(group_id='Admins') or
                                 auth.has_permission('delete', 'customers_subscriptions_credits'))

            header = THEAD(TR(TH(T('Date')),
                              TH(T('Description')),
                              TH(T('Credits')),
                              TH(db.customers_subscriptions_credits.MutationType.label), # MutationType
                              TH(),
                              ))

            table = TABLE(header, _class='table table-striped table-hover')
            for i, row in enumerate(rows):
                repr_row = list(rows[i:i + 1].render())[0]

                csID = row.customers_subscriptions_credits.customers_subscriptions_id
                cuID = self.auth_customer_id

                delete = ''
                edit = ''
                if deletable and delete_permission:
                    confirm_msg = T("Really delete this mutation?")
                    onclick_del = "return confirm('" + confirm_msg + "');"

                    rvars = {'csID':csID,
                             'cuID':cuID,
                             'cscID':row.customers_subscriptions_credits.id}

                    delete = os_gui.get_button('delete_notext',
                                               URL(delete_controller, delete_function, vars=rvars),
                                               onclick=onclick_del,
                                               _class='pull-right')
                if editable and edit_permission:
                    edit = os_gui.get_button('edit',
                        URL('customers', 'subscription_credits_edit', vars=rvars))

                buttons = DIV(edit, delete, _class='pull-right')

                tr = TR(TD(repr_row.customers_subscriptions_credits.MutationDateTime),
                        TD(repr_row.customers_subscriptions_credits.Description),
                        TD(repr_row.customers_subscriptions_credits.MutationAmount),
                        TD(SPAN(repr_row.customers_subscriptions_credits.MutationType)),
                        TD(buttons))

                table.append(tr)

            return table


    def add_credits_month(self, year, month):
        '''
            Add credits for selected month
        '''
        first_day = datetime.date(year, month, 1)
        last_day = get_last_day_month(first_day)

        if self.cs.Startdate <= first_day:
            p_start = first_day
        else:
            p_start = self.cs.Startdate

        if self.cs.Enddate is None or self.cs.Enddate >= last_day:
            p_end = last_day
        else:
            p_end = self.cs.Enddate

        csch = CustomersSubscriptionsCreditsHelper()
        csch.add_subscription_credits_month(
            self.csID,
            self.cs.auth_customer_id,
            year,
            month,
            p_start,
            p_end,
            self.ssu.Classes,
            self.ssu.SubscriptionUnit,
            batch_add=False,
            book_classes=False)


    def _get_allowed_classes_format(self, class_ids):
        '''
            :param class_ids: list of db.classes.id
            :return: html table
        '''
        T = current.globalenv['T']
        db = current.globalenv['db']
        TODAY_LOCAL = current.globalenv['TODAY_LOCAL']

        query = (db.classes.AllowAPI == True) & \
                (db.classes.id.belongs(class_ids)) & \
                (db.classes.Startdate <= TODAY_LOCAL) & \
                ((db.classes.Enddate == None) |
                 (db.classes.Enddate >= TODAY_LOCAL))
        rows = db(query).select(db.classes.ALL,
                                orderby=db.classes.Week_day|db.classes.Starttime|db.classes.school_locations_id)

        header = THEAD(TR(TH(T('Day')),
                          TH(T('Time')),
                          TH(T('Location')),
                          TH(T('Class'))))
        table = TABLE(header, _class='table table-striped table-hover')
        for i, row in enumerate(rows):
            repr_row = list(rows[i:i + 1].render())[0]

            tr = TR(TD(repr_row.Week_day),
                    TD(repr_row.Starttime, ' - ', repr_row.Endtime),
                    TD(repr_row.school_locations_id),
                    TD(repr_row.school_classtypes_id))

            table.append(tr)

        return table


    def get_allowed_classes_enrollment(self, public_only=True, formatted=False):
        '''
            :return: return: list of db.classes.db that are allowed to be enrolled in using this subscription
        '''
        permissions = self.get_class_permissions(public_only=public_only)
        class_ids = []
        for clsID in permissions:
            try:
                if permissions[clsID]['Enroll']:
                    class_ids.append(clsID)
            except KeyError:
                pass

        if not formatted:
            return class_ids
        else:
            return self._get_allowed_classes_format(class_ids)


    def get_allowed_classes_booking(self, public_only=True, formatted=False):
        '''
            :return: return: list of db.classes.db that are allowed to be booked using this subscription
        '''
        permissions = self.get_class_permissions(public_only=public_only)
        class_ids = []
        for clsID in permissions:
            try:
                if permissions[clsID]['ShopBook']:
                    class_ids.append(clsID)
            except KeyError:
                pass


        if not formatted:
            return class_ids
        else:
            return self._get_allowed_classes_format(class_ids)


    def get_allowed_classes_attend(self, public_only=True, formatted=False):
        '''
            :return: return list of db.classes that are allowed to be attended using this subscription
        '''
        permissions = self.get_class_permissions(public_only=public_only)
        class_ids = []
        for clsID in permissions:
            try:
                if permissions[clsID]['Attend']:
                    class_ids.append(clsID)
            except KeyError:
                pass


        if not formatted:
            return class_ids
        else:
            return self._get_allowed_classes_format(class_ids)


    def _get_class_permissions_format(self, permissions):
        '''
            :param permissions: dictionary of class permissions
            :return: html table
        '''
        T = current.globalenv['T']
        db = current.globalenv['db']
        os_gui = current.globalenv['os_gui']
        TODAY_LOCAL = current.globalenv['TODAY_LOCAL']

        class_ids = []
        for clsID in permissions:
            class_ids.append(clsID)

        query = (db.classes.AllowAPI == True) & \
                (db.classes.id.belongs(class_ids)) & \
                (db.classes.Startdate <= TODAY_LOCAL) & \
                ((db.classes.Enddate == None) |
                 (db.classes.Enddate >= TODAY_LOCAL))
        rows = db(query).select(db.classes.ALL,
                                orderby=db.classes.Week_day|db.classes.Starttime|db.classes.school_locations_id)

        header = THEAD(TR(TH(T('Day')),
                          TH(T('Time')),
                          TH(T('Location')),
                          TH(T('Class')),
                          TH(T('Enroll')),
                          TH(T('Book in advance')),
                          TH(T('Attend'))))

        table = TABLE(header, _class='table table-striped table-hover')
        green_check = SPAN(os_gui.get_fa_icon('fa-check'), _class='text-green')

        for i, row in enumerate(rows):
            repr_row = list(rows[i:i + 1].render())[0]

            class_permission = permissions[row.id]
            enroll = class_permission.get('Enroll', '')
            shopbook = class_permission.get('ShopBook', '')
            attend = class_permission.get('Attend', '')

            if enroll:
                enroll = green_check

            if shopbook:
                shopbook = green_check

            if attend:
                attend = green_check

            tr = TR(TD(repr_row.Week_day),
                    TD(repr_row.Starttime, ' - ', repr_row.Endtime),
                    TD(repr_row.school_locations_id),
                    TD(repr_row.school_classtypes_id),
                    TD(enroll),
                    TD(shopbook),
                    TD(attend))

            table.append(tr)

        return table



    def get_class_permissions(self, public_only=True, formatted=False):
        '''
            :return: return list of class permissons (clsID: enroll, book in shop, attend)
        '''
        db = current.globalenv['db']

        # get groups for subscription
        query = (db.school_subscriptions_groups_subscriptions.school_subscriptions_id == self.ssuID)
        rows = db(query).select(db.school_subscriptions_groups_subscriptions.school_subscriptions_groups_id)

        group_ids = []
        for row in rows:
            group_ids.append(row.school_subscriptions_groups_id)

        # get permissions for subscription group
        left = [db.classes.on(db.classes_school_subscriptions_groups.classes_id == db.classes.id)]
        query = (db.classes_school_subscriptions_groups.school_subscriptions_groups_id.belongs(group_ids))

        if public_only:
            query &= (db.classes.AllowAPI == True)

        rows = db(query).select(db.classes_school_subscriptions_groups.ALL,
                                left=left)

        permissions = {}
        for row in rows:
            clsID = row.classes_id
            if clsID not in permissions:
                permissions[clsID] = {}

            if row.Enroll:
                permissions[clsID]['Enroll'] = True
            if row.ShopBook:
                permissions[clsID]['ShopBook'] = True
            if row.Attend:
                permissions[clsID]['Attend'] = True

        if not formatted:
            return permissions
        else:
            return self._get_class_permissions_format(permissions)

    
class Class:
    '''
        Class that gathers useful functions for a class in OpenStudio
    '''
    def __init__(self, clsID, date):
        self.clsID = clsID
        self.date = date

        db = current.globalenv['db']
        self.cls = db.classes(self.clsID)


    def get_name(self, pretty_date=False):
        '''
            Returns class name formatted for general use
        '''
        db = current.globalenv['db']
        T = current.globalenv['T']
        TIME_FORMAT = current.globalenv['TIME_FORMAT']
        DATE_FORMAT = current.globalenv['DATE_FORMAT']

        if pretty_date:
            date = self.date.strftime('%d %B %Y')
        else:
            date = self.date.strftime(DATE_FORMAT)


        record = self.cls
        location = db.school_locations[record.school_locations_id].Name
        classtype = db.school_classtypes[record.school_classtypes_id].Name
        class_name =  date + ' ' + \
                      record.Starttime.strftime(TIME_FORMAT) + ' - ' + \
                      classtype + ' ' + location

        return class_name


    def get_name_shop(self):
        '''
            Returns class name formatted for use in customer profile and shop
        '''
        db = current.globalenv['db']
        T = current.globalenv['T']
        TIME_FORMAT = current.globalenv['TIME_FORMAT']

        record = self.cls
        location = db.school_locations[record.school_locations_id].Name
        classtype = db.school_classtypes[record.school_classtypes_id].Name
        class_name =  self.date.strftime('%d %B %Y') + ' ' + '<br><small>' + \
                      record.Starttime.strftime(TIME_FORMAT) + ' - ' + \
                      record.Endtime.strftime(TIME_FORMAT) + ' ' + \
                      classtype + ' ' + \
                      T('in') + ' ' + location + '</small>'

        return class_name


    def get_price(self):
        '''
            Returns the price for a class
        '''
        db = current.globalenv['db']

        query = (db.classes_price.classes_id == self.clsID) & \
                (db.classes_price.Startdate <= self.date) & \
                ((db.classes_price.Enddate >= self.date) |
                 (db.classes_price.Enddate == None))
        prices = db(query).select(db.classes_price.ALL,
                                  orderby=db.classes_price.Startdate)

        if prices:
            prices = prices.first()
            dropin = prices.Dropin or 0
            trial  = prices.Trial or 0

            trial_tax = db.tax_rates(prices.tax_rates_id_trial)
            dropin_tax = db.tax_rates(prices.tax_rates_id_dropin)

            try:
                trial_tax_rates_id    = trial_tax.id
                dropin_tax_rates_id   = dropin_tax.id
                trial_tax_percentage  = trial_tax.Percentage
                dropin_tax_percentage = dropin_tax.Percentage
            except AttributeError:
                trial_tax_rates_id    = None
                dropin_tax_rates_id   = None
                trial_tax_percentage  = None
                dropin_tax_percentage = None

        else:
            dropin = 0
            trial  = 0
            trial_tax_rates_id    = None
            dropin_tax_rates_id   = None
            trial_tax_percentage  = None
            dropin_tax_percentage = None


        return dict(trial  = trial,
                    dropin = dropin,
                    trial_tax_rates_id   = trial_tax_rates_id,
                    dropin_tax_rates_id   = dropin_tax_rates_id,
                    trial_tax_percentage  = trial_tax_percentage,
                    dropin_tax_percentage = dropin_tax_percentage)


    def get_full(self):
        '''
            Check whether or not this class is full
        '''
        db = current.globalenv['db']

        spaces = self.cls.Maxstudents

        query = (db.classes_attendance.classes_id == self.clsID) & \
                (db.classes_attendance.ClassDate == self.date) & \
                (db.classes_attendance.BookingStatus != 'cancelled')
        filled = db(query).count()

        full = True if filled >= spaces else False

        return full


    def get_full_bookings_shop(self):
        '''
            Check whether there are spaces left for online bookings
        '''
        db = current.globalenv['db']

        spaces = self.cls.MaxOnlineBooking
        query = (db.classes_attendance.classes_id == self.clsID) & \
                (db.classes_attendance.ClassDate == self.date) & \
                (db.classes_attendance.online_booking == True) & \
                (db.classes_attendance.BookingStatus != 'cancelled')
        filled = db(query).count()

        full = True if filled >= spaces else False

        return full


    def get_invoice_order_description(self, attendance_type):
        '''        
            :return: string with a description of the class 
        '''
        DATE_FORMAT = current.globalenv['DATE_FORMAT']
        TIME_FORMAT = current.globalenv['TIME_FORMAT']

        db = current.globalenv['db']
        T  = current.globalenv['T']

        prices = self.get_price()
        if attendance_type == 1:
            price = prices['trial']
            tax_rates_id = prices['trial_tax_rates_id']
            at = T('Trial')
        elif attendance_type == 2:
            price = prices['dropin']
            tax_rates_id = prices['dropin_tax_rates_id']
            at = T('Drop in')


        location =  db.school_locations(self.cls.school_locations_id)
        classtype = db.school_classtypes(self.cls.school_classtypes_id)

        description = self.date.strftime(DATE_FORMAT) + ' ' + \
                      self.cls.Starttime.strftime(TIME_FORMAT) + ' ' + \
                      classtype.Name + ' ' + \
                      location.Name + ' ' + \
                      '(' + at + ')'

        return description


    def add_to_shoppingcart(self, cuID, attendance_type=2):
        """
            Add a workshop product to the shopping cart of a customer
            attendance_type can be 1 for trial class or 2 for drop in class
        """
        db = current.globalenv['db']

        db.customers_shoppingcart.insert(
            auth_customer_id = cuID,
            classes_id = self.clsID,
            ClassDate = self.date,
            AttendanceType = attendance_type
        )


    def is_on_correct_weekday(self):
        """
            Checks if self.date.isoweekday() == self.cls.Week_day
        """
        if self.date.isoweekday() == self.cls.Week_day:
            return True
        else:
            return False


    def is_past(self):
        """
            Return True if NOW_LOCAL > Class start else return False
        """
        import pytz

        db = current.globalenv['db']
        now = current.globalenv['NOW_LOCAL']
        TIMEZONE = current.globalenv['TIMEZONE']

        cls_time = self.cls.Starttime

        class_dt = datetime.datetime(year=self.date.year,
                                     month=self.date.month,
                                     day=self.date.day,
                                     hour=cls_time.hour,
                                     minute=cls_time.minute)
        # localize the class datetime so it can be compared to now
        # class_dt = pytz.utc.localize(class_dt)
        class_dt = pytz.timezone(TIMEZONE).localize(class_dt)

        if class_dt < now:
            return True
        else:
            return False


    def is_cancelled(self):
        """
            Return True if the class is cancelled, else return False
        """
        db = current.globalenv['db']
        query = (db.classes_otc.classes_id == self.clsID) & \
                (db.classes_otc.ClassDate == self.date) & \
                (db.classes_otc.Status == 'cancelled')

        cancelled = True if db(query).count() else False
        return cancelled


    def is_holiday(self):
        """
            Return True if the class is within a holiday, else return False
        """
        db = current.globalenv['db']

        # Query school_holidays table to see if there's a holiday for this location
        left = [db.school_holidays_locations.on(db.school_holidays.id ==
                                                db.school_holidays_locations.school_holidays_id)]
        query = (db.school_holidays.Startdate <= self.date) & \
                (db.school_holidays.Enddate >= self.date) & \
                (db.school_holidays_locations.school_locations_id == self.cls.school_locations_id)

        rows = db(query).select(db.school_holidays.id,
                                left=left)

        holiday = True if len(rows) else False
        return holiday


    def is_taking_place(self):
        """
             Check if the class is not in past, cancelled or in a holiday
             Return True if not in past, cancelled or in holiday, else return False
        """
        correct_weekday = self.is_on_correct_weekday()
        past = self.is_past()
        cancelled = self.is_cancelled()
        holiday = self.is_holiday()

        if not past and not cancelled and not holiday and correct_weekday:
            return True
        else:
            return False


    def is_booked_by_customer(self, cuID):
        """
        :param cuID: db.auth_user.id
        :return: Boolean

        Check if the class is booked by this customer or not
        """
        db = current.globalenv['db']

        query = ((db.classes_attendance.BookingStatus == 'booked') |
                 (db.classes_attendance.BookingStatus == 'attending')) & \
                (db.classes_attendance.classes_id == self.clsID) & \
                (db.classes_attendance.ClassDate == self.date) & \
                (db.classes_attendance.auth_customer_id == cuID)

        rows = db(query).select(db.classes_attendance.id)
        if len(rows) > 0:
            return True
        else:
            return False


    def has_recurring_reservation_spaces(self):
        '''
        Check whether a class has space for more recurring reservations
        :param date: datetime.date
        :return: Boolean
        '''
        db = current.globalenv['db']

        spaces = self.cls.MaxReservationsRecurring

        query = (db.classes_reservation.classes_id == self.clsID) & \
                (db.classes_reservation.ResType == 'recurring') & \
                (db.classes_reservation.Startdate <= self.date) & \
                ((db.classes_reservation.Enddate >= self.date) |
                 (db.classes_reservation.Enddate == None))

        reservations = db(query).count()

        if reservations >= spaces:
            return False
        else:
            return True


    def get_trialclass_allowed_in_shop(self):
        """
        Check whether trial classes in the shop are allowed or not
        :return: Boolean
        """
        if self.cls.AllowShopTrial:
            return True
        else:
            return False


class ClassReservationHelper:
    '''
        This class collects functions classes_reservation that can return or modify multple records at once
    '''
    def get_recurring_reservations_on_date(self, date, by_class=False):
        '''
        :param date: datetime.date
        :return: rows of all recurring reservations on a given date
        '''
        db = current.globalenv['db']

        query = (db.classes_reservation.Startdate <= date) & \
                ((db.classes_reservation.Enddate >= date) |
                 (db.classes_reservation.Enddate == None)) & \
                (db.classes_reservation.ResType == 'recurring')

        return db(query).select(db.classes_reservation.ALL)


class ClassAttendance:
    '''
        This class collects functions related to a class attendance record
    '''
    def __init__(self, clattID):
        db = current.globalenv['db']
        self.id = clattID
        self.row = db.classes_attendance(clattID)


    def get_datetime_start(self):
        '''
            Returns datetime object of class start
        '''
        db = current.globalenv['db']

        pytz = current.globalenv['pytz']
        TIMEZONE = 'Etc/UTC' # Class times in DB be considered local and shouldn't have extra hours added / subtracted

        cls = db.classes(self.row.classes_id)
        date = self.row.ClassDate
        dt_start = datetime.datetime(date.year,
                                     date.month,
                                     date.day,
                                     int(cls.Starttime.hour),
                                     int(cls.Starttime.minute))
        dt_start = pytz.utc.localize(dt_start).astimezone(pytz.timezone(TIMEZONE))

        return dt_start


    def get_cancel_before(self):
        '''
            Calculates datetime of latest cancellation possibility
        '''
        import math
        db = current.globalenv['db']

        cls = db.classes(self.row.classes_id)
        date = self.row.ClassDate

        get_sys_property = current.globalenv['get_sys_property']

        shop_classes_cancellation_limit = get_sys_property('shop_classes_cancellation_limit') or 0


        dt_start = self.get_datetime_start()
        delta = datetime.timedelta(hours=int(shop_classes_cancellation_limit))

        return dt_start - delta


    def get_cancellation_possible(self):
        '''
             Can we still cancel this booking?
             Allow cancellation when within the configures hours limit and not already attending
        '''
        NOW_LOCAL = current.globalenv['NOW_LOCAL']
        cancel_before = self.get_cancel_before()

        if NOW_LOCAL < cancel_before and not self.row.BookingStatus == 'attending':
            return True
        else:
            return False


    def set_status_cancelled(self, force=False):
        '''
            Set status cancelled
        '''
        T = current.globalenv['T']
        db = current.globalenv['db']
        NOW_LOCAL = current.globalenv['NOW_LOCAL']
        return_message = T('Cancelled class')

        # check hours in advance policy
        if self.get_cancellation_possible() or force:
            # Set booking status to cancelled
            self.row.BookingStatus = 'cancelled'
            self.row.update_record()

            # Remove credits taken from customer for attending a class
            query = (db.customers_subscriptions_credits.classes_attendance_id == self.id)
            db(query).delete()
        else:
            return_message = T("This class can no longer be cancelled")

        return return_message


class AttendanceHelper:
    '''
        This class collects common function for attendance in OpenStudio
    '''
    # def get_attending(self, clsID, date, cuID):
    #     '''
    #         Returns wheter or not a customer is attending a class
    #     '''
    #     db = current.globalenv['db']
    #
    #     attending = db.classes_attendance(classes_id       = clsID,
    #                                       ClassDate        = date,
    #                                       auth_customer_id = cuID)
    #
    #     if attending:
    #         return_value = attending
    #     else:
    #         return_value = False
    #
    #     return return_value


    # def get_attending_list(self, clsID, date):
    #     '''
    #         Return list of customers attending a class as list of
    #         db.auth_user.id
    #     '''
    #     db = current.globalenv['db']
    #
    #     query = (db.classes_attendance.classes_id == clsID) & \
    #             (db.classes_attendance.ClassDate  == date)
    #
    #     rows = db(query).select(db.classes_attendance.auth_customer_id)
    #     attending = []
    #     for row in rows:
    #         if not row.auth_customer_id in attending:
    #             attending.append(row.auth_customer_id)
    #
    #     return attending


    def get_attendance_rows(self, clsID, date):
        '''
            :param clsID: db.classes.db
            :param date: class date
            :return: attendance rows
        '''
        db = current.globalenv['db']

        filter_date_teacher_notes = date - datetime.timedelta(days=92)

        fields = [
            db.auth_user.id,
            db.auth_user.trashed,
            db.auth_user.birthday,
            db.auth_user.thumbsmall,
            db.auth_user.first_name,
            db.auth_user.last_name,
            db.auth_user.display_name,
            db.auth_user.email,
            db.classes_reservation.id,
            db.classes_reservation.ResType,
            db.classes_reservation.Startdate,
            db.classes_reservation.Enddate,
            db.invoices.id,
            db.invoices.InvoiceID,
            db.invoices.Status,
            db.invoices.payment_methods_id,
            db.classes_attendance.id,
            db.classes_attendance.AttendanceType,
            db.classes_attendance.online_booking,
            db.classes_attendance.BookingStatus,
            db.classes_attendance.CreatedOn,
            db.auth_user.teacher_notes_count,  # Holds count of recent teacher notes
            db.auth_user.teacher_notes_count_injuries
        ]

        query = '''
            SELECT au.id,
                   au.archived,
                   au.birthday,
                   au.thumbsmall,
                   au.first_name,
                   au.last_name,
                   au.display_name,
                   au.email, 
                   clr.id,
                   clr.restype,
                   clr.Startdate,
                   clr.Enddate,
                   inv.id,
                   inv.InvoiceID,
                   inv.Status,
                   inv.payment_methods_id,
                   clatt.id,
                   clatt.AttendanceType,
                   clatt.online_booking,
                   clatt.BookingStatus,
                   clatt.CreatedOn,
                   ( SELECT COUNT(*) FROM customers_notes cn 
                     WHERE cn.TeacherNote = 'T' AND 
                           cn.auth_customer_id = au.id AND
                           cn.NoteDate >= '{filter_date_teacher_notes}' ),
                   ( SELECT COUNT(*) FROM customers_notes cn 
                     WHERE cn.TeacherNote = 'T' AND 
                           cn.auth_customer_id = au.id AND
                           cn.Injury = 'T' )
            FROM auth_user au
            LEFT JOIN
                ( SELECT id,
                         auth_customer_id,
                         AttendanceType,
                         online_booking,
                         BookingStatus,
                         CreatedOn
                  FROM classes_attendance
                  WHERE ClassDate = '{date}' AND classes_id = {clsID} ) clatt
                ON clatt.auth_customer_id = au.id
            LEFT JOIN
                ( SELECT id,
                         auth_customer_id,
                         classes_id,
                         Startdate,
                         Enddate,
                         ResType,
                         TrialClass
                  FROM classes_reservation
                  WHERE classes_id = {clsID} AND
                        Startdate <= '{date}' AND
                        (Enddate >= '{date}' or Enddate IS NULL)) clr
                ON clr.auth_customer_id = au.id
            LEFT JOIN
                invoices_classes_attendance ica
                ON ica.classes_attendance_id = clatt.id
            LEFT JOIN
                invoices inv ON ica.invoices_id = inv.id
            WHERE clatt.id IS NOT NULL
            ORDER BY au.display_name
        '''.format(date  = date,
                   filter_date_teacher_notes = filter_date_teacher_notes,
                   clsID = clsID)

        rows = db.executesql(query, fields=fields)

        return rows


    def get_attendance_rows_past_days(self, clsID, date, days):
        '''
            :param clsID: db.classes.id 
            :param date: datetime.date
            :param days: int
            :return: 
        '''
        db = current.globalenv['db']
        cache = current.globalenv['cache']

        cls = Class(clsID, date)

        delta = datetime.timedelta(days=days)
        x_days_ago = date - delta

        left = [ db.classes.on(db.classes_attendance.classes_id ==
                               db.classes.id),
                 db.auth_user.on(db.classes_attendance.auth_customer_id ==
                                 db.auth_user.id) ]

        query = (db.classes.school_classtypes_id == cls.cls.school_classtypes_id) & \
                (db.classes.school_locations_id == cls.cls.school_locations_id) & \
                (db.classes_attendance.classes_id == clsID) & \
                (db.classes_attendance.ClassDate <= date) & \
                (db.classes_attendance.ClassDate >= x_days_ago) & \
                (db.auth_user.trashed == False) & \
                (db.auth_user.customer == True)


        rows = db(query).select(db.auth_user.id,
                                db.auth_user.trashed,
                                db.auth_user.birthday,
                                db.auth_user.thumbsmall,
                                db.auth_user.first_name,
                                db.auth_user.last_name,
                                db.auth_user.display_name,
                                db.auth_user.email,
                                db.classes.school_classtypes_id,
                                left=left,
                                orderby=db.auth_user.display_name,
                                distinct=True,
                                cache=(cache.ram, 30))

        return rows


    def get_reservation_rows(self, clsID, date):
        '''
            :param clsID: db.classes.id 
            :param date: datetime.date
            :return: reservation rows for a class
        '''
        db = current.globalenv['db']

        fields = [
            db.auth_user.id,
            db.auth_user.trashed,
            db.auth_user.birthday,
            db.auth_user.thumbsmall,
            db.auth_user.first_name,
            db.auth_user.last_name,
            db.auth_user.display_name,
            db.auth_user.email,
            db.classes_reservation.id,
            db.classes_reservation.ResType,
            db.classes_reservation.Startdate,
            db.classes_reservation.Enddate,
        ]


        query = '''
            SELECT au.id,
                   au.archived,
                   au.birthday,
                   au.thumbsmall,
                   au.first_name,
                   au.last_name,
                   au.display_name,
                   au.email,
                   clr.id,
                   clr.restype,
                   clr.Startdate,
                   clr.Enddate
            FROM auth_user au
            LEFT JOIN
                ( SELECT id,
                         auth_customer_id
                  FROM classes_attendance
                  WHERE ClassDate = '{date}' AND classes_id = {clsID} ) clatt
                ON clatt.auth_customer_id = au.id
            LEFT JOIN
                ( SELECT id,
                         auth_customer_id,
                         classes_id,
                         Startdate,
                         Enddate,
                         ResType,
                         TrialClass
                  FROM classes_reservation
                  WHERE Startdate <= '{date}' AND
                        (Enddate >= '{date}' or Enddate IS NULL)) clr
                ON clr.auth_customer_id = au.id
            WHERE clr.classes_id = '{clsID}'
            ORDER BY clr.TrialClass DESC, au.display_name
        '''.format(date  = date,
                   clsID = clsID)

        rows = db.executesql(query, fields=fields)

        return rows


    def get_attending_list_between(self,
                                   start_date,
                                   end_date):
        """
            Returns distincs a list of customers attending any classes between start_date
            and end_date as a list of db.auth_user_id
        """
        db = current.globalenv['db']

        left = [ db.auth_user.on(db.classes_attendance.auth_customer_id == db.auth_user.id) ]

        query = (db.classes_attendance.ClassDate >= start_date) & \
                (db.classes_attendance.ClassDate <= end_date) & \
                ((db.classes_attendance.AttendanceType == None) |
                 (db.classes_attendance.AttendanceType == 3))

        rows = db(query).select(db.classes_attendance.auth_customer_id,
                                distinct=True,
                                left=left,
                                orderby=db.auth_user.display_name)

        attending = []
        for row in rows:
            attending.append(row.auth_customer_id)

        return attending


    def get_last_attendance(self, customer_ids):
        """
            For each customer id returns the date when the customer last attended
            a class. Returns a dictionary where the key is the customer id and the
            value is the last date when the customer attended a class.

            :param customer_ids: the customers to check
        """
        db = current.globalenv['db']
        max = db.classes_attendance.ClassDate.max()
        having_query = (db.classes_attendance.auth_customer_id.belongs(customer_ids))
        rows = db().select(db.classes_attendance.auth_customer_id,
            max,
            groupby=db.classes_attendance.auth_customer_id,
            having=having_query)

        last_dates = {}
        for row in rows:
            last_dates[row.classes_attendance.auth_customer_id] = row[max]

        return last_dates


    def get_checkin_list_customers_booked(self,
                                          clsID,
                                          date,
                                          class_full=False,
                                          pictures=True,
                                          reservations=True,
                                          invoices=True,
                                          show_notes=True,
                                          show_booking_time=True,
                                          show_subscriptions=True,
                                          manage_checkin=True):
        '''
            :param clsID: db.classes.id
            :param date: Class date
            :return: Table of customers checked in for this class
        '''
        def add_table_row(row,
                          repr_row,
                          pictures=pictures,
                          #manual_enabled=False,
                          #this_class=False,
                          reservations=reservations,
                          show_subscriptions=show_subscriptions,
                          invoices=invoices,
                          show_notes=show_notes,
                          show_booking_time=show_booking_time,
                          #class_full=False
                          ):
            ''''
                Adds a row to the table
            '''
            cuID = row.auth_user.id

            customer = Customer(cuID)
            subscr_cards = ''
            if show_subscriptions:
                subscr_cards = customer.get_subscriptions_and_classcards_formatted(
                    date,
                    new_cards=False,
                    show_subscriptions=show_subscriptions,
                    )

            # check attendance
            if row.classes_attendance.BookingStatus == 'attending':
                links = []
                # Check update permission
                if (auth.has_membership(group_id='Admins') or
                    auth.has_permission('update', 'classes_attendance')):
                    links = [['header', T('Booking status')]]
                    links.append(A(os_gui.get_fa_icon('fa-check-circle-o'),
                                   T('Booked'), ' ',
                                   _href=URL('classes', 'attendance_set_status',
                                             vars={'clattID':row.classes_attendance.id,
                                                   'status':'booked'}),
                                   _class="text-blue"))
                    links.append(A(os_gui.get_fa_icon('fa-ban'),
                                   T('Cancelled'), ' ',
                                   _href=URL('classes', 'attendance_set_status',
                                             vars={'clattID':row.classes_attendance.id,
                                                   'status':'cancelled'}),
                                   _class="text-yellow"))

                # Check delete permission
                if (auth.has_membership(group_id='Admins') or
                    auth.has_permission('delete', 'classes_attendance')):
                    delete_onclick = "return confirm('" + \
                              T('Do you really want to remove this booking?') + "');"

                    links.append('divider')
                    links.append(A(os_gui.get_fa_icon('fa-minus-circle'),
                                   T('Remove'), ' ',
                                   _href=URL('classes', 'attendance_remove', vars={'clattID':row.classes_attendance.id}),
                                   _onclick=delete_onclick,
                                   _class="text-red"))

                btn = os_gui.get_dropdown_menu(
                    links=links,
                    btn_text=T('Actions'),
                    btn_size='btn-sm',
                    btn_icon='actions',
                    menu_class='btn-group pull-right')

                # btn = DIV(_class='btn-group pull-right')
                attending = SPAN(_class='glyphicon glyphicon-ok green very_big_check')


            else:
                attending = SPAN(_class='glyphicon glyphicon-ok grey-light very_big_check')
                btn = ''
                # Check update permission
                if (auth.has_membership(group_id='Admins') or
                    auth.has_permission('update', 'classes_attendance')):

                    checkin = ''
                    if not class_full:
                        checkin = os_gui.get_button('noicon',
                                                    URL('classes', 'attendance_set_status',
                                                        vars={'clattID':row.classes_attendance.id,
                                                              'status':'attending'}),
                                                    title=T('Check in'))

                    links = [['header', T('Booking status')]]
                    links.append(A(os_gui.get_fa_icon('fa-ban'),
                                   T('Cancelled'), ' ',
                                   _href=URL('classes', 'attendance_set_status',
                                             vars={'clattID':row.classes_attendance.id,
                                                   'status':'cancelled'}),
                                   _class="text-yellow"))

                # Check delete permission
                if (auth.has_membership(group_id='Admins') or
                        auth.has_permission('delete', 'classes_attendance')):
                    delete_onclick = "return confirm('" + \
                                     T('Do you really want to remove this booking?') + "');"

                    links.append('divider')
                    links.append(A(os_gui.get_fa_icon('fa-minus-circle'),
                                   T('Remove'), ' ',
                                   _href=URL('classes', 'attendance_remove',
                                             vars={'clattID': row.classes_attendance.id}),
                                   _onclick=delete_onclick,
                                   _class="text-red"))

                dropdown = os_gui.get_dropdown_menu(
                    links=links,
                    btn_text='',
                    btn_size='btn-sm',
                    btn_icon='actions',
                    menu_class='btn-group pull-right')


                if not manage_checkin:
                    # Remove additional options on check-in button for self checkin
                    btn = DIV(checkin, _class='pull-right')
                else:
                    btn = DIV(checkin, dropdown, _class='btn-group pull-right')


                # if not class_full:
                # btn = self.get_signin_buttons(clsID,
                #                               date,
                #                               cuID,
                #                               manual_enabled=manual_enabled)
                # else:
                #     btn = ''

            # Customer picture
            td_pic = ''
            if pictures:
                td_pic = TD(repr_row.auth_user.thumbsmall,
                            _class='os-customer_image_td hidden-xs')


            td_labels = TD(repr_row.classes_attendance.BookingStatus, _class='hidden-xs')
            if reservations and row.classes_reservation.id:
                date_formatted = date.strftime(DATE_FORMAT)
                crID = row.classes_reservation.id

                td_labels.append(SPAN(' ', repr_row.classes_reservation.ResType))

                try:
                    if row.classes_attendance.AttendanceType == 1:
                        td_labels.append(' ')
                        td_labels.append(os_gui.get_label('success', T('Trial class')))

                    elif row.classes_attendance.AttendanceType == 2:
                        td_labels.append(' ')
                        td_labels.append(os_gui.get_label('primary', T('Drop in')))
                except AttributeError:
                    pass


            if show_booking_time:
                td_labels.append(BR())
                td_labels.append(SPAN(T('Booked on'), ' ', repr_row.classes_attendance.CreatedOn,
                                      _class='vsmall_font grey'))

            # Add a small label for online bookings
            try:
                if row.classes_attendance.online_booking:
                    td_labels.append(TD(os_gui.get_label('info', T('Online'))))
            except AttributeError:
                pass


            ##
            # Invoice for drop in or trial class
            ##
            td_inv = ''
            if invoices:
                ih = InvoicesHelper()
                if row.invoices.id:
                    invoice = ih.represent_invoice_for_list(
                        row.invoices.id,
                        repr_row.invoices.InvoiceID,
                        repr_row.invoices.Status,
                        row.invoices.Status,
                        row.invoices.payment_methods_id
                    )
                else:
                    invoice = ''

                td_inv = TD(invoice)

            ##
            # Link to notes page
            ##
            link_notes = ''
            if show_notes:
                notes_text = T('notes')
                if row.auth_user.teacher_notes_count == 1:
                    notes_text = T('note')
                notes_link_text = SPAN(row.auth_user.teacher_notes_count, ' ', T('Recent'), ' ', notes_text)

                count_injuries = row.auth_user.teacher_notes_count_injuries
                if count_injuries > 0:
                    injuries_text = T('Injuries')
                    if count_injuries == 1:
                        injuries_text = T('Injury')
                    notes_link_text.append(BR())
                    notes_link_text.append(SPAN(count_injuries, ' ', injuries_text, _class='smaller_font text-red bold'))

                link_notes = SPAN(A(notes_link_text,
                                    _href=URL('classes', 'attendance_teacher_notes',
                                              vars={'cuID':row.auth_user.id,
                                                    'clsID':clsID,
                                                    'date':date.strftime(DATE_FORMAT)})))

            # TD('notes', row.auth_user.eetra_field_intxtra_field_int),

            tr = TR(TD(attending, _class='very_big_check'),
                    td_pic,
                    TD(SPAN(row.auth_user.display_name, _class='bold'), BR(),
                       subscr_cards),
                    td_labels,
                    td_inv,
                    TD(link_notes),
                    btn)

            table.append(tr)

        ##
        # Start main function
        ##
        T = current.T
        db = current.globalenv['db']
        auth = current.globalenv['auth']
        os_gui = current.globalenv['os_gui']
        DATE_FORMAT = current.globalenv['DATE_FORMAT']

        modals = DIV()

        cls = db.classes(clsID)

        header = THEAD(TR(TH(),
                          TH(),
                          TH(T('Customer')),
                          TH(T('Status')), # Booking Status [and Enrollment]
                          TH(),
                          TH(),
                          TH()))

        rows = self.get_attendance_rows(clsID, date)
        table = TABLE(header, _class='table table-striped table-hover')

        for i, row in enumerate(rows):
            #print row
            repr_row = list(rows[i:i+1].render())[0]
            add_table_row(row,
                          repr_row,
                          reservations=reservations,
                          show_subscriptions=show_subscriptions,
                          invoices=invoices,
                          show_notes=show_notes,
                          show_booking_time=show_booking_time)

        return table


    def get_checkin_list_customers(self,
                                   clsID,
                                   date,
                                   pictures=False,
                                   manual_enabled=False,
                                   this_class=False,
                                   reservations=False,
                                   reservations_cancel=False,
                                   subscriptions=True,
                                   invoices=False,
                                   show_notes=False,
                                   class_full=False):
        '''
            Returns a list of customers who have a reservation or have attended
            a class of this type in the past month

            this_class
            True: look for attendance for this class in the past month
            False: look for attendance for this classtype in the past month
        '''
        def add_table_row(row,
                          repr_row,
                          reservations=False,
                          invoices=False,
                          show_notes=False,
                          modals=None):
            ''''
                Adds a row to the table
            '''
            cuID = row.auth_user.id

            customer = Customer(cuID)
            subscr_cards = customer.get_subscriptions_and_classcards_formatted(
                date,
                new_cards=False,
                show_subscriptions=subscriptions,
                )

            # check attendance
            if cuID in attendance_list:
                btn = DIV(_class='btn-group pull-right')
                attending = SPAN(_class='glyphicon glyphicon-ok green very_big_check')
                date_formatted = date.strftime(DATE_FORMAT)

                notes = ''
                notes_perm = auth.has_membership(group_id='Admins') or \
                             auth.has_permission('read', 'customers_notes_teachers')
                if show_notes and notes_perm:
                    result = self._get_teachers_note_modal(customer.row.id,
                                                           customer.row.display_name,
                                                           modals)
                    modals = result['modals_div']
                    btn.append(result['button'])

                onclick = "return confirm('" + \
                          T('Really check out?') + "');"
                remove = os_gui.get_button('delete_notext',
                                           URL('classes', 'attendance_remove',
                                               vars={'clsID': clsID,
                                                     'cuID': cuID,
                                                     'date': date_formatted}),
                                           title=T('Cancel'),
                                           btn_class='btn-danger',
                                           btn_size='',
                                           onclick=onclick)
                btn.append(remove)


            else:
                attending = SPAN(_class='glyphicon glyphicon-ok grey-light very_big_check')
                if not class_full:
                    btn = self.get_signin_buttons(clsID,
                                                  date,
                                                  cuID,
                                                  manual_enabled=manual_enabled)
                else:
                    btn = ''

            # Customer picture
            td_pic = ''
            if pictures:
                td_pic = TD(repr_row.auth_user.thumbsmall,
                            _class='os-customer_image_td hidden-xs')

            td_res = ''
            if reservations and row.classes_reservation.id:
                date_formatted = date.strftime(DATE_FORMAT)
                crID = row.classes_reservation.id

                td_res = TD(repr_row.classes_reservation.ResType, _class='hidden-xs')

            td_atttype = ''
            try:
                td_atttype = TD()
                if row.classes_attendance.AttendanceType == 1:
                    td_atttype.append(os_gui.get_label('success', T('Trial class')))

                elif row.classes_attendance.AttendanceType == 2:
                    td_atttype.append(os_gui.get_label('primary', T('Drop in')))
            except AttributeError:
                pass

            # Add a small label for online bookings
            td_online_booking = ''
            try:
                if row.classes_attendance.online_booking:
                    td_online_booking = TD(os_gui.get_label('info', T('Online')))
            except AttributeError:
                pass


            td_inv = ''
            if invoices:
                ih = InvoicesHelper()
                if row.invoices.id:
                    invoice = ih.represent_invoice_for_list(
                        row.invoices.id,
                        repr_row.invoices.InvoiceID,
                        repr_row.invoices.Status,
                        row.invoices.Status,
                        row.invoices.payment_methods_id
                    )
                else:
                    invoice = ''

                td_inv = TD(invoice)


            tr = TR(TD(attending, _class='very_big_check'),
                    td_pic,
                    TD(SPAN(row.auth_user.display_name, _class='bold'), BR(),
                       subscr_cards),
                    td_res,
                    td_atttype,
                    td_online_booking,
                    td_inv,
                    btn)

            table.append(tr)

        # Set some values from the globalenv
        T = current.T
        db = current.globalenv['db']
        auth = current.globalenv['auth']
        os_gui = current.globalenv['os_gui']
        DATE_FORMAT = current.globalenv['DATE_FORMAT']

        modals = DIV()

        cls = db.classes(clsID)

        header = THEAD(TR(TH(),
                          TH(),
                          TH(),
                          TH(), # Enrollment
                          TH(),
                          TH()))


        table = TABLE(header,
                      _class='table table-striped table-hover full-width')

        # ## get list of customers attending this class
        rows = self.get_attendance_rows(clsID, date)

        attendance_list = []
        for i, row in enumerate(rows):
            attendance_list.append(row.auth_user.id)

            repr_row = list(rows[i:i+1].render())[0]
            add_table_row(row,
                          repr_row,
                          reservations=True,
                          invoices=invoices,
                          show_notes=show_notes,
                          modals=modals)


        ## get list of reservations
        rows = self.get_reservation_rows(clsID, date)

        reservations_list = []
        for i, row in enumerate(rows):
            if row.auth_user.id in attendance_list:
                continue

            repr_row = list(rows[i:i+1].render())[0]
            add_table_row(row, repr_row, reservations=True)

            reservations_list.append(row.auth_user.id)


        ## get list of customers who have attended this class during the last 2 weeks
        rows = self.get_attendance_rows_past_days(clsID, date, days=15)

        for i, row in enumerate(rows):
            if row.auth_user.id in attendance_list:
                continue

            if row.auth_user.id in reservations_list:
                continue

            repr_row = list(rows[i:i+1].render())[0]
            add_table_row(row, repr_row, reservations=False)

        return DIV(table, modals)


    def get_checkin_list_customers_email_list(self, clsID, date, days=15):
        '''
            :param clsID: db.classes.is 
            :param date: datetime.date
            :param days: int
            :return: list containing email addresses for all people attending, with reservations or expected to attend
        '''
        # Set some values from the globalenv
        db = current.globalenv['db']

        mailing_list = []
        # ## get list of customers attending this class
        rows = self.get_attendance_rows(clsID, date)

        attendance_list = []
        for i, row in enumerate(rows):
            attendance_list.append(row.auth_user.id)

            mailing_list.append([row.auth_user.first_name, row.auth_user.last_name, row.auth_user.email])


        ## get list of reservations
        rows = self.get_reservation_rows(clsID, date)

        reservations_list = []
        for i, row in enumerate(rows):
            if row.auth_user.id in attendance_list:
                continue

            mailing_list.append([row.auth_user.first_name, row.auth_user.last_name, row.auth_user.email])

            reservations_list.append(row.auth_user.id)


        ## get list of customers who have attended this class during the last x days
        rows = self.get_attendance_rows_past_days(clsID, date, days=days)

        for i, row in enumerate(rows):
            if row.auth_user.id in attendance_list:
                continue

            if row.auth_user.id in reservations_list:
                continue

                mailing_list.append([row.auth_user.first_name, row.auth_user.last_name, row.auth_user.email])

        return mailing_list


    def get_checkin_list_customers_email_excel(self, clsID, date, days=15):
        '''
            :param clsID: db.classes.is 
            :param date: datetime.date
            :param days: int
            :return: cStringIO stream containing: 
                list containing email addresses for all people attending, with reservations or expected to attend
        '''
        T = current.globalenv['T']

        import cStringIO, openpyxl
        stream = cStringIO.StringIO()

        title = T('MailingList')
        wb = openpyxl.workbook.Workbook(write_only=True)
        ws = wb.create_sheet(title=title)

        header = [ "First name",
                   "Last name",
                   "Email" ]
        ws.append(header)

        mailing_list = self.get_checkin_list_customers_email_list(clsID, date, days)
        for row in mailing_list:
            ws.append(row)

        wb.save(stream)


        return stream


    def get_signin_buttons(self, clsID, date, cuID, manual_enabled=True):
        '''
            Returns sign in buttons for a class
        '''
        db = current.globalenv['db']
        os_gui = current.globalenv['os_gui']
        DATE_FORMAT = current.globalenv['DATE_FORMAT']
        date_formatted = date.strftime(DATE_FORMAT)

        customer = Customer(cuID)
        # set random id, used for modal classes
        random_id = unicode(int(random.random() * 1000000000000))

        li_link_class = 'btn btn-default btn-lg full-width'

        button = ''
        btn_group = DIV(_class='btn-group pull-right')
        modals = DIV()
        button_text = current.T('Check in')
        # check if not already added
        check = db.classes_attendance(auth_customer_id = cuID,
                                      classes_id       = clsID,
                                      ClassDate        = date)
        if not check:
            # check for subscription
            subscription = ''
            li_subscription = ''
            li_trial = ''
            li_dropin = ''

            rows = customer.get_subscriptions_on_date(date)
            if rows:
                subscription = rows.first()
                csID = subscription.customers_subscriptions.id
                subscription_url = URL('classes', 'attendance_sign_in_subscription',
                                       vars={'cuID'  : cuID,
                                             'clsID' : clsID,
                                             'csID'  : csID,
                                             'date'  : date_formatted})
                li_subscription = LI(A(SPAN(os_gui.get_fa_icon('fa-edit'), ' ', current.T('Subscription')),
                                        _href=subscription_url,
                                        _class=li_link_class))

            ## check for class card
            classcard = ''
            classcard_choose_url = URL('classes',
                                       'attendance_list_classcards',
                                        vars={'cuID'  : cuID,
                                              'clsID' : clsID,
                                              'date'  : date_formatted},
                                        extension='')
            # set default classcard li, which links to page with add button
            li_classcard = LI(A(SPAN(os_gui.get_fa_icon('fa-ticket'), ' ', current.T('Class card')),
                              _href=classcard_choose_url,
                              _class=li_link_class))


            rows = customer.get_classcards(date)
            if rows:
                classcard_count = len(rows)
                classcard = rows.first()
                ccdID = classcard.customers_classcards.id

                classcard_sign_in_url = URL('classes',
                                            'attendance_sign_in_classcard',
                                            vars={'cuID'  : cuID,
                                                  'clsID' : clsID,
                                                  'ccdID' : ccdID,
                                                  'date'  : date_formatted})

                if classcard_count == 1:
                    classcard_url = classcard_sign_in_url
                else: # more than 1 card, allow user to choose
                    classcard_url = classcard_choose_url

                li_classcard = LI(A(SPAN(os_gui.get_fa_icon('fa-ticket'), ' ', current.T('Class card')),
                                    _href=classcard_url,
                                    _class=li_link_class))


            dropin_url = URL('classes', 'attendance_sign_in_dropin',
                             vars={'cuID'  : cuID,
                                   'clsID' : clsID,
                                   'date'  : date_formatted})
            li_dropin = LI(A(SPAN(os_gui.get_fa_icon('fa-level-down'), ' ', current.T('Drop in class')),
                             _href=dropin_url,
                             _class=li_link_class))
            trial_url = URL('classes', 'attendance_sign_in_trialclass',
                             vars={'cuID'  : cuID,
                                   'clsID' : clsID,
                                   'date'  : date_formatted})
            li_trial = LI(A(SPAN(os_gui.get_fa_icon('fa-compass'), ' ', current.T('Trial class')),
                            _href=trial_url,
                            _class=li_link_class))

            if classcard and subscription:
                modal_content = UL(li_subscription,
                                   li_classcard,
                                   _class='check-in_options')
                modal_class   = 'modal_signin_' + random_id
                button_class  = 'btn btn-default btn-checkin'
                modal_title   = current.T('Check in on subscription or class card?')
                result = os_gui.get_modal(button_text   = button_text,
                                          button_class  = button_class,
                                          modal_title   = modal_title,
                                          modal_content = modal_content,
                                          modal_class   = modal_class)
                btn_group.append(result['button'])
                modals.append(result['modal'])
            elif subscription:
                 # subscription button
                button = A(
                    button_text,
                    _href=subscription_url,
                    _class='btn btn-default btn-checkin')
                btn_group.append(button)
            elif classcard:
                button = A(
                    button_text,
                    _href=classcard_url,
                    _class='btn btn-default btn-checkin')
                btn_group.append(button)
            else:
                # drop in or trial class?
                # classes_attendance customers_id & trialclass attendance type count
                message = SPAN(current.T('No subscription or classcard found for this date.'))
                message.append(' ')
                message.append(current.T("Available options:"))
                modal_content = DIV(message,
                                    BR(), BR(),
                                    UL(li_trial,
                                       li_dropin,
                                       _class='check-in_options'))
                modal_class   = 'modal_signin_' + random_id
                button_class  = 'btn btn-default btn-checkin'
                modal_title   = current.T('Sign in as trial class or drop in class?')
                result = os_gui.get_modal(button_text   = button_text,
                                          button_class  = button_class,
                                          modal_title   = modal_title,
                                          modal_content = modal_content,
                                          modal_class   = modal_class)
                btn_group.append(result['button'])
                modals.append(result['modal'])

            ### Button with modal for manual choice ###

            modal_content = UL(_class='check-in_options')
            if li_subscription:
                modal_content.append(li_subscription)
            if li_classcard:
                modal_content.append(li_classcard)
            if li_dropin:
                modal_content.append(li_dropin)
            if li_trial:
                modal_content.append(li_trial)

            modal_class = 'modal_signin_manual_' + random_id
            button_class = 'btn btn-default pull-right'
            button_text = XML(SPAN(_class='glyphicon glyphicon-edit'))
            modal_title = current.T('Manual check in')
            result = os_gui.get_modal(button_text   = button_text,
                                      button_class  = button_class,
                                      button_title  = current.T("Manual check in"),
                                      modal_title   = modal_title,
                                      modal_content = modal_content,
                                      modal_class   = modal_class)
            if manual_enabled:
                btn_group.append(result['button'])
                modals.append(result['modal'])


        return SPAN(btn_group, modals)


    def get_customer_class_booking_options(self,
                                           clsID,
                                           date,
                                           customer,
                                           trial=False,
                                           complementary=False,
                                           list_type='shop',
                                           controller=''):
        """
        :param clsID: db.classes.id
        :param date: datetime.date
        :param date_formatted: datetime.date object formatted with current.globalenv['DATE_FORMAT']
        :param customer: Customer object
        :param: list_type: [shop, attendance, selfcheckin]
        :return:
        """
        def classes_book_options_get_button_book(url):
            """
                Return book button for booking options
            """
            button_text = T('Book')
            if list_type == 'attendance' or list_type == 'selfcheckin':
                button_text = T('Check in')

            button_book = A(SPAN(button_text, ' ', os_gui.get_fa_icon('fa-chevron-right')),
                            _href=url,
                            _class='pull-right btn btn-link')

            return button_book

        T = current.globalenv['T']
        os_gui = current.globalenv['os_gui']
        CURRSYM = current.globalenv['CURRSYM']
        DATE_FORMAT = current.globalenv['DATE_FORMAT']
        get_sys_property = current.globalenv['get_sys_property']

        date_formatted = date.strftime(DATE_FORMAT)

        options = DIV(_class='shop-classes-booking-options row')
        # subscriptions
        subscriptions = customer.get_subscriptions_on_date(date)
        if subscriptions:
            for subscription in subscriptions:
                csID = subscription.customers_subscriptions.id
                # Shop urls are the default for this function when no list_type has been specified
                # Check remaining credits
                credits_remaining = subscription.customers_subscriptions.CreditsRemaining or 0
                recon_classes = subscription.school_subscriptions.ReconciliationClasses
                # Create subscription object
                cs = CustomerSubscription(csID)

                if list_type == 'shop':
                    subscription_permission_check = not int(clsID) in cs.get_allowed_classes_booking(public_only=True)
                else:
                    subscription_permission_check = not int(clsID) in cs.get_allowed_classes_attend(public_only=False)

                if subscription_permission_check:
                    # Check book permission
                    button_book = os_gui.get_button('noicon',
                                                    URL('#'),
                                                    title=SPAN(T("Not allowed for this class")),
                                                    btn_class='btn-link',
                                                    _class='disabled pull-right grey')
                else:

                    if credits_remaining > (recon_classes * -1):
                        url = URL(controller, 'class_book', vars={'clsID': clsID,
                                                      'csID': csID,
                                                      'cuID': customer.row.id,
                                                      'date': date_formatted})
                        button_book = classes_book_options_get_button_book(url)
                    else:
                        button_book = os_gui.get_button('noicon',
                                                        URL('#'),
                                                        title=SPAN(T('No credits remaining')),
                                                        btn_class='btn-link',
                                                        _class='disabled pull-right grey')

                # Check Credits display
                if subscription.school_subscriptions.Unlimited:
                    credits_display = T('Unlimited')
                else:
                    if credits_remaining < 0:
                        credits_display = SPAN(round(credits_remaining, 1), ' ', T('Credits'))
                    else:
                        credits_display = SPAN(round(credits_remaining, 1), ' ',
                                               T('Credits remaining'))

                # let's put everything together
                option = DIV(DIV(T('Subscription'),
                                 _class='col-md-3 bold'),
                             DIV(subscription.school_subscriptions.Name,
                                 SPAN(XML(' &bull; '),
                                      credits_display,
                                      _class='grey'),
                                 _class='col-md-6'),
                             DIV(button_book,
                                 _class='col-md-3'),
                             _class='col-md-10 col-md-offset-1 col-xs-12')

                options.append(option)

        # class cards
        classcards = customer.get_classcards(date)
        if classcards:
            for classcard in classcards:
                ccdID = classcard.customers_classcards.id

                ccd = Classcard(ccdID)
                classes_remaining = ccd.get_classes_remaining_formatted()

                if list_type == 'shop':
                    allowed_classes = ccd.get_allowed_classes_booking()
                else:
                    allowed_classes = ccd.get_allowed_classes_attend(public_only=False)

                if not int(clsID) in allowed_classes:
                    # Check book permission
                    button_book = os_gui.get_button('noicon',
                                                    URL('#'),
                                                    title=SPAN(T("Not allowed for this class")),
                                                    btn_class='btn-link',
                                                    _class='disabled pull-right grey')
                else:
                    url = URL(controller, 'class_book', vars={'clsID': clsID,
                                                  'ccdID': ccdID,
                                                  'cuID': customer.row.id,
                                                  'date': date_formatted})
                    button_book = classes_book_options_get_button_book(url)

                option = DIV(DIV(T('Class card'),
                                 _class='col-md-3 bold'),
                             DIV(classcard.school_classcards.Name, ' ',
                                 SPAN(XML(' &bull; '), T('expires'), ' ',
                                      classcard.customers_classcards.Enddate.strftime(DATE_FORMAT),
                                      XML(' &bull; '), classes_remaining,
                                      _class='small_font grey'),
                                 _class='col-md-6'),
                             DIV(button_book,
                                 _class='col-md-3'),
                             _class='col-md-10 col-md-offset-1 col-xs-12')

                options.append(option)
        else:
            if list_type == 'attendance':
                url = URL('customers', 'classcard_add',
                          vars={'cuID': customer.row.id,
                                'clsID': clsID,
                                'date': date_formatted})
                button_add = A(SPAN(T('Sell card'), ' ', os_gui.get_fa_icon('fa-chevron-right')),
                                _href=url,
                                _class='pull-right btn btn-link')


                option = DIV(DIV(T('Class card'),
                                 _class='col-md-3 bold'),
                             DIV(T('No cards found - sell a new card',),
                                 _class='col-md-6'),
                             DIV(button_add,
                                 _class='col-md-3'),
                             _class='col-md-10 col-md-offset-1 col-xs-12')

                options.append(option)

        # Get class prices
        cls = Class(clsID, date)
        prices = cls.get_price()

        # drop in
        url = URL(controller, 'class_book', vars={'clsID': clsID,
                                      'dropin': 'true',
                                      'cuID': customer.row.id,
                                      'date': date_formatted})
        button_book = classes_book_options_get_button_book(url)

        option = DIV(DIV(T('Drop in'),
                         _class='col-md-3 bold'),
                     DIV(T('Class price:'), ' ', CURRSYM, ' ', format(prices['dropin'], '.2f'),
                         BR(),
                         SPAN(get_sys_property('shop_classes_dropin_message') or '',
                              _class='grey'),
                         _class='col-md-6'),
                     DIV(button_book,
                         _class='col-md-3'),
                     _class='col-md-10 col-md-offset-1 col-xs-12')

        options.append(option)

        # Trial
        # get trial class price
        if trial:
            url = URL(controller, 'class_book', vars={'clsID': clsID,
                                                      'trial': 'true',
                                                      'cuID': customer.row.id,
                                                      'date': date_formatted})
            button_book = classes_book_options_get_button_book(url)

            option = DIV(DIV(T('Trial'),
                             _class='col-md-3 bold'),
                         DIV(T('Class price:'), ' ', CURRSYM, ' ', format(prices['trial'], '.2f'),
                             BR(),
                             SPAN(get_sys_property('shop_classes_trial_message') or '',
                                  _class='grey'),
                             _class='col-md-6'),
                         DIV(button_book,
                             _class='col-md-3'),
                         _class='col-md-10 col-md-offset-1 col-xs-12')

            options.append(option)

        # Complementary
        if complementary:
            options.append(DIV(HR(), _class='col-md-10 col-md-offset-1'))
            url = URL(controller, 'class_book', vars={'clsID': clsID,
                                                      'complementary': 'true',
                                                      'cuID': customer.row.id,
                                                      'date': date_formatted})
            button_book = classes_book_options_get_button_book(url)

            option = DIV(DIV(T('Complementary'),
                             _class='col-md-3 bold'),
                         DIV(T('Give this class for free'),
                             _class='col-md-6'),
                         DIV(button_book,
                             _class='col-md-3'),
                         _class='col-md-10 col-md-offset-1 col-xs-12')

            options.append(option)

        return options


    def _get_teachers_note_modal(self,
                                 cuID,
                                 customers_name,
                                 modals_div):
        '''
            Returns a modal popup for teacher notes
        '''
        T = current.globalenv['T']
        db = current.globalenv['db']
        os_gui = current.globalenv['os_gui']

        notes = LOAD('customers', 'notes.load', ajax=True,
                     vars={'cuID':cuID,
                           'note_type':'teachers'})

        modal_class = 'customers_te_notes_' + unicode(cuID)
        modal_title = SPAN(T('Teacher notes for'), ' ', customers_name)

        query = (db.customers_notes.TeacherNote == True) & \
                (db.customers_notes.auth_customer_id == cuID)
        count_notes = db(query).count()

        notes_text = T('Notes')
        if count_notes == 1:
            notes_text = T('Note ')

        notes_text = SPAN(unicode(count_notes), ' ', notes_text)


        #button_text = XML(SPAN(os_gui.get_fa_icon('fa-sticky-note-o'), ' ', notes_text))
        button_text = XML(notes_text)

        result = os_gui.get_modal(button_text=button_text,
                                  modal_title=modal_title,
                                  modal_content=notes,
                                  modal_class=modal_class,
                                  button_class='btn btn-default btn-checkin')
        modals_div.append(result['modal'])
        button = result['button']

        return dict(modals_div = modals_div,
                    button=button)


    def attendance_sign_in_subscription(self,
                                        cuID,
                                        clsID,
                                        csID,
                                        date,
                                        online_booking=False,
                                        credits_hard_limit=False,
                                        booking_status='booked'):
        '''
            :param cuID: db.auth_user.id
            :param clsID: db.classes.id 
            :param csID: db.customers_subscriptions.id
            :param date: datetime.date
            :return: dict status[ok|fail], message
        '''
        db = current.globalenv['db']
        T = current.globalenv['T']
        DATE_FORMAT = current.globalenv['DATE_FORMAT']
        cache_clear_customers_subscriptions = current.globalenv['cache_clear_customers_subscriptions']


        status = 'fail'
        message = ''
        signed_in = self.attendance_sign_in_check_signed_in(clsID, cuID, date)
        # check credits remaining

        credits_remaining = self._attendance_sign_in_subscription_credits_remaining(csID)
        message_no_credits = T('No credits remaining on this subscription')

        if signed_in:
            message = T("Customer is already checked in")
        elif not credits_remaining and credits_hard_limit:
            # return message, don't sign in
            message = message_no_credits
        else:
            #print 'signing in customer'

            status = 'ok'
            clattID = db.classes_attendance.insert(
                auth_customer_id           = cuID,
                classes_id                 = clsID,
                ClassDate                  = date,
                AttendanceType             = None, # None = subscription
                customers_subscriptions_id = csID,
                online_booking=online_booking,
                BookingStatus=booking_status
                )

            # Take 1 credit
            cls = Class(clsID, date)
            cscID = db.customers_subscriptions_credits.insert(
                customers_subscriptions_id = csID,
                classes_attendance_id = clattID,
                MutationType = 'sub',
                MutationAmount = '1',
                Description = cls.get_name(pretty_date=True)
            )

            #print cscID

            cache_clear_customers_subscriptions(cuID)

            # check subscription classes exceeded
            #result = self._attendance_sign_in_subscription_check_classes_exceeded(csID, clattID, date)
            #if result:
            #    message = result

            # # check credits remaining
            if not credits_remaining:
                message = message_no_credits
            # result = self._attendance_sign_in_subscription_credits_remaining(csID)
            # if result:
            #     message = result

            # check for paused subscription
            result = self._attedance_sign_in_subscription_check_paused(csID, date)
            if result:
                message = result

        return dict(status=status, message=message)


    def _attendance_sign_in_subscription_credits_remaining(self, csID):
        '''
        Check if this subscription has remaining classes, if not, set message

        :param csID:
        :param clattID:
        :param date:
        :return:
        '''
        T = current.globalenv['T']
        db = current.globalenv['db']

        cs = CustomerSubscription(csID)
        balance = cs.get_credits_balance()
        recon_classes = cs.ssu.ReconciliationClasses

        credits_remaining = balance > (recon_classes * -1)

        return credits_remaining


    # def _attendance_sign_in_subscription_check_classes_exceeded(self, csID, clattID, date):
    #     '''
    #         Gets number of weekly classes for a subscription and checks the
    #         attendance table to see if the customer is over the allowed nr of
    #         classes. If so, session.flash is set with a message to notify the user
    #     '''
    #     def insert_exceeded_classes():
    #         db.customers_subscriptions_exceeded.insert(
    #             customers_subscriptions_id = csID,
    #             classes_attendance_id      = clattID,
    #             ClassCount                 = classes_taken
    #         )
    #
    #
    #     T  = current.globalenv['T']
    #     db = current.globalenv['db']
    #     TODAY_LOCAL = current.globalenv['TODAY_LOCAL']
    #
    #     csu = db.customers_subscriptions(csID)
    #     ssu = db.school_subscriptions(csu.school_subscriptions_id)
    #
    #     # Don't do anything if the subscription grants unlimited classes
    #
    #     if ssu.Unlimited:
    #         return
    #
    #     # Ok no unlimited classes, do some checking
    #     if ssu.SubscriptionUnit == 'week':
    #         # check currently used classes
    #         from general_helpers import iso_to_gregorian
    #         iso_week = date.isocalendar()[1]
    #         monday = iso_to_gregorian(date.year, iso_week, 1)
    #         sunday = iso_to_gregorian(date.year, iso_week, 7)
    #
    #         period_start = monday
    #         period_end = sunday
    #     elif ssu.SubscriptionUnit == 'month':
    #         period_start = datetime.date(TODAY_LOCAL.year, TODAY_LOCAL.month, 1)
    #         period_end = get_last_day_month(period_start)
    #
    #     query = (db.classes_attendance.ClassDate >= period_start) & \
    #             (db.classes_attendance.ClassDate <= period_end) & \
    #             (db.classes_attendance.customers_subscriptions_id == csID)
    #
    #     classes_taken = db(query).count()
    #
    #     # check if we should set a message for the user ( None is unlimited )
    #     message = ''
    #     if not ssu.Classes:
    #         message = T("No classes allowed on this subscription")
    #
    #         insert_exceeded_classes()
    #
    #     if ( ssu.Classes and
    #          ssu.SubscriptionUnit == 'week' and
    #          classes_taken > ssu.Classes and
    #          not ssu.Unlimited ):
    #         message = T("Subscription weekly classes exceeded")
    #
    #         insert_exceeded_classes()
    #
    #     # check if we should set a message for the user ( None is unlimited )
    #     if ( ssu.Classes and
    #          ssu.SubscriptionUnit == 'month' and
    #          classes_taken > ssu.Classes and
    #          not ssu.Unlimited ):
    #         message = T("Subscription monthly classes exceeded")
    #
    #         insert_exceeded_classes()
    #
    #     return message


    def _attedance_sign_in_subscription_check_paused(self, csID, date):
        '''
            Check if the subscription if paused on given date, if so, display
            a message for the user
        '''
        T = current.globalenv['T']

        message = ''

        csh = CustomerSubscriptionsHelper(csID)
        paused = csh.get_paused(date)
        if paused:
            message = T("Subscription is paused on this date")

        return message


    def _attendance_sign_in_create_invoice(self,
                                           cuID,
                                           caID,
                                           clsID,
                                           date,
                                           product_type):
        '''
            Creates an invoice for a drop in or trial class
        '''
        db = current.globalenv['db']
        DATE_FORMAT = current.globalenv['DATE_FORMAT']
        T = current.globalenv['T']

        date_formatted = date.strftime(DATE_FORMAT)

        if product_type not in ['trial', 'dropin']:
            raise ValueError('Product type has to be trial or dropin')

        cls = Class(clsID, date)
        prices = cls.get_price()

        if product_type == 'dropin':
            price = prices['dropin']
            tax_percentage = prices['dropin_tax_percentage']
            tax_rates_id = prices['dropin_tax_rates_id']
            description = cls.get_invoice_order_description(2) # 2 = drop in class

        elif product_type == 'trial':
            price = prices['trial']
            tax_percentage = prices['trial_tax_percentage']
            tax_rates_id = prices['dropin_tax_rates_id']
            description = cls.get_invoice_order_description(1) # 1 = trial class

        # check if the price is > 0 when adding an invoice
        if price == 0:
            return

        igpt = db.invoices_groups_product_types(ProductType=product_type)

        iID = db.invoices.insert(
            invoices_groups_id=igpt.invoices_groups_id,
            # classes_attendance_id      = caID,
            Description=T('Class on ') + date_formatted,
            Status='sent'
        )

        # link invoice to attendance
        db.invoices_classes_attendance.insert(
            invoices_id=iID,
            classes_attendance_id=caID
        )

        # create object to set Invoice# and due date
        invoice = Invoice(iID)
        next_sort_nr = invoice.get_item_next_sort_nr()

        iiID = db.invoices_items.insert(
            invoices_id=iID,
            ProductName=product_type,
            Description=description,
            Quantity=1,
            Price=price,
            Sorting=next_sort_nr,
            tax_rates_id=tax_rates_id,
        )

        invoice.set_amounts()
        invoice.link_to_customer(cuID)


    def attendance_sign_in_classcard_recurring(self, cuID, clsID, ccdID, date, date_until, online_booking=False, booking_status='booked'):
        '''
        :param cuID:
        :param clsID:
        :param ccdID:
        :param date:
        :param until_date:
        :param online_booking:
        :param booking_status:
        :return:
        '''
        T = current.globalenv['T']
        TODAY_LOCAL = current.globalenv['TODAY_LOCAL']
        DATE_FORMAT = current.globalenv['DATE_FORMAT']
        get_sys_property = current.globalenv['get_sys_property']

        ccd = Classcard(ccdID)
        ccd_enddate = ccd.classcard.Enddate


        classes_booked = 0
        messages = []

        classes_remaining = ccd.get_classes_remaining()

        shop_classes_advance_booking_limit = get_sys_property('shop_classes_advance_booking_limit')
        # print shop_classes_advance_booking_limit
        book_date_limit = False
        if shop_classes_advance_booking_limit:
            book_date_limit = TODAY_LOCAL + datetime.timedelta(int(shop_classes_advance_booking_limit))


        while date <= date_until:
            # print date
            date_formatted = date.strftime(DATE_FORMAT)
            sign_in_ok = True
            # Check if class is taking place
            cls = Class(clsID, date)
            if cls.is_taking_place() == False:
                sign_in_ok = False
                # print 'class not happening'
                messages.append(T("Class is cancelled or falls within a holiday"))

            # Check online booking spaces available
            if cls.get_full_bookings_shop() == True: # no spaces available
                sign_in_ok = False
                # print "Class full"
                messages.append(T("There are no more spaces for online bookings available for this class"))

            # Check classes remaining
            #if classes_remaining != 'unlimited' or classes_remaining < 1:
            if classes_remaining == 0: # This will pass when it's 'unlimited'
                # print 'no classes remaining'
                date = date_until  # Stop loop
                #TODO: message for customer
                sign_in_ok = False
                messages.append(T('No more classes remaining on this card for class on') + ' ' + date_formatted)

            # Check not past max sign in date
            # print 'booking date limit'
            # print book_date_limit
            if book_date_limit:
                if date > book_date_limit:
                    # print 'class past booking in advance limit'
                    # TODO: message for customer
                    date = date_until # Stop loop
                    sign_in_ok = False

            # Check not past classcard enddate
            if date >= ccd_enddate:
                # print 'class past classcard enddate'
                date = date_until  # Stop loop
                #TODO: message for customer
                sign_in_ok = False
                messages.append(T('Date is past card expiration date:') + ' ' + date_formatted)

            # Check sign in status for fail (if fail, don't add count)
            if sign_in_ok:
                result = self.attendance_sign_in_classcard(cuID, clsID, ccdID, date)
                if not result['status'] == 'fail':
                    messages.append(T("Booked class on") + ' ' + date_formatted)
                    # update remaining classes
                    if not classes_remaining == 'unlimited':
                        classes_remaining -= 1

                    classes_booked += 1

            date += datetime.timedelta(days=7)

        return dict(classes_booked=classes_booked,
                    messages=messages)


    def attendance_sign_in_classcard(self, cuID, clsID, ccdID, date, online_booking=False, booking_status='booked'):
        """
            :param cuID: db.auth_user.id 
            :param clsID: db.classes.id
            :param ccdID: db.customers_classcards.id
            :param date: datetime.date
            :return: 
        """
        db = current.globalenv['db']
        T = current.globalenv['T']

        ccdh = ClasscardsHelper()
        classes_available = ccdh.get_classes_available(ccdID)

        status = 'fail'
        message = ''
        if classes_available:
            signed_in = self.attendance_sign_in_check_signed_in(clsID, cuID, date)
            if signed_in:
                message = T("Already checked in for this class")
            else:
                status = 'success'

                db.classes_attendance.insert(
                    auth_customer_id=cuID,
                    classes_id=clsID,
                    ClassDate=date,
                    AttendanceType=3,  # 3 = classcard
                    customers_classcards_id=ccdID,
                    online_booking=online_booking,
                    BookingStatus=booking_status
                )

                # update class count
                ccdh = ClasscardsHelper()
                ccdh.set_classes_taken(ccdID)
        else:
            message = T("Unable to add, no classes left on card")


        return dict(status=status, message=message)


    def attendance_sign_in_dropin(self,
                                  cuID,
                                  clsID,
                                  date,
                                  online_booking=False,
                                  invoice=True,
                                  booking_status='booked'):
        '''
            :param cuID: db.auth_user.id
            :param clsID: db.classes.id
            :param date: datetime.date
            :return: 
        '''
        db = current.globalenv['db']
        T = current.globalenv['T']

        status = 'fail'
        message = ''
        caID = ''

        signed_in = self.attendance_sign_in_check_signed_in(clsID, cuID, date)
        if signed_in:
            message = T("Already checked in for this class")
        else:
            status = 'ok'
            caID = db.classes_attendance.insert(
                auth_customer_id=cuID,
                classes_id=clsID,
                ClassDate=date,
                AttendanceType=2,  # 2 = drop in class
                online_booking=online_booking,
                BookingStatus=booking_status
            )

            if invoice:
                self._attendance_sign_in_create_invoice(cuID,
                                                        caID,
                                                        clsID,
                                                        date,
                                                        'dropin')

        return dict(status=status, message=message, caID=caID)


    def attendance_sign_in_trialclass(self,
                                      cuID,
                                      clsID,
                                      date,
                                      online_booking=False,
                                      invoice=True,
                                      booking_status='booked'):
        '''
            :param cuID: db.auth_user.id
            :param clsID: db.classes.id
            :param date: datetime.date
            :return: 
        '''
        db = current.globalenv['db']
        T = current.globalenv['T']

        status = 'fail'
        message = ''
        caID = ''

        signed_in = self.attendance_sign_in_check_signed_in(clsID, cuID, date)

        if signed_in:
            message = T("Already checked in for this class")
        else:
            status = 'ok'
            caID = db.classes_attendance.insert(
                auth_customer_id=cuID,
                classes_id=clsID,
                ClassDate=date,
                AttendanceType=1,  # 1 = trial class
                online_booking=online_booking,
                BookingStatus=booking_status
            )

            if invoice:
                self._attendance_sign_in_create_invoice(cuID,
                                                        caID,
                                                        clsID,
                                                        date,
                                                        'trial')

        return dict(status=status, message=message, caID=caID)


    def attendance_sign_in_complementary(self,
                                         cuID,
                                         clsID,
                                         date,
                                         booking_status='booked'):
        '''
            :param cuID: db.auth_user.id
            :param clsID: db.classes.id
            :param date: datetime.date
            :return:
        '''
        db = current.globalenv['db']
        T = current.globalenv['T']

        status = 'fail'
        message = ''
        caID = ''

        signed_in = self.attendance_sign_in_check_signed_in(clsID, cuID, date)

        if signed_in:
            message = T("Already checked in for this class")
        else:
            status = 'ok'
            caID = db.classes_attendance.insert(
                auth_customer_id=cuID,
                classes_id=clsID,
                ClassDate=date,
                AttendanceType=4,  # 4 = Complementary class
                online_booking=False,
                BookingStatus=booking_status
            )


        return dict(status=status, message=message, caID=caID)


    def attendance_sign_in_check_signed_in(self, clsID, cuID, date):
        '''
            Check if a customer isn't already signed in to a class
        '''
        db = current.globalenv['db']
        query = (db.classes_attendance.classes_id == clsID) & \
                (db.classes_attendance.auth_customer_id == cuID) & \
                (db.classes_attendance.ClassDate == date) & \
                (db.classes_attendance.BookingStatus != 'cancelled')

        return db(query).count()


    def attendance_cancel_classes_in_school_holiday(self, shID):
        '''
            :param shID: db.school_holidays.id
            :return: list of db.classes.id
        '''
        db = current.globalenv['db']

        # Get locations
        query = (db.school_holidays_locations.school_holidays_id == shID)
        rows =  db(query).select(db.school_holidays_locations.school_locations_id)
        location_ids = []
        for row in rows:
            location_ids.append(row.school_locations_id)

        # Get classes
        query = (db.classes.school_locations_id.belongs(location_ids))
        rows = db(query).select(db.classes.id)
        class_ids = []
        for row in rows:
            class_ids.append(row.id)

        # Get holiday record and cancel classes
        sh = db.school_holidays(shID)
        self.attendance_cancel_reservations_for_classes(class_ids, sh.Startdate, sh.Enddate)


    def attendance_cancel_reservations_for_classes(self, class_ids, p_start, p_end = None):
        '''
            :param class_ids: list of db.classes.id
            :param p_start: datetime.date
            :param p_end: datetime.date
            :return: None
        '''
        db = current.globalenv['db']

        # in case end period is not specified, assume it's for one day
        if not p_end:
            p_end = p_start

        query = (db.classes_attendance.classes_id.belongs(class_ids)) & \
                (db.classes_attendance.ClassDate >= p_start) & \
                (db.classes_attendance.ClassDate <= p_end)

        db(query).update(BookingStatus='cancelled')

        # Return subscription credits to customers
        csch = CustomersSubscriptionsCreditsHelper()
        csch.refund_credits_in_period(query)


class ReservationHelper:
    '''
        This class collects common functions for reservations in OpenStudio
    '''
    def get_reservation(self, cuID, clsID, date):
        '''
           returns reservation for a customer, if any
        '''
        db = current.globalenv['db']
        os_gui = current.globalenv['os_gui']
        DATE_FORMAT = current.globalenv['DATE_FORMAT']
        date_formatted = date.strftime(DATE_FORMAT)

        query = (db.classes_reservation.auth_customer_id == cuID) & \
                (db.classes_reservation.classes_id == clsID) & \
                (db.classes_reservation.Startdate <= date) & \
                ((db.classes_reservation.Enddate >= date) |
                 (db.classes_reservation.Enddate == None))

        rows = db(query).select(db.classes_reservation.ALL)

        if rows:
            return_value = rows
        else:
            return_value = None

        return return_value


class ClassSchedule:
    def __init__(self, date,
                       filter_id_sys_organization = None,
                       filter_id_school_classtype = None,
                       filter_id_school_location = None,
                       filter_id_school_level = None,
                       filter_id_teacher = None,
                       filter_id_status = None,
                       filter_public = False,
                       sorting = 'starttime',
                       trend_medium = None,
                       trend_high = None):

        self.date = date

        self.filter_id_sys_organization = filter_id_sys_organization
        self.filter_id_school_classtype = filter_id_school_classtype
        self.filter_id_school_location = filter_id_school_location
        self.filter_id_teacher = filter_id_teacher
        self.filter_id_school_level = filter_id_school_level
        self.filter_id_status = filter_id_status
        self.filter_public = filter_public
        self.sorting = sorting
        self.trend_medium = trend_medium
        self.trend_high = trend_high

        self.bookings_open = self._get_bookings_open()


    def _get_bookings_open(self):
        '''
            Returns False if no booking limit is defines, otherwise it returns the date from which
            bookings for this class will be accepted.
        '''
        get_sys_property = current.globalenv['get_sys_property']

        bookings_open = False
        shop_classes_advance_booking_limit = get_sys_property('shop_classes_advance_booking_limit')
        if not shop_classes_advance_booking_limit is None:
            delta = datetime.timedelta(days=int(shop_classes_advance_booking_limit))
            bookings_open = self.date - delta

        return bookings_open


    def _get_day_filter_query(self):
        '''
            Returns the filter query for the schedule
        '''
        where = ''

        if self.filter_id_sys_organization:
            where += 'AND cla.sys_organizations_id = ' + unicode(self.filter_id_sys_organization) + ' '
        if self.filter_id_teacher:
            where += 'AND ((CASE WHEN cotc.auth_teacher_id IS NULL \
                            THEN clt.auth_teacher_id  \
                            ELSE cotc.auth_teacher_id END) = '
            where += unicode(self.filter_id_teacher) + ' '
            where += 'OR (CASE WHEN cotc.auth_teacher_id2 IS NULL \
                          THEN clt.auth_teacher_id2  \
                          ELSE cotc.auth_teacher_id2 END) = '
            where += unicode(self.filter_id_teacher) + ') '
        if self.filter_id_school_classtype:
            where += 'AND (CASE WHEN cotc.school_classtypes_id IS NULL \
                           THEN cla.school_classtypes_id  \
                           ELSE cotc.school_classtypes_id END) = '
            where += unicode(self.filter_id_school_classtype) + ' '
        if self.filter_id_school_location:
            where += 'AND (CASE WHEN cotc.school_locations_id IS NULL \
                           THEN cla.school_locations_id  \
                           ELSE cotc.school_locations_id END) = '
            where += unicode(self.filter_id_school_location) + ' '
        if self.filter_id_school_level:
            where += 'AND cla.school_levels_id = '
            where += unicode(self.filter_id_school_level) + ' '
        if self.filter_public:
            where += "AND cla.AllowAPI = 'T' "
            where += "AND sl.AllowAPI = 'T' "
            where += "AND sct.AllowAPI = 'T' "

        return where


    def _get_day_row_status(self, row):
        '''
            Return status for row
        '''
        status = 'normal'
        status_marker = DIV(_class='status_marker bg_green')
        if row.classes_otc.Status == 'cancelled' or row.school_holidays.id:
            status = 'cancelled'
            status_marker = DIV(_class='status_marker bg_orange')
        elif row.classes_otc.Status == 'open':
            status = 'open'
            status_marker = DIV(_class='status_marker bg_red')
        elif row.classes_teachers.teacher_role == 1:
            status = 'subteacher'
            status_marker = DIV(_class='status_marker bg_blue')

        return dict(status=status, marker=status_marker)


    def _get_day_row_teacher_roles(self, row, repr_row):
        '''
            @return: dict with {teacher_role} and {teacher_role2} as keys
             teacher_role and teacher_role2 are names of teacher with labels
              applied
        '''
        os_gui = current.globalenv['os_gui']
        T = current.globalenv['T']

        teacher_id = row.classes_teachers.auth_teacher_id
        teacher_id2 = row.classes_teachers.auth_teacher_id2
        teacher = repr_row.classes_teachers.auth_teacher_id
        teacher2 = repr_row.classes_teachers.auth_teacher_id2
        teacher_role = row.classes_teachers.teacher_role
        teacher_role2 = row.classes_teachers.teacher_role2

        # set label for teacher role
        if teacher_role == 1:  # sub
            teacher_role = SPAN(os_gui.get_os_label('blue', teacher),
                                _title=T('Sub teacher'))
        elif teacher_role == 2:  # assist
            teacher_role = SPAN(os_gui.get_os_label('yellow', teacher),
                                _title=T("Assistant"))
        elif teacher_role == 3:  # karma
            teacher_role = SPAN(os_gui.get_os_label('purple', teacher),
                                _title=T('Karma teacher'))
        else:
            teacher_role = teacher

        # set label for teacher role 2
        if teacher_role2 == 1:  # sub
            teacher_role2 = SPAN(os_gui.get_os_label('blue', teacher2),
                                 _title=T("Sub teacher"))
        elif teacher_role2 == 2:  # assist
            teacher_role2 = SPAN(os_gui.get_os_label('yellow', teacher2),
                                 _title=T("Assistant"))
        elif teacher_role2 == 3:  # karma
            teacher_role2 = SPAN(os_gui.get_os_label('purple', teacher2),
                                 _title=T('Karma teacher'))
        else:
            teacher_role2 = teacher2

        return dict(teacher_role=teacher_role,
                    teacher_role2=teacher_role2)


    def _get_day_get_table_class_trend_data(self):
        '''
            dict containing trend divs for self.date
        '''
        def average(total, classes_counted):
            try:
                average = float(total / classes_counted)
            except ZeroDivisionError:
                average = float(0)

            return average

        DATE_FORMAT = current.globalenv['DATE_FORMAT']
        db = current.globalenv['db']
        T = current.globalenv['T']
        weekday = self.date.isoweekday()

        date_formatted = self.date.strftime(DATE_FORMAT)

        delta = datetime.timedelta(days=28)
        one_month_ago = self.date - delta
        two_months_ago = one_month_ago - delta

        fields = [
            db.classes.id,
            db.classes.Maxstudents,
            db.classes_schedule_count.Attendance4WeeksAgo,
            db.classes_schedule_count.NRClasses4WeeksAgo,
            db.classes_schedule_count.Attendance8WeeksAgo,
            db.classes_schedule_count.NRClasses8WeeksAgo
        ]

        query = '''
            SELECT cla.id,
                   CASE WHEN cotc.Maxstudents IS NOT NULL
                        THEN cotc.Maxstudents
                        ELSE cla.Maxstudents
                        END AS Maxstudents, 
                   clatt_4w_ago.att_4w,
                   clatt_4w_nrclasses.att_4w_nrclasses,
                   clatt_8w_ago.att_8w,
                   clatt_8w_nrclasses.att_8w_nrclasses
            FROM classes cla
            LEFT JOIN
                ( SELECT id,
                         classes_id,
                         ClassDate,
                         Status,
                         Description,
                         school_locations_id,
                         school_classtypes_id,
                         Starttime,
                         Endtime,
                         auth_teacher_id,
                         teacher_role,
                         auth_teacher_id2,
                         teacher_role2,
                         Maxstudents,
                         MaxOnlinebooking
                  FROM classes_otc
                  WHERE ClassDate = '{class_date}' ) cotc
            ON cla.id = cotc.classes_id            
            LEFT JOIN
                    ( SELECT classes_id, COUNT(*) as att_4w
                      FROM classes_attendance
                      WHERE classes_attendance.Classdate <  '{class_date}' AND
                            classes_attendance.Classdate >= '{one_month_ago}'
                      GROUP BY classes_id
                    ) clatt_4w_ago
                    ON clatt_4w_ago.classes_id = cla.id
                LEFT JOIN
                    ( SELECT classes_id, COUNT(DISTINCT ClassDate) as att_4w_nrclasses
                      FROM classes_attendance
                      WHERE classes_attendance.Classdate <  '{class_date}' AND
                            classes_attendance.Classdate >= '{one_month_ago}'
                      GROUP BY classes_id
                    ) clatt_4w_nrclasses
                    ON clatt_4w_nrclasses.classes_id = cla.id
                LEFT JOIN
                    ( SELECT classes_id, COUNT(*) as att_8w
                      FROM classes_attendance
                      WHERE classes_attendance.Classdate <  '{one_month_ago}' AND
                            classes_attendance.Classdate >= '{two_months_ago}'
                      GROUP BY classes_id
                    ) clatt_8w_ago
                    ON clatt_8w_ago.classes_id = cla.id
                LEFT JOIN
                    ( SELECT classes_id, COUNT(DISTINCT ClassDate) as att_8w_nrclasses
                      FROM classes_attendance
                      WHERE classes_attendance.Classdate <  '{one_month_ago}' AND
                            classes_attendance.Classdate >= '{two_months_ago}'
                      GROUP BY classes_id
                    ) clatt_8w_nrclasses
                    ON clatt_8w_nrclasses.classes_id = cla.id
            WHERE cla.Week_day = '{week_day}' AND
                  cla.Startdate <= '{class_date}' AND
                  (cla.Enddate >= '{class_date}' OR cla.Enddate IS NULL)
            '''.format(class_date=self.date,
                       week_day=weekday,
                       one_month_ago=one_month_ago,
                       two_months_ago=two_months_ago)

        rows = db.executesql(query, fields=fields)

        data = {}

        trend_medium = self.trend_medium
        trend_high = self.trend_high

        for row in rows:
            classes_4w = row.classes_schedule_count.NRClasses4WeeksAgo or 0
            attendance_4w = row.classes_schedule_count.Attendance4WeeksAgo or 0
            avg_4w_ago = average(attendance_4w, classes_4w)
            classes_8w = row.classes_schedule_count.NRClasses8WeeksAgo or 0
            attendance_8w = row.classes_schedule_count.Attendance8WeeksAgo or 0
            avg_8w_ago = average(attendance_8w, classes_8w)

            div = DIV()

            display_class = ''
            capacity = ''

            try:
                avg_att_4w_percentage = (avg_4w_ago / row.classes.Maxstudents) * 100
            except ZeroDivisionError:
                avg_att_4w_percentage = 0
            avg_att_4w_percentage_display = round(avg_att_4w_percentage, 2)

            class_trend_text_color = 'grey'
            if trend_medium:
                capacity = ' - ' + T('Capacity filled: ') + unicode(avg_att_4w_percentage_display) + '%'
                if avg_att_4w_percentage < trend_medium:
                    class_trend_text_color = 'text-red'
                else:
                    class_trend_text_color = 'text-yellow'
            if trend_high:
                capacity = ' - ' + T('Capacity filled: ') + unicode(avg_att_4w_percentage_display) + '%'
                if avg_att_4w_percentage >= trend_high:
                    class_trend_text_color = 'text-green'

            avg_4w_ago_display = DIV(SPAN(int(avg_4w_ago), '/', row.classes.Maxstudents),
                                     _title=T("Average attendance past 4 weeks") + ' ' + capacity,
                                     _class='os-trend_avg_4_weeks inline-block ' + class_trend_text_color)
            try:
                if avg_4w_ago >= avg_8w_ago:
                    # calculate percentual increase
                    increase = avg_4w_ago - avg_8w_ago
                    value = int(round(float(increase / avg_8w_ago) * 100))
                    value = unicode(value) + '%'
                    div = DIV(avg_4w_ago_display, ' ',
                              DIV(_class='os-trend_arrow_up'),
                              SPAN(value, _title=T('Increase during past 4 weeks, compared to 8 weeks ago')),
                              ' ',
                              SPAN(_class='icon user icon-user'))
                else:
                    # calculate percentual decrease
                    decrease = avg_8w_ago - avg_4w_ago
                    value = int(round(float(decrease / avg_8w_ago) * 100))
                    value = unicode(value) + '%'
                    div = DIV(avg_4w_ago_display, ' ',
                              DIV(_class='os-trend_arrow_down'),
                              SPAN(value, _title=T('Decrease during past 4 weeks, compared to 8 weeks ago')),
                              ' ',
                              SPAN(_class='icon user icon-user'))

            except ZeroDivisionError:
                div = ''

            data[row.classes.id] = div

        return data


    def _get_day_get_table_class_trend(self):
        '''
            Generates a div that contains the trend for a class.
            Look at past 4 weeks and compare to the classes before it.
            Take cancelled classes & into account by not counting a class
            when it's date doesn't appear in classes_attendance
        '''
        web2pytest = current.globalenv['web2pytest']
        request = current.globalenv['request']
        auth = current.globalenv['auth']
        T = current.globalenv['T']

        # get attendance data from cache or db

        # Don't cache when running tests
        if web2pytest.is_running_under_test(request, request.application):
            data = self._get_day_get_table_class_trend_data()
        else:
            twelve_hours = 12*60*60
            cache = current.globalenv['cache']
            DATE_FORMAT = current.globalenv['DATE_FORMAT']
            # A key that isn't cleared when schedule changes occur.
            cache_key = 'openstudio_classschedule_trend_get_day_table_' + \
                        self.date.strftime(DATE_FORMAT)

            data = cache.ram(cache_key , lambda: self._get_day_get_table_class_trend_data(), time_expire=twelve_hours)

        return data


    def _get_day_get_table_get_permissions(self):
        """
            :return: dict containing button permissions for a user
        """
        auth = current.globalenv['auth']
        permissions = {}

        if auth.has_membership(group_id='Admins') or \
           auth.has_permission('read', 'classes_attendance'):
            permissions['classes_attendance'] = True
        if auth.has_membership(group_id='Admins') or \
           auth.has_permission('read', 'classes_reservation'):
            permissions['classes_reservation'] = True
        if auth.has_membership(group_id='Admins') or \
           auth.has_permission('read', 'classes_waitinglist'):
            permissions['classes_waitinglist'] = True
        if auth.has_membership(group_id='Admins') or \
           auth.has_permission('read', 'classes_notes'):
            permissions['classes_notes'] = True
        if auth.has_membership(group_id='Admins') or \
           auth.has_permission('create', 'classes_otc'):
            permissions['classes_otc'] = True
        if auth.has_membership(group_id='Admins') or \
           auth.has_permission('update', 'classes'):
            permissions['classes'] = True
        if auth.has_membership(group_id='Admins') or \
           auth.has_permission('delete', 'classes'):
            permissions['classes_delete'] = True

        return permissions


    def _get_day_get_table_get_buttons(self, clsID, date_formatted, permissions):
        '''
            Returns buttons for schedule
            - one button group for edit & attendance buttons
            - separate button for delete
        '''
        os_gui = current.globalenv['os_gui']
        T = current.globalenv['T']
        buttons = DIV(_class='pull-right')

        vars = { 'clsID':clsID,
                 'date' :date_formatted }


        links = [['header', T('Class on') + ' ' + date_formatted]]
        # check Attendance permission
        if permissions.get('classes_attendance', False):
            links.append(A(os_gui.get_fa_icon('fa-check-square-o'), T('Attendance'),
                           _href=URL('attendance', vars=vars)))
        # check Reservations permission
        if permissions.get('classes_reservation', False):
            links.append(
                A(os_gui.get_fa_icon('fa-calendar-check-o'),  T('Enrollments'),
                 _href=URL('reservations', vars=vars)))
        # check Waitinglist permission
        if permissions.get('classes_waitinglist', False):
            links.append(
                A(os_gui.get_fa_icon('fa-calendar-o'), T('Waitinglist'),
                  _href=URL('waitinglist', vars=vars)))
        # check Notes permission
        if permissions.get('classes_notes', False):
            links.append(
                A(os_gui.get_fa_icon('fa-sticky-note-o'), T('Notes'),
                  _href=URL('notes', vars=vars)))
        # check permissions to change this class
        if permissions.get('classes_otc', False):
            links.append(A(os_gui.get_fa_icon('fa-pencil'),
                           T('Edit'),
                           _href=URL('class_edit_on_date', vars=vars)))
        # Check permission to update weekly class
        if permissions.get('classes', False):
            links.append('divider')
            links.append(['header', T('All classes in series')])
            links.append(A(os_gui.get_fa_icon('fa-pencil'),
                           T('Edit'), ' ',
                           _href=URL('class_edit', vars=vars)))

        class_menu = os_gui.get_dropdown_menu(
            links=links,
            btn_text=T('Actions'),
            btn_size='btn-sm',
            btn_icon='actions',
            menu_class='btn-group pull-right')

        remove = ''
        if permissions.get('classes_delete'):
            onclick_remove = "return confirm('" + \
                             T('Do you really want to delete this class?') + \
                             "');"
            remove = os_gui.get_button('delete_notext',
                       URL('class_delete', args=[clsID]),
                       onclick=onclick_remove)

            buttons.append(remove)

        return DIV(buttons, class_menu, _class='pull-right schedule_buttons')


    def _get_day_get_table_get_reservations(self, clsID, date_formatted, row, permissions):
        '''
            Returns tools for schedule
            - reservations
        '''
        auth = current.globalenv['auth']
        T = current.globalenv['T']

        tools = DIV()

        # get bookings count
        res = row.classes_schedule_count.Attendance or 0

        filled = SPAN(res, '/', row.classes.Maxstudents)

        link_class = ''
        if res > row.classes.Maxstudents:
            link_class = 'red'

        reservations = A(SPAN(T('Bookings'), ' ', filled),
                         _href=URL('attendance',
                                   vars={'clsID' : clsID,
                                         'date'  : date_formatted}),
                         _class=link_class)

        if permissions.get('classes_attendance', False):
            tools.append(reservations)

        return tools


    def _get_day_table_get_class_messages(self, row, clsID, date_formatted):
        '''
            Returns messages for a class
        '''
        os_gui = current.globalenv['os_gui']
        auth = current.globalenv['auth']
        T = current.globalenv['T']

        class_messages = []

        if row.school_holidays.Description:
            class_messages.append(
                SPAN(SPAN(_class=os_gui.get_icon('education') + ' grey'), ' ',
                     T('Holiday'), ' (',
                     A(row.school_holidays.Description,
                       _href=URL('schedule', 'holiday_edit',
                                 vars={'shID': row.school_holidays.id})),
                     ')'))

        if row.classes_teachers.teacher_role == 1:
            class_messages.append(T('Subteacher'))

        if row.classes_otc.Status == 'cancelled':
            class_messages.append(T('Cancelled'))

        if row.classes_otc.Status == 'open':
            class_messages.append(T('Open'))

        classes_otc_update_permission = auth.has_membership(group_id='Admins') or \
                                        auth.has_permission('update', 'classes_otc')
        if row.classes_otc.id and classes_otc_update_permission:
            _class = os_gui.get_icon('pencil') + ' grey'
            class_messages.append(
                A(SPAN(SPAN(_class=_class), ' ', T('Edited')),
                  _href=URL('class_edit_on_date',
                            vars={'clsID': clsID,
                                  'date': date_formatted})))

            if row.classes_otc.Description:
                class_messages.append(row.classes_otc.Description)

        num_messages = len(class_messages)
        msgs = SPAN()
        append = msgs.append
        for i, msg in enumerate(class_messages):
            append(msg)
            if i + 1 < num_messages:
                append(' | ')

        return msgs


    def _get_day_list_booking_status(self, row):
        """
            :param row: ClassSchedule.get_day_rows() row
            :return: booking status
        """
        pytz = current.globalenv['pytz']
        TIMEZONE = current.globalenv['TIMEZONE']
        NOW_LOCAL = current.globalenv['NOW_LOCAL']
        TODAY_LOCAL = current.globalenv['TODAY_LOCAL']

        local_tz = pytz.timezone(TIMEZONE)

        dt_start = datetime.datetime(self.date.year,
                                     self.date.month,
                                     self.date.day,
                                     int(row.classes.Starttime.hour),
                                     int(row.classes.Starttime.minute))
        dt_start = local_tz.localize(dt_start)
        dt_end = datetime.datetime(self.date.year,
                                   self.date.month,
                                   self.date.day,
                                   int(row.classes.Endtime.hour),
                                   int(row.classes.Endtime.minute))
        dt_end = local_tz.localize(dt_end)

        status = 'finished'
        if row.classes_otc.Status == 'cancelled' or row.school_holidays.id:
            status = 'cancelled'
        elif dt_start <= NOW_LOCAL and dt_end >= NOW_LOCAL:
            # check start time
            status = 'ongoing'
        elif dt_start >= NOW_LOCAL:
            if not self.bookings_open == False and TODAY_LOCAL < self.bookings_open:
                status = 'not_yet_open'
            else:
                # check spaces for online bookings
                spaces = self._get_day_list_booking_spaces(row)
                if spaces < 1:
                    status = 'full'
                else:
                    status = 'ok'

        return status


    def _get_day_list_booking_spaces(self, row):
        """
        :param row: :param row: ClassSchedule.get_day_rows() row
        :return: int - available online booking spaces for a class
        """
        enrollments = row.classes_schedule_count.Reservations or 0
        enrollment_spaces = row.classes.MaxReservationsRecurring or 0
        enrollment_spaces_left = enrollment_spaces - enrollments

        spaces = row.classes.MaxOnlineBooking or 0
        online_booking = row.classes_schedule_count.OnlineBooking or 0
        #attendance = row.classes_schedule_count.Attendance or 0

        available_spaces = (spaces + enrollment_spaces_left) - online_booking
        if available_spaces < 1:
            available_spaces = 0
        #
        # print '### clsID' + unicode(row.classes.id)
        # print spaces
        # print enrollment_spaces_left
        # print online_booking
        # print available_spaces

        return available_spaces


    def _get_day_rows(self):
        """
            Helper function that returns a dict containing a title for the weekday,
            a date for the class and
            a SQLFORM.grid for a selected day which is within 1 - 7 (ISO standard).
        """
        date = self.date
        DATE_FORMAT = current.globalenv['DATE_FORMAT']
        db = current.globalenv['db']
        weekday = date.isoweekday()

        date_formatted = date.strftime(DATE_FORMAT)

        delta = datetime.timedelta(days=28)
        one_month_ago = date - delta
        two_months_ago = one_month_ago - delta

        if self.sorting == 'location':
            orderby_sql = 'location_name, Starttime'
        elif self.sorting == 'starttime':
            orderby_sql = 'Starttime, location_name'

        fields = [
            db.classes.id,
            db.classes_otc.Status,
            db.classes_otc.Description,
            db.classes.school_locations_id,
            db.school_locations.Name,
            db.classes.school_classtypes_id,
            db.classes.school_levels_id,
            db.classes.Week_day,
            db.classes.Starttime,
            db.classes.Endtime,
            db.classes.Startdate,
            db.classes.Enddate,
            db.classes.Maxstudents,
            db.classes.MaxOnlineBooking,
            db.classes.MaxReservationsRecurring,
            db.classes.AllowAPI,
            db.classes.sys_organizations_id,
            db.classes_otc.id,
            db.classes_teachers.id,
            db.classes_teachers.auth_teacher_id,
            db.classes_teachers.teacher_role,
            db.classes_teachers.auth_teacher_id2,
            db.classes_teachers.teacher_role2,
            db.school_holidays.id,
            db.school_holidays.Description,
            db.classes_schedule_count.Attendance,
            db.classes_schedule_count.OnlineBooking,
            db.classes_schedule_count.Reservations
        ]

        where_filter = self._get_day_filter_query()

        query = '''
        SELECT cla.id,
               CASE WHEN cotc.Status IS NOT NULL
                    THEN cotc.Status
                    ELSE 'normal'
                    END AS Status,
               cotc.Description,
               CASE WHEN cotc.school_locations_id IS NOT NULL
                    THEN cotc.school_locations_id
                    ELSE cla.school_locations_id
                    END AS school_locations_id,
               CASE WHEN cotc.school_locations_id IS NOT NULL
                    THEN slcotc.Name
                    ELSE sl.Name
                    END AS location_name,
               CASE WHEN cotc.school_classtypes_id IS NOT NULL
                    THEN cotc.school_classtypes_id
                    ELSE cla.school_classtypes_id
                    END AS school_classtypes_id,
               cla.school_levels_id,
               cla.Week_day,
               CASE WHEN cotc.Starttime IS NOT NULL
                    THEN cotc.Starttime
                    ELSE cla.Starttime
                    END AS Starttime,
               CASE WHEN cotc.Endtime IS NOT NULL
                    THEN cotc.Endtime
                    ELSE cla.Endtime
                    END AS Endtime,
               cla.Startdate,
               cla.Enddate,
               CASE WHEN cotc.Maxstudents IS NOT NULL
                    THEN cotc.Maxstudents
                    ELSE cla.Maxstudents
                    END AS Maxstudents, 
               CASE WHEN cotc.MaxOnlineBooking IS NOT NULL
                    THEN cotc.MaxOnlineBooking
                    ELSE cla.MaxOnlineBooking
                    END AS MaxOnlineBooking,
               cla.MaxReservationsRecurring,             
               cla.AllowAPI,
               cla.sys_organizations_id,
               cotc.id,
               clt.id,
               CASE WHEN cotc.auth_teacher_id IS NOT NULL
                    THEN cotc.auth_teacher_id
                    ELSE clt.auth_teacher_id
                    END AS auth_teacher_id,
               CASE WHEN cotc.auth_teacher_id IS NOT NULL
                    THEN cotc.teacher_role
                    ELSE clt.teacher_role
                    END AS teacher_role,
               CASE WHEN cotc.auth_teacher_id2 IS NOT NULL
                    THEN cotc.auth_teacher_id2
                    ELSE clt.auth_teacher_id2
                    END AS auth_teacher_id2,
               CASE WHEN cotc.auth_teacher_id2 IS NOT NULL
                    THEN cotc.teacher_role2
                    ELSE clt.teacher_role2
                    END AS teacher_role2,
               sho.id,
               sho.Description,
                /* Count attendance for this class */
               ( SELECT count(clatt.id) as count_att
                 FROM classes_attendance clatt
                 WHERE clatt.classes_id = cla.id AND
                       clatt.ClassDAte ='{class_date}' AND
                       clatt.BookingStatus != 'cancelled') AS count_attendance,
               /* Count of online bookings for this class */
               ( SELECT COUNT(clatt.id) as count_atto
                 FROM classes_attendance clatt
                 WHERE clatt.classes_id = cla.id AND
                       clatt.ClassDate = '{class_date}' AND
                       clatt.BookingStatus != 'cancelled' AND
                       clatt.online_booking = 'T'
                 GROUP BY clatt.classes_id ) as count_clatto,
               /* Count of enrollments (reservations) for this class */
               ( SELECT COUNT(clr.id) as count_clr
                 FROM classes_reservation clr
                 WHERE clr.classes_id = cla.id AND
                       (clr.Startdate <= '{class_date}' AND
                        (clr.Enddate >= '{class_date}' OR clr.Enddate IS NULL))
                 GROUP BY clr.classes_id ) as count_clr
        FROM classes cla
        LEFT JOIN
            ( SELECT id,
                     classes_id,
                     ClassDate,
                     Status,
                     Description,
                     school_locations_id,
                     school_classtypes_id,
                     Starttime,
                     Endtime,
                     auth_teacher_id,
                     teacher_role,
                     auth_teacher_id2,
                     teacher_role2,
                     Maxstudents,
                     MaxOnlinebooking
              FROM classes_otc
              WHERE ClassDate = '{class_date}' ) cotc
            ON cla.id = cotc.classes_id
        LEFT JOIN school_locations sl
            ON sl.id = cla.school_locations_id
        LEFT JOIN school_classtypes sct
            ON sct.id = cla.school_classtypes_id
		LEFT JOIN school_locations slcotc
			ON slcotc.id = cotc.school_locations_id
        LEFT JOIN
            ( SELECT id,
                     classes_id,
                     auth_teacher_id,
                     teacher_role,
                     auth_teacher_id2,
                     teacher_role2
              FROM classes_teachers
              WHERE Startdate <= '{class_date}' AND (
                    Enddate >= '{class_date}' OR Enddate IS NULL)
              ) clt
            ON clt.classes_id = cla.id
        LEFT JOIN
            ( SELECT sh.id, sh.Description, shl.school_locations_id
              FROM school_holidays sh
              LEFT JOIN
                school_holidays_locations shl
                ON shl.school_holidays_id = sh.id
              WHERE sh.Startdate <= '{class_date}' AND
                    sh.Enddate >= '{class_date}') sho
            ON sho.school_locations_id = cla.school_locations_id
        WHERE cla.Week_day = '{week_day}' AND
              cla.Startdate <= '{class_date}' AND
              (cla.Enddate >= '{class_date}' OR cla.Enddate IS NULL)
              {where_filter}
        ORDER BY {orderby_sql}
        '''.format(class_date = date,
                   week_day = weekday,
                   orderby_sql = orderby_sql,
                   where_filter = where_filter,
                   one_month_ago = one_month_ago,
                   two_months_ago = two_months_ago)

        rows = db.executesql(query, fields=fields)

        return rows


    def get_day_rows(self):
        '''
            Get day rows with caching 
        '''
        #web2pytest = current.globalenv['web2pytest']
        #request = current.globalenv['request']

        # # Don't cache when running tests
        # if web2pytest.is_running_under_test(request, request.application):
        #     rows = self._get_day_rows()
        # else:
        #     cache = current.globalenv['cache']
        #     DATE_FORMAT = current.globalenv['DATE_FORMAT']
        #     CACHE_LONG = current.globalenv['CACHE_LONG']
        #     cache_key = 'openstudio_classschedule_get_day_rows_' + self.date.strftime(DATE_FORMAT)
        #     rows = cache.ram(cache_key , lambda: self._get_day_rows(), time_expire=CACHE_LONG)

        rows = self._get_day_rows()

        return rows


    def _get_day_table(self):
        """
            Returns table for today
        """
        os_gui = current.globalenv['os_gui']
        DATE_FORMAT = current.globalenv['DATE_FORMAT']
        ORGANIZATIONS = current.globalenv['ORGANIZATIONS']
        T = current.globalenv['T']
        date_formatted = self.date.strftime(DATE_FORMAT)

        table = TABLE(TR(TH(' ', _class='td_status_marker'), # status marker
                         TH(T('Location'), _class='location'),
                         TH(T('Class type'), _class='classtype'),
                         TH(T('Time'), _class='time'),
                         TH(T('Teacher'), _class='teacher'),
                         TH(T('Level'), _class='level'),
                         TH(T('Public'), _class='api'),
                         TH(T('Trend'), _class='trend'),
                         TH(T('')),
                         _class='os-table_header'),
                      _class='os-schedule')

        rows = self.get_day_rows()

        if len(rows) == 0:
            div_classes=DIV()
        else:
            # Get trend column from cache
            trend_data = self._get_day_get_table_class_trend()
            get_trend_data = trend_data.get

            # avoiding some dots in the loop
            get_status = self._get_day_row_status
            get_teacher_roles = self._get_day_row_teacher_roles
            get_buttons = self._get_day_get_table_get_buttons
            get_reservations = self._get_day_get_table_get_reservations
            get_class_messages = self._get_day_table_get_class_messages

            button_permissions = self._get_day_get_table_get_permissions()

            multiple_organizations = len(ORGANIZATIONS) > 1
            filter_id_status = self.filter_id_status
            msg_no_teacher = SPAN(T('No teacher'), _class='red')

            # Generate list of classes
            for i, row in enumerate(rows):
                repr_row = list(rows[i:i+1].render())[0]
                clsID = row.classes.id

                status_result = get_status(row)
                status = status_result['status']
                status_marker = status_result['marker']

                if filter_id_status and status != filter_id_status:
                    continue

                result = get_teacher_roles(row, repr_row)
                teacher = result['teacher_role']
                teacher2 = result['teacher_role2']

                api = INPUT(value=row.classes.AllowAPI,
                            _type='checkbox',
                            _value='api',
                            _disabled='disabled')

                trend = get_trend_data(row.classes.id, '')
                buttons = get_buttons(clsID, date_formatted, button_permissions)
                reservations = get_reservations(clsID, date_formatted, row, button_permissions)
                class_messages = get_class_messages(row, clsID, date_formatted)

                if multiple_organizations:
                    organization = DIV(repr_row.classes.sys_organizations_id or '',
                                       _class='small_font grey pull-right btn-margin')
                else:
                    organization = ''

                row_class = TR(
                    TD(status_marker),
                    TD(max_string_length(repr_row.classes.school_locations_id, 15)),
                    TD(max_string_length(repr_row.classes.school_classtypes_id, 24)),
                    TD(SPAN(repr_row.classes.Starttime, ' - ', repr_row.classes.Endtime)),
                    TD(teacher if (not status == 'open' and
                                   not row.classes_teachers.auth_teacher_id is None) \
                               else msg_no_teacher),
                    TD(max_string_length(repr_row.classes.school_levels_id, 12)),
                    TD(api),
                    TD(trend),
                    TD(buttons),
                   _class='os-schedule_class')
                row_tools = TR(
                    TD(' '),
                    TD(class_messages, _colspan=3, _class='grey'),
                    TD(teacher2 if not status == 'open' else ''),
                    TD(),
                    TD(),
                    TD(DIV(reservations,
                           _class='os-schedule_links')),
                    TD(organization),
                    _class='os-schedule_links',
                    _id='class_' + unicode(clsID))

                table.append(row_class)
                table.append(row_tools)

        return dict(table=table,
                    weekday=NRtoDay(self.date.isoweekday()),
                    date=date_formatted)


    def get_day_table(self):
        """
            Get day table with caching
        """
        web2pytest = current.globalenv['web2pytest']
        request = current.globalenv['request']
        auth = current.globalenv['auth']

        # Don't cache when running tests
        if web2pytest.is_running_under_test(request, request.application):
            rows = self._get_day_table()
        else:
            cache = current.globalenv['cache']
            DATE_FORMAT = current.globalenv['DATE_FORMAT']
            CACHE_LONG = current.globalenv['CACHE_LONG']
            cache_key = 'openstudio_classschedule_get_day_table_' + \
                        self.date.strftime(DATE_FORMAT) + '_' + \
                        unicode(self.filter_id_school_classtype) + '_' + \
                        unicode(self.filter_id_school_location) + '_' + \
                        unicode(self.filter_id_teacher) + '_' + \
                        unicode(self.filter_id_school_level) + '_' + \
                        unicode(self.filter_id_status) + '_' + \
                        unicode(self.filter_public) + '_' + \
                        self.sorting + '_' + \
                        unicode(self.trend_medium) + '_' + \
                        unicode(self.trend_high)

            rows = cache.ram(cache_key , lambda: self._get_day_table(), time_expire=CACHE_LONG)

        return rows


    def get_day_list(self):
        '''
            Format rows as list
        '''
        os_gui = current.globalenv['os_gui']
        DATE_FORMAT = current.globalenv['DATE_FORMAT']
        T = current.globalenv['T']
        date_formatted = self.date.strftime(DATE_FORMAT)

        rows = self.get_day_rows()

        get_status = self._get_day_row_status

        classes = []
        for i, row in enumerate(rows):
            repr_row = list(rows[i:i+1].render())[0]

            # get status
            status_result = get_status(row)
            status = status_result['status']

            # get teachers
            teacher_id = row.classes_teachers.auth_teacher_id
            teacher_id2 = row.classes_teachers.auth_teacher_id2
            teacher = repr_row.classes_teachers.auth_teacher_id
            teacher2 = repr_row.classes_teachers.auth_teacher_id2
            teacher_role = row.classes_teachers.teacher_role
            teacher_role2 = row.classes_teachers.teacher_role2

            # check filter for teachers
            if self.filter_id_teacher:
                teacher_filter_id = int(self.filter_id_teacher)
                filter_check = (teacher_filter_id == teacher_id or
                                teacher_filter_id == teacher_id2)
                if not filter_check:
                    # break loop if it's not the teacher searched for
                    continue

            # set holidays
            holiday = False
            holiday_description = ''
            if row.school_holidays.id:
                holiday = True
                holiday_description = row.school_holidays.Description

            cancelled = False
            cancelled_description = ''
            if status == 'cancelled':
                cancelled = True
                cancelled_description = row.classes_otc.Description

            subteacher = False
            if ( row.classes_teachers.teacher_role == 1 or
                 row.classes_teachers.teacher_role2 == 1 ):
                subteacher = True

            # shop url
            shop_url = URL('shop', 'classes_book_options', vars={'clsID': row.classes.id,
                                                                 'date' : date_formatted},
                           scheme=True,
                           host=True,
                           extension='')

            # populate class data
            data = dict()
            data['ClassesID'] = row.classes.id
            data['LocationID'] = row.classes.school_locations_id
            data['Location'] = repr_row.classes.school_locations_id
            data['Starttime'] = repr_row.classes.Starttime
            data['time_starttime'] = row.classes.Starttime
            data['Endtime'] = repr_row.classes.Endtime
            data['time_endtime'] = row.classes.Endtime
            data['ClassTypeID'] = row.classes.school_classtypes_id
            data['ClassType'] = repr_row.classes.school_classtypes_id
            data['TeacherID'] = teacher_id
            data['TeacherID2'] = teacher_id2
            data['Teacher'] = teacher
            data['Teacher2'] = teacher2
            data['LevelID'] = row.classes.school_levels_id
            data['Level'] = repr_row.classes.school_levels_id
            data['Subteacher'] = subteacher
            data['Cancelled'] = cancelled
            data['CancelledDescription'] = cancelled_description
            data['Holiday'] = holiday
            data['HolidayDescription'] = holiday_description
            data['MaxStudents'] = row.classes.Maxstudents or 0 # Spaces for a class
            data['CountAttendance'] = row.classes_schedule_count.Attendance or 0
            data['CountAttendanceOnlineBooking'] = row.classes_schedule_count.OnlineBooking or 0
            data['BookingSpacesAvailable'] = self._get_day_list_booking_spaces(row)
            data['BookingStatus'] = self._get_day_list_booking_status(row)
            data['BookingOpen'] = self.bookings_open
            data['LinkShop'] = shop_url

            classes.append(data)

        return classes


class Classcard:
    '''
        Customer classcard
    '''

    def __init__(self, ccdID):
        '''
            Set some initial values
        '''
        db = current.globalenv['db']

        self.ccdID = ccdID
        self.classcard = db.customers_classcards(ccdID)
        self.scdID = self.classcard.school_classcards_id
        self.school_classcard = db.school_classcards(self.scdID)

        self.price = self.school_classcard.Price
        self.name = self.school_classcard.Name
        self.classes = self.school_classcard.Classes
        self.unlimited = self.school_classcard.Unlimited
        self.cuID = self.classcard.auth_customer_id

    def get_name(self):
        '''
            Returns name of classcard
        '''
        return self.school_classcard.Name

    def get_auth_customer_id(self):
        '''
            Returns auth_customer_id
        '''
        return self.classcard.auth_customer_id

    def get_tax_rate_percentage(self):
        '''
            Returns the tax rate percentage for a card
            Returns None if nothing is set
        '''
        db = current.globalenv['db']

        trID = self.school_classcard.tax_rates_id
        if not trID:
            return None

        tax_rate = db.tax_rates(trID)

        return tax_rate.Percentage


    def get_rows_classes_taken(self):
        '''
            Returns rows of classes taken on this card
        '''
        db = current.globalenv['db']

        fields = [
            db.classes_attendance.id,
            db.classes_attendance.ClassDate,
            db.classes.id,
            db.classes.school_locations_id,
            db.classes.school_classtypes_id,
            db.classes.Starttime
        ]

        orderby_sql = 'clatt.ClassDate'

        query = '''
        SELECT clatt.id,
               clatt.ClassDate,
               cla.id,
               CASE WHEN cotc.school_locations_id IS NOT NULL
                    THEN cotc.school_locations_id
                    ELSE cla.school_locations_id
               END AS school_locations_id,
               CASE WHEN cotc.school_classtypes_id IS NOT NULL
                    THEN cotc.school_classtypes_id
                    ELSE cla.school_classtypes_id
               END AS school_classtypes_id,
               CASE WHEN cotc.Starttime IS NOT NULL
                    THEN cotc.Starttime
                    ELSE cla.Starttime
               END AS Starttime
        FROM classes_attendance clatt
        LEFT JOIN classes cla on cla.id = clatt.classes_id
        LEFT JOIN
            ( SELECT id,
                     classes_id,
                     ClassDate,
                     Status,
                     school_locations_id,
                     school_classtypes_id,
                     Starttime,
                     Endtime,
                     auth_teacher_id,
                     teacher_role,
                     auth_teacher_id2,
                     teacher_role2
              FROM classes_otc ) cotc
            ON clatt.classes_id = cotc.classes_id AND clatt.ClassDate = cotc.ClassDate
        WHERE clatt.customers_classcards_id = {ccdID}
        ORDER BY {orderby_sql}
        '''.format(orderby_sql=orderby_sql,
                   ccdID=self.ccdID)

        rows = db.executesql(query, fields=fields)

        return rows


    def get_classes_remaining(self):
        '''
            :return: Remaining classes
        '''
        db = current.globalenv['db']

        if self.unlimited:
            return 'unlimited'

        query = (db.classes_attendance.customers_classcards_id == self.ccdID) & \
                (db.classes_attendance.BookingStatus != 'cancelled')
        used = db(query).count()
        return self.classes - used


    def get_classes_remaining_formatted(self):
        '''
            :return: Representation of remaining classes
        '''
        db = current.globalenv['db']
        T = current.globalenv['T']

        remaining = self.get_classes_remaining()

        if remaining == 'unlimited':
            remaining = T('Unlimited')

        text = T("Classes")
        if remaining == 1:
           text = T("Class")

        return SPAN(unicode(remaining), ' ', text, ' ', T("remaining"))


    def _get_allowed_classes_format(self, class_ids):
        '''
            :param class_ids: list of db.classes.id
            :return: html table
        '''
        T = current.globalenv['T']
        db = current.globalenv['db']
        TODAY_LOCAL = current.globalenv['TODAY_LOCAL']

        query = (db.classes.AllowAPI == True) & \
                (db.classes.id.belongs(class_ids)) & \
                (db.classes.Startdate <= TODAY_LOCAL) & \
                ((db.classes.Enddate == None) |
                 (db.classes.Enddate >= TODAY_LOCAL))
        rows = db(query).select(db.classes.ALL,
                                orderby=db.classes.Week_day|db.classes.Starttime|db.classes.school_locations_id)

        header = THEAD(TR(TH(T('Day')),
                          TH(T('Time')),
                          TH(T('Location')),
                          TH(T('Class'))))
        table = TABLE(header, _class='table table-striped table-hover')
        for i, row in enumerate(rows):
            repr_row = list(rows[i:i + 1].render())[0]

            tr = TR(TD(repr_row.Week_day),
                    TD(repr_row.Starttime, ' - ', repr_row.Endtime),
                    TD(repr_row.school_locations_id),
                    TD(repr_row.school_classtypes_id))

            table.append(tr)

        return table


    # def get_allowed_classes_enrollment(self, public_only=True, formatted=False):
    #     '''
    #         :return: return: list of db.classes.db that are allowed to be enrolled in using this subscription
    #     '''
    #     permissions = self.get_class_permissions(public_only=public_only)
    #     class_ids = []
    #     for clsID in permissions:
    #         try:
    #             if permissions[clsID]['Enroll']:
    #                 class_ids.append(clsID)
    #         except KeyError:
    #             pass
    #
    #     if not formatted:
    #         return class_ids
    #     else:
    #         return self._get_allowed_classes_format(class_ids)


    def get_allowed_classes_booking(self, public_only=True, formatted=False):
        """
            :return: return: list of db.classes.db that are allowed to be booked using this subscription
        """
        permissions = self.get_class_permissions(public_only=public_only)
        class_ids = []
        for clsID in permissions:
            try:
                if permissions[clsID]['ShopBook']:
                    class_ids.append(clsID)
            except KeyError:
                pass


        if not formatted:
            return class_ids
        else:
            return self._get_allowed_classes_format(class_ids)


    def get_allowed_classes_attend(self, public_only=True, formatted=False):
        """
            :return: return list of db.classes that are allowed to be attended using this subscription
        """
        permissions = self.get_class_permissions(public_only=public_only)
        class_ids = []
        for clsID in permissions:
            try:
                if permissions[clsID]['Attend']:
                    class_ids.append(clsID)
            except KeyError:
                pass


        if not formatted:
            return class_ids
        else:
            return self._get_allowed_classes_format(class_ids)


    def _get_class_permissions_format(self, permissions):
        '''
            :param permissions: dictionary of class permissions
            :return: html table
        '''
        T = current.globalenv['T']
        db = current.globalenv['db']
        os_gui = current.globalenv['os_gui']
        TODAY_LOCAL = current.globalenv['TODAY_LOCAL']

        class_ids = []
        for clsID in permissions:
            class_ids.append(clsID)

        query = (db.classes.AllowAPI == True) & \
                (db.classes.id.belongs(class_ids)) & \
                (db.classes.Startdate <= TODAY_LOCAL) & \
                ((db.classes.Enddate == None) |
                 (db.classes.Enddate >= TODAY_LOCAL))
        rows = db(query).select(db.classes.ALL,
                                orderby=db.classes.Week_day|db.classes.Starttime|db.classes.school_locations_id)

        header = THEAD(TR(TH(T('Day')),
                          TH(T('Time')),
                          TH(T('Location')),
                          TH(T('Class')),
                          #TH(T('Enroll')),
                          TH(T('Book in advance')),
                          TH(T('Attend'))))

        table = TABLE(header, _class='table table-striped table-hover')
        green_check = SPAN(os_gui.get_fa_icon('fa-check'), _class='text-green')

        for i, row in enumerate(rows):
            repr_row = list(rows[i:i + 1].render())[0]

            class_permission = permissions[row.id]
            enroll = class_permission.get('Enroll', '')
            shopbook = class_permission.get('ShopBook', '')
            attend = class_permission.get('Attend', '')

            if enroll:
                enroll = green_check

            if shopbook:
                shopbook = green_check

            if attend:
                attend = green_check

            tr = TR(TD(repr_row.Week_day),
                    TD(repr_row.Starttime, ' - ', repr_row.Endtime),
                    TD(repr_row.school_locations_id),
                    TD(repr_row.school_classtypes_id),
                    #TD(enroll),
                    TD(shopbook),
                    TD(attend))

            table.append(tr)

        return table


    def get_class_permissions(self, public_only=True, formatted=False):
        '''
            :return: return list of class permissons (clsID: enroll, book in shop, attend)
        '''
        db = current.globalenv['db']

        # get groups for classcard
        query = (db.school_classcards_groups_classcards.school_classcards_id == self.scdID)
        rows = db(query).select(db.school_classcards_groups_classcards.school_classcards_groups_id)

        group_ids = []
        for row in rows:
            group_ids.append(row.school_classcards_groups_id)


        # get permissions for classcard group
        left = [db.classes.on(db.classes_school_classcards_groups.classes_id == db.classes.id)]
        query = (db.classes_school_classcards_groups.school_classcards_groups_id.belongs(group_ids))

        if public_only:
            query &= (db.classes.AllowAPI == True)

        rows = db(query).select(db.classes_school_classcards_groups.ALL,
                                left=left)

        permissions = {}
        for row in rows:
            clsID = row.classes_id
            if clsID not in permissions:
                permissions[clsID] = {}

            if row.Enroll:
                permissions[clsID]['Enroll'] = True
            if row.ShopBook:
                permissions[clsID]['ShopBook'] = True
            if row.Attend:
                permissions[clsID]['Attend'] = True

        if not formatted:
            return permissions
        else:
            return self._get_class_permissions_format(permissions)


class ClasscardsHelper:
    '''
        Class that contains functions for classcards
    '''

    def set_classes_taken(self, ccdID):
        '''
            Returns payment for a cuID and wspID
        '''
        db = current.globalenv['db']

        query = (db.classes_attendance.customers_classcards_id == ccdID) & \
                (db.classes_attendance.BookingStatus != 'cancelled')
        count = db(query).count()

        classcard = db.customers_classcards(ccdID)
        classcard.ClassesTaken = count
        classcard.update_record()

    def get_classes_taken(self, ccdID):
        '''
            Returns classes taken on a card
        '''
        db = current.globalenv['db']

        query = (db.classes_attendance.customers_classcards_id == ccdID) & \
                (db.classes_attendance.BookingStatus != 'cancelled')
        count = db(query).count()

        return count

    def get_classes_total(self, ccdID):
        '''
            Returns the total classes on a card
        '''
        db = current.globalenv['db']
        classcard = db.customers_classcards(ccdID)
        school_classcard = db.school_classcards(classcard.school_classcards_id)

        if school_classcard.Unlimited:
            return current.T('Unlimited')
        else:
            return school_classcard.Classes

    def get_classes_remaining(self, ccdID):
        '''
            Returns number of classes remaining on a card
        '''
        taken = self.get_classes_taken(ccdID)
        total = self.get_classes_total(ccdID)

        if total == current.T('Unlimited'):
            return total
        else:
            return total - taken

    def get_classes_available(self, ccdID):
        '''
            Returns True if classes are available on a card
            and False if not.
        '''
        remaining = self.get_classes_remaining(ccdID)

        if remaining > 0:
            available = True
        else:
            available = False

        return available

    def represent_validity(self, validity_months=None, validity_days=None):
        '''
            Represent validity for a school_classcard
        '''
        validity = SPAN()

        if validity_months:
            months = SPAN(validity_months, T(" Month"))
            if validity_months > 1:
                months.append(T('s'))
            validity.append(months)
            validity.append(' ')

        if validity_months and validity_days:
            validity.append(T(" and "))

        if validity_days:
            days = SPAN(validity_days, T(" Day"))
            if validity_days > 1:
                days.append(T('s'))
            validity.append(days)

        return validity


class Workshop:
    def __init__(self, wsID):
        self.wsID = wsID

        db = current.globalenv['db']
        query = (db.workshops.id == self.wsID)
        rows = db(query).select(db.workshops.ALL)
        self.workshop = rows.first()
        repr_row = rows.render(0)

        self.Name = self.workshop.Name
        self.Tagline = self.workshop.Tagline or ''
        self.Startdate = self.workshop.Startdate
        self.Startdate_formatted = repr_row.Startdate
        self.Enddate = self.workshop.Enddate
        self.Enddate_formatted = repr_row.Enddate
        self.Starttime = self.workshop.Starttime
        self.Endtime = self.workshop.Endtime
        self.auth_teacher_id = self.workshop.auth_teacher_id
        self.auth_teacher_id2 = self.workshop.auth_teacher_id2
        self.auth_teacher_name = repr_row.auth_teacher_id
        self.auth_teacher_name2 = repr_row.auth_teacher_id2
        self.Preview = self.workshop.Preview
        self.Description = self.workshop.Description
        self.school_locations_id = self.workshop.school_locations_id
        self.school_location = repr_row.school_locations_id
        self.picture = self.workshop.picture
        self.thumbsmall = self.workshop.thumbsmall
        self.thumblarge = self.workshop.thumblarge
        self.PublicWorkshop = self.workshop.PublicWorkshop


    def get_products(self, filter_public = False):
        '''
            :param filter_public: boolean - show only Public products when set to True
            :return: workshop product rows for a workshop
        '''
        db = current.globalenv['db']

        query = (db.workshops_products.workshops_id == self.wsID)
        if filter_public:
            query &= (db.workshops_products.PublicProduct == True)

        rows = db(query).select(db.workshops_products.ALL,
                                orderby = ~db.workshops_products.FullWorkshop)

        return rows


    def get_full_workshop_price(self):
        '''
            :return: price of full workshop product
        '''
        full_ws_product = self.get_products().first()

        return full_ws_product.Price


    def get_activities(self):
        db = current.globalenv['db']

        query = (db.workshops_activities.workshops_id == self.wsID)
        rows = db(query).select(db.workshops_activities.ALL,
                                orderby = db.workshops_activities.Activitydate|\
                                          db.workshops_activities.Starttime)

        return rows


    def update_dates_times(self):
        '''
            After adding/editing/deleting a workshop activity, call this function
            to update the dates & times on the db.workshops record
        '''
        activities = self.get_activities()

        time_from  = None
        time_until = None
        date_from  = None
        date_until = None
        if len(activities) > 0:
            date_from = activities[0].Activitydate
            date_until = activities[0].Activitydate
            time_from = activities[0].Starttime
            time_until = activities[0].Endtime

        if len(activities) > 1:
            date_until = activities[len(activities) - 1].Activitydate
            time_until = activities[len(activities) - 1].Endtime

        self.workshop.Startdate = date_from
        self.workshop.Enddate   = date_until
        self.workshop.Starttime = time_from
        self.workshop.Endtime   = time_until
        self.workshop.update_record()


    def cancel_orders_with_sold_out_products(self):
        '''
            After selling a product online or adding a customer to a product, check whether products aren't sold out.
            If a product is sold out, check for open orders containing the sold out product and cancel them.
        '''
        db = current.globalenv['db']

        products = self.get_products()
        for product in products:
            wsp = WorkshopProduct(product.id)
            if wsp.is_sold_out():
                # Cancel all unpaid orders with this product
                left = [db.customers_orders.on(
                    db.customers_orders_items.customers_orders_id == db.customers_orders.id)]
                query = ((db.customers_orders.Status == 'awaiting_payment') |
                         (db.customers_orders.Status == 'received')) & \
                        (db.customers_orders_items.workshops_products_id == product.id)
                sold_out_rows = db(query).select(db.customers_orders_items.ALL,
                                                 db.customers_orders.ALL,
                                                 left=left)
                for sold_out_row in sold_out_rows:
                    order = Order(sold_out_row.customers_orders.id)
                    order.set_status_cancelled()


class WorkshopsHelper:
    def get_customer_info(self, wsp_cuID, wsID, WorkshopInfo, resend_link=''):
        '''
            Returns checkboxes for payment confirmation and workshop info
            wsp_cuID = workshops_products_customers.id
        '''
        T = current.globalenv['T']

        form = SQLFORM.factory(
            Field('WorkshopInfo', 'boolean',
                  default=WorkshopInfo)
        )

        hidden_field_id = INPUT(_type="hidden",
                                _name="id",
                                _value=wsp_cuID)

        inputs = DIV(
            form.custom.widget.WorkshopInfo, ' ',
            LABEL(T('Event Info'),
                  _for='no_table_WorkshopInfo')
        )

        form = DIV(form.custom.begin,
                   #table,
                   inputs,
                   hidden_field_id,
                   form.custom.end,
                   resend_link)

        return form


    def get_all_workshop_customers(self, wsID, count=False, ids_only=False):
        '''
            Returns a list of gluon.dal.row objects for each customer attending
            a workshop

            If count is set to True, return a count of customers attending
            the workshop
        '''
        # Get list of all customers with email for a workshop
        # Get all workshop_products_ids
        db = current.globalenv['db']
        query = (db.workshops_products.workshops_id == wsID)
        rows = db(query).select(db.workshops_products.id)
        products_ids = []
        for row in rows:
            products_ids.append(row.id)

        wspID = db.workshops_products_customers.workshops_products_id

        query = (wspID.belongs(products_ids))
        customers_rows = []
        left = [db.auth_user.on(db.auth_user.id == \
                                db.workshops_products_customers.auth_customer_id)]
        rows = db(query).select(db.workshops_products_customers.ALL,
                                db.auth_user.id,
                                db.auth_user.trashed,
                                db.auth_user.thumbsmall,
                                db.auth_user.first_name,
                                db.auth_user.last_name,
                                left=left )

        for row in rows:
            if row not in customers_rows:
                customers_rows.append(row)

        if count:
            return_value = len(customers_rows)
        elif ids_only:
            return_value = [row.auth_user.id for row in rows]
        else:
            return_value = customers_rows

        return return_value


    def get_product_sell_buttons(self, cuID, wsID, wspID, cid):
        """
            Returns buttons for workshop_product_sell list type
            This is a select button to select a customer to sell a product to
        """
        db = current.globalenv['db']
        os_gui = current.globalenv['os_gui']

        buttons = DIV(DIV(current.T("Already added"), _class='btn'),
                     _class='btn-group pull-right hidden')

        products_sold = db.workshops_products_customers
        # find full workshop id
        fwspID = self.get_full_workshop_product_id_for_workshop(wsID)

        # check if full workshop has been sold
        check_full_ws = products_sold(workshops_products_id=fwspID,
                                      auth_customer_id=cuID)

        # check if product has been sold
        check = products_sold(workshops_products_id=wspID,
                              auth_customer_id=cuID)

        if not check and not check_full_ws:
            buttons = DIV(os_gui.get_button('add',
                                        URL('events',
                                            'ticket_sell_to_customer',
                                            vars={'cuID' : cuID,
                                                  'wsID' : wsID,
                                                  'wspID': wspID},
                                            extension='')),
                         A(current.T('To waitinglist'),
                           _href=URL('events',
                                     'ticket_sell_to_customer',
                                     vars={'cuID'     : cuID,
                                           'wsID'     : wsID,
                                           'wspID'    : wspID,
                                           'waiting'  : True},
                                     extension=''),
                           _class='btn btn-default btn-sm'),
                        _class='btn-group pull-right')

        return buttons

    def get_full_workshop_product_id_for_workshop(self, wsID):
        '''
            Return id of full workshop product
        '''
        db = current.globalenv['db']
        query = (db.workshops_products.workshops_id == wsID) & \
                (db.workshops_products.FullWorkshop == True)

        return db(query).select().first().id


class WorkshopProduct:
    '''
        Class for workshop products
    '''
    def __init__(self, wspID):
        db = current.globalenv['db']

        self.wspID = int(wspID)
        self.workshop_product = db.workshops_products(self.wspID)
        self.workshop         = db.workshops(self.workshop_product.workshops_id)

        self.name          = self.workshop_product.Name
        self.workshop_name = self.workshop.Name
        self.wsID          = self.workshop.id
        self.tax_rates_id  = self.workshop_product.tax_rates_id

        self._set_price()


    def _set_price(self):
        if self.workshop_product.Price:
            self.price = self.workshop_product.Price
        else:
            self.price = 0


    def get_price(self):
        return self.workshop_product.Price


    def get_tax_rate_percentage(self):
        '''
            Returns the tax percentage for a workshop product, if any
        '''
        db = current.globalenv['db']

        if self.workshop_product.tax_rates_id:
            tax_rate = db.tax_rates(self.workshop_product.tax_rates_id)
            tax_rate_percentage = tax_rate.Percentage
        else:
            tax_rate_percentage = None

        return tax_rate_percentage


    def get_activities(self):
        '''
            Returns all activities for a workshop product
        '''
        db = current.globalenv['db']

        if self.workshop_product.FullWorkshop:
            query = (db.workshops_activities.workshops_id == self.workshop.id)
            left = None
        else:
            query = (db.workshops_products_activities.workshops_products_id == self.wspID)
            left = [ db.workshops_activities.on(
                db.workshops_products_activities.workshops_activities_id ==
                db.workshops_activities.id) ]


        rows = db(query).select(db.workshops_activities.ALL,
                                left=left,
                                orderby=db.workshops_activities.Activitydate|\
                                        db.workshops_activities.Starttime)

        return rows


    def is_sold_to_customer(self, cuID):
        '''
            :param cuID: db.auth_user.id
            :return: True when sold to customer, False when not
        '''
        db = current.globalenv['db']

        query = (db.workshops_products_customers.auth_customer_id == cuID) & \
                (db.workshops_products_customers.workshops_products_id == self.wspID)
        count = db(query).count()

        if count > 0:
            return True
        else:
            return False


    def is_sold_out(self):
        '''
            This function checks if a product is sold out
            It's sold out when any of the activities it contains is completely full
        '''
        def activity_list_customers_get_list_activity_query(wsaID):
            '''
                Returns a query that returns a set of all customers in a specific
                workshop activity, without the full workshop customers
            '''
            query = (db.workshops_activities.id ==
                     db.workshops_products_activities.workshops_activities_id) & \
                    (db.workshops_products_customers.workshops_products_id ==
                     db.workshops_products_activities.workshops_products_id) & \
                    (db.workshops_products_activities.workshops_activities_id ==
                     wsaID) & \
                    (db.workshops_products_customers.Waitinglist == False)

            return query

        def activity_count_reserved(wsaID):
            # count full workshop customers
            query = (db.workshops_products_customers.workshops_products_id == fwsID)
            count_full_ws = db(query).count()
            # count activity customers
            query = activity_list_customers_get_list_activity_query(wsaID)
            count_activity = db(query).count()

            return count_full_ws + count_activity

        db = current.globalenv['db']
        check = False

        fwsID = workshops_get_full_workshop_product_id(self.workshop.id)


        if self.wspID == fwsID:
            # Full workshops check, check if any activity is full
            query = (db.workshops_activities.workshops_id == self.workshop.id)
            rows = db(query).select(db.workshops_activities.ALL)
            for row in rows:
                reserved = activity_count_reserved(row.id)
                if reserved >= row.Spaces:
                    check = True
                    break
        else:
            # Product check, check if any a activity is full
            left = [ db.workshops_activities.on(
                db.workshops_products_activities.workshops_activities_id ==
                db.workshops_activities.id
            )]
            query = (db.workshops_products_activities.workshops_products_id == self.wspID)
            rows = db(query).select(db.workshops_products_activities.ALL,
                                    db.workshops_activities.Spaces,
                                    left=left)
            for row in rows:
                wsaID = row.workshops_products_activities.workshops_activities_id
                reserved = activity_count_reserved(wsaID)
                if reserved >= row.workshops_activities.Spaces:
                    check = True
                    break

        return check


    def add_to_shoppingcart(self, cuID):
        '''
            Add a workshop product to the shopping cart of a customer
        '''
        db = current.globalenv['db']

        db.customers_shoppingcart.insert(
            auth_customer_id = cuID,
            workshops_products_id = self.wspID
        )


    def sell_to_customer(self, cuID, waitinglist=False, invoice=True):
        '''
            Sells a workshop to a customer and creates an invoice
            Creates an invoice when a workshop product is sold
        '''
        db = current.globalenv['db']
        T = current.globalenv['T']

        info = False
        if self.workshop.AutoSendInfoMail:
            info = True

        wspID = self.wspID
        wspcID = db.workshops_products_customers.insert(
            auth_customer_id=cuID,
            workshops_products_id=wspID,
            Waitinglist=waitinglist,
            WorkshopInfo=info)

        ##
        # Add invoice
        ##
        if invoice and not waitinglist and self.price > 0:
            igpt = db.invoices_groups_product_types(ProductType = 'wsp')

            description = self.workshop_name + ' - ' + self.name

            iID = db.invoices.insert(
                invoices_groups_id = igpt.invoices_groups_id,
                Description = description,
                Status = 'sent'
                )

            # link invoice to sold workshop product for customer
            db.invoices_workshops_products_customers.insert(
                invoices_id = iID,
                workshops_products_customers_id = wspcID )

            # create object to set Invoice# and due date
            invoice = Invoice(iID)
            next_sort_nr = invoice.get_item_next_sort_nr()

            price = self.price

            iiID = db.invoices_items.insert(
                invoices_id  = iID,
                ProductName  = T("Event"),
                Description  = description,
                Quantity     = 1,
                Price        = price,
                Sorting      = next_sort_nr,
                tax_rates_id = self.tax_rates_id,
            )

            invoice.set_amounts()
            invoice.link_to_customer(cuID)

        ##
        # Send info mail to customer if we have some practical info
        ##
        if self.workshop.AutoSendInfoMail:
            osmail = OsMail()
            msgID = osmail.render_email_template('workshops_info_mail', workshops_products_customers_id=wspcID)
            osmail.send(msgID, cuID)

        if not waitinglist:
            # Check if sold out
            if self.is_sold_out():
                # Cancel all unpaid orders with a sold out product for this workshop
                ws = Workshop(self.wsID)
                ws.cancel_orders_with_sold_out_products()

        return wspcID


class WorkshopSchedule:
    def __init__(self, filter_date_start,
                       filter_date_end = None,
                       filter_archived = True,
                       filter_only_public = True,
                       sorting = 'date'):

        self.filter_date_start = filter_date_start
        self.filter_date_end = filter_date_end
        self.filter_archived = filter_archived
        self.filter_only_public = filter_only_public

        self.sorting = sorting

    def _get_workshops_rows_filter_query(self):
        '''
            Apply filters to workshops
        '''
        where = ''
        if self.filter_archived:
            where += "AND ws.Archived='F'"
            where += ' '

        if self.filter_only_public:
            where += "AND ws.PublicWorkshop='T'"
            where += ' '

        #TODO: check first activity date as startdate ... or create function in workshops.py that updates dates
        # & times for workshops when an activity is added/updated/deleted.
        if self.filter_date_start:
            where += "AND ws.Startdate >= '" + unicode(self.filter_date_start) + "'"
            where += ' '

        if self.filter_date_end:
            where += "AND ws.Enddate <= " + unicode(self.filter_date_end) + "'"
            where += ' '

        return where


    def _get_workshops_rows_orderby(self):
        '''
            Apply right sorting to rows
        '''
        db = current.globalenv['db']
        orderby = 'ws.Startdate'

        if self.sorting == 'name':
            orderby = 'ws.Name'

        return orderby


    def get_workshops_rows(self):
        '''
            Gets workshop rows
        '''
        db = current.globalenv['db']

        orderby_sql = self._get_workshops_rows_orderby()
        where_filter = self._get_workshops_rows_filter_query()

        fields = [
            db.workshops.id,
            db.workshops.Name,
            db.workshops.Tagline,
            db.workshops.Startdate,
            db.workshops.Enddate,
            db.workshops.Starttime,
            db.workshops.Endtime,
            db.workshops.auth_teacher_id,
            db.workshops.auth_teacher_id2,
            db.workshops.Preview,
            db.workshops.Description,
            db.workshops.school_locations_id,
            db.workshops.picture,
            db.workshops.thumbsmall,
            db.workshops.thumblarge,
            db.workshops_products.Price
        ]

        query = '''
        SELECT ws.id,
               ws.Name,
               ws.Tagline,
               ws.Startdate,
               ws.Enddate,
               ws.Starttime,
               ws.Endtime,
               ws.auth_teacher_id,
               ws.auth_teacher_id2,
               ws.Preview,
               ws.Description,
               ws.school_locations_id,
               ws.picture,
               ws.thumbsmall,
               ws.thumblarge,
               wsp.Price
        FROM workshops ws
        LEFT JOIN
            ( SELECT id, workshops_id, Price FROM workshops_products
              WHERE FullWorkshop = 'T' ) wsp
            ON ws.id = wsp.workshops_id
        WHERE ws.id > 0
              {where_filter}
        ORDER BY {orderby_sql}
        '''.format(orderby_sql = orderby_sql,
                   where_filter = where_filter)

        rows = db.executesql(query, fields=fields)

        return rows


    def get_workshops_list(self):
        '''
            Returns list of workshops
        '''
        rows = self.get_workshops_rows().as_list()


    def _get_workshops_shop(self):
        """
            Format list of workshops in a suitable way for the shop
        """
        def new_workshop_month():
            _class = 'workshops-list-month'

            return DIV(H2(last_day_month.strftime('%B %Y'), _class='center'), _class=_class)

        request = current.globalenv['request']
        os_gui = current.globalenv['os_gui']
        T = current.globalenv['T']
        TODAY_LOCAL = current.globalenv['TODAY_LOCAL']

        rows = self.get_workshops_rows()

        current_month = TODAY_LOCAL
        last_day_month = get_last_day_month(current_month)

        workshops_month = new_workshop_month()
        workshops_month_body = DIV(_class='box-body')


        workshops = DIV()

        for i, row in enumerate(rows):
            repr_row = list(rows[i:i + 1].render())[0]

            more_info = os_gui.get_button('noicon',
                URL('event', vars={'wsID':row.workshops.id}),
                title=T('More info...'),
                btn_class='btn-link',
                btn_size='',
                _class='workshops-list-workshop-more-info')

            # Check if we're going into a later month
            if row.workshops.Startdate > last_day_month:
                if len(workshops_month_body) >= 1:
                    # check if we have more in the month than just the title (the 1 in len())
                    workshops_month.append(DIV(workshops_month_body, _class='box box-solid'))
                    workshops.append(workshops_month)
                last_day_month = get_last_day_month(row.workshops.Startdate)
                workshops_month = new_workshop_month()
                workshops_month_body = DIV(_class='box-body')

            startdate = SPAN(row.workshops.Startdate.strftime('%d %B').lstrip("0").replace(" 0", " "), _class='label_date')
            enddate = ''
            if not row.workshops.Startdate == row.workshops.Enddate:
                enddate = SPAN(row.workshops.Enddate.strftime('%d %B').lstrip("0").replace(" 0", " "), _class='label_date')
            workshop = DIV(
                DIV(DIV(DIV(repr_row.workshops.thumblarge, _class='workshops-list-workshop-image center'),
                        _class='col-xs-12 col-sm-12 col-md-3'),
                        DIV(A(H3(row.workshops.Name), _href=URL('shop', 'event', vars={'wsID':row.workshops.id})),
                            H4(repr_row.workshops.Tagline),
                            DIV(os_gui.get_fa_icon('fa-calendar-o'), ' ',
                                startdate, ' ',
                                repr_row.workshops.Starttime, ' - ',
                                enddate, ' ',
                                repr_row.workshops.Endtime,
                                _class='workshops-list-workshop-date'),
                            DIV(os_gui.get_fa_icon('fa-user-o'), ' ', repr_row.workshops.auth_teacher_id, _class='workshops-list-workshop-teacher'),
                            DIV(os_gui.get_fa_icon('fa-map-marker'), ' ', repr_row.workshops.school_locations_id, _class='workshops-list-workshop-location'),
                            BR(),
                            more_info,
                            _class='col-xs-12 col-sm-12 col-md-9 workshops-list-workshop-info'),
                        _class=''),
                _class='workshops-list-workshop col-md-8 col-md-offset-2 col-xs-12')

            workshops_month_body.append(workshop)

            # if we're at the last row, add the workshops to the page
            if i + 1 == len(rows):
                workshops_month.append(DIV(workshops_month_body, _class='box box-solid'))
                workshops.append(workshops_month)

        return workshops


    def get_workshops_shop(self):
        """
            Use caching when not running as test to return the workshops list in the shop
        """
        web2pytest = current.globalenv['web2pytest']
        request = current.globalenv['request']
        auth = current.globalenv['auth']

        # Don't cache when running tests
        if web2pytest.is_running_under_test(request, request.application):
            rows = self._get_workshops_shop()
        else:
            cache = current.globalenv['cache']
            CACHE_LONG = current.globalenv['CACHE_LONG']
            cache_key = 'openstudio_workshops_workshops_schedule_shop'

            rows = cache.ram(cache_key , lambda: self._get_workshops_shop(), time_expire=CACHE_LONG)

        return rows


class Order:
    '''
        Class containing functions for OpenStudio orders
    '''
    def __init__(self, coID):
        '''
            Init class
        '''
        db = current.globalenv['db']

        self.coID = coID
        self.order = db.customers_orders(coID)


    def set_status_awaiting_payment(self):
        '''
            Change status to 'awaiting_payment'
        '''
        self.order.Status = 'awaiting_payment'
        self.order.update_record()


    def set_status_delivered(self):
        '''
            Change status to 'awaiting_payment'
        '''
        self.order.Status = 'delivered'
        self.order.update_record()


    def set_status_cancelled(self):
        '''
            Change status to 'awaiting_payment'
        '''
        self.order.Status = 'cancelled'
        self.order.update_record()


    def order_item_add_classcard(self, school_classcards_id):
        '''
            :param school_classcards_id: db.school_classcards.id
            :return : db.customers_orders_items.id of inserted item
        '''
        db = current.globalenv['db']
        T  = current.globalenv['T']

        school_classcard = db.school_classcards(school_classcards_id)

        coiID = db.customers_orders_items.insert(
            customers_orders_id  = self.coID,
            school_classcards_id = school_classcards_id,
            ProductName = T('Classcard'),
            Description = school_classcard.Name,
            Quantity = 1,
            Price = school_classcard.Price,
            tax_rates_id = school_classcard.tax_rates_id
        )

        self.set_amounts()

        return coiID


    def order_item_add_workshop_product(self, workshops_products_id):
        '''
            :param workshops_products_id: db.workshops_products.id
            :return: db.customers_orders_items.id of inserted item
        '''
        db = current.globalenv['db']
        T  = current.globalenv['T']

        wsp = db.workshops_products(workshops_products_id)
        ws = db.workshops(wsp.workshops_id)

        coiID = db.customers_orders_items.insert(
            customers_orders_id = self.coID,
            workshops_products_id = workshops_products_id,
            ProductName = T('Event'),
            Description = ws.Name + ' - ' + wsp.Name,
            Quantity = 1,
            Price = wsp.Price,
            tax_rates_id = wsp.tax_rates_id
        )

        self.set_amounts()

        return coiID


    def order_item_add_donation(self, amount, description):
        '''
            :param amount: donation amount
            :param description: donation description
            :return: db.customers_orders.items.id of inserted item 
        '''
        db = current.globalenv['db']
        T  = current.globalenv['T']
        get_sys_property = current.globalenv['get_sys_property']

        sys_property = 'shop_donations_tax_rates_id'
        tax_rates_id = int(get_sys_property(sys_property))


        coiID = db.customers_orders_items.insert(
            customers_orders_id=self.coID,
            Donation=True,
            ProductName=T('Donation'),
            Description=description,
            Quantity=1,
            Price=amount,
            tax_rates_id=tax_rates_id,
        )

        self.set_amounts()

        return coiID


    def order_item_add_class(self, clsID, class_date, attendance_type):
        '''
            :param workshops_products_id: db.workshops_products.id
            :return: db.customers_orders_items.id of inserted item
        '''
        DATE_FORMAT = current.globalenv['DATE_FORMAT']
        TIME_FORMAT = current.globalenv['TIME_FORMAT']
        db = current.globalenv['db']
        T  = current.globalenv['T']

        cls = Class(clsID, class_date)
        prices = cls.get_price()
        if attendance_type == 1:
            price = prices['trial']
            tax_rates_id = prices['trial_tax_rates_id']
        elif attendance_type == 2:
            price = prices['dropin']
            tax_rates_id = prices['dropin_tax_rates_id']

        description = cls.get_invoice_order_description(attendance_type)

        coiID = db.customers_orders_items.insert(
            customers_orders_id = self.coID,
            classes_id = clsID,
            AttendanceType = attendance_type,
            ClassDate = class_date,
            ProductName = T('Class'),
            Description = description,
            Quantity = 1,
            Price = price,
            tax_rates_id = tax_rates_id,
        )

        self.set_amounts()

        return coiID


    def get_order_items_rows(self):
        '''
            :return: db.customers_orders_items rows for order
        '''
        db = current.globalenv['db']

        query = (db.customers_orders_items.customers_orders_id == self.coID)
        rows = db(query).select(db.customers_orders_items.ALL)

        return rows


    def get_amounts(self):
        '''
            Get subtotal, vat and total incl vat
        '''
        db = current.globalenv['db']

        amounts = db.customers_orders_amounts(customers_orders_id = self.coID)

        return amounts


    def set_amounts(self):
        '''
            Set subtotal, vat and total incl vat
        '''
        db = current.globalenv['db']
        # set sums
        sum_subtotal = db.customers_orders_items.TotalPrice.sum()
        sum_vat = db.customers_orders_items.VAT.sum()
        sum_totalvat = db.customers_orders_items.TotalPriceVAT.sum()

        # get info from db
        query = (db.customers_orders_items.customers_orders_id == self.coID)
        rows = db(query).select(sum_subtotal,
                                sum_vat,
                                sum_totalvat)

        sums = rows.first()
        subtotal = sums[sum_subtotal]
        vat = sums[sum_vat]
        total = sums[sum_totalvat]

        if subtotal is None:
            subtotal = 0
        if vat is None:
            vat = 0
        if total is None:
            total = 0

        # check what to do
        amounts = db.customers_orders_amounts(customers_orders_id = self.coID)
        if amounts:
            # update current row
            amounts.TotalPrice = subtotal
            amounts.VAT = vat
            amounts.TotalPriceVAT = total
            amounts.update_record()
        else:
            # insert new row
            db.customers_orders_amounts.insert(
                customers_orders_id = self.coID,
                TotalPrice=subtotal,
                VAT=vat,
                TotalPriceVAT=total
            )


    def deliver(self):
        '''
            Create invoice for order and deliver goods
        '''
        cache_clear_classschedule_api = current.globalenv['cache_clear_classschedule_api']
        get_sys_property = current.globalenv['get_sys_property']
        db = current.globalenv['db']
        T = current.globalenv['T']

        create_invoice = False
        iID = None
        invoice = None
        # Only create an invoice if order amount > 0
        amounts = self.get_amounts()

        sys_property_create_invoice = 'shop_donations_create_invoice'
        create_invoice_for_donations = get_sys_property(sys_property_create_invoice)
        if create_invoice_for_donations == 'on':
            create_invoice_for_donations = True
        else:
            create_invoice_for_donations = False

        if amounts:
            if amounts.TotalPriceVAT > 0:
                if not self.order.Donation or (self.order.Donation and create_invoice_for_donations):
                    create_invoice = True

                    # Create blank invoice
                    igpt = db.invoices_groups_product_types(ProductType='shop')

                    iID = db.invoices.insert(
                        invoices_groups_id=igpt.invoices_groups_id,
                        Description=T('Order #') + unicode(self.coID),
                        Status='sent'
                    )

                    # Link invoice to order
                    db.invoices_customers_orders.insert(
                        customers_orders_id = self.coID,
                        invoices_id = iID
                    )

                    # Call init function for invoices to set Invoice # , etc.
                    invoice = Invoice(iID)
                    invoice.link_to_customer(self.order.auth_customer_id)

        # Add items to the invoice
        rows = self.get_order_items_rows()

        for row in rows:
            ##
            # Only rows where school_classcards_id, workshops_products_id , classes_id or Donation are set
            # are put on the invoice
            ##

            # Check for classcard
            if row.school_classcards_id:
                # Deliver card
                card_start = datetime.date.today()
                scd = SchoolClasscard(row.school_classcards_id)
                ccdID = scd.sell_to_customer(self.order.auth_customer_id,
                                             card_start,
                                             invoice=False)
                # Add card to invoice
                if create_invoice:
                    invoice.item_add_classcard(ccdID)

            # Check for workshop
            if row.workshops_products_id:
                # Deliver workshop product
                wsp = WorkshopProduct(row.workshops_products_id)
                wspcID = wsp.sell_to_customer(self.order.auth_customer_id,
                                              invoice=False)

                # Add workshop product to invoice
                if create_invoice:
                    invoice.item_add_workshop_product(wspcID)

                # Check if sold out
                if wsp.is_sold_out():
                    # Cancel all unpaid orders with a sold out product for this workshop
                    ws = Workshop(wsp.wsID)
                    ws.cancel_orders_with_sold_out_products()

            # Check for classes
            if row.classes_id:
                # Deliver class
                ah = AttendanceHelper()
                if row.AttendanceType == 1:
                    result = ah.attendance_sign_in_trialclass(self.order.auth_customer_id,
                                                              row.classes_id,
                                                              row.ClassDate,
                                                              online_booking=True,
                                                              invoice=False)
                elif row.AttendanceType == 2:
                    result = ah.attendance_sign_in_dropin(self.order.auth_customer_id,
                                                          row.classes_id,
                                                          row.ClassDate,
                                                          online_booking=True,
                                                          invoice=False)

                if create_invoice:
                    invoice.item_add_class(row, result['caID'])

                # Clear api cache to update available spaces
                cache_clear_classschedule_api()

            # Check for donation
            if row.Donation:
                # Add donation line to invoice
                if create_invoice and create_invoice_for_donations:
                    invoice.item_add_donation(row.TotalPriceVAT, row.Description)


        # Notify customer of new invoice
        #if create_invoice:
            #invoice.mail_customer_invoice_created()

        # Update status
        self.set_status_delivered()
        # Notify customer of order delivery
        self._deliver_mail_customer()

        return dict(iID=iID, invoice=invoice)


    def _deliver_mail_customer(self):
        '''
            Notify customer of order delivery
        '''
        osmail = OsMail()
        msgID = osmail.render_email_template('email_template_order_delivered', customers_orders_id=self.coID)

        osmail.send(msgID, self.order.auth_customer_id)


class Invoice:
    '''
        Class that contains functions for an invoice
    '''
    def __init__(self, iID):
        """
            Init function for an invoice
        """
        db = current.globalenv['db']

        self.invoices_id = iID
        self.invoice = db.invoices(iID)
        self.invoice_group = db.invoices_groups(self.invoice.invoices_groups_id)

        if not self.invoice.InvoiceID:
            self._set_invoice_id_duedate_and_amounts()


    def _set_invoice_id_duedate_and_amounts(self):
        '''
            Set invoice id and duedate for an invoice
        '''
        self.invoice.InvoiceID = self._get_next_invoice_id()

        delta = datetime.timedelta(days = self.invoice_group.DueDays)
        self.invoice.DateDue = self.invoice.DateDue + delta

        self.invoice.update_record()

        db = current.globalenv['db']
        db.invoices_amounts.insert(invoices_id = self.invoices_id)


    def _get_next_invoice_id(self):
        '''
            Returns the number for an invoice
        '''
        invoice_id = self.invoice_group.InvoicePrefix

        if self.invoice_group.PrefixYear:
            year = unicode(datetime.date.today().year)
            invoice_id += year

        invoice_id += unicode(self.invoice_group.NextID)

        self.invoice_group.NextID += 1
        self.invoice_group.update_record()

        return invoice_id


    def set_status(self, status):
        '''
            Sets the status of this invoice
        '''
        # check if the status exists:
        actual_status = False
        for status_name, status_text in current.globalenv['invoice_statuses']:
            if status == status_name:
                actual_status = True

        if actual_status:
            self.invoice.Status = status
            self.invoice.update_record()
        else:
            return False


    def set_amounts(self):
        '''
            Set subtotal, vat and total incl vat
        '''
        db = current.globalenv['db']
        # set sums
        sum_subtotal = db.invoices_items.TotalPrice.sum()
        sum_vat = db.invoices_items.VAT.sum()
        sum_totalvat = db.invoices_items.TotalPriceVAT.sum()

        # get info from db
        query = (db.invoices_items.invoices_id == self.invoices_id)
        rows = db(query).select(sum_subtotal,
                                sum_vat,
                                sum_totalvat)

        sums = rows.first()
        subtotal = sums[sum_subtotal]
        vat      = sums[sum_vat]
        total    = sums[sum_totalvat]

        if subtotal is None:
            subtotal = 0
        if vat is None:
            vat = 0
        if total is None:
            total = 0

        # check if we have any payments
        paid = self.get_amount_paid()
        balance = self.get_balance()

        # check what to do
        amounts = db.invoices_amounts(invoices_id = self.invoices_id)
        if amounts:
            # update current row
            amounts.TotalPrice    = subtotal
            amounts.VAT           = vat
            amounts.TotalPriceVAT = total
            amounts.Paid          = paid
            amounts.Balance       = balance
            amounts.update_record()
        else:
            # insert new row
            db.invoices_amounts.insert(
                invoices_id   = self.invoices_id,
                TotalPrice    = subtotal,
                VAT           = vat,
                TotalPriceVAT = total,
                Paid          = paid,
                Balance       = balance)


    def get_amounts(self):
        '''
            Get subtotal, vat and total incl vat
        '''
        db = current.globalenv['db']

        amounts = db.invoices_amounts(invoices_id = self.invoices_id)

        return amounts


    def get_amounts_tax_rates(self, formatted=False):
        '''
            Returns vat for each tax rate as list sorted by tax rate percentage
            format: [ [ Name, Amount ] ]
        '''
        db = current.globalenv['db']
        iID = self.invoices_id
        CURRSYM = current.globalenv['CURRSYM']

        amounts_vat = []
        rows = db().select(db.tax_rates.id, db.tax_rates.Name,
                           orderby=db.tax_rates.Percentage)
        for row in rows:
            sum = db.invoices_items.VAT.sum()
            query = (db.invoices_items.invoices_id == iID) & \
                    (db.invoices_items.tax_rates_id == row.id)

            result = db(query).select(sum).first()

            if not result[sum] is None:
                if formatted:
                    amount = SPAN(CURRSYM, ' ', format(result[sum], '.2f'))
                else:
                    amount = result[sum]
                amounts_vat.append({'Name'   : row.Name,
                                    'Amount' : amount})

        return amounts_vat


    def get_amount_paid(self, formatted=False):
        '''
            Returns the total amount paid
        '''
        db = current.globalenv['db']
        sum = db.invoices_payments.Amount.sum()
        query = (db.invoices_payments.invoices_id == self.invoices_id)

        rows = db(query).select(sum)
        paid = rows.first()[sum]
        if paid is None:
            paid = 0

        if formatted:
            return_value = SPAN(current.globalenv['CURRSYM'], ' ',
                                format(paid, '.2f'))
        else:
            return_value = paid

        return return_value


    def get_balance(self, formatted=False):
        '''
            Returns the balance for an invoice
        '''
        db = current.globalenv['db']
        paid = self.get_amount_paid()
        total = self.get_amounts()['TotalPriceVAT']

        # round numbers first to prevent weird outcomes by decemals
        balance = round(total, 2) - round(paid, 2)

        if formatted:
            return_value = SPAN(current.globalenv['CURRSYM'], ' ',
                                format(balance, '.2f'))
        else:
            return_value = balance

        return return_value


    def get_item_next_sort_nr(self):
        '''
            Returns the next item number for an invoice
            use to set sorting when adding an item
        '''
        db = current.globalenv['db']
        query = (db.invoices_items.invoices_id == self.invoices_id)

        return db(query).count() + 1


    def get_invoice_items_rows(self):
        '''
            :return: db.customers_orders_items rows for order
        '''
        db = current.globalenv['db']

        query = (db.invoices_items.invoices_id == self.invoices_id)
        rows = db(query).select(db.invoices_items.ALL)

        return rows


    def item_add_class(self, order_item_row, caID):
        '''
            :param clsID: db.classes.id
            :param class_date: datetime.date
            :param attendance_type: int 1 or 2 
            :return: db.invoices_items.id
        '''
        DATE_FORMAT = current.globalenv['DATE_FORMAT']
        TIME_FORMAT = current.globalenv['TIME_FORMAT']
        db = current.globalenv['db']
        T  = current.globalenv['T']

        cls = Class(order_item_row.classes_id, order_item_row.ClassDate)

        # link invoice to attendance
        db.invoices_classes_attendance.insert(
            invoices_id=self.invoices_id,
            classes_attendance_id=caID
        )

        # add item to invoice
        next_sort_nr = self.get_item_next_sort_nr()


        iiID = db.invoices_items.insert(
            invoices_id=self.invoices_id,
            ProductName=order_item_row.ProductName,
            Description=order_item_row.Description,
            Quantity=order_item_row.Quantity,
            Price=order_item_row.Price,
            Sorting=next_sort_nr,
            tax_rates_id=order_item_row.tax_rates_id,
        )

        self.set_amounts()

        return iiID


    def item_add_classcard(self, ccdID):
        '''
            :param ccdID: Add customer classcard to invoice
            :return: None
        '''
        db = current.globalenv['db']
        T  = current.globalenv['T']

        classcard = Classcard(ccdID)
        # link invoice to classcard sold to customer
        db.invoices_customers_classcards.insert(
            invoices_id=self.invoices_id,
            customers_classcards_id=ccdID
        )

        # add item to invoice
        next_sort_nr = self.get_item_next_sort_nr()
        price = classcard.price

        iiID = db.invoices_items.insert(
            invoices_id=self.invoices_id,
            ProductName=T("Class card"),
            Description=classcard.name + ' (' + T("Class card") + ' ' + unicode(ccdID) + ')',
            Quantity=1,
            Price=price,
            Sorting=next_sort_nr,
            tax_rates_id=classcard.school_classcard.tax_rates_id,
        )

        self.set_amounts()

        return iiID


    def item_add_workshop_product(self, wspcID):
        '''
            :param wspID: db.workshops_products_id
            :return: db.invoices_items.id
        '''
        db = current.globalenv['db']
        T  = current.globalenv['T']

        wspc = db.workshops_products_customers(wspcID)
        wsp = db.workshops_products(wspc.workshops_products_id)
        ws = db.workshops(wsp.workshops_id)
        # Link invoice to workshop product sold to customer
        db.invoices_workshops_products_customers.insert(
            invoices_id = self.invoices_id,
            workshops_products_customers_id = wspcID
        )

        # Add item to invoice
        next_sort_nr = self.get_item_next_sort_nr()

        iiID = db.invoices_items.insert(
            invoices_id=self.invoices_id,
            ProductName=T('Event'),
            Description=ws.Name + ' - ' + wsp.Name,
            Quantity=1,
            Price=wsp.Price,
            Sorting=next_sort_nr,
            tax_rates_id=wsp.tax_rates_id
        )

        self.set_amounts()

        return iiID


    def item_add_donation(self, amount, description):
        '''
            :param amount: donation amount
            :param description: donation description
            :return: db.customers_orders.items.id of inserted item 
        '''
        db = current.globalenv['db']
        T  = current.globalenv['T']
        get_sys_property = current.globalenv['get_sys_property']

        sys_property = 'shop_donations_tax_rates_id'
        tax_rates_id = int(get_sys_property(sys_property))

        # add item to invoice
        next_sort_nr = self.get_item_next_sort_nr()
        price = amount

        iiID = db.invoices_items.insert(
            invoices_id=self.invoices_id,
            ProductName=T("Donation"),
            Description=description,
            Quantity=1,
            Price=price,
            Sorting=next_sort_nr,
            tax_rates_id=tax_rates_id,
        )

        self.set_amounts()

        return iiID


    def item_add_subscription(self, SubscriptionYear, SubscriptionMonth, description=''):
        """
            :param SubscriptionYear: Year of subscription
            :param SubscriptionMonth: Month of subscription
            :return: db.invoices_items.id
        """
        db = current.globalenv['db']
        DATE_FORMAT = current.globalenv['DATE_FORMAT']

        next_sort_nr = self.get_item_next_sort_nr()

        date = datetime.date(int(SubscriptionYear),
                             int(SubscriptionMonth),
                             1)



        ics = db.invoices_customers_subscriptions(invoices_id = self.invoices_id)
        csID = ics.customers_subscriptions_id
        cs = CustomerSubscription(csID)
        ssuID = cs.ssuID
        ssu = SchoolSubscription(ssuID)
        row = ssu.get_tax_rates_on_date(date)

        if row:
            tax_rates_id = row.school_subscriptions_price.tax_rates_id
        else:
            tax_rates_id = None

        period_start = date
        period_end = get_last_day_month(date)
        price = 0

        # check for alt price
        csap = db.customers_subscriptions_alt_prices
        query = (csap.customers_subscriptions_id == csID) & \
                (csap.SubscriptionYear == SubscriptionYear) & \
                (csap.SubscriptionMonth == SubscriptionMonth)
        csap_rows = db(query).select(csap.ALL)
        if csap_rows:
            csap_row = csap_rows.first()
            price    = csap_row.Amount
            description = csap_row.Description
        else:
            price = ssu.get_price_on_date(date, False)

            broken_period = False
            if cs.startdate > date and cs.startdate <= period_end:
                # Start later in month
                broken_period = True
                period_start = cs.startdate
                delta = period_end - cs.startdate
                cs_days = delta.days + 1
                total_days = period_end.day

            if cs.enddate:
                if cs.enddate >= date and cs.enddate < period_end:
                    # End somewhere in month
                    broken_period = True

                    delta = cs.enddate - date
                    cs_days = delta.days + 1
                    total_days = period_end.day

                    period_end = cs.enddate

            if broken_period:
                price = round(float(cs_days) / float(total_days) * float(price), 2)

            if not description:
                description = cs.name + ' ' + period_start.strftime(DATE_FORMAT) + ' - ' + period_end.strftime(DATE_FORMAT)

        iiID = db.invoices_items.insert(
            invoices_id  = self.invoices_id,
            ProductName  = current.T("Subscription") + ' ' + unicode(csID),
            Description  = description,
            Quantity     = 1,
            Price        = price,
            Sorting      = next_sort_nr,
            tax_rates_id = tax_rates_id,
        )

        self.set_amounts()

        return iiID


    def payment_add(self,
                    amount,
                    date,
                    payment_methods_id,
                    note=None,
                    mollie_payment_id=None):
        '''
            Add payment to invoice
        '''
        db = current.globalenv['db']

        ipID = db.invoices_payments.insert(
            invoices_id = self.invoices_id,
            Amount = amount,
            PaymentDate = date,
            payment_methods_id = payment_methods_id,
            Note = note,
            mollie_payment_id = mollie_payment_id
        )

        self.is_paid()

        return ipID


    def is_paid(self):
        '''
            Check if the status should be changed to 'paid'
        '''
        db = current.globalenv['db']

        # Update invoice status
        sum_payments = db.invoices_payments.Amount.sum()
        query = (db.invoices_payments.invoices_id == self.invoices_id)
        # sum
        amount_paid = db(query).select(sum_payments).first()[sum_payments]
        # Decimal
        amount_paid = Decimal(amount_paid)
        # Rounded to 2 decimals
        amount_paid = Decimal(amount_paid.quantize(Decimal('.01'),
                                                   rounding=ROUND_HALF_UP))


        invoice_amounts = db.invoices_amounts(invoices_id=self.invoices_id)
        invoice_amount = Decimal(invoice_amounts.TotalPriceVAT)
        invoice_total = Decimal(invoice_amount.quantize(Decimal('.01'),
                                                        rounding=ROUND_HALF_UP))

        if amount_paid >= invoice_total:
            self.invoice.Status = 'paid'
            self.invoice.update_record()
            return True
        else:
            return False


    def set_customer_info(self, cuID):
        """
            Set customer information for an invoice
        """
        customer = Customer(cuID)

        address = ''
        if customer.row.address:
            address = ''.join([address, customer.row.address, '\n'])
        if customer.row.city:
            address = ''.join([address, customer.row.city, ' '])
        if customer.row.postcode:
            address = ''.join([address, customer.row.postcode, '\n'])
        if customer.row.country:
            address = ''.join([address, customer.row.country])

        list_name = customer.row.full_name
        if customer.row.company:
            list_name = customer.row.company

        self.invoice.update_record(
            CustomerCompany = customer.row.company,
            CustomerName = customer.row.full_name,
            CustomerListName = list_name,
            CustomerAddress = address,
        )


    def link_to_customer(self, cuID):
        """
            Link invoice to customer
        """
        db = current.globalenv['db']
        # Insert link
        db.invoices_customers.insert(
            invoices_id = self.invoices_id,
            auth_customer_id = cuID
        )

        # Set customer info
        self.set_customer_info(cuID)


    def link_to_customer_subscription(self, csID):
        """
            Link invoice to customer subscription
        """
        db = current.globalenv['db']
        db.invoices_customers_subscriptions.insert(
            invoices_id = self.invoices_id,
            customers_subscriptions_id = csID
        )


    def get_linked_customer_id(self):
        """
            Returns auth.user.id of account linked to this invoice
            :return: auth.user.id
        """
        db = current.globalenv['db']

        query = (db.invoices_customers.invoices_id == self.invoices_id)
        rows = db(query).select(db.invoices_customers.auth_customer_id)

        if rows:
            return rows.first().auth_customer_id
        else:
            return None


    def get_linked_customer_subscription_id(self):
        """
            Returns auth.user.id of account linked to this invoice
            :return: auth.user.id
        """
        db = current.globalenv['db']

        query = (db.invoices_customers_subscriptions.invoices_id == self.invoices_id)
        rows = db(query).select(db.invoices_customers_subscriptions.customers_subscriptions_id)

        if rows:
            return rows.first().customers_subscriptions_id
        else:
            return None


class InvoicesHelper:
    """
        Contains functions for invoices usefull in multiple controllers
    """
    def _add_get_form_permissions_check(self):
        """
            Check if the currently logged in user is allowed to create
            invoices
        """
        auth = current.globalenv['auth']
        if not (auth.has_membership(group_id='Admins') or
                auth.has_permission('create', 'invoices')):
            redirect(URL('default', 'user', args=['not_authorized']))


    def _add_get_form_set_default_values_customer(self, customer):
        """
        :param customer: Customer object
        :return: None
        """
        db = current.globalenv['db']
        address = ''
        if customer.row.address:
            address = ''.join([address, customer.row.address, '\n'])
        if customer.row.city:
            address = ''.join([address, customer.row.city, ' '])
        if customer.row.postcode:
            address = ''.join([address, customer.row.postcode, '\n'])
        if customer.row.country:
            address = ''.join([address, customer.row.country])

        db.invoices.CustomerCompany.default = customer.row.company
        db.invoices.CustomerName.default = customer.row.full_name
        db.invoices.CustomerAddress.default = address


    def _add_get_form_enable_minimal_fields(self):
        """
            Only enable the bare minimum of fields
        """
        db = current.globalenv['db']

        for field in db.invoices:
            field.readable=False
            field.writable=False

        db.invoices.invoices_groups_id.readable = True
        db.invoices.invoices_groups_id.writable = True
        db.invoices.Description.readable = True
        db.invoices.Description.writable = True


    def _add_get_form_enable_subscription_fields(self, csID):
        """
            Enable fields required for subscriptions
        """
        db = current.globalenv['db']

        cs = CustomerSubscription(csID)
        db.invoices.customers_subscriptions_id.default = csID
        db.invoices.payment_methods_id.default = cs.payment_methods_id
        db.invoices.SubscriptionYear.readable = True
        db.invoices.SubscriptionYear.writable = True
        db.invoices.SubscriptionMonth.readable = True
        db.invoices.SubscriptionMonth.writable = True


    def add_get_form(self, cuID,
                           csID = None,
                           subscription_year = '',
                           subscription_month = '',
                           full_width = True):
        """
            Returns add form for an invoice
        """
        self._add_get_form_permissions_check()

        db = current.globalenv['db']
        T  = current.globalenv['T']

        customer = Customer(cuID)
        self._add_get_form_set_default_values_customer(customer)
        self._add_get_form_enable_minimal_fields()
        if csID:
            self._add_get_form_enable_subscription_fields(csID)

        form = SQLFORM(db.invoices, formstyle='bootstrap3_stacked')

        elements = form.elements('input, select, textarea')
        for element in elements:
            element['_form'] = "invoice_add"

        form_element = form.element('form')
        form['_id'] = 'invoice_add'

        if form.process().accepted:
            iID = form.vars.id
            invoice = Invoice(iID) # This sets due date and Invoice#
            invoice.link_to_customer(cuID)
            self._add_reset_list_status_filter()

            if csID:
                invoice.link_to_customer_subscription(csID)
                invoice.item_add_subscription(
                    form.vars.SubscriptionYear,
                    form.vars.SubscriptionMonth
                )

            redirect(URL('invoices', 'edit', vars={'iID':iID}))


        # So the grids display the fields normally
        for field in db.invoices:
            field.readable=True


        # crud = current.globalenv['crud']
        # # request = current.globalenv['request']
        #
        # create_onaccept = [ self._add_set_invoice_nr_and_due_date,
        #                     self._add_reset_list_status_filter]
        #
        # # set all fields as unreadable/writable
        # for field in db.invoices:
        #     field.readable=False
        #     field.writable=False
        #
        # db.invoices.invoices_groups_id.readable = True
        # db.invoices.invoices_groups_id.writable = True
        # db.invoices.Description.readable = True
        # db.invoices.Description.writable = True
        #
        # #TODO: Fill customer name, address and contact fields
        # #TODO: Link new invoice to customer
        # db.invoices.auth_customer_id.default = cuID
        # if csID:
        #     cs = CustomerSubscription(csID)
        #     db.invoices.customers_subscriptions_id.default = csID
        #     db.invoices.payment_methods_id.default = cs.payment_methods_id
        #     db.invoices.SubscriptionYear.readable = True
        #     db.invoices.SubscriptionYear.writable = True
        #     db.invoices.SubscriptionMonth.readable = True
        #     db.invoices.SubscriptionMonth.writable = True
        #     # add invoice item after form accepts
        #     create_onaccept.append(self._add_create_subscription_invoice_item)
        #
        # #TODO: Link invoice to db.invoices_customers_subscriptions
        # db.invoices.customers_subscriptions_id.default = csID
        #
        # crud.messages.submit_button = T("Save")
        # crud.messages.record_created = T("Added invoice")
        # crud.settings.create_onaccept=create_onaccept
        # crud.settings.create_next = '/invoices/edit/?iID=[id]'
        # form = crud.create(db.invoices)
        #
        # elements = form.elements('input, select, textarea')
        # for element in elements:
        #     element['_form'] = "invoice_add"
        #
        # form_element = form.element('form')
        # form['_id'] = 'invoice_add'
        #
        # # So the grids display the fields normally
        # for field in db.invoices:
        #     field.readable=True
        #     #field.writable=False
        #
        # if full_width:
        #     # make the inputs in the table full width
        #     table = form.element('table')
        #     table['_class'] = 'full-width'
        #
        return form


    def add_get_modal(self, crud_form):
        '''
            Returns add modal for new invoice
        '''
        T = current.globalenv['T']
        os_gui = current.globalenv['os_gui']
        gen_passwd = current.globalenv['generate_password']

        modal_class = gen_passwd()
        button_text = XML(SPAN(SPAN(_class='fa fa-plus'), ' ',
                          T("Add")))
        result = os_gui.get_modal(button_text=button_text,
                                  button_title=T("Add invoice"),
                                  modal_title=T("Add invoice"),
                                  modal_content=crud_form,
                                  modal_footer_content=os_gui.get_submit_button('invoice_add'),
                                  modal_class=modal_class,
                                  button_class='btn-sm')

        button = SPAN(result['button'], result['modal'])

        return dict(button=button, modal_class=modal_class)


    def _add_reset_list_status_filter(self):
        '''
            Reset session variable that holds status for filter
        '''
        session = current.globalenv['session']
        session.invoices_list_status = None


    def _list_set_status_filter(self):
        '''
            Sets session variable that holds the status for the filter
        '''
        request = current.globalenv['request']
        session = current.globalenv['session']
        # status definitions
        if 'status' in request.vars:
            session.invoices_list_status = request.vars['status']
        elif session.invoices_list_status:
            pass
        else:
            session.invoices_list_status = 'all'


    def list_get_status_filter(self):
        '''
            Returns the filter for Invoice statuses
        '''
        self._list_set_status_filter()

        invoice_statuses = current.globalenv['invoice_statuses']
        invoice_statuses.append(['overdue', current.T('Overdue')])

        session = current.globalenv['session']

        form = SQLFORM.factory(
            Field('status',
                default=session.invoices_list_status,
                requires=IS_IN_SET(invoice_statuses,
                                   zero=current.T('All statuses')))
        )

        form = form.element('form')
        form['_class'] = 'no-margin-top no-margin-bottom'

        select = form.element('#no_table_status')
        select['_onchange'] = 'this.form.submit();'

        #TODO: The pull-right class causes it to not work on mobile, should be reworked to show up
        # should be reworked nicely without the pull-right at some point.
        form = DIV(form.custom.begin,
                   form.custom.widget.status,
                   form.custom.end,
                   _class='pull-right',
                   _id='invoices_status_filter')

        return form


    def represent_invoice_for_list(self, iID,
                                         invoice_id,
                                         repr_status,
                                         status,
                                         payment_methods_id):
        '''
            Represent invoice for lists
            invoice_id and status should come from a rendered row (repr_row)
        '''
        os_gui = current.globalenv['os_gui']

        payment = self.represent_invoice_for_list_get_payment(iID, status, payment_methods_id)

        invoice = DIV(repr_status,
                      os_gui.get_button(
                            'noicon',
                            URL('invoices', 'edit', vars={'iID':iID},
                                extension=''),
                            title=invoice_id,
                            btn_size = 'btn-sm',
                            btn_class = 'btn-link'), BR(),
                      payment)

        return invoice


    def represent_invoice_for_list_get_payment(self, iID, status, payment_methods_id):
        '''
            Get add payment modal when no payment is found, or just
            show the information for the payments found.
        '''
        db = current.globalenv['db']
        os_gui = current.globalenv['os_gui']
        T = current.globalenv['T']

        query = (db.invoices_payments.invoices_id == iID)
        rows = db(query).select(db.invoices_payments.ALL)

        # show payments
        payments = DIV()
        for row in rows.render():
            payments.append(SPAN(
                row.PaymentDate, ': ',
                row.Amount,
                BR(),
                _class = 'grey small_font'
            ))

        if status == 'sent':
            if payment_methods_id == 3: # direct debit
                payments.append(SPAN(T('Direct debit'), _class='grey small_font'))
            else:
                # show add payment
                content = LOAD('invoices', 'payment_add.load', ajax=True,
                                vars={'iID':iID})

                invoice = db.invoices(iID)
                title = current.T('Add payment for invoice') + ' #' + \
                        invoice.InvoiceID

                button_text = XML(SPAN(
                    SPAN(_class='fa fa-credit-card'), ' ',
                    current.T('Add payment'),
                    _class='small_font grey'
                ))

                form_id = 'form_payment_add_' + unicode(iID)

                result = os_gui.get_modal(button_text=button_text,
                                          button_title=current.T("Add payment"),
                                          modal_title=title,
                                          modal_content=content,
                                          modal_footer_content=os_gui.get_submit_button(form_id),
                                          modal_class='form_payment_add_' + unicode(iID),
                                          button_class='btn-xs invoice_list_add_payment')

                payments.append(SPAN(result['button'], result['modal']))

        return payments


    # def payment_add_get_form(self, iID):
    #     '''
    #         Add payments for an invoice
    #     '''
    #     db = current.globalenv['db']
    #     T = current.globalenv['T']
    #     crud = current.globalenv['crud']
    #
    #     invoice = db.invoices(iID)
    #
    #     ## default values
    #     db.invoices_payments.invoices_id.default = iID
    #     # amount
    #     amounts = db.invoices_amounts(invoices_id=iID)
    #     try:
    #         db.invoices_payments.Amount.default = amounts.TotalPriceVAT
    #     except AttributeError:
    #         pass
    #     # payment method
    #     try:
    #         payment_info = db.customers_payment_info(
    #             auth_customer_id=invoice.auth_customer_id)
    #         default_method = payment_info.payment_methods_id
    #         db.invoices_payments.payment_methods_id.default = default_method
    #     except AttributeError:
    #         pass
    #
    #     # # if session.invoices_payment_add_back == 'invoices_invoice_payments':
    #     # #     # Don't redirect client side here, stay in the modal on the invoice edit page
    #     # #     create_next = URL('invoices', 'invoice_payments', vars={'iID':iID})
    #     # # else:
    #     # create_next = '/invoices/payment_add_redirect_oncreate?ipID=[id]'
    #
    #     crud.messages.submit_button = T("Save")
    #     crud.messages.record_created = T("Saved")
    #     #crud.settings.create_next = create_next
    #     #crud.settings.create_onaccept = [payment_add_update_status]
    #     form = crud.create(db.invoices_payments)
    #
    #     form['_action'] = URL('invoices', 'payment_add', vars={'iID':iID})
    #     form['_name'] = "fpa_" + unicode(iID)
    #
    #     form_id = 'form_payment_add_' + unicode(iID)
    #     form_element = form.element('form')
    #     form['_id'] = form_id
    #
    #     elements = form.elements('input, select, textarea')
    #     for element in elements:
    #         element['_form'] = form_id
    #
    #     return form


    def list_invoices(self, cuID=None, csID=None, search_enabled=False, group_filter_enabled=False):
        db = current.globalenv['db']
        auth = current.globalenv['auth']
        session = current.globalenv['session']
        grid_ui = current.globalenv['grid_ui']
        DATE_FORMAT = current.globalenv['DATE_FORMAT']
        from general_helpers import datestr_to_python
        T = current.globalenv['T']

        session.invoices_invoice_payment_add_back = None

        # disable unused fields
        db.invoices.id.readable = False
        db.invoices.invoices_groups_id.readable = False
        db.invoices.Footer.readable = False
        db.invoices.Note.readable = False
        db.invoices.Terms.readable = False

        links = [dict(header=T("Balance"),
                      body=self._list_invoices_get_balance),
                 self._list_invoices_get_buttons]
        left = [db.invoices_amounts.on(db.invoices_amounts.invoices_id ==
                                       db.invoices.id),
                db.invoices_customers.on(
                    db.invoices_customers.invoices_id == db.invoices.id),
                db.invoices_customers_subscriptions.on(
                    db.invoices_customers_subscriptions.invoices_id ==
                    db.invoices.id)
                ]

        fields = [db.invoices.Status,
                  db.invoices.InvoiceID,
                  db.invoices.Description,
                  db.invoices.DateCreated,
                  db.invoices.DateDue,
                  db.invoices_amounts.TotalPriceVAT]

        query = (db.invoices.id > 0)
        # Status filter
        query = self._list_invoices_get_status_query(query)
        if search_enabled:
            query = self._list_invoices_get_search_query(query)
        if group_filter_enabled:
            query = self._list_invoices_get_groups_query(query)

        # General list, list for customer or list for subscription
        if not cuID and not csID:
            # list all invoices
            db.invoices.auth_customer_id.readable = True
            fields.insert(2, db.invoices.CustomerListName)

        if cuID:
            query &= (db.invoices_customers.auth_customer_id == cuID)
        if csID:
            query &= (db.invoices_customers_subscriptions.customers_subscriptions_id == csID)
            fields.insert(3, db.invoices.SubscriptionMonth)
            fields.insert(4, db.invoices.SubscriptionYear)

        delete_permission = auth.has_membership(group_id='Admins') or \
                            auth.has_permission('delete', 'invoices')

        headers = {'invoices.auth_customer_id': T("Customer")}

        grid = SQLFORM.grid(query,
                            links=links,
                            left=left,
                            field_id=db.invoices.id,
                            fields=fields,
                            headers=headers,
                            create=False,
                            editable=False,
                            details=False,
                            searchable=False,
                            deletable=delete_permission,
                            csv=False,
                            # maxtextlengths=maxtextlengths,
                            orderby=~db.invoices.id,
                            ui=grid_ui)
        grid.element('.web2py_counter', replace=None)  # remove the counter
        grid.elements('span[title=Delete]', replace=None)  # remove text from delete button

        form_search = ''
        content = DIV()
        if search_enabled:
            #response.js = 'set_form_classes();' # we're no longer in a loaded component
            request = current.globalenv['request']
            if 'search' in request.vars:
                session.invoices_list_invoices_search = request.vars['search']
                # date_created_from = datestr_to_python(DATE_FORMAT, request.vars['date_created_from'])
                # session.invoices_list_invoices_date_created_from = date_created_from
                try:
                    date_created_from = datestr_to_python(DATE_FORMAT, request.vars['date_created_from'])
                    session.invoices_list_invoices_date_created_from = date_created_from
                except (ValueError, AttributeError):
                    session.invoices_list_invoices_date_created_from = None
                try:
                    date_created_until = datestr_to_python(DATE_FORMAT, request.vars['date_created_until'])
                    session.invoices_list_invoices_date_created_until = date_created_until
                except (ValueError, AttributeError):
                    session.invoices_list_invoices_date_created_until = None
                try:
                    date_due_from = datestr_to_python(DATE_FORMAT, request.vars['date_due_from'])
                    session.invoices_list_invoices_date_due_from = date_due_from
                except (ValueError, AttributeError):
                    session.invoices_list_invoices_date_due_from = None
                try:
                    date_due_until = datestr_to_python(DATE_FORMAT, request.vars['date_due_until'])
                    session.invoices_list_invoices_date_due_until = date_due_until
                except (ValueError, AttributeError):
                    session.invoices_list_invoices_date_due_until = None

                keys = ['search', 'date_created_from', 'date_created_until', 'date_due_from', 'date_due_until']
                for key in keys:
                    try:
                        del request.vars[key]
                    except KeyError:
                        pass

                # redirect to update page
                redirect(URL(vars=request.vars))

            form_search = self._list_invoices_get_form_search()
            content.append(form_search)

        form_groups = ''
        if group_filter_enabled:
            if 'invoices_groups_id' in request.vars:
                session.invoices_list_invoices_group = request.vars['invoices_groups_id']

                try:
                    del request.vars['invoices_groups_id']
                except KeyError:
                    pass

                # redirect to update page
                redirect(URL(vars=request.vars))

        # always add the grid
        content.append(grid)
        return content


    def _list_invoices_get_form_search(self):
        '''
            Returns search form for invoices page
        '''
        T = current.globalenv['T']
        session = current.globalenv['session']
        DATE_FORMAT = current.globalenv['DATE_FORMAT']

        form = SQLFORM.factory(
            Field('search',
                  default=session.invoices_list_invoices_search,
                  label=T('')),
            Field('date_created_from', 'date',
                  requires=IS_EMPTY_OR(IS_DATE_IN_RANGE(format=DATE_FORMAT,
                                                        minimum=datetime.date(1900, 1, 1),
                                                        maximum=datetime.date(2999, 1, 1))),
                  # ),
                  default=session.invoices_list_invoices_date_created_from),
            Field('date_created_until', 'date',
                  requires=IS_EMPTY_OR(IS_DATE_IN_RANGE(format=DATE_FORMAT,
                                                        minimum=datetime.date(1900, 1, 1),
                                                        maximum=datetime.date(2999, 1, 1))),
                  # ),
                  default=session.invoices_list_invoices_date_created_until),
            Field('date_due_from', 'date',
                  requires=IS_EMPTY_OR(IS_DATE_IN_RANGE(format=DATE_FORMAT,
                                                        minimum=datetime.date(1900, 1, 1),
                                                        maximum=datetime.date(2999, 1, 1))),
                  # ),
                  default=session.invoices_list_invoices_date_due_from),
            Field('date_due_until', 'date',
                  requires=IS_EMPTY_OR(IS_DATE_IN_RANGE(format=DATE_FORMAT,
                                                        minimum=datetime.date(1900, 1, 1),
                                                        maximum=datetime.date(2999, 1, 1))),
                  # ),
                  default=session.invoices_list_invoices_date_due_until),
            submit_button=T('Go')
        )

        search = form.element('#no_table_search')
        search['_class'] += ' margin-right'
        search['_placeholder'] = T("Invoice #")

        btn_clear = A(T("Clear"),
                      _href=URL('invoices', 'list_invoices_clear_search',
                                vars={'search_enabled': True}),
                      _title=T("Clear search"),
                      _class='btn btn-default')

        form = DIV(
            DIV(
                form.custom.begin,
                DIV(LABEL(T('Search')), BR(), form.custom.widget.search, _class='col-md-2'),
                DIV(DIV(DIV(LABEL(T('Date from')), BR(), form.custom.widget.date_created_from, _class='col-md-2'),
                        DIV(LABEL(T('Date until')), BR(), form.custom.widget.date_created_until, _class='col-md-2'),
                        DIV(LABEL(T('Due from')), BR(), form.custom.widget.date_due_from, _class='col-md-2'),
                        DIV(LABEL(T('Due until')), BR(), form.custom.widget.date_due_until, _class='col-md-2'),
                        DIV(LABEL(T('Filter')), BR(), form.custom.submit, btn_clear, _class='col-md-3'),
                        _class='row'),
                    _class='col-md-8'),
                form.custom.end,
                _class='row'),
        )

        return form

        ##################### test code begin


        # @auth.requires(auth.has_membership(group_id='Admins') or \
        #                auth.has_permission('read', 'invoices'))
        # def _list_invoices_clear_search(self):
        #     '''
        #         Clears search for invoices page
        #     '''
        #     session.invoices_list_invoices_search = None
        #     session.invoices_list_invoices_date_created_from = None
        #     session.invoices_list_invoices_date_created_until = None
        #     session.invoices_list_invoices_date_due_from = None
        #     session.invoices_list_invoices_date_due_until = None
        #
        #     redirect(URL('list_invoices', vars=request.vars))

    def _list_invoices_get_status_query(self, query):
        '''
            Returns status query
        '''
        db = current.globalenv['db']
        session = current.globalenv['session']

        if session.invoices_list_status == 'draft':
            query &= (db.invoices.Status == 'draft')
        if session.invoices_list_status == 'sent':
            query &= (db.invoices.Status == 'sent')
        if session.invoices_list_status == 'paid':
            query &= (db.invoices.Status == 'paid')
        if session.invoices_list_status == 'cancelled':
            query &= (db.invoices.Status == 'cancelled')
        if session.invoices_list_status == 'overdue':
            query &= (db.invoices.Status == 'sent')
            query &= (db.invoices.DateDue < datetime.date.today())

        return query

    def _list_invoices_get_search_query(self, query):
        '''
            Adds search for invoice number to query
        '''
        db = current.globalenv['db']
        session = current.globalenv['session']

        if session.invoices_list_invoices_search:
            search = session.invoices_list_invoices_search.strip()
            query &= (db.invoices.InvoiceID.like('%' + search + '%'))

        if session.invoices_list_invoices_date_created_from:
            query &= (db.invoices.DateCreated >= session.invoices_list_invoices_date_created_from)

        if session.invoices_list_invoices_date_created_until:
            query &= (db.invoices.DateCreated <= session.invoices_list_invoices_date_created_until)

        if session.invoices_list_invoices_date_due_from:
            query &= (db.invoices.DateDue >= session.invoices_list_invoices_date_due_from)

        if session.invoices_list_invoices_date_due_until:
            query &= (db.invoices.DateDue <= session.invoices_list_invoices_date_due_until)

        return query

    def _list_invoices_get_groups_query(self, query):
        '''
            Adds filter for invoice group to query
        '''
        if session.invoices_list_invoices_group:
            query &= (db.invoices.invoices_groups_id == session.invoices_list_invoices_group)

    def _list_invoices_get_buttons(self, row):
        '''
            Group all links for invoices into .btn-group
        '''
        auth = current.globalenv['auth']
        os_gui = current.globalenv['os_gui']
        T = current.globalenv['T']

        iID = row.invoices.id
        modals = SPAN()
        links = DIV(_class='btn-group')
        buttons = SPAN(links, modals)

        if auth.has_membership(group_id='Admins') or \
                auth.has_permission('create', 'invoices_payments'):
            links.append(self._list_invoices_get_link_add_payment(iID))

            #result = self._list_invoices_get_link_add_payment(iID)
            #links.append(result['button'])
            #modals.append(result['modal'])

        if auth.has_membership(group_id='Admins') or \
                auth.has_permission('read', 'invoices'):
            pdf = os_gui.get_button('print',
                                    URL('invoices', 'pdf', vars={'iID': iID}))
            links.append(pdf)

        # if auth.has_membership(group_id='Admins') or \
        #    auth.has_permission('read', 'auth_user'):
        #     customer = os_gui.get_button('user',
        #         URL('customers', 'edit', args=row.invoices.auth_customer_id,
        #             extension=''))
        #     links.append(customer)

        if auth.has_membership(group_id='Admins') or \
                auth.has_permission('update', 'invoices'):
            edit = os_gui.get_button('edit',
                                     URL('invoices', 'edit', vars={'iID': iID}, extension=''),
                                     tooltip=T('Edit Invoice'))
            links.append(edit)

        return buttons

    def _list_invoices_get_balance(self, row):
        '''
            Retuns the balance for an invoice
        '''
        iID = row.invoices.id
        invoice = Invoice(iID)

        return invoice.get_balance(formatted=True)


    def _list_invoices_get_link_add_payment(self, iID):
        '''
            Returns an button and modal to add a payment for an invoice
        '''
        T = current.globalenv['T']
        os_gui = current.globalenv['os_gui']

        button = os_gui.get_button('credit-card',
                                   URL('invoices', 'payment_add', vars={'iID':iID}),
                                   tooltip=T('Add payment'))

        return button

        #
        # db = current.globalenv['db']
        # T = current.globalenv['T']
        # os_gui = current.globalenv['os_gui']
        #
        # content = LOAD('invoices', 'payment_add', ajax=False, ajax_trap=True, extension='load',
        #                 vars={'iID': iID})
        # #content = self.payment_add_get_form(iID)
        #
        # invoice = db.invoices(iID)
        # title = T('Add payment for invoice') + ' #' + invoice.InvoiceID
        #
        # button_text = os_gui.get_modal_button_icon('credit-card')
        #
        # form_id = 'form_payment_add_' + unicode(iID)
        #
        # result = os_gui.get_modal(button_text=button_text,
        #                           button_title=T("Add payment"),
        #                           modal_title=title,
        #                           modal_content=content,
        #                           modal_footer_content=os_gui.get_submit_button(form_id),
        #                           modal_class=form_id,
        #                           # modal_id='modal_payment_add_' + unicode(iID),
        #                           button_class='btn-sm')
        #
        #
        #
        # return result

        #payments.append(SPAN(result['button'], result['modal']))


            #################### test code end



            # def get_add_modal(self, cuID,
    #                         pbiID        = None,
    #                         csID         = None,
    #                         button_text  = current.T('Add'),
    #                         button_class = 'btn-sm'
    #                         ):
    #     '''
    #         Return button and modal to add a new invoice
    #     '''
    #     os_gui = current.globalenv['os_gui']
    #     gen_passwd = current.globalenv['generate_password']
    #
    #     vars = {'cuID':cuID}
    #     if pbiID:
    #         vars['pbiID'] = pbiID
    #     if csID:
    #         vars['csID'] = csID
    #
    #     add = LOAD('invoices', 'add.load', ajax=True, vars=vars)
    #
    #     button_text = XML(SPAN(SPAN(_class='glyphicon glyphicon-plus'), ' ',
    #                       button_text))
    #
    #     # get 30 chars of randomness for modal class
    #     modal_class = gen_passwd()
    #
    #     result = os_gui.get_modal(button_text=button_text,
    #                               modal_title=current.T('Add invoice'),
    #                               modal_content=add,
    #                               modal_class=modal_class,
    #                               button_class=button_class)
    #     modal = result['modal']
    #     button = result['button']
    #
    #     return DIV(button, modal)


class School:
    def get_classcards(self, public_only=True):
        '''
            :param public_only: Defines whether or not to show only public classcards, True by default
                                False means all cards are returned
            Returns classcards for school
        '''
        db = current.globalenv['db']

        query = (db.school_classcards.Archived == False)
        if public_only:
            query &= (db.school_classcards.PublicCard == True)

        return db(query).select(db.school_classcards.ALL,
                                orderby=db.school_classcards.Trialcard|
                                        db.school_classcards.Name)


    def get_classcards_formatted(self, public_only=True, per_row=3, link_type=None):
        '''
            :param public_only: show only public cards - Default: True
            :param per_row: Number of cards in each row - Default 4. Allowed values: [3, 4]
            :param link_type: Specified what kind of link will be shown in the footer of each classcard.
                Allowed values: ['backend', 'shop']
                - backend adds a modal to choose date
                - shop adds a button to add the card to the shopping cart
            Returns classcards formatted in BS3 style

        '''
        def get_validity(row):
            '''
                takes a db.school_classcards() row as argument
            '''
            validity = SPAN(unicode(row.Validity), ' ')

            validity_in = represent_validity_units(row.ValidityUnit, row)
            if row.Validity == 1:  # Cut the last 's"
                validity_in = validity_in[:-1]

            validity.append(validity_in)

            return validity


        os_gui = current.globalenv['os_gui']
        T = current.globalenv['T']


        if per_row == 3:
            card_class = 'col-md-4'
        elif per_row == 4:
            card_class = 'col-md-3'
        else:
            raise ValueError('Incompatible value: per_row has to be 3 or 4')

        rows = self.get_classcards(public_only=public_only)

        cards = DIV()
        display_row = DIV(_class='row')
        row_item = 0

        for i, row in enumerate(rows):
            repr_row = list(rows[i:i + 1].render())[0]

            card_name = max_string_length(row.Name, 37)
            validity  = get_validity(row)

            card_content = TABLE(TR(TD(T('Validity')),
                                    TD(validity)),
                                 TR(TD(T('Classes')),
                                    TD(repr_row.Classes)),
                                 TR(TD(T('Price')),
                                    TD(repr_row.Price)),
                                 TR(TD(T('Description')),
                                    TD(repr_row.Description or '')),
                                 _class='table')


            if row.Trialcard:
                panel_class = 'box-success'
            else:
                panel_class = 'box-primary'


            footer_content = ''
            if link_type == 'shop':
                footer_content = self._get_classcards_formatted_button_to_cart(row.id)


            card = DIV(os_gui.get_box_table(card_name,
                                            card_content,
                                            panel_class,
                                            show_footer=True,
                                            footer_content=footer_content),
                       _class=card_class)

            display_row.append(card)

            row_item += 1

            if row_item == per_row or i == (len(rows) - 1):
                cards.append(display_row)
                display_row = DIV(_class='row')

                row_item = 0

        return cards


    def _get_classcards_formatted_button_to_cart(self, scdID):
        '''
            Get button to add card to shopping cart
        '''
        os_gui = current.globalenv['os_gui']
        T = current.globalenv['T']

        return A(SPAN(os_gui.get_fa_icon('fa-shopping-cart'), ' ', T('Add to cart')),
                 _href=URL('classcard_add_to_cart', vars={'scdID':scdID}))


    def _get_subscriptions_formatted_button_to_cart(self, ssuID):
        '''
            Get button to add card to shopping cart
        '''
        os_gui = current.globalenv['os_gui']
        T = current.globalenv['T']

        return A(SPAN(os_gui.get_fa_icon('fa-shopping-cart'), ' ', T('Get this subscription')),
                 _href=URL('subscription_terms', vars={'ssuID':ssuID}))


    def get_subscriptions(self, public_only=True):
        '''
            :param public: boolean, defines whether to show only public or all subscriptions
            :return: list of school_subscriptions
        '''
        db = current.globalenv['db']

        query = (db.school_subscriptions.id > 0)

        if public_only:
            query &= (db.school_subscriptions.PublicSubscription == True)

        rows = db(query).select(db.school_subscriptions.ALL,
                                orderby=~db.school_subscriptions.SortOrder|db.school_subscriptions.Name)

        return rows


    def get_subscriptions_formatted(self, per_row=3, public_only=True, link_type='shop'):
        '''
            :param public: boolean, defines whether to show only public or all subscriptions
            :return: list of school_subscriptions formatted for shop
        '''
        os_gui = current.globalenv['os_gui']
        T = current.globalenv['T']

        if per_row == 3:
            card_class = 'col-md-4'
        elif per_row == 4:
            card_class = 'col-md-3'
        else:
            raise ValueError('Incompatible value: per_row has to be 3 or 4')

        rows = self.get_subscriptions(public_only=public_only)

        subscriptions = DIV()
        display_row = DIV(_class='row')
        row_item = 0

        for i, row in enumerate(rows):
            repr_row = list(rows[i:i + 1].render())[0]

            ssu = SchoolSubscription(row.id)
            name = max_string_length(row.Name, 33)

            classes = ''
            if row.Unlimited:
                classes = T('Unlimited')
            elif row.SubscriptionUnit == 'week':
                classes = SPAN(unicode(row.Classes) + ' / ' + T('Week'))
            elif row.SubscriptionUnit == 'month':
                classes = SPAN(unicode(row.Classes) + ' / ' + T('Month'))

            subscription_content = TABLE(TR(TD(T('Classes')),
                                            TD(classes)),
                                         TR(TD(T('Monthly')),
                                            TD(ssu.get_price_on_date(datetime.date.today()))),
                                         TR(TD(T('Description')),
                                            TD(row.Description or '')),
                                         _class='table')

            panel_class = 'box-primary'


            footer_content = ''
            if link_type == 'shop':
                footer_content = self._get_subscriptions_formatted_button_to_cart(row.id)


            subscription = DIV(os_gui.get_box_table(name,
                                                    subscription_content,
                                                    panel_class,
                                                    show_footer=True,
                                                    footer_content=footer_content),
                               _class=card_class)

            display_row.append(subscription)

            row_item += 1

            if row_item == per_row or i == (len(rows) - 1):
                subscriptions.append(display_row)
                display_row = DIV(_class='row')
                row_item = 0

        return subscriptions


class SchoolClasscard:
    '''
        Class that contains functions for a class card
    '''
    def __init__(self, scdID):
        '''
            Class init function which sets ssuID
        '''
        self.scdID = scdID


    def get_validity_formatted(self):
        '''
            :return: Validity for school classcard
        '''
        T  = current.globalenv['T']
        db = current.globalenv['db']

        row = db.school_classcards(self.scdID)
        validity = SPAN(unicode(row.Validity), ' ')

        validity_in = represent_validity_units(row.ValidityUnit, row)
        if row.Validity == 1: # Cut the last 's"
            validity_in = validity_in[:-1]

        validity.append(validity_in)

        return validity


    def add_to_shoppingcart(self, auth_user_id):
        '''
            :param auth_user_id: db.auth_user.id
        '''
        db = current.globalenv['db']

        db.customers_shoppingcart.insert(
            auth_customer_id     = auth_user_id,
            school_classcards_id = self.scdID
        )


    def sell_to_customer(self, auth_user_id, date_start, note=None, invoice=True):
        '''
            :param auth_user_id: Sell classcard to customer
        '''
        db = current.globalenv['db']
        cache_clear_customers_classcards = current.globalenv['cache_clear_customers_classcards']

        ccdID = db.customers_classcards.insert(
            auth_customer_id = auth_user_id,
            school_classcards_id = self.scdID,
            Startdate = date_start,
            Enddate = self.sell_to_customer_get_enddate(date_start),
            Note = note
        )

        cache_clear_customers_classcards(auth_user_id)

        if invoice:
            self.sell_to_customer_create_invoice(ccdID)

        return ccdID


    def sell_to_customer_create_invoice(self, ccdID):
        '''
            Add an invoice after adding a classcard
        '''
        db = current.globalenv['db']
        T = current.globalenv['T']

        classcard = Classcard(ccdID)

        igpt = db.invoices_groups_product_types(ProductType='classcard')

        iID = db.invoices.insert(
            invoices_groups_id=igpt.invoices_groups_id,
            Description=classcard.get_name(),
            Status='sent'
        )

        # link invoice to classcard
        db.invoices_customers_classcards.insert(
            invoices_id=iID,
            customers_classcards_id=ccdID)

        # create object to set Invoice# and due date
        invoice = Invoice(iID)
        next_sort_nr = invoice.get_item_next_sort_nr()

        price = classcard.price

        iiID = db.invoices_items.insert(
            invoices_id=iID,
            ProductName=T("Class card"),
            Description=T("Class card") + ' ' + unicode(ccdID),
            Quantity=1,
            Price=price,
            Sorting=next_sort_nr,
            tax_rates_id=classcard.school_classcard.tax_rates_id,
        )

        invoice.set_amounts()
        invoice.link_to_customer(classcard.get_auth_customer_id())


    def sell_to_customer_get_enddate(self, date_start):
        '''
           Calculate and set enddate when adding a classcard
           :param ccdID: db.customers_classcards.id
           :return : enddate for a classcard
        '''
        def add_months(sourcedate, months):
            month = sourcedate.month - 1 + months
            year = int(sourcedate.year + month / 12)
            month = month % 12 + 1
            last_day_new = calendar.monthrange(year, month)[1]
            day = min(sourcedate.day, last_day_new)

            ret_val = datetime.date(year, month, day)

            last_day_source = calendar.monthrange(sourcedate.year,
                                                  sourcedate.month)[1]

            if sourcedate.day == last_day_source and last_day_source > last_day_new:
                return ret_val
            else:
                delta = datetime.timedelta(days=1)
                return ret_val - delta

        db = current.globalenv['db']

        # get info
        card = db.school_classcards(self.scdID)

        if card.ValidityUnit == 'months':
            # check for and add months
            months = card.Validity
            if months:
                enddate = add_months(date_start, months)
        else:
            if card.ValidityUnit == 'weeks':
                days = card.Validity * 7
            else:
                days = card.Validity

            delta_days = datetime.timedelta(days=days)
            enddate = (date_start + delta_days) - datetime.timedelta(days=1)

        return enddate


class StaffSchedule:
    def __init__(self, date,
                       filter_id_school_shifts = None,
                       filter_id_school_locations = None,
                       filter_id_employee = None,
                       filter_id_status = None,
                       sorting = 'starttime'):
        self.date = date

        self.filter_id_school_shifts = filter_id_school_shifts
        self.filter_id_school_locations = filter_id_school_locations
        self.filter_id_employee = filter_id_employee
        self.filter_id_status = filter_id_status
        self.sorting = sorting

    def _get_day_filter_query(self):
        '''
            Returns the filter query for the schedule
        '''
        where = ''
        if self.filter_id_school_shifts:
            where += 'AND sh.school_shifts_id = '
            where += unicode(self.filter_id_school_shifts) + ' '

        if self.filter_id_school_locations:
            where += 'AND sh.school_locations_id = '
            where += unicode(self.filter_id_school_locations) + ' '

        return where


    # def _get_day_row_status(self, row):
    #     '''
    #         Return status for row
    #     '''
    #     status = 'normal'
    #     status_marker = DIV(_class='status_marker bg_green')
    #     if row.shifts_cancelled.id:
    #         status = 'cancelled'
    #         status_marker = DIV(_class='status_marker bg_orange')
    #     elif row.shifts_open.id:
    #         status = 'open'
    #         status_marker = DIV(_class='status_marker bg_red')
    #     elif row.shifts_sub.id:
    #         status = 'sub'
    #         status_marker = DIV(_class='status_marker bg_blue')
    #
    #     return dict(status=status, marker=status_marker)


    def _get_day_row_status(self, row):
        '''
            Return status for row
        '''
        status = 'normal'
        status_marker = DIV(_class='status_marker bg_green')
        if row.shifts_otc.Status == 'cancelled':
            status = 'cancelled'
            status_marker = DIV(_class='status_marker bg_orange')
        elif row.shifts_otc.Status == 'open':
            status = 'open'
            status_marker = DIV(_class='status_marker bg_red')
        elif not row.shifts_otc.auth_employee_id == None:
            status = 'sub'
            status_marker = DIV(_class='status_marker bg_blue')

        return dict(status=status, marker=status_marker)


    def _get_day_row_staff(self, row, repr_row, status):
        '''
            Return employees for a row
        '''
        if status == 'sub':
            employee_id  = row.shifts_otc.auth_employee_id
            employee_id2 = row.shifts_otc.auth_employee_id2
            employee     = repr_row.shifts_otc.auth_employee_id
            employee2    = repr_row.shifts_otc.auth_employee_id2
        else:
            employee_id  = row.shifts_staff.auth_employee_id
            employee_id2 = row.shifts_staff.auth_employee_id2
            employee     = repr_row.shifts_staff.auth_employee_id
            employee2    = repr_row.shifts_staff.auth_employee_id2

        return dict(employee_id  = employee_id,
                    employee_id2 = employee_id2,
                    employee     = employee,
                    employee2    = employee2)


    def _get_day_table_buttons(self, shID):
        '''
            Returns edit & delete buttons for schedule
        '''
        auth = current.globalenv['auth']
        os_gui = current.globalenv['os_gui']
        T = current.globalenv['T']
        DATE_FORMAT = current.globalenv['DATE_FORMAT']
        date_formatted = self.date.strftime(DATE_FORMAT)

        buttons = DIV(_class='pull-right')

        vars = { 'shID' : shID, 'date' : date_formatted }


        if auth.has_membership(group_id='Admins') or \
           auth.has_permission('delete', 'shifts'):

            onclick_remove = "return confirm('" + \
                             T('Do you really want to delete this shift?') + \
                             "');"
            remove = os_gui.get_button('delete_notext',
                       URL('shift_delete', vars=vars),
                       onclick=onclick_remove,
                       _class='pull-right')

            buttons.append(remove)

        edit_menu = ''
        links = [['header', T('Shift on') + ' ' + date_formatted]]

        # check permissions to change this class
        if auth.has_membership(group_id='Admins') or \
                auth.has_permission('create', 'shifts_otc'):
            links.append(A(os_gui.get_fa_icon('fa-pencil'),
                           T('Edit'),
                           _href=URL('shift_edit_on_date', vars=vars)))
        # Check permission to update weekly class
        if auth.has_membership(group_id='Admins') or \
                auth.has_permission('update', 'shifts'):
            links.append('divider')
            links.append(['header', T('All shifts in series')])

            # Add edit link
            links.append(A(os_gui.get_fa_icon('fa-pencil'),
                           T('Edit'), ' ',
                           _href=URL('shift_edit', vars=vars)))

            # Add staff link
            if auth.has_membership(group_id='Admind') or \
               auth.has_permission('read', 'shifts_staff'):
                links.append(A(os_gui.get_fa_icon('fa-user'),
                               T('Staff'), ' ',
                               _href=URL('shift_staff', vars=vars)))


            # edit_links = [
            #     [ A(os_gui.get_fa_icon('fa-calendar-o'), T('This shift'),
            #              SPAN('(' + date_formatted + ')',
            #                   _class='vsmall_font grey'),
            #         _href=URL('shift_edit_on_date', vars=vars)),
            #       A(os_gui.get_fa_icon('fa-calendar'), T('All shifts'),
            #         _href=URL('shift_edit', vars=vars)) ]]

            #edit = os_gui.get_button('edit', URL('class_edit', vars=vars))
            edit_menu = os_gui.get_dropdown_menu(
                links,
                btn_text = '',
                btn_size = 'btn-sm',
                btn_icon = 'pencil',
                menu_class = 'btn-group pull-right'
            )

            buttons.append(edit_menu)

        return buttons


    def get_day_rows(self):
        '''
            Helper function that returns a dict containing a title for the weekday,
            a date for the class and
            a SQLFORM.grid for a selected day which is within 1 - 7 (ISO standard).
        '''
        date = self.date
        DATE_FORMAT = current.globalenv['DATE_FORMAT']
        db = current.globalenv['db']
        weekday = date.isoweekday()

        date_formatted = date.strftime(DATE_FORMAT)

        if self.sorting == 'location':
            orderby_sql = 'sl.Name, sh.Starttime'
        elif self.sorting == 'starttime':
            orderby_sql = 'sh.Starttime, sl.Name'

        fields = [
            db.shifts.id,
            db.shifts_otc.Status,
            db.shifts_otc.Description,
            db.shifts.school_locations_id,
            db.shifts.school_shifts_id,
            db.shifts.Week_day,
            db.shifts.Starttime,
            db.shifts.Endtime,
            db.shifts.Startdate,
            db.shifts.Enddate,
            db.shifts_otc.id,
            db.shifts_otc.auth_employee_id,
            db.shifts_otc.auth_employee_id2,
            db.shifts_staff.id,
            db.shifts_staff.auth_employee_id,
            db.shifts_staff.auth_employee_id2,
        ]

        where_filter = self._get_day_filter_query()

        query = '''
        SELECT sh.id,
               CASE WHEN sotc.Status IS NOT NULL
                    THEN sotc.Status
                    ELSE 'normal'
                    END AS Status,
               sotc.Description,
               CASE WHEN sotc.school_locations_id IS NOT NULL
                    THEN sotc.school_locations_id
                    ELSE sh.school_locations_id
                    END AS school_locations_id,
               CASE WHEN sotc.school_shifts_id IS NOT NULL
                    THEN sotc.school_shifts_id
                    ELSE sh.school_shifts_id
                    END AS school_shifts_id,
               sh.Week_day,
               CASE WHEN sotc.Starttime IS NOT NULL
                    THEN sotc.Starttime
                    ELSE sh.Starttime
                    END AS Starttime,
               CASE WHEN sotc.Endtime IS NOT NULL
                    THEN sotc.Endtime
                    ELSE sh.Endtime
                    END AS Endtime,
               sh.Startdate,
               sh.Enddate,
               sotc.id,
               sotc.auth_employee_id,
               sotc.auth_employee_id2,
               shs.id,
               shs.auth_employee_id,
               shs.auth_employee_id2
        FROM shifts sh
        LEFT JOIN school_locations sl
            ON sl.id = sh.school_locations_id
        LEFT JOIN
            ( SELECT id,
                     shifts_id,
                     ShiftDate,
                     Status,
                     Description,
                     school_locations_id,
                     school_shifts_id,
                     Starttime,
                     Endtime,
                     auth_employee_id,
                     auth_employee_id2
              FROM shifts_otc
              WHERE ShiftDate = '{shift_date}' ) sotc
            ON sh.id = sotc.shifts_id
        LEFT JOIN
            ( SELECT id,
                     shifts_id,
                     auth_employee_id,
                     auth_employee_id2
              FROM shifts_staff
              WHERE Startdate <= '{shift_date}' AND (
                    Enddate >= '{shift_date}' OR Enddate IS NULL)
              ) shs
            ON sh.id = shs.shifts_id
        WHERE sh.Week_day = '{week_day}' AND
              sh.Startdate <= '{shift_date}' AND
              (sh.Enddate >= '{shift_date}' OR sh.Enddate IS NULL)
              {where_filter}
        ORDER BY {orderby_sql}
        '''.format(shift_date = date,
                   week_day = weekday,
                   orderby_sql = orderby_sql,
                   where_filter = where_filter)

        rows = db.executesql(query, fields=fields)

        return rows


    def get_day_table(self):
        '''
            Calls the schedule_get_day_rows function and formats the rows
            in a desktop friendly table
        '''
        DATE_FORMAT = current.globalenv['DATE_FORMAT']
        T = current.globalenv['T']
        auth = current.globalenv['auth']

        rows = self.get_day_rows()

        if len(rows) == 0:
            table = DIV()
        else:
            table = TABLE(TR(TH(' ', _class='td_status_marker'), # status marker
                             TH(T('Location'), _class='location'),
                             TH(T('Shift'), _class='classtype'),
                             TH(T('Time'), _class='time'),
                             TH(T('Employee'), _class='teacher'),
                             TH(T('')),
                             _class='os-table_header'),
                          _class='os-schedule')
            # Generate list of classes
            for i, row in enumerate(rows):
                repr_row = list(rows[i:i+1].render())[0]
                shID = row.shifts.id

                get_status = self._get_day_row_status(row)
                status = get_status['status']
                status_marker = get_status['marker']
                # filter status
                if self.filter_id_status:
                    if status != self.filter_id_status:
                        continue

                result = self._get_day_row_staff(row, repr_row, status)
                employee_id  = result['employee_id']
                employee_id2 = result['employee_id2']
                employee     = result['employee']
                employee2    = result['employee2'] or ''

                # check filter for employees
                if self.filter_id_employee:
                    self.filter_id_employee = int(self.filter_id_employee)
                    filter_check = (self.filter_id_employee == employee_id or
                                    self.filter_id_employee == employee_id2)
                    if not filter_check:
                        # break loop if it's not the employee searched for
                        continue

                location = max_string_length(
                    repr_row.shifts.school_locations_id, 15)
                shift_type = max_string_length(
                    repr_row.shifts.school_shifts_id, 24)
                time = SPAN(repr_row.shifts.Starttime, ' - ',
                            repr_row.shifts.Endtime)

                buttons = self._get_day_table_buttons(shID)

                row_class = TR(TD(status_marker),
                               TD(location),
                               TD(shift_type),
                               TD(time),
                               TD(employee),
                               TD(buttons),
                               _class='os-schedule_class')
                row_tools = TR(TD(' '),
                               TD(_colspan=3),
                               TD(employee2, _class='grey'),
                               TD(),
                               _class='os-schedule_links',
                               _id='class_' + unicode(shID))

                table.append(row_class)
                table.append(row_tools)

        return dict(table=table,
                    weekday=NRtoDay(self.date.isoweekday()),
                    date=self.date.strftime(DATE_FORMAT))


    def get_day_list(self, show_id = False):
        '''
            Format rows as list
        '''
        os_gui = current.globalenv['os_gui']
        DATE_FORMAT = current.globalenv['DATE_FORMAT']
        T = current.globalenv['T']
        date_formatted = self.date.strftime(DATE_FORMAT)

        rows = self.get_day_rows()

        shifts = []
        for i, row in enumerate(rows):
            repr_row = list(rows[i:i+1].render())[0]

            # get status
            status_result = self._get_day_row_status(row)
            status = status_result['status']

            # get employees
            result = self._get_day_row_staff(row, repr_row, status)
            employee_id  = result['employee_id']
            employee_id2 = result['employee_id2']
            employee     = result['employee']
            employee2    = result['employee2']

            # check filter for employees
            if self.filter_id_employee:
                self.filter_id_employee = int(self.filter_id_employee)
                filter_check = (self.filter_id_employee == employee_id or
                                self.filter_id_employee == employee_id2)
                if not filter_check:
                    # break loop if it's not the employee searched for
                    continue


            # get cancelled
            cancelled = False
            if status == 'cancelled':
                cancelled = True

            # get sub
            sub = False
            if status == 'sub':
                subteacher = True

            # populate class data
            data = dict()
            if show_id:
                data['id'] = row.shifts.id
            data['LocationID'] = row.shifts.school_locations_id
            data['Location'] = repr_row.shifts.school_locations_id
            data['Starttime'] = repr_row.shifts.Starttime
            data['Endtime'] = repr_row.shifts.Endtime
            data['ShiftID'] = row.shifts.school_shifts_id
            data['Shift'] = repr_row.shifts.school_shifts_id
            data['EmployeeID'] = employee_id
            data['EmployeeID2'] = employee_id2
            data['Employee'] = employee
            data['Employee2'] = employee2
            data['Sub'] = sub
            data['Cancelled'] = cancelled

            shifts.append(data)

        return shifts


#TODO: update shift status
class ShiftStatus:
    '''
        Class to manage status of a shift
    '''
    def __init__(self, shID, date):
        self.date = date
        self.shID = shID


    def _get_sotcID(self):
        '''
            :return: db.shifts_otc.id
        '''
        db = current.globalenv['db']

        query = (db.shifts_otc.shifts_id == self.shID) & \
                (db.shifts_otc.ShiftDate == self.date)
        rows = db(query).select(db.shifts_otc.id)

        sotcID = None
        if len(rows):
            sotcID = rows.first().id

        return sotcID


    def _set_status(self, status):
        '''
            :param status: ['open' or 'cancelled']
            :return: None
        '''
        db = current.globalenv['db']
        T = current.globalenv['T']

        sotcID = self.set_normal()

        row = db.shifts_otc(sotcID)
        if row:
            row.Status = status
            row.update_record()
        else:
            db.shifts_otc.insert(
                shifts_id = self.shID,
                ShiftDate = self.date,
                Status    = status
            )

    def set_normal(self):
        '''
            Remove status if found
        '''
        db = current.globalenv['db']

        sotcID = self._get_sotcID()

        row = db.shifts_otc(sotcID)
        if row:
            row.Status = None
            row.update_record()

        return sotcID


    def set_open(self):
        '''
            Change status to open
        '''
        self._set_status('open')


    def set_cancelled(self):
        '''
            Change status to cancelled
        '''
        self._set_status('cancelled')


class SchoolSubscription:
    '''
        Class that contains functions for school subscriptions
    '''
    def __init__(self, ssuID):
        '''
            Class init function which sets ssuID
        '''
        db = current.globalenv['db']

        self.ssuID = ssuID


    def _set_dbinfo(self):
        '''
            Gets information about the subscription from db and adds it
            to the object
        '''
        db = current.globalenv['db']

        row = db.school_subscriptions(self.ssuID)

        self.Name               = row.Name
        self.Classes            = row.Classes
        self.SubscriptionUnit   = row.SubscriptionUnit
        self.Archived           = row.Archived
        self.Terms              = row.Terms


    def get_price_on_date(self, date, formatted=True):
        '''
            Returns the price for a subscription on a given date
        '''
        db = current.globalenv['db']

        price = ''
        query = (db.school_subscriptions_price.school_subscriptions_id ==
                 self.ssuID) & \
                (db.school_subscriptions_price.Startdate <= date) & \
                ((db.school_subscriptions_price.Enddate >= date) |
                 (db.school_subscriptions_price.Enddate == None))

        rows = db(query).select(db.school_subscriptions_price.ALL,
                                orderby=db.school_subscriptions_price.Startdate)
        if len(rows):
            if formatted:
                repr_row = list(rows[0:1].render())[0] # first row
                price = repr_row.Price
            else:
                row = rows.first()
                price = row.Price

        if not price:
            price = 0

        return price

    def get_tax_rates_on_date(self, date):
        '''
            Returns tax rates on date
        '''
        db = current.globalenv['db']

        left = [ db.tax_rates.on(db.school_subscriptions_price.tax_rates_id ==
                                 db.tax_rates.id) ]

        query = (db.school_subscriptions_price.school_subscriptions_id ==
                 self.ssuID) & \
                (db.school_subscriptions_price.Startdate <= date) & \
                ((db.school_subscriptions_price.Enddate >= date) |
                 (db.school_subscriptions_price.Enddate == None))

        rows = db(query).select(db.school_subscriptions.ALL,
                                db.school_subscriptions_price.ALL,
                                db.tax_rates.ALL,
                                left=left,
                                orderby=db.school_subscriptions_price.Startdate)

        if rows:
            row = rows.first()
        else:
            row = None

        return row

    def get_name(self):
        '''
            Returns the name of the subscription
        '''
        self._set_dbinfo()

        return self.Name


    def get_classes_formatted(self):
        '''
            SPAN object containing
        '''
        T = current.globalenv['T']
        self._set_dbinfo()

        classes_text = T('classes')
        if self.Classes == 1:
            classes_text = T('class')

        classes = ''
        if self.SubscriptionUnit == 'week':
            classes = SPAN(unicode(self.Classes) + ' ' + classes_text + ' ' + T('a') + ' ' + T('week'))
        elif self.SubscriptionUnit == 'month':
            classes = SPAN(unicode(self.Classes) + ' ' + classes_text + ' ' + T('a') + ' ' + T('month'))

        return classes


class OsMail:
    def send(self, msgID, cuID): # Used to be 'mail_customer()'
        '''
            Send a message to a customer
            returns True when a mail is sent and False when it failed
        '''
        db = current.globalenv['db']
        MAIL = current.globalenv['MAIL']

        customer = db.auth_user(cuID)
        message = db.messages(msgID)

        check = MAIL.send(
        to=customer.email,
        subject=message.msg_subject,
        # If reply_to is omitted, then mail.settings.sender is used
        reply_to=None,
        message=message.msg_content)

        if check:
            status = 'sent'
            rvalue = True
        else:
            status = 'fail'
            rvalue = False
        db.customers_messages.insert(auth_customer_id = cuID,
                                     messages_id      = msgID,
                                     Status           = status)

        return rvalue


    def _render_email_template_order(self, template_content, customers_orders_id):
        '''
            :param customers_orders_id:
            :return: mail body for order_received & order_delivered
        '''
        def get_row(value_left, value_right, first=False, total=False):
            border = ''
            font_weight = ''
            if first:
                border = "border-top: 1px dashed #aaaaaa;"

            if total:
                border = "border-top: 1px solid #eaeaea; border-bottom: 1px dashed #aaaaaa;"
                font_weight = "font-weight:bold;"

            tr = TR(TD(
                TABLE(TR(TD(TABLE(TR(TD(TABLE(TR(TD(value_left, # left column
                                                    _align="left", _style="font-family: Arial, sans-serif; color: #333333; font-size: 16px; " + font_weight)),
                                              _cellpadding="0", _cellspacing="0", _border="0", _width="100%"),
                                        _style="padding: 0 0 10px 0;")),
                                  _cellpadding="0", _cellspacing="0", _border="0", _width="47%", _style="width:47%;", _align="left"),
                            TABLE(TR(TD(TABLE(TR(TD(value_right, # right column
                                                    _align="right", _style="font-family: Arial, sans-serif; color: #333333; font-size: 16px;  " + font_weight)),
                                              _cellpadding="0", _cellspacing="0", _border="0", _width="100%"),
                                        _style="padding: 0 0 10px 0;")),
                                  _cellpadding="0", _cellspacing="0", _border="0", _width="47%", _style="width:47%;", _align="right"),
                            _valign="top", _class="mobile-wrapper")),
                      _cellspacing="0", _cellpadding="0", _border="0", _width="100%"),
                _style="padding: 10px 0 0 0; " + border))

            return tr


        T = current.globalenv['T']
        DATETIME_FORMAT = current.globalenv['DATETIME_FORMAT']
        represent_float_as_amount = current.globalenv['represent_float_as_amount']

        order = Order(customers_orders_id)
        item_rows = order.get_order_items_rows()
        order_items = TABLE(_border="0", _cellspacing="0", _cellpadding="0", _width="100%", _style="max-width: 500px;", _class="responsive-table")
        for i, row in enumerate(item_rows):
            repr_row = list(item_rows[i:i + 1].render())[0]

            first = False
            if i == 0:
                first = True

            tr = get_row(SPAN(row.ProductName, ' ', row.Description), repr_row.TotalPriceVAT, first)
            order_items.append(tr)

        # add total row
        amounts = order.get_amounts()
        total_row = get_row(T('Total'), represent_float_as_amount(amounts.TotalPriceVAT), total=True)
        order_items.append(total_row)

        # TODO: Add to manual & button on page available variables;

        # return XML(order_items)
        return XML(template_content.format(order_id=order.order.id,
                                           order_date=order.order.DateCreated.strftime(DATETIME_FORMAT),
                                           order_status=order.order.Status,
                                           order_items=order_items,
                                           link_profile_orders=URL('profile', 'orders', scheme=True, host=True),
                                           link_profile_invoices=URL('profile', 'invoices', scheme=True, host=True)))

    #
    # def _render_email_template_invoice(self, template_content, invoices_id):
    #     '''
    #         :param template_content: html template code from db.sys_properties
    #         :param invoices_id: db.invoices.id
    #         :return: mail body for invoice
    #     '''
    #     T = current.globalenv['T']
    #     DATETIME_FORMAT = current.globalenv['DATETIME_FORMAT']
    #
    #     invoice = Invoice(invoices_id)
    #     item_rows = invoice.get_invoice_items_rows()
    #     header = THEAD(TR(TH(T('Item'), _style='padding: 8px; font-size: 10px;'),
    #                       TH(T('Description'), _style='padding: 8px; font-size: 10px;'),
    #                       TH(T('Quantity'), _style='padding: 8px; font-size: 10px;'),
    #                       TH(T('Price'), _style='padding: 8px; font-size: 10px;'),
    #                       TH(T('Subtotal'), _style='padding: 8px; font-size: 10px;'),
    #                       TH(T('VAT'), _style='padding: 8px; font-size: 10px;'),
    #                       TH(T('Total'), _style='padding: 8px; font-size: 10px;')))
    #     invoice_items = TABLE(header, _style='margin-left: auto; margin-right: auto; ')
    #     for i, row in enumerate(item_rows):
    #         repr_row = list(item_rows[i:i + 1].render())[0]
    #         invoice_items.append(TR(
    #             TD(row.ProductName, _style='padding: 8px; font-size: 10px;'),
    #             TD(row.Description, _style='padding: 8px; font-size: 10px;'),
    #             TD(row.Quantity, _style='padding: 8px; font-size: 10px;'),
    #             TD(repr_row.Price, _style='padding: 8px; font-size: 10px;'),
    #             TD(repr_row.TotalPrice, _style='padding: 8px; font-size: 10px;'),
    #             TD(repr_row.VAT, _style='padding: 8px; font-size: 10px;'),
    #             TD(repr_row.TotalPriceVAT, _style='padding: 8px; font-size: 10px;'),
    #         ))
    #
    #     # TODO: Add to manual & button on page available variables;
    #     return XML(template_content.format(invoice_id=invoice.invoice.InvoiceID,
    #                                        invoice_date_created=invoice.invoice.DateCreated.strftime(DATETIME_FORMAT),
    #                                        invoice_date_due=invoice.invoice.DateDue.strftime(DATETIME_FORMAT),
    #                                        invoice_items=invoice_items,
    #                                        link_profile_invoices=URL('profile', 'invoices', scheme=True, host=True)))


    # def _render_email_template_payment(self, template_content, invoices_payments_id):
    #     '''
    #         :param template_content: html template code from db.sys_properties
    #         :param invoices_id: db.invoices_payments_id
    #         :return: mail body for invoice
    #     '''
    #     db = current.globalenv['db']
    #     T = current.globalenv['T']
    #     DATE_FORMAT = current.globalenv['DATE_FORMAT']
    #     CURRSYM = current.globalenv['CURRSYM']
    #
    #     payment = db.invoices_payments(invoices_payments_id)
    #     invoice = Invoice(payment.invoices_id)
    #
    #     # TODO: Add to manual & button on page available variables;
    #     return XML(template_content.format(invoice_id=invoice.invoice.InvoiceID,
    #                                        payment_amount=SPAN(CURRSYM, ' ', format(payment.Amount, '.2f')),
    #                                        payment_date=payment.PaymentDate.strftime(DATE_FORMAT),
    #                                        link_profile_invoices=URL('profile', 'invoices', scheme=True, host=True)))


    def _render_email_template_payment_recurring_failed(self, template_content):
        '''
            :param template_content: html template code from db.sys_properties
            :param invoices_id: db.invoices_payments_id
            :return: mail body for invoice
        '''
        db = current.globalenv['db']
        T = current.globalenv['T']
        DATE_FORMAT = current.globalenv['DATE_FORMAT']

        # TODO: Add to manual & button on page available variables;
        return XML(template_content.format(link_profile_invoices=URL('profile', 'invoices', scheme=True, host=True)))


    def _render_email_workshops_info_mail(self, wspc, wsp, ws):
        '''
        :param template_content: Mail content
        :param workshops_products_id: db.workshops_products.id
        :return: mail body for workshop
        '''
        db = current.globalenv['db']
        T = current.globalenv['T']
        DATE_FORMAT = current.globalenv['DATE_FORMAT']
        TIME_FORMAT = current.globalenv['TIME_FORMAT']
        customer = Customer(wspc.auth_customer_id)

        try:
            time_info = TR(TH(T('Date')),
                           TD(ws.Startdate.strftime(DATE_FORMAT), ' ', ws.Starttime.strftime(TIME_FORMAT), ' - ',
                              ws.Enddate.strftime(DATE_FORMAT), ' ', ws.Endtime.strftime(TIME_FORMAT),
                              _align="left"))
        except AttributeError:
            time_info = ''

        description = TABLE(TR(TH(T('Ticket')),
                               TD(wsp.Name, _align="left")),
                            time_info,
                            _cellspacing="0", _cellpadding='5px', _width='100%', border="0")

        wsm = db.workshops_mail(workshops_id=ws.id)
        try:
            content = wsm.MailContent
        except AttributeError:
            content = ''


        image = IMG(_src=URL('default', 'download', ws.picture, scheme=True, host=True))

        return dict(content=DIV(image, BR(), BR(), XML(content)), description=description)


    def render_email_template(self,
                              email_template,
                              subject='',
                              template_content=None,
                              customers_orders_id=None,
                              invoices_id=None,
                              invoices_payments_id=None,
                              workshops_products_customers_id=None,
                              return_html=False):
        '''
            Renders default email template
        '''
        db = current.globalenv['db']
        T = current.globalenv['T']
        DATETIME_FORMAT = current.globalenv['DATETIME_FORMAT']

        get_sys_property = current.globalenv['get_sys_property']
        request = current.globalenv['request']
        response = current.globalenv['response']

        title = ''
        description = ''
        comments = ''

        logo = self._render_email_template_get_logo()

        template = os.path.join(request.folder, 'views', 'templates/email/default.html')
        if template_content is None:
            # Get email template from settings
            template_content = get_sys_property(email_template)

        if email_template == 'email_template_order_received' or email_template == 'email_template_order_delivered':
            if email_template == 'email_template_order_received':
                subject = T('Order received')
            else:
                subject = T('Order delivered')
            # do some pre-processing to show the correct order info
            content = self._render_email_template_order(template_content, customers_orders_id)
        # elif email_template == 'email_template_invoice_created':
        #     subject = T('Invoice')
        #     content = self._render_email_template_invoice(template_content, invoices_id)
        # elif email_template == 'email_template_payment_received':
        #     subject = T('Payment received')
        #     content = self._render_email_template_payment(template_content, invoices_payments_id)
        elif email_template == 'email_template_payment_recurring_failed':
            subject = T('Recurring payment failed')
            content = self._render_email_template_payment_recurring_failed(template_content)
        elif email_template == 'workshops_info_mail':
            wspc = db.workshops_products_customers(workshops_products_customers_id)
            wsp = db.workshops_products(wspc.workshops_products_id)
            ws = db.workshops(wsp.workshops_id)
            subject = ws.Name
            title = ws.Name
            result = self._render_email_workshops_info_mail(wspc, wsp, ws)
            content = result['content']
            description = result['description']
        else:
            template = os.path.join(request.folder, 'views', 'templates/email/default_simple.html')
            content = XML(template_content)
            subject = subject

        footer = XML(get_sys_property('email_template_sys_footer'))

        message =  response.render(template,
                                   dict(logo=logo,
                                        title=title,
                                        description=description,
                                        content=content,
                                        comments=comments,
                                        footer=footer))

        if return_html:
            return message
        else:
            msgID = db.messages.insert(
                msg_content = message,
                msg_subject = subject
            )

            return msgID


    def _render_email_template_get_logo(self):
        '''
            Returns logo for email template
        '''
        request = current.globalenv['request']

        branding_logo = os.path.join(request.folder,
                                     'static',
                                     'plugin_os-branding',
                                     'logos',
                                     'branding_logo_invoices.png')
        if os.path.isfile(branding_logo):
            abs_url = '%s://%s/%s/%s' % (request.env.wsgi_url_scheme,
                                         request.env.http_host,
                                         'static',
                            'plugin_os-branding/logos/branding_logo_invoices.png')
            logo_img = IMG(_src=abs_url)

        else:
            logo_img = ''

        return logo_img


class OsScheduler:
    def set_tasks(self):
        '''
            Queue all tasks here
            Call during setup & migration
        '''
        scheduler = current.globalenv['scheduler']

        today = datetime.date.today()
        start_time = datetime.datetime(today.year,
                                       today.month,
                                       today.day,
                                       00,
                                       01)  # Do stuff at 1 minute past midnight

        # clean up first
        self._remove_tasks()
        # add all tasks
        scheduler.queue_task('daily',
                             start_time=start_time,
                             timeout=1800, # Run for max half an hour
                             prevent_drift=True,
                             period=24*60*60, # once a day
                             repeats=0, # Every day
                             )

    def _remove_tasks(self):
        '''
            Removes all scheduled tasks
        '''
        db = current.globalenv['db']

        query = (db.scheduler_task.id > 0)
        db(query).delete()


class ReportsHelper:
    def get_query_subscriptions_new_in_month(self, date):
        """
            Returns query for new subscriptions
        """
        firstdaythismonth = datetime.date(date.year, date.month, 1)
        next_month = date.replace(day=28) + datetime.timedelta(days=4)  # this will never fail
        lastdaythismonth = next_month - datetime.timedelta(days=next_month.day)

        query = """SELECT cu.id,
                          cu.archived,
                          cu.thumbsmall,
                          cu.birthday,
                          cu.display_name,
                          cu.date_of_birth,
                          csu.school_subscriptions_id,
                          csu.startdate,
                          csu.payment_methods_id
                   FROM auth_user cu
                   LEFT JOIN
                       (SELECT auth_customer_id,
                               startdate,
                               enddate,
                               school_subscriptions_id,
                               payment_methods_id
                        FROM customers_subscriptions
                        GROUP BY auth_customer_id) csu
                   ON cu.id = csu.auth_customer_id
                   LEFT JOIN school_subscriptions ssu
                   ON ssu.id = csu.school_subscriptions_id
                   ,
                   (SELECT min(startdate) startdate,
                                          auth_customer_id
                    FROM customers_subscriptions
                    GROUP BY auth_customer_id) chk
                   WHERE chk.startdate = csu.startdate AND
                         chk.auth_customer_id = csu.auth_customer_id AND
                         csu.startdate >= '{firstdaythismonth}' AND csu.startdate <= '{lastdaythismonth}'
                   ORDER BY ssu.Name,
                            cu.display_name
                            DESC""".format(firstdaythismonth=firstdaythismonth,
                                           lastdaythismonth=lastdaythismonth)
        return query


class ShopBrands:
    def __init__(self, show_archive=False):
        self.show_archive = show_archive


    def list(self):
        """
            :return: List of shop brands (gluon.dal.rows)
        """
        db = current.globalenv['db']

        query = (db.shop_brands.Archived == self.show_archive)
        rows = db(query).select(db.shop_brands.ALL,
                                orderby=db.shop_brands.Name)

        return rows


    def list_formatted(self):
        """
            :return: HTML table with shop brands
        """
        T = current.globalenv['T']
        os_gui = current.globalenv['os_gui']
        auth = current.globalenv['auth']

        header = THEAD(TR(TH(T('Brand')),
                          TH(T('Description')),
                          TH()))
        table = TABLE(header, _class='table table-striped table-hover')

        permission_edit = (auth.has_membership(group_id='Admins') or
                           auth.has_permission('update', 'shop_brands'))

        rows = self.list()
        for row in rows:
            buttons = ''
            edit = ''
            archive = ''
            vars = {'sbID':row.id}

            if permission_edit:
                edit = os_gui.get_button('edit',
                    URL('shop_manage', 'brand_edit', vars=vars))
                archive = os_gui.get_button('archive',
                    URL('shop_manage', 'brand_archive', vars=vars))
                buttons = DIV(edit, archive, _class='pull-right')

            tr = TR(
                TD(os_gui.max_string_length(row.Name, 30)),
                TD(os_gui.max_string_length(row.Description, 60)),
                TD(buttons)
            )

            table.append(tr)

        return table


class ShopProductsSet:
    def __init__(self, spsID):
        db = current.globalenv['db']
        self.spsID = spsID
        self.row = db.shop_products_sets(self.spsID)


    def options(self):
        """
            :return: list of options for a products set
        """
        db = current.globalenv['db']

        query = (db.shop_products_sets_options.shop_products_sets_id ==
                 self.spsID)
        return db(query).select(db.shop_products_sets_options.ALL,
                                orderby=db.shop_products_sets_options.Name)


    def get_option_names(self):
        """
            :return: dict mapping ids to option names
        """
        options = self.options()
        names = {}
        for option in options:
            names[option.id] = option.Name

        return names


    def options_with_values(self):
        """
            :return: list of options with values for a products set
        """
        db = current.globalenv['db']

        options = {}
        for option in self.options():
            query = (db.shop_products_sets_options_values.shop_products_sets_options_id ==
                     option.id)
            rows = db(query).select(db.shop_products_sets_options_values.ALL,
                                    orderby=db.shop_products_sets_options_values.Name)
            values = []
            for row in rows:
                values.append(int(row.id))

            options[option.id] = {
                'name': option.Name,
                'values': values
            }

        return options


    def get_value_names(self):
        """
             :return: dict[db.shop_products_sets_options_values.id] = name
        """
        db = current.globalenv['db']

        option_ids = []
        for option in self.options():
            option_ids.append(option.id)

        query = (db.shop_products_sets_options_values.shop_products_sets_options_id.belongs(option_ids))
        rows = db(query).select(db.shop_products_sets_options_values.id,
                                db.shop_products_sets_options_values.Name)
        value_names = {}
        for row in rows:
            value_names[row.id] = row.Name

        return value_names


    def get_linked_products(self):
        """
        :return: list containing ids of linked products
        """
        db = current.globalenv['db']

        query = (db.shop_products.shop_products_sets_id == self.spsID)
        rows = db(query).select(db.shop_products.id)
        ids = []
        for row in rows:
            ids.append(row.id)

        return ids


    def insert_variants(self, enabled=True):
        """
        insert (missing) variants for all products linked to this set
        :param enabled: boolean
        :return: None
        """
        linked_products = self.get_linked_products()
        for shop_products_id in linked_products:
            self.insert_variants_for_product(shop_products_id,
                                             enabled=enabled)


    def insert_variants_for_product(self, shop_products_id, enabled=True):
        """
        :param shop_products_id: db.shop_products.id
        :param enabled: boolean
        :return: None
        """
        from itertools import product, combinations, permutations

        db = current.globalenv['db']
        options = self.options_with_values()
        value_names = self.get_value_names()

        values = []
        for key in options:
            values.append(options[key]['values'])
        variants = list(product(*values))

        for i, variant in enumerate(variants):
            variant_code = '-'.join(str(value) for value in variant)
            variant_name = ''
            for value in variant:
                option_name = ''
                value_name = value_names.get(value, '')
                for key in options:
                    if value in options[key]['values']:
                        option_name = options[key]['name']
                        if len(variant_name):
                            variant_name += ', '
                        variant_name += option_name + ': ' + value_name
                        break

            query = (db.shop_products_variants.VariantCode == variant_code) & \
                    (db.shop_products_variants.shop_products_id == shop_products_id)
            count = db(query).count()
            if not count:
                db.shop_products_variants.insert(
                    Enabled=enabled,
                    shop_products_id = shop_products_id,
                    Name = variant_name,
                    DefaultVariant = True if not i else False,
                    VariantCode = variant_code
                )


class ShopProductsSets:
    def list(self):
        """
            :return: List of shop products_sets (gluon.dal.rows)
        """
        db = current.globalenv['db']

        query = (db.shop_products_sets)
        rows = db(query).select(db.shop_products_sets.ALL,
                                orderby=db.shop_products_sets.Name)

        return rows


    def list_formatted(self):
        """
            :return: HTML table with shop products_sets
        """
        T = current.globalenv['T']
        os_gui = current.globalenv['os_gui']
        auth = current.globalenv['auth']

        header = THEAD(TR(TH(T('Product set')),
                          TH(T('Description')),
                          TH()))
        table = TABLE(header, _class='table table-striped table-hover')

        permission_options = (auth.has_membership(group_id='Admins') or
                              auth.has_permission('read', 'shop_products_sets_options'))
        permission_edit = (auth.has_membership(group_id='Admins') or
                           auth.has_permission('update', 'shop_products_sets'))
        permission_delete = (auth.has_membership(group_id='Admins') or
                             auth.has_permission('delete', 'shop_products_sets'))

        onclick_delete = "return confirm('" \
            + T('Do you really want to delete this product set?') + ' ' \
            + T('It will remove all product variants in products linked to this set.') \
            + "');"


        rows = self.list()
        for row in rows:
            buttons = ''
            edit = ''
            delete = ''
            vars = {'spsID':row.id}
            buttons = DIV(_class="pull-right")

            if permission_options:
                options = os_gui.get_button(
                    'noicon',
                    URL('shop_manage', 'products_set_options', vars=vars),
                    title=T('Options')
                )
                buttons.append(options)

            if permission_edit:
                edit = os_gui.get_button('edit',
                    URL('shop_manage', 'products_set_edit', vars=vars))
                buttons.append(edit)
            if permission_delete:
                delete = os_gui.get_button('delete_notext',
                    URL('shop_manage', 'products_set_delete', vars=vars),
                    onclick = onclick_delete)
                buttons.append(delete)

            tr = TR(
                TD(os_gui.max_string_length(row.Name, 30)),
                TD(os_gui.max_string_length(row.Description, 60)),
                TD(buttons)
            )

            table.append(tr)

        return table


class ShopProductsSetsOptions:
    def __init__(self,
                 products_sets_id,
                 url_list):
        self.products_sets_id = products_sets_id
        self.url_list = url_list


    def has_linked_products(self):
        """
            :return: boolean
        """
        db = current.globalenv['db']

        query = (db.shop_products.shop_products_sets_id ==
                 self.products_sets_id)

        return True if db(query).count() else False


    def list(self):
        """
            :return: List of shop products sets options
        """
        db = current.globalenv['db']

        query = (db.shop_products_sets_options.shop_products_sets_id ==
                 db.shop_products_sets.id)
        rows = db(query).select(db.shop_products_sets_options.ALL,
                                orderby=db.shop_products_sets_options.Name)

        return rows


    def list_formatted(self):
        """
            :return: HTML table with shop products sets options
        """
        T = current.globalenv['T']
        os_gui = current.globalenv['os_gui']
        auth = current.globalenv['auth']

        linked_products = self.has_linked_products()

        header = THEAD(TR(TH(T('Option')),
                          TH(T('Values')),
                          TH()))
        table = TABLE(header, _class='table')

        if linked_products:
            permission_delete = False
        else:
            permission_delete = (auth.has_membership(group_id='Admins') or
                                 auth.has_permission('delete', 'shop_products_options'))

        onclick_delete = "return confirm('" \
            + T('Do you really want to delete this option?') + "');"

        rows = self.list()
        for row in rows:
            buttons = DIV()
            delete = ''
            vars = {'spsoID':row.id}

            if permission_delete:
                delete = os_gui.get_button('delete_notext',
                    URL('shop_manage',
                        'shop_products_sets_options_delete',
                        vars=vars),
                    onclick=onclick_delete,
                    _class='pull-right')
                buttons.append(delete)

            tr = TR(
                TD(os_gui.max_string_length(row.Name, 30)),
                TD(self._list_formatted_get_option_values(row.id,
                                                          self.url_list)),
                TD(buttons)
            )

            table.append(tr)

        if not linked_products:
            table.append(TR(TD(self._list_formatted_get_form_add())))

        return table


    def _list_formatted_get_option_values(self, options_id, url_list):
        """
            :return: returns a list of option values for an option
        """
        spsov = ShopProductsSetsOptionsValues(options_id, url_list)
        return spsov.list_formatted()


    def _list_formatted_get_form_add(self):
        """
            :return: CRUD form to add an option
        """
        from os_forms import OsForms

        T = current.globalenv['T']
        db = current.globalenv['db']

        db.shop_products_sets_options.Name.label = ''
        db.shop_products_sets_options.shop_products_sets_id.default = \
            self.products_sets_id

        os_forms = OsForms()
        result = os_forms.get_crud_form_create(
            db.shop_products_sets_options,
            self.url_list,
            submit_button=T("Add option"),
            form_id="AddOption"
        )

        return DIV(result['form'], result['submit'])


class ShopProductsSetsOptionsValues:
    def __init__(self, options_id, url_list):
        self.options_id = options_id
        self.url_list = url_list


    def list(self):
        """
            :return: List of shop products sets options values (gluon.dal.rows)
        """
        db = current.globalenv['db']

        query = (db.shop_products_sets_options_values.shop_products_sets_options_id == \
                 self.options_id)
        rows = db(query).select(db.shop_products_sets_options_values.ALL,
                                orderby=db.shop_products_sets_options_values.Name)

        return rows


    def list_formatted(self):
        """
            :return: HTML table with shop categories
        """
        T = current.globalenv['T']
        os_gui = current.globalenv['os_gui']
        auth = current.globalenv['auth']

        table = TABLE(_class='table')

        permission_create = (auth.has_membership(group_id='Admins') or
                             auth.has_permission('create', 'shop_products_options_values'))
        permission_delete = (auth.has_membership(group_id='Admins') or
                             auth.has_permission('delete', 'shop_products_options_values'))
        onclick_delete = "return confirm('" \
            + T('Do you really want to delete this option value?') + "');"

        rows = self.list()
        for row in rows:
            buttons = DIV()
            delete = ''
            vars = {'spsovID':row.id}

            if permission_delete:
                delete = os_gui.get_button('delete_notext',
                    URL('shop_manage',
                        'shop_products_sets_options_value_delete',
                        vars=vars),
                    onclick=onclick_delete,
                    _class='pull-right')
                buttons.append(delete)

            tr = TR(
                TD(os_gui.max_string_length(row.Name, 30)),
                TD(buttons)
            )

            table.append(tr)

        table.append(TR(TD(self._list_formatted_get_form_add())))

        return table


    def _list_formatted_get_form_add(self):
        """
            :return: CRUD form to add an option
        """
        from os_forms import OsForms

        T = current.globalenv['T']
        db = current.globalenv['db']
        request = current.globalenv['request']

        # make sure the value is saved for the right option
        if 'shop_products_sets_options_id' in request.vars:
            options_id = request.vars['shop_products_sets_options_id']
        else:
            options_id = self.options_id

        db.shop_products_sets_options_values.Name.label = ''
        db.shop_products_sets_options_values.shop_products_sets_options_id.default = \
            options_id

        form_id = "AddValue_" + unicode(self.options_id)

        os_forms = OsForms()
        result = os_forms.get_crud_form_create(
            db.shop_products_sets_options_values,
            self.url_list,
            submit_button=T("Add value"),
            form_id=form_id,
            onaccept=[self._product_set_options_update_variants]
        )

        form = result['form']
        field_id = INPUT(_type='hidden',
                         _value=self.options_id,
                         _form=form_id,
                         _name='shop_products_sets_options_id')

        form.insert(0, field_id)

        return DIV(form, result['submit'])


    def _product_set_options_update_variants(self, form):
        """
        :param form:
        :return:
        """
        db = current.globalenv['db']

        spsovID = form.vars.id
        option = db.shop_products_sets_options(self.options_id)
        spsID = option.shop_products_sets_id
        product_set = ShopProductsSet(spsID)
        product_set.insert_variants(enabled=False)


class ShopCategories:
    def __init__(self, show_archive=False):
        self.show_archive = show_archive


    def list(self):
        """
            :return: List of shop categories (gluon.dal.rows)
        """
        db = current.globalenv['db']

        query = (db.shop_categories.Archived == self.show_archive)
        rows = db(query).select(db.shop_categories.ALL,
                                orderby=db.shop_categories.Name)

        return rows


    def list_formatted(self):
        """
            :return: HTML table with shop categories
        """
        T = current.globalenv['T']
        os_gui = current.globalenv['os_gui']
        auth = current.globalenv['auth']

        header = THEAD(TR(TH(T('Category')),
                          TH(T('Description')),
                          TH()))
        table = TABLE(header, _class='table table-striped table-hover')

        permission_edit = (auth.has_membership(group_id='Admins') or
                           auth.has_permission('update', 'shop_categories'))

        rows = self.list()
        for row in rows:
            buttons = ''
            edit = ''
            archive = ''
            vars = {'scID':row.id}

            if permission_edit:
                edit = os_gui.get_button('edit',
                    URL('shop_manage', 'category_edit', vars=vars))
                archive = os_gui.get_button('archive',
                    URL('shop_manage', 'category_archive', vars=vars))
                buttons = DIV(edit, archive, _class='pull-right')

            tr = TR(
                TD(os_gui.max_string_length(row.Name, 30)),
                TD(os_gui.max_string_length(row.Description, 60)),
                TD(buttons)
            )

            table.append(tr)

        return table


class ShopSuppliers:
    def __init__(self, show_archive=False):
        self.show_archive = show_archive


    def list(self):
        """
            :return: List of shop suppliers (gluon.dal.rows)
        """
        db = current.globalenv['db']

        query = (db.shop_suppliers.Archived == self.show_archive)
        rows = db(query).select(db.shop_suppliers.ALL,
                                orderby=db.shop_suppliers.Name)

        return rows


    def list_formatted(self):
        """
            :return: HTML table with shop brands
        """
        T = current.globalenv['T']
        os_gui = current.globalenv['os_gui']
        auth = current.globalenv['auth']

        header = THEAD(TR(TH(T('Name')),
                          TH(T('Description')),
                          TH(T('Contact')),
                          TH(T('Phone')),
                          TH(T('Email')),
                          TH()))
        table = TABLE(header, _class='table table-striped table-hover')

        permission_edit = (auth.has_membership(group_id='Admins') or
                           auth.has_permission('update', 'shop_suppliers'))

        rows = self.list()
        for row in rows:
            buttons = ''
            edit = ''
            archive = ''
            vars = {'supID':row.id}

            if permission_edit:
                edit = os_gui.get_button('edit',
                    URL('shop_manage', 'supplier_edit', vars=vars))
                archive = os_gui.get_button('archive',
                    URL('shop_manage', 'supplier_archive', vars=vars))
                buttons = DIV(edit, archive, _class='pull-right')

            tr = TR(
                TD(os_gui.max_string_length(row.Name, 30)),
                TD(os_gui.max_string_length(row.Description, 30)),
                TD(os_gui.max_string_length(row.ContactName, 20),
                   _title=row.ContactName),
                TD(os_gui.max_string_length(row.ContactPhone, 15)),
                TD(os_gui.max_string_length(row.ContactEmail, 32),
                   _title=row.ContactEmail),
                TD(buttons)
            )

            table.append(tr)

        return table


class ShopProducts:
    def list(self):
        """
            :return: List of shop products (gluon.dal.rows)
        """
        db = current.globalenv['db']

        rows = db(db.shop_products).select(db.shop_products.ALL,
                                           orderby=db.shop_products.Name)

        return rows


    def list_formatted(self):
        """
            :return: HTML table with shop products
        """
        T = current.globalenv['T']
        os_gui = current.globalenv['os_gui']
        auth = current.globalenv['auth']

        header = THEAD(TR(TH(),
                          TH(T('Name')),
                          TH(T('Description')),
                          TH()))
        table = TABLE(header, _class='table table-striped table-hover')

        permission_variants = (auth.has_membership(group_id='Admins') or
                               auth.has_permission('read', 'shop_products_variants'))
        permission_edit = (auth.has_membership(group_id='Admins') or
                           auth.has_permission('update', 'shop_products'))
        permission_delete = (auth.has_membership(group_id='Admins') or
                             auth.has_permission('delete', 'shop_products'))

        onclick_delete = "return confirm('" \
            + T('Do you really want to delete this product?') + "');"

        rows = self.list()
        for i, row in enumerate(rows):
            repr_row = list(rows[i:i + 1].render())[0]
            buttons = DIV(_class='pull-right')
            vars = {'spID':row.id}

            if permission_variants:
                variants = os_gui.get_button('noicon',
                    URL('shop_manage', 'product_variants', vars=vars),
                    title=T('Variants'))
                buttons.append(variants)
            if permission_edit:
                edit = os_gui.get_button('edit',
                    URL('shop_manage', 'product_edit', vars=vars))
                buttons.append(edit)
            if permission_delete:
                delete = os_gui.get_button('delete_notext',
                    URL('shop_manage', 'product_delete', vars=vars),
                    onclick=onclick_delete)
                buttons.append(delete)

            tr = TR(
                TD(repr_row.thumbsmall),
                TD(os_gui.max_string_length(row.Name, 30)),
                TD(os_gui.max_string_length(row.Description, 30)),
                TD(buttons)
            )

            table.append(tr)

        return table


class ShopProduct:
    def __init__(self, spID):
        """
            :param spID: db.shop_products.id
        """
        db = current.globalenv['db']

        self.id = spID
        self.row = db.shop_products(self.id)


    def count_variants(self):
        """
            :return: integer - number of variants for this product
        """
        db = current.globalenv['db']
        query = (db.shop_products_variants.shop_products_id == self.id)

        return db(query).count()


    def add_default_variant(self):
        """
            Create default variant for a product without a product set
            :return: None
        """
        T = current.globalenv['T']
        db = current.globalenv['db']

        db.shop_products_variants.insert(
            Enabled=True,
            shop_products_id = self.id,
            Name = T('Default'),
            DefaultVariant = True
        )


    # def add_product_set_variants(self):
    #     """
    #     :param spsID:
    #     :return: None
    #     """
    #     db = current.globalenv['db']


    def has_products_set(self):
        """
        :return: boolean
        """
        return True if self.row.shop_products_sets_id else False


class ShopProductsVariant:
    def __init__(self, shop_products_variants_id):
        db = current.globalenv['db']

        self.id = shop_products_variants_id
        self.row = db.shop_products_variants(self.id)


    def set_default(self):
        """
            Set this product variant as default for a product
        """
        db = current.globalenv['db']

        query = (db.shop_products_variants.shop_products_id ==
                 self.row.shop_products_id)
        db(query).update(DefaultVariant=False)

        self.row.DefaultVariant = True
        self.row.update_record()


    def disable(self):
        """
            Disable variant
        """
        self.row.Enabled = False
        self.row.update_record()


    def enable(self):
        """
            Enable variant
        """
        self.row.Enabled = True
        self.row.update_record()


class ShopProductsVariants:
    def __init__(self, shop_products_id):
        self.shop_products_id = shop_products_id

    def list(self):
        """
            :return: List of shop product variants(gluon.dal.rows)
        """
        db = current.globalenv['db']

        query = (db.shop_products_variants.shop_products_id ==
                 self.shop_products_id)
        rows = db(query).select(db.shop_products_variants.ALL,
                                orderby=db.shop_products_variants.Name)

        return rows


    def list_formatted(self):
        """
            :return: HTML table with shop products variants
        """
        T = current.globalenv['T']
        os_gui = current.globalenv['os_gui']
        auth = current.globalenv['auth']

        product = ShopProduct(self.shop_products_id)

        header = THEAD(TR(TH(),
                          TH(T('Name')),
                          TH(T('Price')),
                          TH(T('Article Code')),
                          TH(T('Keep stock')),
                          TH(T('Stock shop')),
                          TH(T('Stock warehouse')),
                          TD(),
                          TH()))
        table = TABLE(header, _class='table table-striped table-hover')
        table_disabled = TABLE(header, _class='table table-striped table-hover')

        permission_edit = (auth.has_membership(group_id='Admins') or
                           auth.has_permission('update', 'shop_products_variants'))
        permission_delete = (auth.has_membership(group_id='Admins') or
                             auth.has_permission('delete', 'shop_products_variants'))

        onclick_delete = self._list_formatted_get_onclick_delete()

        rows = self.list()
        for i, row in enumerate(rows):
            repr_row = list(rows[i:i + 1].render())[0]

            default = self._list_formatted_get_label_default(T, os_gui, row)
            buttons = self._list_formatted_get_buttons(
                permission_edit,
                permission_delete,
                onclick_delete,
                T,
                os_gui,
                row
            )

            tr = TR(
                TD(repr_row.thumbsmall),
                TD(os_gui.max_string_length(row.Name, 50)),
                TD(repr_row.Price),
                TD(repr_row.ArticleCode),
                TD(repr_row.KeepStock),
                TD(row.StockShop),
                TD(row.StockWarehouse),
                TD(default),
                TD(buttons)
            )

            if row.Enabled:
                table.append(tr)
            else:
                table_disabled.append(tr)

        if product.has_products_set():
            return DIV(table, H4(T('Disabled')), table_disabled)
        else:
            return table


    def _list_formatted_get_label_default(self, T, os_gui, row):
        """

        """
        default = ''
        if row.DefaultVariant:
            default = os_gui.get_label('success', T('Default'))

        return default


    def _list_formatted_get_buttons(self,
                                    permission_edit,
                                    permission_delete,
                                    onclick_delete,
                                    T,
                                    os_gui,
                                    row):
        """
            :return:
        """
        buttons = DIV(_class='pull-right')
        vars = {'spvID': row.id, 'spID': self.shop_products_id}

        if row.Enabled:
            if permission_delete:
                disabled = False if not row.DefaultVariant else True
                delete = os_gui.get_button('delete_notext',
                                           URL('shop_manage', 'product_variant_delete',
                                               vars=vars),
                                           onclick=onclick_delete,
                                           _class='pull-right',
                                           _disabled=disabled)
                buttons.append(delete)

            if permission_edit:
                edit = self._list_formatted_get_buttons_edit(
                    T,
                    os_gui,
                    row,
                    vars
                )
                buttons.append(edit)
        else:
            buttons.append(A(T('Enable'),
                             _href=URL('shop_manage',
                                       'product_variant_enable',
                                       vars=vars)))
        return buttons


    def _list_formatted_get_buttons_edit(self, T, os_gui, row, vars):
        """
            Return edit drop down
        """
        edit = A(os_gui.get_fa_icon('fa-pencil'),
                 T('Edit'),
                 _href=URL('shop_manage', 'product_variant_edit',
                           vars=vars))
        set_default = ''
        if not row.DefaultVariant:
            set_default = A(os_gui.get_fa_icon('fa-check-circle'),
                            T('Set default'),
                            _href=URL('shop_manage', 'product_variant_set_default',
                                      vars=vars))
        links = [
            edit,
            set_default
        ]

        dd = os_gui.get_dropdown_menu(
            links=links,
            btn_text=T('Actions'),
            btn_size='btn-sm',
            btn_icon='actions',
            menu_class='btn-group pull-right')

        return dd


    def _list_formatted_get_onclick_delete(self):
        """
            :return: onclick delete for
        """
        T = current.globalenv['T']
        product = ShopProduct(self.shop_products_id)
        if product.has_products_set():
            delete_message = T('Do you really want to disable this variant?')
        else:
            delete_message = T('Do you really want to delete this variant?')
        onclick_delete = "return confirm('" \
            + delete_message + "');"

        return onclick_delete
